from flask import Flask, render_template_string, request, redirect, url_for, flash, session, send_file, g, Response
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re, io, json, os
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime

# ══════════════════════════════════════════════════════════════════════════════
# WERSJONOWANIE PARSERA I LOKALIZACJA PLIKÓW
# ══════════════════════════════════════════════════════════════════════════════
PARSER_VERSION  = 1   # Inkrementuj przy każdej zmianie logiki parsowania
MATCH_FILES_DIR = os.environ.get("MATCH_FILES_DIR", "/app/match_files")
os.makedirs(MATCH_FILES_DIR, exist_ok=True)


app = Flask(__name__)

_secret_key = os.environ.get("SECRET_KEY")
if not _secret_key:
    raise RuntimeError("SECRET_KEY environment variable is required — set it in your .env file")
app.secret_key = _secret_key

app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "false").lower() == "true"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_NAME"] = "basketkolcz_session"
app.config["PERMANENT_SESSION_LIFETIME"] = 86400 * 30  # 30 dni
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

@app.before_request
def _portal_only():
    if os.environ.get("PORTAL_ONLY", "").lower() != "true":
        return
    p = request.path
    if p == "/portal" or p.startswith("/portal/") or p.startswith("/static/") or p == "/favicon.ico":
        return
    return Response("Forbidden", 403)

import functools, hashlib

# ── Dane użytkownika (hashed password) ────────────────────────────────────────
USERS = {
    "kosma.kolcz@gmail.com": {
        "name": "Kosma Kołcz",
        "password_hash": hashlib.sha256("88614855_Basket".encode()).hexdigest(),
    }
}

def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("landing"))
        return f(*args, **kwargs)
    return decorated

@app.before_request
def keep_session_alive():
    """Odświeżaj sesję przy każdym żądaniu — zapobiega wylogowaniu."""
    session.modified = True
    if session.get("logged_in"):
        session.permanent = True


DRUZYNY_JS = '(function(){\nvar DB=window.DB_INIT||[];\nvar PLAYERS=window.PLAYERS_INIT||{};\nvar KI=-1,KEXT=false,KN="",KSEZON="",STEP=-1;\nvar TEAM_KI=-1,TEAM_S="",TEAM_D="";\nvar EDIT_PI=-1,pendDel=-1,pendDelPl=-1;\n\nfunction saveDB(){\n  fetch("/druzyny/save",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({kluby:DB,players:PLAYERS})});\n}\n\nfunction esc(s){\n  return String(s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;").replace(/"/g,"&quot;");\n}\n\nfunction draw(){\n  var root=document.getElementById("tree-root");\n  if(!root)return;\n  if(!DB.length){\n    root.innerHTML=\'<div style="padding:40px;text-align:center;color:#bbb;font-size:12px">Brak klub\\u00f3w. Kliknij + Dodaj.</div>\';\n    return;\n  }\n  var h="";\n  for(var ki=0;ki<DB.length;ki++){\n    var k=DB[ki],ext=k.ext||false;\n    h+=\'<div style="margin-bottom:10px;border-radius:9px;border:1px solid #eee;overflow:hidden">\';\n    h+=\'<div style="background:#1a2b4a;padding:8px 12px;display:flex;justify-content:space-between;align-items:center">\';\n    h+=\'<div style="display:flex;align-items:center;gap:7px;font-size:12px;font-weight:500;color:#fff">\';\n    h+=\'<span style="width:8px;height:8px;border-radius:50%;background:\'+(ext?\'#5DCAA5\':\'#EF9F27\')+\';display:inline-block"></span>\'+esc(k.name)+\'</div>\';\n    h+=\'<div style="position:relative">\';\n    h+=\'<button class="mbtn" data-ki="\'+ki+\'" style="background:rgba(255,255,255,.18);border:none;color:#fff;cursor:pointer;padding:3px 9px;border-radius:5px;font-size:15px">&#8943;</button>\';\n    h+=\'<div id="mdd\'+ki+\'" style="display:none;position:absolute;top:calc(100% + 4px);right:0;background:#fff;border:1px solid #e0e0e0;border-radius:8px;min-width:110px;z-index:200;box-shadow:0 4px 16px rgba(0,0,0,.1)">\';\n    h+=\'<div class="dde" data-ki="\'+ki+\'" style="padding:9px 14px;cursor:pointer;font-size:12px;color:#222">Edytuj</div>\';\n    h+=\'<div style="height:1px;background:#f0f0f0"></div>\';\n    h+=\'<div class="ddd" data-ki="\'+ki+\'" style="padding:9px 14px;cursor:pointer;font-size:12px;color:#A32D2D">Usu\\u0144</div>\';\n    h+=\'</div></div></div>\';\n    var sezony=k.sezony||{};\n    for(var s in sezony){\n      h+=\'<div style="padding:5px 12px;background:#f8f9fa;border-top:1px solid #eee;font-size:11px;font-weight:500;color:#666">\'+esc(s)+\'</div>\';\n      h+=\'<div style="padding:6px 12px 8px;display:flex;flex-wrap:wrap;gap:5px">\';\n      var tt=sezony[s]||[];\n      if(!tt.length) h+=\'<span style="font-size:11px;color:#ccc">brak dru\\u017cyn</span>\';\n      for(var ti=0;ti<tt.length;ti++){\n        h+=\'<button class="tbtn" data-ki="\'+ki+\'" data-s="\'+encodeURIComponent(s)+\'" data-d="\'+encodeURIComponent(tt[ti])+\'" style="display:inline-flex;align-items:center;border:1px solid #e0e0e0;border-radius:20px;padding:4px 12px;font-size:11px;color:#222;cursor:pointer;background:#fff">\'+esc(tt[ti])+\' &#8594;</button>\';\n      }\n      h+=\'</div>\';\n    }\n    h+=\'</div>\';\n  }\n  root.innerHTML=h;\n  bind();\n}\n\nfunction bind(){\n  document.querySelectorAll(".mbtn").forEach(function(b){\n    b.addEventListener("click",function(e){\n      e.stopPropagation();\n      var ki=parseInt(this.dataset.ki);\n      document.querySelectorAll("[id^=mdd]").forEach(function(d){d.style.display="none";});\n      document.getElementById("mdd"+ki).style.display="block";\n    });\n  });\n  document.querySelectorAll(".dde").forEach(function(el){el.addEventListener("click",function(){doEdit(parseInt(this.dataset.ki));});});\n  document.querySelectorAll(".ddd").forEach(function(el){el.addEventListener("click",function(){askDel(parseInt(this.dataset.ki));});});\n  document.querySelectorAll(".tbtn").forEach(function(b){\n    b.addEventListener("click",function(){\n      goTeam(parseInt(this.dataset.ki),decodeURIComponent(this.dataset.s),decodeURIComponent(this.dataset.d));\n    });\n  });\n}\n\ndocument.addEventListener("click",function(){\n  document.querySelectorAll("[id^=mdd]").forEach(function(d){d.style.display="none";});\n});\n\nfunction idle(){\n  STEP=-1;KI=-1;TEAM_KI=-1;\n  document.getElementById("sbar").style.display="none";\n  document.getElementById("rfoot").style.display="none";\n  var rb=document.getElementById("rbody");\n  rb.style.padding="40px 20px";rb.style.textAlign="center";\n  rb.innerHTML="Kliknij &#8943; przy klubie aby edytowa\\u0107. Kliknij + Dodaj aby doda\\u0107 nowy klub.";\n}\n\nfunction panel(body,foot,bar){\n  var rb=document.getElementById("rbody");rb.style.padding="0";rb.style.textAlign="";rb.innerHTML=body;\n  var rf=document.getElementById("rfoot");rf.innerHTML=foot;rf.style.display=foot?"flex":"none";\n  var sb=document.getElementById("sbar");\n  if(bar){sb.innerHTML=bar;sb.style.display="flex";}else{sb.style.display="none";}\n}\n\nfunction sbarH(){\n  var steps=["Klub","Sezon","Dru\\u017cyna","Gotowe"];\n  var h=\'<div style="display:flex;align-items:center;width:100%;padding:10px 14px 6px">\';\n  for(var i=0;i<steps.length;i++){\n    if(i>0) h+=\'<div style="flex:1;height:1px;background:\'+(i<=STEP?\'#1a2b4a\':\'#e0e0e0\')+\';margin-bottom:10px"></div>\';\n    var bg=i<STEP?\'#1a2b4a\':i===STEP?\'#EF9F27\':\'#f0f0f0\',tc=i<=STEP?\'#fff\':\'#aaa\';\n    h+=\'<div style="display:flex;flex-direction:column;align-items:center">\';\n    h+=\'<div style="width:22px;height:22px;border-radius:50%;background:\'+bg+\';color:\'+tc+\';display:flex;align-items:center;justify-content:center;font-size:10px;font-weight:600">\'+(i<STEP?\'&#10003;\':i+1)+\'</div>\';\n    h+=\'<div style="font-size:9px;color:#aaa;margin-top:2px">\'+steps[i]+\'</div></div>\';\n  }\n  return h+\'</div>\';\n}\n\nfunction fld(id,lbl,ph,val){\n  return \'<div style="margin-bottom:10px"><div style="font-size:11px;color:#666;margin-bottom:3px">\'+lbl+\'</div>\'\n    +\'<input id="\'+id+\'" placeholder="\'+ph+\'" value="\'+esc(val||\'\')+\'" style="width:100%;padding:7px 10px;border:1px solid #e0e0e0;border-radius:8px;font-size:12px;box-sizing:border-box"></div>\';\n}\n\nfunction showStep(){\n  var b="",f="";\n  if(STEP===0){\n    b=\'<div style="padding:14px 16px">\';\n    b+=\'<div style="font-size:14px;font-weight:600;color:#1a2b4a;margin-bottom:2px">Nowy klub</div>\';\n    b+=\'<div style="font-size:12px;color:#888;margin-bottom:14px">Podaj nazw\\u0119 organizacji.</div>\';\n    b+=fld("f0","Nazwa klubu","np. MKS Katowice",KN);\n    b+=\'<div style="display:flex;gap:8px">\';\n    b+=\'<div id="tc" style="flex:1;padding:7px;text-align:center;border:1px solid \'+(KEXT?\'#e0e0e0\':\'#1a2b4a\')+\';border-radius:8px;cursor:pointer;font-size:11px;background:\'+(KEXT?\'#fff\':\'#E6F1FB\')+\';color:\'+(KEXT?\'#888\':\'#0C447C\')+\'">Klub sportowy</div>\';\n    b+=\'<div id="te" style="flex:1;padding:7px;text-align:center;border:1px solid \'+(KEXT?\'#1a2b4a\':\'#e0e0e0\')+\';border-radius:8px;cursor:pointer;font-size:11px;background:\'+(KEXT?\'#E6F1FB\':\'#fff\')+\';color:\'+(KEXT?\'#0C447C\':\'#888\')+\'">Kadra</div>\';\n    b+=\'</div></div>\';\n    f=\'<button id="sc" style="flex:1;background:none;border:1px solid #ddd;color:#888;padding:8px;border-radius:8px;cursor:pointer;font-size:12px">Anuluj</button>\';\n    f+=\'<button id="sn" style="flex:2;background:#EF9F27;color:#fff;border:none;padding:8px;border-radius:8px;cursor:pointer;font-size:12px;font-weight:600">Dalej</button>\';\n  } else if(STEP===1){\n    b=\'<div style="padding:14px 16px">\';\n    b+=\'<div style="font-size:14px;font-weight:600;color:#1a2b4a;margin-bottom:2px">Dodaj sezon</div>\';\n    b+=\'<div style="font-size:12px;color:#888;margin-bottom:14px">Dla: <b>\'+esc(KN)+\'</b></div>\';\n    b+=fld("f1","Sezon","np. 2025/2026",KSEZON)+\'</div>\';\n    f=\'<button id="sb" style="flex:1;background:none;border:1px solid #ddd;color:#888;padding:8px;border-radius:8px;cursor:pointer;font-size:12px">Wstecz</button>\';\n    f+=\'<button id="sn" style="flex:2;background:#EF9F27;color:#fff;border:none;padding:8px;border-radius:8px;cursor:pointer;font-size:12px;font-weight:600">Dalej</button>\';\n  } else if(STEP===2){\n    b=\'<div style="padding:14px 16px">\';\n    b+=\'<div style="font-size:14px;font-weight:600;color:#1a2b4a;margin-bottom:2px">Dodaj dru\\u017cyn\\u0119</div>\';\n    b+=\'<div style="font-size:12px;color:#888;margin-bottom:14px">\'+esc(KN)+\' &middot; \'+esc(KSEZON)+\'</div>\';\n    b+=fld("f2","Nazwa dru\\u017cyny","np. U14","");\n    b+=fld("f2r","Rocznik (opcjonalnie)","np. 2012","")+\'</div>\';\n    f=\'<button id="sb" style="flex:1;background:none;border:1px solid #ddd;color:#888;padding:8px;border-radius:8px;cursor:pointer;font-size:12px">Wstecz</button>\';\n    f+=\'<button id="sn" style="flex:2;background:#EF9F27;color:#fff;border:none;padding:8px;border-radius:8px;cursor:pointer;font-size:12px;font-weight:600">Zapisz</button>\';\n  } else {\n    b=\'<div style="text-align:center;padding:30px 12px">\';\n    b+=\'<div style="font-size:14px;font-weight:600;color:#1a2b4a;margin-bottom:4px">Dodano!</div>\';\n    b+=\'<div style="font-size:12px;color:#888">\'+esc(KN)+\' &middot; \'+esc(KSEZON)+\'</div></div>\';\n    f=\'<button id="sm" style="flex:1;background:#1a2b4a;color:#fff;border:none;padding:8px;border-radius:8px;cursor:pointer;font-size:12px;font-weight:600">+ kolejna dru\\u017cyna</button>\';\n    f+=\'<button id="sd" style="flex:2;background:#EF9F27;color:#fff;border:none;padding:8px;border-radius:8px;cursor:pointer;font-size:12px;font-weight:600">Gotowe</button>\';\n  }\n  panel(b,f,sbarH());\n  var tc=document.getElementById("tc"),te=document.getElementById("te");\n  if(tc) tc.addEventListener("click",function(){KEXT=false;showStep();});\n  if(te) te.addEventListener("click",function(){KEXT=true;showStep();});\n  var sn=document.getElementById("sn");\n  if(sn) sn.addEventListener("click",function(){\n    if(STEP===0){\n      var v=document.getElementById("f0");KN=v?v.value.trim():"";\n      if(!KN){if(v)v.style.borderColor="#E24B4A";return;}\n      DB.push({name:KN,ext:KEXT,sezony:{}});KI=DB.length-1;saveDB();draw();STEP=1;KSEZON="";showStep();\n    } else if(STEP===1){\n      var v=document.getElementById("f1");KSEZON=v?v.value.trim():"";\n      if(!KSEZON){if(v)v.style.borderColor="#E24B4A";return;}\n      if(!DB[KI].sezony[KSEZON])DB[KI].sezony[KSEZON]=[];\n      saveDB();draw();STEP=2;showStep();\n    } else if(STEP===2){\n      var vd=document.getElementById("f2"),vr=document.getElementById("f2r");\n      var d=vd?vd.value.trim():"";\n      if(!d){if(vd)vd.style.borderColor="#E24B4A";return;}\n      var r=vr?vr.value.trim():"";\n      DB[KI].sezony[KSEZON].push(d+(r?" (rocznik "+r+")":""));\n      saveDB();draw();STEP=3;showStep();\n    }\n  });\n  var sb=document.getElementById("sb");\n  if(sb) sb.addEventListener("click",function(){STEP--;showStep();});\n  var sc=document.getElementById("sc");if(sc) sc.addEventListener("click",idle);\n  var sm=document.getElementById("sm");if(sm) sm.addEventListener("click",function(){STEP=2;showStep();});\n  var sd=document.getElementById("sd");if(sd) sd.addEventListener("click",idle);\n}\n\ndocument.getElementById("btn-add").addEventListener("click",function(){\n  STEP=0;KI=-1;KEXT=false;KN="";KSEZON="";showStep();\n});\n\nfunction askDel(ki){\n  document.querySelectorAll("[id^=mdd]").forEach(function(d){d.style.display="none";});\n  pendDel=ki;\n  document.getElementById("ov-title").textContent="Usun\\u0105\\u0107 \\u201e"+DB[ki].name+"\\u201d?";\n  document.getElementById("ov-msg").textContent="Klub zostanie usuni\\u0119ty. Operacji nie mo\\u017cna cofn\\u0105\\u0107.";\n  document.getElementById("ov-ok").onclick=function(){DB.splice(pendDel,1);saveDB();closeOv();draw();idle();};\n  document.getElementById("ov-del").style.display="flex";\n}\ndocument.getElementById("ov-cancel").addEventListener("click",closeOv);\nfunction closeOv(){document.getElementById("ov-del").style.display="none";}\n\nfunction doEdit(ki){\n  document.querySelectorAll("[id^=mdd]").forEach(function(d){d.style.display="none";});\n  KI=ki;var k=DB[KI];\n  var b=\'<div style="padding:14px 16px;overflow-y:auto;max-height:460px">\';\n  b+=\'<div style="font-size:14px;font-weight:600;color:#1a2b4a;margin-bottom:12px">Edycja: \'+esc(k.name)+\'</div>\';\n  b+=\'<div style="background:#f8f9fa;border-radius:9px;padding:12px;margin-bottom:12px">\';\n  b+=\'<input id="ename" value="\'+esc(k.name)+\'" style="width:100%;padding:7px 10px;border:1px solid #e0e0e0;border-radius:7px;font-size:12px;margin-bottom:8px;box-sizing:border-box">\';\n  b+=\'<div style="display:flex;gap:8px">\';\n  b+=\'<div id="ec" style="flex:1;padding:6px;text-align:center;border:1px solid \'+(k.ext?\'#e0e0e0\':\'#1a2b4a\')+\';border-radius:7px;cursor:pointer;font-size:11px;background:\'+(k.ext?\'#fff\':\'#E6F1FB\')+\';color:\'+(k.ext?\'#888\':\'#0C447C\')+\'">Klub sportowy</div>\';\n  b+=\'<div id="ee" style="flex:1;padding:6px;text-align:center;border:1px solid \'+(k.ext?\'#1a2b4a\':\'#e0e0e0\')+\';border-radius:7px;cursor:pointer;font-size:11px;background:\'+(k.ext?\'#E6F1FB\':\'#fff\')+\';color:\'+(k.ext?\'#0C447C\':\'#888\')+\'">Kadra</div>\';\n  b+=\'</div></div>\';\n  var sk=Object.keys(k.sezony||{});\n  for(var si=0;si<sk.length;si++){\n    var s=sk[si],tt=k.sezony[s]||[];\n    b+=\'<div style="margin-bottom:8px;border:1px solid #eee;border-radius:8px;overflow:hidden">\';\n    b+=\'<div style="display:flex;align-items:center;gap:6px;padding:7px 10px;background:#f8f9fa">\';\n    b+=\'<input value="\'+esc(s)+\'" id="sn\'+si+\'" style="flex:1;padding:5px 9px;border:1px solid #e0e0e0;border-radius:6px;font-size:12px">\';\n    b+=\'<button class="ds" data-s="\'+encodeURIComponent(s)+\'" style="background:none;border:none;cursor:pointer;font-size:11px;color:#A32D2D;padding:3px 6px">Usu\\u0144</button></div>\';\n    b+=\'<div style="padding:8px 10px"><div style="display:flex;flex-wrap:wrap;gap:5px;margin-bottom:6px">\';\n    for(var di=0;di<tt.length;di++){\n      b+=\'<div style="display:inline-flex;align-items:center;gap:3px;border:1px solid #eee;border-radius:14px;padding:3px 9px;font-size:11px">\'+esc(tt[di]);\n      b+=\'<button class="dd2" data-s="\'+encodeURIComponent(s)+\'" data-di="\'+di+\'" style="background:none;border:none;cursor:pointer;color:#bbb;font-size:12px;line-height:1">&#10005;</button></div>\';\n    }\n    if(!tt.length) b+=\'<span style="font-size:11px;color:#ccc">Brak dru\\u017cyn</span>\';\n    b+=\'</div><div style="display:flex;gap:5px">\';\n    b+=\'<input placeholder="+ Nowa dru\\u017cyna..." id="nd\'+si+\'" style="flex:1;padding:5px 9px;border:1px solid #e0e0e0;border-radius:6px;font-size:11px">\';\n    b+=\'<button class="ad" data-s="\'+encodeURIComponent(s)+\'" data-si="\'+si+\'" style="background:#1a2b4a;color:#fff;border:none;padding:5px 10px;border-radius:6px;font-size:11px;cursor:pointer">Dodaj</button></div></div></div>\';\n  }\n  b+=\'<div style="display:flex;gap:6px;margin-top:4px">\';\n  b+=\'<input id="nsi" placeholder="Nowy sezon, np. 2026/2027" style="flex:1;padding:6px 10px;border:1px solid #e0e0e0;border-radius:7px;font-size:12px">\';\n  b+=\'<button id="as" style="background:#EF9F27;color:#fff;border:none;padding:6px 12px;border-radius:7px;font-size:12px;cursor:pointer;white-space:nowrap">Dodaj sezon</button></div></div>\';\n  var f=\'<button id="ec2" style="flex:1;background:none;border:1px solid #ddd;color:#888;padding:8px;border-radius:8px;cursor:pointer;font-size:12px">Anuluj</button>\';\n  f+=\'<button id="es" style="flex:2;background:#EF9F27;color:#fff;border:none;padding:8px;border-radius:8px;cursor:pointer;font-size:12px;font-weight:600">Zapisz</button>\';\n  panel(b,f,null);\n  var ec=document.getElementById("ec"),ee=document.getElementById("ee");\n  if(ec) ec.addEventListener("click",function(){DB[KI].ext=false;doEdit(KI);});\n  if(ee) ee.addEventListener("click",function(){DB[KI].ext=true;doEdit(KI);});\n  document.querySelectorAll(".ds").forEach(function(b){b.addEventListener("click",function(){delete DB[KI].sezony[decodeURIComponent(this.dataset.s)];saveDB();doEdit(KI);draw();});});\n  document.querySelectorAll(".dd2").forEach(function(b){b.addEventListener("click",function(){DB[KI].sezony[decodeURIComponent(this.dataset.s)].splice(parseInt(this.dataset.di),1);saveDB();doEdit(KI);draw();});});\n  document.querySelectorAll(".ad").forEach(function(b){b.addEventListener("click",function(){var s=decodeURIComponent(this.dataset.s),si=parseInt(this.dataset.si);var v=document.getElementById("nd"+si);if(!v||!v.value.trim())return;DB[KI].sezony[s].push(v.value.trim());saveDB();doEdit(KI);draw();});});\n  var as2=document.getElementById("as");\n  if(as2) as2.addEventListener("click",function(){var v=document.getElementById("nsi");if(!v||!v.value.trim())return;if(!DB[KI].sezony[v.value.trim()])DB[KI].sezony[v.value.trim()]=[];saveDB();doEdit(KI);draw();});\n  document.getElementById("ec2").addEventListener("click",idle);\n  document.getElementById("es").addEventListener("click",function(){\n    var n=document.getElementById("ename");if(n&&n.value.trim())DB[KI].name=n.value.trim();\n    var sk2=Object.keys(DB[KI].sezony||{}),ns={};\n    for(var i=0;i<sk2.length;i++){var sn=document.getElementById("sn"+i);var k2=sn&&sn.value.trim()?sn.value.trim():sk2[i];ns[k2]=DB[KI].sezony[sk2[i]];}\n    DB[KI].sezony=ns;saveDB();draw();idle();\n  });\n}\n\nfunction goTeam(ki,s,d){\n  TEAM_KI=ki;TEAM_S=s;TEAM_D=d;\n  var k=String(ki);\n  if(!PLAYERS[k])PLAYERS[k]={};\n  if(!PLAYERS[k][s])PLAYERS[k][s]={};\n  if(!PLAYERS[k][s][d])PLAYERS[k][s][d]=[];\n  renderTeam();\n}\n\nfunction renderTeam(){\n  var k=DB[TEAM_KI],ps=PLAYERS[String(TEAM_KI)][TEAM_S][TEAM_D];\n  var h=\'<div style="padding:10px 14px;border-bottom:1px solid #f0f0f0;display:flex;align-items:center;gap:8px">\';\n  h+=\'<button id="tb" style="background:none;border:none;cursor:pointer;color:#888;font-size:12px;padding:0">&#8592; Wstecz</button>\';\n  h+=\'<div style="width:1px;height:14px;background:#e0e0e0"></div>\';\n  h+=\'<div style="flex:1"><div style="font-size:13px;font-weight:600;color:#1a2b4a">\'+esc(TEAM_D)+\'</div>\';\n  h+=\'<div style="font-size:11px;color:#888">\'+esc(k?k.name:\'\')+\'&middot;\'+esc(TEAM_S)+\'</div></div></div>\';\n  h+=\'<div style="display:flex;flex-direction:column;overflow:hidden;flex:1">\';\n  h+=\'<div style="padding:8px 10px;border-bottom:1px solid #f0f0f0;display:flex;gap:7px;align-items:center;flex-shrink:0">\';\n  h+=\'<div style="flex:1;position:relative;display:flex;align-items:center">\';\n  h+=\'<svg style="position:absolute;left:8px;width:12px;height:12px;stroke:#aaa;fill:none;stroke-width:2;pointer-events:none" viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>\';\n  h+=\'<input id="pl-q" placeholder="Szukaj..." style="width:100%;padding:5px 8px 5px 26px;border:1px solid #e0e0e0;border-radius:7px;font-size:11px;background:#f8f9fa">\';\n  h+=\'</div>\';\n  h+=\'<button id="pl-sort" style="display:flex;align-items:center;gap:3px;padding:5px 9px;border:1px solid #e0e0e0;border-radius:7px;background:#f8f9fa;color:#1a2b4a;font-size:11px;cursor:pointer;white-space:nowrap"><svg style="width:10px;height:10px;stroke:currentColor;fill:none;stroke-width:2" viewBox="0 0 24 24"><line x1="3" y1="6" x2="21" y2="6"/><line x1="3" y1="12" x2="15" y2="12"/><line x1="3" y1="18" x2="9" y2="18"/></svg><span id="pl-sl">A\\u2192Z</span></button>\';\n  h+=\'<button id="pl-af" style="padding:5px 9px;border:1px solid #e0e0e0;border-radius:7px;background:#f8f9fa;color:#666;font-size:11px;cursor:pointer;white-space:nowrap">Aktywni</button>\';\n  h+=\'<button id="apb" style="background:#EF9F27;color:#fff;border:none;padding:5px 10px;border-radius:7px;cursor:pointer;font-size:11px;font-weight:500;white-space:nowrap">+ Dodaj</button>\';\n  h+=\'</div>\';\n  h+=\'<div style="padding:3px 12px;border-bottom:1px solid #f5f5f5"><span id="pl-cnt" style="font-size:10px;color:#aaa">\'+ ps.length +\' zawodnik\\u00f3w</span></div>\';\n  h+=\'<div id="pl-list" style="overflow-y:auto;flex:1;min-height:80px"></div>\';\n  h+=\'<div style="padding:10px 14px;border-top:1px solid #f0f0f0;flex-shrink:0">\';\n  h+=\'<div id="dz" style="border:1.5px dashed #e0e0e0;border-radius:9px;padding:12px;text-align:center;cursor:pointer;margin-bottom:8px">\';\n  h+=\'<div style="font-size:12px;font-weight:500;color:#444;margin-bottom:1px">Przeciągnij plik lub kliknij</div>\';\n  h+=\'<div style="font-size:11px;color:#aaa">Obsługiwane: .xlsx, .csv</div></div>\';\n  h+=\'<div style="display:flex;gap:8px">\';\n  h+=\'<a href="/template/sklad" style="flex:1;background:none;border:1px solid #e0e0e0;color:#666;padding:5px;border-radius:7px;font-size:11px;text-align:center;text-decoration:none;display:block">↓ Szablon Excel</a>\';\n  h+=\'<a href="/template/sklad?fmt=csv" style="flex:1;background:none;border:1px solid #e0e0e0;color:#666;padding:5px;border-radius:7px;font-size:11px;text-align:center;text-decoration:none;display:block">↓ Szablon CSV</a>\';\n  h+=\'</div></div></div>\';\n  panel(h,"",null);\n  document.getElementById("tb").addEventListener("click",idle);\n  var apb=document.getElementById("apb");\n  if(apb) apb.addEventListener("click",function(){\n    EDIT_PI=-1;\n    document.getElementById("pm-title").textContent="Nowy zawodnik";\n    document.getElementById("pm-sub").textContent=TEAM_D+" \\u00b7 "+TEAM_S;\n    document.getElementById("pm-imie").value="";\n    document.getElementById("pm-nazw").value="";\n    document.getElementById("pm-num").value="";\n    document.getElementById("pm-poz").value="";\n    var akt=document.getElementById("pm-aktywny");\n    if(akt) akt.value="1";\n    document.getElementById("ov-player").style.display="flex";\n  });\n  document.querySelectorAll(".dp").forEach(function(b){b.addEventListener("click",function(e){\n    e.stopPropagation();\n    pendDelPl=parseInt(this.dataset.pi);\n    document.getElementById("ov-title").textContent="Usun\\u0105\\u0107 zawodnika?";\n    document.getElementById("ov-msg").textContent="Operacji nie mo\\u017cna cofn\\u0105\\u0107.";\n    document.getElementById("ov-ok").onclick=function(){PLAYERS[String(TEAM_KI)][TEAM_S][TEAM_D].splice(pendDelPl,1);closeOv();saveDB();renderTeam();};\n    document.getElementById("ov-del").style.display="flex";\n  });});\n  document.querySelectorAll(".ep").forEach(function(row){row.addEventListener("click",function(){\n    var pi=parseInt(this.dataset.pi);\n    var pl=PLAYERS[String(TEAM_KI)][TEAM_S][TEAM_D][pi];\n    if(!pl)return;\n    EDIT_PI=pi;\n    document.getElementById("pm-title").textContent="Edytuj zawodnika";\n    document.getElementById("pm-sub").textContent=TEAM_D+" \\u00b7 "+TEAM_S;\n    document.getElementById("pm-imie").value=pl.imie||"";\n    document.getElementById("pm-nazw").value=pl.nazwisko||"";\n    document.getElementById("pm-num").value=pl.num||"";\n    var sel=document.getElementById("pm-poz");\n    sel.value=pl.poz||"";\n    if(sel.value!==pl.poz) sel.value="";\n    var akt=document.getElementById("pm-aktywny");\n    if(akt) akt.value=(pl.aktywny===false?"0":"1");\n    document.getElementById("ov-player").style.display="flex";\n  });});\n  var dz=document.getElementById("dz");\n  if(dz){\n    dz.addEventListener("click",function(){var inp=document.createElement("input");inp.type="file";inp.accept=".xlsx,.csv";inp.addEventListener("change",function(){if(this.files[0])importF(this.files[0]);});inp.click();});\n    dz.addEventListener("dragover",function(e){e.preventDefault();this.style.borderColor="#EF9F27";this.style.background="#FAEEDA";});\n    dz.addEventListener("dragleave",function(){this.style.borderColor="#e0e0e0";this.style.background="";});\n    dz.addEventListener("drop",function(e){e.preventDefault();this.style.borderColor="#e0e0e0";this.style.background="";var f=e.dataTransfer.files[0];if(f)importF(f);});\n  }\n  // init state and render list\n  window._plSortAsc=true;\n  window._plFilterAkt=false;\n  var qEl=document.getElementById("pl-q");\n  if(qEl) qEl.addEventListener("input",renderList);\n  var sortBtn=document.getElementById("pl-sort");\n  if(sortBtn) sortBtn.addEventListener("click",plSort);\n  var afBtn=document.getElementById("pl-af");\n  if(afBtn) afBtn.addEventListener("click",plFilter);\n  renderList();\n}\n\nfunction plSort(){\n  window._plSortAsc=!window._plSortAsc;\n  var lbl=document.getElementById("pl-sl");\n  if(lbl)lbl.textContent=window._plSortAsc?"A→Z":"Z→A";\n  renderList();\n}\n\nfunction plFilter(){\n  window._plFilterAkt=!window._plFilterAkt;\n  var btn=document.getElementById("pl-af");\n  if(btn){btn.textContent=window._plFilterAkt?"Wszyscy":"Aktywni";btn.style.background=window._plFilterAkt?"#E1F5EE":"#f8f9fa";btn.style.color=window._plFilterAkt?"#0F6E56":"#666";btn.style.borderColor=window._plFilterAkt?"#5DCAA5":"#e0e0e0";}\n  renderList();\n}\n\nfunction renderList(){\n  var listEl=document.getElementById("pl-list");\n  if(!listEl)return;\n  var ps=PLAYERS[String(TEAM_KI)][TEAM_S][TEAM_D]||[];\n  var qEl=document.getElementById("pl-q");\n  var q=qEl?(qEl.value||"").trim().toLowerCase():"";\n  var filtered=ps.map(function(p,i){return{p:p,i:i};}).filter(function(x){\n    if(window._plFilterAkt&&x.p.aktywny===false)return false;\n    if(!q)return true;\n    return((x.p.imie||"")+" "+(x.p.nazwisko||"")).toLowerCase().indexOf(q)>=0;\n  });\n  filtered.sort(function(a,b){\n    var ca=((a.p.nazwisko||"")+" "+(a.p.imie||"")).toLowerCase();\n    var cb=((b.p.nazwisko||"")+" "+(b.p.imie||"")).toLowerCase();\n    return window._plSortAsc?ca.localeCompare(cb,"pl"):cb.localeCompare(ca,"pl");\n  });\n  var cnt=document.getElementById("pl-cnt");\n  if(cnt)cnt.textContent=filtered.length+" zawodników";\n  if(!filtered.length){\n    var msg=q?"Brak wynik\\u00f3w.":"Brak zawodnik\\u00f3w.";\n    listEl.innerHTML=\'<div style="text-align:center;padding:20px;color:#ccc;font-size:12px">\'+msg+\'</div>\';\n    return;\n  }\n  var h="";\n  filtered.forEach(function(x){\n    var p=x.p,pi=x.i;\n    var akt=p.aktywny!==false;\n    var full=esc(p.imie)+" "+esc(p.nazwisko);\n    if(q){\n      var re=new RegExp("("+q.replace(/[.*+?^${}()|[\\]\\\\]/g,"\\\\$&")+")","gi");\n      full=full.replace(re,"<mark style=\\"background:#FAEEDA;color:#633806;border-radius:2px;padding:0 1px\\">$1</mark>");\n    }\n    h+=\'<div class="ep" data-pi="\'+pi+\'" style="display:flex;align-items:center;gap:10px;padding:8px 2px;border-bottom:1px solid #f5f5f5;cursor:pointer">\';\n    h+=\'<div style="width:28px;height:28px;border-radius:50%;background:#1a2b4a;color:#fff;font-size:11px;font-weight:600;display:flex;align-items:center;justify-content:center;flex-shrink:0">\'+p.num+\'</div>\';\n    h+=\'<div style="flex:1"><div style="font-size:13px;font-weight:500;color:\'+(akt?"#222":"#999")+\'">\'+full+\'</div>\';\n    h+=\'<div style="display:flex;align-items:center;gap:5px"><span style="font-size:11px;color:#aaa">\'+esc(p.poz||"—")+\'</span>\';\n    h+=\'<span style="font-size:9px;font-weight:600;padding:1px 6px;border-radius:10px;background:\'+(akt?"#E1F5EE":"#f5f5f5")+\';color:\'+(akt?"#0F6E56":"#aaa")+\'">\'+(akt?"Aktywny":"Nieaktywny")+\'</span></div></div>\';\n    h+=\'<svg style="width:13px;height:13px;stroke:#ccc;fill:none;stroke-width:2;flex-shrink:0" viewBox="0 0 24 24"><path d="M11 4H4a2 2 0 0 0-2 2v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2v-7"/><path d="M18.5 2.5a2.121 2.121 0 0 1 3 3L12 15l-4 1 1-4 9.5-9.5z"/></svg>\';\n    h+=\'<button class="dp" data-pi="\'+pi+\'" style="background:none;border:none;cursor:pointer;padding:4px;color:#ddd;font-size:15px">&#128465;</button></div>\';\n  });\n  listEl.innerHTML=h;\n  // rebind ep and dp\n  listEl.querySelectorAll(".ep").forEach(function(row){row.addEventListener("click",function(){\n    var pi=parseInt(this.dataset.pi);\n    var pl=PLAYERS[String(TEAM_KI)][TEAM_S][TEAM_D][pi];\n    if(!pl)return;\n    EDIT_PI=pi;\n    document.getElementById("pm-title").textContent="Edytuj zawodnika";\n    document.getElementById("pm-sub").textContent=TEAM_D+" · "+TEAM_S;\n    document.getElementById("pm-imie").value=pl.imie||"";\n    document.getElementById("pm-nazw").value=pl.nazwisko||"";\n    document.getElementById("pm-num").value=pl.num||"";\n    var sel=document.getElementById("pm-poz");sel.value=pl.poz||"";if(sel.value!==pl.poz)sel.value="";\n    var akt=document.getElementById("pm-aktywny");if(akt)akt.value=(pl.aktywny===false?"0":"1");\n    document.getElementById("ov-player").style.display="flex";\n  });});\n  listEl.querySelectorAll(".dp").forEach(function(b){b.addEventListener("click",function(e){\n    e.stopPropagation();\n    pendDelPl=parseInt(this.dataset.pi);\n    document.getElementById("ov-title").textContent="Usunąć zawodnika?";\n    document.getElementById("ov-msg").textContent="Operacji nie można cofnąć.";\n    document.getElementById("ov-ok").onclick=function(){PLAYERS[String(TEAM_KI)][TEAM_S][TEAM_D].splice(pendDelPl,1);closeOv();saveDB();renderTeam();};\n    document.getElementById("ov-del").style.display="flex";\n  });});\n}\n\ndocument.getElementById("pm-cancel").addEventListener("click",function(){document.getElementById("ov-player").style.display="none";});\ndocument.getElementById("pm-save").addEventListener("click",function(){\n  var im=document.getElementById("pm-imie").value.trim();\n  var nz=document.getElementById("pm-nazw").value.trim();\n  if(!im||!nz){document.getElementById("pm-imie").style.borderColor="#E24B4A";return;}\n  var poz=document.getElementById("pm-poz").value;\n  var num=parseInt(document.getElementById("pm-num").value)||0;\n  var aktEl=document.getElementById("pm-aktywny");\n  var aktywny=aktEl?aktEl.value!=="0":true;\n  if(EDIT_PI>=0){\n    PLAYERS[String(TEAM_KI)][TEAM_S][TEAM_D][EDIT_PI]={imie:im,nazwisko:nz,num:num,poz:poz,aktywny:aktywny};\n  } else {\n    PLAYERS[String(TEAM_KI)][TEAM_S][TEAM_D].push({imie:im,nazwisko:nz,num:num,poz:poz,aktywny:aktywny});\n  }\n  EDIT_PI=-1;\n  document.getElementById("ov-player").style.display="none";\n  saveDB();\n  renderTeam();\n});\n\nfunction importF(file){\n  var ext=file.name.split(".").pop().toLowerCase();\n  if(ext==="csv"){\n    var r=new FileReader();r.onload=function(e){\n      var lines=e.target.result.split(/\\r?\\n/),added=0;\n      for(var i=1;i<lines.length;i++){\n        var c=lines[i].split(",");if(c.length<3)continue;\n        var im=(c[1]||"").trim(),nz=(c[2]||"").trim();if(!im&&!nz)continue;\n        if(im.toLowerCase()==="imie"||im.toLowerCase()==="imię")continue;\n        PLAYERS[String(TEAM_KI)][TEAM_S][TEAM_D].push({imie:im,nazwisko:nz,num:parseInt(c[3])||0,poz:(c[4]||"").trim()});added++;\n      }\n      alert("Wczytano "+added+" zawodnik\\u00f3w.");saveDB();renderTeam();\n    };r.readAsText(file,"UTF-8");\n  } else if(ext==="xlsx"){\n    var r=new FileReader();r.onload=function(e){\n      try{\n        var wb=XLSX.read(new Uint8Array(e.target.result),{type:"array"});\n        var rows=XLSX.utils.sheet_to_json(wb.Sheets[wb.SheetNames[0]],{header:1,defval:""});\n        var added=0;\n        for(var i=1;i<rows.length;i++){\n          var row=rows[i],im=String(row[1]||"").trim(),nz=String(row[2]||"").trim();\n          if(!im&&!nz)continue;\n          if(im.toLowerCase()==="imie"||im.toLowerCase()==="imię")continue;\n          PLAYERS[String(TEAM_KI)][TEAM_S][TEAM_D].push({imie:im,nazwisko:nz,num:parseInt(row[3])||0,poz:String(row[4]||"").trim()});added++;\n        }\n        alert("Wczytano "+added+" zawodnik\\u00f3w.");saveDB();renderTeam();\n      }catch(err){alert("B\\u0142\\u0105d: "+err.message);}\n    };r.readAsArrayBuffer(file);\n  } else {\n    alert("Obs\\u0142ugiwane: .xlsx, .csv");\n  }\n}\n\ndraw();\n})();\n'


@app.route("/court-preview")
def court_preview():
    """Podgląd boiska shooting chart (bez autoryzacji — tylko dev)."""
    _base = os.path.dirname(os.path.abspath(__file__))
    _fp = os.path.join(_base, "static", "court_edit.html")
    if os.path.exists(_fp):
        with open(_fp, encoding="utf-8") as _f:
            return _f.read(), 200, {"Content-Type": "text/html; charset=utf-8"}
    return "Brak pliku court_edit.html", 404

@app.route("/static/img/<path:filename>")
def serve_static_img(filename):
    """Serwuje obrazki statyczne z folderu static/img (fallback gdy /app/static nie jest zamontowany)."""
    import mimetypes
    _allowed = {"court_attack.png", "court_defense.png", "gtk_logo.png", "app_logo.png"}
    if filename not in _allowed:
        from flask import abort
        abort(404)
    _base = os.path.dirname(os.path.abspath(__file__))
    _fp = os.path.join(_base, "static", "img", filename)
    if os.path.exists(_fp):
        from flask import send_file as _sf
        mt, _ = mimetypes.guess_type(_fp)
        return _sf(_fp, mimetype=mt or "image/png")
    # Fallback: sprawdź względem CWD
    _fp2 = os.path.join(os.getcwd(), "static", "img", filename)
    if os.path.exists(_fp2):
        from flask import send_file as _sf2
        return _sf2(_fp2, mimetype="image/png")
    from flask import abort
    abort(404)


@app.route("/start")
def landing():
    return render_template_string(LANDING_HTML)


@app.route("/player")
def player_dashboard():
    return redirect(url_for("portal"))


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        email    = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = USERS.get(email)
        if user and user["password_hash"] == hashlib.sha256(password.encode()).hexdigest():
            session.permanent = True
            session["logged_in"] = True
            session["user_name"] = user["name"]
            session["user_email"] = email
            next_url = request.args.get("next") or url_for("index")
            return redirect(next_url)
        error = "Nieprawidłowy login lub hasło."
    return render_template_string(LOGIN_HTML, error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


LANDING_HTML = """<!DOCTYPE html>
<html lang="pl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Basket Kołcz — Wybierz panel</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #f0f2f7;
    min-height: 100vh;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    padding: 2rem 1rem;
  }
  .logo { text-align: center; margin-bottom: 2.5rem; }
  .logo-icon { font-size: 2.8rem; line-height: 1; margin-bottom: .5rem; }
  .logo-name { font-size: 1.8rem; font-weight: 800; color: #1a2b4a; letter-spacing: -.5px; }
  .logo-name span { color: #EF9F27; }
  .logo-sub { font-size: .82rem; color: #888; margin-top: 4px; }
  .cards {
    display: flex;
    gap: 1.5rem;
    flex-wrap: wrap;
    justify-content: center;
    width: 100%;
    max-width: 700px;
  }
  .card {
    flex: 1 1 280px;
    max-width: 320px;
    background: #fff;
    border-radius: 18px;
    padding: 2.4rem 2rem 2rem;
    text-align: center;
    box-shadow: 0 4px 28px rgba(0,0,0,.08);
    cursor: pointer;
    text-decoration: none;
    transition: transform .18s, box-shadow .18s;
    border: 2px solid transparent;
    display: block;
    position: relative;
    overflow: hidden;
  }
  .card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 5px;
  }
  .card.coach::before { background: #1a2b4a; }
  .card.player::before { background: #EF9F27; }
  .card:hover { transform: translateY(-6px); box-shadow: 0 14px 40px rgba(0,0,0,.13); }
  .card.coach:hover { border-color: #1a2b4a; }
  .card.player:hover { border-color: #EF9F27; }
  .card-icon { font-size: 3.2rem; margin-bottom: 1rem; line-height: 1; }
  .card-title { font-size: 1.25rem; font-weight: 700; margin-bottom: .5rem; }
  .card.coach .card-title { color: #1a2b4a; }
  .card.player .card-title { color: #EF9F27; }
  .card-desc { font-size: .82rem; color: #888; line-height: 1.6; margin-bottom: 1.4rem; }
  .card-btn {
    display: inline-block;
    padding: 9px 28px;
    border-radius: 30px;
    font-size: .85rem;
    font-weight: 600;
    text-decoration: none;
    transition: opacity .15s;
  }
  .card-btn:hover { opacity: .85; }
  .card.coach .card-btn { background: #1a2b4a; color: #fff; }
  .card.player .card-btn { background: #EF9F27; color: #fff; }
  .badge-soon {
    display: inline-block;
    background: #fff4e5;
    color: #EF9F27;
    font-size: .7rem;
    font-weight: 700;
    padding: 3px 12px;
    border-radius: 20px;
    letter-spacing: .5px;
    text-transform: uppercase;
  }
  .footer { margin-top: 2.5rem; font-size: .75rem; color: #bbb; }
</style>
</head>
<body>
<div class="logo">
  <div class="logo-icon">🏀</div>
  <div class="logo-name">Basket <span>Kołcz</span></div>
  <div class="logo-sub">Analytics Platform — wybierz swój panel</div>
</div>

<div class="cards">
  <div class="card coach">
    <div class="card-icon">📋</div>
    <div class="card-title">Pulpit Trenera</div>
    <div class="card-desc">Analizy meczów, statystyki zawodników, raporty taktyczne i zarządzanie sezonem.</div>
    <a href="/login" class="card-btn">Zaloguj się</a>
  </div>
  <a href="/portal" class="card player" style="text-decoration:none">
    <div class="card-icon">🏃</div>
    <div class="card-title">Pulpit Zawodnika</div>
    <div class="card-desc">Twoje statystyki, postępy drużyny i statystyki indywidualne sezonu.</div>
    <span class="card-btn" style="background:#EF9F27;color:#fff">Wejdź</span>
  </a>
</div>

<div class="footer">© 2025 made by Kosma Kołcz</div>
</body>
</html>"""

PLAYER_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Player Dashboard — Under Construction</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #f0f2f5;
    min-height: 100vh;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    padding: 2rem 1rem;
    text-align: center;
  }
  .hard-hat { font-size: 7rem; display: block; margin: 0 auto 1rem; animation: wiggle 2s infinite; }
  @keyframes wiggle { 0%,100%{transform:rotate(-5deg)} 50%{transform:rotate(5deg)} }
  h1 { font-size: 2rem; font-weight: 800; color: #1a2b4a; margin-bottom: .5rem; }
  .sub { font-size: 1rem; color: #888; margin-bottom: 1.8rem; max-width: 400px; line-height: 1.6; }
  .workers { font-size: 2.5rem; margin-bottom: .8rem; letter-spacing: .3rem; }
  .tag {
    display: inline-block; background: #fff4e5; color: #f97316;
    font-weight: 700; font-size: .75rem; padding: 3px 14px;
    border-radius: 20px; letter-spacing: .5px; margin-bottom: 1.8rem;
  }
  .progress-bar { width: 280px; height: 10px; background: #e5e7eb; border-radius: 99px; overflow: hidden; margin: 0 auto 1rem; }
  .progress-fill { height: 100%; width: 15%; background: linear-gradient(90deg,#f97316,#fbbf24); border-radius: 99px; animation: load 3s infinite; }
  @keyframes load { 0%{width:5%} 50%{width:60%} 100%{width:5%} }
  .note { font-size: .75rem; color: #bbb; margin-bottom: 1.5rem; }
  .back { display: inline-block; margin-top: .5rem; padding: .55rem 1.4rem; background: #1a2b4a; color: #fff; border-radius: 8px; text-decoration: none; font-size: .88rem; font-weight: 600; transition: background .15s; }
  .back:hover { background: #263f6a; }
</style>
</head>
<body>
<span class="hard-hat">🏗️</span>
<div class="workers">👷 👷‍♀️ 🦺</div>
<h1>Under Construction</h1>
<div class="sub">
  Our best engineers (and a few basketballs) are working hard on the Player Dashboard.<br>
  <strong>It's not ready yet — but trust us, it'll be worth the wait.</strong>
</div>
<div class="tag">🚧 &nbsp;WORK IN PROGRESS &nbsp;🚧</div>
<div class="progress-bar"><div class="progress-fill"></div></div>
<div class="note">Loading awesomeness... please hold</div>
<a href="/" class="back">← Back to Home</a>
</body>
</html>"""

LOGIN_HTML = """<!DOCTYPE html>
<html lang="pl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Logowanie — Basket Kołcz</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
         background: #f0f2f5; display: flex; align-items: center;
         justify-content: center; min-height: 100vh; }
  .card { background: #fff; border-radius: 12px; padding: 2.5rem 2rem;
          width: 100%; max-width: 380px; box-shadow: 0 4px 24px rgba(0,0,0,.08); }
  .logo { text-align: center; margin-bottom: 1.5rem; }
  .logo-dot { color: #f97316; font-size: 1.5rem; }
  .logo-name { font-size: 1.4rem; font-weight: 700; color: #1a2b4a; }
  .logo-sub { font-size: .78rem; color: #888; margin-top: 2px; }
  label { display: block; font-size: .82rem; font-weight: 500;
          color: #444; margin-bottom: 4px; margin-top: 1rem; }
  input { width: 100%; padding: .55rem .75rem; border: 1px solid #d1d5db;
          border-radius: 6px; font-size: .9rem; outline: none; transition: border .15s; }
  input:focus { border-color: #1a2b4a; }
  .btn { width: 100%; margin-top: 1.5rem; padding: .65rem;
         background: #1a2b4a; color: #fff; border: none; border-radius: 6px;
         font-size: .95rem; font-weight: 600; cursor: pointer; transition: background .15s; }
  .btn:hover { background: #263f6a; }
  .error { background: #fff0f0; color: #c0392b; border: 1px solid #f5c6cb;
           border-radius: 6px; padding: .6rem .8rem; font-size: .82rem; margin-top: 1rem; }
</style>
</head>
<body>
<div class="card">
  <div class="logo">
    <div><span class="logo-dot">●</span> <span class="logo-name">Basket Kołcz</span></div>
    <div class="logo-sub">Analytics Platform</div>
  </div>
  <form method="POST">
    <label for="email">Email</label>
    <input type="email" id="email" name="email" placeholder="twoj@email.com"
           autocomplete="username" required>
    <label for="password">Hasło</label>
    <input type="password" id="password" name="password" placeholder="••••••••"
           autocomplete="current-password" required>
    {% if error %}
    <div class="error">{{ error }}</div>
    {% endif %}
    <button class="btn" type="submit">Zaloguj się</button>
  </form>
</div>
</body>
</html>
"""

DATABASE_URL = os.environ.get("DATABASE_URL", "")

# ══════════════════════════════════════════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════════════════════════════════════════

def get_db():
    if "db" not in g:
        try:
            g.db = psycopg2.connect(
                DATABASE_URL,
                cursor_factory=RealDictCursor,
                connect_timeout=10,
                options="-c statement_timeout=25000"
            )
            g.db.autocommit = False
        except psycopg2.OperationalError as e:
            raise RuntimeError(f"Nie mozna polaczyc sie z baza danych: {e}")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db:
        try: db.rollback()
        except: pass
        try: db.close()
        except: pass

def init_db():
    db = get_db()
    cur = db.cursor()
    # Tabele główne
    cur.execute("""
    CREATE TABLE IF NOT EXISTS matches (
        id SERIAL PRIMARY KEY,
        sezon VARCHAR(20) NOT NULL DEFAULT '',
        data_meczu DATE,
        przeciwnik VARCHAR(100) NOT NULL,
        nazwa_gtk VARCHAR(100) DEFAULT '',
        rozgrywki VARCHAR(100) DEFAULT '',
        runda VARCHAR(50) DEFAULT '',
        kolejka VARCHAR(50) DEFAULT '',
        miejsce VARCHAR(20) DEFAULT '',
        wynik_gtk INTEGER DEFAULT 0,
        wynik_opp INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS match_stats (
        id SERIAL PRIMARY KEY,
        match_id INTEGER REFERENCES matches(id) ON DELETE CASCADE,
        druzyna VARCHAR(10) NOT NULL,
        kwarta INTEGER NOT NULL,
        pts INTEGER DEFAULT 0, poss INTEGER DEFAULT 0,
        p2m INTEGER DEFAULT 0, p2a INTEGER DEFAULT 0,
        p3m INTEGER DEFAULT 0, p3a INTEGER DEFAULT 0,
        ftm INTEGER DEFAULT 0, fta INTEGER DEFAULT 0,
        br INTEGER DEFAULT 0, fd INTEGER DEFAULT 0,
        ast INTEGER DEFAULT 0, oreb INTEGER DEFAULT 0, dreb INTEGER DEFAULT 0,
        stl INTEGER DEFAULT 0, blk INTEGER DEFAULT 0,
        d2m INTEGER DEFAULT 0, d2a INTEGER DEFAULT 0, przerw INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS player_stats (
        id SERIAL PRIMARY KEY,
        match_id INTEGER REFERENCES matches(id) ON DELETE CASCADE,
        druzyna VARCHAR(10) NOT NULL,
        nr INTEGER NOT NULL,
        pts INTEGER DEFAULT 0, p2m INTEGER DEFAULT 0, p2a INTEGER DEFAULT 0,
        p3m INTEGER DEFAULT 0, p3a INTEGER DEFAULT 0,
        ftm INTEGER DEFAULT 0, fta INTEGER DEFAULT 0,
        ast INTEGER DEFAULT 0, oreb INTEGER DEFAULT 0, dreb INTEGER DEFAULT 0,
        br INTEGER DEFAULT 0, fd INTEGER DEFAULT 0, finishes INTEGER DEFAULT 0,
        stl INTEGER DEFAULT 0, blk INTEGER DEFAULT 0,
        time_sum REAL DEFAULT 0, time_cnt INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS timing_stats (
        id SERIAL PRIMARY KEY,
        match_id INTEGER REFERENCES matches(id) ON DELETE CASCADE,
        druzyna VARCHAR(10) NOT NULL, bucket VARCHAR(10) NOT NULL,
        made2 INTEGER DEFAULT 0, att2 INTEGER DEFAULT 0,
        made3 INTEGER DEFAULT 0, att3 INTEGER DEFAULT 0,
        br INTEGER DEFAULT 0, ftm INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS roster (
        id SERIAL PRIMARY KEY,
        imie VARCHAR(50) NOT NULL,
        nazwisko VARCHAR(50) NOT NULL DEFAULT '',
        pseudonim VARCHAR(30) DEFAULT '',
        aktywny BOOLEAN DEFAULT TRUE,
        created_at TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS player_aliases (
        id SERIAL PRIMARY KEY,
        roster_id INTEGER REFERENCES roster(id) ON DELETE CASCADE,
        nr INTEGER NOT NULL,
        sezon VARCHAR(20) DEFAULT '',
        UNIQUE(roster_id, nr, sezon)
    );
    CREATE TABLE IF NOT EXISTS settings (
        key VARCHAR(50) PRIMARY KEY,
        value TEXT
    );
    CREATE TABLE IF NOT EXISTS lineup_stats (
        id SERIAL PRIMARY KEY,
        match_id INTEGER REFERENCES matches(id) ON DELETE CASCADE,
        druzyna VARCHAR(10) NOT NULL,
        lineup VARCHAR(30) NOT NULL,
        pts INTEGER DEFAULT 0,
        poss INTEGER DEFAULT 0,
        p2m INTEGER DEFAULT 0, p2a INTEGER DEFAULT 0,
        p3m INTEGER DEFAULT 0, p3a INTEGER DEFAULT 0,
        ftm INTEGER DEFAULT 0, fta INTEGER DEFAULT 0,
        br INTEGER DEFAULT 0, fd INTEGER DEFAULT 0,
        ast INTEGER DEFAULT 0, oreb INTEGER DEFAULT 0,
        dreb INTEGER DEFAULT 0, stl INTEGER DEFAULT 0, blk INTEGER DEFAULT 0
    );
    INSERT INTO settings (key, value) VALUES ('gtk_name', 'GTK') ON CONFLICT DO NOTHING;
    INSERT INTO settings (key, value) VALUES ('current_season', '') ON CONFLICT DO NOTHING;
    INSERT INTO settings (key, value) VALUES ('current_klub', '') ON CONFLICT DO NOTHING;
    INSERT INTO settings (key, value) VALUES ('current_druzyna', '') ON CONFLICT DO NOTHING;
    INSERT INTO settings (key, value) VALUES ('kluby_json', '[]') ON CONFLICT DO NOTHING;
    """)
    db.commit()

    # ALTER TABLE — każdy osobno z savepoint
    alters = [
        "ALTER TABLE settings ALTER COLUMN value TYPE TEXT",
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS file_path TEXT DEFAULT NULL",
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS parser_version INTEGER DEFAULT 0",
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS team_id INTEGER REFERENCES teams(id) ON DELETE SET NULL",
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS nazwa_gtk VARCHAR(100) DEFAULT ''",
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS rozgrywki VARCHAR(100) DEFAULT ''",
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS runda VARCHAR(50) DEFAULT ''",
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS kolejka VARCHAR(50) DEFAULT ''",
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS miejsce VARCHAR(20) DEFAULT ''",
        "ALTER TABLE player_stats ADD COLUMN IF NOT EXISTS roster_id INTEGER REFERENCES roster(id) ON DELETE SET NULL",
        "ALTER TABLE player_stats ADD COLUMN IF NOT EXISTS stl INTEGER DEFAULT 0",
        "ALTER TABLE player_stats ADD COLUMN IF NOT EXISTS blk INTEGER DEFAULT 0",
        "ALTER TABLE match_stats ADD COLUMN IF NOT EXISTS ast INTEGER DEFAULT 0",
        "ALTER TABLE match_stats ADD COLUMN IF NOT EXISTS oreb INTEGER DEFAULT 0",
        "ALTER TABLE match_stats ADD COLUMN IF NOT EXISTS dreb INTEGER DEFAULT 0",
        "ALTER TABLE match_stats ADD COLUMN IF NOT EXISTS stl INTEGER DEFAULT 0",
        "ALTER TABLE match_stats ADD COLUMN IF NOT EXISTS blk INTEGER DEFAULT 0",
        "ALTER TABLE match_stats ADD COLUMN IF NOT EXISTS d2m INTEGER DEFAULT 0",
        "ALTER TABLE match_stats ADD COLUMN IF NOT EXISTS d2a INTEGER DEFAULT 0",
        "ALTER TABLE match_stats ADD COLUMN IF NOT EXISTS przerw INTEGER DEFAULT 0",
        "ALTER TABLE timing_stats ADD COLUMN IF NOT EXISTS kwarta INTEGER DEFAULT 0",
        "ALTER TABLE timing_stats ADD COLUMN IF NOT EXISTS br INTEGER DEFAULT 0",
        "ALTER TABLE timing_stats ADD COLUMN IF NOT EXISTS ftm INTEGER DEFAULT 0",
        "ALTER TABLE timing_stats ADD COLUMN IF NOT EXISTS poss_ft INTEGER DEFAULT 0",
        "ALTER TABLE player_stats ADD COLUMN IF NOT EXISTS time_sum REAL DEFAULT 0",
        "ALTER TABLE player_stats ADD COLUMN IF NOT EXISTS time_cnt INTEGER DEFAULT 0",
        """CREATE TABLE IF NOT EXISTS score_flow (
            id SERIAL PRIMARY KEY,
            match_id INTEGER REFERENCES matches(id) ON DELETE CASCADE,
            kwarta INTEGER NOT NULL,
            czas_sek REAL NOT NULL,
            pts_gtk INTEGER NOT NULL,
            pts_opp INTEGER NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS clubs (
            id SERIAL PRIMARY KEY,
            name VARCHAR(100) NOT NULL UNIQUE,
            ext BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS seasons (
            id SERIAL PRIMARY KEY,
            club_id INTEGER REFERENCES clubs(id) ON DELETE CASCADE,
            name VARCHAR(50) NOT NULL,
            UNIQUE(club_id, name)
        )""",
        """CREATE TABLE IF NOT EXISTS teams (
            id SERIAL PRIMARY KEY,
            season_id INTEGER REFERENCES seasons(id) ON DELETE CASCADE,
            name VARCHAR(100) NOT NULL,
            UNIQUE(season_id, name)
        )""",
        """CREATE TABLE IF NOT EXISTS players (
            id SERIAL PRIMARY KEY,
            team_id INTEGER REFERENCES teams(id) ON DELETE CASCADE,
            imie VARCHAR(50) NOT NULL,
            nazwisko VARCHAR(50) NOT NULL DEFAULT '',
            numer INTEGER DEFAULT 0,
            pozycja VARCHAR(30) DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS player_match_map (
            id SERIAL PRIMARY KEY,
            match_id INTEGER REFERENCES matches(id) ON DELETE CASCADE,
            team_id INTEGER REFERENCES teams(id) ON DELETE CASCADE,
            nr INTEGER NOT NULL,
            player_id INTEGER REFERENCES players(id) ON DELETE SET NULL,
            UNIQUE(match_id, nr)
        )""",
        "ALTER TABLE player_stats ADD COLUMN IF NOT EXISTS player_id INTEGER REFERENCES players(id) ON DELETE SET NULL",
        """ALTER TABLE matches ADD COLUMN IF NOT EXISTS team_id INTEGER REFERENCES teams(id) ON DELETE SET NULL""",
        """ALTER TABLE matches ADD COLUMN IF NOT EXISTS team_name_a VARCHAR(100) DEFAULT ''""",
        """ALTER TABLE matches ADD COLUMN IF NOT EXISTS team_name_b VARCHAR(100) DEFAULT ''""",
        "ALTER TABLE lineup_stats ADD COLUMN IF NOT EXISTS ast   INTEGER DEFAULT 0",
        "ALTER TABLE lineup_stats ADD COLUMN IF NOT EXISTS oreb  INTEGER DEFAULT 0",
        "ALTER TABLE lineup_stats ADD COLUMN IF NOT EXISTS dreb  INTEGER DEFAULT 0",
        "ALTER TABLE lineup_stats ADD COLUMN IF NOT EXISTS stl   INTEGER DEFAULT 0",
        "ALTER TABLE lineup_stats ADD COLUMN IF NOT EXISTS blk   INTEGER DEFAULT 0",
        """CREATE TABLE IF NOT EXISTS shot_zones (
            id SERIAL PRIMARY KEY,
            match_id INTEGER REFERENCES matches(id) ON DELETE CASCADE,
            druzyna VARCHAR(10) NOT NULL,
            nr INTEGER NOT NULL,
            zone INTEGER NOT NULL,
            made INTEGER DEFAULT 0,
            att INTEGER DEFAULT 0
        )""",
    ]
    for sql in alters:
        try:
            cur.execute("SAVEPOINT sp")
            cur.execute(sql)
            cur.execute("RELEASE SAVEPOINT sp")
            db.commit()
        except Exception:
            cur.execute("ROLLBACK TO SAVEPOINT sp")
            db.commit()
    cur.close()


def get_setting(key):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT value FROM settings WHERE key=%s", (key,))
    row = cur.fetchone()
    cur.close()
    return row['value'] if row else None

def set_setting(key, value):
    db = get_db()
    cur = db.cursor()
    cur.execute("INSERT INTO settings (key,value) VALUES (%s,%s) ON CONFLICT (key) DO UPDATE SET value=%s",
                (key, value, value))
    db.commit()
    cur.close()

def get_portal_context():
    """Zwraca kontekst portalu — wyłącznie portal_* settings, bez fallbacku na current_*."""
    klub  = get_setting("portal_klub")    or ""
    sezon = get_setting("portal_sezon")   or ""
    druz  = get_setting("portal_druzyna") or ""
    return klub, sezon, druz

# ══════════════════════════════════════════════════════════════════════════════
# PARSER (identyczny jak w app_v2.py)
# ══════════════════════════════════════════════════════════════════════════════

ACTION_2PM = {"2","2+1","2+0","2+1/1W","2+0/1W","2D","2D+1","2D+0/1W","2D+1/1W"}
ACTION_3PM = {"3","3+1","3+0","3+1/1W","3+0/1W"}
ACTION_BR  = {"BR"}
ACTION_F   = {
    "F",
    "2+1", "2+0", "3+1", "3+0",
    "2D+1", "2D+0/1W", "2D+1/1W",
    "0D+0/2W", "0D+1/2W", "0D+2/2W",
    "1/2W", "2/2W", "0/2W",
    "1/3W", "2/3W", "3/3W", "0/3W",
    "1/1W", "0/1W",
    "1/2WL", "2/2WL", "0/2WL",
}
ACTION_STL = {"STL"}
ACTION_BLK = {"BLK"}
TEAM_TOKEN = "$"  # znak drużynowy — akcja bez przypisania do zawodnika
BUCKETS    = ["0s","1-4s","5-8s","9-12s","13-16s","17-20s","21-24s"]

def extract_ft(code):
    # Format X/YW — np. 1/2W, 2/2W, 0/3W
    m = re.match(r'^(\d+)/(\d+)W', code)
    if m: return int(m.group(1)), int(m.group(2))
    m2 = re.search(r'(\d+)/(\d+)W', code)
    if m2: return int(m2.group(1)), int(m2.group(2))
    # Plus-one: 2+1, 3+1, 2D+1, 2+1/1W, 3+1/1W → celny 1/1 RW
    if re.search(r'\+1(/1W)?$', code): return 1, 1
    # Plus-zero: 2+0, 3+0, 2D+0, 2+0/1W, 3+0/1W → niecelny 0/1 RW
    if re.search(r'\+0(/1W)?$', code): return 0, 1
    return 0, 0

def time_bucket(t):
    if t == 0:    return "0s"
    if t <= 4:    return "1-4s"
    if t <= 8:    return "5-8s"
    if t <= 12:   return "9-12s"
    if t <= 16:   return "13-16s"
    if t <= 20:   return "17-20s"
    return "21-24s"

def parse_team_sheet(ws, sheet_type='A'):
    """
    sheet_type='A': arkusz ataku wlasnej druzyny (tej wybranej przy wgrywaniu)
      - L: jedna wartosc — cyfra=zawodnik wlasny+druzyna, $=tylko druzyna (AST)
      - M: wiele wartosci (;) — cyfra=zawodnik wlasny+druzyna, $=tylko druzyna (OREB)
      - O: $ = przechwyt druzyny B (rywala) — zliczany w opp_quarter[kwarta]["stl"]
      - N, P: ignorowane (dotycza rywala, nie tej druzyny)

    sheet_type='B': arkusz ataku rywala / obrony wlasnej druzyny
      - N: jedna wartosc — cyfra=zawodnik wlasny+druzyna, $=tylko druzyna (DREB)
      - O: jedna wartosc — cyfra=zawodnik wlasny+druzyna, $=tylko druzyna (STL)
      - P: jedna wartosc — cyfra=zawodnik wlasny+druzyna, $=tylko druzyna (BLK)
      - L, M: ignorowane (S = atak rywala, nie wlasnej druzyny)
      - E-I: piatka wlasnej druzyny na boisku podczas obrony
      - K: bez znaczenia dla zaleznosci B;C;D
    """
    stats = {
        "quarter": defaultdict(lambda: {"ftm":0,"fta":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"br":0,"fd":0,"poss":0,"pts":0,"d2m":0,"d2a":0,"przerw":0,"stl":0,"blk":0,"oreb":0,"dreb":0,"ast":0}),
        "players": defaultdict(lambda: {"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"fd":0,"br":0,"finishes":0,"ast":0,"oreb":0,"dreb":0,"stl":0,"blk":0,"time_sum":0,"time_cnt":0}),
        "timing":  {q: {b: {"2PT":{"made":0,"miss":0},"3PT":{"made":0,"miss":0},"br":0,"ftm":0,"poss_ft":0} for b in BUCKETS} for q in [0,1,2,3,4]},
        "lineups": defaultdict(lambda: {"pts":0,"poss":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"br":0,"fd":0,"stl":0,"blk":0,"oreb":0,"dreb":0,"ast":0}),
        "flow":    [],
        # zones[nr][zone] = {"made": int, "att": int}
        "zones":   defaultdict(lambda: defaultdict(lambda: {"made": 0, "att": 0})),
        # opp_quarter: STL druzyny B zliczane z kol O arkusza A ($ = przechwyt rywala)
        "opp_quarter": defaultdict(lambda: {"stl": 0}),
    }
    current_q = 1
    current_lineup = []

    def _nr(v):
        try: return int(str(v).strip().lstrip("#"))
        except: return None

    def _tok(raw):
        """Jedna lub wiele wartosci oddzielonych ;"""
        return [x.strip() for x in str(raw).split(";") if str(x).strip()] if raw else []

    def _assign_own(val_list, q_key, stats_key, q, lineup):
        """Przypisz do zawodnika WLASNEJ druzyny i druzyny. Uzyj dla arkusza A: L,M; arkusza B: N,O,P."""
        for v in val_list:
            if v == TEAM_TOKEN:
                q[q_key] += 1
            else:
                nr = _nr(v)
                if nr is not None:
                    stats["players"][nr][stats_key] += 1
                    q[q_key] += 1
        if any(v == TEAM_TOKEN or _nr(v) is not None for v in val_list):
            if len(lineup) == 5:
                lk = "-".join(str(x) for x in sorted(lineup))
                stats["lineups"][lk][q_key] = stats["lineups"][lk].get(q_key, 0) + 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row[:4]): continue

        # Kwarta
        if row[0] is not None:
            try: current_q = int(str(row[0]).replace("*","").strip())
            except: current_q = 1

        # Skład (kol E-I) — w obu arkuszach
        for i in range(4, 9):
            val = row[i] if len(row) > i and row[i] is not None else None
            if val is not None:
                s = str(val).strip()
                if ";" in s:
                    parts = s.split(";")
                    try:
                        old_p, new_p = int(parts[0].strip()), int(parts[1].strip())
                        try: current_lineup[current_lineup.index(old_p)] = new_p
                        except ValueError: current_lineup.append(new_p)
                    except: pass
                else:
                    try:
                        p = int(s)
                        if p not in current_lineup: current_lineup.append(p)
                    except: pass
        if len(current_lineup) > 5: current_lineup = current_lineup[-5:]

        # Kolumny tekstowe
        raw_b = str(row[1])  if len(row)>1  and row[1]  is not None else ""
        raw_c = str(row[2])  if row[2]  is not None else ""
        raw_d = str(row[3])  if len(row)>3  and row[3]  is not None else ""
        raw_k = str(row[10]) if len(row)>10 and row[10] is not None else ""
        raw_l = str(row[11]) if len(row)>11 and row[11] is not None else ""
        raw_m = str(row[12]) if len(row)>12 and row[12] is not None else ""
        raw_n = str(row[13]) if len(row)>13 and row[13] is not None else ""
        raw_o = str(row[14]) if len(row)>14 and row[14] is not None else ""
        raw_p = str(row[15]) if len(row)>15 and row[15] is not None else ""

        codes     = [c.strip().upper() for c in raw_c.split(";") if c.strip()]
        times     = [t.strip() for t in raw_b.replace(";",",").split(",") if t.strip()]
        # K powiazane z B;C tylko w arkuszu A
        finishers = [f.strip() for f in raw_k.split(";") if f.strip()] if sheet_type == 'A' else []

        q = stats["quarter"][current_q]
        q["poss"] += 1

        # ── Kolumny per-posiadanie wg sheet_type ─────────────────────────
        if sheet_type == 'A':
            # L: AST — jedna wartosc, cyfra/$ → zawodnik wlasny / druzyna wlasna
            _assign_own(_tok(raw_l)[:1], "ast",  "ast",  q, current_lineup)
            # M: OREB — wiele wartosci, cyfra/$ → zawodnik wlasny / druzyna wlasna
            _assign_own(_tok(raw_m),     "oreb", "oreb", q, current_lineup)
            # O: $ = przechwyt druzyny B (rywala) — zliczamy w opp_quarter
            if TEAM_TOKEN in _tok(raw_o):
                stats["opp_quarter"][current_q]["stl"] += 1
            # N, P: ignorowane (dotycza rywala)
        else:  # sheet_type == 'B'
            # L: AST rywala — kazdy $ lub cyfra = +1 asysta dla druzyny B
            for _v in _tok(raw_l):
                q["ast"] += 1
            # M: OREB rywala — kazdy $ lub cyfra = +1 zbiorka ofensywna druzyny B
            for _v in _tok(raw_m):
                q["oreb"] += 1
            # N: DREB — cyfra/$ → zawodnik wlasny / druzyna wlasna
            _assign_own(_tok(raw_n)[:1], "dreb", "dreb", q, current_lineup)
            # O: STL — cyfra/$ → zawodnik wlasny / druzyna wlasna
            _assign_own(_tok(raw_o)[:1], "stl",  "stl",  q, current_lineup)
            # P: BLK — cyfra/$ → zawodnik wlasny / druzyna wlasna
            _assign_own(_tok(raw_p)[:1], "blk",  "blk",  q, current_lineup)

        # ── Czas gry: kol B per wiersz, niezależnie od kol C ────────────
        _row_t = 0
        for _ts in raw_b.replace(";", ",").split(","):
            try: _row_t += float(_ts.strip())
            except: pass
        if _row_t > 0 and current_lineup:
            for _p_nr in current_lineup:
                stats["players"][_p_nr]["time_sum"] += _row_t
                stats["players"][_p_nr]["time_cnt"] += 1

        # ── Per-akcja: kody z kol C ───────────────────────────────────────
        for ai, code in enumerate(codes):
            t_val = 0
            if ai < len(times):
                try: t_val = float(times[ai])
                except: pass
            bucket = time_bucket(t_val)

            finisher = _nr(finishers[ai]) if ai < len(finishers) else None

            pts = 0

            # FD — niezależnie od innych bloków
            if code in ACTION_F:
                q["fd"] += 1
                if finisher is not None:
                    stats["players"][finisher]["fd"] += 1

            if code in ACTION_2PM:
                q["p2m"]+=1; q["p2a"]+=1; pts=2
                if code in ("2D","2D+1","2D+0/1W","2D+1/1W"): q["d2m"]+=1; q["d2a"]+=1
                stats["timing"][current_q][bucket]["2PT"]["made"]+=1
                if finisher is not None:
                    stats["players"][finisher]["p2m"]+=1; stats["players"][finisher]["p2a"]+=1
                    stats["players"][finisher]["finishes"]+=1
                    # time_sum zbierany dla całego składu powyżej

            elif code in ("0/2","0/2D","0D+0/2W","0D+1/2W","0D+2/2W"):
                q["p2a"]+=1
                if "D" in code: q["d2a"]+=1
                stats["timing"][current_q][bucket]["2PT"]["miss"]+=1
                if finisher is not None:
                    stats["players"][finisher]["p2a"]+=1; stats["players"][finisher]["finishes"]+=1
                    # time_sum zbierany dla całego składu powyżej

            elif code in ACTION_3PM:
                q["p3m"]+=1; q["p3a"]+=1; pts=3
                stats["timing"][current_q][bucket]["3PT"]["made"]+=1
                if finisher is not None:
                    stats["players"][finisher]["p3m"]+=1; stats["players"][finisher]["p3a"]+=1
                    stats["players"][finisher]["finishes"]+=1
                    # time_sum zbierany dla całego składu powyżej

            elif code == "0/3":
                q["p3a"]+=1
                stats["timing"][current_q][bucket]["3PT"]["miss"]+=1
                if finisher is not None:
                    stats["players"][finisher]["p3a"]+=1; stats["players"][finisher]["finishes"]+=1
                    # time_sum zbierany dla całego składu powyżej

            elif code in ACTION_BR:
                q["br"]+=1
                stats["timing"][current_q][bucket]["br"]+=1
                if finisher is not None: stats["players"][finisher]["br"]+=1

            elif code in ACTION_STL:
                pass  # STL z kol O (sheet_type B) — obsługiwane per-posiadanie przez _assign_own

            elif code in ACTION_BLK:
                pass  # BLK z kol P (sheet_type B) — obsługiwane per-posiadanie przez _assign_own

            elif code == "P":
                q["przerw"]+=1

            # Strefa rzutu (kolumna D) — tylko arkusz A, kody trafionych/pudłowanych rzutów z pola
            if sheet_type == 'A' and raw_d:
                try:
                    zones_raw = [z.strip() for z in raw_d.split(";") if z.strip()]
                    zone_val = int(zones_raw[ai]) if ai < len(zones_raw) else None
                except (ValueError, IndexError):
                    zone_val = None
                if zone_val is not None and 1 <= zone_val <= 28:
                    is_shot_made = code in ACTION_2PM or code in ACTION_3PM
                    is_shot_att  = is_shot_made or code in ("0/2","0/2D","0D+0/2W","0D+1/2W","0D+2/2W","0/3")
                    if is_shot_att:
                        player_key = finisher if finisher is not None else 0  # 0 = druzyna (brak finishera)
                        stats["zones"][player_key][zone_val]["att"] += 1
                        if is_shot_made:
                            stats["zones"][player_key][zone_val]["made"] += 1

            # Rzuty wolne
            ftm, fta = extract_ft(code)
            if fta > 0:
                q["ftm"]+=ftm; q["fta"]+=fta; pts+=ftm
                if ftm > 0:
                    stats["timing"][current_q][bucket]["ftm"]+=ftm
                if fta > 0:  # poss_ft = każde posiadanie z rzutami wolnymi (niezależnie od skuteczności)
                    stats["timing"][current_q][bucket]["poss_ft"]+=1
                if finisher is not None:
                    stats["players"][finisher]["ftm"]+=ftm
                    stats["players"][finisher]["fta"]+=fta

            q["pts"] += pts

            # Flow
            if pts > 0:
                total_pts = sum(stats["quarter"][qn].get("pts",0) for qn in range(1,5))
                stats["flow"].append((current_q, t_val, total_pts))

            # Lineup tracking
            if len(current_lineup) == 5:
                lk = "-".join(str(x) for x in sorted(current_lineup))
                lu = stats["lineups"][lk]
                lu["poss"] += 1
                lu["pts"]  += pts
                if code in ACTION_F:
                    lu["fd"] += 1
                if code in ACTION_2PM:
                    lu["p2m"]+=1; lu["p2a"]+=1
                elif code in ("0/2","0/2D","0D+0/2W","0D+1/2W","0D+2/2W"):
                    lu["p2a"]+=1
                elif code in ACTION_3PM:
                    lu["p3m"]+=1; lu["p3a"]+=1
                elif code == "0/3":
                    lu["p3a"]+=1
                elif code in ACTION_BR:
                    lu["br"]+=1
                if fta > 0:
                    lu["ftm"]+=ftm; lu["fta"]+=fta

    return stats

def build_gtk_def_lineups(ws_gtk, ws_opp):
    """DEF lineup stats GTK: arkusz rywala (ws_opp) zawiera w kolumnach E-I
    zawodnikow GTK na boisku podczas obrony - identyczna struktura jak arkusz GTK.
    Wystarczy sparsowac ws_opp przez parse_team_sheet i zwrocic jego lineups.
    """
    stats_opp_sheet = parse_team_sheet(ws_opp, sheet_type="B")
    return stats_opp_sheet["lineups"]

def build_lineup_section_html(cur, match_id, nr_name_map):
    """Sekcja piątek w PDF — jedna tabela OFF+DEF+NET, identyczna jak w app, sortowana wg POSS malejąco."""
    try:
        cur.execute("""SELECT * FROM lineup_stats WHERE match_id=%s AND druzyna='gtk'
                       ORDER BY poss DESC""", (match_id,))
        lu_off = list(cur.fetchall())
        cur.execute("""SELECT * FROM lineup_stats WHERE match_id=%s AND druzyna='gtk_def'
                       ORDER BY poss DESC""", (match_id,))
        lu_def = list(cur.fetchall())
    except Exception:
        return ""
    if not lu_off and not lu_def:
        return ""

    def_map_p = {lu["lineup"]: lu for lu in lu_def}
    off_rtg = {lu["lineup"]: lu["pts"]*100/lu["poss"] for lu in lu_off if int(lu.get("poss",0) or 0)>0}
    def_rtg = {lu["lineup"]: lu["pts"]*100/lu["poss"] for lu in lu_def if int(lu.get("poss",0) or 0)>0}

    def lu_name(s):
        return " · ".join(nr_name_map.get(str(n), f"#{n}") for n in s.split("-"))

    # Style nagłówka jak w app
    th  = "background:#1a2b4a;color:#fff;font-size:7px;font-weight:500;padding:3px 3px;text-align:center;white-space:nowrap"
    thl = "background:#1a2b4a;color:#fff;font-size:7px;font-weight:500;padding:3px 5px;text-align:left;white-space:nowrap"
    thg = "background:#1a2b4a;color:rgba(255,255,255,.55);font-size:6px;letter-spacing:.3px;padding:3px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center"
    thz = "background:#152236;color:rgba(255,255,255,.55);font-size:6px;letter-spacing:.3px;padding:3px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.1);text-align:center"
    ths = "background:#1a2b4a;color:rgba(255,255,255,.75);font-size:6px;font-weight:500;padding:2px 3px 3px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center"
    thzs= "background:#152236;color:rgba(255,255,255,.65);font-size:6px;font-weight:500;padding:2px 3px 3px;border-bottom:0.5px solid rgba(255,255,255,.1);text-align:center"
    than= "background:#412402;color:#FAC775;font-size:7px;font-weight:500;padding:3px 3px;text-align:center"
    vm  = "vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2)"

    header = (
        f'<thead><tr>'
        f'<th style="{thl};{vm}" rowspan="3">Skład</th>'
        f'<th style="{th};{vm}" rowspan="3">POSS</th>'
        f'<th style="{th};{vm}" rowspan="3">PKT</th>'
        f'<th style="{thg}" colspan="2">2PT</th>'
        f'<th style="{thg}" colspan="2">3PT</th>'
        f'<th style="{thg}" colspan="2">FT</th>'
        f'<th style="{thz}" colspan="2">ZB</th>'
        f'<th style="{th};{vm}" rowspan="3">AST</th>'
        f'<th style="{th};{vm}" rowspan="3">TO</th>'
        f'<th style="{th};{vm}" rowspan="3">STL</th>'
        f'<th style="{th};{vm}" rowspan="3">BLK</th>'
        f'<th style="{thg}" colspan="2">OFF</th>'
        f'<th style="{thg}" colspan="2">DEF</th>'
        f'<th style="{than}" colspan="3">NET RTG</th>'
        f'</tr><tr>'
        f'<th style="{ths}">M/A</th><th style="{ths}">%</th>'
        f'<th style="{ths}">M/A</th><th style="{ths}">%</th>'
        f'<th style="{ths}">M/A</th><th style="{ths}">%</th>'
        f'<th style="{thzs}">A</th><th style="{thzs}">O</th>'
        f'<th style="{ths}">eFG%</th><th style="{ths}">PPP</th>'
        f'<th style="{ths}">eFG%</th><th style="{ths}">PPP</th>'
        f'<th style="{than}">ORtg</th><th style="{than}">DRtg</th><th style="{than}">Net</th>'
        f'</tr></thead>'
    )

    rows = ""
    for i, lu in enumerate(sorted(lu_off, key=lambda x: int(x.get("poss",0) or 0), reverse=True)):
        k    = lu["lineup"]
        d_lu = def_map_p.get(k, {})
        p2m=int(lu.get("p2m",0) or 0); p2a=int(lu.get("p2a",0) or 0)
        p3m=int(lu.get("p3m",0) or 0); p3a=int(lu.get("p3a",0) or 0)
        ftm=int(lu.get("ftm",0) or 0); fta=int(lu.get("fta",0) or 0)
        pts=int(lu.get("pts",0) or 0); poss=int(lu.get("poss",0) or 0)
        br =int(lu.get("br",0)  or 0)
        oreb=int(lu.get("oreb",0) or 0)
        dreb=int(d_lu.get("dreb",0) or lu.get("dreb",0) or 0)
        ast =int(lu.get("ast",0) or 0)
        stl =int(d_lu.get("stl",0) or lu.get("stl",0) or 0)
        blk =int(d_lu.get("blk",0) or lu.get("blk",0) or 0)
        fga = p2a+p3a
        efg_v = round((p2m+1.5*p3m)/fga*100) if fga else None
        ppp_v = pts/poss if poss else None
        efg_s = f"{efg_v}%" if efg_v is not None else "—"
        ppp_s = f"{ppp_v:.2f}" if ppp_v is not None else "—"
        p2pct = f"{p2m/p2a:.0%}" if p2a else "—"
        p3pct = f"{p3m/p3a:.0%}" if p3a else "—"
        ftpct = f"{ftm/fta:.0%}" if fta else "—"
        # DEF stats
        dp2m=int(d_lu.get("p2m",0) or 0); dp2a=int(d_lu.get("p2a",0) or 0)
        dp3m=int(d_lu.get("p3m",0) or 0); dp3a=int(d_lu.get("p3a",0) or 0)
        dposs=int(d_lu.get("poss",0) or 0); dpts=int(d_lu.get("pts",0) or 0)
        dfga=dp2a+dp3a
        defg_v = round((dp2m+1.5*dp3m)/dfga*100) if dfga else None
        dppp_v = dpts/dposs if dposs else None
        defg_s = f"{defg_v}%" if defg_v is not None else "—"
        dppp_s = f"{dppp_v:.2f}" if dppp_v is not None else "—"
        # NET RTG
        ortg_v = off_rtg.get(k); drtg_v = def_rtg.get(k)
        net_v  = round(ortg_v-drtg_v,1) if (ortg_v is not None and drtg_v is not None) else None
        ortg_s = f"{ortg_v:.1f}" if ortg_v is not None else "—"
        drtg_s = f"{drtg_v:.1f}" if drtg_v is not None else "—"
        net_s  = f"{net_v:+.1f}" if net_v is not None else "—"
        net_c  = "#0F6E56" if (net_v is not None and net_v>0) else ("#A32D2D" if (net_v is not None and net_v<0) else "#888")
        ppp_c  = "#0F6E56" if ppp_v and ppp_v>=0.9 else ("#A32D2D" if ppp_v and ppp_v<0.7 else "#444")
        bg = "#f8f9ff" if i%2==0 else "#fff"
        rows += (
            f'<tr style="background:{bg};font-size:7.5px">'
            f'<td style="text-align:left;padding:2px 5px">{lu_name(k)}</td>'
            f'<td style="text-align:center;padding:2px 3px">{poss}</td>'
            f'<td style="text-align:center;font-weight:700;color:#1a2b4a;padding:2px 3px">{pts}</td>'
            f'<td style="text-align:center;padding:2px 3px">{p2m}/{p2a}</td>'
            f'<td style="text-align:center;padding:2px 3px">{p2pct}</td>'
            f'<td style="text-align:center;padding:2px 3px">{p3m}/{p3a}</td>'
            f'<td style="text-align:center;padding:2px 3px">{p3pct}</td>'
            f'<td style="text-align:center;padding:2px 3px">{ftm}/{fta}</td>'
            f'<td style="text-align:center;padding:2px 3px">{ftpct}</td>'
            f'<td style="text-align:center;padding:2px 3px">{oreb}</td>'
            f'<td style="text-align:center;padding:2px 3px">{dreb}</td>'
            f'<td style="text-align:center;padding:2px 3px">{ast}</td>'
            f'<td style="text-align:center;padding:2px 3px">{br}</td>'
            f'<td style="text-align:center;padding:2px 3px">{stl}</td>'
            f'<td style="text-align:center;padding:2px 3px">{blk}</td>'
            f'<td style="text-align:center;padding:2px 3px;color:#444">{efg_s}</td>'
            f'<td style="text-align:center;font-weight:700;padding:2px 3px;color:{ppp_c}">{ppp_s}</td>'
            f'<td style="text-align:center;padding:2px 3px;color:#888">{defg_s}</td>'
            f'<td style="text-align:center;padding:2px 3px;color:#888">{dppp_s}</td>'
            f'<td style="text-align:center;padding:2px 3px;color:#9FE1CB">{ortg_s}</td>'
            f'<td style="text-align:center;padding:2px 3px;color:#F09595">{drtg_s}</td>'
            f'<td style="text-align:center;font-weight:700;padding:2px 3px;color:{net_c}">{net_s}</td>'
            f'</tr>'
        )

    return (
        '<div style="margin-bottom:8px">'
        '<div style="font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:#1a2b4a;margin-bottom:4px">Piątki</div>'
        '<div style="overflow-x:auto">'
        '<table style="width:100%;border-collapse:collapse;table-layout:auto">'
        + header +
        f'<tbody>{rows}</tbody>'
        '</table></div></div>'
    ) if rows else ""


def build_nr_name_map(cur, match_id):
    """Buduje mapę nr→nazwisko dla zawodników GTK w meczu.
    Próbuje kolejno 4 źródła, każde uzupełnia brakujące numery.
    """
    result = {}
    try:
        # Źródło 1: players przez player_id (nowa struktura)
        cur.execute("""SELECT ps.nr, p.imie, p.nazwisko
                       FROM player_stats ps
                       JOIN players p ON ps.player_id = p.id
                       WHERE ps.match_id=%s AND ps.druzyna='gtk'""", (match_id,))
        for row in cur.fetchall():
            result[str(row["nr"])] = f"{row['nazwisko']} {row['imie'][0]}."
    except Exception:
        pass
    try:
        # Źródło 2: roster przez roster_id (stara struktura)
        cur.execute("""SELECT ps.nr, r.imie, r.nazwisko
                       FROM player_stats ps
                       JOIN roster r ON ps.roster_id = r.id
                       WHERE ps.match_id=%s AND ps.druzyna='gtk'""", (match_id,))
        for row in cur.fetchall():
            if str(row["nr"]) not in result:
                result[str(row["nr"])] = f"{row['nazwisko']} {row['imie'][0]}."
    except Exception:
        pass
    try:
        # Źródło 3: players przez numer+team_id meczu (bez przypisania player_id)
        cur.execute("""SELECT ps.nr, p.imie, p.nazwisko
                       FROM player_stats ps
                       JOIN matches m ON m.id = ps.match_id
                       JOIN players p ON p.team_id = m.team_id AND p.numer = ps.nr
                       WHERE ps.match_id=%s AND ps.druzyna='gtk'""", (match_id,))
        for row in cur.fetchall():
            if str(row["nr"]) not in result:
                result[str(row["nr"])] = f"{row['nazwisko']} {row['imie'][0]}."
    except Exception:
        pass
    try:
        # Źródło 4: player_aliases przez nr i sezon meczu
        cur.execute("""SELECT pa.nr, r.imie, r.nazwisko
                       FROM player_aliases pa
                       JOIN roster r ON pa.roster_id = r.id
                       JOIN matches m ON m.id = %s
                       WHERE pa.sezon = m.sezon""", (match_id,))
        for row in cur.fetchall():
            if str(row["nr"]) not in result:
                result[str(row["nr"])] = f"{row['nazwisko']} {row['imie'][0]}."
    except Exception:
        pass
    return result

def suma_quarters(stats):
    s = defaultdict(int)
    for qn in [1,2,3,4]:
        for k,v in stats["quarter"].get(qn,{}).items():
            s[k] += v
    return dict(s)

def calc_kpi(d):
    fga = d.get("p2a",0) + d.get("p3a",0)
    pts = d.get("pts",0)
    poss = max(d.get("poss",1),1)
    fta = d.get("fta",0); ftm = d.get("ftm",0)
    pm2 = d.get("p2m",0); pa2 = d.get("p2a",0)
    pm3 = d.get("p3m",0); pa3 = d.get("p3a",0)
    def pct(n,d): return f"{n/d:.1%}" if d else "-"
    efg  = (pm2+1.5*pm3)/fga if fga else None
    ts   = pts/(2*(fga+0.44*fta)) if (fga+fta) else None
    return {
        "efg":   pct(pm2+1.5*pm3,fga) if fga else "-",
        "ts":    f"{ts:.1%}" if ts else "-",
        "ortg":  f"{pts*100/poss:.1f}",
        "ppp":   f"{pts/poss:.2f}",
        "topct": f"{d.get('br',0)/poss:.1%}",
        "ftr":   f"{fta/fga:.2f}" if fga else "-",
        "p2_pct":pct(pm2,pa2),
        "p3_pct":pct(pm3,pa3),
        "ft_pct":pct(ftm,fta),
    }


def calc_play_time(match_id):
    """Szacowany czas gry zawodników GTK z pliku Excel.
    Sumuje czasy akcji (kol B) per zawodnik (kol E-I) z obu arkuszy.
    Zwraca dict {nr: sekundy}
    """
    import os as _os
    path = _os.path.join(MATCH_FILES_DIR, f"{match_id}.xlsx")
    if not _os.path.exists(path):
        return {}
    def _parse_secs(v):
        if v is None: return 0
        try:
            total = 0
            for p in str(v).replace(",", ".").split(";"):
                p = p.strip()
                if p.replace(".", "").isdigit():
                    total += float(p)
            return total
        except: return 0
    def _parse_lineup(row):
        nrs = set()
        for col in row[4:9]:
            if col is None: continue
            for p in str(col).split(";"):
                p = p.strip()
                try: nrs.add(int(float(p)))
                except: pass
        return nrs
    def _sheet_secs(ws):
        rows = list(ws.iter_rows(values_only=True))[1:]
        lineup_rows = [(i, row) for i, row in enumerate(rows)
                       if any(row[j] is not None for j in range(4, 9))]
        secs = {}
        for si, (li, lr) in enumerate(lineup_rows):
            nxt = lineup_rows[si+1][0] if si+1 < len(lineup_rows) else len(rows)
            t = sum(_parse_secs(rows[j][1]) for j in range(li, nxt))
            for nr in _parse_lineup(lr):
                secs[nr] = secs.get(nr, 0) + t
        return secs
    try:
        import openpyxl as _opx
        wb = _opx.load_workbook(path, read_only=True, data_only=True)
        report = validate_workbook(wb)
        name_a, name_b = report["names"]
        s_a = _sheet_secs(wb[name_a]) if name_a in wb.sheetnames else {}
        s_b = _sheet_secs(wb[name_b]) if name_b in wb.sheetnames else {}
        wb.close()
        result = {}
        for nr in set(s_a) | set(s_b):
            result[nr] = (s_a.get(nr, 0) + s_b.get(nr, 0)) * 1.22
        return result
    except:
        return {}

def save_match_to_db(przeciwnik, nazwa_gtk, sezon, data_meczu, stats_gtk, stats_opp,
                     rozgrywki="", runda="", kolejka="", miejsce="", def_lineups=None,
                     team_name_a="", team_name_b="",
                     team_id=None, nr_to_player=None):
    db = get_db()
    cur = db.cursor()
    suma_gtk = suma_quarters(stats_gtk)
    suma_opp = suma_quarters(stats_opp)

    # Wstaw mecz
    # Konwertuj pusty string na None — PostgreSQL nie akceptuje "" jako DATE
    _data = data_meczu if data_meczu and str(data_meczu).strip() else None
    cur.execute("""
        INSERT INTO matches (sezon, data_meczu, przeciwnik, nazwa_gtk, rozgrywki, runda, kolejka, miejsce, wynik_gtk, wynik_opp, team_name_a, team_name_b)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
    """, (sezon, _data, przeciwnik, nazwa_gtk,
          rozgrywki or "", runda or "", kolejka or "", miejsce or "",
          suma_gtk.get("pts",0), suma_opp.get("pts",0),
          team_name_a or nazwa_gtk, team_name_b or przeciwnik))
    match_id = cur.fetchone()["id"]

    # Statystyki per kwarta
    # oreb GTK pochodzi z arkusza A (stats_gtk), dreb GTK z arkusza B (stats_opp)
    # stl/blk GTK z arkusza B (stats_opp)
    # stl OPP pochodzi z kolumny O arkusza A ($ = przechwyt rywala) → stats_gtk["opp_quarter"]
    for qn in [1,2,3,4]:
        qg = dict(stats_gtk["quarter"].get(qn,{}))
        qo = dict(stats_opp["quarter"].get(qn,{}))
        # Uzupełnij GTK o dane obronne z arkusza B
        qg["dreb"] = qo.get("dreb", 0)
        qg["stl"]  = qo.get("stl",  0)
        qg["blk"]  = qo.get("blk",  0)
        # STL dla OPP: jedyna poprawna logika to $ w kolumnie O arkusza A
        qo["stl"]  = stats_gtk.get("opp_quarter", {}).get(qn, {}).get("stl", 0)
        for druzyna, qd in [("gtk", qg), ("opp", qo)]:
            cur.execute("""
                INSERT INTO match_stats (match_id,druzyna,kwarta,pts,poss,p2m,p2a,p3m,p3a,ftm,fta,br,fd,ast,oreb,dreb,stl,blk,d2m,d2a,przerw)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (match_id, druzyna, qn,
                  qd.get("pts",0), qd.get("poss",0),
                  qd.get("p2m",0), qd.get("p2a",0),
                  qd.get("p3m",0), qd.get("p3a",0),
                  qd.get("ftm",0), qd.get("fta",0),
                  qd.get("br",0),  qd.get("fd",0),
                  qd.get("ast",0), qd.get("oreb",0), qd.get("dreb",0),
                  qd.get("stl",0), qd.get("blk",0),
                  qd.get("d2m",0), qd.get("d2a",0), qd.get("przerw",0)))

    # Zawodnicy — dreb/stl/blk GTK z arkusza B (stats_opp), oreb/ast z arkusza A (stats_gtk)
    opp_players = dict(stats_opp["players"])  # nr → dane obronne
    for druzyna, stats in [("gtk", stats_gtk), ("opp", stats_opp)]:
        for nr, pd in stats["players"].items():
            pts = pd.get("p2m",0)*2 + pd.get("p3m",0)*3 + pd.get("ftm",0)
            # Dla druzyny gtk: uzupelnij dreb/stl/blk z arkusza B
            if druzyna == "gtk" and int(nr) in opp_players:
                op = opp_players[int(nr)]
                dreb = op.get("dreb",0); stl = op.get("stl",0); blk = op.get("blk",0)
            else:
                dreb = pd.get("dreb",0); stl = pd.get("stl",0); blk = pd.get("blk",0)
            cur.execute("""
                INSERT INTO player_stats (match_id,druzyna,nr,pts,p2m,p2a,p3m,p3a,ftm,fta,ast,oreb,dreb,br,fd,finishes,time_sum,time_cnt,stl,blk)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (match_id, druzyna, int(nr), pts,
                  pd.get("p2m",0), pd.get("p2a",0),
                  pd.get("p3m",0), pd.get("p3a",0),
                  pd.get("ftm",0), pd.get("fta",0),
                  pd.get("ast",0), pd.get("oreb",0),
                  dreb, pd.get("br",0),
                  pd.get("fd",0), pd.get("finishes",0),
                  pd.get("time_sum",0), pd.get("time_cnt",0),
                  stl, blk))

    # Timing Akcji — zapis do timing_stats per kwarta i bucket
    for druzyna, stats in [("gtk", stats_gtk), ("opp", stats_opp)]:
        for q in [0,1,2,3,4]:
            for b in BUCKETS:
                td = stats["timing"][q][b]
                made2 = td["2PT"]["made"]
                miss2 = td["2PT"]["miss"]
                made3 = td["3PT"]["made"]
                miss3 = td["3PT"]["miss"]
                br_v     = td.get("br", 0)
                ftm_v    = td.get("ftm", 0)
                poss_ft_v= td.get("poss_ft", 0)
                att2  = made2 + miss2
                att3  = made3 + miss3
                if att2 + att3 + br_v + ftm_v == 0:
                    continue
                cur.execute("""
                    INSERT INTO timing_stats (match_id, druzyna, bucket, kwarta, made2, att2, made3, att3, br, ftm, poss_ft)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (match_id, druzyna, b, q,
                      made2, att2, made3, att3, br_v, ftm_v, poss_ft_v))


    # Lineup stats
    for druzyna, stats in [("gtk", stats_gtk), ("opp", stats_opp)]:
        for lineup_key, lu in stats.get("lineups", {}).items():
            # STL/BLK dla GTK są w def_lineups (arkusz B, kol O/P)
            _def_lu = (def_lineups or {}).get(lineup_key, {}) if druzyna == "gtk" else {}
            cur.execute("""
                INSERT INTO lineup_stats (match_id,druzyna,lineup,pts,poss,p2m,p2a,p3m,p3a,ftm,fta,br,fd,ast,oreb,dreb,stl,blk)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (match_id, druzyna, lineup_key,
                  lu.get("pts",0), lu.get("poss",0),
                  lu.get("p2m",0), lu.get("p2a",0),
                  lu.get("p3m",0), lu.get("p3a",0),
                  lu.get("ftm",0), lu.get("fta",0),
                  lu.get("br",0),  lu.get("fd",0),
                  lu.get("ast",0), lu.get("oreb",0),
                  _def_lu.get("dreb",0) or lu.get("dreb",0),
                  _def_lu.get("stl",0) or lu.get("stl",0),
                  _def_lu.get("blk",0) or lu.get("blk",0)))
    if def_lineups:
        for lineup_key, lu in def_lineups.items():
            cur.execute("""
                INSERT INTO lineup_stats (match_id,druzyna,lineup,pts,poss,p2m,p2a,p3m,p3a,ftm,fta,br,fd,ast,oreb,dreb,stl,blk)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (match_id, "gtk_def", lineup_key,
                  lu.get("pts",0), lu.get("poss",0),
                  lu.get("p2m",0), lu.get("p2a",0),
                  lu.get("p3m",0), lu.get("p3a",0),
                  lu.get("ftm",0), lu.get("fta",0),
                  lu.get("br",0),  lu.get("fd",0),
                  lu.get("ast",0), lu.get("oreb",0),
                  lu.get("dreb",0),lu.get("stl",0), lu.get("blk",0)))

    # Shot zones
    for druzyna, st in [("gtk", stats_gtk), ("opp", stats_opp)]:
        for nr, zone_map in st.get("zones", {}).items():
            for zone, zd in zone_map.items():
                if zd["att"] > 0:
                    cur.execute("""
                        INSERT INTO shot_zones (match_id, druzyna, nr, zone, made, att)
                        VALUES (%s,%s,%s,%s,%s,%s)
                    """, (match_id, druzyna, int(nr), int(zone), zd["made"], zd["att"]))

    # Score flow
    flow_gtk = stats_gtk.get("flow", [])
    flow_opp = stats_opp.get("flow", [])
    # Build cumulative score flow from gtk perspective
    flow_map_gtk = {(q, t): p for q, t, p in flow_gtk}
    flow_map_opp = {(q, t): p for q, t, p in flow_opp}
    all_times = sorted(set(list(flow_map_gtk.keys()) + list(flow_map_opp.keys())))
    last_gtk = 0; last_opp = 0
    for (q, t) in all_times:
        last_gtk = flow_map_gtk.get((q, t), last_gtk)
        last_opp = flow_map_opp.get((q, t), last_opp)
        cur.execute(
            "INSERT INTO score_flow (match_id, kwarta, czas_sek, pts_gtk, pts_opp) VALUES (%s,%s,%s,%s,%s)",
            (match_id, q, t, last_gtk, last_opp)
        )

    # Set team_id on the match
    if team_id:
        try:
            cur.execute("UPDATE matches SET team_id=%s WHERE id=%s", (int(team_id), match_id))
        except Exception:
            pass

    # Assign player_id from nr_to_player mapping
    if nr_to_player:
        for nr_str, player_id_val in nr_to_player.items():
            if not player_id_val:
                continue
            try:
                cur.execute(
                    "UPDATE player_stats SET player_id=%s WHERE match_id=%s AND druzyna='gtk' AND nr=%s",
                    (int(player_id_val), match_id, int(nr_str))
                )
            except Exception:
                pass

    db.commit()
    cur.close()
    return match_id


CSS = """
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
/* ── ZMIENNE ── */
:root{
  --navy:#1a2b4a; --navy2:#2e5090;
  --gtk:#1a6b3c;  --gtk-light:#e8f5e9;
  --opp:#8b1a1a;  --opp-light:#ffebee;
  --gold:#EF9F27;
  --bg:#f0f2f7;
  --card:#fff;
  --radius:12px;
  --shadow:0 1px 6px rgba(0,0,0,.08);
}

/* ── RESET ── */
*{box-sizing:border-box}
body{margin:0;background:var(--bg);font-family:'Segoe UI',system-ui,sans-serif;font-size:.9rem;color:#222}

/* ══════════════════════════════════════════
   SIDEBAR (desktop)
══════════════════════════════════════════ */
.sidebar{
  position:fixed;top:0;left:0;height:100vh;width:240px;
  background:var(--navy);z-index:1000;
  display:flex;flex-direction:column;
  overflow-y:auto;overflow-x:hidden;
  transition:transform .3s ease;
}
.sidebar-logo{
  padding:1.1rem 1rem .9rem;
  border-bottom:1px solid #ffffff18;
  flex-shrink:0;
}
.sidebar-logo .brand{font-size:1.05rem;font-weight:700;color:#fff;letter-spacing:.3px}
.sidebar-logo .brand span{color:var(--gold)}
.sidebar-logo .sub{font-size:.68rem;color:#ffffff55;margin-top:2px}
.nav-season{
  margin:.6rem .6rem .3rem;
  padding:.5rem .7rem;
  background:#ffffff0d;
  border-radius:8px;
  font-size:.73rem;color:#ffffff77;
}
.nav-season strong{color:var(--gold);display:block;font-size:.8rem;margin-bottom:1px}
.nav-section{
  padding:.5rem 1rem .2rem;
  font-size:.6rem;text-transform:uppercase;
  letter-spacing:1.2px;color:#ffffff33;font-weight:700;
}
.nav-item-link{
  display:flex;align-items:center;gap:.6rem;
  padding:.58rem .9rem;
  color:#ffffffaa;text-decoration:none;
  font-size:.84rem;border-radius:8px;
  margin:1px .5rem;transition:.15s;
  white-space:nowrap;
}
.nav-item-link:hover{background:#ffffff14;color:#fff}
.nav-item-link.active{background:#EF9F2720;color:var(--gold);font-weight:600}
.nav-item-link .icon{width:20px;text-align:center;font-size:.95rem;flex-shrink:0}

/* ══════════════════════════════════════════
   TOPBAR (mobile)
══════════════════════════════════════════ */
.topbar{
  display:none;
  position:fixed;top:0;left:0;right:0;height:56px;
  background:var(--navy);z-index:1001;
  align-items:center;padding:0 1rem;gap:.75rem;
}
.topbar .t-brand{font-size:1rem;font-weight:700;color:#fff}
.topbar .t-brand span{color:var(--gold)}
.hamburger{
  background:none;border:none;cursor:pointer;
  padding:6px;border-radius:6px;
  display:flex;flex-direction:column;gap:5px;
}
.hamburger span{display:block;width:22px;height:2px;background:#fff;border-radius:2px;transition:.25s}
.hamburger.open span:nth-child(1){transform:translateY(7px) rotate(45deg)}
.hamburger.open span:nth-child(2){opacity:0}
.hamburger.open span:nth-child(3){transform:translateY(-7px) rotate(-45deg)}
.sidebar-overlay{
  display:none;position:fixed;inset:0;background:#00000055;z-index:999;
}

/* ══════════════════════════════════════════
   MAIN CONTENT
══════════════════════════════════════════ */
.main-content{
  margin-left:240px;
  min-height:100vh;
  padding:1.5rem 1.25rem;
  max-width:1400px;
}

/* ══════════════════════════════════════════
   CARDS & STATS
══════════════════════════════════════════ */
.card{
  border:none;border-radius:var(--radius);
  box-shadow:var(--shadow);background:var(--card);
}
.stat-card{
  background:var(--card);border-radius:var(--radius);
  padding:.85rem .75rem;text-align:center;
  box-shadow:var(--shadow);
}
.stat-val{font-size:1.5rem;font-weight:700;color:var(--navy);line-height:1.1}
.stat-val.sm{font-size:1.05rem}
.stat-lbl{font-size:.65rem;color:#999;text-transform:uppercase;letter-spacing:.5px;margin-top:.2rem}

/* ══════════════════════════════════════════
   TABLES
══════════════════════════════════════════ */
.table th{
  background:var(--navy);color:#fff;
  font-size:.75rem;font-weight:600;
  border:none;padding:.42rem .55rem;
  white-space:nowrap;
}
.table td{
  font-size:.8rem;vertical-align:middle;
  padding:.36rem .55rem;
}
.table-hover tbody tr:hover{background:#f0f4ff}
.table-responsive{-webkit-overflow-scrolling:touch}

/* ══════════════════════════════════════════
   MISC UI
══════════════════════════════════════════ */
.hero{
  background:linear-gradient(135deg,var(--navy),var(--navy2));
  color:#fff;border-radius:14px;padding:1.25rem 1.5rem;
}
.page-title{font-size:1.2rem;font-weight:700;color:var(--navy);margin-bottom:.75rem}
.badge-win{background:#c8e6c9;color:var(--gtk);font-size:.72rem;padding:3px 10px;border-radius:20px;font-weight:700}
.badge-loss{background:#ffcdd2;color:var(--opp);font-size:.72rem;padding:3px 10px;border-radius:20px;font-weight:700}
.badge-draw{background:#e0e0e0;color:#555;font-size:.72rem;padding:3px 10px;border-radius:20px;font-weight:700}
.gtk-color{color:var(--gtk);font-weight:700}
.opp-color{color:var(--opp);font-weight:700}
.upload-zone{
  border:2px dashed #c5cfe8;border-radius:14px;
  padding:2rem 1.5rem;text-align:center;
  background:#fff;cursor:pointer;transition:.2s;
}
.upload-zone:hover{background:#f0f4ff;border-color:var(--navy)}
.nav-tabs .nav-link{color:#666;font-size:.82rem;padding:.4rem .8rem}
.nav-tabs .nav-link.active{color:var(--navy);font-weight:600;border-bottom:2px solid var(--navy)}
.section-hdr{
  font-size:.66rem;text-transform:uppercase;letter-spacing:1px;
  color:#aaa;font-weight:700;margin:.6rem 0 .35rem;
  padding-bottom:.25rem;border-bottom:1px solid #f0f0f0;
}
.flash-msg{padding:.6rem 1rem;border-radius:8px;margin-bottom:.75rem;font-size:.85rem}
.flash-success{background:#e8f5e9;color:#1a6b3c;border:1px solid #a5d6a7}
.flash-error{background:#ffebee;color:#8b1a1a;border:1px solid #ef9a9a}

/* ══════════════════════════════════════════
   RESPONSIVE BREAKPOINTS
══════════════════════════════════════════ */

/* ══════════════════════════════════════════
   RESPONSIVE BREAKPOINTS
══════════════════════════════════════════ */

/* Tablet / pół ekranu (< 992px) — sidebar tylko ikony */
@media(max-width:991px){
  .sidebar{width:56px}
  .sidebar-logo{padding:.75rem 0;text-align:center}
  .sidebar .brand-text,.sidebar .sub,
  .nav-section,.nav-season{display:none}
  .nav-item-link span.brand-text{display:none}
  .nav-item-link{
    justify-content:center;
    padding:.65rem 0;
    margin:2px 6px;
    border-radius:8px;
    position:relative;
  }
  .nav-item-link .icon{
    width:28px;height:28px;
    display:flex;align-items:center;justify-content:center;
    font-size:1.2rem;
    margin:0;
  }
  /* Tooltip przy hover */
  .nav-item-link::after{
    content:attr(data-label);
    position:absolute;left:62px;top:50%;transform:translateY(-50%);
    background:#1a2b4a;color:#fff;
    font-size:.75rem;font-weight:600;
    padding:4px 10px;border-radius:6px;
    white-space:nowrap;pointer-events:none;
    opacity:0;transition:opacity .15s;
    box-shadow:0 2px 8px rgba(0,0,0,.25);
    z-index:9999;
  }
  .nav-item-link:hover::after{opacity:1}
  .nav-item-link.active .icon{
    background:#EF9F2720;border-radius:6px;
  }
  .main-content{margin-left:56px;padding:1.25rem 1rem}
  /* Logo — tylko kropka */
  .sidebar-logo .brand{font-size:1.2rem;text-align:center;display:block}
}

/* Mobile (< 768px) — sidebar chowana, topbar widoczny */
@media(max-width:767px){
  .sidebar{transform:translateX(-100%);width:240px}
  .sidebar.mobile-open{transform:translateX(0)}
  .sidebar .brand-text,.sidebar .sub,
  .nav-section,.nav-item-link span,
  .nav-season{display:block}
  .nav-item-link{
    justify-content:flex-start;
    padding:.58rem .9rem;margin:1px .5rem;
  }
  .nav-item-link::after{display:none}
  .nav-item-link .icon{width:20px;height:auto;font-size:.95rem}
  .topbar{display:flex}
  .main-content{margin-left:0;padding:4.5rem 1rem 1.5rem}
  .hero{padding:1rem 1.1rem}
  .stat-val{font-size:1.2rem}
  .page-title{font-size:1.05rem}
  .table th,.table td{font-size:.72rem;padding:.3rem .4rem}
  /* Sekcje — wyrównanie kontrolek */
  .d-flex.justify-content-between.flex-wrap{gap:.5rem}
  /* Kolumny Bootstrap — stos na mobile */
  .col-lg-7,.col-lg-5,.col-md-6,.col-lg-4,.col-lg-8{flex:0 0 100%;max-width:100%}
  /* Karty stat — 3 w rzędzie */
  .col-6{flex:0 0 50%;max-width:50%}
  .col-4{flex:0 0 50%;max-width:50%}
  /* Stat-karty kompaktowe */
  .stat-card{padding:.6rem .5rem}
  /* Wzmianka scroll dla tabel */
  .table-responsive::before{
    content:"← przewijaj →";
    display:block;font-size:.6rem;color:#aaa;
    text-align:center;padding:3px 0 2px;
    display:none /* wyłączone domyślnie, włączane w overflowing */
  }
}

/* Bardzo małe (< 480px) */
@media(max-width:479px){
  .main-content{padding:4.5rem .65rem 1.5rem}
  .hero{padding:.85rem .9rem}
  .stat-card{padding:.55rem .4rem}
  .stat-val{font-size:.95rem}
  .stat-val.sm{font-size:.85rem}
  .stat-lbl{font-size:.58rem}
  .card .card-body{padding:.65rem .55rem !important}
  .btn-sm{font-size:.7rem;padding:.22rem .45rem}
  .nav-tabs .nav-link{font-size:.72rem;padding:.3rem .45rem}
  .table th,.table td{font-size:.66rem;padding:.25rem .3rem}
  /* 3 kolumny stat kart na bardzo małym */
  .col-6{flex:0 0 50%;max-width:50%}
  .col-3{flex:0 0 50%;max-width:50%}
  /* Pill dropdown kompaktowy */
  #perSelect{font-size:.72rem;padding:5px 28px 5px 12px}
  /* Przycisk toggle kompaktowy */
  #statModeAvg,#statModeSum{font-size:.68rem;padding:4px 10px}
}
.nav-submenu{background:#ffffff08;border-radius:6px;margin:2px 6px 4px}
.nav-group>.nav-item-link{border-radius:8px;display:flex;align-items:center}
@media(max-width:991px){.nav-group .brand-text,.nav-submenu{display:none!important}}
</style>
"""

def nav(active="home"):
    import json as _j
    klub    = get_setting("current_klub") or ""
    season  = get_setting("current_season") or ""
    druzyna = get_setting("current_druzyna") or ""
    try:
        kj    = get_setting("kluby_json") or "[]"
        kluby = _j.loads(kj)
    except:
        kluby = []

    active_klub  = next((k for k in kluby if k["name"] == klub), None)
    sezony_keys  = list(active_klub.get("sezony", {}).keys()) if active_klub else []
    druzyna_list = (active_klub.get("sezony", {}).get(season, [])
                   if active_klub and season else [])

    def pill(fid, lbl, val, opts, ph):
        items = ""
        if opts:
            for o in opts:
                sc  = " selected" if o == val else ""
                oe  = o.replace("'", "\\'")
                items += ('<div class="dd-opt' + sc + '" '
                          'onclick="pickCtx(\'' + fid + '\',\'' + oe + '\')">'
                          + o + '</div>')
        else:
            items = '<div class="dd-opt disabled">' + ph + '</div>'
        cur   = ('<div class="ctx-val">'   + val + '</div>') if val else ('<div class="ctx-empty">' + ph + '</div>')
        chev  = ('<svg id="chev-' + fid + '" class="ctx-chev" viewBox="0 0 24 24"'
                 ' fill="none" stroke="rgba(255,255,255,.3)" stroke-width="2">'
                 '<polyline points="6 9 12 15 18 9"/></svg>') if opts else ""
        oc    = 'onclick="togglePill(\'' + fid + '\')"' if opts else ""
        return ('<div class="ctx-wrap" id="pw-' + fid + '">'
                '<div class="ctx-hdr" ' + oc + '>'
                '<div><div class="ctx-lbl">' + lbl + '</div>' + cur + '</div>'
                + chev +
                '</div>'
                '<div id="dd-' + fid + '" class="ctx-dd" style="display:none">'
                + items + '</div></div>')

    p_klub  = pill("klub",    "Klub",
                   klub,    [k["name"] for k in kluby],
                   "\u2014 brak klub\u00f3w \u2014" if not kluby else "\u2014 wybierz klub \u2014")
    p_sezon = pill("sezon",   "Sezon",   season,  sezony_keys,  "\u2014 wybierz sezon \u2014")
    p_druz  = pill("druzyna", "Dru\u017cyna", druzyna, druzyna_list, "\u2014 wybierz sezon \u2014")

    hint = ""

    def ni(key, href, svg, label):
        s = ("background:rgba(239,159,39,.18);color:#EF9F27;font-weight:600;"
             if active == key else "color:rgba(255,255,255,.55);")
        return ('<a href="' + href + '" style="display:flex;align-items:center;gap:8px;'
                'padding:8px 10px;border-radius:7px;margin:1px 6px;'
                + s + 'font-size:11px;text-decoration:none">'
                '<svg width="15" height="15" viewBox="0 0 24 24" fill="none"'
                ' stroke="currentColor" stroke-width="1.8" style="flex-shrink:0">'
                + svg + '</svg>'
                '<span class="brand-text">' + label + '</span></a>')

    so    = active in ("season", "players")
    ss    = ("background:rgba(239,159,39,.18);color:#EF9F27;font-weight:600;"
             if so else "color:rgba(255,255,255,.55);")
    s_rot  = "90deg" if so  else "0deg"
    s_disp = "block" if so  else "none"
    uname  = session.get("user_name", "")

    parts = [
        '<div class="topbar">',
        '<button class="hamburger" id="hamburger" aria-label="Menu"'
        ' onclick="toggleSidebar()"><span></span><span></span><span></span></button>',
        '<div class="t-brand"><span>&#x25cf;</span> Basket Kołcz</div>',
        '</div>',
        '<div class="sidebar-overlay" id="sidebarOverlay" onclick="toggleSidebar()"></div>',
        '<style>',
        '.ctx-wrap{position:relative;margin-bottom:3px}',
        '.ctx-hdr{display:flex;align-items:center;justify-content:space-between;padding:7px 10px;background:rgba(255,255,255,.07);border-radius:8px;transition:background .15s}',
        '.ctx-hdr[onclick]{cursor:pointer}',
        '.ctx-hdr[onclick]:hover{background:rgba(255,255,255,.11)}',
        '.ctx-hdr.is-open{border-radius:8px 8px 0 0;background:rgba(255,255,255,.1)}',
        '.ctx-lbl{font-size:9px;color:rgba(255,255,255,.35);text-transform:uppercase;letter-spacing:.4px;margin-bottom:1px}',
        '.ctx-val{font-size:11px;color:#fff;font-weight:500}',
        '.ctx-empty{font-size:11px;color:rgba(255,255,255,.25);font-style:italic}',
        '.ctx-chev{width:12px;height:12px;transition:transform .2s;flex-shrink:0}',
        '.ctx-dd{background:#152032;border-radius:0 0 8px 8px;border-top:0.5px solid rgba(255,255,255,.08);overflow:hidden;z-index:50}',
        '.dd-opt{padding:8px 12px;cursor:pointer;font-size:12px;color:rgba(255,255,255,.65);transition:background .1s}',
        '.dd-opt:hover{background:rgba(255,255,255,.07);color:#fff}',
        '.dd-opt.selected{color:#EF9F27;font-weight:500}',
        '.dd-opt.disabled{color:rgba(255,255,255,.2);font-style:italic;cursor:default;font-size:11px}',
        '</style>',
        '<div class="sidebar" id="sidebar">',
        '<div class="sidebar-logo"><div style="display:flex;align-items:center;gap:10px">',
        '<img src="data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCABIAEgDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwDl3OpYw0jke6Cm7b08kf8AjgqwstznjZ/37FSLNeEjBOc8AIK/N7+h95YqiG77of8Avika3uv+eb/98U/UNeWwk8u91GGKTIHlkZfngDaMnn6VRPiyPc+2K9kCRmRm+zBFCh9hYlyvG4bfrxVKnNq6RDqQTtcne3ve0Tn/AIAKQWmogf6g/ioqpd+M7a0t7e4ube7jiueIG2xN5p6YAVyevHseDipz4pgS5W3nh1C3maR41R7NiSyffA27s47npVexqW2/pb/cL2sL2v8A09vvHm01Ltb5+i//AF6Y9vqS9bXr6rV2HUmvUVrC4juBjJEbKSPqOo/GiT+0CfmWQfgKjVOzRorPVGc1tqGOYMf8Boq4Vu+4kz9BRRzegcpqTvCkMjQN5sgUlIy20Mewzg4+tcjPc6nqcEcrXCqbe4RrzTwnlAKAxeJ3HzD5RuVyQj8AY6V2WdPH/LST9a474o3WlWelQTW9stzqlxILW1LIW25OSSvR8cEKQfmINXhU5zUFu+v9fpZmVeSpxc5apdP6/W6OaudT8P21xd6ZYXF5fWLu80ltZxOXWQgfOkwcfu/u8SbtpBGO9dPLBrNzHJ4h1G00vTkubUW3mfbGeSRH2sFUQqQzbwxXac/ORz1p2kaVd+HdLgs7dpLfUr8ny7WBl+03Dj70k8xB2IuckJwvQFial8NeF7AeIJdMummvI9HtIRlp5ArXMxZ3kADfLheABjhjSxGJo2ck9F13bWibWyV3ZX96/fRm+Hwte6jJay6bJPVpPduyu7e7btqjnR4N1SGLR7nUbjRtLtNKlZ7WO9keNpg0vmbXGTt54xknHXmk1nVLq3mm1DULSPU5JIoY5bjTblv3qRytJsIZAVRiVDbcnCe5rqIdFsNI1uU6nNfWkk87fYNUS5YqEY5WBt+4Iy9BkYYd88V0Mp8T6d8ySxa5AOsbgW9yB7Efu3PsQv1rKpmV5JytK+32d3d2evXe7WvTQ2p5ZaLUbxto7e9srK606bWi9OutzywajpuqOL+ylLva77xlVjb3V5eSH503nC7R6KS20ADkmtnR9W1lUdbu2kW4AQjTJD5lyePmKsMsoz0EoH+9UvimKzljuvEmgRqLu2/5CWn3EOwyKOTHNEe+MlW9uCa6nwxb6AbCLVtORh9uhRw8shkcJjIjy2SFX0rulWpuipKLttZ7p727a73te2zR58qFSnV5XJPrdbNbX73WzV7X3T6p5MjRq2JULAEqycj2OKK0pZLc5Hm/5/KivP5mdVhnkWfrJ+Zrh/ijdWGjah4a1cSFpLTUN5hbkvHgbiPpgfnXToL7uJD+H/1q4X4t6Pe3VlDfR2zyyqGWRzkiGJFZzgY4yRye+AK7cvhGWIipvR3X3qxx4+U40HKG6s/uaZ3Hw+urbW5NV8SpKs0lxdvawn/nlbxn5EHpnJc+pPtU9o40/wCIuoxTEJHqtjDPCzHAZ4co659drKfpXz14b8R614cuWn0i9kty+PMQgMkn+8p4P1qW61uXXNUFz4pv9Suo+cCErlfZVb5VH0FdFThmo61R8/7tq3mrWsrbaWXXUilxTT9hTXJ+8Tu9bJ3vd331TfTQ+obmK3uLdre5jilhmUq0cgBWRT2weornLaWTw3rFrol1M8uk6gTHYSSsS9vIBnyGY8lSPuE8jGD2rw27m8JSWwCWniaAoMQu91FID/wEqMfgay/7f1gWaWX9pXL28UyzxLI24xuvRlJ6fhWGH4XnyuPPo901b0as3qv+AdGI4rp8ylyarZp39U7paP8A4J6J8Q/E9k099Akoj1bT9sUM4GftETjDwt6gZzg9CMjvXWeC47O28I6VAkxmC2y5dH+Uk8nHHqSPwrwGNZru7VAWknnkABJ5ZmPr9TX0D4Nsba28MWKy2zW8zxCSaIMQEkI+bA7ZPOPUmu7McHTweGhTi+v6fle7+bPPwWYVMfip1ZKyt+v52svkjQP2Un7zD3LH/CipWjsv7r/99GivDPYLK3M5GcKf+A1HcTPJE8MyRPG6lWUqcMCMEVREj4xvz9RS+Y5/i/8AHafLYR5/4p+H8cstk2mMwhSL7PIp5ZBzsfp8wGQGHXAzXnd5Y3Wkag9tqFo8c0Z5ViQvscjqD6ivoMHj736VX1Cxs9QgeC8gSeN0KMGXJweuD1Hr9RXtYXOalL3ai5l+J42LyalV96n7r/A8IvryGW3wpJkIwSFxn255C+3em6JYy385gtLaS6uSpKqB8q/7R9h/nNekaT8NtKtrppr+5lvEDkxw7dq7c8bj1Y/kK66ztNPsFdbOKG2DnLeWm3J969CvndKKtSV3+B5+HyOrN3quy/E5XwB4KSwhjv8AWIla8DI8UQOfJ2kkZ9WJOfbArt5C3OA/4mq++PvKD+H/ANak8yH/AJ6D8q+dxFepiJ882fR4fD08PBQghzGQD7rdfUUUwuuMhifworGxvct7ZB2zTG83sKKKkQ0ibqKYTP7/AJA0UUJjGtJMMZXd3xij5mXPkSgfSiimIYyg9Y5P++aiKqD0kH1ooqkFwKgjjcPoaKKKoD//2Q==" alt="logo"'
        ' style="width:36px;height:36px;border-radius:50%;object-fit:cover;flex-shrink:0">',
        '<div class="brand-text"><div class="brand">Basket <span>Kołcz</span></div>'
        '<div class="sub">Analytics Platform</div></div>',
        '</div></div>',
        '<div class="nav-section brand-text">Nawigacja</div>',
        ni("home", "/",
           '<rect x="3" y="3" width="7" height="7" rx="1"/><rect x="14" y="3" width="7" height="7" rx="1"/>'
           '<rect x="3" y="14" width="7" height="7" rx="1"/><rect x="14" y="14" width="7" height="7" rx="1"/>',
           "Pulpit"),
        ni("druzyny", "/druzyny",
           '<rect x="3" y="3" width="18" height="18" rx="2"/><path d="M9 12h6M12 9v6"/>',
           "Struktura klub\u00f3w"),

        '<div style="display:flex;align-items:center;gap:8px;padding:8px 10px;border-radius:7px;margin:1px 6px;' + ss + 'font-size:11px;cursor:pointer" onclick="toggleSub(&quot;stats-sub&quot;)">',
        '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" style="flex-shrink:0"><polyline points="22 12 18 12 15 21 9 3 6 12 2 12"/></svg>',
        '<span class="brand-text">Statystyki</span>',
        '<svg id="arr-stats" style="margin-left:auto;width:10px;height:10px;transition:transform .2s;transform:rotate(' + s_rot + ')" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="9 18 15 12 9 6"/></svg>',
        '</div>',
        '<div id="stats-sub" class="brand-text" style="display:' + s_disp + '">',
        ni("season",  "/sezon",     '<polyline points="22 12 18 12 15 21 9 3 6 12 2 12"/>',  "Statystyki drużynowe"),
        ni("players", "/zawodnicy", '<line x1="18" y1="20" x2="18" y2="10"/><line x1="12" y1="20" x2="12" y2="4"/><line x1="6" y1="20" x2="6" y2="14"/>', "Statystyki indywidualne"),
        '</div>',
        ni("history",  "/historia",   '<rect x="3" y="4" width="18" height="18" rx="2"/><line x1="3" y1="10" x2="21" y2="10"/>',  "Historia meczów"),
        ni("settings", "/ustawienia", '<circle cx="12" cy="12" r="3"/><path d="M19.07 4.93a10 10 0 0 1 0 14.14M4.93 4.93a10 10 0 0 0 0 14.14"/>',  "Ustawienia"),
        '<div style="height:0.5px;background:rgba(255,255,255,.08);margin:6px 10px 2px"></div>',
        '<div class="nav-section brand-text">Kontekst</div>',
        '<div style="padding:0 6px 4px">',
        p_klub, p_sezon, p_druz, hint,
        # Szybkie linki do filtrowanych widoków

        '</div>',
        '<div style="padding:4px 10px 4px">''<a href="/template/zapis" style="display:flex;align-items:center;justify-content:center;gap:6px;padding:7px 10px;border-radius:6px;background:rgba(255,255,255,.08);border:0.5px solid rgba(255,255,255,.15);text-decoration:none;width:100%"><svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="rgba(255,255,255,.7)" stroke-width="2"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg><span style="font-size:11px;color:rgba(255,255,255,.8)">Pobierz szablon meczu</span></a></div>',
        '<div style="padding:2px 10px 8px"><a href="/portal" target="_blank" style="display:flex;align-items:center;justify-content:center;gap:6px;padding:7px 10px;border-radius:6px;background:rgba(239,159,39,.15);border:0.5px solid rgba(239,159,39,.4);text-decoration:none;width:100%"><svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="#EF9F27" stroke-width="2"><path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"/><circle cx="12" cy="7" r="4"/></svg><span style="font-size:11px;color:#EF9F27;font-weight:600">Podgląd zawodnika</span></a></div>',
        '<div style="margin-top:auto;padding:1rem .75rem;border-top:1px solid rgba(255,255,255,.08)">',
        '<div style="font-size:.72rem;color:rgba(255,255,255,.5);margin-bottom:4px">Zalogowany jako</div>',
        '<div style="font-size:.8rem;color:#fff;font-weight:500;margin-bottom:8px">' + uname + '</div>',
        '<a href="/logout" style="display:block;text-align:center;padding:6px;background:rgba(255,255,255,.1);color:rgba(255,255,255,.7);border-radius:6px;font-size:.75rem;text-decoration:none;">Wyloguj</a>',
        '</div></div>',
        '<script>',
        'function toggleSidebar(){var s=document.getElementById("sidebar"),o=document.getElementById("sidebarOverlay"),h=document.getElementById("hamburger");s.classList.toggle("mobile-open");o.style.display=s.classList.contains("mobile-open")?"block":"none";h.classList.toggle("open");}',
        'function toggleSub(id){var el=document.getElementById(id),key=id==="team-sub"?"team":"stats",arr=document.getElementById("arr-"+key),open=el.style.display==="block";el.style.display=open?"none":"block";if(arr)arr.style.transform=open?"rotate(0deg)":"rotate(90deg)";}',
        'function togglePill(f){var dd=document.getElementById("dd-"+f),hdr=dd?dd.previousElementSibling:null,chev=document.getElementById("chev-"+f),isOpen=dd&&dd.style.display==="block";["klub","sezon","druzyna"].forEach(function(x){var d=document.getElementById("dd-"+x),h=d?d.previousElementSibling:null,c=document.getElementById("chev-"+x);if(d)d.style.display="none";if(h)h.classList.remove("is-open");if(c)c.style.transform="rotate(0deg)";});if(!isOpen&&dd){dd.style.display="block";if(hdr)hdr.classList.add("is-open");if(chev)chev.style.transform="rotate(180deg)";}  }',
        'function pickCtx(field,val){fetch("/set_context_ajax",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({field:field,value:val})}).then(function(){window.location.reload();});}',
        'window.addEventListener("resize",function(){if(window.innerWidth>767){document.getElementById("sidebar").classList.remove("mobile-open");document.getElementById("sidebarOverlay").style.display="none";document.getElementById("hamburger").classList.remove("open");}});',
        'document.addEventListener("click",function(e){if(!e.target.closest(".ctx-wrap")){["klub","sezon","druzyna"].forEach(function(f){var d=document.getElementById("dd-"+f),h=d?d.previousElementSibling:null,c=document.getElementById("chev-"+f);if(d)d.style.display="none";if(h)h.classList.remove("is-open");if(c)c.style.transform="rotate(0deg)";});}});',
        '</script>',
    ]
    return "\n".join(parts)


def html_response(html):
    """Zwraca HTML jako Response z pominięciem Jinja2."""
    return Response(html, mimetype='text/html')


def base(content, scripts="", active="home"):
    # Flash messages
    flash_html = ""
    try:
        from flask import get_flashed_messages
        msgs = get_flashed_messages(with_categories=True)
        for cat, msg in msgs:
            css = "flash-success" if cat == "success" else "flash-error"
            flash_html += f'<div class="{css} flash-msg">{msg}</div>'
    except: pass

    return f"""<!DOCTYPE html>
<html lang="pl"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<meta name="apple-mobile-web-app-capable" content="yes">
<title>Basket Kołcz Analytics</title>
{CSS}
</head>
<body>
{nav(active)}
<div class="main-content">
{flash_html}
{content}
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<script>
function initSortable(tableId, skipCol) {{
  var tbl = document.getElementById(tableId);
  if (!tbl) return;
  var state = {{col: -1, asc: true}};
  var ths = tbl.querySelectorAll('thead th');
  ths.forEach(function(th, ci) {{
    if (ci === skipCol) return;
    th.style.cursor = 'pointer';
    th.style.userSelect = 'none';
    th.style.whiteSpace = 'nowrap';
    th.addEventListener('click', function() {{
      var asc = (state.col === ci) ? !state.asc : false;
      state = {{col: ci, asc: asc}};
      ths.forEach(function(h, i) {{
        var base = h.getAttribute('data-label') || h.textContent.replace(/[\u25b2\u25bc]/g,'').trim();
        h.setAttribute('data-label', base);
        h.textContent = base + (i === ci ? (asc ? ' \u25b2' : ' \u25bc') : '');
      }});
      var tbody = tbl.querySelector('tbody');
      var rows = Array.from(tbody.querySelectorAll('tr'));
      rows.sort(function(a, b) {{
        var ca = a.cells[ci] ? a.cells[ci].textContent.trim() : '';
        var cb = b.cells[ci] ? b.cells[ci].textContent.trim() : '';
        var na = parseFloat(ca.replace('%','').replace('\u2014','').replace(',','.'));
        var nb = parseFloat(cb.replace('%','').replace('\u2014','').replace(',','.'));
        if (!isNaN(na) && !isNaN(nb)) return asc ? na-nb : nb-na;
        var fa = ca.match(/^([0-9]+)[/]([0-9]+)$/), fb = cb.match(/^([0-9]+)[/]([0-9]+)$/);
        if (fa && fb) {{ var va=parseInt(fa[1])/parseInt(fa[2]), vb=parseInt(fb[1])/parseInt(fb[2]); return asc ? va-vb : vb-va; }}
        return asc ? ca.localeCompare(cb) : cb.localeCompare(ca);
      }});
      rows.forEach(function(r) {{ tbody.appendChild(r); }});
    }});
  }});
}}
</script>
{scripts}
</body></html>"""

# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — STRONA GŁÓWNA
# ══════════════════════════════════════════════════════════════════════════════


@app.route("/")
@login_required
def index():
    try: init_db()
    except: pass

    user_name  = session.get("user_name", "Trener")
    first_name = user_name.split()[0] if user_name else "Trener"

    def dot(color, text):
        return ('<li style="font-size:10px;color:#666;padding:3px 0;display:flex;align-items:center;gap:5px">'
                '<span style="width:5px;height:5px;border-radius:50%;background:' + color + ';flex-shrink:0;display:inline-block"></span>'
                + text + '</li>')

    def feature_card(icon_bg, icon_stroke, icon_path, title, items):
        dots = "".join(dot(c, t) for c, t in items)
        return (
            '<div style="background:#fff;border-radius:10px;border:0.5px solid rgba(0,0,0,.07);padding:14px 16px">'
            '<div style="width:30px;height:30px;border-radius:8px;background:' + icon_bg + ';display:flex;align-items:center;justify-content:center;margin-bottom:10px">'
            '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="' + icon_stroke + '" stroke-width="2">' + icon_path + '</svg></div>'
            '<div style="font-size:12px;font-weight:600;color:#1a2b4a;margin-bottom:8px">' + title + '</div>'
            '<ul style="list-style:none;padding:0;margin:0">' + dots + '</ul>'
            '</div>'
        )

    def step_box(num, num_color, title, desc, link_label, link_href, link_color):
        return (
            '<div style="flex:1;background:rgba(255,255,255,.06);border-radius:8px;padding:12px 14px;border:0.5px solid rgba(255,255,255,.08)">'
            '<div style="display:flex;align-items:center;gap:8px;margin-bottom:10px">'
            '<div style="width:22px;height:22px;border-radius:50%;background:' + num_color + ';display:flex;align-items:center;justify-content:center;font-size:10px;font-weight:700;color:#fff;flex-shrink:0">' + str(num) + '</div>'
            '<div style="font-size:11px;font-weight:600;color:#fff">' + title + '</div></div>'
            '<div style="font-size:10px;color:rgba(255,255,255,.5);line-height:1.6;margin-bottom:10px">' + desc + '</div>'
            '<a href="' + link_href + '" style="font-size:10px;color:' + link_color + ';font-weight:600;text-decoration:none">' + link_label + '</a>'
            '</div>'
        )

    arrow = '<div style="color:rgba(255,255,255,.25);font-size:18px;align-self:center;flex-shrink:0">\u203a</div>'

    s1 = step_box(1, '#1D9E75',
        'Dodaj klub i dru\u017cyny',
        'Skonfiguruj struktur\u0119 \u2014 klub, sezon, dru\u017cyn\u0119. Bez tego aplikacja nie wie dla kogo zbiera\u0107 dane.',
        'Struktura klub\u00f3w \u2192', '/druzyny', '#1D9E75')
    s2 = step_box(2, '#378ADD',
        'Dodaj zawodnik\u00f3w',
        'Uzupe\u0142nij sk\u0142ad r\u0119cznie lub zaimportuj z pliku CSV. Numery koszulek musz\u0105 zgadza\u0107 si\u0119 z kodowaniem.',
        'Struktura klub\u00f3w \u2192', '/druzyny', '#378ADD')
    s3 = step_box(3, '#EF9F27',
        'Pobierz szablon meczu',
        'Pobierz SZABLON_MECZU_v3.xlsx z sidebara. Wype\u0142nij META i zakoduj akcje obu dru\u017cyn.',
        '\u2193 Pobierz szablon', '/template/zapis', '#EF9F27')
    s4 = step_box(4, '#534AB7',
        'Wgraj mecz i analizuj',
        'Wgraj plik przez \u201eWgraj mecz\u201d. Statystyki, pi\u0105tki i profil zawodnika s\u0105 dost\u0119pne od razu.',
        'Wgraj mecz \u2192', '/upload', '#AFA9EC')

    c1 = feature_card('#E6F1FB', '#185FA5',
        '<polyline points="22 12 18 12 15 21 9 3 6 12 2 12"/>',
        'Analiza mecz\u00f3w i statystyk',
        [('#1D9E75', 'Statystyki dru\u017cyny \u2014 ORtg, PPP, eFG%, TS%'),
         ('#1D9E75', 'Profile indywidualne zawodnik\u00f3w'),
         ('#1D9E75', 'Pi\u0105tki OFF / DEF / NET RTG'),
         ('#EF9F27', 'Historia wszystkich mecz\u00f3w sezonu')])
    c2 = feature_card('#FAEEDA', '#854F0B',
        '<path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M23 21v-2a4 4 0 0 0-3-3.87M16 3.13a4 4 0 0 1 0 7.75"/>',
        'Zarz\u0105dzanie dru\u017cyn\u0105',
        [('#1D9E75', 'Wiele klub\u00f3w, sezon\u00f3w i dru\u017cyn'),
         ('#1D9E75', 'Sk\u0142ad zawodnik\u00f3w z pozycjami'),
         ('#1D9E75', 'Import zbiorczy z Excel / CSV'),
         ('#EF9F27', 'Prze\u0142\u0105czanie kontekstu z sidebara')])
    c3 = feature_card('#E1F5EE', '#0F6E56',
        '<path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="12" y1="18" x2="12" y2="12"/><line x1="9" y1="15" x2="15" y2="15"/>',
        'Szablony i import danych',
        [('#1D9E75', 'Szablon zapisu meczu v3 .xlsx'),
         ('#1D9E75', 'Szablon sk\u0142adu zawodnik\u00f3w .csv'),
         ('#1D9E75', 'Automatyczny odczyt daty z META'),
         ('#EF9F27', 'Eksport raport\u00f3w do Excel')])

    
    court_svg = '<svg width="140" height="88" viewBox="0 0 160 100" fill="none" stroke="rgba(255,255,255,.15)" stroke-width="1.3" style="flex-shrink:0"><rect x="3" y="3" width="154" height="94" rx="2"/><line x1="80" y1="3" x2="80" y2="97"/><circle cx="80" cy="50" r="10"/><circle cx="80" cy="50" r="1.8" fill="rgba(255,255,255,.15)" stroke="none"/><rect x="3" y="34" width="26" height="32" rx="1"/><path d="M29 34 A16 16 0 0 1 29 66"/><line x1="3" y1="18" x2="38" y2="18"/><path d="M38 18 A34 34 0 0 1 38 82"/><line x1="38" y1="82" x2="3" y2="82"/><rect x="131" y="34" width="26" height="32" rx="1"/><path d="M131 34 A16 16 0 0 0 131 66"/><line x1="157" y1="18" x2="122" y2="18"/><path d="M122 18 A34 34 0 0 0 122 82"/><line x1="122" y1="82" x2="157" y2="82"/></svg>'

    content = (
        # ── HERO ──
        '<div style="background:#1a2b4a;border-radius:11px;padding:20px 24px;display:flex;align-items:center;'
        'justify-content:space-between;margin-bottom:14px;overflow:hidden">'
        '<div>'
        '<div style="font-size:19px;font-weight:600;color:#fff;margin-bottom:4px">'
        'Cześć, <span style="color:#EF9F27">' + first_name + '</span>!</div>'
        '<div style="font-size:11px;color:rgba(255,255,255,.45);line-height:1.7;max-width:400px">'
        'Basket Kołcz Analytics Platform — zarządzaj drużynami, analizuj mecze, śledź zawodników.'
        '</div></div>' + court_svg + '</div>'

        # ── SEKCJA LABEL ──
        + '<div style="font-size:10px;font-weight:600;color:#aaa;text-transform:uppercase;'
        'letter-spacing:.6px;margin-bottom:10px">Konfiguracja — zacznij tutaj</div>'

        # ── INSTRUKCJA (ciemna) ──
        + '<div style="background:#1a2b4a;border-radius:10px;padding:16px 20px;margin-bottom:14px">'

        # Nagłówek instrukcji
        + '<div style="display:flex;align-items:center;gap:8px;margin-bottom:14px;'
        'padding-bottom:12px;border-bottom:0.5px solid rgba(255,255,255,.08)">'
        '<div style="width:26px;height:26px;border-radius:7px;background:rgba(239,159,39,.15);'
        'display:flex;align-items:center;justify-content:center">'
        '<svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="#EF9F27" stroke-width="2">'
        '<circle cx="12" cy="12" r="10"/><line x1="12" y1="16" x2="12" y2="12"/>'
        '<line x1="12" y1="8" x2="12.01" y2="8"/></svg></div>'
        '<div style="font-size:12px;font-weight:600;color:#fff">Jak zacząć — instrukcja krok po kroku</div>'
        '</div>'

        # 4 kroki ze strzałkami
        + '<div style="display:flex;align-items:stretch;gap:8px">'
        + s1 + arrow + s2 + arrow + s3 + arrow + s4
        + '</div>'

        # Nota o kontekście
        + '<div style="margin-top:12px;padding:10px 14px;background:rgba(255,255,255,.04);'
        'border-radius:6px;border-left:2px solid rgba(239,159,39,.4);display:flex;align-items:center;gap:8px">'
        '<svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="#EF9F27" stroke-width="2" style="flex-shrink:0">'
        '<circle cx="12" cy="12" r="10"/><line x1="12" y1="16" x2="12" y2="12"/>'
        '<line x1="12" y1="8" x2="12.01" y2="8"/></svg>'
        '<div style="font-size:10px;color:rgba(255,255,255,.5)">'
        'Pamiętaj: przed analizą ustaw kontekst w sidebarze (klub → sezon → drużyna). '
        'Wszystkie widoki filtrują się wg aktywnego kontekstu.'
        '</div></div>'
        + '</div>'

        # ── CO MOŻESZ ROBIĆ ──
        + '<div style="font-size:10px;font-weight:600;color:#aaa;text-transform:uppercase;'
        'letter-spacing:.6px;margin-bottom:10px">Co możesz robić w aplikacji</div>'
        + '<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px">'
        + c1 + c2 + c3
        + '</div>'
    )

    return html_response(base(content, active="home"))

# ══════════════════════════════════════════════════════════════════════════════
# WALIDACJA
# ══════════════════════════════════════════════════════════════════════════════

VALID_CODES = {
    # Podstawowe kody punktowe
    "2","0/2","3","0/3","BR","P","F","T",
    # Rzuty wolne standalone (1/1W, 0/1W — nowy szablon v2)
    "1/1W","0/1W",
    # Defensywa — nowy szablon v2
    "STL","BLK",
    # Kody złożone (rzut + rzuty wolne)
    "2+1/1W","2+0/1W","3+1/1W","3+0/1W",
    # Stare kody — zachowane dla wstecznej kompatybilności
    "2+1","2+0","3+1","3+0",
    "2D","0/2D","2D+1","2D+0/1W","2D+1/1W",
    "1/2W","2/2W","0/2W",
    "1/3W","2/3W","3/3W","0/3W",
    "1/2WL","2/2WL","0/2WL",
    "1/1WT","0/1WT",
    "0D+0/2W","0D+1/2W","0D+2/2W"
}

def read_meta(wb):
    """Odczytaj arkusz META i zwróć słownik danych"""
    meta = {}
    for sn in wb.sheetnames:
        if sn.upper() == "META":
            ws = wb[sn]
            for row in ws.iter_rows(min_row=2, values_only=False):
                if not row[0].value: continue
                key = str(row[0].value).strip().lower()
                # Obsłuż datetime z Excela
                raw_val = row[1].value if len(row) > 1 else None
                from datetime import datetime as _dt, date as _date
                if isinstance(raw_val, (_dt, _date)):
                    val = raw_val  # zachowaj jako obiekt — obsłużony w _do_save
                else:
                    val = str(raw_val).strip() if raw_val is not None else ""
                if "drużyna a" in key or "druzyna a" in key: meta["nazwa_a"] = val
                if "drużyna b" in key or "druzyna b" in key: meta["nazwa_b"] = val
                if "wynik a"   in key: meta["wynik_a"] = str(val) if val else ""
                if "wynik b"   in key: meta["wynik_b"] = str(val) if val else ""
                if "data"      in key: meta["data"] = raw_val  # surowa wartość
                if "rozgrywki" in key: meta["rozgrywki"] = val
                if "runda"     in key: meta["runda"] = val
                if "kolejka"   in key: meta["kolejka"] = val
                if "miejsce"   in key: meta["miejsce"] = val
                if "uwagi"     in key: meta["uwagi"] = val
                if "nazwa pliku" in key: meta["nazwa_pliku"] = val
            break
    return meta

def validate_workbook(wb):
    """
    Zwraca dict:
      errors   — błędy krytyczne (lista str)
      warnings — ostrzeżenia (lista str)
      info     — informacje (lista str)
      meta     — dane z META
      names    — (name_a, name_b)
    """
    errors = []; warnings = []; info = []

    sheets_data = [s for s in wb.sheetnames if s.upper() not in ("META","KODY","LEGENDA")]

    if len(sheets_data) < 2:
        errors.append(f"Brakuje arkuszy danych — znaleziono {len(sheets_data)}, wymagane 2 (GTK + przeciwnik)")
        return {"errors":errors,"warnings":warnings,"info":info,"meta":{},"names":("","") }

    name_a, name_b = sheets_data[0], sheets_data[1]
    meta = read_meta(wb)

    # ── 1. Nazwy drużyn ────────────────────────────────────────────────────
    meta_a = meta.get("nazwa_a","")
    meta_b = meta.get("nazwa_b","")
    if meta_a and meta_a not in ("TWOJA_DRUZYNA","TWOJA_DRUŻYNA","-",""):
        if meta_a.upper() != name_a.upper():
            warnings.append(
                f"Nazwa drużyny A w META (<b>{meta_a}</b>) "
                f"różni się od nazwy arkusza (<b>{name_a}</b>). "
                f"Aplikacja użyje nazwy arkusza."
            )
        else:
            info.append(f"✓ Nazwa drużyny A: <b>{name_a}</b> — zgodna z META")
    if meta_b and meta_b not in ("RYWAL","-",""):
        if meta_b.upper() != name_b.upper():
            warnings.append(
                f"Nazwa drużyny B w META (<b>{meta_b}</b>) "
                f"różni się od nazwy arkusza (<b>{name_b}</b>). "
                f"Aplikacja użyje nazwy arkusza."
            )
        else:
            info.append(f"✓ Nazwa drużyny B: <b>{name_b}</b> — zgodna z META")

    # ── 2. Parsuj i sprawdź wyniki ─────────────────────────────────────────
    stats_a = parse_team_sheet(wb[name_a], sheet_type="A")
    stats_b = parse_team_sheet(wb[name_b], sheet_type="B")
    suma_a  = suma_quarters(stats_a)
    suma_b  = suma_quarters(stats_b)
    pts_a   = suma_a.get("pts",0)
    pts_b   = suma_b.get("pts",0)

    # Wynik końcowy z META
    try:
        meta_pts_a = int(meta.get("wynik_a","")) if meta.get("wynik_a") else None
        meta_pts_b = int(meta.get("wynik_b","")) if meta.get("wynik_b") else None
    except: meta_pts_a = meta_pts_b = None

    if meta_pts_a is not None:
        q_pts_a = [stats_a["quarter"].get(q,{}).get("pts",0) for q in [1,2,3,4]]
        if meta_pts_a != pts_a:
            errors.append({
                "msg": f"Wynik <b>{name_a}</b> w META: <b>{meta_pts_a}</b> pkt, suma z kodowania: <b>{pts_a}</b> pkt (różnica: {pts_a-meta_pts_a:+d})",
                "quarters": q_pts_a,
                "total": pts_a,
                "meta": meta_pts_a,
            })
        else:
            info.append(f"✓ Wynik {name_a}: <b>{pts_a}</b> pkt — zgodny z META")

    if meta_pts_b is not None:
        q_pts_b = [stats_b["quarter"].get(q,{}).get("pts",0) for q in [1,2,3,4]]
        if meta_pts_b != pts_b:
            errors.append({
                "msg": f"Wynik <b>{name_b}</b> w META: <b>{meta_pts_b}</b> pkt, suma z kodowania: <b>{pts_b}</b> pkt (różnica: {pts_b-meta_pts_b:+d})",
                "quarters": q_pts_b,
                "total": pts_b,
                "meta": meta_pts_b,
            })
        else:
            info.append(f"✓ Wynik {name_b}: <b>{pts_b}</b> pkt — zgodny z META")

    # ── 3. Wynik per kwarta ────────────────────────────────────────────────
    for sheet_name, stats in [(name_a, stats_a), (name_b, stats_b)]:
        total_check = sum(stats["quarter"].get(q,{}).get("pts",0) for q in [1,2,3,4])
        suma = suma_quarters(stats)
        if total_check != suma.get("pts",0):
            warnings.append(
                f"⚠️ {sheet_name}: suma kwart ({total_check}) "
                f"≠ łączna suma punktów ({suma.get('pts',0)})"
            )
        else:
            q_pts = [stats["quarter"].get(q,{}).get("pts",0) for q in [1,2,3,4]]
            info.append(f"✓ {sheet_name} per kwarta: {q_pts[0]}+{q_pts[1]}+{q_pts[2]}+{q_pts[3]} = <b>{total_check}</b>")

    # ── 4. Brakujące dane (puste kolumny A/B/C) ────────────────────────────
    for sheet_name in [name_a, name_b]:
        ws = wb[sheet_name]
        empty_rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=200, values_only=True), 2):
            if not any(v is not None for v in row[:4]): break
            missing = []
            if row[0] is None: missing.append("Kwarta(A)")
            if row[2] is None: missing.append("Kod(C)")
            if missing:
                empty_rows.append(f"wiersz {i}: brak {', '.join(missing)}")
        if empty_rows:
            sample = empty_rows[:5]
            warnings.append(
                f"⚠️ {sheet_name} — brakujące dane w {len(empty_rows)} wierszach: "
                f"{'; '.join(sample)}"
                + (f" (i {len(empty_rows)-5} więcej...)" if len(empty_rows)>5 else "")
            )
        else:
            info.append(f"✓ {sheet_name}: wszystkie wiersze mają wymagane dane")

    # ── 5. Nieznane kody akcji ─────────────────────────────────────────────
    for sheet_name in [name_a, name_b]:
        ws = wb[sheet_name]
        unknown = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=500, values_only=True), 2):
            if not any(v is not None for v in row[:4]): break
            raw_c = str(row[2]).strip() if row[2] is not None else ""
            if not raw_c: continue
            for code in [c.strip().upper() for c in raw_c.split(";") if c.strip()]:
                if code not in VALID_CODES:
                    if code not in unknown: unknown[code] = []
                    unknown[code].append(i)
        if unknown:
            details = "; ".join(
                f"<b>{k}</b> (wiersze: {', '.join(map(str,v[:3]))}{'...' if len(v)>3 else ''})"
                for k,v in list(unknown.items())[:8]
            )
            errors.append(
                f"❌ {sheet_name} — nieznane kody akcji: {details}"
            )
        else:
            info.append(f"✓ {sheet_name}: wszystkie kody akcji są prawidłowe")

    # ── 6. Finishery (kolumna K) ──────────────────────────────────────────
    for sheet_name in [name_a, name_b]:
        ws = wb[sheet_name]
        total_actions = 0
        missing_fin = 0
        hash_prefix = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=500, values_only=True), 2):
            if not any(v is not None for v in row[:4]): break
            raw_c = str(row[2]).strip() if len(row)>2 and row[2] is not None else ""
            if not raw_c: continue
            codes = [c.strip() for c in raw_c.split(";") if c.strip()]
            fins  = str(row[10]).strip().split(";") if len(row)>10 and row[10] is not None else []
            for ai, code in enumerate(codes):
                if code.upper() in ("BR","P","F","STL","BLK","T"): continue  # nie wymagają finishera
                total_actions += 1
                fin_raw = fins[ai].strip() if ai < len(fins) else ""
                if not fin_raw:
                    missing_fin += 1
                elif "#" in fin_raw:
                    hash_prefix += 1
        if hash_prefix > 0:
            warnings.append(
                f"⚠️ {sheet_name} — finishery (kol K) mają prefix '#' ({hash_prefix} akcji). "
                f"Parser usunie '#' automatycznie — dane zawodników zostaną wczytane."
            )
        elif total_actions > 0 and missing_fin == total_actions:
            warnings.append(
                f"⚠️ {sheet_name} — brak finisherów (kol K pusta). "
                f"Statystyki indywidualne zawodników nie będą dostępne."
            )
        elif missing_fin > 0:
            warnings.append(
                f"⚠️ {sheet_name} — brak finishera w {missing_fin}/{total_actions} akcjach. "
                f"Niektóre statystyki zawodników mogą być niepełne."
            )
        else:
            if total_actions > 0:
                info.append(f"✓ {sheet_name}: finishery kompletne ({total_actions} akcji)")

    # ── 7. Asysty i zbiórki ────────────────────────────────────────────────
    ws_a = wb[name_a]
    has_ast = has_oreb = has_dreb = False
    for row in ws_a.iter_rows(min_row=2, max_row=500, values_only=True):
        if not any(v is not None for v in row[:4]): break
        if len(row)>11 and row[11] is not None and str(row[11]).strip(): has_ast = True
        if len(row)>12 and row[12] is not None and str(row[12]).strip(): has_oreb = True
        if len(row)>13 and row[13] is not None and str(row[13]).strip(): has_dreb = True
    missing_cols = []
    if not has_ast:  missing_cols.append("Asysta (kol L)")
    if not has_oreb: missing_cols.append("Zbiórkа OFF (kol M)")
    if not has_dreb: missing_cols.append("Zbiórkа DEF (kol N)")
    if missing_cols:
        warnings.append(
            f"⚠️ {name_a} — brak danych: {', '.join(missing_cols)}. "
            f"Statystyki AST/OREB/DREB będą wynosiły 0."
        )
    else:
        info.append(f"✓ {name_a}: asysty i zbiórki są wypełnione")

    # ── 8. Arkusz META — kompletność ──────────────────────────────────────
    meta_missing = []
    if not meta.get("data"):       meta_missing.append("data meczu")
    if not meta.get("wynik_a"):    meta_missing.append("wynik A")
    if not meta.get("wynik_b"):    meta_missing.append("wynik B")
    if not meta.get("rozgrywki"):  meta_missing.append("rozgrywki")
    if not meta.get("runda"):      meta_missing.append("runda/kolejka")
    if not meta.get("miejsce"):    meta_missing.append("miejsce")
    if meta_missing:
        warnings.append(
            f"⚠️ META — brakujące pola: {', '.join(meta_missing)}. "
            f"Mecz zostanie zapisany bez tych danych."
        )
    else:
        info.append(f"✓ META: wszystkie pola wypełnione")

    # ── 9. Liczba wierszy z danymi ─────────────────────────────────────────
    for sheet_name in [name_a, name_b]:
        ws = wb[sheet_name]
        n_rows = sum(1 for row in ws.iter_rows(min_row=2, max_row=1000, values_only=True)
                     if any(v is not None and str(v).strip() for v in row[:4]))
        info.append(f"✓ {sheet_name}: {n_rows} zakodowanych posiadań")

    return {
        "errors":   errors,
        "warnings": warnings,
        "info":     info,
        "meta":     meta,
        "names":    (name_a, name_b),
        "pts":      (pts_a, pts_b),
    }

# ══════════════════════════════════════════════════════════════════════════════
# UPLOAD
# ══════════════════════════════════════════════════════════════════════════════

import tempfile, base64

def _do_save(wb, name_gtk, name_opp, sezon, data_meczu, team_id=None, nr_to_player=None):
    """Właściwy zapis meczu do bazy"""
    # Konwertuj pusty string na None
    if data_meczu is not None and str(data_meczu).strip() == "":
        data_meczu = None
    try:
        init_db()
    except: pass

    meta = read_meta(wb)

    # Nazwy z META
    display_gtk = meta.get("nazwa_a","") or name_gtk
    display_opp = meta.get("nazwa_b","") or name_opp
    if display_gtk in ("TWOJA_DRUZYNA","TWOJA_DRUŻYNA","-",""): display_gtk = name_gtk
    if display_opp in ("RYWAL","-",""): display_opp = name_opp

    # Data — META ma priorytet nad formularzem
    meta_data = meta.get("data")
    if meta_data:
        from datetime import datetime as dt2, date
        # Jeśli Excel zwrócił obiekt datetime/date
        if isinstance(meta_data, (dt2, date)):
            data_meczu = meta_data.strftime('%Y-%m-%d') if isinstance(meta_data, dt2) else meta_data.isoformat()
        else:
            raw = str(meta_data).strip().replace(" ","").replace("\n","")
            # Spróbuj różnych formatów daty
            parsed = None
            for fmt in ('%d.%m.%Y','%d/%m/%Y','%Y-%m-%d','%d-%m-%Y',
                        '%Y.%m.%d','%d.%m.%y','%d/%m/%y'):
                try:
                    parsed = dt2.strptime(raw, fmt).strftime('%Y-%m-%d')
                    break
                except: pass
            if parsed:
                data_meczu = parsed
            # Jeśli nie udało się sparsować — zostaw datę z formularza
    
    # Sezon z META jeśli dostępny
    if meta.get("rozgrywki") and not sezon:
        sezon = meta["rozgrywki"]

    stats_gtk = parse_team_sheet(wb[name_gtk], sheet_type="A")
    stats_opp = parse_team_sheet(wb[name_opp], sheet_type="B")
    try:
        def_lineups = build_gtk_def_lineups(wb[name_gtk], wb[name_opp])
    except Exception:
        def_lineups = None
    match_id  = save_match_to_db(
        display_opp, display_gtk, sezon, data_meczu,
        stats_gtk, stats_opp,
        rozgrywki=str(meta.get("rozgrywki","") or ""),
        runda=str(meta.get("runda","") or ""),
        kolejka=str(meta.get("kolejka","") or ""),
        miejsce=str(meta.get("miejsce","") or ""),
        def_lineups=def_lineups,
        team_name_a=display_gtk,
        team_name_b=display_opp,
        team_id=team_id,
        nr_to_player=nr_to_player,
    )
    # Zapisz oryginalny plik xlsx na dysku
    try:
        import io as _io2
        file_path = None
        wb_bytes_out = _io2.BytesIO()
        wb.save(wb_bytes_out)
        wb_bytes_out.seek(0)
        file_path = os.path.join(MATCH_FILES_DIR, f"{match_id}.xlsx")
        with open(file_path, "wb") as fp:
            fp.write(wb_bytes_out.read())
        # Zapisz ścieżkę i wersję parsera w bazie
        db2 = get_db(); cur2 = db2.cursor()
        cur2.execute("UPDATE matches SET file_path=%s, parser_version=%s WHERE id=%s",
                     (file_path, PARSER_VERSION, match_id))
        db2.commit(); cur2.close()
    except Exception as _fe:
        pass  # Błąd zapisu pliku nie blokuje meczu

    for k in ['pt','ps','pd','vr']: session.pop(k, None)
    session["last_match_id"] = match_id
    flash(f"✓ Mecz {display_gtk} vs {display_opp} zapisany pomyślnie!","success")
    return redirect(url_for("mecz", match_id=match_id))


# Tymczasowy storage dla plików oczekujących na potwierdzenie
# Używamy katalogu /tmp (dostępny na Render)
PENDING_DIR = "/tmp/basketkolcz_pending"
os.makedirs(PENDING_DIR, exist_ok=True)




@app.route("/upload", methods=["GET"])
@login_required
def upload_page():
    try: init_db()
    except: pass
    import json as _j
    db = get_db(); cur = db.cursor()

    # Build clubs→seasons→teams tree for dropdown
    cur.execute("""
        SELECT c.id as cid, c.name as cname,
               s.id as sid, s.name as sname,
               t.id as tid, t.name as tname
        FROM clubs c
        JOIN seasons s ON s.club_id=c.id
        JOIN teams t ON t.season_id=s.id
        ORDER BY c.name, s.name, t.name
    """)
    tree_rows = cur.fetchall()

    team_options = ""
    for r in tree_rows:
        label = f"{r['cname']} · {r['sname']} · {r['tname']}"
        team_options += f'<option value="{r["tid"]}">{label}</option>'

    # Last used team_id from settings
    last_team = get_setting("last_upload_team") or ""
    cur.close()

    content = f"""
<div style="max-width:600px;margin:0 auto">
<div style="background:#fff;border-radius:12px;border:0.5px solid rgba(0,0,0,.08);overflow:hidden;margin-bottom:16px">

  <!-- Nagłówek -->
  <div style="background:#1a2b4a;padding:12px 16px;display:flex;align-items:center;justify-content:space-between">
    <span style="color:#fff;font-size:11px;font-weight:600;letter-spacing:.5px;text-transform:uppercase">Wgrywanie meczu</span>
    <a href="/template/zapis" style="display:inline-flex;align-items:center;gap:5px;font-size:10px;font-weight:600;color:#AFA9EC;text-decoration:none">
      <svg width="10" height="10" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
      Pobierz szablon
    </a>
  </div>

  <form method="POST" action="/upload" enctype="multipart/form-data" id="upload-form">
  <input type="file" name="file" id="file-input" accept=".xlsx" required style="display:none" onchange="handleFileDrop(this.files[0])">

  <!-- ══════════════════════════════════════
       KROK 1 — PLIK
  ══════════════════════════════════════ -->
  <div style="padding:14px 16px;border-bottom:0.5px solid #f0f0f0;display:flex;align-items:flex-start;gap:10px">
    <div id="num1" style="width:22px;height:22px;border-radius:50%;background:#EF9F27;color:#fff;font-size:10px;font-weight:700;display:flex;align-items:center;justify-content:center;flex-shrink:0;margin-top:2px">1</div>
    <div style="flex:1">
      <div style="font-size:12px;font-weight:500;color:#1a2b4a;margin-bottom:8px">Plik meczu (.xlsx)</div>
      <div id="drop-zone"
        onclick="document.getElementById('file-input').click()"
        ondragover="dzOver(event)" ondragleave="dzLeave(event)" ondrop="dzDrop(event)"
        style="border:2px dashed #d0d7e2;border-radius:10px;padding:24px 16px;text-align:center;cursor:pointer;background:#f8faff;transition:border-color .15s,background .15s">
        <div style="width:36px;height:36px;background:#E6F1FB;border-radius:9px;display:flex;align-items:center;justify-content:center;margin:0 auto 10px">
          <svg width="17" height="17" viewBox="0 0 24 24" fill="none" stroke="#378ADD" stroke-width="2"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
        </div>
        <div style="font-size:12px;font-weight:600;color:#1a2b4a;margin-bottom:3px">Przeciągnij i upuść plik tutaj</div>
        <div style="font-size:11px;color:#aaa;margin-bottom:10px">lub kliknij aby wybrać z dysku</div>
        <div style="display:inline-flex;align-items:center;gap:5px;padding:6px 14px;border-radius:7px;background:#1a2b4a;color:#fff;font-size:11px;font-weight:600">
          <svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="#fff" stroke-width="2"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
          Wybierz plik
        </div>
      </div>
      <div id="file-chosen" style="display:none;align-items:center;gap:10px;padding:10px 12px;background:#E1F5EE;border:1px solid #9FE1CB;border-radius:8px">
        <div style="width:30px;height:30px;background:#1D9E75;border-radius:6px;display:flex;align-items:center;justify-content:center;flex-shrink:0">
          <svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="#fff" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/></svg>
        </div>
        <div style="flex:1;min-width:0">
          <div id="fname" style="font-size:12px;font-weight:600;color:#085041;overflow:hidden;text-overflow:ellipsis;white-space:nowrap"></div>
          <div id="fsize" style="font-size:10px;color:#0F6E56;margin-top:1px"></div>
        </div>
        <div id="file-loading" style="display:none;font-size:10px;color:#0F6E56">⏳ Analizowanie…</div>
        <button type="button" onclick="removeFile()" style="font-size:10px;color:#888;background:rgba(0,0,0,.06);border:none;border-radius:4px;padding:4px 8px;cursor:pointer;flex-shrink:0">✕ Usuń</button>
      </div>
    </div>
  </div>

  <!-- ══════════════════════════════════════
       KROK 2 — WALIDACJA
  ══════════════════════════════════════ -->
  <div id="step-valid" style="display:none;padding:14px 16px;border-bottom:0.5px solid #f0f0f0;flex-direction:column;gap:0">
    <div style="display:flex;align-items:flex-start;gap:10px">
      <div id="num2" style="width:22px;height:22px;border-radius:50%;background:#ccc;color:#fff;font-size:10px;font-weight:700;display:flex;align-items:center;justify-content:center;flex-shrink:0;margin-top:2px">2</div>
      <div style="flex:1">
        <div style="font-size:12px;font-weight:500;color:#1a2b4a;margin-bottom:8px">Walidacja pliku</div>
        <div id="valid-content"></div>
      </div>
    </div>
  </div>

  <!-- ══════════════════════════════════════
       KROK 3 — META + DRUŻYNA + ZAWODNICY
  ══════════════════════════════════════ -->
  <div id="step-3" style="display:none;padding:14px 16px;border-bottom:0.5px solid #f0f0f0;flex-direction:column;gap:0">
    <div style="display:flex;align-items:flex-start;gap:10px">
      <div id="num3" style="width:22px;height:22px;border-radius:50%;background:#ccc;color:#fff;font-size:10px;font-weight:700;display:flex;align-items:center;justify-content:center;flex-shrink:0;margin-top:2px">3</div>
      <div style="flex:1">

        <!-- META -->
        <div style="font-size:12px;font-weight:500;color:#1a2b4a;margin-bottom:8px">Dane meczu (META)</div>
        <div id="meta-grid" style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:12px"></div>

        <!-- separator -->
        <div style="height:0.5px;background:#f0f0f0;margin:10px 0"></div>

        <!-- DRUŻYNA -->
        <div style="font-size:12px;font-weight:500;color:#1a2b4a;margin-bottom:6px">Drużyna</div>
        <select name="team_id" id="team-sel"
                style="width:100%;padding:7px 10px;border:0.5px solid #ddd;border-radius:8px;font-size:12px;background:#fff;outline:none;margin-bottom:12px"
                onchange="onTeamChange(this.value)">
          <option value="">— wybierz drużynę —</option>
          {team_options}
        </select>

        <!-- separator -->
        <div style="height:0.5px;background:#f0f0f0;margin:4px 0 10px"></div>

        <!-- ZAWODNICY -->
        <div style="display:flex;align-items:center;margin-bottom:8px">
          <div style="font-size:12px;font-weight:500;color:#1a2b4a;flex:1">Przypisz zawodników</div>
          <div id="assign-hint" style="font-size:10px;color:#aaa">Najpierw wybierz drużynę</div>
        </div>
        <div id="roster-wrap">
          <div style="text-align:center;padding:10px 0;font-size:11px;color:#bbb">Wybierz drużynę aby przypisać zawodników</div>
        </div>
        <input type="hidden" name="mappings_json" id="mappings-json" value="{{}}">

      </div>
    </div>
  </div>

  <!-- FOOTER -->
  <div id="step-footer" style="display:none;padding:14px 16px;justify-content:flex-end">
    <button type="submit"
            style="background:#EF9F27;color:#fff;border:none;border-radius:8px;padding:9px 26px;font-size:13px;font-weight:600;cursor:pointer">
      Zapisz mecz
    </button>
  </div>

  </form>
</div>
</div>

<script>
var ROSTER = [];
var NR_STATS = {{}};

/* ── Drag & drop ── */
function dzOver(e) {{
  e.preventDefault();
  var dz=document.getElementById('drop-zone');
  dz.style.borderColor='#378ADD'; dz.style.background='#E6F1FB';
}}
function dzLeave(e) {{
  var dz=document.getElementById('drop-zone');
  dz.style.borderColor='#d0d7e2'; dz.style.background='#f8faff';
}}
function dzDrop(e) {{
  e.preventDefault(); dzLeave(e);
  var f=e.dataTransfer.files[0];
  if(f && f.name.endsWith('.xlsx')) {{
    var dt=new DataTransfer(); dt.items.add(f);
    document.getElementById('file-input').files=dt.files;
    handleFileDrop(f);
  }}
}}

/* ── Krok 1: Plik wybrany → wywołaj preview ── */
function handleFileDrop(f) {{
  if(!f) return;
  document.getElementById('drop-zone').style.display='none';
  document.getElementById('file-chosen').style.display='flex';
  document.getElementById('fname').textContent=f.name;
  document.getElementById('fsize').textContent=(f.size/1024).toFixed(1)+' KB';
  document.getElementById('file-loading').style.display='inline';
  document.getElementById('num1').style.background='#1D9E75';

  // Reset późniejszych kroków
  ['step-valid','step-3','step-footer'].forEach(function(id) {{
    var el=document.getElementById(id); if(el) el.style.display='none';
  }});
  NR_STATS={{}};

  var fd=new FormData();
  fd.append('file', f);
  fetch('/upload/preview', {{method:'POST', body:fd}})
    .then(function(r){{ return r.json(); }})
    .then(function(data){{
      document.getElementById('file-loading').style.display='none';
      if(data.error){{ showValidError(data.error); return; }}
      NR_STATS = data.stats || {{}};
      showValidation(data);
    }})
    .catch(function(e){{
      document.getElementById('file-loading').style.display='none';
      showValidError('Błąd połączenia: '+e);
    }});
}}

function removeFile() {{
  document.getElementById('file-input').value='';
  document.getElementById('drop-zone').style.display='block';
  document.getElementById('file-chosen').style.display='none';
  document.getElementById('num1').style.background='#EF9F27';
  ['step-valid','step-3','step-footer'].forEach(function(id) {{
    var el=document.getElementById(id); if(el) el.style.display='none';
  }});
  NR_STATS={{}};
}}

/* ── Krok 2: Walidacja ── */
function showValidError(msg) {{
  var sv=document.getElementById('step-valid');
  sv.style.display='flex';
  document.getElementById('num2').style.background='#e53935';
  document.getElementById('valid-content').innerHTML=
    '<div style="background:#fff0f0;border:0.5px solid #f5c6cb;border-radius:7px;padding:9px 12px;font-size:11px;color:#8b1a1a">⛔ '+msg+'</div>';
}}

function showValidation(data) {{
  var errors   = data.errors   || [];
  var warnings = data.warnings || [];
  var info     = data.info     || [];
  var hasErr   = errors.length > 0;
  var hasWarn  = warnings.length > 0;
  var h = '';

  // Błędy krytyczne
  errors.forEach(function(e) {{
    var msg = (e.type==='dict') ? e.msg : e.msg;
    var qstr = '';
    if(e.type==='dict' && e.quarters && e.quarters.length) {{
      qstr = ' <span style="font-size:10px;color:#999">(Q: '+e.quarters.join('+')+'='+e.total+' vs META:'+e.meta+')</span>';
    }}
    h += '<div style="background:#fff0f0;border:0.5px solid #f5c6cb;border-radius:7px;padding:7px 10px;font-size:11px;color:#8b1a1a;margin-bottom:5px">⛔ <span>'+msg+'</span>'+qstr+'</div>';
  }});

  // Ostrzeżenia
  warnings.forEach(function(w) {{
    h += '<div style="background:#fffbea;border:0.5px solid #f5e199;border-radius:7px;padding:7px 10px;font-size:11px;color:#856404;margin-bottom:5px">⚠ '+w+'</div>';
  }});

  // Info (zwinięte jeśli brak błędów)
  if(info.length) {{
    if(!hasErr && !hasWarn) {{
      h += '<div style="background:#f0fff4;border:0.5px solid #9fe1cb;border-radius:7px;padding:7px 10px;font-size:11px;color:#085041">✓ Plik poprawny — '+info.length+' sprawdzeń bez błędów</div>';
    }} else {{
      var ih = info.map(function(i){{ return '<div style="font-size:11px;color:#555;padding:2px 0">'+i+'</div>'; }}).join('');
      h += '<details style="margin-top:4px"><summary style="font-size:11px;color:#888;cursor:pointer">Szczegóły ('+info.length+')</summary><div style="padding:6px 0">'+ih+'</div></details>';
    }}
  }}

  var sv=document.getElementById('step-valid');
  sv.style.display='flex';
  document.getElementById('valid-content').innerHTML = h || '<div style="font-size:11px;color:#aaa">Brak danych walidacyjnych</div>';
  document.getElementById('num2').style.background = hasErr ? '#e53935' : (hasWarn ? '#EF9F27' : '#1D9E75');

  // Krok 3 — zawsze odblokuj (błędy nie blokują, tylko ostrzegają)
  showStep3(data.meta || {{}});
}}

/* ── Krok 3: META + Drużyna + Zawodnicy ── */
function showStep3(meta) {{
  var fields = [
    ['Drużyna A', meta.nazwa_a || '—'],
    ['Drużyna B', meta.nazwa_b || '—'],
    ['Data',      meta.data    || '—'],
    ['Wynik',     (meta.wynik_a && meta.wynik_b) ? meta.wynik_a+' : '+meta.wynik_b : '—'],
    ['Rozgrywki', meta.rozgrywki || '—'],
    ['Runda / Kolejka', meta.runda || '—'],
    ['Miejsce',   meta.miejsce  || '—'],
  ];
  var h='';
  fields.forEach(function(f) {{
    var empty=(f[1]==='—');
    h+='<div style="background:#f8faff;border-radius:7px;padding:7px 10px;border:0.5px solid '+(empty?'#f5e199':'#e3e8f0')+'">';
    h+='<div style="font-size:9px;color:#aaa;text-transform:uppercase;letter-spacing:.4px;margin-bottom:2px">'+f[0]+'</div>';
    h+='<div style="font-size:12px;font-weight:600;color:'+(empty?'#b07800':'#1a2b4a')+'">'+f[1]+'</div>';
    h+='</div>';
  }});
  document.getElementById('meta-grid').innerHTML=h;

  var s3=document.getElementById('step-3');
  s3.style.display='flex';
  document.getElementById('num3').style.background='#EF9F27';
  document.getElementById('step-footer').style.display='flex';
}}

function onTeamChange(team_id) {{
  if(!team_id) {{
    ROSTER=[];
    document.getElementById('assign-hint').textContent='Najpierw wybierz drużynę';
    document.getElementById('assign-hint').style.color='#aaa';
    document.getElementById('roster-wrap').innerHTML='<div style="text-align:center;padding:10px 0;font-size:11px;color:#bbb">Wybierz drużynę aby przypisać zawodników</div>';
    return;
  }}
  fetch('/api/team_players/'+team_id)
    .then(function(r){{return r.json();}})
    .then(function(data){{
      ROSTER=(data.players||[]).sort(function(a,b){{
        return (a.nazwisko+' '+a.imie).localeCompare(b.nazwisko+' '+b.imie,'pl');
      }});
      renderMappings();
    }});
}}

function renderMappings() {{
  var nrs=Object.keys(NR_STATS);
  var wrap=document.getElementById('roster-wrap');
  var hint=document.getElementById('assign-hint');
  if(!nrs.length) {{
    if(wrap) wrap.innerHTML='<div style="text-align:center;padding:10px 0;font-size:11px;color:#bbb">Brak zawodników w pliku</div>';
    return;
  }}
  if(hint) {{ hint.textContent=nrs.length+' zawodników'; hint.style.color='#888'; }}
  var h='';
  nrs.sort(function(a,b){{return parseInt(a)-parseInt(b);}}).forEach(function(nr) {{
    var st=NR_STATS[nr];
    h+='<div style="display:grid;grid-template-columns:58px 80px 14px 1fr 46px;gap:6px;padding:6px 0;align-items:center;border-bottom:0.5px solid #f5f5f5">';
    h+='<div><div style="width:28px;height:28px;border-radius:50%;background:#1a2b4a;color:#fff;font-size:11px;font-weight:600;display:flex;align-items:center;justify-content:center">#'+nr+'</div></div>';
    h+='<div style="font-size:11px;color:#888">'+st.pts+' pkt</div>';
    h+='<div style="text-align:center;color:#ccc;font-size:13px">›</div>';
    h+='<select id="sel-'+nr+'" data-nr="'+nr+'" onchange="updateMappings()" style="width:100%;padding:5px 8px;border:0.5px solid #e0e0e0;border-radius:7px;font-size:12px;background:#fff">';
    h+='<option value="">— nieprzypisany</option>';
    ROSTER.forEach(function(p) {{
      h+='<option value="'+p.id+'">'+p.nazwisko+' '+p.imie+'</option>';
    }});
    h+='</select>';
    h+='<div id="st-'+nr+'" style="text-align:center"><span style="font-size:10px;padding:2px 6px;border-radius:10px;background:#f5f5f5;color:#aaa">—</span></div>';
    h+='</div>';
  }});
  if(wrap) wrap.innerHTML=h;
  updateMappings();
}}

function updateMappings() {{
  var nrs=Object.keys(NR_STATS), mappings={{}}, assigned=0;
  nrs.forEach(function(nr) {{
    var sel=document.getElementById('sel-'+nr);
    var st=document.getElementById('st-'+nr);
    if(!sel) return;
    var val=sel.value;
    mappings[nr]=val||null;
    if(val) {{
      assigned++;
      sel.style.borderColor='#5DCAA5'; sel.style.background='#E1F5EE';
      st.innerHTML='<span style="font-size:10px;padding:2px 6px;border-radius:10px;background:#E1F5EE;color:#085041">✓</span>';
    }} else {{
      sel.style.borderColor='#e0e0e0'; sel.style.background='#fff';
      st.innerHTML='<span style="font-size:10px;padding:2px 6px;border-radius:10px;background:#f5f5f5;color:#aaa">—</span>';
    }}
  }});
  document.getElementById('mappings-json').value=JSON.stringify(mappings);
  var hint=document.getElementById('assign-hint');
  if(hint) {{
    hint.textContent=assigned+'/'+nrs.length+' przypisanych';
    hint.style.color=assigned===nrs.length?'#1D9E75':'#EF9F27';
  }}
}}
</script>
"""
    return html_response(base(content, active="upload"))


@app.route("/upload/preview", methods=["POST"])
@login_required
def upload_preview():
    """Parse xlsx — walidacja + META + statystyki zawodników."""
    import json as _j
    if "file" not in request.files:
        return _j.dumps({"error": "Brak pliku"})
    f = request.files["file"]
    try:
        import openpyxl, io as _io
        file_bytes = f.read()
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), data_only=True)

        # --- Walidacja (validate_workbook parsuje też arkusze) ---
        report = validate_workbook(wb)

        # Serializuj errors (mogą być dict lub str)
        def _ser_err(e):
            if isinstance(e, dict):
                return {"type":"dict", "msg": e.get("msg",""), "quarters": e.get("quarters",[]), "total": e.get("total",0), "meta": e.get("meta",0)}
            return {"type":"str", "msg": str(e)}

        # --- META ---
        raw_meta = report.get("meta") or {}
        from datetime import datetime as _dt2, date as _date2
        _md = raw_meta.get("data")
        if isinstance(_md, (_dt2, _date2)):
            meta_date = _md.strftime('%d.%m.%Y')
        elif _md:
            meta_date = str(_md).strip()
        else:
            meta_date = ""

        meta_out = {
            "nazwa_a":   str(raw_meta.get("nazwa_a","")).strip(),
            "nazwa_b":   str(raw_meta.get("nazwa_b","")).strip(),
            "wynik_a":   str(raw_meta.get("wynik_a","")).strip(),
            "wynik_b":   str(raw_meta.get("wynik_b","")).strip(),
            "data":      meta_date,
            "rozgrywki": str(raw_meta.get("rozgrywki","")).strip(),
            "runda":     str(raw_meta.get("runda","")).strip(),
            "miejsce":   str(raw_meta.get("miejsce","")).strip(),
        }

        # --- Statystyki zawodników (arkusz A, już sparsowany) ---
        from collections import defaultdict
        player_stats = defaultdict(lambda: {"pts":0,"p2":0,"p3":0,"actions":0})
        sheets_data = [s for s in wb.sheetnames if s.upper() not in ("META","KODY","LEGENDA")]
        sheet_a = sheets_data[0] if sheets_data else None
        if sheet_a:
            try:
                stats = parse_team_sheet(wb[sheet_a], sheet_type="A")
                for nr_str, pd in stats["players"].items():
                    try:
                        nr = int(nr_str)
                        if 0 <= nr <= 99:
                            k = str(nr)
                            pts = pd.get("p2m",0)*2 + pd.get("p3m",0)*3 + pd.get("ftm",0)
                            player_stats[k]["pts"]     += pts
                            player_stats[k]["p2"]      += pd.get("p2a",0)
                            player_stats[k]["p3"]      += pd.get("p3a",0)
                            player_stats[k]["actions"] += pd.get("finishes",0)
                    except: continue
            except: pass

        return _j.dumps({
            "stats":    {k: dict(v) for k,v in player_stats.items()},
            "meta":     meta_out,
            "errors":   [_ser_err(e) for e in report.get("errors",[])],
            "warnings": [str(w) for w in report.get("warnings",[])],
            "info":     [str(i) for i in report.get("info",[])],
            "names":    list(report.get("names", ("",""))),
        })
    except Exception as e:
        return _j.dumps({"error": str(e)})


@app.route("/api/team_players/<int:team_id>")
@login_required
def api_team_players(team_id):
    """Return sorted player list for a team."""
    import json as _j
    try:
        db = get_db(); cur = db.cursor()
        cur.execute("""
            SELECT id, imie, nazwisko, numer, pozycja
            FROM players WHERE team_id=%s
            ORDER BY nazwisko, imie
        """, (team_id,))
        players = [dict(r) for r in cur.fetchall()]
        cur.close()
        return _j.dumps({"players": players})
    except Exception as e:
        return _j.dumps({"players": [], "error": str(e)})



@app.route("/upload", methods=["POST"])
@login_required
def upload():
    if "file" not in request.files:
        flash("Nie wybrano pliku","error"); return redirect(url_for("upload_page"))
    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        flash("Plik musi być .xlsx","error"); return redirect(url_for("upload_page"))

    team_id      = request.form.get("team_id", None)
    mappings_raw = request.form.get("mappings_json", "{}")
    import json as _jmap
    try:    nr_to_player = _jmap.loads(mappings_raw)
    except: nr_to_player = {}

    sezon = request.form.get("sezon", get_setting("current_season") or "")

    try:
        file_bytes = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        # ── Odczytaj datę z META ──────────────────────────────────────────
        meta = read_meta(wb)
        data_meczu = None
        if meta.get("data"):
            raw = str(meta["data"]).strip().replace(" ","")
            for fmt in ('%d.%m.%Y','%d/%m/%Y','%Y-%m-%d','%d-%m-%Y','%Y.%m.%d'):
                try:
                    from datetime import datetime as dt2
                    data_meczu = dt2.strptime(raw, fmt).strftime('%Y-%m-%d')
                    break
                except: pass
            if not data_meczu:
                # Fallback — dateutil
                try:
                    from dateutil import parser as dp
                    data_meczu = dp.parse(raw, dayfirst=True).strftime('%Y-%m-%d')
                except: pass

        # ── Walidacja ─────────────────────────────────────────────────────
        report = validate_workbook(wb)
        has_issues = bool(report["errors"] or report["warnings"])

        if has_issues:
            import uuid
            token = str(uuid.uuid4())
            tmp_path = os.path.join(PENDING_DIR, f"{token}.xlsx")
            with open(tmp_path, "wb") as fp:
                fp.write(file_bytes)

            import re as _re
            def clean(s):
                if isinstance(s, dict):
                    return {
                        "msg": _re.sub(r'<[^>]+>', '', s.get("msg",""))[:150],
                        "quarters": s.get("quarters",[]),
                        "total": s.get("total",0),
                        "meta": s.get("meta",None),
                    }
                return _re.sub(r'<[^>]+>', '', str(s))[:120]

            for k in ['pt','ps','pd','vr']: session.pop(k, None)
            session["pt"] = token
            session["ps"] = sezon
            session["pd"] = data_meczu or ""
            session["pteam"] = team_id or ""
            session["pmap"]  = mappings_raw or "{}"
            session["vr"] = {
                "e": [clean(e) for e in report["errors"][:8]],
                "w": [clean(w) for w in report["warnings"][:15]],
                "i": [clean(i) for i in report["info"][:15]],
                "n": list(report["names"]),
                "p": list(report["pts"]),
            }
            return redirect(url_for("validation_report"))

        # ── Brak problemów — zapisz od razu ──────────────────────────────
        return _do_save(wb, report["names"][0], report["names"][1], sezon, data_meczu, team_id=team_id, nr_to_player=nr_to_player)

    except Exception as e:
        try:
            get_db().rollback()
        except: pass
        flash(f"Błąd wgrywania: {str(e)}","error")
        return redirect(url_for("index"))


@app.route("/walidacja/pobierz-z-bledami")
@login_required
def download_with_errors():
    """Pobierz oryginalny plik z zaznaczonymi błędami na czerwono"""
    token = session.get("pt","")
    if not token:
        flash("Sesja wygasła — wgraj plik ponownie","error")
        return redirect(url_for("index"))

    tmp_path = os.path.join(PENDING_DIR, f"{token}.xlsx")
    if not os.path.exists(tmp_path):
        flash("Plik tymczasowy wygasł — wgraj ponownie","error")
        return redirect(url_for("index"))

    try:
        with open(tmp_path, "rb") as fp:
            file_bytes = fp.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))

        RED_FILL   = PatternFill("solid", fgColor="FFCDD2")
        RED_FONT   = Font(color="B71C1C", bold=True)
        RED_BORDER = Border(
            left=Side(style="medium", color="E53935"),
            right=Side(style="medium", color="E53935"),
            top=Side(style="medium", color="E53935"),
            bottom=Side(style="medium", color="E53935"),
        )
        COMMENT_FILL = PatternFill("solid", fgColor="FFEBEE")

        report = validate_workbook(wb)

        # Zbierz błędy per arkusz per wiersz
        error_cells = {}  # {(sheet_name, row): [opisy]}

        sheets_data = [s for s in wb.sheetnames if s.upper() not in ("META","KODY","LEGENDA")]
        if len(sheets_data) >= 2:
            name_a, name_b = sheets_data[0], sheets_data[1]

            for sheet_name in [name_a, name_b]:
                ws = wb[sheet_name]

                # ── Wynik końcowy (cały arkusz)
                suma = suma_quarters(parse_team_sheet(ws, sheet_type="A" if sheet_name==name_a else "B"))
                meta = read_meta(wb)
                key = "wynik_a" if sheet_name == name_a else "wynik_b"
                try:
                    meta_pts = int(meta.get(key, "")) if meta.get(key) else None
                    coded_pts = suma.get("pts", 0)
                    if meta_pts is not None and meta_pts != coded_pts:
                        # Zaznacz nagłówek arkusza (wiersz 1)
                        k = (sheet_name, 1)
                        if k not in error_cells: error_cells[k] = []
                        error_cells[k].append(
                            f"Wynik w META: {meta_pts}, kodowanie: {coded_pts} (różnica: {coded_pts-meta_pts:+d})"
                        )
                except: pass

                # ── Brakujące dane (kolumny A/B/C)
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=500, values_only=False), 2):
                    if not any(c.value is not None for c in row[:4]): break
                    missing = []
                    if row[0].value is None: missing.append("Kwarta(A)")
                    if row[2].value is None: missing.append("Kod(C)")
                    if missing:
                        k = (sheet_name, i)
                        if k not in error_cells: error_cells[k] = []
                        error_cells[k].append(f"Brak: {', '.join(missing)}")

                # ── Nieznane kody akcji
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=500, values_only=False), 2):
                    if not any(c.value is not None for c in row[:4]): break
                    raw_c = str(row[2].value).strip() if row[2].value is not None else ""
                    if not raw_c: continue
                    bad_codes = [c.strip().upper() for c in raw_c.split(";")
                                 if c.strip() and c.strip().upper() not in VALID_CODES]
                    if bad_codes:
                        k = (sheet_name, i)
                        if k not in error_cells: error_cells[k] = []
                        error_cells[k].append(f"Nieznany kod: {', '.join(bad_codes)}")

        # ── Zaznacz błędy w pliku ───────────────────────────────────────────
        for (sheet_name, row_idx), descs in error_cells.items():
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]

            if row_idx == 1:
                # Błąd wyniku — zaznacz wiersz nagłówka i dodaj komentarz w A1
                for col in range(1, 15):
                    c = ws.cell(row_idx, col)
                    c.fill = PatternFill("solid", fgColor="FFCCBC")
                # Wpisz info w wolnym miejscu
                ws.cell(1, 16).value = "⚠ " + " | ".join(descs)
                ws.cell(1, 16).font = Font(color="BF360C", bold=True, size=9)
                ws.cell(1, 16).fill = PatternFill("solid", fgColor="FFF3E0")
            else:
                # Błąd w wierszu danych — zaznacz cały wiersz
                for col in range(1, 12):
                    c = ws.cell(row_idx, col)
                    c.fill = RED_FILL
                    if col in [1, 3]:  # Kwarta i Kod — główne kolumny z błędem
                        c.font = RED_FONT
                        c.border = RED_BORDER
                # Opis błędu w kolumnie O
                desc_cell = ws.cell(row_idx, 15)
                desc_cell.value = "❌ " + " | ".join(descs)
                desc_cell.font = Font(color="B71C1C", bold=True, size=9)
                desc_cell.fill = COMMENT_FILL

        # ── Dodaj podsumowanie punktów per kwarta (prawa strona) ──────────────
        HDR_BLUE  = PatternFill("solid", fgColor="1A2B4A")
        HDR_WHITE = Font(color="FFFFFF", bold=True, size=9)
        SUM_FILL  = PatternFill("solid", fgColor="E3F2FD")
        SUM_FONT  = Font(bold=True, size=9, color="0C447C")
        OK_FILL   = PatternFill("solid", fgColor="E8F5E9")
        ERR_FILL  = PatternFill("solid", fgColor="FFEBEE")

        meta = read_meta(wb)

        for idx, sheet_name in enumerate(sheets_data):
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            stats = parse_team_sheet(ws, sheet_type="A" if idx==0 else "B")

            # Znajdź ostatni wiersz z danymi
            last_row = 1
            for r in ws.iter_rows(min_row=2, max_row=600, values_only=True):
                if not any(v is not None for v in r[:4]): break
                last_row += 1

            # Kolumny podsumowania — zaczynamy od P (16)
            COL_START = 17  # kolumna Q

            # Nagłówki
            for ci, lbl in enumerate(["KWARTA","PKT","2PM/A","3PM/A","FTM/A","TO","POSS"]):
                c = ws.cell(1, COL_START + ci, lbl)
                c.fill = HDR_BLUE; c.font = HDR_WHITE
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(COL_START + ci)].width = 9

            ws.column_dimensions[get_column_letter(COL_START)].width = 8

            # Dane per kwarta
            total_pts = 0
            for qi, qn in enumerate([1,2,3,4]):
                qd = stats["quarter"].get(qn, {})
                r = 2 + qi
                pts_q = qd.get("pts",0)
                total_pts += pts_q

                vals = [
                    f"{qn}Q",
                    pts_q,
                    f"{qd.get('p2m',0)}/{qd.get('p2a',0)}",
                    f"{qd.get('p3m',0)}/{qd.get('p3a',0)}",
                    f"{qd.get('ftm',0)}/{qd.get('fta',0)}",
                    qd.get("br",0),
                    qd.get("poss",0),
                ]
                for ci, v in enumerate(vals):
                    c = ws.cell(r, COL_START + ci, v)
                    c.fill = SUM_FILL; c.font = Font(size=9)
                    c.alignment = Alignment(horizontal="center")

            # Wiersz SUMA
            r_sum = 6
            ws.cell(r_sum, COL_START, "SUMA").fill = PatternFill("solid", fgColor="1A2B4A")
            ws.cell(r_sum, COL_START).font = Font(color="FFFFFF", bold=True, size=9)
            ws.cell(r_sum, COL_START).alignment = Alignment(horizontal="center")

            suma_all = suma_quarters(stats)
            suma_vals = [
                total_pts,
                f"{suma_all.get('p2m',0)}/{suma_all.get('p2a',0)}",
                f"{suma_all.get('p3m',0)}/{suma_all.get('p3a',0)}",
                f"{suma_all.get('ftm',0)}/{suma_all.get('fta',0)}",
                suma_all.get("br",0),
                suma_all.get("poss",0),
            ]
            for ci, v in enumerate(suma_vals):
                c = ws.cell(r_sum, COL_START + 1 + ci, v)
                c.font = Font(bold=True, size=9); c.alignment = Alignment(horizontal="center")
                c.fill = PatternFill("solid", fgColor="BBDEFB")

            # Porównanie z META
            meta_key = "wynik_a" if idx == 0 else "wynik_b"
            try:
                meta_pts = int(meta.get(meta_key,"")) if meta.get(meta_key) else None
            except: meta_pts = None

            r_meta = 8
            if meta_pts is not None:
                ws.cell(r_meta, COL_START, "META").fill = PatternFill("solid", fgColor="37474F")
                ws.cell(r_meta, COL_START).font = Font(color="FFFFFF", bold=True, size=9)
                ws.cell(r_meta, COL_START).alignment = Alignment(horizontal="center")

                match = meta_pts == total_pts
                fill = OK_FILL if match else ERR_FILL
                font_col = "1B5E20" if match else "B71C1C"

                ws.cell(r_meta, COL_START+1, meta_pts).fill = fill
                ws.cell(r_meta, COL_START+1).font = Font(bold=True, size=9, color=font_col)
                ws.cell(r_meta, COL_START+1).alignment = Alignment(horizontal="center")

                ws.cell(r_meta, COL_START+2, "✓ OK" if match else f"❌ RÓŻNICA: {total_pts-meta_pts:+d}")
                ws.cell(r_meta, COL_START+2).fill = fill
                ws.cell(r_meta, COL_START+2).font = Font(bold=True, size=9, color=font_col)
                ws.merge_cells(start_row=r_meta, start_column=COL_START+2,
                               end_row=r_meta, end_column=COL_START+5)

        # ── Dodaj legendę na początku każdego arkusza ──────────────────────
        for sheet_name in sheets_data:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            # Kolumna O nagłówek
            hdr = ws.cell(1, 15)
            hdr.value = "BŁĘDY WALIDACJI"
            hdr.font = Font(color="FFFFFF", bold=True, size=9)
            hdr.fill = PatternFill("solid", fgColor="C62828")
            ws.column_dimensions["O"].width = 40

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        names = session.get("vr", {}).get("n", ["plik","plik"])
        filename = f"BLEDY_{names[0]}_vs_{names[1]}.xlsx".replace(" ","_")
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        flash(f"Błąd generowania pliku: {str(e)}", "error")
        return redirect(url_for("validation_report"))


@app.route("/walidacja")
@login_required
def validation_report():
    vr = session.get("vr")
    if not vr:
        return redirect(url_for("index"))

    errors   = vr.get("e",[])
    warnings = vr.get("w",[])
    info     = vr.get("i",[])
    names    = vr.get("n",["?","?"])
    pts      = vr.get("p",[0,0])
    has_errors = len(errors) > 0

    # Renderuj błędy — obsługa zarówno dict (z kwartami) jak i str
    def render_error(e, idx):
        if isinstance(e, dict):
            msg      = e.get("msg","")
            quarters = e.get("quarters",[])
            total    = e.get("total",0)
            meta_pts = e.get("meta",None)
            # Pasek per kwarta
            q_html = ""
            if quarters:
                q_items = " | ".join(
                    f'<span style="font-weight:700;color:#{'1a6b3c' if quarters[i]==max(quarters) else 'b71c1c' if quarters[i]==min(quarters) else '444'}">'
                    f'{i+1}Q — {quarters[i]}</span>'
                    for i in range(len(quarters))
                )
                q_html = f"""
                <div style="margin-top:.5rem;padding:.4rem .6rem;background:#fff;border-radius:6px;border:1px solid #ffcdd2;font-size:.82rem">
                  {q_items}
                  <span style="margin-left:.5rem;color:#888;font-size:.75rem">= {total} pkt (META: {meta_pts})</span>
                </div>"""
            return f"""
            <div class="val-item val-error" style="flex-direction:column;cursor:pointer" onclick="this.querySelector('.qdetail').style.display=this.querySelector('.qdetail').style.display=='none'?'block':'none'">
              <div style="display:flex;gap:.75rem;align-items:flex-start;width:100%">
                <span class="val-icon">🚫</span>
                <span style="flex:1">{msg}</span>
                <span style="font-size:.75rem;color:#c62828;flex-shrink:0">▼ rozwiń</span>
              </div>
              <div class="qdetail" style="display:none">{q_html}</div>
            </div>"""
        else:
            return f'<div class="val-item val-error"><span class="val-icon">🚫</span><span>{e}</span></div>'

    def render_warning(w):
        return f'<div class="val-item val-warning"><span class="val-icon">⚠️</span><span>{w}</span></div>'

    def render_info(i):
        return f'<div class="val-item val-info"><span class="val-icon">✓</span><span>{i}</span></div>'

    errors_html   = "".join(render_error(e, i) for i,e in enumerate(errors))
    warnings_html = "".join(render_warning(w) for w in warnings)
    info_html     = "".join(render_info(i) for i in info)

    content = f"""
<div class="page-title">🔍 Raport walidacji pliku</div>

<div class="card mb-3 p-3">
  <div class="d-flex gap-3 align-items-center flex-wrap">
    <div>
      <div style="font-size:.8rem;color:#888">Drużyny</div>
      <div class="fw-bold">{names[0]} vs {names[1]}</div>
    </div>
    <div>
      <div style="font-size:.8rem;color:#888">Wynik z kodowania</div>
      <div class="fw-bold">{pts[0]} : {pts[1]}</div>
    </div>
    <div class="ms-auto d-flex gap-2 flex-wrap">
      <a href="/" class="btn btn-outline-secondary btn-sm">← Anuluj</a>
      {'<a href="/walidacja/pobierz-z-bledami" class="btn btn-outline-danger btn-sm fw-bold">📥 Pobierz plik z błędami</a>' if has_errors else ''}
      {'<span class="btn btn-secondary btn-sm disabled">Zapisz (popraw błędy)</span>' if has_errors else
       '<form method="POST" action="/upload/force" style="display:inline"><button type="submit" class="btn btn-success btn-sm fw-bold">✓ Zapisz mimo ostrzeżeń</button></form>'}
    </div>
  </div>
</div>

<style>
.val-section{{margin-bottom:1rem}}
.val-section-title{{font-size:.72rem;text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-bottom:.5rem;padding:.3rem .6rem;border-radius:6px}}
.val-item{{display:flex;gap:.75rem;align-items:flex-start;padding:.5rem .75rem;border-radius:8px;margin-bottom:.3rem;font-size:.85rem;line-height:1.4}}
.val-icon{{font-size:1rem;flex-shrink:0;margin-top:1px}}
.val-error{{background:#fff0f0;border-left:3px solid #e53935}}
.val-warning{{background:#fffde7;border-left:3px solid #f9a825}}
.val-info{{background:#f1f8e9;border-left:3px solid #43a047}}
</style>

<div class="card p-3">
  {'<div class="val-section"><div class="val-section-title" style="background:#ffebee;color:#c62828">🚫 Błędy krytyczne (' + str(len(errors)) + ') — kliknij aby rozwinąć</div>' + errors_html + '</div>' if errors else ''}
  {'<div class="val-section"><div class="val-section-title" style="background:#fff8e1;color:#f57f17">⚠️ Ostrzeżenia (' + str(len(warnings)) + ')</div>' + warnings_html + '</div>' if warnings else ''}
  {'<div class="val-section"><div class="val-section-title" style="background:#e8f5e9;color:#2e7d32">✓ Poprawne (' + str(len(info)) + ')</div>' + info_html + '</div>' if info else ''}
</div>

{'<div class="card p-3 mt-2" style="background:#fff0f0;border:1px solid #ffcdd2"><b>Plik zawiera błędy krytyczne.</b> Popraw plik i wgraj ponownie.</div>' if has_errors else '<div class="card p-3 mt-2" style="background:#fffde7;border:1px solid #fff176"><b>Plik zawiera ostrzeżenia.</b> Możesz zapisać mimo ostrzeżeń lub poprawić plik.</div>'}
"""
    return html_response(base(content, active="home"))


@app.route("/upload/force", methods=["POST"])
@login_required
def upload_force():
    """Zapisz plik mimo ostrzeżeń"""
    token      = session.get("pt","")
    sezon      = session.get("ps", get_setting("current_season") or "")
    data_meczu = session.get("pd") or None  # "" → None dla DATE
    team_id    = session.get("pteam") or None
    import json as _jf
    try:    nr_to_player = _jf.loads(session.get("pmap","{}"))
    except: nr_to_player = {}

    if not token:
        flash("Sesja wygasła — wgraj plik ponownie","error")
        return redirect(url_for("index"))

    tmp_path = os.path.join(PENDING_DIR, f"{token}.xlsx")
    if not os.path.exists(tmp_path):
        flash("Plik tymczasowy wygasł — wgraj ponownie","error")
        return redirect(url_for("index"))

    try:
        with open(tmp_path, "rb") as fp:
            file_bytes = fp.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        report = validate_workbook(wb)
        # Usuń plik tymczasowy
        try: os.remove(tmp_path)
        except: pass
        return _do_save(wb, report["names"][0], report["names"][1], sezon, data_meczu, team_id=team_id, nr_to_player=nr_to_player)
    except Exception as e:
        try: get_db().rollback()
        except: pass
        flash(f"Błąd zapisu: {str(e)}","error")
        return redirect(url_for("index"))

# ══════════════════════════════════════════════════════════════════════════════
# HISTORIA MECZÓW
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/debug-players")
@login_required  
def debug_players():
    import json as _j
    pj = get_setting("players_json") or "{}"
    kj = get_setting("kluby_json") or "[]"
    try:    players = _j.loads(pj)
    except: players = {}
    try:    kluby = _j.loads(kj)
    except: kluby = []
    
    # Count players
    total = 0
    details = []
    for ki, seasons in players.items():
        kname = kluby[int(ki)]["name"] if int(ki) < len(kluby) else f"klub_{ki}"
        for sname, teams in seasons.items():
            for tname, plist in teams.items():
                total += len(plist)
                details.append(f"{kname} · {sname} · {tname}: {len(plist)} zawodników")
    
    content = f"""
<div class="page-title">Debug: players_json</div>
<div class="card p-3">
<p>Łącznie zawodników w players_json: <b>{total}</b></p>
<pre style="font-size:11px">{"<br>".join(details) or "BRAK DANYCH"}</pre>
<hr>
<form method="POST" action="/admin/restore-players">
  <button class="btn btn-warning">Przywróć zawodników z players_json do tabeli players</button>
</form>
</div>"""
    return html_response(base(content, active="settings"))


@app.route("/admin/restore-players", methods=["POST"])
@login_required
def restore_players():
    import json as _j
    try:
        db = get_db(); cur = db.cursor()
        pj = get_setting("players_json") or "{}"
        kj = get_setting("kluby_json") or "[]"
        try:    all_players = _j.loads(pj)
        except: all_players = {}
        try:    kluby = _j.loads(kj)
        except: kluby = []
        
        restored = 0
        for ki_str, seasons in all_players.items():
            try: ki = int(ki_str)
            except: continue
            if ki >= len(kluby): continue
            kname = kluby[ki].get("name","")
            
            cur.execute("SELECT id FROM clubs WHERE name=%s", (kname,))
            cr = cur.fetchone()
            if not cr: continue
            club_id = cr["id"]
            
            for sname, teams in seasons.items():
                cur.execute("SELECT id FROM seasons WHERE club_id=%s AND name=%s",
                            (club_id, sname))
                sr = cur.fetchone()
                if not sr: continue
                season_id = sr["id"]
                
                for tname, plist in teams.items():
                    cur.execute("SELECT id FROM teams WHERE season_id=%s AND name=%s",
                                (season_id, tname))
                    tr = cur.fetchone()
                    if not tr: continue
                    team_id = tr["id"]
                    
                    for pl in plist:
                        imie = pl.get("imie","")
                        nazw = pl.get("nazwisko","")
                        if not imie and not nazw: continue
                        cur.execute("""
                            INSERT INTO players (team_id, imie, nazwisko, numer, pozycja)
                            VALUES (%s,%s,%s,%s,%s)
                        """, (team_id, imie, nazw,
                              int(pl.get("num",0)), pl.get("poz","")))
                        restored += 1
        
        db.commit()
        cur.close()
        flash(f"Przywrócono {restored} zawodników.", "success")
    except Exception as e:
        try: get_db().rollback()
        except: pass
        flash(f"Błąd: {e}", "error")
    return redirect(url_for("debug_players"))


@app.route("/admin/reset-clubs", methods=["POST"])
@login_required
def reset_clubs():
    """Czyści tabele clubs/seasons/teams/players i resynchronizuje z kluby_json."""
    import json as _j
    try:
        db = get_db(); cur = db.cursor()
        # Tylko struktura — NIE usuwamy players (zawodnicy muszą zostać)
        # Najpierw odepnij player_match_map od teams
        cur.execute("DELETE FROM player_match_map")
        # Odepnij players od teams tymczasowo (zachowaj rekordy)
        cur.execute("UPDATE players SET team_id=NULL WHERE team_id IS NOT NULL")
        cur.execute("DELETE FROM teams")
        cur.execute("DELETE FROM seasons")
        cur.execute("DELETE FROM clubs")
        db.commit()

        # Resync from kluby_json
        kj = get_setting("kluby_json") or "[]"
        try:    kluby = _j.loads(kj)
        except: kluby = []

        for klub in kluby:
            kname = klub.get("name","")
            kext  = bool(klub.get("ext", False))
            if not kname: continue
            cur.execute("""INSERT INTO clubs (name, ext) VALUES (%s,%s)
                           ON CONFLICT (name) DO UPDATE SET ext=%s RETURNING id""",
                        (kname, kext, kext))
            club_id = cur.fetchone()["id"]
            for sname, teams in klub.get("sezony",{}).items():
                cur.execute("""INSERT INTO seasons (club_id, name) VALUES (%s,%s)
                               ON CONFLICT (club_id, name) DO NOTHING RETURNING id""",
                            (club_id, sname))
                sr = cur.fetchone()
                if not sr:
                    cur.execute("SELECT id FROM seasons WHERE club_id=%s AND name=%s",
                                (club_id, sname))
                    sr = cur.fetchone()
                season_id = sr["id"]
                for tname in teams:
                    cur.execute("""INSERT INTO teams (season_id, name) VALUES (%s,%s)
                                   ON CONFLICT (season_id, name) DO NOTHING RETURNING id""",
                                (season_id, tname))
                    tr = cur.fetchone()
                    if not tr:
                        cur.execute("SELECT id FROM teams WHERE season_id=%s AND name=%s",
                                    (season_id, tname))
                        tr = cur.fetchone()
                    if tr:
                        team_id_new = tr["id"]
                        # Re-link players from players_json backup
                        import json as _j2
                        pj = get_setting("players_json") or "{}"
                        try:    all_players = _j2.loads(pj)
                        except: all_players = {}
                        ki_str = str(kluby.index(klub))
                        team_players = (all_players.get(ki_str) or {}).get(sname,{}).get(tname,[])
                        for pl in team_players:
                            cur.execute("""INSERT INTO players (team_id, imie, nazwisko, numer, pozycja)
                                           VALUES (%s,%s,%s,%s,%s)""",
                                        (team_id_new, pl.get("imie",""), pl.get("nazwisko",""),
                                         int(pl.get("num",0)), pl.get("poz","")))
        db.commit()
        cur.close()
        flash("Struktura klubów odświeżona. Zawodnicy zachowani.", "success")
    except Exception as e:
        try: get_db().rollback()
        except: pass
        flash(f"Błąd: {e}", "error")
    return redirect(url_for("ustawienia"))


@app.route("/admin/match-files")
@login_required
def admin_match_files():
    try: init_db()
    except: pass
    db = get_db(); cur = db.cursor()
    try:
        cur.execute("""
            SELECT m.id, m.sezon, m.data_meczu, m.przeciwnik, m.nazwa_gtk,
                   COALESCE(m.parser_version, 0) as parser_version,
                   m.file_path
            FROM matches m
            ORDER BY m.data_meczu DESC NULLS LAST, m.created_at DESC
        """)
        matches = cur.fetchall()
    except: matches = []
    cur.close()

    outdated = [m for m in matches if (m.get("parser_version") or 0) < PARSER_VERSION]
    no_file  = [m for m in matches if not m.get("file_path") or not os.path.exists(m.get("file_path",""))]

    rows = ""
    for m in matches:
        dt    = m["data_meczu"].strftime("%d.%m.%Y") if m["data_meczu"] else "—"
        pv    = m.get("parser_version") or 0
        fp    = m.get("file_path") or ""
        has_f = bool(fp and os.path.exists(fp))
        fsize = f"{os.path.getsize(fp)/1024:.0f} KB" if has_f else "—"
        is_old = pv < PARSER_VERSION
        pv_badge = (f'<span style="background:#fff3cd;color:#856404;padding:2px 7px;border-radius:8px;font-size:10px">v{pv} → v{PARSER_VERSION}</span>'
                    if is_old else
                    f'<span style="background:#e8f5e9;color:#1a6b3c;padding:2px 7px;border-radius:8px;font-size:10px">v{pv} ✓</span>')
        file_badge = (f'<span style="font-size:10px;color:#1a6b3c">{fsize}</span>'
                      if has_f else
                      f'<span style="font-size:10px;color:#8b1a1a">brak pliku</span>')
        reparse_btn = (f'<form method="POST" action="/admin/reparse-match/{m["id"]}" style="display:inline">'
                       f'<button class="btn btn-sm btn-outline-warning" style="font-size:.72rem">Przelicz</button></form>'
                       if has_f else
                       f'<button class="btn btn-sm btn-outline-secondary" style="font-size:.72rem" disabled>Brak pliku</button>')
        del_btn = (f'<form method="POST" action="/admin/delete-match-file/{m["id"]}" style="display:inline" onsubmit="return confirm(''Usunąć plik? Dane w bazie zostają.'')">'
                   f'<button class="btn btn-sm btn-outline-danger" style="font-size:.72rem">Usuń plik</button></form>' if has_f else "")
        rows += (f'<tr><td style="font-size:.8rem">{dt}</td>'
                 f'<td style="font-size:.8rem">{m.get("nazwa_gtk","GTK")} vs {m["przeciwnik"]}</td>'
                 f'<td style="font-size:.8rem">{m["sezon"] or "—"}</td>'
                 f'<td>{pv_badge}</td><td>{file_badge}</td><td>{reparse_btn} {del_btn}</td></tr>')

    warn = (f'<div class="alert alert-warning" style="font-size:.85rem">⚠ {len(outdated)} meczów ma starszą wersję parsera.'
            f' <form method="POST" action="/admin/reparse-all" style="display:inline">'
            f'<button class="btn btn-warning btn-sm ms-2">Przelicz wszystkie</button></form></div>') if outdated else ""

    html = f"""
<div class="page-title">&#128196; Pliki meczów i wersjonowanie</div>
<div style="display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap">
  <div class="card p-3" style="flex:1;min-width:150px">
    <div style="font-size:11px;color:#888;margin-bottom:2px">Wersja parsera</div>
    <div style="font-size:24px;font-weight:500;color:#1a2b4a">v{PARSER_VERSION}</div>
  </div>
  <div class="card p-3" style="flex:1;min-width:150px">
    <div style="font-size:11px;color:#888;margin-bottom:2px">Do przeliczenia</div>
    <div style="font-size:24px;font-weight:500;color:{"#8b1a1a" if outdated else "#1a6b3c"}">{len(outdated)}</div>
  </div>
  <div class="card p-3" style="flex:1;min-width:150px">
    <div style="font-size:11px;color:#888;margin-bottom:2px">Pliki na dysku</div>
    <div style="font-size:24px;font-weight:500;color:#1a2b4a">{len(matches)-len(no_file)} / {len(matches)}</div>
  </div>
  <div class="card p-3" style="flex:1;min-width:150px;background:#f8f9fa">
    <div style="font-size:11px;color:#888;margin-bottom:2px">Folder plików</div>
    <div style="font-size:11px;font-family:monospace;word-break:break-all;color:#444">{MATCH_FILES_DIR}</div>
  </div>
</div>
{warn}
<div class="card"><div class="card-body p-2">
  <table class="table table-sm mb-0">
    <thead><tr><th>Data</th><th>Mecz</th><th>Sezon</th><th>Parser</th><th>Plik</th><th>Akcje</th></tr></thead>
    <tbody>{rows if rows else '<tr><td colspan="6" class="text-center text-muted py-3">Brak meczów</td></tr>'}</tbody>
  </table>
</div></div>"""
    return html_response(base(html, active="settings"))


@app.route("/admin/reparse-match/<int:match_id>", methods=["POST"])
@login_required
def reparse_match(match_id):
    try:
        db = get_db(); cur = db.cursor()
        cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
        m = cur.fetchone()
        if not m:
            flash("Mecz nie istnieje.", "error")
            return redirect(url_for("admin_match_files"))
        fp = m.get("file_path","")
        if not fp or not os.path.exists(fp):
            flash("Brak pliku na dysku.", "error")
            return redirect(url_for("admin_match_files"))

        # === Zachowaj dane których nie chcemy tracić ===
        saved_team_id = m.get("team_id")

        # Zachowaj przypisania zawodników (player_id/roster_id per nr)
        cur.execute("""SELECT nr, player_id, roster_id
                       FROM player_stats
                       WHERE match_id=%s AND druzyna='gtk'
                         AND (player_id IS NOT NULL OR roster_id IS NOT NULL)""", (match_id,))
        saved_assignments = {row["nr"]: {"player_id": row["player_id"], "roster_id": row["roster_id"]}
                             for row in cur.fetchall()}

        # Zachowaj player_match_map
        cur.execute("SELECT player_id, nr, team_id FROM player_match_map WHERE match_id=%s", (match_id,))
        saved_pmm = cur.fetchall()

        import openpyxl, io as _io2
        with open(fp,"rb") as f: wb = openpyxl.load_workbook(_io2.BytesIO(f.read()), data_only=True)
        report = validate_workbook(wb)

        # Data: preferuj META z excela (może być zmieniona), fallback na DB
        meta_fresh = read_meta(wb)
        from datetime import datetime as _dt2, date as _date2
        _meta_date = meta_fresh.get("data")
        if _meta_date:
            if isinstance(_meta_date, (_dt2, _date2)):
                reparse_date = _meta_date.strftime('%Y-%m-%d') if isinstance(_meta_date, _dt2) else _meta_date.isoformat()
            else:
                reparse_date = None
                for _fmt in ('%d.%m.%Y','%d/%m/%Y','%Y-%m-%d','%d-%m-%Y'):
                    try: reparse_date = _dt2.strptime(str(_meta_date).strip(), _fmt).strftime('%Y-%m-%d'); break
                    except: pass
        else:
            reparse_date = str(m["data_meczu"]) if m["data_meczu"] else None

        # Usuń stare dane tego meczu
        cur.execute("DELETE FROM score_flow WHERE match_id=%s",       (match_id,))
        cur.execute("DELETE FROM lineup_stats WHERE match_id=%s",     (match_id,))
        cur.execute("DELETE FROM timing_stats WHERE match_id=%s",     (match_id,))
        cur.execute("DELETE FROM player_stats WHERE match_id=%s",     (match_id,))
        cur.execute("DELETE FROM match_stats WHERE match_id=%s",      (match_id,))
        cur.execute("DELETE FROM shot_zones WHERE match_id=%s",       (match_id,))
        cur.execute("DELETE FROM player_match_map WHERE match_id=%s", (match_id,))
        cur.execute("DELETE FROM matches WHERE id=%s", (match_id,))
        db.commit(); cur.close()

        # Przeparsuj od nowa
        _stats_gtk_r = parse_team_sheet(wb[report["names"][0]], "A")
        _stats_opp_r = parse_team_sheet(wb[report["names"][1]], "B")
        _def_lineups_r = build_gtk_def_lineups(wb[report["names"][0]], wb[report["names"][1]])
        new_id = save_match_to_db(
            report["names"][1], report["names"][0],
            m["sezon"], reparse_date,
            _stats_gtk_r, _stats_opp_r,
            rozgrywki=m.get("rozgrywki",""), runda=m.get("runda",""),
            kolejka=m.get("kolejka",""),
            miejsce=m.get("miejsce",""),
            def_lineups=_def_lineups_r,
            team_id=saved_team_id,
        )

        # === Przywróć przypisania zawodników w player_stats ===
        if saved_assignments:
            db3 = get_db(); cur3 = db3.cursor()
            for nr, asgn in saved_assignments.items():
                try:
                    if asgn.get("player_id"):
                        cur3.execute("UPDATE player_stats SET player_id=%s WHERE match_id=%s AND druzyna='gtk' AND nr=%s",
                                     (asgn["player_id"], new_id, nr))
                    elif asgn.get("roster_id"):
                        cur3.execute("UPDATE player_stats SET roster_id=%s WHERE match_id=%s AND druzyna='gtk' AND nr=%s",
                                     (asgn["roster_id"], new_id, nr))
                except Exception:
                    pass
            db3.commit(); cur3.close()

        # === Przywróć player_match_map ===
        if saved_pmm:
            db4 = get_db(); cur4 = db4.cursor()
            for row in saved_pmm:
                try:
                    cur4.execute(
                        "INSERT INTO player_match_map (match_id, player_id, nr, team_id) "
                        "VALUES (%s,%s,%s,%s) ON CONFLICT DO NOTHING",
                        (new_id, row["player_id"], row["nr"], row.get("team_id"))
                    )
                except Exception:
                    pass
            db4.commit(); cur4.close()

        # Przenieś plik na nowe match_id
        new_fp = os.path.join(MATCH_FILES_DIR, f"{new_id}.xlsx")
        os.rename(fp, new_fp)
        db2 = get_db(); cur2 = db2.cursor()
        cur2.execute("UPDATE matches SET file_path=%s, parser_version=%s WHERE id=%s",
                     (new_fp, PARSER_VERSION, new_id))
        db2.commit(); cur2.close()
        flash(f"Mecz przeliczony pomyślnie. Zawodnicy zachowani.", "success")
    except Exception as e:
        try: get_db().rollback()
        except: pass
        flash(f"Błąd przeliczania: {e}", "error")
    return redirect(url_for("admin_match_files"))


@app.route("/admin/reparse-all", methods=["POST"])
@login_required
def reparse_all():
    db = get_db(); cur = db.cursor()
    try:
        cur.execute("SELECT id FROM matches WHERE COALESCE(parser_version,0) < %s AND file_path IS NOT NULL",
                    (PARSER_VERSION,))
        ids = [r["id"] for r in cur.fetchall()]
        cur.close()
    except: ids = []
    ok = err = 0
    for mid in ids:
        try:
            import requests as _rq
            # Wywołaj reparse przez wewnętrzny redirect
            from flask import test_request_context
            with app.test_request_context():
                pass
            # Prościej — bezpośrednie wywołanie
            db2 = get_db(); cur2 = db2.cursor()
            cur2.execute("SELECT * FROM matches WHERE id=%s", (mid,))
            m = cur2.fetchone(); cur2.close()
            if not m: continue
            fp = m.get("file_path","")
            if not fp or not os.path.exists(fp): continue
            import openpyxl, io as _io2
            with open(fp,"rb") as f: wb = openpyxl.load_workbook(_io2.BytesIO(f.read()), data_only=True)
            report = validate_workbook(wb)
            # Zachowaj team_id i przypisanych zawodników
            saved_team_id2 = m.get("team_id")
            cur_pmm = get_db().cursor()
            cur_pmm.execute("SELECT player_id, nr, team_id FROM player_match_map WHERE match_id=%s", (mid,))
            saved_pmm2 = cur_pmm.fetchall(); cur_pmm.close()

            cur3 = get_db().cursor()
            for tbl in ["score_flow","lineup_stats","timing_stats","player_stats","match_stats","shot_zones","player_match_map"]:
                cur3.execute(f"DELETE FROM {tbl} WHERE match_id=%s", (mid,))
            cur3.execute("DELETE FROM matches WHERE id=%s", (mid,))
            get_db().commit(); cur3.close()
            _sg = parse_team_sheet(wb[report["names"][0]], "A")
            _so = parse_team_sheet(wb[report["names"][1]], "B")
            _dl = build_gtk_def_lineups(wb[report["names"][0]], wb[report["names"][1]])
            new_id = save_match_to_db(
                report["names"][1], report["names"][0],
                m["sezon"], str(m["data_meczu"]) if m["data_meczu"] else None,
                _sg, _so,
                rozgrywki=m.get("rozgrywki",""), runda=m.get("runda",""),
                kolejka=m.get("kolejka",""),
                miejsce=m.get("miejsce",""),
                def_lineups=_dl,
                team_id=saved_team_id2,
            )
            # Przywróć przypisanych zawodników
            if saved_pmm2:
                cur_r = get_db().cursor()
                for row in saved_pmm2:
                    try:
                        cur_r.execute(
                            "INSERT INTO player_match_map (match_id, player_id, nr, team_id) "
                            "VALUES (%s,%s,%s,%s) ON CONFLICT DO NOTHING",
                            (new_id, row["player_id"], row["nr"], row.get("team_id"))
                        )
                    except Exception:
                        pass
                get_db().commit(); cur_r.close()

            new_fp = os.path.join(MATCH_FILES_DIR, f"{new_id}.xlsx")
            os.rename(fp, new_fp)
            cur4 = get_db().cursor()
            cur4.execute("UPDATE matches SET file_path=%s, parser_version=%s WHERE id=%s",
                         (new_fp, PARSER_VERSION, new_id))
            get_db().commit(); cur4.close()
            ok += 1
        except: err += 1
    flash(f"Przeliczono {ok} meczów. Błędów: {err}.", "success" if not err else "warning")
    return redirect(url_for("admin_match_files"))


@app.route("/admin/delete-match-file/<int:match_id>", methods=["POST"])
@login_required
def delete_match_file(match_id):
    try:
        db = get_db(); cur = db.cursor()
        cur.execute("SELECT file_path FROM matches WHERE id=%s", (match_id,))
        row = cur.fetchone()
        if row and row.get("file_path") and os.path.exists(row["file_path"]):
            os.remove(row["file_path"])
        cur.execute("UPDATE matches SET file_path=NULL WHERE id=%s", (match_id,))
        db.commit(); cur.close()
        flash("Plik usunięty. Dane meczu w bazie zostały.", "success")
    except Exception as e:
        flash(f"Błąd: {e}", "error")
    return redirect(url_for("admin_match_files"))


@app.route("/admin/reset-matches", methods=["POST"])
@login_required
def reset_matches():
    """Usuwa wszystkie mecze i powiązane statystyki z bazy."""
    import json as _j
    try:
        db = get_db(); cur = db.cursor()
        cur.execute("DELETE FROM player_match_map")
        cur.execute("DELETE FROM score_flow")
        cur.execute("DELETE FROM lineup_stats")
        cur.execute("DELETE FROM timing_stats")
        cur.execute("DELETE FROM player_stats")
        cur.execute("DELETE FROM match_stats")
        cur.execute("DELETE FROM matches")
        db.commit()
        cur.close()
        flash("Wszystkie mecze i statystyki zostały usunięte.", "success")
    except Exception as e:
        try: get_db().rollback()
        except: pass
        flash(f"Błąd: {e}", "error")
    return redirect(url_for("index"))


@app.route("/historia")
@login_required
def historia():
    try: init_db()
    except: pass

    # Kontekst z sidebara jako domyślne filtry
    ctx_sezon   = get_setting("current_season") or ""
    ctx_klub    = get_setting("current_klub") or ""
    ctx_druzyna = get_setting("current_druzyna") or ""

    # URL params nadpisują kontekst jeśli podane
    sezon_filter      = request.args.get("sezon", ctx_sezon)
    druzyna_filter    = request.args.get("druzyna", "")  # team_id
    data_od           = request.args.get("data_od","")
    data_do           = request.args.get("data_do","")
    przeciwnik_filter = request.args.get("przeciwnik","").strip().lower()
    show_all          = request.args.get("all","")  # "1" = ignoruj kontekst

    if show_all:
        sezon_filter   = request.args.get("sezon","")
        druzyna_filter = ""

    db = get_db(); cur = db.cursor()

    # Rozwiąż team_id z kontekstu jeśli nie podano jawnie
    ctx_team_id = None
    if not show_all and ctx_druzyna and ctx_sezon and ctx_klub:
        try:
            cur.execute("""
                SELECT t.id FROM teams t
                JOIN seasons s ON t.season_id=s.id
                JOIN clubs c ON s.club_id=c.id
                WHERE c.name=%s AND s.name=%s AND t.name=%s
            """, (ctx_klub, ctx_sezon, ctx_druzyna))
            tr = cur.fetchone()
            if tr: ctx_team_id = tr["id"]
        except: pass

    # Buduj WHERE
    conditions = []
    params = []
    if sezon_filter:
        conditions.append("m.sezon=%s"); params.append(sezon_filter)
    if druzyna_filter:
        conditions.append("m.team_id=%s"); params.append(int(druzyna_filter))
    elif ctx_team_id and not show_all:
        # Pokaż mecze przypisane do drużyny LUB mecze bez team_id z pasującą nazwą klubu
        conditions.append("(m.team_id=%s OR (m.team_id IS NULL AND (m.team_name_a=%s OR m.nazwa_gtk=%s)))")
        params.append(ctx_team_id); params.append(ctx_klub); params.append(ctx_klub)
    if data_od:
        conditions.append("m.data_meczu >= %s"); params.append(data_od)
    if data_do:
        conditions.append("m.data_meczu <= %s"); params.append(data_do)
    if przeciwnik_filter:
        conditions.append("LOWER(m.przeciwnik) LIKE %s"); params.append(f"%{przeciwnik_filter}%")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    try:
        cur.execute(f"""
            SELECT m.id, m.sezon, m.data_meczu, m.przeciwnik,
                   COALESCE(m.rozgrywki,'') as rozgrywki,
                   COALESCE(m.runda,'') as runda,
                   COALESCE(m.miejsce,'') as miejsce,
                   m.wynik_gtk, m.wynik_opp,
                   t.name as team_name,
                   s.name as season_name,
                   c.name as club_name
            FROM matches m
            LEFT JOIN teams t ON m.team_id=t.id
            LEFT JOIN seasons s ON t.season_id=s.id
            LEFT JOIN clubs c ON s.club_id=c.id
            {where}
            ORDER BY m.data_meczu DESC NULLS LAST, m.created_at DESC
        """, params)
    except:
        cur.execute(f"SELECT * FROM matches m {where} ORDER BY m.created_at DESC", params)
    matches = cur.fetchall()

    # Sezony dla aktywnego klubu z bazy seasons (spójne z kontekstem)
    sezony_db = []
    if ctx_klub:
        try:
            cur.execute("""
                SELECT s.name FROM seasons s
                JOIN clubs c ON s.club_id=c.id
                WHERE c.name=%s ORDER BY s.name DESC
            """, (ctx_klub,))
            sezony_db = [r["name"] for r in cur.fetchall()]
        except: pass
    if not sezony_db:
        cur.execute("SELECT DISTINCT sezon FROM matches ORDER BY sezon DESC")
        sezony_db = [r["sezon"] for r in cur.fetchall()]
    sezony = sezony_db
    cur.execute("SELECT DISTINCT przeciwnik FROM matches ORDER BY przeciwnik")
    przeciwnicy = [r["przeciwnik"] for r in cur.fetchall()]
    cur.close()

    # Etykieta aktywnego kontekstu
    ctx_label = ""
    if not show_all and (ctx_klub or ctx_sezon or ctx_druzyna):
        parts = [p for p in [ctx_klub, ctx_sezon, ctx_druzyna] if p]
        ctx_label = " · ".join(parts)

    rows = ""
    for i, m in enumerate(matches):
        wynik = f"{m['wynik_gtk']} : {m['wynik_opp']}"
        if m['wynik_gtk'] > m['wynik_opp']:   badge = '<span class="badge-win">W</span>'
        elif m['wynik_gtk'] < m['wynik_opp']: badge = '<span class="badge-loss">P</span>'
        else:                                  badge = '<span class="badge-draw">R</span>'
        dt      = m['data_meczu'].strftime('%d.%m.%Y') if m['data_meczu'] else '—'
        rozg    = m.get('rozgrywki','') or '—'
        runda   = m.get('runda','') or '—'
        miejsce = m.get('miejsce','') or '—'
        team_lbl = m.get('team_name') or '—'
        bg = "background:#f8f9ff" if i%2==0 else ""
        rows += f"""<tr style="{bg}" id="mrow-{m['id']}">
            <td style="width:32px;text-align:center"><input type="checkbox" class="cmp-chk" data-id="{m['id']}" onchange="cmpToggle(this)" style="width:14px;height:14px;cursor:pointer"></td>
            <td style="width:44px">{badge}</td>
            <td style="font-size:.82rem;font-weight:600">{dt}</td>
            <td style="font-size:.78rem;color:#666">{team_lbl}</td>
            <td style="font-size:.78rem;color:#666">{rozg}</td>
            <td><a href="/mecz/{m['id']}" class="fw-bold text-decoration-none" style="color:#1a2b4a">{m['przeciwnik']}</a></td>
            <td style="font-size:.78rem;color:#666">{miejsce}</td>
            <td class="text-center"><span style="font-size:.95rem;font-weight:700">{wynik}</span></td>
            <td class="text-center">
              <a href="/mecz/{m['id']}" class="btn btn-outline-primary btn-sm" style="font-size:.72rem">Raport</a>
              <a href="/mecz/{m['id']}/delete" class="btn btn-outline-danger btn-sm ms-1" style="font-size:.72rem"
                 onclick="return confirm('Usunąć ten mecz?')">&#10005;</a>
            </td>
        </tr>"""

    # Badge kontekstu — pokazuje aktywny klub/sezon/drużynę
    if ctx_klub and ctx_sezon and ctx_druzyna:
        ctx_badge = (f'<span style="font-size:11px;background:#E6F1FB;color:#0C447C;padding:3px 10px;border-radius:12px;font-weight:500">'
                     f'{ctx_klub} &middot; {ctx_sezon} &middot; {ctx_druzyna}</span>')
    elif ctx_sezon:
        ctx_badge = f'<span style="font-size:11px;background:#E6F1FB;color:#0C447C;padding:3px 10px;border-radius:12px;font-weight:500">{ctx_sezon}</span>'
    else:
        ctx_badge = '<span style="font-size:11px;color:#aaa">Brak kontekstu — ustaw w sidebarze</span>'
    opp_opts    = "".join([f'<option value="{p}" {"selected" if p.lower()==przeciwnik_filter else ""}>{p}</option>' for p in przeciwnicy])
    season_opts = "".join([f'<option value="{s}" {"selected" if s==sezon_filter else ""}>{s}</option>' for s in sezony])

    ctx_bar = ""
    if ctx_label:
        ctx_bar = f"""
<div style="display:flex;align-items:center;gap:10px;padding:8px 14px;background:#E6F1FB;border-radius:8px;margin-bottom:10px;font-size:12px">
  <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#185FA5" stroke-width="2"><circle cx="12" cy="12" r="10"/><line x1="12" y1="8" x2="12" y2="12"/><line x1="12" y1="16" x2="12.01" y2="16"/></svg>
  <span style="color:#185FA5;font-weight:500">Aktywny kontekst:</span>
  <span style="color:#0C447C">{ctx_label}</span>
  <a href="/historia?all=1" style="margin-left:auto;font-size:11px;color:#185FA5;text-decoration:none">Pokaż wszystkie mecze &#8594;</a>
</div>"""

    content = f"""
<div class="page-title">&#128203; Historia meczów</div>

{ctx_bar}

<div class="card p-3 mb-3">
  <form method="GET">
    <div class="row g-2 align-items-end">
      <div class="col-auto">
        <label class="form-label mb-1" style="font-size:.75rem;font-weight:600">Sezon</label>
        <select name="sezon" class="form-select form-select-sm" style="width:110px">
          <option value="">Wszystkie</option>
          {season_opts}
        </select>
      </div>
      <div class="col-auto">
        <label class="form-label mb-1" style="font-size:.75rem;font-weight:600">Data od</label>
        <input type="date" name="data_od" class="form-control form-control-sm" value="{data_od}" style="width:140px">
      </div>
      <div class="col-auto">
        <label class="form-label mb-1" style="font-size:.75rem;font-weight:600">Data do</label>
        <input type="date" name="data_do" class="form-control form-control-sm" value="{data_do}" style="width:140px">
      </div>
      <div class="col-auto">
        <label class="form-label mb-1" style="font-size:.75rem;font-weight:600">Przeciwnik</label>
        <select name="przeciwnik" class="form-select form-select-sm" style="width:180px">
          <option value="">Wszyscy</option>
          {opp_opts}
        </select>
      </div>
      <div class="col-auto">
        <button type="submit" class="btn btn-primary btn-sm">Filtruj</button>
        <a href="/historia" class="btn btn-outline-secondary btn-sm ms-1">Wyczyść</a>
      </div>
      <div class="col-auto ms-auto">
        <span style="font-size:.82rem;color:#888;line-height:2.2">{len(matches)} meczów</span>
        <a href="/porownaj" class="btn btn-sm ms-2" style="background:#534AB7;color:#fff;font-size:.75rem">&#9654; Porównaj drużyny</a>
      </div>
    </div>
  </form>
</div>

<div class="card">
  <div class="card-body p-2">
    <div class="table-responsive">
      <table class="table table-hover mb-0">
        <thead><tr>
          <th style="width:32px"><input type="checkbox" id="cmp-all" onclick="cmpAll(this)" style="width:14px;height:14px;cursor:pointer" title="Zaznacz wszystkie"></th>
          <th style="width:28px"></th>
          <th style="cursor:pointer" onclick="sortTable(2)">Data &#8597;</th>
          <th>Drużyna</th>
          <th>Rozgrywki</th>
          <th style="cursor:pointer" onclick="sortTable(5)">Przeciwnik &#8597;</th>
          <th>Miejsce</th>
          <th class="text-center">Wynik</th>
          <th class="text-center">Akcje</th>
        </tr></thead>
        <tbody id="matchTable">
          {rows if rows else '<tr><td colspan="9" class="text-center text-muted py-4">Brak meczów spełniających kryteria</td></tr>'}
        </tbody>
      </table>
  </div>
  <div id="cmp-bar" style="display:none;padding:8px 12px;background:#534AB7;border-radius:0 0 8px 8px;display:flex;align-items:center;justify-content:space-between">
    <span style="color:#fff;font-size:12px">Zaznaczono: <strong id="cmp-count">0</strong> mecze</span>
    <button onclick="window.location='/porownaj?ids='+Array.from(document.querySelectorAll('.cmp-chk:checked')).map(c=>c.dataset.id).join(',')" style="background:#fff;color:#534AB7;border:none;border-radius:6px;padding:5px 14px;font-size:12px;font-weight:600;cursor:pointer">Porównaj zaznaczone &#8594;</button>
  </div>
    </div>
  </div>
</div>"""

    scripts = """<script>
let sortDir = {};
function sortTable(col) {
    const tbody = document.getElementById('matchTable');
    const rows = Array.from(tbody.querySelectorAll('tr'));
    sortDir[col] = !sortDir[col];
    rows.sort((a, b) => {
        const av = a.cells[col]?.textContent.trim() || '';
        const bv = b.cells[col]?.textContent.trim() || '';
        // Sortowanie daty DD.MM.YYYY
        if (col === 2) {
            const toDate = s => { const p=s.split('.'); return p.length===3?new Date(p[2],p[1]-1,p[0]):new Date(0); }
            return sortDir[col] ? toDate(bv)-toDate(av) : toDate(av)-toDate(bv);
        }
        return sortDir[col] ? bv.localeCompare(av,'pl') : av.localeCompare(bv,'pl');
    });
    rows.forEach(r => tbody.appendChild(r));
}
function gtkTab(id) {
    ['gtk_q','gtk_p','gtk_l','gtk_t'].forEach(function(p) {
        var pane = document.getElementById('gpane-'+p);
        var btn  = document.getElementById('gtktab-'+p);
        if (!pane || !btn) return;
        var active = p === id;
        pane.style.display = active ? 'block' : 'none';
        btn.style.borderBottomColor = active ? '#1a2b4a' : 'transparent';
        btn.style.color = active ? '#1a2b4a' : '#666';
        btn.style.fontWeight = active ? '500' : 'normal';
    });
}
function oppTab(id) {
    ['opp_q','opp_t'].forEach(function(p) {
        var pane = document.getElementById('opane-'+p);
        var btn  = document.getElementById('opptab-'+p);
        if (!pane || !btn) return;
        var active = p === id;
        pane.style.display = active ? 'block' : 'none';
        btn.style.borderBottomColor = active ? '#1a2b4a' : 'transparent';
        btn.style.color = active ? '#1a2b4a' : '#666';
        btn.style.fontWeight = active ? '500' : 'normal';
    });
}
function timSwitch(druzyna, view) {
    var sum = document.getElementById("tim-sum-" + druzyna);
    var q   = document.getElementById("tim-q-"   + druzyna);
    var btnSum = document.getElementById("tim-btn-sum-" + druzyna);
    var btnQ   = document.getElementById("tim-btn-q-"   + druzyna);
    if (!sum || !q) return;
    if (view === "sum") {
        sum.style.display = ""; q.style.display = "none";
        btnSum.style.borderBottomColor = "#1a2b4a"; btnSum.style.color = "#1a2b4a"; btnSum.style.fontWeight = "500";
        btnQ.style.borderBottomColor = "transparent"; btnQ.style.color = "#666"; btnQ.style.fontWeight = "normal";
    } else {
        sum.style.display = "none"; q.style.display = "";
        btnQ.style.borderBottomColor = "#1a2b4a"; btnQ.style.color = "#1a2b4a"; btnQ.style.fontWeight = "500";
        btnSum.style.borderBottomColor = "transparent"; btnSum.style.color = "#666"; btnSum.style.fontWeight = "normal";
    }
}
</script>"""

    scripts += """<script>
var _cmpIds = new Set();
function cmpToggle(el) {
    var id = el.dataset.id;
    if (el.checked) { _cmpIds.add(id); } else { _cmpIds.delete(id); }
    var bar = document.getElementById('cmp-bar');
    var cnt = document.getElementById('cmp-count');
    if (_cmpIds.size >= 2) {
        bar.style.display = 'flex';
        cnt.textContent = _cmpIds.size;
    } else {
        bar.style.display = 'none';
    }
    if (_cmpIds.size > 4) { el.checked = false; _cmpIds.delete(id); alert('Maksymalnie 4 mecze'); }
}
function cmpAll(el) {
    document.querySelectorAll('.cmp-chk').forEach(function(c) {
        c.checked = el.checked;
        if (el.checked) _cmpIds.add(c.dataset.id); else _cmpIds.delete(c.dataset.id);
    });
    var bar = document.getElementById('cmp-bar');
    var cnt = document.getElementById('cmp-count');
    bar.style.display = _cmpIds.size >= 2 ? 'flex' : 'none';
    if (cnt) cnt.textContent = _cmpIds.size;
}
</script>"""
    return html_response(base(content, scripts, active="history"))

# ══════════════════════════════════════════════════════════════════════════════
# RAPORT MECZU
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/mecz/<int:match_id>")
@login_required
def mecz(match_id):
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
    m = cur.fetchone()
    if not m: flash("Mecz nie istnieje","error"); return redirect(url_for("historia"))

    gtk_name = (m.get("team_name_a","") or m.get("nazwa_gtk","") or "").strip() or get_setting("gtk_name") or "GTK"
    name_opp = (m.get("team_name_b","") or m["przeciwnik"])

    cur.execute("SELECT * FROM match_stats WHERE match_id=%s ORDER BY kwarta", (match_id,))
    all_stats = cur.fetchall()

    cur.execute("SELECT * FROM player_stats WHERE match_id=%s", (match_id,))
    all_players = cur.fetchall()

    cur.execute("SELECT * FROM timing_stats WHERE match_id=%s", (match_id,))
    all_timing = cur.fetchall()

    # Piątki
    try:
        cur.execute("""SELECT * FROM lineup_stats WHERE match_id=%s AND druzyna='gtk'
                       ORDER BY poss DESC""", (match_id,))
        all_lineups = list(cur.fetchall())
    except:
        all_lineups = []

    try:
        cur.execute("""SELECT * FROM lineup_stats WHERE match_id=%s AND druzyna='gtk_def'
                       ORDER BY poss DESC""", (match_id,))
        all_lineups_def = list(cur.fetchall())
    except:
        all_lineups_def = []

    # Oblicz NetRtg per piątka: ORtg_off - DRtg_def
    _off_map = {}
    for _lu in all_lineups:
        _k = _lu["lineup"]; _p = int(_lu.get("poss",0) or 0); _pts = int(_lu.get("pts",0) or 0)
        if _p > 0: _off_map[_k] = _pts * 100 / _p
    _def_map = {}
    for _lu in all_lineups_def:
        _k = _lu["lineup"]; _p = int(_lu.get("poss",0) or 0); _pts = int(_lu.get("pts",0) or 0)
        if _p > 0: _def_map[_k] = _pts * 100 / _p
    # Wstrzyknij net_rtg do piątek OFF
    for _lu in all_lineups:
        _k = _lu["lineup"]
        if _k in _off_map and _k in _def_map:
            _lu["net_rtg"] = round(_off_map[_k] - _def_map[_k], 1)
        else:
            _lu["net_rtg"] = None
    # Zbuduj listę NET posortowaną po NetRtg
    _seen_net = {}
    for _lu in all_lineups + all_lineups_def:
        _k = _lu["lineup"]
        if _k not in _seen_net:
            _seen_net[_k] = dict(_lu)
            _seen_net[_k]["ortg"] = _off_map.get(_k)
            _seen_net[_k]["drtg"] = _def_map.get(_k)
            _net = (_off_map[_k] - _def_map[_k]) if (_k in _off_map and _k in _def_map) else None
            _seen_net[_k]["net_rtg"] = round(_net, 1) if _net is not None else None
    all_lineups_net = sorted(_seen_net.values(),
        key=lambda x: (x["net_rtg"] is not None, x["net_rtg"] if x["net_rtg"] is not None else 0),
        reverse=True)

    # Score flow
    try:
        cur.execute("""SELECT kwarta, czas_sek, pts_gtk, pts_opp
                       FROM score_flow WHERE match_id=%s
                       ORDER BY kwarta, czas_sek""", (match_id,))
        flow_rows = list(cur.fetchall())
    except:
        flow_rows = []

    # Mapa roster_id → "Nazwisko I." dla GTK
    try:
        # Najpierw spróbuj z players (Struktura klubów przez player_id)
        cur.execute("""SELECT ps.id as ps_id, p.imie, p.nazwisko
                       FROM player_stats ps
                       JOIN players p ON ps.player_id = p.id
                       WHERE ps.match_id=%s AND ps.druzyna='gtk'""", (match_id,))
        roster_map = {row["ps_id"]: f"{row['nazwisko']} {row['imie'][0]}." for row in cur.fetchall()}
        # Uzupełnij z roster (stary moduł) dla tych bez player_id
        if not roster_map:
            cur.execute("""SELECT ps.id as ps_id, r.imie, r.nazwisko
                           FROM player_stats ps
                           JOIN roster r ON ps.roster_id = r.id
                           WHERE ps.match_id=%s AND ps.druzyna='gtk'""", (match_id,))
            roster_map = {row["ps_id"]: f"{row['nazwisko']} {row['imie'][0]}." for row in cur.fetchall()}
    except:
        roster_map = {}

    # Mapa nr → nazwisko dla piątek (wszystkie źródła)
    nr_name_map = build_nr_name_map(cur, match_id)
    cur.close()

    # ── Momentum kwart ────────────────────────────────────────────────────────
    def quarter_momentum():
        rows = ""
        for qn in [1, 2, 3, 4]:
            qg = next((r for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==qn), {})
            qo = next((r for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==qn), {})
            pg = qg.get("pts", 0) or 0
            po = qo.get("pts", 0) or 0
            if pg > po:
                res = f'<span style="background:#e8f5e9;color:#1a5c2a;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600">W +{pg-po}</span>'
            elif po > pg:
                res = f'<span style="background:#ffebee;color:#8b1a1a;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600">L -{po-pg}</span>'
            else:
                res = '<span style="background:#f5f5f5;color:#666;padding:2px 8px;border-radius:12px;font-size:11px">=</span>'
            q_labels = {1:"#c8e6c9", 2:"#bbdefb", 3:"#fff9c4", 4:"#fce4ec"}
            q_text   = {1:"#1a5c2a", 2:"#0d47a1", 3:"#f57f17", 4:"#880e4f"}
            efg_g = f"{(qg.get('p2m',0)+1.5*qg.get('p3m',0))/(qg.get('p2a',0)+qg.get('p3a',0)):.0%}" if (qg.get('p2a',0)+qg.get('p3a',0)) else "-"
            efg_o = f"{(qo.get('p2m',0)+1.5*qo.get('p3m',0))/(qo.get('p2a',0)+qo.get('p3a',0)):.0%}" if (qo.get('p2a',0)+qo.get('p3a',0)) else "-"
            rows += f"""<tr>
              <td><span style="background:{q_labels[qn]};color:{q_text[qn]};padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600">{qn}Q</span></td>
              <td class="text-center">{res}</td>
              <td class="text-center fw-bold" style="color:#1a2b4a">{pg}</td>
              <td class="text-center" style="color:#888">{po}</td>
              <td class="text-center">{efg_g}</td>
              <td class="text-center" style="color:#888">{efg_o}</td>
              <td class="text-center">{qg.get('br',0) or 0}</td>
              <td class="text-center" style="color:#888">{qo.get('br',0) or 0}</td>
              <td class="text-center">{qg.get('poss',0) or 0}</td>
            </tr>"""
        q_results = []
        for qn in [1,2,3,4]:
            qg = next((r for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==qn), {})
            qo = next((r for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==qn), {})
            pg = qg.get("pts",0) or 0; po = qo.get("pts",0) or 0
            q_results.append("W" if pg>po else ("L" if po>pg else "="))
        streak_html = ""
        for r in q_results:
            col = "#1a6b3c" if r=="W" else ("#8b1a1a" if r=="L" else "#888")
            streak_html += f'<span style="display:inline-flex;align-items:center;justify-content:center;width:28px;height:28px;border-radius:50%;background:{col};color:#fff;font-size:11px;font-weight:600;margin-right:4px">{r}</span>'
        best = 0; cur_streak = 0
        for r in q_results:
            cur_streak = cur_streak+1 if r=="W" else 0
            best = max(best, cur_streak)
        best_html = f'<span style="font-size:11px;color:#888">Najlepsza seria: <b style="color:#1a6b3c">{best} kwart{" z rzędu" if best>1 else ""}</b></span>' if best >= 1 else ""
        return f"""
        <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;flex-wrap:wrap">
          {streak_html}{best_html}
        </div>
        <div class="table-responsive">
        <table class="table table-hover mb-0" style="font-size:.82rem">
          <thead><tr>
            <th>Q</th><th class="text-center">Wynik</th>
            <th class="text-center" style="color:#1a2b4a">PKT</th>
            <th class="text-center" style="color:#888">PKT rywal</th>
            <th class="text-center" style="color:#1a2b4a">eFG%</th>
            <th class="text-center" style="color:#888">eFG% rywal</th>
            <th class="text-center" style="color:#1a2b4a">TO</th>
            <th class="text-center" style="color:#888">TO rywal</th>
            <th class="text-center">POSS</th>
          </tr></thead>
          <tbody>{rows}</tbody>
        </table></div>"""

    # ── Clutch stats ──────────────────────────────────────────────────────────
    def clutch_stats():
        import math as _math
        # Clutch = ostatnie 1/3 posiadań Q4 (i dogrywek), zaokrąglone w górę.
        # Np. 25 posiadań → ceil(25/3) = 9 clutch posiadań → ratio = 9/25 = 0.36
        # Każda drużyna ma swój własny ratio oparty na jej liczbie posiadań.
        clutch_qtrs = sorted({r["kwarta"] for r in all_stats
                               if r.get("kwarta", 0) >= 4 and r.get("druzyna") in ("gtk","opp")})
        if not clutch_qtrs:
            return '<p class="text-muted p-3 mb-0" style="font-size:.82rem">Brak danych Q4/OT w tym meczu.</p>'

        keys_g = ["pts_g","poss_g","p2m_g","p2a_g","p3m_g","p3a_g","ftm_g","fta_g",
                  "br_g","fd_g","ast_g","oreb_g","dreb_g","stl_g","blk_g","d2m_g","d2a_g"]
        keys_o = [k.replace("_g","_o") for k in keys_g]
        c = {k: 0 for k in keys_g + keys_o}
        clutch_poss_g_total = 0  # do etykiety w stopce
        clutch_poss_o_total = 0

        for qn in clutch_qtrs:
            qg = next((r for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==qn), {})
            qo = next((r for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==qn), {})
            poss_g = qg.get("poss", 0) or 0
            poss_o = qo.get("poss", 0) or 0
            # Ostatnie 1/3 posiadań (ceil), ratio = clutch_poss / total_poss
            cp_g = _math.ceil(poss_g / 3) if poss_g else 0
            cp_o = _math.ceil(poss_o / 3) if poss_o else 0
            ratio_g = cp_g / poss_g if poss_g else 0
            ratio_o = cp_o / poss_o if poss_o else 0
            clutch_poss_g_total += cp_g
            clutch_poss_o_total += cp_o
            def _sc(d, k, ratio): return round((d.get(k, 0) or 0) * ratio)
            for k in ["pts","p2m","p2a","p3m","p3a","ftm","fta","br","fd","ast","oreb","dreb","stl","blk","d2m","d2a"]:
                c[f"{k}_g"] += _sc(qg, k, ratio_g)
                c[f"{k}_o"] += _sc(qo, k, ratio_o)
            # poss zliczamy wprost jako liczbę clutch posiadań
            c["poss_g"] += cp_g
            c["poss_o"] += cp_o

        def pct(m, a):  return f"{round(m/a*100)}%" if a else "—"
        def ppp(p, po): return f"{p/po:.2f}" if po else "—"
        def topct(br, poss): return f"{round(br/poss*100)}%" if poss else "—"

        efg_g = pct(c["p2m_g"]+int(1.5*c["p3m_g"]), c["p2a_g"]+c["p3a_g"])
        efg_o = pct(c["p2m_o"]+int(1.5*c["p3m_o"]), c["p2a_o"]+c["p3a_o"])
        p2p_g = pct(c["p2m_g"], c["p2a_g"]); p2p_o = pct(c["p2m_o"], c["p2a_o"])
        p3p_g = pct(c["p3m_g"], c["p3a_g"]); p3p_o = pct(c["p3m_o"], c["p3a_o"])
        ftp_g = pct(c["ftm_g"], c["fta_g"]); ftp_o = pct(c["ftm_o"], c["fta_o"])
        ppp_g = ppp(c["pts_g"], c["poss_g"]); ppp_o = ppp(c["pts_o"], c["poss_o"])
        to_g  = topct(c["br_g"], c["poss_g"]); to_o = topct(c["br_o"], c["poss_o"])
        ortg_g = round(c["pts_g"]/c["poss_g"]*100) if c["poss_g"] else 0
        ortg_o = round(c["pts_o"]/c["poss_o"]*100) if c["poss_o"] else 0
        net_g  = ortg_g - ortg_o; net_o = ortg_o - ortg_g
        net_gs = ("+"+str(net_g)) if net_g>0 else str(net_g)
        net_os = ("+"+str(net_o)) if net_o>0 else str(net_o)

        wc     = "GTK lepsza" if c["pts_g"]>c["pts_o"] else ("Rywal lepszy" if c["pts_o"]>c["pts_g"] else "Remis")
        wc_bg  = "#e8f5e9" if c["pts_g"]>c["pts_o"] else ("#ffebee" if c["pts_o"]>c["pts_g"] else "#f5f5f5")
        wc_col = "#1a6b3c" if c["pts_g"]>c["pts_o"] else ("#8b1a1a" if c["pts_o"]>c["pts_g"] else "#666")

        def col(vg, vo, higher=True):
            try:
                fg=float(str(vg).replace('%','').replace('—','0').replace('+',''))
                fo=float(str(vo).replace('%','').replace('—','0').replace('+',''))
                if higher: return ("#1a6b3c" if fg>fo else "#8b1a1a" if fg<fo else "#888"), \
                                  ("#1a6b3c" if fo>fg else "#8b1a1a" if fo<fg else "#888")
                else:      return ("#1a6b3c" if fg<fo else "#8b1a1a" if fg>fo else "#888"), \
                                  ("#1a6b3c" if fo<fg else "#8b1a1a" if fo>fg else "#888")
            except: return "#1a2b4a","#1a2b4a"

        qtrs_label = "+".join(f"Q{q}" for q in sorted(clutch_qtrs))

        # ── Scoreboard ────────────────────────────────────────────────────────
        win_gtk = c["pts_g"] > c["pts_o"]
        win_opp = c["pts_o"] > c["pts_g"]
        score_c_g = "#1a6b3c" if win_gtk else ("#8b1a1a" if win_opp else "#856404")
        score_c_o = "#1a6b3c" if win_opp else ("#8b1a1a" if win_gtk else "#856404")
        ortg_gs = str(ortg_g) if c["poss_g"] else "—"
        ortg_os = str(ortg_o) if c["poss_o"] else "—"
        net_color = "#1a6b3c" if net_g > 0 else ("#8b1a1a" if net_g < 0 else "#856404")

        # ── Efficiency strip ────────────────────────────────────────────────
        def eff_pill2(lbl, vg, vo, higher=True, neutral=False):
            if neutral: cg = co = "#1a2b4a"
            else: cg, co = col(vg, vo, higher=higher)
            return (f'<div style="flex:1;background:#f4f6fb;border-radius:10px;padding:10px 8px;text-align:center;min-width:0;border:0.5px solid #e3e8f0">'
                    f'<div style="font-size:9px;color:#999;text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px">{lbl}</div>'
                    f'<div style="display:flex;align-items:center;justify-content:center;gap:6px">'
                    f'<span style="font-size:15px;font-weight:700;color:{cg}">{vg}</span>'
                    f'<span style="font-size:9px;color:#bbb">vs</span>'
                    f'<span style="font-size:15px;font-weight:700;color:{co}">{vo}</span>'
                    f'</div></div>')

        eff_strip = (eff_pill2("eFG%",  efg_g,  efg_o) +
                     eff_pill2("POSS",  c["poss_g"], c["poss_o"], neutral=True) +
                     eff_pill2("TO%",   to_g,   to_o,  higher=False) +
                     eff_pill2("PPP",   ppp_g,  ppp_o))

        # ── Face-off bars ────────────────────────────────────────────────────
        def face_bar(lbl, mg, ag, mo, ao, pts_type=""):
            pg = pct(mg, ag); po = pct(mo, ao)
            max_w = 46
            wg = round(float(str(pg).replace('%','') or 0) / 100 * max_w) if ag else 0
            wo = round(float(str(po).replace('%','') or 0) / 100 * max_w) if ao else 0
            badge_bg = "#2e5090" if pts_type=="3PT" else ("#856404" if pts_type=="FT" else "#1a6b3c")
            return f"""
<div style="margin-bottom:14px">
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:5px">
    <span style="font-size:10px;font-weight:700;color:#fff;background:{badge_bg};border-radius:4px;padding:1px 7px">{lbl}</span>
    <span style="font-size:10px;color:#888">{mg}/{ag}</span>
    <span style="flex:1;height:1px;background:#e8ecf0"></span>
    <span style="font-size:10px;color:#888">{mo}/{ao}</span>
  </div>
  <div style="display:flex;align-items:center;gap:4px">
    <span style="font-size:13px;font-weight:700;color:{score_c_g};width:38px;text-align:right">{pg}</span>
    <div style="flex:1;display:flex;height:20px;border-radius:4px;overflow:hidden;background:#eef0f5">
      <div style="flex:1;display:flex;justify-content:flex-end;align-items:center">
        <div style="width:{wg}%;height:100%;background:#1a6b3c;border-radius:3px 0 0 3px"></div>
      </div>
      <div style="width:2px;background:#fff;flex-shrink:0"></div>
      <div style="flex:1;display:flex;align-items:center">
        <div style="width:{wo}%;height:100%;background:#8b1a1a;border-radius:0 3px 3px 0"></div>
      </div>
    </div>
    <span style="font-size:13px;font-weight:700;color:{score_c_o};width:38px">{po}</span>
  </div>
</div>"""

        face_bars = (face_bar("2PT", c["p2m_g"],c["p2a_g"],c["p2m_o"],c["p2a_o"],"2PT") +
                     face_bar("3PT", c["p3m_g"],c["p3a_g"],c["p3m_o"],c["p3a_o"],"3PT") +
                     face_bar("FT",  c["ftm_g"],c["fta_g"],c["ftm_o"],c["fta_o"],"FT"))

        # ── Stat rows (mini-table) ───────────────────────────────────────────
        def srow(lbl, vg, vo, higher=True, neutral=False):
            if neutral: cg_s = co_s = "#1a2b4a"
            else:
                rg, ro = col(vg, vo, higher=higher)
                cg_s = rg; co_s = ro
            return (f'<div style="display:flex;align-items:center;padding:5px 0;border-bottom:0.5px solid #eceef2">'
                    f'<span style="flex:1;font-size:12px;font-weight:600;color:{cg_s};text-align:right;padding-right:12px">{vg}</span>'
                    f'<span style="width:90px;font-size:10px;color:#999;text-align:center;text-transform:uppercase;letter-spacing:.4px;flex-shrink:0">{lbl}</span>'
                    f'<span style="flex:1;font-size:12px;font-weight:600;color:{co_s};text-align:left;padding-left:12px">{vo}</span>'
                    f'</div>')

        stat_rows = (
            srow("AST",    c["ast_g"],  c["ast_o"]) +
            srow("STL",    c["stl_g"],  c["stl_o"]) +
            srow("BLK",    c["blk_g"],  c["blk_o"]) +
            srow("ZB Off", c["oreb_g"], c["oreb_o"]) +
            srow("ZB Def", c["dreb_g"], c["dreb_o"]) +
            srow("DOB M/A",f"{c['d2m_g']}/{c['d2a_g']}", f"{c['d2m_o']}/{c['d2a_o']}", neutral=True) +
            srow("FD",     c["fd_g"],   c["fd_o"]) +
            srow("TO",     c["br_g"],   c["br_o"],  higher=False) +
            srow("TO%",    to_g,        to_o,       higher=False) +
            srow("NETrtg", net_gs,      net_os)
        )

        wc_bg2  = "#e8f5e9" if win_gtk else ("#ffebee" if win_opp else "#fff8e1")
        wc_col2 = "#1a6b3c" if win_gtk else ("#8b1a1a" if win_opp else "#856404")

        return f"""
<div style="border-radius:12px;overflow:hidden;border:0.5px solid #e3e8f0;background:#fff">

  <!-- HEADER -->
  <div style="padding:10px 16px;display:flex;align-items:center;gap:10px;border-bottom:0.5px solid #eceef2;background:#f8f9fc">
    <span style="background:#EF9F27;color:#fff;padding:2px 10px;border-radius:20px;font-size:11px;font-weight:700;letter-spacing:.4px">{qtrs_label} CLUTCH</span>
    <span style="font-size:11px;color:#aaa">Ostatnie 3 min Q4</span>
    <span style="margin-left:auto;background:{wc_bg2};color:{wc_col2};padding:2px 10px;border-radius:20px;font-size:11px;font-weight:600">{wc} w clutch</span>
  </div>

  <!-- SCOREBOARD -->
  <div style="display:grid;grid-template-columns:1fr auto 1fr;align-items:center;padding:18px 16px;gap:12px;border-bottom:0.5px solid #eceef2">
    <div style="text-align:right">
      <div style="font-size:11px;font-weight:600;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">{gtk_name}</div>
      <div style="font-size:48px;font-weight:800;color:{score_c_g};line-height:1">{c["pts_g"]}</div>
      <div style="font-size:11px;color:#aaa;margin-top:5px">ORtg <span style="color:#555;font-weight:600">{ortg_gs}</span></div>
    </div>
    <div style="text-align:center;padding:0 8px">
      <div style="font-size:11px;color:#bbb;font-weight:400;margin-bottom:8px">vs</div>
      <div style="background:#f4f6fb;border-radius:8px;padding:6px 12px;border:0.5px solid #e3e8f0">
        <div style="font-size:9px;color:#aaa;text-transform:uppercase;letter-spacing:.4px">NETrtg</div>
        <div style="font-size:18px;font-weight:800;color:{net_color}">{net_gs}</div>
      </div>
    </div>
    <div style="text-align:left">
      <div style="font-size:11px;font-weight:600;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">{name_opp}</div>
      <div style="font-size:48px;font-weight:800;color:{score_c_o};line-height:1">{c["pts_o"]}</div>
      <div style="font-size:11px;color:#aaa;margin-top:5px">ORtg <span style="color:#555;font-weight:600">{ortg_os}</span></div>
    </div>
  </div>

  <!-- EFFICIENCY STRIP -->
  <div style="display:flex;gap:6px;padding:10px 16px;border-bottom:0.5px solid #eceef2;background:#fafbfd">
    {eff_strip}
  </div>

  <!-- SHOOTING FACE-OFF -->
  <div style="padding:14px 16px;border-bottom:0.5px solid #eceef2">
    <div style="font-size:9px;color:#aaa;text-transform:uppercase;letter-spacing:.6px;margin-bottom:10px;display:flex;align-items:center;gap:8px">
      <span style="color:#1a6b3c;font-weight:700">{gtk_name}</span>
      <span style="flex:1;height:1px;background:#eceef2"></span>
      <span>Rzuty z pola</span>
      <span style="flex:1;height:1px;background:#eceef2"></span>
      <span style="color:#8b1a1a;font-weight:700">{name_opp}</span>
    </div>
    {face_bars}
  </div>

  <!-- STAT TABLE -->
  <div style="padding:10px 16px 14px">
    <div style="font-size:9px;color:#aaa;text-transform:uppercase;letter-spacing:.6px;margin-bottom:8px">Szczeg\u00f3\u0142owe statystyki</div>
    <div style="display:flex;align-items:center;margin-bottom:4px">
      <span style="flex:1;font-size:10px;font-weight:700;color:#1a6b3c;text-align:right;padding-right:12px">{gtk_name}</span>
      <span style="width:90px;flex-shrink:0"></span>
      <span style="flex:1;font-size:10px;font-weight:700;color:#8b1a1a;text-align:left;padding-left:12px">{name_opp}</span>
    </div>
    {stat_rows}
  </div>

  <div style="padding:4px 16px 8px;font-size:10px;color:#bbb;border-top:0.5px solid #eceef2;background:#fafbfd">
    Ostatnie 1/3 posa\u0107 Q4 (zaokr. w g\u00f3r\u0119) \u2014 GTK: {clutch_poss_g_total} posa\u0107, Rywal: {clutch_poss_o_total} posa\u0107
  </div>
</div>"""

    def build_suma(druzyna):
        s = {"pts":0,"poss":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"br":0,"fd":0,
             "ast":0,"oreb":0,"dreb":0,"stl":0,"blk":0,"d2m":0,"d2a":0,"przerw":0}
        for row in all_stats:
            if row["druzyna"] == druzyna:
                for k in s: s[k] += row.get(k,0) or 0
        return s

    suma_gtk = build_suma("gtk")
    suma_opp = build_suma("opp")
    kpi_gtk  = calc_kpi(suma_gtk)
    kpi_opp  = calc_kpi(suma_opp)

    dt = m['data_meczu'].strftime('%d.%m.%Y') if m['data_meczu'] else ""

    # KPI cards
    def kpi_cards(suma, kpi, druzyna="gtk"):
        reb = (suma.get("oreb",0) or 0) + (suma.get("dreb",0) or 0)
        ast = suma.get("ast",0) or 0
        ks = 'background:#f4f6fb;border-radius:8px;padding:14px 12px;text-align:center'
        kv = 'font-size:22px;font-weight:500;color:#1a2b4a'
        kl = 'font-size:10px;color:#999;text-transform:uppercase;letter-spacing:.5px;margin-top:2px'
        ka = 'background:#E6F1FB;border-radius:8px;padding:14px 12px;text-align:center'
        kav = 'font-size:22px;font-weight:500;color:#0C447C'
        kal = 'font-size:10px;color:#185FA5;text-transform:uppercase;letter-spacing:.5px;margin-top:2px'
        def pct_bar(val_str):
            try: v = float(str(val_str).replace("%",""))
            except: v = 0
            w = min(int(v), 100)
            return f'<div style="flex:1;height:5px;background:#e0e0e0;border-radius:3px;overflow:hidden;margin-left:8px"><div style="width:{w}%;height:100%;background:#185FA5;border-radius:3px"></div></div>'
        row1 = f'''<div style="display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:8px;margin-bottom:8px">
  <div style="{ks}"><div style="{kv}">{suma.get("pts",0)}</div><div style="{kl}">Punkty</div></div>
  <div style="{ks}"><div style="{kv}">{ast}</div><div style="{kl}">Asysty</div></div>
  <div style="{ks}"><div style="{kv}">{reb}</div><div style="{kl}">Zbiórki</div></div>
  <div style="{ks}"><div style="{kv}">{suma.get("br",0)}</div><div style="{kl}">Straty</div></div>
</div>'''
        row2 = f'''<div style="display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:8px;margin-bottom:8px">
  <div style="{ka}"><div style="{kav}">{kpi["ortg"]}</div><div style="{kal}">ORtg</div></div>
  <div style="{ka}"><div style="{kav}">{kpi["ppp"]}</div><div style="{kal}">PPP</div></div>
  <div style="{ka}"><div style="{kav}">{kpi["efg"]}</div><div style="{kal}">eFG%</div></div>
  <div style="{ka}"><div style="{kav}">{kpi["ts"]}</div><div style="{kal}">TS%</div></div>
</div>'''
        row3 = f'''<div style="display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:8px;margin-bottom:8px">
  <div style="{ks};display:flex;align-items:center;padding:12px">
    <div><div style="{kl}">2PT%</div><div style="{kv};font-size:18px">{kpi["p2_pct"]}</div></div>
    {pct_bar(kpi["p2_pct"])}
  </div>
  <div style="{ks};display:flex;align-items:center;padding:12px">
    <div><div style="{kl}">3PT%</div><div style="{kv};font-size:18px">{kpi["p3_pct"]}</div></div>
    {pct_bar(kpi["p3_pct"])}
  </div>
  <div style="{ks};display:flex;align-items:center;padding:12px">
    <div><div style="{kl}">FT%</div><div style="{kv};font-size:18px">{kpi["ft_pct"]}</div></div>
    {pct_bar(kpi["ft_pct"])}
  </div>
</div>'''
        return row1 + row2 + row3

    # Per kwarta tabela
    def q_table(druzyna):
        opp_druzyna = "opp" if druzyna == "gtk" else "gtk"
        rows = ""
        for qn in [1,2,3,4]:
            qd  = next((r for r in all_stats if r["druzyna"]==druzyna and r["kwarta"]==qn), {})
            oqd = next((r for r in all_stats if r["druzyna"]==opp_druzyna and r["kwarta"]==qn), {})
            ast  = qd.get("ast",0)  or 0
            oreb = qd.get("oreb",0) or 0
            dreb = qd.get("dreb",0) or 0
            prz  = qd.get("przerw",0) or 0
            to_  = qd.get("br",0)   or 0
            dob  = qd.get("d2m",0)  or 0
            doba = qd.get("d2a",0)  or 0
            poss = qd.get("poss",0) or 0
            fd_  = qd.get("fd",0)   or 0
            p2m  = qd.get("p2m",0)  or 0; p2a = qd.get("p2a",0) or 0
            p3m  = qd.get("p3m",0)  or 0; p3a = qd.get("p3a",0) or 0
            ftm  = qd.get("ftm",0)  or 0; fta = qd.get("fta",0) or 0
            pts  = qd.get("pts",0)  or 0
            opp_pts  = oqd.get("pts",0)  or 0
            opp_poss = oqd.get("poss",0) or 0
            efg_v = round((p2m+1.5*p3m)/(p2a+p3a)*100) if (p2a+p3a) else None
            efg_s = ("%d%%" % efg_v) if efg_v is not None else "—"
            efg_cls = "mgood" if efg_v and efg_v>=50 else ("mbad" if efg_v and efg_v<35 else "")
            ftp_s = ("%d%%" % round(ftm/fta*100)) if fta else "—"
            p2p_s = ("%d%%" % round(p2m/p2a*100)) if p2a else "—"
            p3p_s = ("%d%%" % round(p3m/p3a*100)) if p3a else "—"
            ortg_s = ("%.1f" % (pts*100/poss)) if poss else "—"
            drtg_s = ("%.1f" % (opp_pts*100/opp_poss)) if opp_poss else "—"
            qd_cls = {1:"mq1",2:"mq2",3:"mq3",4:"mq4"}[qn]
            td = 'style="text-align:center;padding:7px 10px"'
            rows += (
                "<tr>"
                '<td style="text-align:left;padding:7px 10px"><span class="mqdot ' + qd_cls + '">' + str(qn) + 'Q</span></td>'
                '<td ' + td + ' style="text-align:center;padding:7px 10px;font-weight:500">' + str(pts) + "</td>"
                '<td ' + td + '>' + str(p2m) + "/" + str(p2a) + "</td>"
                '<td ' + td + '>' + p2p_s + "</td>"
                '<td ' + td + '>' + str(p3m) + "/" + str(p3a) + "</td>"
                '<td ' + td + '>' + p3p_s + "</td>"
                '<td ' + td + '>' + str(ftm) + "/" + str(fta) + "</td>"
                '<td ' + td + '>' + ftp_s + "</td>"
                '<td ' + td + '>' + str(oreb) + "</td>"
                '<td ' + td + '>' + str(dreb) + "</td>"
                '<td ' + td + '>' + str(ast) + "</td>"
                '<td ' + td + '>' + str(prz) + "</td>"
                '<td ' + td + '>' + str(to_) + "</td>"
                '<td ' + td + '>' + str(dob) + "</td>"
                '<td ' + td + '>' + str(doba) + "</td>"
                '<td ' + td + '>' + str(poss) + "</td>"
                '<td ' + td + '>' + str(fd_) + "</td>"
                '<td ' + td + ' class="' + efg_cls + '">' + efg_s + "</td>"
                '<td ' + td + '>' + ortg_s + "</td>"
                '<td ' + td + '>' + drtg_s + "</td>"
                "</tr>"
            )
        def _sum(field):
            return sum(r.get(field,0) or 0 for r in all_stats if r["druzyna"]==druzyna and r["kwarta"] in [1,2,3,4])
        def _osum(field):
            return sum(r.get(field,0) or 0 for r in all_stats if r["druzyna"]==opp_druzyna and r["kwarta"] in [1,2,3,4])
        sp2m=_sum("p2m");sp2a=_sum("p2a");sp3m=_sum("p3m");sp3a=_sum("p3a")
        sftm=_sum("ftm");sfta=_sum("fta");spts=_sum("pts");sposs=_sum("poss")
        sast=_sum("ast");soreb=_sum("oreb");sdreb=_sum("dreb");sprz=_sum("przerw")
        sto=_sum("br");sdob=_sum("d2m");sdoba=_sum("d2a");sfd=_sum("fd")
        sopp_pts=_osum("pts");sopp_poss=_osum("poss")
        sefg_v = round((sp2m+1.5*sp3m)/(sp2a+sp3a)*100) if (sp2a+sp3a) else None
        sefg_s = ("%d%%" % sefg_v) if sefg_v is not None else "—"
        sefg_c = "mgood" if sefg_v and sefg_v>=50 else ("mbad" if sefg_v and sefg_v<35 else "")
        sftp_s = ("%d%%" % round(sftm/sfta*100)) if sfta else "—"
        sp2p_s = ("%d%%" % round(sp2m/sp2a*100)) if sp2a else "—"
        sp3p_s = ("%d%%" % round(sp3m/sp3a*100)) if sp3a else "—"
        sortg_s = ("%.1f" % (spts*100/sposs)) if sposs else "—"
        sdrtg_s = ("%.1f" % (sopp_pts*100/sopp_poss)) if sopp_poss else "—"
        td = 'style="text-align:center;padding:7px 10px"'
        rows += (
            '<tr class="msrow">'
            '<td style="text-align:left;padding:7px 10px"><span class="mqdot mqs">&Sigma;</span></td>'
            '<td ' + td + '>' + str(spts) + "</td>"
            '<td ' + td + '>' + str(sp2m) + "/" + str(sp2a) + "</td>"
            '<td ' + td + '>' + sp2p_s + "</td>"
            '<td ' + td + '>' + str(sp3m) + "/" + str(sp3a) + "</td>"
            '<td ' + td + '>' + sp3p_s + "</td>"
            '<td ' + td + '>' + str(sftm) + "/" + str(sfta) + "</td>"
            '<td ' + td + '>' + sftp_s + "</td>"
            '<td ' + td + '>' + str(soreb) + "</td>"
            '<td ' + td + '>' + str(sdreb) + "</td>"
            '<td ' + td + '>' + str(sast) + "</td>"
            '<td ' + td + '>' + str(sprz) + "</td>"
            '<td ' + td + '>' + str(sto) + "</td>"
            '<td ' + td + '>' + str(sdob) + "</td>"
            '<td ' + td + '>' + str(sdoba) + "</td>"
            '<td ' + td + '>' + str(sposs) + "</td>"
            '<td ' + td + '>' + str(sfd) + "</td>"
            '<td ' + td + ' class="' + sefg_c + '">' + sefg_s + "</td>"
            '<td ' + td + '>' + sortg_s + "</td>"
            '<td ' + td + '>' + sdrtg_s + "</td>"
            "</tr>"
        )
        # Nagłówki — jeden spójny kolor #1a2b4a, trzy wiersze
        TH  = 'background:#1a2b4a;color:#fff;font-size:10px;font-weight:500;padding:8px 10px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.15)'
        THL = TH + ';text-align:left'
        GRP = 'background:#1a2b4a;color:rgba(255,255,255,.55);font-size:9px;font-weight:400;letter-spacing:.4px;padding:5px 10px 2px;text-align:center;white-space:nowrap;border-bottom:0.5px solid rgba(255,255,255,.12)'
        SUB = 'background:#1a2b4a;color:rgba(255,255,255,.85);font-size:10px;font-weight:500;padding:2px 10px 7px;text-align:center;white-space:nowrap;border-bottom:0.5px solid rgba(255,255,255,.2)'
        return (
            '<div class="mtw"><table style="width:100%;border-collapse:collapse;font-size:12px;table-layout:fixed">'
            '<colgroup><col style="width:32px"><col style="width:38px"><col style="width:54px"><col style="width:40px"><col style="width:54px"><col style="width:40px"><col style="width:54px"><col style="width:40px"><col style="width:38px"><col style="width:38px"><col style="width:38px"><col style="width:38px"><col style="width:36px"><col style="width:34px"><col style="width:34px"><col style="width:44px"><col style="width:36px"><col style="width:44px"><col style="width:50px"><col style="width:50px"></colgroup>'
            '<thead>'
            # Wiersz 1: główne kolumny i grupy
            '<tr>'
            '<th style="' + THL + '" rowspan="3">Q</th>'
            '<th style="' + TH  + '" rowspan="3">PKT</th>'
            '<th style="' + GRP + '" colspan="2">2PT</th>'
            '<th style="' + GRP + '" colspan="2">3PT</th>'
            '<th style="' + GRP + '" colspan="2">FT</th>'
            '<th style="' + GRP + '" colspan="2">ZB</th>'
            '<th style="' + TH  + '" rowspan="3">AST</th>'
            '<th style="' + TH  + '" rowspan="3">PRZ</th>'
            '<th style="' + TH  + '" rowspan="3">TO</th>'
            '<th style="' + GRP + '" colspan="2">DOB</th>'
            '<th style="' + TH  + '" rowspan="3">POSS</th>'
            '<th style="' + TH  + '" rowspan="3">FD</th>'
            '<th style="' + TH  + '" rowspan="3">eFG%</th>'
            '<th style="' + TH  + '" rowspan="3">ORtg</th>'
            '<th style="' + TH  + '" rowspan="3">DRtg</th>'
            '</tr>'
            # Wiersz 2: pod-nagłówki
            '<tr>'
            '<th style="' + SUB + '">M/A</th><th style="' + SUB + '">%</th>'
            '<th style="' + SUB + '">M/A</th><th style="' + SUB + '">%</th>'
            '<th style="' + SUB + '">M/A</th><th style="' + SUB + '">%</th>'
            '<th style="' + SUB + '">A</th><th style="' + SUB + '">O</th>'
            '<th style="' + SUB + '">M</th><th style="' + SUB + '">A</th>'
            '</tr>'
            '</thead><tbody>' + rows + '</tbody></table></div>'
        )

    # Czas gry z pliku Excel — musi być przed p_table()
    play_time_secs = calc_play_time(match_id)

    def p_table(druzyna):
        nr_to_name_map = {}
        try:
            cur_p = db.cursor()
            cur_p.execute("""
                SELECT ps.nr,
                       COALESCE(p.imie, r.imie, '')        AS imie,
                       COALESCE(p.nazwisko, r.nazwisko, '') AS nazwisko,
                       COALESCE(CAST(p.numer AS TEXT), CAST(ps.nr AS TEXT)) AS numer,
                       CASE WHEN p.id IS NOT NULL THEN 1
                            WHEN r.id IS NOT NULL THEN 2
                            ELSE 3 END AS prio
                FROM player_stats ps
                LEFT JOIN players p ON ps.player_id = p.id
                LEFT JOIN roster  r ON ps.roster_id  = r.id
                WHERE ps.match_id = %s AND ps.druzyna = %s
                  AND (p.id IS NOT NULL OR r.id IS NOT NULL)
                ORDER BY ps.nr, prio
            """, (match_id, druzyna))
            for row in cur_p.fetchall():
                nr_int = int(row["nr"])
                if nr_int not in nr_to_name_map:
                    pi  = row["imie"]    or ""
                    pn  = row["nazwisko"] or ""
                    pnr = str(row["numer"] or row["nr"])
                    nr_to_name_map[nr_int] = (pi, pn, pnr)
            cur_p.close()
        except Exception:
            pass

        team_poss = sum(r.get("poss",0) or 0 for r in all_stats if r["druzyna"]==druzyna)

        rows = ""
        players_d = [r for r in all_players if r["druzyna"]==druzyna]
        players_d.sort(key=lambda x: x.get("pts",0) or 0, reverse=True)
        for idx_p, pd in enumerate(players_d):
            nr = pd.get("nr","?")
            try: nr_int = int(nr)
            except: nr_int = -1

            if nr_int in nr_to_name_map:
                pi, pn, pnr = nr_to_name_map[nr_int]
                # Tylko nazwisko i inicjał imienia - bez numeru
                if pi and pn:
                    name_str = pi[:1] + ". " + pn
                elif pn:
                    name_str = pn
                else:
                    name_str = "#" + str(nr)
            else:
                name_str = "#" + str(nr)

            p2m=pd.get("p2m",0) or 0; p2a=pd.get("p2a",0) or 0
            p3m=pd.get("p3m",0) or 0; p3a=pd.get("p3a",0) or 0
            ftm=pd.get("ftm",0) or 0; fta=pd.get("fta",0) or 0
            pts=pd.get("pts",0) or 0
            ast=pd.get("ast",0) or 0
            oreb=pd.get("oreb",0) or 0; dreb=pd.get("dreb",0) or 0
            br=pd.get("br",0) or 0
            stl=pd.get("stl",0) or 0; blk=pd.get("blk",0) or 0
            fd=pd.get("fd",0) or 0
            fin=pd.get("finishes",0) or 0
            # Czas gry z pliku Excel (calc_play_time) - priorytet
            # calc_play_time już stosuje ×1.22; nr=0 jest poprawnym numerem zawodnika
            _nr_val = pd.get("nr")
            _pt_secs = play_time_secs.get(int(_nr_val), 0) if _nr_val is not None else 0
            if _pt_secs > 0:
                czas_s = f"{int(_pt_secs)//60}:{int(_pt_secs)%60:02d}"
            else:
                czas_s = "—"
            fga=p2a+p3a
            usg = round((fga+0.44*fta+br)/max(team_poss,1)*100,1) if team_poss else 0
            efg_v = round((p2m+1.5*p3m)/fga*100) if fga else None
            efg_s = ("%d%%" % efg_v) if efg_v is not None else "—"
            efg_cls = "mgood" if efg_v and efg_v>=50 else ("mbad" if efg_v and efg_v<35 else "mmuted" if efg_v is None else "")
            tsd = 2*(fga+0.44*fta)
            ts_s = ("%d%%" % round(pts/tsd*100)) if tsd else "—"
            p2p_s = ("%d%%" % round(p2m/p2a*100)) if p2a else "—"
            p3p_s = ("%d%%" % round(p3m/p3a*100)) if p3a else "—"
            ftp_s = ("%d%%" % round(ftm/fta*100)) if fta else "—"
            to_cls = " class=\"mbad\"" if br>=4 else ""
            fin_cls = " class=\"mgood\"" if fin>=10 else ""
            # Naprzemienne tło wierszy — delikatny kontrast
            row_bg = 'background:#fafbfd' if idx_p % 2 == 1 else ''
            row_style = f'style="{row_bg}"' if row_bg else ''
            td_base = f'border-bottom:0.5px solid #eceef2;padding:7px 4px;text-align:center;{row_bg}'
            td_left = f'border-bottom:0.5px solid #eceef2;padding:7px 7px;text-align:left;overflow:hidden;text-overflow:ellipsis;{row_bg}'
            rows += (
                "<tr " + row_style + ">"
                '<td style="' + td_left + '">' + name_str + "</td>"
                '<td style="' + td_base + '" data-v="' + str(int(_pt_secs)) + '">' + czas_s + "</td>"
                '<td style="' + td_base + 'font-weight:500" data-v="' + str(pts) + '">' + str(pts) + "</td>"
                '<td style="' + td_base + '" data-v="' + str(p2m*100+p2a) + '">' + str(p2m) + "/" + str(p2a) + "</td>"
                '<td style="' + td_base + '" data-v="' + (str(round(p2m/p2a*100)) if p2a else "-1") + '">' + p2p_s + "</td>"
                '<td style="' + td_base + '" data-v="' + str(p3m*100+p3a) + '">' + str(p3m) + "/" + str(p3a) + "</td>"
                '<td style="' + td_base + '" data-v="' + (str(round(p3m/p3a*100)) if p3a else "-1") + '">' + p3p_s + "</td>"
                '<td style="' + td_base + '" data-v="' + str(ftm*100+fta) + '">' + str(ftm) + "/" + str(fta) + "</td>"
                '<td style="' + td_base + '" data-v="' + (str(round(ftm/fta*100)) if fta else "-1") + '">' + ftp_s + "</td>"
                '<td style="' + td_base + '" data-v="' + str(oreb) + '">' + str(oreb) + "</td>"
                '<td style="' + td_base + '" data-v="' + str(dreb) + '">' + str(dreb) + "</td>"
                '<td style="' + td_base + '" data-v="' + str(ast) + '">' + str(ast) + "</td>"
                '<td style="' + td_base + '"' + to_cls + ' data-v="' + str(br) + '">' + str(br) + "</td>"
                '<td style="' + td_base + '" data-v="' + str(stl) + '">' + str(stl) + "</td>"
                '<td style="' + td_base + '" data-v="' + str(blk) + '">' + str(blk) + "</td>"
                '<td style="' + td_base + '" data-v="' + str(fd) + '">' + str(fd) + "</td>"
                '<td style="' + td_base + '" class="' + efg_cls + '" data-v="' + (str(efg_v) if efg_v is not None else "-1") + '">' + efg_s + "</td>"
                '<td style="' + td_base + '" data-v="' + (str(round(pts/tsd*100)) if tsd else "-1") + '">' + ts_s + "</td>"
                '<td style="' + td_base + '" data-v="' + str(usg) + '">' + str(usg) + "%</td>"
                '<td style="' + td_base + '"' + fin_cls + ' data-v="' + str(fin) + '">' + str(fin) + "</td>"
                "</tr>"
            )

        th_vm  = 'style="background:#1a2b4a;color:#fff;font-size:10px;font-weight:500;padding:6px 4px;text-align:center;border-bottom:0.5px solid rgba(255,255,255,.2);vertical-align:middle;overflow:hidden"'
        th_l   = 'style="background:#1a2b4a;color:#fff;font-size:10px;font-weight:500;padding:6px 7px;text-align:left;border-bottom:0.5px solid rgba(255,255,255,.2);vertical-align:middle"'
        th_grp = 'style="background:#1a2b4a;color:rgba(255,255,255,.6);font-size:9px;letter-spacing:.3px;padding:4px 4px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center"'
        th_sub = 'style="background:#1a2b4a;color:rgba(255,255,255,.8);font-size:10px;font-weight:500;padding:2px 4px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center"'
        th_zb  = 'style="background:#1a2b4a;color:rgba(255,255,255,.6);font-size:9px;letter-spacing:.3px;padding:4px 4px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center"'
        th_zbs = 'style="background:#1a2b4a;color:rgba(255,255,255,.8);font-size:10px;font-weight:500;padding:2px 4px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center"'

        colgroup = (
            '<colgroup>'
            '<col style="width:110px">'  
            '<col style="width:36px">'  
            '<col style="width:34px">'  
            '<col style="width:44px"><col style="width:32px">'  
            '<col style="width:44px"><col style="width:32px">'  
            '<col style="width:44px"><col style="width:32px">'  
            '<col style="width:28px"><col style="width:28px">'  
            '<col style="width:28px">'  
            '<col style="width:28px">'  
            '<col style="width:28px">'  
            '<col style="width:28px">'  
            '<col style="width:28px">'  
            '<col style="width:36px">'  
            '<col style="width:32px">'  
            '<col style="width:42px">'  
            '<col style="width:28px">'  
            '</colgroup>'
        )

        return (
            '<div class="mtw"><table style="width:100%;border-collapse:collapse;font-size:12px;table-layout:fixed">'
            + colgroup +
            '<thead>'
            '<tr>'
            '<th ' + th_l  + ' rowspan="3">Zawodnik</th>'
            '<th ' + th_vm + ' rowspan="3" style="cursor:pointer;user-select:none" onclick="sortP(this,1)" title="Czas gry (szac.)">MIN (szac.) <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_vm + ' rowspan="3" id="pth-pts" style="cursor:pointer;user-select:none" onclick="sortP(this,2)">PTS <span class=\'sort-arrow\'>▼</span></th>'
            '<th ' + th_grp + ' colspan="2">2PT</th>'
            '<th ' + th_grp + ' colspan="2">3PT</th>'
            '<th ' + th_grp + ' colspan="2">FT</th>'
            '<th ' + th_zb  + ' colspan="2">ZB</th>'
            '<th ' + th_vm + ' rowspan="3" style="cursor:pointer;user-select:none" onclick="sortP(this,11)">AST <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_vm + ' rowspan="3" style="cursor:pointer;user-select:none" onclick="sortP(this,12)">TO <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_vm + ' rowspan="3" style="cursor:pointer;user-select:none" onclick="sortP(this,13)">STL <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_vm + ' rowspan="3" style="cursor:pointer;user-select:none" onclick="sortP(this,14)">BLK <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_vm + ' rowspan="3" style="cursor:pointer;user-select:none" onclick="sortP(this,15)">FD <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_vm + ' rowspan="3" style="cursor:pointer;user-select:none" onclick="sortP(this,16)">eFG% <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_vm + ' rowspan="3" style="cursor:pointer;user-select:none" onclick="sortP(this,17)">TS% <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_vm + ' rowspan="3" style="cursor:pointer;user-select:none" onclick="sortP(this,18)">USG% <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_vm + ' rowspan="3" style="cursor:pointer;user-select:none" onclick="sortP(this,19)">FIN <span class=\'sort-arrow\'></span></th>'
            '</tr>'
            '<tr>'
            '<th ' + th_sub + ' style="cursor:pointer;user-select:none" onclick="sortP(this,3)">M/A <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_sub + ' style="cursor:pointer;user-select:none" onclick="sortP(this,4)">% <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_sub + ' style="cursor:pointer;user-select:none" onclick="sortP(this,5)">M/A <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_sub + ' style="cursor:pointer;user-select:none" onclick="sortP(this,6)">% <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_sub + ' style="cursor:pointer;user-select:none" onclick="sortP(this,7)">M/A <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_sub + ' style="cursor:pointer;user-select:none" onclick="sortP(this,8)">% <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_zbs + ' style="cursor:pointer;user-select:none" onclick="sortP(this,9)">A <span class=\'sort-arrow\'></span></th>'
            '<th ' + th_zbs + ' style="cursor:pointer;user-select:none" onclick="sortP(this,10)">O <span class=\'sort-arrow\'></span></th>'
            '</tr>'
            '</thead>'
            '<tbody>' + rows + '</tbody></table></div>'
            '<script>(function(){var _pd={};window.sortP=function(th,col){var tbl=th.closest(\'table\');if(!tbl)return;var tb=tbl.querySelector(\'tbody\');var rows=Array.from(tb.querySelectorAll(\'tr\'));var k=\'p-\'+col;_pd[k]=!_pd[k];var asc=_pd[k];rows.sort(function(a,b){var av=parseFloat(a.children[col]&&a.children[col].dataset.v);var bv=parseFloat(b.children[col]&&b.children[col].dataset.v);if(isNaN(av))av=asc?Infinity:-Infinity;if(isNaN(bv))bv=asc?Infinity:-Infinity;return asc?av-bv:bv-av;});rows.forEach(function(r,i){r.style.background=i%2===0?\'#f8f9ff\':\'#fff\';tb.appendChild(r);});tbl.querySelectorAll(\'.sort-arrow\').forEach(function(s){s.textContent=\'\';});var arrow=th.querySelector(\'.sort-arrow\');if(arrow)arrow.textContent=asc?\' ▲\':\' ▼\';};})()</script>'
        )

    def lineup_rows_html(lineups, mode="off"):
        """Generuje wiersze tabeli piątek — pełne kolumny."""
        rows = ""
        for i, lu in enumerate(sorted(lineups, key=lambda x: int(x.get("poss",0) or 0), reverse=True)):
            p2m=int(lu.get("p2m",0) or 0); p2a=int(lu.get("p2a",0) or 0)
            p3m=int(lu.get("p3m",0) or 0); p3a=int(lu.get("p3a",0) or 0)
            ftm=int(lu.get("ftm",0) or 0); fta=int(lu.get("fta",0) or 0)
            pts=int(lu.get("pts",0) or 0); poss=int(lu.get("poss",0) or 0)
            br =int(lu.get("br",0)  or 0)
            oreb=int(lu.get("oreb",0) or 0)
            _def_lu_d = def_map.get(lu["lineup"], {})
            dreb=int(_def_lu_d.get("dreb",0) or lu.get("dreb",0) or 0)
            ast =int(lu.get("ast",0)  or 0)
            stl =int(_def_lu_d.get("stl",0)  or lu.get("stl",0)  or 0)
            blk =int(_def_lu_d.get("blk",0)  or lu.get("blk",0)  or 0)
            fga = p2a + p3a
            efg_v = round((p2m+1.5*p3m)/fga*100) if fga else None
            ppp_v = pts/poss if poss else None
            efg   = f"{efg_v}%" if efg_v is not None else "—"
            ppp   = f"{ppp_v:.2f}" if ppp_v is not None else "—"
            p2pct = f"{p2m/p2a:.0%}" if p2a else "—"
            p3pct = f"{p3m/p3a:.0%}" if p3a else "—"
            ftpct = f"{ftm/fta:.0%}" if fta else "—"
            if mode == "off":
                ppp_c = "#0F6E56" if ppp_v and ppp_v>=0.9 else ("#A32D2D" if ppp_v and ppp_v<0.7 else "inherit")
                efg_c = "#0F6E56" if efg_v and efg_v>=50 else ("#A32D2D" if efg_v and efg_v<35 else "inherit")
            else:
                ppp_c = "#0F6E56" if ppp_v and ppp_v<0.7 else ("#A32D2D" if ppp_v and ppp_v>=0.9 else "inherit")
                efg_c = "#0F6E56" if efg_v and efg_v<35 else ("#A32D2D" if efg_v and efg_v>=50 else "inherit")
            bg = "#f8f9ff" if i%2==0 else "#fff"
            skladniki = " · ".join(nr_name_map.get(n, f"#{n}") for n in lu["lineup"].split("-"))
            efg_s = str(efg_v) if efg_v is not None else "-1"
            ppp_s = f"{ppp_v*100:.1f}" if ppp_v is not None else "-1"
            p2_s  = f"{p2m/p2a*100:.0f}" if p2a else "-1"
            p3_s  = f"{p3m/p3a*100:.0f}" if p3a else "-1"
            ft_s  = f"{ftm/fta*100:.0f}" if fta else "-1"
            rows += f"""<tr style="background:{bg}">
                <td style="font-size:.78rem;text-align:left;padding-left:8px">{skladniki}</td>
                <td class="text-center" data-v="{poss}">{poss}</td>
                <td class="text-center fw-bold" style="color:#1a2b4a" data-v="{pts}">{pts}</td>
                <td class="text-center" data-v="{p2m*100+p2a}">{p2m}/{p2a}</td>
                <td class="text-center" data-v="{p2_s}">{p2pct}</td>
                <td class="text-center" data-v="{p3m*100+p3a}">{p3m}/{p3a}</td>
                <td class="text-center" data-v="{p3_s}">{p3pct}</td>
                <td class="text-center" data-v="{ftm*100+fta}">{ftm}/{fta}</td>
                <td class="text-center" data-v="{ft_s}">{ftpct}</td>
                <td class="text-center" data-v="{oreb}">{oreb}</td>
                <td class="text-center" data-v="{dreb}">{dreb}</td>
                <td class="text-center" data-v="{ast}">{ast}</td>
                <td class="text-center" data-v="{br}" style="color:{'#A32D2D' if br>=4 else 'inherit'}">{br}</td>
                <td class="text-center" data-v="{stl}">{stl}</td>
                <td class="text-center" data-v="{blk}">{blk}</td>
                <td class="text-center" data-v="{efg_s}" style="color:{efg_c}">{efg}</td>
                <td class="text-center fw-bold" data-v="{ppp_s}" style="color:{ppp_c}">{ppp}</td>
            </tr>"""
        return rows

    def lineup_table():
        if not all_lineups and not all_lineups_def:
            return '<p class="text-muted p-3 mb-0" style="font-size:.82rem">Brak danych piątek — wgraj mecz ponownie aby wygenerować.</p>'

        # Zbierz dane DEF i NET per lineup key
        def_map  = {lu["lineup"]: lu for lu in all_lineups_def}
        off_rtg  = {lu["lineup"]: lu["pts"]*100/lu["poss"] for lu in all_lineups if int(lu.get("poss",0) or 0)>0}
        def_rtg  = {lu["lineup"]: lu["pts"]*100/lu["poss"] for lu in all_lineups_def if int(lu.get("poss",0) or 0)>0}

        rows = ""
        for i, lu in enumerate(sorted(all_lineups, key=lambda x: int(x.get("poss",0) or 0), reverse=True)):
            p2m=int(lu.get("p2m",0) or 0); p2a=int(lu.get("p2a",0) or 0)
            p3m=int(lu.get("p3m",0) or 0); p3a=int(lu.get("p3a",0) or 0)
            ftm=int(lu.get("ftm",0) or 0); fta=int(lu.get("fta",0) or 0)
            pts=int(lu.get("pts",0) or 0); poss=int(lu.get("poss",0) or 0)
            br =int(lu.get("br",0)  or 0)
            oreb=int(lu.get("oreb",0) or 0); dreb=int(lu.get("dreb",0) or 0)
            ast =int(lu.get("ast",0)  or 0)
            stl =int(lu.get("stl",0)  or 0)
            blk =int(lu.get("blk",0)  or 0)
            fd_ =int(lu.get("fd",0)   or 0)
            fga = p2a + p3a
            efg_v  = round((p2m+1.5*p3m)/fga*100) if fga else None
            ppp_v  = pts/poss if poss else None
            efg_s  = f"{efg_v}%" if efg_v is not None else "—"
            ppp_s  = f"{ppp_v:.2f}" if ppp_v is not None else "—"
            p2pct  = f"{p2m/p2a:.0%}" if p2a else "—"
            p3pct  = f"{p3m/p3a:.0%}" if p3a else "—"
            ftpct  = f"{ftm/fta:.0%}" if fta else "—"
            ppp_c  = "#0F6E56" if ppp_v and ppp_v>=0.9 else ("#A32D2D" if ppp_v and ppp_v<0.7 else "inherit")
            efg_c  = "#0F6E56" if efg_v and efg_v>=50 else ("#A32D2D" if efg_v and efg_v<35 else "inherit")

            # DEF stats dla tej piątki
            dlu = def_map.get(lu["lineup"], {})
            dp2m=int(dlu.get("p2m",0) or 0); dp2a=int(dlu.get("p2a",0) or 0)
            dp3m=int(dlu.get("p3m",0) or 0); dp3a=int(dlu.get("p3a",0) or 0)
            dp_fga = dp2a+dp3a
            defg_v = round((dp2m+1.5*dp3m)/dp_fga*100) if dp_fga else None
            dposs  = int(dlu.get("poss",0) or 0)
            dpts   = int(dlu.get("pts",0) or 0)
            dppp_v = dpts/dposs if dposs else None
            defg_s = f"{defg_v}%" if defg_v is not None else "—"
            dppp_s = f"{dppp_v:.2f}" if dppp_v is not None else "—"
            defg_c = "#0F6E56" if defg_v and defg_v<35 else ("#A32D2D" if defg_v and defg_v>=50 else "inherit")
            dppp_c = "#0F6E56" if dppp_v and dppp_v<0.7 else ("#A32D2D" if dppp_v and dppp_v>=0.9 else "inherit")

            # NET RTG
            k = lu["lineup"]
            ortg_v = off_rtg.get(k)
            drtg_v = def_rtg.get(k)
            net_v  = round(ortg_v - drtg_v, 1) if (ortg_v is not None and drtg_v is not None) else None
            ortg_s = f"{ortg_v:.1f}" if ortg_v is not None else "—"
            drtg_s = f"{drtg_v:.1f}" if drtg_v is not None else "—"
            net_s  = f"{net_v:+.1f}" if net_v is not None else "—"
            net_c  = "#0F6E56" if net_v and net_v>0 else ("#A32D2D" if net_v and net_v<0 else "#888")
            ortg_c = "#0F6E56" if ortg_v and ortg_v>=90 else ("#A32D2D" if ortg_v and ortg_v<70 else "inherit")
            drtg_c = "#0F6E56" if drtg_v and drtg_v<70 else ("#A32D2D" if drtg_v and drtg_v>=90 else "inherit")

            bg = "#f8f9ff" if i%2==0 else "#fff"
            br_c = "#A32D2D" if br>=4 else "inherit"
            efg_sv = str(efg_v) if efg_v is not None else "-1"
            ppp_sv = f"{ppp_v*100:.1f}" if ppp_v is not None else "-1"
            defg_sv = str(defg_v) if defg_v is not None else "-1"
            dppp_sv = f"{dppp_v*100:.1f}" if dppp_v is not None else "-1"
            net_sv = f"{net_v:.1f}" if net_v is not None else "-999"
            skladniki = " · ".join(nr_name_map.get(n, f"#{n}") for n in lu["lineup"].split("-"))

            # MIN (szac.) — średni czas gry składu = suma czasów zawodników / 5
            _lu_nrs = [int(n) for n in lu["lineup"].split("-") if n.isdigit()]
            _lu_secs = [play_time_secs.get(n, 0) for n in _lu_nrs]
            _lu_secs_valid = [s for s in _lu_secs if s > 0]
            if _lu_secs_valid:
                _avg_secs = sum(_lu_secs_valid) / len(_lu_nrs)
                _min_str  = f"{int(_avg_secs)//60}:{int(_avg_secs)%60:02d}"
                _min_sv   = f"{_avg_secs:.1f}"
            else:
                _min_str = "—"
                _min_sv  = "-1"

            rows += f"""<tr style="background:{bg}">
                <td style="font-size:.78rem;text-align:left;padding-left:8px">{skladniki}</td>
                <td class="text-center" data-v="{poss}">{poss}</td>
                <td class="text-center" style="color:#633806;font-weight:500" data-v="{_min_sv}">{_min_str}</td>
                <td class="text-center fw-bold" style="color:#1a2b4a" data-v="{pts}">{pts}</td>
                <td class="text-center" data-v="{p2m*100+p2a}">{p2m}/{p2a}</td>
                <td class="text-center" data-v="{int(p2m/p2a*100) if p2a else -1}">{p2pct}</td>
                <td class="text-center" data-v="{p3m*100+p3a}">{p3m}/{p3a}</td>
                <td class="text-center" data-v="{int(p3m/p3a*100) if p3a else -1}">{p3pct}</td>
                <td class="text-center" data-v="{ftm*100+fta}">{ftm}/{fta}</td>
                <td class="text-center" data-v="{int(ftm/fta*100) if fta else -1}">{ftpct}</td>
                <td class="text-center" data-v="{oreb}">{oreb}</td>
                <td class="text-center" data-v="{dreb}">{dreb}</td>
                <td class="text-center" data-v="{ast}">{ast}</td>
                <td class="text-center" data-v="{br}" style="color:{br_c}">{br}</td>
                <td class="text-center" data-v="{stl}">{stl}</td>
                <td class="text-center" data-v="{blk}">{blk}</td>
                <td class="text-center" data-v="{fd_}">{fd_}</td>
                <td class="text-center" data-v="{efg_sv}" style="color:{efg_c}">{efg_s}</td>
                <td class="text-center fw-bold" data-v="{ppp_sv}" style="color:{ppp_c}">{ppp_s}</td>
                <td class="text-center" data-v="{defg_sv}" style="color:{defg_c}">{defg_s}</td>
                <td class="text-center fw-bold" data-v="{dppp_sv}" style="color:{dppp_c}">{dppp_s}</td>
                <td class="text-center" data-v="{ortg_v or -999}" style="color:{ortg_c}">{ortg_s}</td>
                <td class="text-center" data-v="{drtg_v or 999}" style="color:{drtg_c}">{drtg_s}</td>
                <td class="text-center fw-bold" data-v="{net_sv}" style="color:{net_c}">{net_s}</td>
            </tr>"""

        no_data = '<tr><td colspan="22" class="text-muted text-center p-3" style="font-size:.82rem">Brak danych — wgraj mecz ponownie.</td></tr>'

        # Style nagłówków
        th  = 'background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 3px;text-align:center;white-space:nowrap;cursor:pointer;user-select:none'
        thl = 'background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 8px;text-align:left;white-space:nowrap'
        thg = 'background:#1a2b4a;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center'
        ths = 'background:#1a2b4a;color:rgba(255,255,255,.8);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center;cursor:pointer;user-select:none'
        thz = 'background:#152236;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.1);text-align:center'
        thzs= 'background:#152236;color:rgba(255,255,255,.75);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center;cursor:pointer;user-select:none'
        thn = 'background:#412402;color:#FAC775;font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center'
        thns= 'background:#412402;color:#FAC775;font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center;cursor:pointer;user-select:none'
        vm  = 'vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2)'

        cg = ('<colgroup>'
            '<col style="width:190px">'
            '<col style="width:36px"><col style="width:46px"><col style="width:34px">'
            '<col style="width:40px"><col style="width:30px">'
            '<col style="width:40px"><col style="width:30px">'
            '<col style="width:40px"><col style="width:30px">'
            '<col style="width:26px"><col style="width:26px">'
            '<col style="width:28px"><col style="width:28px"><col style="width:28px"><col style="width:28px"><col style="width:28px">'
            '<col style="width:34px"><col style="width:34px">'
            '<col style="width:34px"><col style="width:34px">'
            '<col style="width:38px"><col style="width:38px"><col style="width:42px">'
            '</colgroup>')

        hdr = (f'<thead>'
            f'<tr>'
            f'<th style="{thl};{vm}" rowspan="3">Skład</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="luSort(this,1)">POSS ↕</th>'
            f'<th style="{th};{vm};cursor:pointer;user-select:none" rowspan="3" onclick="luSort(this,2)" title="Szac. czas gry składu (suma MIN zawodników ÷ 5)">MIN<br>(szac.)</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="luSort(this,4)">PKT ↕</th>'
            f'<th style="{thg}" colspan="2">2PT</th>'
            f'<th style="{thg}" colspan="2">3PT</th>'
            f'<th style="{thg}" colspan="2">FT</th>'
            f'<th style="{thz}" colspan="2">ZB</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="luSort(this,12)">AST ↕</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="luSort(this,13)">TO ↕</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="luSort(this,14)">STL ↕</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="luSort(this,15)">BLK ↕</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="luSort(this,16)">FD ↕</th>'
            f'<th style="{thg}" colspan="2">OFF</th>'
            f'<th style="{thg}" colspan="2">DEF</th>'
            f'<th style="{thn}" colspan="3">NET RTG</th>'
            f'</tr>'
            f'<tr>'
            f'<th style="{ths}" onclick="luSort(this,4)">M/A ↕</th><th style="{ths}" onclick="luSort(this,5)">% ↕</th>'
            f'<th style="{ths}" onclick="luSort(this,6)">M/A ↕</th><th style="{ths}" onclick="luSort(this,7)">% ↕</th>'
            f'<th style="{ths}" onclick="luSort(this,8)">M/A ↕</th><th style="{ths}" onclick="luSort(this,9)">% ↕</th>'
            f'<th style="{thzs}" onclick="luSort(this,10)">A ↕</th><th style="{thzs}" onclick="luSort(this,11)">O ↕</th>'
            f'<th style="{ths}" onclick="luSort(this,17)">eFG% ↕</th><th style="{ths}" onclick="luSort(this,18)">PPP ↕</th>'
            f'<th style="{ths}" onclick="luSort(this,19)">eFG% ↕</th><th style="{ths}" onclick="luSort(this,20)">PPP ↕</th>'
            f'<th style="{thns}" onclick="luSort(this,21)">ORtg ↕</th><th style="{thns}" onclick="luSort(this,22)">DRtg ↕</th><th style="{thns}" onclick="luSort(this,23)">Net ↕</th>'
            f'</tr>'
            f'</thead>')

        js = """<script>
(function(){
  var _d={};
  window.luSort=function(th,col){
    var tbl=document.getElementById('tbl-lu-off'); if(!tbl) return;
    var tb=tbl.querySelector('tbody');
    var rows=Array.from(tb.querySelectorAll('tr'));
    var k='lu-'+col; _d[k]=!_d[k]; var asc=_d[k];
    rows.sort(function(a,b){
      var av=parseFloat(a.children[col]&&a.children[col].dataset.v);
      var bv=parseFloat(b.children[col]&&b.children[col].dataset.v);
      if(isNaN(av))av=asc?Infinity:-Infinity;
      if(isNaN(bv))bv=asc?Infinity:-Infinity;
      return asc?av-bv:bv-av;
    });
    rows.forEach(function(r,i){r.style.background=i%2===0?'#f8f9ff':'#fff';tb.appendChild(r);});
    tbl.querySelectorAll('th').forEach(function(h){h.innerHTML=h.innerHTML.replace(/ [▲▼]/,' ↕');});
    th.innerHTML=th.innerHTML.replace(' ↕',asc?' ▲':' ▼');
  };
})();
</script>"""

        sub = 'font-size:8px;color:#888;margin-bottom:6px'
        return (
            f'<div style="{sub}">PPP OFF: <span style="color:#0F6E56">≥0.90 dobry</span> / <span style="color:#A32D2D">&lt;0.70 słaby</span> &nbsp;·&nbsp; '
            f'PPP DEF: <span style="color:#0F6E56">&lt;0.70 dobry</span> / <span style="color:#A32D2D">≥0.90 słaby</span> &nbsp;·&nbsp; sortowanie: POSS malejąco</div>'
            f'<div class="table-responsive"><table id="tbl-lu-off" class="table table-hover mb-0" style="table-layout:fixed;min-width:900px">'
            f'{cg}{hdr}'
            f'<tbody>{rows or no_data}</tbody></table></div>'
            + js
        )


    # Przebieg meczu
    def flow_chart():
        if not flow_rows:
            return '<p class="text-muted p-3 mb-0" style="font-size:.82rem">Brak danych przebiegu — wgraj mecz ponownie aby wygenerować.</p>'

        # Konwertuj do JSON dla Chart.js
        # Dodaj punkt startowy (0,0,0) i końcowy
        pts_final_gtk = m["wynik_gtk"] or 0
        pts_final_opp = m["wynik_opp"] or 0

        # Punkty na osi X: globalny czas (kwarta*600 + (600-czas_sek))
        # W koszykówce czas biegnie od 10:00 do 0:00 per kwartę
        labels = []
        diff_data = []
        gtk_data = []
        opp_data = []

        # Punkt startowy
        labels.append(0)
        diff_data.append(0)
        gtk_data.append(0)
        opp_data.append(0)

        for row in flow_rows:
            q = row["kwarta"]
            t = row["czas_sek"] or 0
            # Globalny czas: kwarta 1 = 0-600s, kwarta 2 = 600-1200s itd.
            # czas_sek to ile sekund pozostało w kwarcie → upłynęło = 600-t
            elapsed = (q - 1) * 600 + max(0, 600 - int(t))
            labels.append(elapsed)
            g = row["pts_gtk"]
            o = row["pts_opp"]
            diff_data.append(g - o)
            gtk_data.append(g)
            opp_data.append(o)

        # Punkt końcowy
        labels.append(2400)
        diff_data.append(pts_final_gtk - pts_final_opp)
        gtk_data.append(pts_final_gtk)
        opp_data.append(pts_final_opp)

        # Wykryj runy (serie ≥5 pkt bez odpowiedzi)
        runs_gtk = []
        runs_opp = []
        streak_g = streak_o = 0
        streak_g_start = streak_o_start = 0
        prev_g = prev_o = 0
        for i, (g, o) in enumerate(zip(gtk_data, opp_data)):
            dg = g - prev_g
            do_ = o - prev_o
            if dg > 0 and do_ == 0:
                if streak_g == 0: streak_g_start = labels[i]
                streak_g += dg; streak_o = 0
            elif do_ > 0 and dg == 0:
                if streak_o == 0: streak_o_start = labels[i]
                streak_o += do_; streak_g = 0
            else:
                if streak_g >= 5: runs_gtk.append((streak_g_start, labels[i], streak_g))
                if streak_o >= 5: runs_opp.append((streak_o_start, labels[i], streak_o))
                streak_g = streak_o = 0
            prev_g, prev_o = g, o

        # Przygotuj annotations dla runs
        annotations_js = ""
        for s, e, pts in runs_gtk[:3]:
            mid = (s + e) // 2
            annotations_js += f"""
            'run_g_{s}': {{type:'box',xMin:{s},xMax:{e},yMin:-2,yMax:2,backgroundColor:'rgba(26,107,60,0.08)',borderColor:'rgba(26,107,60,0.3)',borderWidth:1,label:{{display:true,content:'GTK {pts}-0',font:{{size:9}},color:'#1a6b3c',position:'start'}}}},"""
        for s, e, pts in runs_opp[:3]:
            annotations_js += f"""
            'run_o_{s}': {{type:'box',xMin:{s},xMax:{e},yMin:-2,yMax:2,backgroundColor:'rgba(139,26,26,0.08)',borderColor:'rgba(139,26,26,0.3)',borderWidth:1,label:{{display:true,content:'OPP {pts}-0',font:{{size:9}},color:'#8b1a1a',position:'start'}}}},"""

        import json
        labels_js = json.dumps(labels)
        diff_js = json.dumps(diff_data)
        gtk_js = json.dumps(gtk_data)
        opp_js = json.dumps(opp_data)
        max_diff = max(abs(d) for d in diff_data) if diff_data else 10
        y_max = max(max_diff + 5, 10)

        return f"""
        <div style="font-size:.72rem;color:#aaa;margin-bottom:.5rem">
          Różnica punktowa w czasie meczu ·
          <span style="color:#1a6b3c;font-weight:600">{gtk_name}</span> vs
          <span style="color:#8b1a1a;font-weight:600">{name_opp}</span>
        </div>
        <div style="position:relative;height:200px">
          <canvas id="flowChart"></canvas>
        </div>
        <div style="display:flex;gap:16px;margin-top:8px;font-size:.75rem;flex-wrap:wrap">
          <div style="font-weight:600;color:var(--color-text-secondary)">Najdłuższe runy:</div>
          {''.join(f'<span style="background:#e8f5e9;color:#1a5c2a;padding:2px 8px;border-radius:12px">{p}-0 GTK</span>' for _,__,p in sorted(runs_gtk,key=lambda x:-x[2])[:3])}
          {''.join(f'<span style="background:#ffebee;color:#8b1a1a;padding:2px 8px;border-radius:12px">{p}-0 OPP</span>' for _,__,p in sorted(runs_opp,key=lambda x:-x[2])[:3])}
        </div>
        {momentum_table()}
        {clutch_stats()}
        <script>
        (function(){{
          var _flowInited = false;

          function initFlowChart() {{
            if (_flowInited) return;
            var canvas = document.getElementById('flowChart');
            if (!canvas || canvas.offsetWidth === 0) return;
            _flowInited = true;

            var ctx = canvas.getContext('2d');
            var labels  = {labels_js};
            var diffData = {diff_js};
            var gtkData  = {gtk_js};
            var oppData  = {opp_js};

            var gradGreen = ctx.createLinearGradient(0,0,0,200);
            gradGreen.addColorStop(0,'rgba(26,107,60,0.18)');
            gradGreen.addColorStop(1,'rgba(26,107,60,0)');
            var gradRed = ctx.createLinearGradient(0,0,0,200);
            gradRed.addColorStop(0,'rgba(139,26,26,0)');
            gradRed.addColorStop(1,'rgba(139,26,26,0.18)');

            new Chart(ctx, {{
              type: 'line',
              data: {{
                labels: labels,
                datasets: [{{
                  label: 'Różnica',
                  data: diffData,
                  borderColor: '#1a2b4a',
                  borderWidth: 2,
                  pointRadius: 0,
                  pointHoverRadius: 4,
                  fill: {{
                    target: {{value: 0}},
                    above: gradGreen,
                    below: gradRed
                  }},
                  tension: 0.3
                }}]
              }},
              options: {{
                responsive: true,
                maintainAspectRatio: false,
                interaction: {{mode:'index', intersect:false}},
                plugins: {{
                  legend: {{display: false}},
                  tooltip: {{
                    callbacks: {{
                      title: function(items) {{
                        var s = items[0].parsed.x;
                        var q = Math.floor(s/600)+1;
                        var min = Math.floor((s%600)/60);
                        var sec = (s%600)%60;
                        return q+'Q ' + min + ':' + String(sec).padStart(2,'0');
                      }},
                      label: function(item) {{
                        var idx = item.dataIndex;
                        var d = item.parsed.y;
                        return 'Różnica: '+(d>0?'+':'')+d+'  ('+gtkData[idx]+':'+oppData[idx]+')';
                      }}
                    }}
                  }}
                }},
                scales: {{
                  x: {{
                    type: 'linear',
                    min: 0, max: 2400,
                    ticks: {{
                      stepSize: 600,
                      callback: function(v) {{ return ['','1Q','2Q','3Q','4Q'][v/600]||''; }},
                      font: {{size:10}}
                    }},
                    grid: {{color:'rgba(0,0,0,0.06)'}}
                  }},
                  y: {{
                    min: -{y_max}, max: {y_max},
                    ticks: {{font:{{size:10}}}},
                    grid: {{
                      color: function(c) {{
                        return c.tick.value===0 ? 'rgba(0,0,0,0.25)' : 'rgba(0,0,0,0.06)';
                      }}
                    }}
                  }}
                }}
              }}
            }});
          }}  // koniec initFlowChart

          // Odpal gdy zakładka Przebieg staje się widoczna
          var tabBtn = document.querySelector('[data-bs-target="#tabFLOW"]');
          if (tabBtn) {{
            tabBtn.addEventListener('shown.bs.tab', function() {{
              initFlowChart();
            }});
          }}
          // Fallback: jeśli zakładka jest już aktywna przy załadowaniu
          setTimeout(initFlowChart, 200);
        }})();
        </script>"""

    # Clutch stats — ostatnie 5 min Q4 przy różnicy ≤5
    def momentum_table():
        if not all_stats:
            return '<p class="text-muted p-3 mb-0" style="font-size:.82rem">Brak danych.</p>'

        rows_html = ""
        q_results = []
        for qn in [1, 2, 3, 4]:
            # Pobierz punkty bezpośrednio z match_stats per kwarta
            qs_gtk = next((r for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==qn), {})
            qs_opp = next((r for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==qn), {})
            q_gtk = qs_gtk.get("pts", 0) or 0
            q_opp = qs_opp.get("pts", 0) or 0
            q_winner = "gtk" if q_gtk > q_opp else ("opp" if q_opp > q_gtk else "tie")
            q_results.append(q_winner)

            def _efg(druzyna, qn):
                qs = next((r for r in all_stats if r["druzyna"]==druzyna and r["kwarta"]==qn), {})
                p2a=qs.get("p2a",0) or 0; p3a=qs.get("p3a",0) or 0
                p2m=qs.get("p2m",0) or 0; p3m=qs.get("p3m",0) or 0
                return str(round((p2m+1.5*p3m)/(p2a+p3a)*100)) + "%" if (p2a+p3a) else "—"
            def _to(druzyna, qn):
                return next((r for r in all_stats if r["druzyna"]==druzyna and r["kwarta"]==qn), {}).get("br",0) or 0
            def _poss(druzyna, qn):
                return next((r for r in all_stats if r["druzyna"]==druzyna and r["kwarta"]==qn), {}).get("poss",0) or 0

            qd_cls = {1:"mq1",2:"mq2",3:"mq3",4:"mq4"}[qn]
            if q_winner=="gtk":
                res_html = '<span class="win">W +%d</span>' % (q_gtk-q_opp)
            elif q_winner=="opp":
                res_html = '<span class="loss">L -%d</span>' % (q_opp-q_gtk)
            else:
                res_html = '<span style="background:#eee;color:#666;border-radius:6px;padding:2px 8px;font-size:11px">=</span>'

            efg_g = _efg("gtk",qn); efg_o = _efg("opp",qn)
            to_g  = _to("gtk",qn);  to_o  = _to("opp",qn)
            ps_g  = _poss("gtk",qn); ps_o = _poss("opp",qn)

            def efg_int(s):
                try: return int(s.rstrip("%"))
                except: return -1

            g_efg = "good" if efg_g!="—" and efg_o!="—" and efg_int(efg_g)>efg_int(efg_o) else ("bad" if efg_g!="—" and efg_o!="—" and efg_int(efg_g)<efg_int(efg_o) else "")
            o_efg = "good" if efg_g!="—" and efg_o!="—" and efg_int(efg_o)>efg_int(efg_g) else ("bad" if efg_g!="—" and efg_o!="—" and efg_int(efg_o)<efg_int(efg_g) else "")
            g_to  = "good" if to_g < to_o else ("bad" if to_g > to_o else "")
            o_to  = "good" if to_o < to_g else ("bad" if to_o > to_g else "")
            g_pkt = "good" if q_gtk > q_opp else ("bad" if q_gtk < q_opp else "")
            o_pkt = "good" if q_opp > q_gtk else ("bad" if q_opp < q_gtk else "")

            rows_html += (
                '<tr>'
                '<td><span class="mqdot ' + qd_cls + '">' + str(qn) + 'Q</span></td>'
                '<td style="text-align:center">' + res_html + '</td>'
                '<td style="text-align:center" class="' + g_pkt + '">' + str(q_gtk) + '</td>'
                '<td style="text-align:center" class="' + g_efg + '">' + efg_g + '</td>'
                '<td style="text-align:center" class="' + g_to + '">' + str(to_g) + '</td>'
                '<td style="text-align:center">' + str(ps_g) + '</td>'
                '<td style="text-align:center;border-left:2px solid #e0e0e0" class="' + o_pkt + '">' + str(q_opp) + '</td>'
                '<td style="text-align:center" class="' + o_efg + '">' + efg_o + '</td>'
                '<td style="text-align:center" class="' + o_to + '">' + str(to_o) + '</td>'
                '<td style="text-align:center">' + str(ps_o) + '</td>'
                '</tr>'
            )

        streak_html = ""
        for r in q_results:
            col = "#0F6E56" if r=="gtk" else ("#A32D2D" if r=="opp" else "#888")
            lbl = "W" if r=="gtk" else ("L" if r=="opp" else "=")
            streak_html += ('<span style="display:inline-flex;align-items:center;justify-content:center;'
                           'width:28px;height:28px;border-radius:50%;background:' + col + ';color:#fff;'
                           'font-size:11px;font-weight:500;margin-right:4px">' + lbl + '</span>')
        best = cur_s = 0
        for r in q_results:
            cur_s = cur_s+1 if r=="gtk" else 0
            best = max(best, cur_s)
        best_html = ('<span style="font-size:12px;color:#666">Najlepsza seria: '
                    '<span style="color:#0F6E56;font-weight:500">' + str(best) +
                    (' kwarty' if best>1 else ' kwarta') + '</span></span>') if best>=1 else ""

        gtk_s = gtk_name.upper()[:16]
        opp_s = name_opp.upper()[:14]
        th_qw  = "padding:8px 12px;font-size:10px;font-weight:500;text-transform:uppercase;letter-spacing:.5px;color:#888;border-bottom:0.5px solid #e0e0e0;background:#fff;white-space:nowrap"
        th_grp = "padding:8px 12px 4px;font-size:10px;font-weight:500;letter-spacing:.4px;text-align:center;background:#fff;border-bottom:none;white-space:nowrap"
        th_sub = "padding:4px 12px 8px;font-size:10px;font-weight:500;text-align:center;border-bottom:0.5px solid #e0e0e0;background:#fff;color:#888;white-space:nowrap"

        return (
            '<div style="padding:12px 16px;display:flex;align-items:center;gap:8px;flex-wrap:wrap;border-bottom:0.5px solid #e0e0e0">'
            + streak_html + best_html +
            '</div>'
            '<div style="overflow-x:auto">'
            '<table style="width:100%;border-collapse:collapse;font-size:12px;min-width:540px">'
            '<thead>'
            '<tr>'
            '<th rowspan="2" style="' + th_qw + ';text-align:left;vertical-align:middle">Q</th>'
            '<th rowspan="2" style="' + th_qw + ';text-align:center;vertical-align:middle">Wynik</th>'
            '<th colspan="4" style="' + th_grp + ';color:#185FA5">' + gtk_s + '</th>'
            '<th colspan="4" style="' + th_grp + ';color:#A32D2D;border-left:2px solid #e0e0e0">' + opp_s + '</th>'
            '</tr>'
            '<tr>'
            '<th style="' + th_sub + '">PKT</th>'
            '<th style="' + th_sub + '">eFG%</th>'
            '<th style="' + th_sub + '">TO</th>'
            '<th style="' + th_sub + '">POSS</th>'
            '<th style="' + th_sub + ';border-left:2px solid #e0e0e0">PKT</th>'
            '<th style="' + th_sub + '">eFG%</th>'
            '<th style="' + th_sub + '">TO</th>'
            '<th style="' + th_sub + '">POSS</th>'
            '</tr>'
            '</thead>'
            '<tbody>' + rows_html + '</tbody>'
            '</table></div>'
        )

    def tim_table(druzyna):
        qcolors = {1:"mq1",2:"mq2",3:"mq3",4:"mq4"}

        def get_td(q, b):
            return next((r for r in all_timing
                         if r["druzyna"]==druzyna and r["bucket"]==b
                         and (r.get("kwarta") or 0)==q), {})

        def calc_efg(m2, a2, m3, a3, ftm_v):
            att = a2 + a3
            if att == 0 and ftm_v > 0:
                return 100
            if att == 0:
                return None
            base = round((m2 + 1.5*m3) / att * 100)
            if ftm_v > 0:
                return max(base, round((m2 + 1.5*m3 + 1) / (att + 1) * 100))
            return base

        def calc_udane(m2, m3, ftm_v):
            # Akcja udana = trafiony 2PT lub 3PT lub ≥1 celny rzut wolny
            return m2 + m3 + (1 if ftm_v > 0 else 0)

        def calc_nieudane(m2, a2, m3, a3, br_v):
            # Akcja nieudana = chybiony 2PT lub 3PT lub strata
            return (a2 - m2) + (a3 - m3) + br_v

        def pill(efg):
            if efg is None:
                return '<span style="color:#aaa">&#8212;</span>'
            c = "pill-good" if efg >= 50 else ("pill-bad" if efg < 35 else "pill-neu")
            return '<span class="epill %s">%d%%</span>' % (c, efg)

        # Sumy per bucket
        sum_rows = []
        for b in BUCKETS:
            m2=a2=m3=a3=br_v=ftm_v=poss_ft_v=0
            for q in [1,2,3,4]:
                td = get_td(q, b)
                m2+=td.get("made2",0); a2+=td.get("att2",0)
                m3+=td.get("made3",0); a3+=td.get("att3",0)
                br_v+=td.get("br",0);  ftm_v+=td.get("ftm",0)
                poss_ft_v+=td.get("poss_ft",0)
            td0 = get_td(0, b)
            if td0 and not any(get_td(q,b) for q in [1,2,3,4]):
                m2=td0.get("made2",0); a2=td0.get("att2",0)
                m3=td0.get("made3",0); a3=td0.get("att3",0)
                br_v=td0.get("br",0);  ftm_v=td0.get("ftm",0)
                poss_ft_v=td0.get("poss_ft",0)
            # fallback dla starych meczów: poss_ft nie było zapisane w DB
            eff_poss_ft = poss_ft_v if poss_ft_v > 0 else (1 if ftm_v > 0 else 0)
            udane    = calc_udane(m2, m3, eff_poss_ft)
            nieudane = calc_nieudane(m2, a2, m3, a3, br_v)
            total    = udane + nieudane
            efg      = calc_efg(m2, a2, m3, a3, ftm_v)
            sum_rows.append((b, m2, a2, m3, a3, eff_poss_ft, br_v, udane, nieudane, total, efg))

        th  = 'style="padding:9px 14px;font-size:10px;font-weight:500;text-align:right;background:#1a2b4a;color:#fff;white-space:nowrap;text-transform:uppercase;letter-spacing:.4px"'
        thl = 'style="padding:9px 14px;font-size:10px;font-weight:500;text-align:left;background:#1a2b4a;color:#fff;white-space:nowrap;text-transform:uppercase;letter-spacing:.4px;width:80px"'
        tdr = 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:right;font-size:12px"'
        tdl = 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:left;font-size:12px;font-weight:500"'
        tdu = 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:right;font-size:12px;font-weight:500;color:#085041"'
        tdn = 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:right;font-size:12px;font-weight:500;color:#A32D2D"'

        tdc = 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:center;font-size:12px"'
        tdc_b = 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:center;font-size:12px;font-weight:700"'
        tdlc = 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:center;font-size:12px;font-weight:500"'

        s_rows = ""
        for b,m2,a2,m3,a3,ftm_v,br_v,udane,nieudane,total,efg_v in sum_rows:
            chyb2 = a2 - m2
            chyb3 = a3 - m3
            p2 = str(m2) if m2 > 0 else '<span style="color:#aaa">&#8212;</span>'
            p3 = str(m3) if m3 > 0 else '<span style="color:#aaa">&#8212;</span>'
            ft_s  = str(ftm_v) if ftm_v > 0 else '<span style="color:#aaa">&#8212;</span>'
            c2_s  = str(chyb2) if chyb2 > 0 else '<span style="color:#aaa">&#8212;</span>'
            c3_s  = str(chyb3) if chyb3 > 0 else '<span style="color:#aaa">&#8212;</span>'
            br_s  = str(br_v)  if br_v  > 0 else '<span style="color:#aaa">&#8212;</span>'
            tot_s = ('<b>%d</b>' % total) if total else '<span style="color:#aaa">&#8212;</span>'
            efg_s = pill(efg_v) if (a2+a3) else '<span style="color:#aaa">&#8212;</span>'
            s_rows += ('<tr>'
                       '<td %s>%s</td>'
                       '<td %s>%s</td><td %s>%s</td><td %s>%s</td>'
                       '<td %s>%s</td><td %s>%s</td><td %s>%s</td>'
                       '<td %s>%s</td><td %s>%s</td>'
                       '</tr>'
                       % (tdlc, b,
                          tdr, p2, tdr, p3, tdu, ft_s,
                          tdn, c2_s, tdn, c3_s, tdn, br_s,
                          tdc_b, tot_s, tdc, efg_s))

        # BR per kwarta z match_stats
        q_to_vals = {}
        for q in [1,2,3,4]:
            qs = next((r for r in all_stats if r["druzyna"]==druzyna and r["kwarta"]==q), {})
            q_to_vals[q] = qs.get("br",0) or 0

        q_rows = ""
        for q in [1,2,3,4]:
            qcls = qcolors[q]
            to_q = q_to_vals[q]
            # Nagłówek kwarty — bez BR, samo oznaczenie kwarty
            q_rows += ('<tr style="background:#f0f2f7"><td colspan="9" style="padding:5px 14px;border-bottom:0.5px solid #e0e0e0">'
                       '<span class="mqdot %s">%dQ</span>'
                       '</td></tr>' % (qcls, q))
            has = False
            br_q_remaining = to_q  # BR z match_stats dla tej kwarty
            for b in BUCKETS:
                td = get_td(q, b)
                m2=td.get("made2",0); a2=td.get("att2",0)
                m3=td.get("made3",0); a3=td.get("att3",0)
                br_v=td.get("br",0);  ftm_v=td.get("ftm",0)
                poss_ft_v=td.get("poss_ft",0)
                if a2+a3+br_v+ftm_v == 0: continue
                has = True
                eff_poss_ft = poss_ft_v if poss_ft_v > 0 else (1 if ftm_v > 0 else 0)
                udane    = calc_udane(m2, m3, eff_poss_ft)
                nieudane = calc_nieudane(m2, a2, m3, a3, br_v)
                total    = udane + nieudane
                efg_v    = calc_efg(m2, a2, m3, a3, ftm_v)
                chyb2    = a2 - m2
                chyb3    = a3 - m3
                p2 = str(m2) if m2 > 0 else "&#8212;"
                p3 = str(m3) if m3 > 0 else "&#8212;"
                ft_s  = str(eff_poss_ft) if eff_poss_ft > 0 else "&#8212;"
                c2_s  = str(chyb2) if chyb2 > 0 else "&#8212;"
                c3_s  = str(chyb3) if chyb3 > 0 else "&#8212;"
                br_s  = str(br_v)  if br_v  > 0 else "&#8212;"
                tot_s = ('<b>%d</b>' % total) if total else "&#8212;"
                efg_s = pill(efg_v) if (a2+a3) else "&#8212;"
                q_rows += ('<tr>'
                           '<td style="padding:8px 14px 8px 24px;border-bottom:0.5px solid #e0e0e0;font-size:12px;color:#888;text-align:center">%s</td>'
                           '<td %s>%s</td><td %s>%s</td><td %s>%s</td>'
                           '<td %s>%s</td><td %s>%s</td><td %s>%s</td>'
                           '<td %s>%s</td><td %s>%s</td>'
                           '</tr>' % (b,
                                      tdr,p2, tdr,p3, tdu,ft_s,
                                      tdn,c2_s, tdn,c3_s, tdn,br_s,
                                      tdc_b,tot_s, tdc,efg_s))
            if not has:
                q_rows += '<tr><td colspan="9" style="padding:6px 14px 6px 24px;border-bottom:0.5px solid #e0e0e0;font-size:11px;color:#aaa">brak danych</td></tr>'

        tbl_wrap = 'overflow-x:auto;border:0.5px solid #e0e0e0;border-radius:10px;overflow:hidden'
        tbl_s = 'width:100%;border-collapse:collapse;font-size:12px'

        # Nagłówki — dwupoziomowe: Czas | Udane (2PT/3PT/FT≥1) | Nieudane (2PT/3PT/BR) | Razem | eFG%
        th_c  = 'style="background:#1a2b4a;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:right;white-space:nowrap"'
        th_g1 = 'style="background:#085041;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:center;white-space:nowrap;border-bottom:1px solid #5DCAA5"'
        th_r1 = 'style="background:#A32D2D;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:center;white-space:nowrap;border-bottom:1px solid #F09595"'
        th_g2 = 'style="background:#0F6E56;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:right;white-space:nowrap"'
        th_r2 = 'style="background:#791F1F;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:right;white-space:nowrap"'
        th_mid = 'style="background:#1a2b4a;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:center;white-space:nowrap;vertical-align:middle"'
        hdr = ('<tr>'
               '<th rowspan="2" ' + th_mid + ' style="background:#1a2b4a;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:center;white-space:nowrap;vertical-align:middle;width:80px">Czas</th>'
               '<th colspan="3" ' + th_g1 + '>Udane</th>'
               '<th colspan="3" ' + th_r1 + '>Nieudane</th>'
               '<th rowspan="2" ' + th_mid + '>Razem</th>'
               '<th rowspan="2" ' + th_mid + '>eFG%</th>'
               '</tr><tr>'
               '<th ' + th_g2 + '>2PT</th>'
               '<th ' + th_g2 + '>3PT</th>'
               '<th ' + th_g2 + '>FT≥1</th>'
               '<th ' + th_r2 + '>2PT</th>'
               '<th ' + th_r2 + '>3PT</th>'
               '<th ' + th_r2 + '>BR</th>'
               '</tr>')

        out  = '<div id="tim-wrap-' + druzyna + '">'
        out += '<div style="display:flex;border-bottom:0.5px solid #e0e0e0;margin-bottom:12px">'
        out += '<button onclick="timSwitch(\'%s\',\'sum\')" id="tim-btn-sum-%s" style="font-size:12px;padding:8px 14px;border:none;background:none;border-bottom:2px solid #1a2b4a;font-weight:500;color:#1a2b4a;cursor:pointer;margin-bottom:-1px">Suma</button>' % (druzyna,druzyna)
        out += '<button onclick="timSwitch(\'%s\',\'q\')" id="tim-btn-q-%s" style="font-size:12px;padding:8px 14px;border:none;background:none;border-bottom:2px solid transparent;color:#888;cursor:pointer;margin-bottom:-1px">Per kwarta</button>' % (druzyna,druzyna)
        out += '</div>'
        out += '<div id="tim-sum-' + druzyna + '"><div style="' + tbl_wrap + '"><table style="' + tbl_s + '">'
        out += '<thead>' + hdr + '</thead>'
        out += '<tbody>%s</tbody></table></div></div>' % s_rows
        out += '<div id="tim-q-' + druzyna + '" style="display:none"><div style="' + tbl_wrap + '"><table style="' + tbl_s + '">'
        out += '<thead>' + hdr + '</thead>'
        out += '<tbody>%s</tbody></table></div></div></div>' % q_rows
        return out

    pts_q_gtk = [next((r["pts"] for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==q),0) for q in [1,2,3,4]]
    pts_q_opp = [next((r["pts"] for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==q),0) for q in [1,2,3,4]]

    # Dane do wykresu flow
    import json as _json
    try:
        flow_gtk_js = _json.dumps([r["pts_gtk"] for r in flow_rows])
        flow_opp_js = _json.dumps([r["pts_opp"] for r in flow_rows])
    except Exception:
        flow_gtk_js = "[]"
        flow_opp_js = "[]"

    content = f"""
<div class="d-flex justify-content-between align-items-center mb-3 flex-wrap gap-2">
  <div>
    <div class="page-title mb-0">{gtk_name} vs {name_opp}</div>
    <div style="font-size:.8rem;color:#888">{dt} · Sezon {m['sezon']}</div>
  </div>
  <div class="d-flex gap-2">
    <a href="/historia" class="btn btn-outline-secondary btn-sm">← Historia</a>
    <a href="/mecz/{match_id}/edytuj" class="btn btn-outline-primary btn-sm">✏️ Przypisz zawodników</a>
    <a href="/mecz/{match_id}/shooting-chart" class="btn btn-outline-dark btn-sm">🏀 Shooting Chart</a>
    <div class="dropdown">
      <button class="btn btn-warning btn-sm fw-bold dropdown-toggle" type="button" data-bs-toggle="dropdown" aria-expanded="false">
        ⬇ Raport
      </button>
      <ul class="dropdown-menu dropdown-menu-end">
        <li><a class="dropdown-item" href="/mecz/{match_id}/export/xlsx">📊 Pobierz Excel</a></li>
        <li><a class="dropdown-item" href="/mecz/{match_id}/export/pdf">📄 Pobierz PDF</a></li>
        <li><hr class="dropdown-divider"></li>
        <li>
          <form method="POST" action="/admin/reparse-match/{match_id}" style="margin:0">
            <button type="submit" class="dropdown-item" onclick="return confirm('Przeliczyć mecz ponownie z pliku Excel?\nDane w bazie zostaną zastąpione.')"
              style="color:#856404">
              🔄 Przelicz mecz
            </button>
          </form>
        </li>

      </ul>
    </div>
  </div>
</div>

<div class="hero mb-3">
  <div class="d-flex justify-content-between align-items-center flex-wrap gap-3">
    <div>
      <div style="font-size:.8rem;opacity:.7">{gtk_name}</div>
      <div style="font-size:2rem;font-weight:700">{m['wynik_gtk']}</div>
    </div>
    <div class="text-center">
      {'<span class="badge-win" style="font-size:.9rem;padding:6px 14px">WYGRANA</span>' if m['wynik_gtk']>m['wynik_opp'] else '<span class="badge-loss" style="font-size:.9rem;padding:6px 14px">PRZEGRANA</span>' if m['wynik_gtk']<m['wynik_opp'] else '<span class="badge-draw" style="font-size:.9rem;padding:6px 14px">REMIS</span>'}
    </div>
    <div class="text-end">
      <div style="font-size:.8rem;opacity:.7">{name_opp}</div>
      <div style="font-size:2rem;font-weight:700">{m['wynik_opp']}</div>
    </div>
  </div>
  <div style="border-top:1px solid rgba(255,255,255,.15);margin-top:8px;padding-top:8px;display:flex;justify-content:center;gap:10px">
    {"".join(f'<div style="text-align:center"><div style="font-size:8px;opacity:.55;letter-spacing:.5px;margin-bottom:3px">{q}Q</div><div style="background:rgba(255,255,255,.13);border-radius:6px;padding:4px 12px;font-size:13px;font-weight:700;letter-spacing:1px">{pts_q_gtk[q-1]} : {pts_q_opp[q-1]}</div></div>' for q in [1,2,3,4])}
  </div>
</div>

<ul class="nav nav-tabs mb-2" id="mainTabs">
  <li class="nav-item"><button class="nav-link active" data-bs-toggle="tab" data-bs-target="#tabGTK">{gtk_name}</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tabOPP">{name_opp}</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tabCMP">Porównanie</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tabFLOW">Przebieg</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tabCLUTCH">Clutch</button></li>
</ul>

<div class="tab-content">

<div class="tab-pane fade show active" id="tabGTK">
  {kpi_cards(suma_gtk, kpi_gtk, "gtk")}
  <div style="display:flex;gap:4px;margin:8px 0 4px;border-bottom:1px solid #dee2e6;padding-bottom:0">
    <button onclick="gtkTab('gtk_q')" id="gtktab-gtk_q" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid #1a2b4a;font-weight:500;color:#1a2b4a;cursor:pointer;margin-bottom:-1px">Per kwarta</button>
    <button onclick="gtkTab('gtk_p')" id="gtktab-gtk_p" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid transparent;color:#666;cursor:pointer;margin-bottom:-1px">Statystyki</button>
    <button onclick="gtkTab('gtk_l')" id="gtktab-gtk_l" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid transparent;color:#666;cursor:pointer;margin-bottom:-1px">Piątki</button>
    <button onclick="gtkTab('gtk_t')" id="gtktab-gtk_t" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid transparent;color:#666;cursor:pointer;margin-bottom:-1px">Timing akcji</button>
  </div>
  <div id="gpane-gtk_q" style="display:block"><div class="card mt-1"><div class="card-body" style="padding:0">{q_table('gtk')}</div></div></div>
  <div id="gpane-gtk_p" style="display:none"><div class="card mt-1"><div class="card-body" style="padding:0">{p_table('gtk')}</div></div></div>
   <div id="gpane-gtk_l" style="display:none"><div class="card mt-1"><div class="card-body" style="padding:0">{lineup_table()}</div></div></div>
  <div id="gpane-gtk_t" style="display:none"><div class="card mt-1"><div class="card-body p-2">{tim_table('gtk')}</div></div></div>
</div>

<div class="tab-pane fade" id="tabOPP">
  {kpi_cards(suma_opp, kpi_opp, "opp")}
  <div style="display:flex;gap:4px;margin:8px 0 4px;border-bottom:1px solid #dee2e6;padding-bottom:0">
    <button onclick="oppTab('opp_q')" id="opptab-opp_q" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid #1a2b4a;font-weight:500;color:#1a2b4a;cursor:pointer;margin-bottom:-1px">Per kwarta</button>
    <button onclick="oppTab('opp_t')" id="opptab-opp_t" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid transparent;color:#666;cursor:pointer;margin-bottom:-1px">Timing akcji</button>
  </div>
  <div id="opane-opp_q" style="display:block"><div class="card mt-1"><div class="card-body" style="padding:0">{q_table('opp')}</div></div></div>
  <div id="opane-opp_t" style="display:none"><div class="card mt-1"><div class="card-body p-2">{tim_table('opp')}</div></div></div>
</div>

<div class="tab-pane fade" id="tabCMP">
  <div style="margin:10px 0 8px;font-size:.78rem;color:#888">Kliknij metrykę aby rozwinąć wykres per kwarta</div>
  <div id="cmp-accordion"></div>
</div>

<div class="tab-pane fade" id="tabFLOW">
  <div class="card mt-1"><div class="card-body p-2"><div style="background:#f4f6fb;border-radius:8px;padding:10px 12px;margin-bottom:8px;display:flex;gap:16px;align-items:center"><div style="display:flex;align-items:center;gap:6px"><div style="width:20px;height:3px;background:#1a6b3c;border-radius:2px"></div><span style="font-size:12px;color:#666">{gtk_name}</span></div><div style="display:flex;align-items:center;gap:6px"><div style="width:20px;height:3px;background:#c0392b;border-radius:2px"></div><span style="font-size:12px;color:#666">{name_opp}</span></div><span style="font-size:11px;color:#aaa;margin-left:auto">Skumulowane punkty w czasie meczu</span></div><div style="border:0.5px solid #e0e0e0;border-radius:8px;padding:12px;margin-bottom:10px"><canvas id="mFlowChart" style="width:100%;height:200px"></canvas></div><div style="border:0.5px solid #e0e0e0;border-radius:8px;overflow:hidden">{momentum_table()}</div></div></div>
</div>

<div class="tab-pane fade" id="tabCLUTCH">
  <div class="card mt-1"><div class="card-body p-2">{clutch_stats()}</div></div>
</div>

</div>

</div>"""

    import json as _json
    def _qval(druzyna, field, q):
        r = next((x for x in all_stats if x['druzyna']==druzyna and x['kwarta']==q), {})
        return r.get(field, 0) or 0
    def _qefg(druzyna, q):
        p2m=_qval(druzyna,'p2m',q); p3m=_qval(druzyna,'p3m',q)
        p2a=_qval(druzyna,'p2a',q); p3a=_qval(druzyna,'p3a',q)
        return round((p2m+1.5*p3m)/(p2a+p3a)*100,1) if (p2a+p3a) else 0
    def _qortg(druzyna, q):
        pts=_qval(druzyna,'pts',q); poss=_qval(druzyna,'poss',q)
        return round(pts/poss*100,1) if poss else 0
    def _qppp(druzyna, q):
        pts=_qval(druzyna,'pts',q); poss=_qval(druzyna,'poss',q)
        return round(pts/poss,2) if poss else 0
    def _qpct(druzyna, made, att, q):
        m=_qval(druzyna,made,q); a=_qval(druzyna,att,q)
        return round(m/a*100,1) if a else 0
    def _qreb(druzyna, q):
        return (_qval(druzyna,'oreb',q) or 0)+(_qval(druzyna,'dreb',q) or 0)
    def _ast(druzyna):
        return sum(r.get('ast',0) or 0 for r in all_stats if r['druzyna']==druzyna)
    def _stl(druzyna):
        return sum(r.get("stl",0) or 0 for r in all_stats if r["druzyna"]==druzyna)
    def _fin(druzyna):
        m = sum(int(r.get("d2m",0) or 0) for r in all_stats if r["druzyna"]==druzyna)
        a = sum(int(r.get("d2a",0) or 0) for r in all_stats if r["druzyna"]==druzyna)
        return (m, a)
    def _netrtg_q(q):
        g_pts=_qval("gtk","pts",q); g_poss=_qval("gtk","poss",q)
        o_pts=_qval("opp","pts",q); o_poss=_qval("opp","poss",q)
        ortg = g_pts/g_poss*100 if g_poss else 0
        drtg = o_pts/o_poss*100 if o_poss else 0
        return round(ortg - drtg, 1)
    _g_ortg = float(kpi_gtk["ortg"]) if kpi_gtk["ortg"] != "-" else 0
    _o_ortg = float(kpi_opp["ortg"]) if kpi_opp["ortg"] != "-" else 0
    _netrtg_total = round(_g_ortg - _o_ortg, 1)
    cmp_list = [
        ("Punkty",        suma_gtk.get("pts",0),       suma_opp.get("pts",0),       False,
         [_qval("gtk","pts",q) for q in [1,2,3,4]], [_qval("opp","pts",q) for q in [1,2,3,4]]),
        ("Zbiórki",       (suma_gtk.get("oreb",0) or 0)+(suma_gtk.get("dreb",0) or 0),
                          (suma_opp.get("oreb",0) or 0)+(suma_opp.get("dreb",0) or 0), False,
         [_qreb("gtk",q) for q in [1,2,3,4]], [_qreb("opp",q) for q in [1,2,3,4]]),
        ("Asysty",        _ast("gtk"),                 _ast("opp"),                 False,
         [_qval("gtk","ast",q) for q in [1,2,3,4]], [_qval("opp","ast",q) for q in [1,2,3,4]]),
        ("Przechwyty",    _stl("gtk"),                 _stl("opp"),                 False,
         [_qval("gtk","stl",q) for q in [1,2,3,4]], [_qval("opp","stl",q) for q in [1,2,3,4]]),
        ("Straty (TO)",   suma_gtk.get("br",0),        suma_opp.get("br",0),        True,
         [_qval("gtk","br",q) for q in [1,2,3,4]], [_qval("opp","br",q) for q in [1,2,3,4]]),
        ("Dobitki",       f"{_fin('gtk')[0]}/{_fin('gtk')[1]}",  f"{_fin('opp')[0]}/{_fin('opp')[1]}",  True,
         [_qval("gtk","d2m",q) for q in [1,2,3,4]],
         [_qval("opp","d2m",q) for q in [1,2,3,4]]),
        ("NetRtg",        _netrtg_total,               -_netrtg_total,              False,
         [_netrtg_q(q) for q in [1,2,3,4]], [-_netrtg_q(q) for q in [1,2,3,4]]),
        ("PPP",           kpi_gtk["ppp"],              kpi_opp["ppp"],              False,
         [_qppp("gtk",q) for q in [1,2,3,4]], [_qppp("opp",q) for q in [1,2,3,4]]),
        ("TS%",           kpi_gtk["ts"],               kpi_opp["ts"],               False,
         [round(_qval("gtk","pts",q)/(2*(_qval("gtk","p2a",q)+_qval("gtk","p3a",q)+0.44*_qval("gtk","fta",q)))*100,1) if (_qval("gtk","p2a",q)+_qval("gtk","p3a",q)+_qval("gtk","fta",q)) else 0 for q in [1,2,3,4]],
         [round(_qval("opp","pts",q)/(2*(_qval("opp","p2a",q)+_qval("opp","p3a",q)+0.44*_qval("opp","fta",q)))*100,1) if (_qval("opp","p2a",q)+_qval("opp","p3a",q)+_qval("opp","fta",q)) else 0 for q in [1,2,3,4]]),
        ("eFG%",          kpi_gtk["efg"],              kpi_opp["efg"],              False,
         [_qefg("gtk",q) for q in [1,2,3,4]], [_qefg("opp",q) for q in [1,2,3,4]]),
    ]
    cmp_js = _json.dumps([{"lbl":l,"g":str(g),"o":str(o),"low":low,"gq":gq,"oq":oq}
                          for l,g,o,low,gq,oq in cmp_list])

    scripts = f"""<script>
const gtkName = {_json.dumps(gtk_name)};
const oppName = {_json.dumps(name_opp)};
const CMP_DATA = {cmp_js};

var _cmpCharts = {{}};
(function() {{
  var wrap = document.getElementById('cmp-accordion');
  if (!wrap) return;
  CMP_DATA.forEach(function(m, i) {{
    var gNum = parseFloat(String(m.g).replace('%','')) || 0;
    var oNum = parseFloat(String(m.o).replace('%','')) || 0;
    var total = gNum + oNum || 1;
    var gPct = Math.round(gNum / total * 100);
    var oPct = 100 - gPct;
    var diff = gNum - oNum;
    var ds = (diff > 0 ? '+' : '') + (diff % 1 === 0 ? Math.round(diff) : diff.toFixed(1));
    var gBetter = m.low ? gNum < oNum : gNum > oNum;
    var dc = diff === 0 ? 'color:#888' : (gBetter ? 'color:#1a6b3c' : 'color:#8b1a1a');
    var el = document.createElement('div');
    el.style.cssText = 'border:0.5px solid #e0e0e0;border-radius:8px;margin-bottom:8px;overflow:hidden';
    el.innerHTML =
      '<div style="display:flex;align-items:center;gap:0;padding:10px 14px;cursor:pointer;background:#fafafa" onclick="toggleCmp('+i+',this)">' +
        '<span style="font-weight:500;min-width:90px;font-size:.85rem">'+m.lbl+'</span>' +
        '<div style="display:flex;align-items:center;gap:10px;flex:1">' +
          '<div style="text-align:right;min-width:52px"><div style="font-size:9px;color:#185FA5;text-transform:uppercase;letter-spacing:.3px">'+ gtkName +'</div><div style="font-size:15px;font-weight:600;color:#0C447C">'+m.g+'</div></div>' +
          '<div style="flex:1;height:10px;background:#e8e8e8;border-radius:5px;overflow:hidden;display:flex">' +
            '<div style="flex:'+gPct+';background:#185FA5;height:100%;border-radius:5px 0 0 5px"></div>' +
            '<div style="flex:'+oPct+';background:#c0392b;height:100%;border-radius:0 5px 5px 0"></div>' +
          '</div>' +
          '<div style="text-align:left;min-width:52px"><div style="font-size:9px;color:#c0392b;text-transform:uppercase;letter-spacing:.3px">'+ oppName +'</div><div style="font-size:15px;font-weight:600;color:#a32d2d">'+m.o+'</div></div>' +
        '</div>' +
        '<span id="cmp-chev-'+i+'" style="font-size:11px;color:#aaa;margin-left:10px">&#9660;</span>' +
      '</div>' +
      '<div id="cmp-body-'+i+'" style="display:none;padding:0 14px 14px;border-top:0.5px solid #e8e8e8">' +
        '<canvas id="cmp-chart-'+i+'" height="100"></canvas>' +
      '</div>';
    wrap.appendChild(el);
  }});
}})();

function toggleCmp(i, header) {{
  var body = document.getElementById('cmp-body-'+i);
  var chev = document.getElementById('cmp-chev-'+i);
  var isOpen = body.style.display !== 'none';
  body.style.display = isOpen ? 'none' : 'block';
  chev.style.transform = isOpen ? '' : 'rotate(180deg)';
  if (!isOpen && !_cmpCharts[i]) {{
    var m = CMP_DATA[i];
    var ctx = document.getElementById('cmp-chart-'+i);
    if (!ctx) return;
    _cmpCharts[i] = new Chart(ctx, {{
      type: 'bar',
      data: {{
        labels: ['1Q','2Q','3Q','4Q'],
        datasets: [
          {{label: gtkName, data: m.gq, backgroundColor: '#185FA5cc', borderColor: '#185FA5', borderWidth: 1, borderRadius: 4, barPercentage: .45}},
          {{label: oppName, data: m.oq, backgroundColor: '#c0392bcc', borderColor: '#c0392b', borderWidth: 1, borderRadius: 4, barPercentage: .45}}
        ]
      }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{legend: {{position:'top',labels:{{font:{{size:11}},boxWidth:12}}}},tooltip:{{mode:'index',intersect:false}}}},
        scales: {{x:{{grid:{{display:false}},ticks:{{font:{{size:11}}}}}},y:{{beginAtZero:true,grid:{{color:'rgba(0,0,0,.05)'}},ticks:{{font:{{size:10}}}}}}}}
      }}
    }});
  }}
}}

function gtkTab(id) {{
    ['gtk_q','gtk_p','gtk_l','gtk_t'].forEach(function(p) {{
        var pane = document.getElementById('gpane-'+p);
        var btn  = document.getElementById('gtktab-'+p);
        if (!pane || !btn) return;
        var active = p === id;
        pane.style.display = active ? 'block' : 'none';
        btn.style.borderBottomColor = active ? '#1a2b4a' : 'transparent';
        btn.style.color = active ? '#1a2b4a' : '#666';
        btn.style.fontWeight = active ? '500' : 'normal';
    }});
}}

function oppTab(id) {{
    ['opp_q','opp_t'].forEach(function(p) {{
        var pane = document.getElementById('opane-'+p);
        var btn  = document.getElementById('opptab-'+p);
        if (!pane || !btn) return;
        var active = p === id;
        pane.style.display = active ? 'block' : 'none';
        btn.style.borderBottomColor = active ? '#1a2b4a' : 'transparent';
        btn.style.color = active ? '#1a2b4a' : '#666';
        btn.style.fontWeight = active ? '500' : 'normal';
    }});
}}

// Flow chart
(function() {{
  var gtkData = {flow_gtk_js};
  var oppData = {flow_opp_js};
  if (!gtkData.length) return;
  var labels = gtkData.map(function(_,i){{return i;}});
  function initFlow() {{
    var canvas = document.getElementById('mFlowChart');
    if (!canvas) return;
    if (window._flowChart) {{ window._flowChart.destroy(); }}
    window._flowChart = new Chart(canvas, {{
      type: 'line',
      data: {{
        labels: labels,
        datasets: [
          {{label: '{gtk_name}', data: gtkData, borderColor:'#1a6b3c', backgroundColor:'transparent', borderWidth:2, pointRadius:0, tension:.3}},
          {{label: '{name_opp}', data: oppData, borderColor:'#c0392b', backgroundColor:'transparent', borderWidth:2, pointRadius:0, tension:.3}}
        ]
      }},
      options: {{
        responsive:true, maintainAspectRatio:false,
        plugins:{{legend:{{display:false}},tooltip:{{mode:'index',intersect:false}}}},
        scales:{{x:{{display:false}},y:{{beginAtZero:true,grid:{{color:'rgba(0,0,0,.05)'}},ticks:{{font:{{size:10}},stepSize:10}}}}}}
      }}
    }});
  }}
  // Inicjalizuj gdy zakładka Przebieg (główna) staje się widoczna
  var flowTab = document.querySelector('[data-bs-target="#tabFLOW"]');
  if (flowTab) {{
    flowTab.addEventListener('shown.bs.tab', function() {{ setTimeout(initFlow, 50); }});
  }}
}})();


function timSwitch(druzyna, view) {{
    var sum = document.getElementById("tim-sum-" + druzyna);
    var q   = document.getElementById("tim-q-"   + druzyna);
    var btnSum = document.getElementById("tim-btn-sum-" + druzyna);
    var btnQ   = document.getElementById("tim-btn-q-"   + druzyna);
    if (!sum || !q) return;
    if (view === "sum") {{
        sum.style.display = ""; q.style.display = "none";
        if(btnSum){{btnSum.style.borderBottomColor="#1a2b4a";btnSum.style.color="#1a2b4a";btnSum.style.fontWeight="500";}}
        if(btnQ){{btnQ.style.borderBottomColor="transparent";btnQ.style.color="#666";btnQ.style.fontWeight="normal";}}
    }} else {{
        sum.style.display = "none"; q.style.display = "";
        if(btnQ){{btnQ.style.borderBottomColor="#1a2b4a";btnQ.style.color="#1a2b4a";btnQ.style.fontWeight="500";}}
        if(btnSum){{btnSum.style.borderBottomColor="transparent";btnSum.style.color="#666";btnSum.style.fontWeight="normal";}}
    }}
}}
// Domyślne sortowanie tabeli zawodników wg PTS malejąco
document.addEventListener('DOMContentLoaded', function() {{
  ['gtk','opp'].forEach(function(druzyna) {{
    var pane = document.getElementById('gpane-'+druzyna+'_p');
    if (!pane) return;
    var tbl = pane.querySelector('table');
    if (!tbl) return;
    var tb = tbl.querySelector('tbody');
    if (!tb) return;
    var rows = Array.from(tb.querySelectorAll('tr'));
    rows.sort(function(a,b) {{
      var av = parseFloat(a.children[2] && a.children[2].dataset.v) || 0;
      var bv = parseFloat(b.children[2] && b.children[2].dataset.v) || 0;
      return bv - av;
    }});
    rows.forEach(function(r,i) {{
      r.style.background = i%2===0 ? '#f8f9ff' : '#fff';
      tb.appendChild(r);
    }});
  }});
}});
</script>"""

    return html_response(base(content, scripts, active="history"))


@app.route("/mecz/<int:match_id>/delete")
@login_required
def mecz_delete(match_id):
    db = get_db(); cur = db.cursor()
    cur.execute("DELETE FROM matches WHERE id=%s", (match_id,))
    db.commit(); cur.close()
    flash("Mecz usunięty","success")
    return redirect(url_for("historia"))

# ══════════════════════════════════════════════════════════════════════════════
# RAPORT TRENERSKI
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/mecz/<int:match_id>/raport-trenerski")
@login_required
def raport_trenerski(match_id):
    db = get_db(); cur = db.cursor()
    # Mapa nr → nazwisko (wszystkie źródła)
    nr_name_map = build_nr_name_map(cur, match_id)


    cur.execute("SELECT * FROM match_stats WHERE match_id=%s ORDER BY kwarta", (match_id,))
    all_stats = list(cur.fetchall())
    cur.execute("SELECT * FROM player_stats WHERE match_id=%s AND druzyna='gtk'", (match_id,))
    all_players = list(cur.fetchall())

    # Mapa nr → nazwisko (wszystkie źródła)
    nr_name_map = build_nr_name_map(cur, match_id)


    # Piątki OFF + DEF → NET
    try:
        cur.execute("SELECT * FROM lineup_stats WHERE match_id=%s AND druzyna='gtk' ORDER BY poss DESC", (match_id,))
        lu_off = list(cur.fetchall())
        cur.execute("SELECT * FROM lineup_stats WHERE match_id=%s AND druzyna='gtk_def' ORDER BY poss DESC", (match_id,))
        lu_def = list(cur.fetchall())
    except: lu_off = lu_def = []

    off_rtg = {lu["lineup"]: lu["pts"]*100/lu["poss"] for lu in lu_off if lu.get("poss",0)>0}
    def_rtg = {lu["lineup"]: lu["pts"]*100/lu["poss"] for lu in lu_def if lu.get("poss",0)>0}
    net_lineups = []
    seen = {}
    for lu in lu_off:
        k = lu["lineup"]
        if k not in seen:
            seen[k] = True
            net = round(off_rtg[k] - def_rtg[k], 1) if k in off_rtg and k in def_rtg else None
            net_lineups.append({"lineup": k, "poss": lu["poss"], "ortg": off_rtg.get(k), "drtg": def_rtg.get(k), "net": net})
    net_lineups.sort(key=lambda x: (x["net"] is not None, x["net"] if x["net"] else 0), reverse=True)

    cur.close()

    def build_suma(druzyna):
        s = {"pts":0,"poss":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"br":0,"fd":0}
        for row in all_stats:
            if row["druzyna"] == druzyna:
                for k in s: s[k] += row.get(k,0) or 0
        return s

    sg = build_suma("gtk"); so = build_suma("opp")
    kg = calc_kpi(sg); ko = calc_kpi(so)
    net_rtg_total = f"{float(kg['ortg']) - float(ko['ortg']):.1f}" if kg["ortg"]!="-" and ko["ortg"]!="-" else "—"

    def pct_bar(val_g, val_o, higher_is_better=True):
        try:
            vg = float(str(val_g).split('/')[0].replace('%',''))
            vo = float(str(val_o).split('/')[0].replace('%',''))
            g_win = (vg > vo) if higher_is_better else (vg < vo)
            o_win = (vo > vg) if higher_is_better else (vo < vg)
            gc = "#1a6b3c" if g_win else ("#8b1a1a" if o_win else "#555")
            oc = "#8b1a1a" if g_win else ("#1a6b3c" if o_win else "#555")
            return f'<td style="color:{gc};font-weight:{"700" if g_win else "400"}">{val_g}</td><td style="color:{oc};font-weight:{"700" if o_win else "400"}">{val_o}</td>'
        except: return f'<td>{val_g}</td><td>{val_o}</td>'

    # Tabela per kwarta
    q_rows = ""
    for qn in [1,2,3,4]:
        qg = next((r for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==qn), {})
        qo = next((r for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==qn), {})
        fg = qg.get("pts",0) or 0; fo = qo.get("pts",0) or 0
        win_g = fg > fo; win_o = fo > fg
        qcolors = {1:"#c8e6c9",2:"#bbdefb",3:"#fff9c4",4:"#fce4ec"}
        q_rows += f"""<tr>
          <td style="background:{qcolors[qn]};font-weight:600;font-size:11px;padding:4px 8px">{qn}Q</td>
          <td style="color:{'#1a6b3c' if win_g else '#333'};font-weight:{'700' if win_g else '400'};text-align:center">{fg}</td>
          <td style="color:{'#8b1a1a' if win_o else '#333'};font-weight:{'700' if win_o else '400'};text-align:center">{fo}</td>
          <td style="text-align:center;font-size:11px">{qg.get('p2m',0)}/{qg.get('p2a',0)}</td>
          <td style="text-align:center;font-size:11px">{qg.get('p3m',0)}/{qg.get('p3a',0)}</td>
          <td style="text-align:center;font-size:11px">{qg.get('ftm',0)}/{qg.get('fta',0)}</td>
          <td style="text-align:center;font-size:11px">{qg.get('br',0)}</td>
          <td style="text-align:center;font-size:11px">{qg.get('poss',0)}</td>
        </tr>"""

    # Top 5 zawodników po punktach
    top_players = sorted(all_players, key=lambda x: x.get("pts",0) or 0, reverse=True)[:6]
    p_rows = ""
    for pd in top_players:
        name = nr_name_map.get(str(pd["nr"]), f"#{pd['nr']}")
        fga = (pd.get("p2a",0) or 0) + (pd.get("p3a",0) or 0)
        fta = pd.get("fta",0) or 0
        pts = pd.get("pts",0) or 0
        efg = f"{((pd.get('p2m',0) or 0)+1.5*(pd.get('p3m',0) or 0))/fga:.0%}" if fga else "—"
        usg = f"{(fga+0.44*fta+(pd.get('br',0) or 0))/(sg['poss']):.0%}" if sg['poss'] else "—"
        p_rows += f"""<tr>
          <td style="font-size:11px;font-weight:600">{name}</td>
          <td style="text-align:center;font-weight:700;color:#1a2b4a">{pts}</td>
          <td style="text-align:center;font-size:11px">{pd.get('p2m',0)}/{pd.get('p2a',0)}</td>
          <td style="text-align:center;font-size:11px">{pd.get('p3m',0)}/{pd.get('p3a',0)}</td>
          <td style="text-align:center;font-size:11px">{pd.get('ftm',0)}/{fta}</td>
          <td style="text-align:center">{efg}</td>
          <td style="text-align:center">{usg}</td>
          <td style="text-align:center;font-size:11px">{pd.get('ast',0)}</td>
          <td style="text-align:center;font-size:11px">{pd.get('br',0)}</td>
        </tr>"""

    # Top piątki NET
    net_rows = ""
    for lu in net_lineups[:5]:
        k = lu["lineup"]
        skl = " · ".join(nr_name_map.get(n, f"#{n}") for n in k.split("-"))
        net_val = lu["net"]
        net_str = f"{net_val:+.1f}" if net_val is not None else "—"
        net_col = "#1a6b3c" if net_val and net_val>0 else ("#8b1a1a" if net_val and net_val<0 else "#888")
        ortg_str = f"{lu['ortg']:.1f}" if lu.get("ortg") else "—"
        drtg_str = f"{lu['drtg']:.1f}" if lu.get("drtg") else "—"
        net_rows += f"""<tr>
          <td style="font-size:10px">{skl}</td>
          <td style="text-align:center;font-size:11px">{lu['poss']}</td>
          <td style="text-align:center;font-size:11px;color:#1a6b3c">{ortg_str}</td>
          <td style="text-align:center;font-size:11px;color:#8b1a1a">{drtg_str}</td>
          <td style="text-align:center;font-weight:700;color:{net_col}">{net_str}</td>
        </tr>"""

    wynik_g = m["wynik_gtk"]; wynik_o = m["wynik_opp"]
    wygrana = wynik_g > wynik_o

    html = f"""<!DOCTYPE html>
<html lang="pl">
<head>
<meta charset="UTF-8">
<title>Raport trenerski — {gtk_name} vs {name_opp}</title>
<style>
  @page {{ size: A4 landscape; margin: 12mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; font-size: 12px; color: #222; background: #fff; }}
  .page {{ width: 100%; }}
  .header {{ display: flex; justify-content: space-between; align-items: flex-start;
             border-bottom: 3px solid #1a2b4a; padding-bottom: 8px; margin-bottom: 10px; }}
  .header-left h1 {{ font-size: 20px; font-weight: 700; color: #1a2b4a; }}
  .header-left p {{ font-size: 11px; color: #666; margin-top: 2px; }}
  .score-box {{ background: #1a2b4a; color: #fff; border-radius: 8px; padding: 8px 20px; text-align: center; }}
  .score-box .score {{ font-size: 28px; font-weight: 700; }}
  .score-box .result {{ font-size: 11px; margin-top: 2px;
                        color: {'#7fff7f' if wygrana else '#ff9999'}; font-weight: 600; }}
  .grid-3 {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px; }}
  .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }}
  .card {{ border: 1px solid #ddd; border-radius: 6px; padding: 8px; }}
  .card-title {{ font-size: 10px; font-weight: 700; color: #666; text-transform: uppercase;
                 letter-spacing: .04em; margin-bottom: 6px; border-bottom: 1px solid #eee; padding-bottom: 4px; }}
  .kpi-grid {{ display: grid; grid-template-columns: repeat(3,1fr); gap: 5px; }}
  .kpi {{ background: #f4f6fb; border-radius: 4px; padding: 5px 6px; text-align: center; }}
  .kpi-val {{ font-size: 16px; font-weight: 700; color: #1a2b4a; }}
  .kpi-lbl {{ font-size: 9px; color: #888; text-transform: uppercase; }}
  .kpi-net {{ background: #e8f5e9; }}
  .kpi-net .kpi-val {{ color: #1a6b3c; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th {{ background: #1a2b4a; color: #fff; font-size: 10px; padding: 4px 6px; text-align: center; font-weight: 600; }}
  th:first-child {{ text-align: left; }}
  td {{ padding: 4px 6px; border-bottom: .5px solid #eee; font-size: 11px; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:nth-child(even) {{ background: #f8f9ff; }}
  .footer {{ margin-top: 10px; border-top: 1px solid #ddd; padding-top: 6px;
             display: flex; justify-content: space-between; font-size: 9px; color: #aaa; }}
  @media print {{ body {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }} }}
</style>
</head>
<body>
<div class="page">

<!-- NAGŁÓWEK -->
<div class="header">
  <div class="header-left">
    <h1>{gtk_name} vs {name_opp}</h1>
    <p>{dt} &nbsp;·&nbsp; {m.get('rozgrywki','')} &nbsp;·&nbsp; Runda {m.get('runda','')} &nbsp;·&nbsp; {m.get('miejsce','').capitalize()}</p>
    <p style="margin-top:6px;font-size:10px;color:#aaa">Raport trenerski · Basket Kołcz Analytics</p>
  </div>
  <div class="score-box">
    <div class="score">{wynik_g} : {wynik_o}</div>
    <div class="result">{'WYGRANA' if wygrana else 'PORAŻKA'}</div>
  </div>
</div>

<!-- KPI + KWARTY + ZAWODNICY -->
<div class="grid-3" style="margin-bottom:10px">

  <!-- KPI -->
  <div class="card">
    <div class="card-title">Kluczowe metryki</div>
    <div class="kpi-grid">
      <div class="kpi"><div class="kpi-val">{kg['efg']}</div><div class="kpi-lbl">eFG%</div></div>
      <div class="kpi"><div class="kpi-val">{kg['ts']}</div><div class="kpi-lbl">TS%</div></div>
      <div class="kpi"><div class="kpi-val">{kg['ortg']}</div><div class="kpi-lbl">ORtg</div></div>
      <div class="kpi"><div class="kpi-val">{ko['ortg']}</div><div class="kpi-lbl">DRtg</div></div>
      <div class="kpi"><div class="kpi-val">{kg['ppp']}</div><div class="kpi-lbl">PPP</div></div>
      <div class="kpi kpi-net"><div class="kpi-val">{net_rtg_total}</div><div class="kpi-lbl">NetRtg</div></div>
      <div class="kpi"><div class="kpi-val">{sg.get('poss',0)}</div><div class="kpi-lbl">POSS</div></div>
      <div class="kpi"><div class="kpi-val">{sg.get('br',0)}</div><div class="kpi-lbl">Straty</div></div>
      <div class="kpi"><div class="kpi-val">{kg['ft_pct']}</div><div class="kpi-lbl">FT%</div></div>
    </div>
    <table style="margin-top:8px">
      <thead><tr><th>Metryka</th><th>{gtk_name}</th><th>{name_opp}</th></tr></thead>
      <tbody>
        <tr><td>eFG%</td>{pct_bar(kg['efg'], ko['efg'])}</tr>
        <tr><td>ORtg</td>{pct_bar(kg['ortg'], ko['ortg'])}</tr>
        <tr><td>2PT%</td>{pct_bar(kg['p2_pct'], ko['p2_pct'])}</tr>
        <tr><td>3PT%</td>{pct_bar(kg['p3_pct'], ko['p3_pct'])}</tr>
        <tr><td>FT%</td>{pct_bar(kg['ft_pct'], ko['ft_pct'])}</tr>
        <tr><td>Straty/POSS</td>{pct_bar(kg['topct'], ko['topct'], False)}</tr>
      </tbody>
    </table>
  </div>

  <!-- PER KWARTA -->
  <div class="card">
    <div class="card-title">{gtk_name} — per kwarta</div>
    <table>
      <thead><tr><th>Q</th><th>GTK</th><th>OPP</th><th>2PM/A</th><th>3PM/A</th><th>FTM/A</th><th>TO</th><th>POSS</th></tr></thead>
      <tbody>{q_rows}</tbody>
    </table>
  </div>

  <!-- ZAWODNICY -->
  <div class="card">
    <div class="card-title">Top zawodnicy</div>
    <table>
      <thead><tr><th>Zawodnik</th><th>PTS</th><th>2PM/A</th><th>3PM/A</th><th>FTM/A</th><th>eFG%</th><th>USG%</th><th>AST</th><th>TO</th></tr></thead>
      <tbody>{p_rows}</tbody>
    </table>
  </div>

</div>

<!-- PIĄTKI NET -->
<div class="card">
  <div class="card-title">Piątki — Net Rating (ORtg OFF − DRtg DEF) · top 5</div>
  {'<p style="font-size:10px;color:#aaa;padding:4px 0">Brak danych NET — wgraj mecz ponownie.</p>' if not net_lineups else f"""
  <table>
    <thead><tr><th>Skład</th><th>POSS</th><th>ORtg</th><th>DRtg</th><th>NetRtg</th></tr></thead>
    <tbody>{net_rows}</tbody>
  </table>"""}
</div>

<div class="footer">
  <span>Basket Kołcz Analytics · basketkolcz.onrender.com</span>
  <span>Aby zapisać jako PDF: Ctrl+P → Zapisz jako PDF · Orientacja: pozioma · Marginesy: minimalne</span>
  <span>Wygenerowano: {dt}</span>
</div>

</div>
</body>
</html>"""

    return html

# ══════════════════════════════════════════════════════════════════════════════
# SHOOTING CHART / HOT ZONES
# ══════════════════════════════════════════════════════════════════════════════

# Współrzędne środków stref na obrazku (% szerokości, % wysokości)
# Atak: strefy 1-16, Obrona: strefy 17-28
# Obrazek: koszyk na dole, widok z góry
ZONE_META = {
    # --- ATAK (2PT: 1-6, 3PT: 7-16) ---
    1:  {"name": "Paint – Pod Koszem",       "pts": 2, "cx": 50.0, "cy": 83.0},
    2:  {"name": "Baseline Right – MidRange", "pts": 2, "cx": 74.0, "cy": 76.0},
    3:  {"name": "Baseline Left – MidRange",  "pts": 2, "cx": 26.0, "cy": 76.0},
    4:  {"name": "Paint – Midrange",          "pts": 2, "cx": 50.0, "cy": 68.0},
    5:  {"name": "Right Wing – MidRange",     "pts": 2, "cx": 70.0, "cy": 60.0},
    6:  {"name": "Left Wing – MidRange",      "pts": 2, "cx": 30.0, "cy": 60.0},
    7:  {"name": "Corner Right – 3PT",        "pts": 3, "cx": 93.0, "cy": 80.0},
    8:  {"name": "Corner Left – 3PT",         "pts": 3, "cx":  7.0, "cy": 80.0},
    9:  {"name": "Right Wing – 3PT",          "pts": 3, "cx": 84.0, "cy": 58.0},
    10: {"name": "Left Wing – 3PT",           "pts": 3, "cx": 16.0, "cy": 58.0},
    11: {"name": "Deep Right – 3PT",          "pts": 3, "cx": 78.0, "cy": 38.0},
    12: {"name": "Deep Left – 3PT",           "pts": 3, "cx": 22.0, "cy": 38.0},
    13: {"name": "TOP – 3PT",                 "pts": 3, "cx": 50.0, "cy": 34.0},
    14: {"name": "Deep TOP – 3PT",            "pts": 3, "cx": 50.0, "cy": 20.0},
    15: {"name": "HalfCourt Right",           "pts": 3, "cx": 67.0, "cy": 10.0},
    16: {"name": "HalfCourt Left",            "pts": 3, "cx": 33.0, "cy": 10.0},
    # --- OBRONA (strefy 17-28, przeciwnik strzela) ---
    17: {"name": "HalfCourt def – Right",     "pts": 3, "cx": 67.0, "cy": 10.0},
    18: {"name": "HalfCourt def – Left",      "pts": 3, "cx": 33.0, "cy": 10.0},
    19: {"name": "Right Wing – MidRange def", "pts": 2, "cx": 70.0, "cy": 60.0},
    20: {"name": "Left Wing – MidRange def",  "pts": 2, "cx": 30.0, "cy": 60.0},
    21: {"name": "Right Wing – 3PT def",      "pts": 3, "cx": 84.0, "cy": 58.0},
    22: {"name": "Left Wing – 3PT def",       "pts": 3, "cx": 16.0, "cy": 58.0},
    23: {"name": "Corner Right – 3PT def",    "pts": 3, "cx": 93.0, "cy": 80.0},
    24: {"name": "Corner Left – 3PT def",     "pts": 3, "cx":  7.0, "cy": 80.0},
    25: {"name": "Baseline Right def",        "pts": 2, "cx": 74.0, "cy": 76.0},
    26: {"name": "Baseline Left def",         "pts": 2, "cx": 26.0, "cy": 76.0},
    27: {"name": "Paint – Midrange def",      "pts": 2, "cx": 50.0, "cy": 60.0},
    28: {"name": "Paint – Pod Koszem def",    "pts": 2, "cx": 50.0, "cy": 83.0},
}

# Kształty stref — punkty polygonu w przestrzeni 0-100 (viewBox boiska)
# Koszyk na dole (y≈90), linia połowy na górze (y≈3)
ZONE_POLYGONS = {
    # --- 2PT ---
    1:  "37,72 63,72 63,97 37,97",       # Paint – Pod Koszem
    2:  "63,72 85,72 85,97 63,97",       # Baseline Right MR
    3:  "15,72 37,72 37,97 15,97",       # Baseline Left MR
    4:  "37,58 63,58 63,72 37,72",       # Paint – Midrange
    5:  "63,45 85,45 85,72 63,72",       # Right Wing MR
    6:  "15,45 37,45 37,72 15,72",       # Left Wing MR
    # --- 3PT ---
    7:  "85,58 97,58 97,97 85,97",       # Corner Right 3PT
    8:  "3,58  15,58 15,97 3,97",        # Corner Left 3PT
    9:  "85,35 97,35 97,58 85,58",       # Right Wing 3PT
    10: "3,35  15,35 15,58 3,58",        # Left Wing 3PT
    11: "63,25 97,25 97,45 63,45",       # Deep Right 3PT
    12: "3,25  37,25 37,45 3,45",        # Deep Left 3PT
    13: "37,25 63,25 63,45 37,45",       # TOP 3PT
    14: "37,12 63,12 63,25 37,25",       # Deep TOP 3PT
    15: "50,3  97,3  97,12 50,12",       # HalfCourt Right
    16: "3,3   50,3  50,12 3,12",        # HalfCourt Left
}


def _zone_fill_color(pct, att):
    """Zwraca (fill, stroke, text_color) dla wypełnionej strefy.
    Progi: 90%+ zielony | 70-89% jasno-zielony | 50-69% żółty | 35-49% jasno-czerwony | <35% czerwony
    """
    if att == 0:
        return "rgba(180,195,220,0.20)", "rgba(140,155,180,0.35)", "#999"
    if pct >= 0.90:
        return "#1b8a4e", "#126336", "#fff"
    elif pct >= 0.70:
        return "#6ab04c", "#4a8034", "#fff"
    elif pct >= 0.50:
        return "#f6c90e", "#c9a200", "#222"
    elif pct >= 0.35:
        return "#f47b20", "#c05a10", "#fff"
    else:
        return "#d9263a", "#a81a2b", "#fff"


def _zone_box_style(pct, att):
    """Zwraca (fill, stroke, text_color) dla białej ramki statystyk (używane w obronie)."""
    if att == 0:
        return "rgba(255,255,255,0.55)", "#ccc", "#aaa"
    if pct >= 0.60:
        return "rgba(220,242,220,0.95)", "#2e7d32", "#1b5e20"
    elif pct >= 0.50:
        return "rgba(232,245,233,0.95)", "#43a047", "#2e7d32"
    elif pct >= 0.40:
        return "rgba(255,253,231,0.95)", "#f9a825", "#b07800"
    elif pct >= 0.30:
        return "rgba(255,243,224,0.95)", "#ef6c00", "#bf360c"
    else:
        return "rgba(255,235,238,0.95)", "#e53935", "#b71c1c"


def _build_full_court_svg(zone_data):
    """SVG boiska NBA (połowa, kosz na górze). Skala 10px = 1 stopa.
    ViewBox: 0 0 520 340.
    Baseline: y=20. Kosz: (260,73). Klucz: x=180-340, FT linia y=222.
    Łuk 3PT: r=237.5px, narożniki x=40/480 → y=160. Kosz: 5'3" od linii końcowej.
    """
    # ── Strefy ataku (polygony w px, viewBox 520×340) ─────────────────────────
    # Tuple: (polygon_points, label_x, label_y, name, pts_type)
    ATK_ZONES = {
        # ── 2PT ──
        1:  ("180,20 340,20 340,113 180,113",      260,  67, "Paint – Pod Koszem",      2),
        2:  ("340,20 480,20 480,113 340,113",       410,  67, "Baseline Right – MR",     2),
        3:  ("40,20 180,20 180,113 40,113",         110,  67, "Baseline Left – MR",      2),
        4:  ("180,113 340,113 340,222 180,222",     260, 168, "Paint – Midrange",         2),
        5:  ("340,113 480,113 480,222 340,222",     410, 168, "Right Wing – MR",          2),
        6:  ("40,113 180,113 180,222 40,222",       110, 168, "Left Wing – MR",           2),
        # ── 3PT ──
        7:  ("480,20 510,20 510,160 480,160",       495,  90, "Corner Right – 3PT",      3),
        8:  ("10,20 40,20 40,160 10,160",            25,  90, "Corner Left – 3PT",       3),
        9:  ("480,160 510,160 510,222 480,222",     495, 191, "Right Wing – 3PT",         3),
        10: ("10,160 40,160 40,222 10,222",          25, 191, "Left Wing – 3PT",          3),
        11: ("340,222 510,222 510,295 340,295",     425, 259, "Deep Right – 3PT",         3),
        12: ("10,222 180,222 180,295 10,295",        95, 259, "Deep Left – 3PT",          3),
        13: ("180,222 340,222 340,295 180,295",     260, 259, "TOP – 3PT",                3),
        14: ("180,295 340,295 340,322 180,322",     260, 309, "Deep TOP – 3PT",           3),
        15: ("260,322 510,322 510,338 260,338",     385, 330, "HalfCourt Right",          3),
        16: ("10,322 260,322 260,338 10,338",       135, 330, "HalfCourt Left",           3),
    }

    zone_els = []
    for z, (pts, lx, ly, name, _) in ATK_ZONES.items():
        zd   = zone_data.get(z, {"made": 0, "att": 0})
        made = int(zd.get("made", 0) or 0)
        att  = int(zd.get("att",  0) or 0)
        pct  = made / att if att > 0 else None

        fill, stroke_c, tcol = _zone_fill_color(pct if pct is not None else 0, att)
        pct_txt = f"{pct*100:.0f}%" if pct is not None else "—"
        ma_txt  = f"{made}/{att}"

        # Wąskie strefy narożnikowe — mniejszy font
        narrow = z in (7, 8, 9, 10)
        fs1 = "7" if narrow else "11"
        fs2 = "6" if narrow else "9.5"

        tooltip = f'<title>{name}: {made}/{att} ({pct_txt})</title>'
        poly    = f'<polygon points="{pts}" fill="{fill}" stroke="{stroke_c}" stroke-width="0.7"/>'

        if att > 0:
            texts = (
                f'<text x="{lx}" y="{ly-5}" text-anchor="middle" dominant-baseline="middle"'
                f' font-size="{fs1}" font-weight="800" font-family="Arial,sans-serif"'
                f' fill="{tcol}">{ma_txt}</text>'
                f'<text x="{lx}" y="{ly+6}" text-anchor="middle" dominant-baseline="middle"'
                f' font-size="{fs2}" font-weight="600" font-family="Arial,sans-serif"'
                f' fill="{tcol}">{pct_txt}</text>'
            )
        else:
            texts = (
                f'<text x="{lx}" y="{ly}" text-anchor="middle" dominant-baseline="middle"'
                f' font-size="9" font-weight="400" font-family="Arial,sans-serif"'
                f' fill="rgba(100,110,130,0.6)">{z}</text>'
            )

        zone_els.append(f'<g class="zg">{tooltip}{poly}{texts}</g>')

    zones_html = "".join(zone_els)

    # ── Linie boiska NBA (10px/ft) ─────────────────────────────────────────────
    # Baseline y=20 | Kosz (260,73) | Klucz x=180-340 | FT y=222
    # Narożnik 3PT x=40/480 do y=160 | Łuk 3PT r=237.5 od (40,160) do (480,160)
    court_lines = """
  <!-- Obramowanie boiska -->
  <rect x="10" y="20" width="500" height="318" fill="none" stroke="#555" stroke-width="2"/>
  <!-- Linia końcowa (baseline) — pogrubiona -->
  <line x1="10" y1="20" x2="510" y2="20" stroke="#333" stroke-width="2.5"/>

  <!-- Klucz (lane/key) — 16' szer. × 20.25' gł. -->
  <rect x="180" y="20" width="160" height="202" fill="none" stroke="#444" stroke-width="1.8"/>

  <!-- Linia rzutów wolnych (FT line) — 15' od kosza -->
  <line x1="180" y1="222" x2="340" y2="222" stroke="#444" stroke-width="1.8"/>

  <!-- Koło RW — dolna półkola (w kluczu, przerywana) -->
  <path d="M 200,222 A 60,60 0 0,0 320,222"
        fill="none" stroke="#444" stroke-width="1.5" stroke-dasharray="7,5"/>
  <!-- Koło RW — górna półkola (poza kluczem, ciągła) -->
  <path d="M 200,222 A 60,60 0 0,1 320,222"
        fill="none" stroke="#444" stroke-width="1.8"/>

  <!-- Strefa ograniczona (restricted area) — 4' od kosza -->
  <path d="M 220,73 A 40,40 0 0,1 300,73"
        fill="none" stroke="#444" stroke-width="1.5"/>

  <!-- 3PT — narożniki (proste), 22' od kosza, 14' od baseline -->
  <line x1="40"  y1="20" x2="40"  y2="160" stroke="#444" stroke-width="1.8"/>
  <line x1="480" y1="20" x2="480" y2="160" stroke="#444" stroke-width="1.8"/>

  <!-- 3PT — łuk: środek=(260,73), r=237.5, od (40,160) przez (260,311) do (480,160) -->
  <path d="M 40,160 A 237.5,237.5 0 1,0 480,160"
        fill="none" stroke="#444" stroke-width="2"/>

  <!-- Hash marki — lewa strona klucza -->
  <line x1="168" y1="95"  x2="180" y2="95"  stroke="#444" stroke-width="1.2"/>
  <line x1="168" y1="133" x2="180" y2="133" stroke="#444" stroke-width="1.2"/>
  <line x1="168" y1="168" x2="180" y2="168" stroke="#444" stroke-width="1.2"/>
  <line x1="168" y1="205" x2="180" y2="205" stroke="#444" stroke-width="1.2"/>
  <!-- Hash marki — prawa strona klucza -->
  <line x1="340" y1="95"  x2="352" y2="95"  stroke="#444" stroke-width="1.2"/>
  <line x1="340" y1="133" x2="352" y2="133" stroke="#444" stroke-width="1.2"/>
  <line x1="340" y1="168" x2="352" y2="168" stroke="#444" stroke-width="1.2"/>
  <line x1="340" y1="205" x2="352" y2="205" stroke="#444" stroke-width="1.2"/>

  <!-- Tablica (backboard) — 4' od baseline, 6' szeroka -->
  <line x1="230" y1="59" x2="290" y2="59" stroke="#222" stroke-width="2.8"/>
  <!-- Kosz (hoop) — 5'3" od baseline, obwódka pomarańczowa -->
  <circle cx="260" cy="73" r="9" fill="none" stroke="#e07000" stroke-width="2.2"/>
  <!-- Siatka (symboliczna) -->
  <line x1="256" y1="82" x2="255" y2="90" stroke="rgba(0,0,0,0.25)" stroke-width="0.8"/>
  <line x1="260" y1="82" x2="260" y2="91" stroke="rgba(0,0,0,0.25)" stroke-width="0.8"/>
  <line x1="264" y1="82" x2="265" y2="90" stroke="rgba(0,0,0,0.25)" stroke-width="0.8"/>"""

    return (
        '<svg viewBox="0 0 520 340" xmlns="http://www.w3.org/2000/svg"'
        ' style="width:100%;display:block;border-radius:10px;box-shadow:0 2px 12px rgba(0,0,0,.13)">'
        '<style>'
        '.zg polygon{cursor:pointer;transition:opacity .18s,filter .18s;opacity:.78}'
        '.zg:hover polygon{opacity:1;filter:brightness(1.08) drop-shadow(0 0 3px rgba(0,0,0,.45))}'
        '.zg text{pointer-events:none}'
        '</style>'
        '<rect width="520" height="340" fill="#f5f0e8" rx="10"/>'
        + zones_html
        + court_lines
        + '</svg>'
    )


def _build_court_svg(zone_data, court_type="attack"):
    """Generuje SVG overlay — kolorowe strefy z hover i statystykami.
    zone_data: {zone_nr: {"made": int, "att": int}}
    Używa viewBox 0 0 100 100 (procenty obrazka).
    """
    zone_range = range(1, 17) if court_type == "attack" else range(17, 29)
    elements = []

    css = (
        '<style>'
        '.zg polygon{cursor:pointer;transition:opacity .18s,filter .18s;opacity:.72}'
        '.zg:hover polygon{opacity:1;filter:brightness(1.1) drop-shadow(0 0 1.5px rgba(0,0,0,.55))}'
        '.zg text{pointer-events:none;font-family:Arial,sans-serif}'
        '</style>'
    )

    for z in zone_range:
        zd   = zone_data.get(z, {"made": 0, "att": 0})
        made = int(zd.get("made", 0) or 0)
        att  = int(zd.get("att",  0) or 0)
        pct  = made / att if att > 0 else None
        meta = ZONE_META.get(z, {"cx": 50.0, "cy": 50.0, "name": f"Z{z}"})
        cx   = meta["cx"]; cy = meta["cy"]

        pts_str = ZONE_POLYGONS.get(z, "")
        if not pts_str:
            continue

        fill, stroke, tcol = _zone_fill_color(pct if pct is not None else 0, att)
        pct_txt = f"{pct*100:.0f}%" if pct is not None else "—"
        ma_txt  = f"{made}/{att}"

        tooltip = f'<title>{meta["name"]}: {made}/{att} ({pct_txt})</title>'
        poly    = f'<polygon points="{pts_str}" fill="{fill}" stroke="{stroke}" stroke-width="0.5"/>'

        texts = ""
        if att > 0:
            texts = (
                f'<text x="{cx}" y="{cy - 2.2}" text-anchor="middle" '
                f'dominant-baseline="middle" font-size="3.8" font-weight="800" fill="{tcol}">{ma_txt}</text>'
                f'<text x="{cx}" y="{cy + 2.4}" text-anchor="middle" '
                f'dominant-baseline="middle" font-size="3.2" font-weight="600" fill="{tcol}">{pct_txt}</text>'
            )
        else:
            # strefa bez danych — delikatna etykieta numeru
            texts = (
                f'<text x="{cx}" y="{cy}" text-anchor="middle" '
                f'dominant-baseline="middle" font-size="3.0" font-weight="400" fill="#aab">{z}</text>'
            )

        elements.append(f'<g class="zg">{tooltip}{poly}{texts}</g>')

    return (
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100"'
        ' preserveAspectRatio="none"'
        ' style="position:absolute;top:0;left:0;width:100%;height:100%">'
        + css + "".join(elements) + "</svg>"
    )


@app.route("/mecz/<int:match_id>/shooting-chart")
@login_required
def shooting_chart(match_id):
    db = get_db()
    cur = db.cursor()

    # Mecz
    cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
    m = cur.fetchone()
    if not m:
        cur.close()
        return "Mecz nie znaleziony", 404

    gtk_name = get_setting("gtk_name") or "GTK"
    opp_name = m["przeciwnik"] or "OPP"

    # Zawodnicy GTK z tego meczu
    cur.execute("""
        SELECT DISTINCT nr FROM player_stats
        WHERE match_id=%s AND druzyna='gtk' AND nr > 0
        ORDER BY nr
    """, (match_id,))
    gtk_players = [r["nr"] for r in cur.fetchall()]

    # Wybrani zawodnicy (GET params: p1..p5)
    selected = []
    for i in range(1, 6):
        v = request.args.get(f"p{i}", "")
        if v and v.isdigit() and int(v) in gtk_players:
            selected.append(int(v))
    # domyślnie: pierwszych 1 zawodnik
    if not selected and gtk_players:
        selected = [gtk_players[0]]
    selected = selected[:5]

    # Shot zone data per zawodnik dla GTK (atak) i OPP (obrona)
    def load_zones(druzyna, nrs):
        if not nrs:
            return {}
        placeholders = ",".join(["%s"] * len(nrs))
        cur.execute(f"""
            SELECT nr, zone, SUM(made) AS made, SUM(att) AS att
            FROM shot_zones
            WHERE match_id=%s AND druzyna=%s AND nr IN ({placeholders})
            GROUP BY nr, zone
        """, [match_id, druzyna] + list(nrs))
        result = {}
        for row in cur.fetchall():
            n = row["nr"]
            if n not in result:
                result[n] = {}
            result[n][row["zone"]] = {"made": row["made"], "att": row["att"]}
        return result

    # Dane per zawodnik
    atk_data = load_zones("gtk", selected)
    # Agregat drużyny (nr=0 lub sumuj wszystkich)
    cur.execute("""
        SELECT zone, SUM(made) AS made, SUM(att) AS att
        FROM shot_zones WHERE match_id=%s AND druzyna='gtk'
        GROUP BY zone
    """, (match_id,))
    team_atk = {r["zone"]: {"made": r["made"], "att": r["att"]} for r in cur.fetchall()}

    # OPP atak (defense view)
    cur.execute("""
        SELECT zone, SUM(made) AS made, SUM(att) AS att
        FROM shot_zones WHERE match_id=%s AND druzyna='opp'
        GROUP BY zone
    """, (match_id,))
    team_def = {r["zone"]: {"made": r["made"], "att": r["att"]} for r in cur.fetchall()}

    cur.close()

    # Widok: tylko atak
    view_mode = "attack"

    # Buduj panele zawodników
    def player_label(nr):
        return f"#{nr}"

    def zone_summary_table(zone_data, court_type):
        zone_range = range(1, 17) if court_type == "attack" else range(17, 29)
        rows = []
        for z in zone_range:
            zd = zone_data.get(z, {"made": 0, "att": 0})
            made = zd["made"]; att = zd["att"]
            pct = f"{made/att*100:.1f}%" if att > 0 else "—"
            name = ZONE_META.get(z, {}).get("name", f"Z{z}")
            pts_type = "3PT" if ZONE_META.get(z, {}).get("pts", 2) == 3 else "2PT"
            row_style = ""
            if att > 0:
                p = made / att
                if p >= 0.90:   row_style = "background:rgba(27,138,78,.12)"
                elif p >= 0.70: row_style = "background:rgba(106,176,76,.12)"
                elif p >= 0.50: row_style = "background:rgba(246,201,14,.12)"
                elif p >= 0.35: row_style = "background:rgba(244,123,32,.12)"
                else:           row_style = "background:rgba(217,38,58,.10)"
            rows.append(f"""<tr style="{row_style}">
              <td class="text-muted small">{z}</td>
              <td>{name}</td>
              <td class="text-center"><span class="badge bg-secondary">{pts_type}</span></td>
              <td class="text-center fw-bold">{made}/{att}</td>
              <td class="text-center">{pct}</td>
            </tr>""")
        return "".join(rows)

    # Wybór court_type
    if view_mode == "defense":
        court_img = "/static/img/court_defense.png"
        court_type = "defense"
        display_data = team_def
        title_suffix = f"{opp_name} — strefy rzutów (obrona {gtk_name})"
    else:
        court_img = "/static/img/court_attack.png"
        court_type = "attack"
        title_suffix = f"{gtk_name} — strefy rzutów (atak)"
        if len(selected) == 1:
            display_data = atk_data.get(selected[0], {})
        else:
            # Sumuj wszystkich wybranych
            display_data = {}
            for nr in selected:
                for z, zd in atk_data.get(nr, {}).items():
                    if z not in display_data:
                        display_data[z] = {"made": 0, "att": 0}
                    display_data[z]["made"] += zd["made"]
                    display_data[z]["att"]  += zd["att"]

    court_svg = _build_full_court_svg(display_data)
    table_rows = zone_summary_table(display_data, "attack")

    # Player selector checkboxes
    def player_checks():
        html = ""
        for nr in gtk_players:
            checked = "checked" if nr in selected else ""
            html += f"""<div class="form-check form-check-inline">
              <input class="form-check-input player-check" type="checkbox" value="{nr}" id="pc{nr}" {checked}>
              <label class="form-check-label" for="pc{nr}">#{nr}</label>
            </div>"""
        return html

    # Totals
    total_atk_made = sum(v["made"] for v in team_atk.values())
    total_atk_att  = sum(v["att"]  for v in team_atk.values())
    total_def_made = sum(v["made"] for v in team_def.values())
    total_def_att  = sum(v["att"]  for v in team_def.values())
    total_atk_pct  = f"{total_atk_made/total_atk_att*100:.1f}%" if total_atk_att else "—"
    total_def_pct  = f"{total_def_made/total_def_att*100:.1f}%" if total_def_att else "—"

    # Legenda kolor → opis (nowe progi)
    legend_items = [
        ("#1b8a4e", "#fff", "≥ 90%"),
        ("#6ab04c", "#fff", "70 – 89%"),
        ("#f6c90e", "#222", "50 – 69%"),
        ("#f47b20", "#fff", "35 – 49%"),
        ("#d9263a", "#fff", "< 35%"),
    ]
    legend_html = "".join(
        f'<span style="display:inline-flex;align-items:center;gap:5px;font-size:11px;color:#444">'
        f'<span style="width:28px;height:16px;background:{bg};border-radius:4px;display:inline-block;'
        f'box-shadow:0 1px 3px rgba(0,0,0,.2)"></span>'
        f'<span style="color:{tc}">{label}</span></span>'
        for bg, tc, label in legend_items
    )

    html = f"""<!doctype html><html lang="pl"><head>
<meta charset="utf-8">
<title>Shooting Chart — {m['przeciwnik']}</title>
{CSS}
<style>
.court-wrap{{display:block;width:100%}}
@media print{{
  .no-print{{display:none!important}}
  @page{{margin:8mm;size:A4 landscape}}
}}
</style>
</head><body>
{nav("home")}
<div class="main-content">

<!-- Nagłówek -->
<div class="page-header d-flex align-items-center gap-3 mb-3 no-print">
  <a href="/mecz/{match_id}" class="btn btn-sm btn-outline-secondary">&larr; Mecz</a>
  <h4 class="mb-0">Shooting Chart</h4>
  <span class="text-muted" style="font-size:13px">{m.get('data_meczu','') or ''} &middot; vs {opp_name}</span>
  <div class="ms-auto">
    <span style="background:#1a2b4a;color:#fff;padding:5px 14px;border-radius:20px;font-size:12px;font-weight:600">Atak</span>
  </div>
</div>

<!-- KPI pasek -->
<div class="d-flex gap-2 mb-3 flex-wrap">
  <div style="background:#f4f6fb;border-radius:8px;padding:8px 18px;text-align:center;min-width:110px">
    <div style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:.4px">FG% (atak)</div>
    <div style="font-size:22px;font-weight:700;color:#1a6b3c">{total_atk_pct}</div>
    <div style="font-size:11px;color:#aaa">{total_atk_made}/{total_atk_att}</div>
  </div>
</div>

<!-- Wybór zawodników (tylko atak) -->
<div id="player-selector" class="mb-3 no-print" {'style="display:none"' if view_mode=='defense' else ''}>
  <div style="font-size:12px;font-weight:600;color:#555;margin-bottom:6px">Zawodnicy GTK (maks. 5):</div>
  <div style="display:flex;flex-wrap:wrap;gap:6px;align-items:center">
    {player_checks()}
    <button class="btn btn-sm btn-primary" onclick="applyPlayers()">Pokaż</button>
  </div>
</div>

<!-- Główny układ: boisko (szersze) + tabela -->
<div class="row g-3 align-items-start">

  <!-- Boisko -->
  <div class="col-lg-7 col-md-6">
    <div style="font-size:11px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.4px;margin-bottom:6px">{title_suffix}</div>
    <div class="court-wrap">
      {court_svg}
    </div>
    <!-- Legenda -->
    <div class="d-flex gap-2 flex-wrap mt-2 no-print" style="margin-top:10px">
      {legend_html}
    </div>
  </div>

  <!-- Tabela stref -->
  <div class="col-lg-5 col-md-6">
    <div style="font-size:11px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.4px;margin-bottom:6px">Statystyki per strefa</div>
    <div class="card" style="border:0.5px solid #e3e8f0">
      <div class="card-body p-0">
        <table class="table table-sm mb-0" style="font-size:12px">
          <thead>
            <tr style="background:#1a2b4a;color:#fff">
              <th style="width:24px;padding:7px 8px">#</th>
              <th style="padding:7px 8px">Strefa</th>
              <th class="text-center" style="padding:7px 8px">Typ</th>
              <th class="text-center" style="padding:7px 8px">M/A</th>
              <th class="text-center" style="padding:7px 8px">FG%</th>
            </tr>
          </thead>
          <tbody>
            {table_rows}
          </tbody>
        </table>
      </div>
    </div>
  </div>

</div>
</div><!-- /main-content -->

<script>
function applyPlayers() {{
  var checked = Array.from(document.querySelectorAll('.player-check:checked')).map(c=>c.value).slice(0,5);
  if (!checked.length) return;
  var url = new URL(window.location.href);
  ['p1','p2','p3','p4','p5'].forEach(k=>url.searchParams.delete(k));
  checked.forEach((v,i)=>url.searchParams.set('p'+(i+1), v));
  window.location.href = url.toString();
}}
</script>
</body></html>"""

    return html


# ══════════════════════════════════════════════════════════════════════════════
# STATYSTYKI SEZONU
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/sezon")
@login_required
def sezon():
    # Sezon i kontekst WYŁĄCZNIE z sidebara — brak override przez URL
    ctx_klub    = get_setting("current_klub") or ""
    ctx_sezon   = get_setting("current_season") or ""
    ctx_druzyna = get_setting("current_druzyna") or ""
    sezon_filter   = ctx_sezon
    team_id_filter = ""
    db = get_db(); cur = db.cursor()

    # Pobierz team_id z kontekstu sidebara
    if ctx_klub and ctx_sezon and ctx_druzyna:
        try:
            cur.execute("""
                SELECT t.id FROM teams t
                JOIN seasons s ON t.season_id=s.id
                JOIN clubs c ON s.club_id=c.id
                WHERE c.name=%s AND s.name=%s AND t.name=%s
            """, (ctx_klub, ctx_sezon, ctx_druzyna))
            tr = cur.fetchone()
            if tr: team_id_filter = str(tr["id"])
        except: pass

    # Sezony dla aktywnego klubu z bazy seasons (spójne z kontekstem)
    sezony_db = []
    if ctx_klub:
        try:
            cur.execute("""
                SELECT s.name FROM seasons s
                JOIN clubs c ON s.club_id=c.id
                WHERE c.name=%s ORDER BY s.name DESC
            """, (ctx_klub,))
            sezony_db = [r["name"] for r in cur.fetchall()]
        except: pass
    if not sezony_db:
        cur.execute("SELECT DISTINCT sezon FROM matches ORDER BY sezon DESC")
        sezony_db = [r["sezon"] for r in cur.fetchall()]
    sezony = sezony_db

    # Filtruj liczbę meczów po sezon + opcjonalnie team_id
    if team_id_filter:
        cur.execute("SELECT COUNT(*) as cnt FROM matches WHERE sezon=%s AND team_id=%s",
                    (sezon_filter, int(team_id_filter)))
    else:
        cur.execute("SELECT COUNT(*) as cnt FROM matches WHERE sezon=%s", (sezon_filter,))
    n_matches = cur.fetchone()["cnt"]

    if n_matches == 0:
        cur.close()
        content = f"""
<div class="page-title">&#128202; Statystyki sezonu</div>
<div class="card p-3" style="border:2px dashed #e2e8f0;background:#f8faff;text-align:center;padding:3rem 2rem">
  <div style="font-size:3rem;margin-bottom:1rem">🏀</div>
  <div style="font-size:1.1rem;font-weight:700;color:#1a2b4a;margin-bottom:.5rem">Brak danych do wyświetlenia</div>
  <div style="font-size:.85rem;color:#888;margin-bottom:1.5rem;max-width:400px;margin-left:auto;margin-right:auto;line-height:1.6">
    Nie znaleziono meczów dla wybranego kontekstu.<br>
    Wgraj pliki <b>zapis.xlsx</b> aby zobaczyć statystyki drużyny.
  </div>
  <a href="/upload" style="display:inline-block;padding:.5rem 1.4rem;background:#1a2b4a;color:#fff;border-radius:8px;text-decoration:none;font-size:.85rem;font-weight:600">
    ➕ Wgraj mecz
  </a>
</div>"""
        return html_response(base(content, active="season"))

    # Buduj filtr WHERE dla agregatów
    _team_cond = " AND m.team_id=%s" if team_id_filter else ""
    _team_params_gtk = [sezon_filter] + ([int(team_id_filter)] if team_id_filter else [])
    _team_params_opp = [sezon_filter] + ([int(team_id_filter)] if team_id_filter else [])

    # Agregaty GTK
    cur.execute(f"""
        SELECT
            SUM(pts) as pts, SUM(poss) as poss,
            SUM(p2m) as p2m, SUM(p2a) as p2a,
            SUM(p3m) as p3m, SUM(p3a) as p3a,
            SUM(ftm) as ftm, SUM(fta) as fta,
            SUM(br)  as br,  SUM(fd)  as fd,
            COALESCE(SUM(ast),0)  as ast,
            COALESCE(SUM(oreb),0) as oreb,
            COALESCE(SUM(dreb),0) as dreb,
            COALESCE(SUM(stl),0)  as stl,
            COALESCE(SUM(blk),0)  as blk
        FROM match_stats ms
        JOIN matches m ON ms.match_id = m.id
        WHERE m.sezon=%s AND ms.druzyna='gtk'{_team_cond}
    """, _team_params_gtk)
    gtk_tot = dict(cur.fetchone())

    # Agregaty OPP
    cur.execute(f"""
        SELECT
            SUM(pts) as pts, SUM(poss) as poss,
            SUM(p2m) as p2m, SUM(p2a) as p2a,
            SUM(p3m) as p3m, SUM(p3a) as p3a,
            SUM(ftm) as ftm, SUM(fta) as fta,
            SUM(br)  as br,  SUM(fd)  as fd,
            COALESCE(SUM(ast),0)  as ast,
            COALESCE(SUM(oreb),0) as oreb,
            COALESCE(SUM(dreb),0) as dreb,
            COALESCE(SUM(stl),0)  as stl,
            COALESCE(SUM(blk),0)  as blk
        FROM match_stats ms
        JOIN matches m ON ms.match_id = m.id
        WHERE m.sezon=%s AND ms.druzyna='opp'{_team_cond}
    """, _team_params_opp)
    opp_tot = dict(cur.fetchone())

    # Agregaty player_stats — STL, BLK, AST, OREB, DREB
    try:
        cur.execute(f"""
            SELECT
                SUM(ps.stl)  as stl,  SUM(ps.blk)  as blk,
                SUM(ps.ast)  as ast,  SUM(ps.oreb) as oreb,
                SUM(ps.dreb) as dreb
            FROM player_stats ps
            JOIN matches m ON ps.match_id=m.id
            WHERE m.sezon=%s AND ps.druzyna='gtk'{_team_cond}
        """, _team_params_gtk)
        _pl = cur.fetchone()
        gtk_pl = dict(_pl) if _pl else {}
    except Exception:
        gtk_pl = {}

    try:
        cur.execute(f"""
            SELECT
                SUM(ps.stl)  as stl,  SUM(ps.blk)  as blk,
                SUM(ps.ast)  as ast,  SUM(ps.oreb) as oreb,
                SUM(ps.dreb) as dreb
            FROM player_stats ps
            JOIN matches m ON ps.match_id=m.id
            WHERE m.sezon=%s AND ps.druzyna='opp'{_team_cond}
        """, _team_params_opp)
        _pl = cur.fetchone()
        opp_pl = dict(_pl) if _pl else {}
    except Exception:
        opp_pl = {}

    def avg_pl(d, k):
        v = d.get(k) or 0
        return f"{v/n_matches:.1f}" if n_matches else "—"

    # Wyniki meczów
    if team_id_filter:
        cur.execute("SELECT wynik_gtk, wynik_opp FROM matches WHERE sezon=%s AND team_id=%s",
                    (sezon_filter, int(team_id_filter)))
    else:
        cur.execute("SELECT wynik_gtk, wynik_opp FROM matches WHERE sezon=%s", (sezon_filter,))
    results = cur.fetchall()
    wins   = sum(1 for r in results if r["wynik_gtk"] > r["wynik_opp"])
    losses = sum(1 for r in results if r["wynik_gtk"] < r["wynik_opp"])
    draws  = n_matches - wins - losses

    # Shot timing per sezon
    cur.execute("""
        SELECT ts.druzyna, ts.bucket,
               SUM(ts.made2) as made2, SUM(ts.att2) as att2,
               SUM(ts.made3) as made3, SUM(ts.att3) as att3
        FROM timing_stats ts
        JOIN matches m ON ts.match_id=m.id
        WHERE m.sezon=%s
        GROUP BY ts.druzyna, ts.bucket
    """, (sezon_filter,))
    timing_rows = cur.fetchall()

    # Piątki sezonu — agregacja PER MECZ po nazwiskach z przypisania tego meczu
    season_lineups_off = []
    season_lineups_def = []
    season_lineups_net = []
    try:
        # ── Krok 1: Buduj mapę per mecz: match_id → {nr_str → "Nazwisko I."} ──
        match_nr_name = {}  # {match_id: {nr: name}}

        # Źródło A: players przez player_id (nowa struktura)
        try:
            cur.execute("""
                SELECT ps.match_id, ps.nr, p.imie, p.nazwisko
                FROM player_stats ps
                JOIN players p ON ps.player_id = p.id
                JOIN matches m ON ps.match_id = m.id
                WHERE m.sezon = %s AND ps.druzyna = 'gtk'
            """, (sezon_filter,))
            for row in cur.fetchall():
                mid = row["match_id"]
                if mid not in match_nr_name:
                    match_nr_name[mid] = {}
                match_nr_name[mid][str(row["nr"])] = f"{row['nazwisko']} {row['imie'][0]}."
        except Exception:
            pass

        # Źródło B: roster przez roster_id (stara struktura) — uzupełnia brakujące
        try:
            cur.execute("""
                SELECT ps.match_id, ps.nr, r.imie, r.nazwisko
                FROM player_stats ps
                JOIN roster r ON ps.roster_id = r.id
                JOIN matches m ON ps.match_id = m.id
                WHERE m.sezon = %s AND ps.druzyna = 'gtk'
            """, (sezon_filter,))
            for row in cur.fetchall():
                mid = row["match_id"]
                if mid not in match_nr_name:
                    match_nr_name[mid] = {}
                nr_s = str(row["nr"])
                if nr_s not in match_nr_name[mid]:
                    match_nr_name[mid][nr_s] = f"{row['nazwisko']} {row['imie'][0]}."
        except Exception:
            pass

        # ── Krok 2: Pobierz lineupy per (mecz, lineup, druzyna) ────────────────
        cur.execute("""
            SELECT ls.match_id, ls.lineup, ls.druzyna,
                   SUM(ls.pts)  as pts,  SUM(ls.poss) as poss,
                   SUM(ls.p2m)  as p2m,  SUM(ls.p2a)  as p2a,
                   SUM(ls.p3m)  as p3m,  SUM(ls.p3a)  as p3a,
                   SUM(ls.ftm)  as ftm,  SUM(ls.fta)  as fta,
                   SUM(ls.br)   as br,   SUM(ls.fd)   as fd,
                   SUM(ls.ast)  as ast,  SUM(ls.oreb) as oreb,
                   SUM(ls.dreb) as dreb, SUM(ls.stl)  as stl,
                   SUM(ls.blk)  as blk
            FROM lineup_stats ls
            JOIN matches m ON ls.match_id = m.id
            WHERE m.sezon = %s AND ls.druzyna IN ('gtk', 'gtk_def')
            GROUP BY ls.match_id, ls.lineup, ls.druzyna
        """, (sezon_filter,))
        all_lu_rows = cur.fetchall()

        # ── Krok 3: Agreguj w Pythonie po kluczu = posortowane nazwiska z TEGO meczu
        _FIELDS = ["pts","poss","p2m","p2a","p3m","p3a","ftm","fta","br","fd","ast","oreb","dreb","stl","blk"]
        agg_off = {}  # {(name,...): {label, pts, poss, ..., match_ids: set}}
        agg_def = {}

        for row in all_lu_rows:
            mid = row["match_id"]
            nr_name = match_nr_name.get(mid, {})
            nrs = row["lineup"].split("-")
            names = []
            for nr in nrs:
                name = nr_name.get(nr, f"#{nr}")
                names.append(name)
            key = tuple(sorted(names))          # klucz = posortowane nazwiska z tego meczu
            label = " · ".join(sorted(names))   # etykieta = to samo
            target = agg_off if row["druzyna"] == "gtk" else agg_def
            if key not in target:
                target[key] = {"label": label, "match_ids": set()}
                for f in _FIELDS:
                    target[key][f] = 0
            for f in _FIELDS:
                target[key][f] += int(row[f] or 0)
            target[key]["match_ids"].add(mid)

        # Przelicz n_games z setu meczów
        for target in [agg_off, agg_def]:
            for v in target.values():
                v["n_games"] = len(v["match_ids"])
                del v["match_ids"]

        season_lineups_off = sorted(agg_off.values(), key=lambda x: x["poss"], reverse=True)
        season_lineups_def = sorted(agg_def.values(), key=lambda x: x["poss"], reverse=True)

        # ── Krok 4: Oblicz ORtg / DRtg / Net RTG ──────────────────────────────
        s_off_map = {lu["label"]: lu["pts"]*100/lu["poss"]
                     for lu in season_lineups_off if lu["poss"] > 0}
        s_def_map = {lu["label"]: lu["pts"]*100/lu["poss"]
                     for lu in season_lineups_def if lu["poss"] > 0}

        for lu in season_lineups_off:
            lbl = lu["label"]
            lu["ortg"] = round(s_off_map[lbl], 1) if lbl in s_off_map else None
            lu["drtg"] = round(s_def_map[lbl], 1) if lbl in s_def_map else None
            lu["net_rtg"] = round(s_off_map[lbl] - s_def_map[lbl], 1) if (lbl in s_off_map and lbl in s_def_map) else None

        for lu in season_lineups_def:
            lbl = lu["label"]
            lu["net_rtg"] = round(s_off_map[lbl] - s_def_map[lbl], 1) if (lbl in s_off_map and lbl in s_def_map) else None

        # Zbuduj NET lineups sezonu
        seen_s_net = {}
        for lu in season_lineups_off:
            lbl = lu["label"]
            if lbl not in seen_s_net:
                seen_s_net[lbl] = dict(lu)
        season_lineups_net = sorted(seen_s_net.values(),
            key=lambda x: (x["net_rtg"] is not None, x["net_rtg"] if x["net_rtg"] is not None else 0),
            reverse=True)
    except Exception:
        pass

    cur.close()

    gtk_name = get_setting("gtk_name") or "GTK"

    def avg(d, k):
        v = d.get(k) or 0
        return round(v / n_matches, 1)

    def avg_kpi(d):
        k = calc_kpi(d)
        return k

    gtk_kpi = calc_kpi(gtk_tot)
    opp_kpi = calc_kpi(opp_tot)

    # Net RTG = ORtg - DRtg
    try:
        ortg_val = float(gtk_kpi["ortg"])
        drtg_val = float(opp_kpi["ortg"])
        net_rtg_val = round(ortg_val - drtg_val, 1)
        net_rtg_str = f"+{net_rtg_val}" if net_rtg_val > 0 else str(net_rtg_val)
    except:
        net_rtg_val = 0; net_rtg_str = "—"
        ortg_val = 0; drtg_val = 0

    def cmp_row(lbl, vg, vo, higher_is_better=True, new=False):
        try:
            fg = float(str(vg).split('/')[0].replace('%','').replace('-','0'))
            fo = float(str(vo).split('/')[0].replace('%','').replace('-','0'))
            sg = "font-weight:700;color:#1a6b3c" if (higher_is_better and fg>fo) or (not higher_is_better and fg<fo) else ""
            so = "font-weight:700;color:#8b1a1a" if (higher_is_better and fo>fg) or (not higher_is_better and fo<fg) else ""
            # Pasek porównawczy
            total = abs(fg) + abs(fo)
            pct_g = int(fg/total*100) if total else 50
            bar_col = "#1D9E75" if higher_is_better else "#E24B4A"
        except:
            sg=so=""; pct_g=50; bar_col="#1D9E75"
        new_badge = ""
        bar_html = f'<div style="height:5px;border-radius:3px;background:var(--color-border-tertiary,#e0e0e0);overflow:hidden"><div style="height:100%;width:{pct_g}%;background:{bar_col};border-radius:3px"></div></div>'
        return f"<tr><td><b>{lbl}</b>{new_badge}</td><td class='text-center' style='{sg}'>{vg}</td><td class='text-center' style='{so}'>{vo}</td><td style='width:70px;padding:8px 6px'>{bar_html}</td></tr>"

    def section_row(lbl):
        return f"<tr style='background:var(--color-background-secondary,#f8f9fa)'><td colspan='4' style='font-size:10px;font-weight:600;color:#888;text-transform:uppercase;letter-spacing:.07em;padding:8px 8px 4px;border-top:0.5px solid #e0e0e0'>{lbl}</td></tr>"

    kpi_rows = (
        section_row("Atak / mecz") +
        cmp_row("PKT",        avg(gtk_tot,"pts"),        avg(opp_tot,"pts")) +
        cmp_row("AST",        avg_pl(gtk_pl,"ast"),      avg(opp_tot,"ast")) +
        cmp_row("OREB",       avg_pl(gtk_pl,"oreb"),     avg(opp_tot,"oreb")) +
        cmp_row("TO",         avg(gtk_tot,"br"),         avg(opp_tot,"br"),     False) +
        cmp_row("Faule Wymuszone", avg(gtk_tot,"fd"),    avg(opp_tot,"fd")) +
        cmp_row("Posiadania",  avg(gtk_tot,"poss"),       avg(opp_tot,"poss")) +
        cmp_row("PPP",               gtk_kpi["ppp"],            opp_kpi["ppp"]) +
        cmp_row("ORtg",              gtk_kpi["ortg"],           opp_kpi["ortg"]) +
        cmp_row("FT Rate",           gtk_kpi["ftr"],            opp_kpi["ftr"]) +
        cmp_row("2PT%",              gtk_kpi["p2_pct"],         opp_kpi["p2_pct"]) +
        cmp_row("3PT%",              gtk_kpi["p3_pct"],         opp_kpi["p3_pct"]) +
        cmp_row("FT%",               gtk_kpi["ft_pct"],         opp_kpi["ft_pct"]) +
        cmp_row("eFG%",              gtk_kpi["efg"],            opp_kpi["efg"]) +
        cmp_row("TS%",               gtk_kpi["ts"],             opp_kpi["ts"]) +
        section_row("Obrona / mecz") +
        cmp_row("DREB",       avg_pl(gtk_pl,"dreb"),     avg(opp_tot,"dreb")) +
        cmp_row("STL",        avg_pl(gtk_pl,"stl"),      avg(opp_tot,"stl")) +
        cmp_row("BLK",        avg_pl(gtk_pl,"blk"),      avg(opp_tot,"blk")) +
        cmp_row("Faule",      avg(opp_tot,"fd"),         avg(gtk_tot,"fd"),     False) +
        cmp_row("DRtg",              opp_kpi["ortg"],           gtk_kpi["ortg"],       False)
    )

    # Zsumowane statystyki sezonu (do wiersza arc-tiles)
    tot_pts = int(gtk_tot.get("pts")  or 0)
    tot_reb = int((gtk_pl.get("oreb") or 0) + (gtk_pl.get("dreb") or 0))
    tot_ast = int(gtk_pl.get("ast")   or 0)
    tot_stl = int(gtk_pl.get("stl")   or 0)
    tot_blk = int(gtk_pl.get("blk")   or 0)

    # Timing tabela
    def timing_row(bucket):
        gd = next((r for r in timing_rows if r["druzyna"]=="gtk" and r["bucket"]==bucket), {})
        od = next((r for r in timing_rows if r["druzyna"]=="opp" and r["bucket"]==bucket), {})
        gm2=int(gd.get("made2",0) or 0); ga2=int(gd.get("att2",0) or 0)
        gm3=int(gd.get("made3",0) or 0); ga3=int(gd.get("att3",0) or 0)
        om2=int(od.get("made2",0) or 0); oa2=int(od.get("att2",0) or 0)
        om3=int(od.get("made3",0) or 0); oa3=int(od.get("att3",0) or 0)
        gm=gm2+gm3; ga=ga2+ga3
        om=om2+om3; oa=oa2+oa3
        ge=f"{gm/ga:.0%}" if ga else "—"
        oe=f"{om/oa:.0%}" if oa else "—"
        # Paski
        max_att = max(
            max((int(r.get("att2",0) or 0)+int(r.get("att3",0) or 0)) for r in timing_rows if r["druzyna"]=="gtk") if timing_rows else 1,
            1
        )
        gbar = int(ga/max_att*80) if ga else 0
        obar = int(oa/max_att*80) if oa else 0
        return f"""<tr>
            <td class="fw-bold" style="font-size:.85rem;background:#fff">{bucket}</td>
            <td style="font-size:.8rem;background:#f0fff4;text-align:center">{gm}/{ga}</td>
            <td style="font-weight:700;color:#1a6b3c;background:#f0fff4;text-align:center">{ge}</td>
            <td style="font-size:.75rem;color:#555;background:#f0fff4;text-align:center">{gm2}/{ga2} | {gm3}/{ga3}</td>
            <td style="padding:6px 8px;background:#fff"><div style="height:8px;width:{gbar}px;background:#1a6b3c;border-radius:4px"></div></td>
            <td style="padding:6px 8px;background:#fff"><div style="height:8px;width:{obar}px;background:#8b1a1a;border-radius:4px"></div></td>
            <td style="font-size:.75rem;color:#555;background:#fff5f5;text-align:center">{om2}/{oa2} | {om3}/{oa3}</td>
            <td style="font-weight:700;color:#8b1a1a;background:#fff5f5;text-align:center">{oe}</td>
            <td style="font-size:.8rem;background:#fff5f5;text-align:center">{om}/{oa}</td>
        </tr>"""

    tim_rows = "".join(timing_row(b) for b in BUCKETS)

    def season_lineup_rows(lineups, def_map):
        if not lineups:
            return ""
        rows = ""
        for i, lu in enumerate(lineups):
            ng   = lu.get("n_games", 1) or 1
            def _a(k): return lu.get(k, 0) or 0
            poss = _a("poss"); pts = _a("pts")
            p2m  = _a("p2m"); p2a = _a("p2a")
            p3m  = _a("p3m"); p3a = _a("p3a")
            ftm  = _a("ftm"); fta = _a("fta")
            br   = _a("br");  fd  = _a("fd")
            ast  = _a("ast"); oreb = _a("oreb")
            dreb = _a("dreb"); stl = _a("stl"); blk = _a("blk")
            fga  = p2a + p3a
            ppp_v  = pts / poss if poss else None
            efg_v  = round((p2m + 1.5*p3m) / fga * 100) if fga else None
            p2pct  = f"{p2m/p2a:.0%}" if p2a else "—"
            p3pct  = f"{p3m/p3a:.0%}" if p3a else "—"
            ftpct  = f"{ftm/fta:.0%}" if fta else "—"
            efg_s  = f"{efg_v}%" if efg_v is not None else "—"
            ppp_s  = f"{ppp_v:.2f}" if ppp_v is not None else "—"
            ppp_c  = "#0F6E56" if ppp_v and ppp_v>=0.9 else ("#A32D2D" if ppp_v and ppp_v<0.7 else "inherit")
            efg_c  = "#0F6E56" if efg_v and efg_v>=50 else ("#A32D2D" if efg_v and efg_v<35 else "inherit")
            def _f(v): return f"{v/ng:.1f}"
            bg = "#f8f9ff" if i%2==0 else "#fff"
            br_c = "#A32D2D" if br/ng >= 4 else "inherit"

            # DEF stats dla tej piątki
            dlu  = def_map.get(lu["label"], {})
            def _da(k): return dlu.get(k, 0) or 0
            dp2m = _da("p2m"); dp2a = _da("p2a")
            dp3m = _da("p3m"); dp3a = _da("p3a")
            dposs= _da("poss"); dpts = _da("pts")
            dfga = dp2a + dp3a
            defg_v = round((dp2m + 1.5*dp3m) / dfga * 100) if dfga else None
            dppp_v = dpts / dposs if dposs else None
            defg_s = f"{defg_v}%" if defg_v is not None else "—"
            dppp_s = f"{dppp_v:.2f}" if dppp_v is not None else "—"
            defg_c = "#0F6E56" if defg_v and defg_v<35 else ("#A32D2D" if defg_v and defg_v>=50 else "inherit")
            dppp_c = "#0F6E56" if dppp_v and dppp_v<0.7 else ("#A32D2D" if dppp_v and dppp_v>=0.9 else "inherit")

            # NET RTG
            ortg_v = lu.get("ortg")
            drtg_v = lu.get("drtg")
            net_v  = lu.get("net_rtg")
            ortg_s = f"{ortg_v:.1f}" if ortg_v is not None else "—"
            drtg_s = f"{drtg_v:.1f}" if drtg_v is not None else "—"
            net_s  = f"{net_v:+.1f}" if net_v is not None else "—"
            ortg_c = "#0F6E56" if ortg_v and ortg_v>=90 else ("#A32D2D" if ortg_v and ortg_v<70 else "inherit")
            drtg_c = "#0F6E56" if drtg_v and drtg_v<70 else ("#A32D2D" if drtg_v and drtg_v>=90 else "inherit")
            net_c  = "#0F6E56" if net_v and net_v>0 else ("#A32D2D" if net_v and net_v<0 else "#888")

            rows += f"""<tr style="background:{bg}">
                <td style="font-size:.65rem;text-align:left;padding-left:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{lu["label"]}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{poss/ng:.1f}">{_f(poss)}</td>
                <td class="text-center" style="color:#633806;font-size:.72rem;white-space:nowrap" data-v="{ng}">{ng}</td>
                <td class="text-center fw-bold" style="color:#1a2b4a;white-space:nowrap" data-v="{pts/ng:.1f}">{_f(pts)}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{(p2m*100+p2a)/ng:.0f}">{p2m/ng:.1f}/{p2a/ng:.1f}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{int(p2m/p2a*100) if p2a else -1}">{p2pct}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{(p3m*100+p3a)/ng:.0f}">{p3m/ng:.1f}/{p3a/ng:.1f}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{int(p3m/p3a*100) if p3a else -1}">{p3pct}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{(ftm*100+fta)/ng:.0f}">{ftm/ng:.1f}/{fta/ng:.1f}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{int(ftm/fta*100) if fta else -1}">{ftpct}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{oreb/ng:.1f}">{_f(oreb)}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{dreb/ng:.1f}">{_f(dreb)}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{ast/ng:.1f}">{_f(ast)}</td>
                <td class="text-center" style="white-space:nowrap;color:{br_c}" data-v="{br/ng:.1f}">{_f(br)}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{stl/ng:.1f}">{_f(stl)}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{blk/ng:.1f}">{_f(blk)}</td>
                <td class="text-center" style="white-space:nowrap" data-v="{fd/ng:.1f}">{_f(fd)}</td>
                <td class="text-center" style="white-space:nowrap;color:{efg_c}" data-v="{efg_v or -1}">{efg_s}</td>
                <td class="text-center fw-bold" style="white-space:nowrap;color:{ppp_c}" data-v="{ppp_v*100 if ppp_v else -1}">{ppp_s}</td>
                <td class="text-center" style="white-space:nowrap;color:{defg_c}" data-v="{defg_v or -1}">{defg_s}</td>
                <td class="text-center fw-bold" style="white-space:nowrap;color:{dppp_c}" data-v="{dppp_v*100 if dppp_v else -1}">{dppp_s}</td>
                <td class="text-center" style="white-space:nowrap;color:{ortg_c}" data-v="{ortg_v or -999}">{ortg_s}</td>
                <td class="text-center" style="white-space:nowrap;color:{drtg_c}" data-v="{drtg_v or 999}">{drtg_s}</td>
                <td class="text-center fw-bold" style="white-space:nowrap;color:{net_c}" data-v="{net_v if net_v is not None else -999}">{net_s}</td>
            </tr>"""
        return rows

    def season_lineup_table_html():
        if not season_lineups_off:
            return '<p class="text-muted p-3 mb-0" style="font-size:.82rem">Brak danych piątek OFF w tym sezonie.</p>'
        # Zbuduj mapę label→DEF stats
        def_map = {lu["label"]: lu for lu in season_lineups_def}

        th  = 'background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 3px;text-align:center;white-space:nowrap;cursor:pointer;user-select:none'
        thl = 'background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 8px;text-align:left;white-space:nowrap'
        thg = 'background:#1a2b4a;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center'
        ths = 'background:#1a2b4a;color:rgba(255,255,255,.8);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center;cursor:pointer;user-select:none'
        thz = 'background:#152236;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.1);text-align:center'
        thzs= 'background:#152236;color:rgba(255,255,255,.75);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center;cursor:pointer;user-select:none'
        thd = 'background:#152236;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.1);text-align:center'
        thds= 'background:#152236;color:rgba(255,255,255,.75);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center;cursor:pointer;user-select:none'
        thn = 'background:#412402;color:#FAC775;font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center'
        thns= 'background:#412402;color:#FAC775;font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center;cursor:pointer;user-select:none'
        vm  = 'vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2)'
        cg = ('<colgroup>'
            '<col style="width:190px">'
            '<col style="width:38px"><col style="width:30px"><col style="width:34px">'
            '<col style="width:50px"><col style="width:30px">'
            '<col style="width:50px"><col style="width:30px">'
            '<col style="width:50px"><col style="width:30px">'
            '<col style="width:26px"><col style="width:26px">'
            '<col style="width:28px"><col style="width:28px"><col style="width:28px"><col style="width:28px"><col style="width:28px">'
            '<col style="width:34px"><col style="width:34px">'
            '<col style="width:34px"><col style="width:34px">'
            '<col style="width:38px"><col style="width:38px"><col style="width:42px">'
            '</colgroup>')
        hdr = (f'<thead>'
            f'<tr>'
            f'<th style="{thl};{vm};cursor:pointer;user-select:none" rowspan="3" onclick="sluToggleSkład()" title="Kliknij aby rozwinąć/zwinąć skład">Skład <span id="slu-expand-icon" style="font-size:9px;opacity:.6">⇔</span></th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="sluSort(this,1)">POSS ↕</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="sluSort(this,2)" title="Liczba meczów">G</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="sluSort(this,3)">PKT ↕</th>'
            f'<th style="{thg}" colspan="2">2PT</th>'
            f'<th style="{thg}" colspan="2">3PT</th>'
            f'<th style="{thg}" colspan="2">FT</th>'
            f'<th style="{thz}" colspan="2">ZB</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="sluSort(this,12)">AST ↕</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="sluSort(this,13)">TO ↕</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="sluSort(this,14)">STL ↕</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="sluSort(this,15)">BLK ↕</th>'
            f'<th style="{th};{vm}" rowspan="3" onclick="sluSort(this,16)">FD ↕</th>'
            f'<th style="{thg}" colspan="2">OFF</th>'
            f'<th style="{thd}" colspan="2">DEF</th>'
            f'<th style="{thn}" colspan="3">NET RTG</th>'
            f'</tr>'
            f'<tr>'
            f'<th style="{ths}" onclick="sluSort(this,4)">M/A ↕</th><th style="{ths}" onclick="sluSort(this,5)">% ↕</th>'
            f'<th style="{ths}" onclick="sluSort(this,6)">M/A ↕</th><th style="{ths}" onclick="sluSort(this,7)">% ↕</th>'
            f'<th style="{ths}" onclick="sluSort(this,8)">M/A ↕</th><th style="{ths}" onclick="sluSort(this,9)">% ↕</th>'
            f'<th style="{thzs}" onclick="sluSort(this,10)">A ↕</th><th style="{thzs}" onclick="sluSort(this,11)">O ↕</th>'
            f'<th style="{ths}" onclick="sluSort(this,17)">eFG% ↕</th><th style="{ths}" onclick="sluSort(this,18)">PPP ↕</th>'
            f'<th style="{thds}" onclick="sluSort(this,19)">eFG% ↕</th><th style="{thds}" onclick="sluSort(this,20)">PPP ↕</th>'
            f'<th style="{thns}" onclick="sluSort(this,21)">ORtg ↕</th><th style="{thns}" onclick="sluSort(this,22)">DRtg ↕</th><th style="{thns}" onclick="sluSort(this,23)">Net ↕</th>'
            f'</tr>'
            f'</thead>')
        js = """<script>
(function(){
  var _d={};
  var _expanded=false;
  window.sluSort=function(th,col){
    var tbl=document.getElementById('tbl-slu'); if(!tbl) return;
    var tb=tbl.querySelector('tbody');
    var rows=Array.from(tb.querySelectorAll('tr'));
    var k='slu-'+col; _d[k]=!_d[k]; var asc=_d[k];
    rows.sort(function(a,b){
      var av=parseFloat(a.children[col]&&a.children[col].dataset.v);
      var bv=parseFloat(b.children[col]&&b.children[col].dataset.v);
      if(isNaN(av))av=asc?Infinity:-Infinity;
      if(isNaN(bv))bv=asc?Infinity:-Infinity;
      return asc?av-bv:bv-av;
    });
    rows.forEach(function(r,i){r.style.background=i%2===0?'#f8f9ff':'#fff';tb.appendChild(r);});
    tbl.querySelectorAll('th').forEach(function(h){h.innerHTML=h.innerHTML.replace(/ [▲▼]/,' ↕');});
    th.innerHTML=th.innerHTML.replace(' ↕',asc?' ▲':' ▼');
  };
  window.sluToggleSkład=function(){
    var tbl=document.getElementById('tbl-slu'); if(!tbl) return;
    _expanded=!_expanded;
    tbl.querySelectorAll('tbody tr td:first-child').forEach(function(td){
      td.style.whiteSpace=_expanded?'normal':'nowrap';
      td.style.overflow=_expanded?'visible':'hidden';
      td.style.textOverflow=_expanded?'clip':'ellipsis';
    });
    var icon=document.getElementById('slu-expand-icon');
    if(icon) icon.textContent=_expanded?'⇱':'⇔';
  };
})();
</script>"""
        legend = '<div style="font-size:8px;color:#888;margin-bottom:6px">Wartości uśrednione per mecz &nbsp;·&nbsp; PPP OFF: <span style="color:#0F6E56">≥0.90 dobry</span> / <span style="color:#A32D2D">&lt;0.70 słaby</span> &nbsp;·&nbsp; PPP DEF: <span style="color:#0F6E56">&lt;0.70 dobry</span> / <span style="color:#A32D2D">≥0.90 słaby</span> &nbsp;·&nbsp; sortowanie: POSS malejąco</div>'
        return (legend +
            f'<div class="table-responsive"><table id="tbl-slu" class="table table-hover mb-0" style="table-layout:fixed;min-width:1100px">'
            f'{cg}{hdr}'
            f'<tbody>{season_lineup_rows(season_lineups_off, def_map)}</tbody></table></div>'
            + js)

    pts_per_match_gtk = [0,0,0,0]
    pts_per_match_opp = [0,0,0,0]

    # Badge kontekstu — pokazuje aktywny klub/sezon/drużynę
    if ctx_klub and ctx_sezon and ctx_druzyna:
        ctx_badge = (f'<span style="font-size:11px;background:#E6F1FB;color:#0C447C;padding:3px 10px;border-radius:12px;font-weight:500">'
                     f'{ctx_klub} &middot; {ctx_sezon} &middot; {ctx_druzyna}</span>')
    elif ctx_sezon:
        ctx_badge = f'<span style="font-size:11px;background:#E6F1FB;color:#0C447C;padding:3px 10px;border-radius:12px;font-weight:500">{ctx_sezon}</span>'
    else:
        ctx_badge = '<span style="font-size:11px;color:#aaa">Brak kontekstu — ustaw w sidebarze</span>'

    # ── Duel row helper ──────────────────────────────────────────────────
    def duel_row(lbl, vg, vo, higher_is_better=True, neutral=False):
        try:
            fg = float(str(vg).replace('%','').replace('—','0').replace('+','') or 0)
            fo = float(str(vo).replace('%','').replace('—','0').replace('+','') or 0)
            total = abs(fg) + abs(fo)
            pct_g = max(5, min(95, int(fg / total * 100))) if total else 50
            pct_o = 100 - pct_g
            gtk_wins = (higher_is_better and fg > fo) or (not higher_is_better and fg < fo)
            opp_wins = (higher_is_better and fo > fg) or (not higher_is_better and fo < fg)
            cv_g = "#1a6b3c" if gtk_wins else ("#aaa" if opp_wins else "#555")
            cv_o = "#8b1a1a" if opp_wins else "#aaa"
            fw_g = "700" if gtk_wins else "400"
            fw_o = "700" if opp_wins else "400"
            if neutral:
                bg_g = "#aaa"; bg_o = "#aaa"
            elif higher_is_better:
                bg_g = "#1D9E75"; bg_o = "#E24B4A"
            else:
                bg_g = "#1D9E75" if gtk_wins else "#E24B4A"
                bg_o = "#1D9E75" if opp_wins else "#E24B4A"
        except Exception:
            pct_g=50; pct_o=50; cv_g="#1a6b3c"; cv_o="#aaa"; fw_g=fw_o="400"
            bg_g="#1D9E75"; bg_o="#E24B4A"
        bar = (f'<div style="display:flex;align-items:center;height:10px">'
               f'<div style="height:10px;border-radius:3px 0 0 3px;background:{bg_g};width:{pct_g}%"></div>'
               f'<div style="width:1px;height:14px;background:#ccc;flex-shrink:0"></div>'
               f'<div style="height:10px;border-radius:0 3px 3px 0;background:{bg_o};width:{pct_o}%"></div>'
               f'</div>')
        return (f'<div style="display:grid;grid-template-columns:62px 1fr 82px;align-items:center;gap:8px;'
                f'padding:6px 0;border-bottom:.5px solid #f0f0f0">'
                f'<div style="font-size:.95rem;font-weight:{fw_g};color:{cv_g};text-align:right">{vg}</div>'
                f'<div><div style="font-size:.6rem;font-weight:600;color:#666;text-align:center;margin-bottom:3px">{lbl}</div>{bar}</div>'
                f'<div style="font-size:.95rem;font-weight:{fw_o};color:{cv_o};text-align:left">{vo}</div>'
                f'</div>')

    def duel_card_header(left_lbl, right_lbl):
        return (f'<div style="display:grid;grid-template-columns:62px 1fr 82px;gap:8px;'
                f'font-size:.68rem;font-weight:700;text-transform:uppercase;letter-spacing:.07em;'
                f'margin-bottom:6px">'
                f'<div style="color:#1a6b3c;text-align:right">{left_lbl}</div>'
                f'<div></div>'
                f'<div style="color:#8b1a1a;word-break:break-word;line-height:1.2">{right_lbl}</div>'
                f'</div>')

    _duel_hdr   = duel_card_header(gtk_name, "Przeciwnicy")
    _duel_hdr_s = duel_card_header(gtk_name, "Prz.")

    # Scoring duel card
    scoring_duel_html = (
        _duel_hdr +
        duel_row("PKT",          avg(gtk_tot,"pts"),       avg(opp_tot,"pts")) +
        duel_row("AST",          avg_pl(gtk_pl,"ast"),     avg(opp_tot,"ast")) +
        duel_row("OREB",         avg_pl(gtk_pl,"oreb"),    avg(opp_tot,"oreb")) +
        duel_row("TO ↓",         avg(gtk_tot,"br"),        avg(opp_tot,"br"),        False) +
        duel_row("FD",           avg(gtk_tot,"fd"),        avg(opp_tot,"fd")) +
        duel_row("PPP",          gtk_kpi["ppp"],           opp_kpi["ppp"]) +
        duel_row("POSS",         avg(gtk_tot,"poss"),      avg(opp_tot,"poss"),      True, True)
    )

    # Defense duel card
    defense_duel_html = (
        _duel_hdr_s +
        duel_row("DREB",         avg_pl(gtk_pl,"dreb"),    avg(opp_tot,"dreb")) +
        duel_row("STL",          avg_pl(gtk_pl,"stl"),     avg(opp_tot,"stl")) +
        duel_row("BLK",          avg_pl(gtk_pl,"blk"),     avg(opp_tot,"blk")) +
        duel_row("Faule ↓",      avg(opp_tot,"fd"),        avg(gtk_tot,"fd"),        False) +
        duel_row("DRtg ↓",       opp_kpi["ortg"],          gtk_kpi["ortg"],          False)
    )

    # Shooting efficiency card
    def eff_row(lbl, vg, vo):
        try:
            fg = float(str(vg).replace('%','') or 0)
            fo = float(str(vo).replace('%','') or 0)
            total = abs(fg) + abs(fo)
            pg = max(5, min(95, int(fg/total*100))) if total else 50
            po = 100 - pg
            gtk_w = fg > fo
            cv_g = "#1a6b3c" if gtk_w else "#aaa"
            cv_o = "#8b1a1a" if fo > fg else "#aaa"
            fw_g = "700" if gtk_w else "400"
            fw_o = "700" if fo > fg else "400"
        except Exception:
            pg=po=50; cv_g="#1a6b3c"; cv_o="#aaa"; fw_g=fw_o="400"
        return (f'<div style="display:grid;grid-template-columns:46px 1fr 46px;align-items:center;gap:6px;'
                f'padding:5px 0;border-bottom:.5px solid #f0f0f0">'
                f'<div style="font-size:.88rem;font-weight:{fw_g};color:{cv_g};text-align:right">{vg}</div>'
                f'<div><div style="font-size:.6rem;font-weight:600;color:#666;text-align:center;margin-bottom:2px">{lbl}</div>'
                f'<div style="display:flex;align-items:center;height:8px">'
                f'<div style="height:8px;border-radius:3px 0 0 3px;background:#1D9E75;width:{pg}%"></div>'
                f'<div style="width:1px;height:12px;background:#ccc;flex-shrink:0"></div>'
                f'<div style="height:8px;border-radius:0 3px 3px 0;background:#E24B4A;width:{po}%"></div>'
                f'</div></div>'
                f'<div style="font-size:.88rem;font-weight:{fw_o};color:{cv_o};text-align:left">{vo}</div>'
                f'</div>')

    shooting_eff_html = (
        duel_card_header(gtk_name, "Prz.") +
        eff_row("2PT%",   gtk_kpi["p2_pct"],  opp_kpi["p2_pct"]) +
        eff_row("3PT%",   gtk_kpi["p3_pct"],  opp_kpi["p3_pct"]) +
        eff_row("FT%",    gtk_kpi["ft_pct"],  opp_kpi["ft_pct"]) +
        eff_row("eFG%",   gtk_kpi["efg"],     opp_kpi["efg"]) +
        eff_row("TS%",    gtk_kpi["ts"],      opp_kpi["ts"]) +
        eff_row("FT Rate",gtk_kpi["ftr"],     opp_kpi["ftr"])
    )

    content = f"""
<div class="d-flex justify-content-between align-items-center mb-3 flex-wrap gap-2">
  <div class="page-title mb-0">📊 Statystyki drużyny</div>
  <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
    {ctx_badge}
    <a href="/ustawienia" style="font-size:11px;color:#aaa;text-decoration:none" title="Zmień kontekst w sidebarze">⚙ zmień kontekst</a>
  </div>
</div>

<div class="row g-2 mb-2">
  <!-- Bilans meczów (merged) -->
  <div class="col-12 col-md-6">
    <div class="stat-card" style="display:flex;flex-direction:row;align-items:center;justify-content:center;gap:20px;padding:.85rem 1.2rem">
      <div style="text-align:center">
        <div style="display:flex;align-items:flex-end;gap:10px">
          <div>
            <div style="font-size:2.1rem;font-weight:800;color:#1a6b3c;line-height:1">{wins}</div>
            <div style="font-size:.6rem;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-top:2px">W</div>
          </div>
          <div style="font-size:1.8rem;color:#ddd;font-weight:300;line-height:1.1">|</div>
          <div>
            <div style="font-size:2.1rem;font-weight:800;color:#8b1a1a;line-height:1">{losses}</div>
            <div style="font-size:.6rem;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-top:2px">P</div>
          </div>
        </div>
      </div>
      <div style="border-left:1px solid #eee;padding-left:20px;text-align:center">
        <div style="font-size:1.5rem;font-weight:700;color:var(--navy);line-height:1.1">{wins/n_matches:.0%}</div>
        <div style="font-size:.65rem;color:#999;text-transform:uppercase;letter-spacing:.5px;margin-top:.2rem">Win% &middot; {n_matches} mecze</div>
      </div>
    </div>
  </div>
  <!-- ORtg -->
  <div class="col-6 col-md-3">
    <div class="stat-card">
      <div class="stat-val" style="color:#1a6b3c">{gtk_kpi["ortg"]}</div>
      <div class="stat-lbl">ORtg</div>
    </div>
  </div>
  <!-- DRtg -->
  <div class="col-6 col-md-3">
    <div class="stat-card">
      <div class="stat-val" style="color:#8b1a1a">{opp_kpi["ortg"]}</div>
      <div class="stat-lbl">DRtg</div>
    </div>
  </div>
</div>

<div class="row g-2 mb-3">
  <div class="col"><div class="stat-card"><div class="stat-val">{tot_pts}</div><div class="stat-lbl">Punkty</div></div></div>
  <div class="col"><div class="stat-card"><div class="stat-val">{tot_reb}</div><div class="stat-lbl">Zbiórki</div></div></div>
  <div class="col"><div class="stat-card"><div class="stat-val">{tot_ast}</div><div class="stat-lbl">Asysty</div></div></div>
  <div class="col"><div class="stat-card"><div class="stat-val">{tot_stl}</div><div class="stat-lbl">Przechwyty</div></div></div>
  <div class="col"><div class="stat-card"><div class="stat-val">{tot_blk}</div><div class="stat-lbl">Bloki</div></div></div>
</div>

<ul class="nav nav-tabs mb-2">
  <li class="nav-item"><button class="nav-link active" data-bs-toggle="tab" data-bs-target="#sMetrics">Metryki średnie</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#sPiatki">Piątki</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#sTiming">Timing Akcji</button></li>
</ul>

<div class="tab-content">

<div class="tab-pane fade show active" id="sMetrics">
  <div class="row g-3">

    <!-- Atak / mecz -->
    <div class="col-lg-4">
      <div class="card h-100"><div class="card-body p-3">
        <div class="section-hdr">Atak / mecz</div>
        {scoring_duel_html}
      </div></div>
    </div>

    <!-- Obrona / mecz -->
    <div class="col-lg-4">
      <div class="card h-100"><div class="card-body p-3">
        <div class="section-hdr">Obrona / mecz</div>
        {defense_duel_html}
      </div></div>
    </div>

    <!-- Skuteczność rzutów -->
    <div class="col-lg-4">
      <div class="card h-100"><div class="card-body p-3">
        <div class="section-hdr">Skuteczność rzutów</div>
        {shooting_eff_html}
      </div></div>
    </div>
  </div>
</div>

<div class="tab-pane fade" id="sPiatki">
  <div class="card mt-1"><div class="card-body p-2">
    <div class="section-hdr">Piątki sezonu — statystyki uśrednione per mecz</div>
    {season_lineup_table_html()}
  </div></div>
</div>

<div class="tab-pane fade" id="sTiming">
  <div class="card mt-1"><div class="card-body p-2">
    <div class="section-hdr">Timing Akcji — skuteczność według czasu posiadania (zegar 24s)</div>
    <div class="d-flex gap-3 mb-2" style="font-size:.78rem">
      <span><span style="display:inline-block;width:12px;height:8px;background:#1a6b3c;border-radius:2px;margin-right:4px"></span>{gtk_name}</span>
      <span><span style="display:inline-block;width:12px;height:8px;background:#8b1a1a;border-radius:2px;margin-right:4px"></span>Przeciwnicy</span>
    </div>
    <div class="table-responsive">
    <table class="table table-hover mb-0">
      <thead>
        <tr>
          <th rowspan="2" style="vertical-align:middle;background:#1a2b4a;color:#fff">Czas</th>
          <th colspan="3" style="background:#e8f5e9;color:#1a6b3c;text-align:center;border-bottom:2px solid #1a6b3c">{gtk_name}</th>
          <th colspan="2" style="background:#fff;border-bottom:2px solid #dee2e6"></th>
          <th colspan="3" style="background:#ffebee;color:#8b1a1a;text-align:center;border-bottom:2px solid #8b1a1a">Przeciwnicy</th>
        </tr>
        <tr>
          <th style="background:#e8f5e9;color:#1a6b3c;text-align:center;font-size:.72rem">Celne/Próby</th>
          <th style="background:#e8f5e9;color:#1a6b3c;text-align:center;font-size:.72rem">Eff%</th>
          <th style="background:#e8f5e9;color:#1a6b3c;text-align:center;font-size:.72rem">2PT | 3PT</th>
          <th style="background:#fff;text-align:center;font-size:.72rem;width:90px;color:#555">{gtk_name}</th>
          <th style="background:#fff;text-align:center;font-size:.72rem;width:90px;color:#555">Przeciwnicy</th>
          <th style="background:#ffebee;color:#8b1a1a;text-align:center;font-size:.72rem">2PT | 3PT</th>
          <th style="background:#ffebee;color:#8b1a1a;text-align:center;font-size:.72rem">Eff%</th>
          <th style="background:#ffebee;color:#8b1a1a;text-align:center;font-size:.72rem">Celne/Próby</th>
        </tr>
      </thead>
      <tbody>{tim_rows}</tbody>
    </table>
    </div>
  </div></div>
</div>

</div>"""

    scripts = ""

    return html_response(base(content, scripts, active="season"))

# ══════════════════════════════════════════════════════════════════════════════
# ZAWODNICY SEZONU
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/zawodnicy")
@login_required
def zawodnicy():
    sezon_filter   = request.args.get("sezon", get_setting("current_season") or "")
    team_id_filter = request.args.get("team_id", "")
    db = get_db(); cur = db.cursor()

    # Pobierz team_id z kontekstu sidebara jeśli nie podano w URL
    if not team_id_filter:
        ctx_klub    = get_setting("current_klub") or ""
        ctx_sezon   = get_setting("current_season") or ""
        ctx_druzyna = get_setting("current_druzyna") or ""
        if ctx_klub and ctx_sezon and ctx_druzyna:
            try:
                cur.execute("""
                    SELECT t.id FROM teams t
                    JOIN seasons s ON t.season_id=s.id
                    JOIN clubs c ON s.club_id=c.id
                    WHERE c.name=%s AND s.name=%s AND t.name=%s
                """, (ctx_klub, ctx_sezon, ctx_druzyna))
                tr = cur.fetchone()
                if tr: team_id_filter = str(tr["id"])
            except: pass

    # Sezony dla aktywnego klubu z bazy seasons (spójne z kontekstem)
    sezony_db = []
    if ctx_klub:
        try:
            cur.execute("""
                SELECT s.name FROM seasons s
                JOIN clubs c ON s.club_id=c.id
                WHERE c.name=%s ORDER BY s.name DESC
            """, (ctx_klub,))
            sezony_db = [r["name"] for r in cur.fetchall()]
        except: pass
    if not sezony_db:
        cur.execute("SELECT DISTINCT sezon FROM matches ORDER BY sezon DESC")
        sezony_db = [r["sezon"] for r in cur.fetchall()]
    sezony = sezony_db

    if team_id_filter:
        cur.execute("SELECT COUNT(*) as cnt FROM matches WHERE sezon=%s AND team_id=%s",
                    (sezon_filter, int(team_id_filter)))
    else:
        cur.execute("SELECT COUNT(*) as cnt FROM matches WHERE sezon=%s", (sezon_filter,))
    n_matches = cur.fetchone()["cnt"]

    # Agreguj TYLKO po roster_id — nieprzypisani osobno per numer
    try:
        cur.execute("""
            SELECT
                grp_id, nazwa,
                SUM(pts) as pts, SUM(p2m) as p2m, SUM(p2a) as p2a,
                SUM(p3m) as p3m, SUM(p3a) as p3a,
                SUM(ftm) as ftm, SUM(fta) as fta,
                SUM(ast) as ast, SUM(oreb) as oreb, SUM(dreb) as dreb,
                SUM(br) as br, SUM(fd) as fd, SUM(finishes) as finishes,
                SUM(stl) as stl, SUM(blk) as blk,
                COUNT(DISTINCT match_id) as mecze,
                BOOL_OR(ma_nieprzypisane) as ma_nieprzypisane,
                SUM(team_poss) as team_poss,
                SUM(time_sum) as time_sum, SUM(time_cnt) as time_cnt
            FROM (
                SELECT
                    CASE WHEN p.id IS NOT NULL THEN 'p_'||p.id::text
                         WHEN r.id IS NOT NULL THEN r.id::text
                         ELSE 'nr_'||ps.nr::text END as grp_id,
                    CASE WHEN p.id IS NOT NULL
                         THEN p.nazwisko || ' ' || p.imie
                         WHEN r.id IS NOT NULL
                         THEN r.nazwisko || ' ' || r.imie
                         ELSE '— nieprzypisany #' || ps.nr::text
                    END as nazwa,
                    ps.match_id,
                    SUM(ps.pts) as pts, SUM(ps.p2m) as p2m, SUM(ps.p2a) as p2a,
                    SUM(ps.p3m) as p3m, SUM(ps.p3a) as p3a,
                    SUM(ps.ftm) as ftm, SUM(ps.fta) as fta,
                    SUM(ps.ast) as ast, SUM(ps.oreb) as oreb, SUM(ps.dreb) as dreb,
                    SUM(ps.br) as br, SUM(ps.fd) as fd, SUM(ps.finishes) as finishes,
                    SUM(COALESCE(ps.stl,0)) as stl, SUM(COALESCE(ps.blk,0)) as blk,
                    (p.id IS NULL AND r.id IS NULL) as ma_nieprzypisane,
                    COALESCE(mp.poss, 0) as team_poss,
                    SUM(COALESCE(ps.time_sum,0)) as time_sum,
                    SUM(COALESCE(ps.time_cnt,0)) as time_cnt
                FROM player_stats ps
                JOIN matches m ON ps.match_id=m.id
                LEFT JOIN players p ON ps.player_id=p.id
                LEFT JOIN roster r ON ps.roster_id=r.id
                LEFT JOIN (
                    SELECT match_id, SUM(poss) as poss
                    FROM match_stats WHERE druzyna='gtk'
                    GROUP BY match_id
                ) mp ON mp.match_id=ps.match_id
                WHERE m.sezon=%s AND ps.druzyna='gtk'
                AND (%s::int IS NULL OR m.team_id=%s::int)
                GROUP BY p.id, p.imie, p.nazwisko, r.id, r.imie, r.nazwisko, ps.nr, ps.match_id, mp.poss
            ) sub
            GROUP BY grp_id, nazwa
            ORDER BY BOOL_OR(ma_nieprzypisane) ASC, SUM(pts) DESC
        """, (sezon_filter, team_id_filter or None, team_id_filter or None))
    except Exception:
        try: get_db().rollback()
        except: pass
        cur = get_db().cursor()
        cur.execute("""
            SELECT ps.nr::text as grp_id,
                   '— nieprzypisany #'||ps.nr::text as nazwa,
                   SUM(ps.pts) as pts, SUM(ps.p2m) as p2m, SUM(ps.p2a) as p2a,
                   SUM(ps.p3m) as p3m, SUM(ps.p3a) as p3a,
                   SUM(ps.ftm) as ftm, SUM(ps.fta) as fta,
                   SUM(ps.ast) as ast, SUM(ps.oreb) as oreb, SUM(ps.dreb) as dreb,
                   SUM(ps.br) as br, SUM(ps.fd) as fd, SUM(ps.finishes) as finishes,
                   SUM(COALESCE(ps.stl,0)) as stl, SUM(COALESCE(ps.blk,0)) as blk,
                   COUNT(DISTINCT ps.match_id) as mecze,
                   TRUE as ma_nieprzypisane
            FROM player_stats ps
            JOIN matches m ON ps.match_id=m.id
            WHERE m.sezon=%s AND ps.druzyna='gtk'
            GROUP BY ps.nr ORDER BY SUM(ps.pts) DESC
        """, (sezon_filter,))
    players = cur.fetchall()
    cur.close()

    # ── Preload calc_play_time dla sezonu (niezależnie od DB time_sum) ──────
    _pt_by_grp = {}
    try:
        _cur_pt = get_db().cursor()
        if team_id_filter:
            _cur_pt.execute("SELECT id FROM matches WHERE sezon=%s AND team_id=%s",
                            (sezon_filter, int(team_id_filter)))
        else:
            _cur_pt.execute("SELECT id FROM matches WHERE sezon=%s", (sezon_filter,))
        _match_ids_pt = [_r["id"] for _r in _cur_pt.fetchall()]
        _pt_all = {}  # {(match_id, nr): secs×1.22}
        for _mid in _match_ids_pt:
            for _nr, _secs in calc_play_time(_mid).items():
                _pt_all[(_mid, int(_nr))] = _secs
        _q_args = (sezon_filter, int(team_id_filter)) if team_id_filter else (sezon_filter,)
        _cur_pt.execute(f"""
            SELECT DISTINCT
                CASE WHEN p.id IS NOT NULL THEN 'p_'||p.id::text
                     WHEN r.id IS NOT NULL THEN r.id::text
                     ELSE 'nr_'||ps.nr::text END as grp_id,
                ps.match_id, ps.nr
            FROM player_stats ps
            JOIN matches m ON ps.match_id=m.id
            LEFT JOIN players p ON ps.player_id=p.id
            LEFT JOIN roster r ON ps.roster_id=r.id
            WHERE m.sezon=%s AND ps.druzyna='gtk'
            {"AND m.team_id=%s" if team_id_filter else ""}
        """, _q_args)
        for _gr in _cur_pt.fetchall():
            _g = _gr["grp_id"]
            _s = _pt_all.get((_gr["match_id"], int(_gr["nr"] or 0)), 0)
            _pt_by_grp[_g] = _pt_by_grp.get(_g, 0) + _s
        _cur_pt.close()
    except Exception:
        pass

    rows = ""
    for i, p in enumerate(players):
        def _i(k): return int(p.get(k,0) or 0)
        n    = max(_i("mecze"), 1)
        pm2  = _i("p2m"); p2a = _i("p2a")
        pm3  = _i("p3m"); p3a = _i("p3a")
        ftm  = _i("ftm"); fta = _i("fta")
        pts  = _i("pts")
        ast  = _i("ast"); oreb = _i("oreb"); dreb = _i("dreb")
        br   = _i("br");  fd   = _i("fd");   fin  = _i("finishes")
        stl  = _i("stl"); blk  = _i("blk")
        fga  = p2a + p3a
        # Rate stats (use totals)
        efg  = f"{(pm2+1.5*pm3)/fga*100:.1f}%" if fga else "—"
        ts   = f"{pts/(2*(fga+0.44*fta))*100:.1f}%" if (fga+fta) else "—"
        tposs = _i("team_poss")
        usg  = f"{(fga + 0.44*fta + br) / tposs*100:.1f}%" if tposs else "—"
        p2pct = f"{pm2/p2a*100:.1f}%" if p2a else "—"
        p3pct = f"{pm3/p3a*100:.1f}%" if p3a else "—"
        ftpct = f"{ftm/fta*100:.1f}%" if fta else "—"
        # Averages per game
        def _avg(v): return f"{v/n:.1f}"
        # MIN szac.: calc_play_time (×1.22 already), fallback DB time_sum
        _grp_id = p.get("grp_id", "")
        _pt_total = _pt_by_grp.get(_grp_id, 0)
        if _pt_total:
            _avg_min = _pt_total / n / 60
        else:
            tsum = float(p.get("time_sum") or 0)
            _avg_min = (tsum / n / 60) * 1.22 if tsum else 0
        if _avg_min:
            min_pg = f"{int(_avg_min)}:{int((_avg_min % 1) * 60):02d}"
        else:
            min_pg = "—"
        # Color coding
        efg_c = "#0F6E56" if fga and (pm2+1.5*pm3)/fga>=0.5 else ("#A32D2D" if fga and (pm2+1.5*pm3)/fga<0.35 else "inherit")
        nie = p.get('ma_nieprzypisane')
        nazwa = p.get('nazwa','?')
        grp_id = p.get('grp_id','')
        bg = "background:#fff8e1" if nie else ("background:#f8f9ff" if i%2==0 else "")
        warn = ' <span title="Przypisz zawodnika w raporcie meczu" style="color:#f9a825;font-size:.75rem">⚠</span>' if nie else ''
        if not nie and grp_id:
            if grp_id.startswith('p_'):
                pid = grp_id[2:]
                nazwa_cell = f'<a href="/zawodnik/{pid}?sezon={sezon_filter}" style="color:#1a2b4a;text-decoration:none;font-weight:600">{nazwa}</a>'
            elif grp_id.isdigit():
                nazwa_cell = f'<a href="/zawodnik/{grp_id}?sezon={sezon_filter}" style="color:#1a2b4a;text-decoration:none;font-weight:600">{nazwa}</a>'
            else:
                nazwa_cell = f'<span style="font-weight:500">{nazwa}</span>'
        else:
            nazwa_cell = f'{nazwa}{warn}'
        _total_min = _avg_min * n  # łączne minuty zawodnika w sezonie
        rows += f"""<tr style="{bg}" data-n="{n}" data-min="{_total_min:.3f}"
            data-pts="{pts}" data-p2m="{pm2}" data-p2a="{p2a}"
            data-p3m="{pm3}" data-p3a="{p3a}" data-ftm="{ftm}" data-fta="{fta}"
            data-oreb="{oreb}" data-dreb="{dreb}" data-ast="{ast}" data-br="{br}"
            data-stl="{stl}" data-blk="{blk}" data-fd="{fd}">
            <td style="white-space:nowrap;font-size:.82rem">{nazwa_cell}</td>
            <td class="text-center" style="color:#633806;white-space:nowrap" data-cell="min">{min_pg}</td>
            <td class="text-center" style="font-size:.75rem;color:#aaa;white-space:nowrap">{n}</td>
            <td class="text-center fw-bold" style="color:#1a2b4a;white-space:nowrap" data-cell="pts">{_avg(pts)}</td>
            <td class="text-center" style="white-space:nowrap" data-cell="p2ma">{pm2/n:.1f}/{p2a/n:.1f}</td>
            <td class="text-center" style="white-space:nowrap">{p2pct}</td>
            <td class="text-center" style="white-space:nowrap" data-cell="p3ma">{pm3/n:.1f}/{p3a/n:.1f}</td>
            <td class="text-center" style="white-space:nowrap">{p3pct}</td>
            <td class="text-center" style="white-space:nowrap" data-cell="ftma">{ftm/n:.1f}/{fta/n:.1f}</td>
            <td class="text-center" style="white-space:nowrap">{ftpct}</td>
            <td class="text-center" style="white-space:nowrap" data-cell="oreb">{_avg(oreb)}</td>
            <td class="text-center" style="white-space:nowrap" data-cell="dreb">{_avg(dreb)}</td>
            <td class="text-center" style="white-space:nowrap" data-cell="zbs">{_avg(oreb+dreb)}</td>
            <td class="text-center" style="white-space:nowrap" data-cell="ast">{_avg(ast)}</td>
            <td class="text-center" style="white-space:nowrap" data-cell="br">{_avg(br)}</td>
            <td class="text-center" style="white-space:nowrap" data-cell="stl">{_avg(stl)}</td>
            <td class="text-center" style="white-space:nowrap" data-cell="blk">{_avg(blk)}</td>
            <td class="text-center" style="white-space:nowrap" data-cell="fd">{_avg(fd)}</td>
            <td class="text-center" style="white-space:nowrap;color:{efg_c}"><b>{efg}</b></td>
            <td class="text-center" style="white-space:nowrap">{ts}</td>
            <td class="text-center" style="white-space:nowrap"><b>{usg}</b></td>
        </tr>"""

    season_opts = "".join([f'<option value="{s}" {"selected" if s==sezon_filter else ""}>{s}</option>' for s in sezony])
    gtk_name = get_setting("gtk_name") or "GTK"

    th_s  = 'background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;cursor:pointer;user-select:none;vertical-align:middle'
    th_l  = 'background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 8px;text-align:left;white-space:nowrap;vertical-align:middle'
    th_g  = 'background:#1a2b4a;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center'
    th_gs = 'background:#1a2b4a;color:rgba(255,255,255,.8);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center;cursor:pointer;user-select:none;white-space:nowrap'
    th_z  = 'background:#152236;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.1);text-align:center'
    th_zs = 'background:#152236;color:rgba(255,255,255,.75);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center;cursor:pointer;user-select:none;white-space:nowrap'
    vm    = 'vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2)'

    def zth(label, col, style=None):
        s = style or th_s
        return f'<th style="{s}" onclick="sortZaw({col})"><span id="thz_{col}">{label}</span></th>'

    thead = f"""<thead>
      <tr>
        <th style="{th_l};{vm}" rowspan="3">Zawodnik</th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortZaw(1)"><span id="thz_1">MIN<br>(szac.)</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortZaw(2)"><span id="thz_2">G</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortZaw(3)"><span id="thz_3">PTS ↓</span></th>
        <th style="{th_g}" colspan="2">2PT</th>
        <th style="{th_g}" colspan="2">3PT</th>
        <th style="{th_g}" colspan="2">FT</th>
        <th style="{th_z}" colspan="3">ZB</th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortZaw(13)"><span id="thz_13">AST</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortZaw(14)"><span id="thz_14">TO</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortZaw(15)"><span id="thz_15">STL</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortZaw(16)"><span id="thz_16">BLK</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortZaw(17)"><span id="thz_17">FD</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortZaw(18)"><span id="thz_18">eFG%</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortZaw(19)"><span id="thz_19">TS%</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortZaw(20)"><span id="thz_20">USG%</span></th>
      </tr>
      <tr>
        <th style="{th_gs}" onclick="sortZaw(4)"><span id="thz_4">M/A</span></th>
        <th style="{th_gs}" onclick="sortZaw(5)"><span id="thz_5">%</span></th>
        <th style="{th_gs}" onclick="sortZaw(6)"><span id="thz_6">M/A</span></th>
        <th style="{th_gs}" onclick="sortZaw(7)"><span id="thz_7">%</span></th>
        <th style="{th_gs}" onclick="sortZaw(8)"><span id="thz_8">M/A</span></th>
        <th style="{th_gs}" onclick="sortZaw(9)"><span id="thz_9">%</span></th>
        <th style="{th_zs}" onclick="sortZaw(10)"><span id="thz_10">A</span></th>
        <th style="{th_zs}" onclick="sortZaw(11)"><span id="thz_11">O</span></th>
        <th style="{th_zs}" onclick="sortZaw(12)"><span id="thz_12">S</span></th>
      </tr>
    </thead>"""

    content = f"""
<div class="d-flex justify-content-between align-items-center mb-3 flex-wrap gap-2">
  <div style="border-left:3px solid #1a2b4a;padding-left:11px">
    <div style="font-size:.95rem;font-weight:700;color:#1a2b4a;line-height:1.2">Statystyki indywidualne</div>
    <div id="per36-desc" style="font-size:.67rem;color:#bbb;margin-top:2px">Wartości uśrednione &middot; G = liczba meczów</div>
  </div>
  <div style="display:flex;align-items:center;gap:10px">
    <!-- Przełącznik Średnie / Sumaryczne -->
    <div style="display:flex;background:#1a2b4a;border-radius:20px;padding:2px;gap:0;box-shadow:0 2px 8px rgba(0,0,0,.18)">
      <button id="statModeAvg" onclick="setStatMode('avg')"
        style="border:none;border-radius:18px;padding:5px 13px;font-size:.74rem;font-weight:600;
               cursor:pointer;background:#fff;color:#1a2b4a;transition:.18s;white-space:nowrap">Średnie</button>
      <button id="statModeSum" onclick="setStatMode('sum')"
        style="border:none;border-radius:18px;padding:5px 13px;font-size:.74rem;font-weight:600;
               cursor:pointer;background:transparent;color:rgba(255,255,255,.7);transition:.18s;white-space:nowrap">Sumaryczne</button>
    </div>
    <!-- Pill dropdown Per -->
    <div style="position:relative;display:inline-block">
      <select id="perSelect" onchange="setPer(this.value)"
        style="background:#fff;color:#1a2b4a;border:none;border-radius:20px;
               padding:6px 32px 6px 16px;font-size:.78rem;font-weight:600;
               cursor:pointer;appearance:none;-webkit-appearance:none;
               outline:none;box-shadow:0 2px 8px rgba(0,0,0,.18);
               letter-spacing:.2px;min-width:115px">
        <option value="game" selected>Per Mecz</option>
        <option value="36">Per 36 min</option>
        <option value="40">Per 40 min</option>
        <option value="100">Per 100 pos</option>
      </select>
      <span style="position:absolute;right:12px;top:50%;transform:translateY(-50%);
                   pointer-events:none;color:#1a2b4a;font-size:.6rem">▼</span>
    </div>
  </div>
</div>
<div class="card">
  <div class="card-body p-2">
    <div class="table-responsive">
      <table class="table table-hover mb-0" id="zawTable" style="min-width:900px">
        {thead}
        <tbody id="zawBody">
          {rows if rows else '<tr><td colspan="21" class="text-center text-muted py-4">Brak danych zawodników</td></tr>'}
        </tbody>
      </table>
    </div>
  </div>
</div>"""

    scripts = """<script>
let _zawDir = {};
let _perMode = 'game';
let _statMode = 'avg';

function f1(v){ return v.toFixed(1); }

function setStatMode(mode) {
    _statMode = mode;
    const avgBtn = document.getElementById('statModeAvg');
    const sumBtn = document.getElementById('statModeSum');
    if(avgBtn && sumBtn) {
        if(mode === 'avg') {
            avgBtn.style.background = '#fff';  avgBtn.style.color = '#1a2b4a';
            sumBtn.style.background = 'transparent'; sumBtn.style.color = 'rgba(255,255,255,.7)';
        } else {
            sumBtn.style.background = '#fff';  sumBtn.style.color = '#1a2b4a';
            avgBtn.style.background = 'transparent'; avgBtn.style.color = 'rgba(255,255,255,.7)';
        }
    }
    const perSel = document.getElementById('perSelect');
    if(perSel) {
        perSel.disabled = (mode === 'sum');
        perSel.parentElement.style.opacity = (mode === 'sum') ? '0.38' : '1';
        perSel.parentElement.style.pointerEvents = (mode === 'sum') ? 'none' : '';
    }
    if(mode === 'avg') {
        setPer(_perMode);
    } else {
        const desc = document.getElementById('per36-desc');
        if(desc) desc.textContent = 'Wartości sumaryczne za cały sezon  ·  G = liczba meczów';
        const tbody = document.getElementById('zawBody');
        if(!tbody) return;
        tbody.querySelectorAll('tr').forEach(function(row) {
            row.querySelectorAll('[data-cell]').forEach(function(td) {
                const cell = td.dataset.cell;
                const iv = k => parseInt(row.dataset[k]) || 0;
                if(cell === 'pts')       td.textContent = iv('pts');
                else if(cell === 'p2ma') td.textContent = iv('p2m') + '/' + iv('p2a');
                else if(cell === 'p3ma') td.textContent = iv('p3m') + '/' + iv('p3a');
                else if(cell === 'ftma') td.textContent = iv('ftm') + '/' + iv('fta');
                else if(cell === 'oreb') td.textContent = iv('oreb');
                else if(cell === 'dreb') td.textContent = iv('dreb');
                else if(cell === 'zbs')  td.textContent = iv('oreb') + iv('dreb');
                else if(cell === 'ast')  td.textContent = iv('ast');
                else if(cell === 'br')   td.textContent = iv('br');
                else if(cell === 'stl')  td.textContent = iv('stl');
                else if(cell === 'blk')  td.textContent = iv('blk');
                else if(cell === 'fd')   td.textContent = iv('fd');
                else if(cell === 'min')  {
                    const m = parseFloat(row.dataset.min) || 0;
                    td.textContent = m ? Math.floor(m) + ':' + String(Math.round((m%1)*60)).padStart(2,'0') : '—';
                }
            });
        });
    }
}

function setPer(mode) {
    _perMode = mode;
    if(_statMode === 'sum') return;
    const desc = document.getElementById('per36-desc');
    const labels = {
        'game': 'Wartości uśrednione per mecz w którym zawodnik zagrał  ·  G = liczba meczów',
        '36':   'Statystyki przeliczone na 36 minut gry  ·  G = liczba meczów',
        '40':   'Statystyki przeliczone na 40 minut gry  ·  G = liczba meczów',
        '100':  'Statystyki przeliczone na 100 posiadań  ·  G = liczba meczów'
    };
    if(desc) desc.textContent = labels[mode] || labels['game'];

    const tbody = document.getElementById('zawBody');
    if(!tbody) return;
    tbody.querySelectorAll('tr').forEach(function(row) {
        const n   = parseFloat(row.dataset.n)   || 1;
        const min = parseFloat(row.dataset.min)  || 0;
        const has = min > 0;
        const scale = mode === '36' ? 36 : mode === '40' ? 40 : mode === '100' ? 100 : 0;

        function pg(key) {
            const v = parseFloat(row.dataset[key]) || 0;
            if(scale === 0) return f1(v / n);
            if(!has) return '—';
            return f1(v / min * scale);
        }
        function pgma(km, ka) {
            const vm = parseFloat(row.dataset[km]) || 0;
            const va = parseFloat(row.dataset[ka]) || 0;
            if(scale === 0) return f1(vm/n) + '/' + f1(va/n);
            if(!has) return '—/—';
            return f1(vm/min*scale) + '/' + f1(va/min*scale);
        }

        row.querySelectorAll('[data-cell]').forEach(function(td) {
            const cell = td.dataset.cell;
            if(cell === 'pts')       td.textContent = pg('pts');
            else if(cell === 'p2ma') td.textContent = pgma('p2m','p2a');
            else if(cell === 'p3ma') td.textContent = pgma('p3m','p3a');
            else if(cell === 'ftma') td.textContent = pgma('ftm','fta');
            else if(cell === 'oreb') td.textContent = pg('oreb');
            else if(cell === 'dreb') td.textContent = pg('dreb');
            else if(cell === 'zbs')  { const o=parseFloat(row.dataset.oreb)||0, d=parseFloat(row.dataset.dreb)||0; td.textContent = scale===0 ? f1((o+d)/n) : (has ? f1((o+d)/min*scale) : '—'); }
            else if(cell === 'ast')  td.textContent = pg('ast');
            else if(cell === 'br')   td.textContent = pg('br');
            else if(cell === 'stl')  td.textContent = pg('stl');
            else if(cell === 'blk')  td.textContent = pg('blk');
            else if(cell === 'fd')   td.textContent = pg('fd');
        });
    });
}

function sortZaw(col) {
    const tbody = document.getElementById('zawBody');
    if(!tbody) return;
    const rows = Array.from(tbody.querySelectorAll('tr'));
    _zawDir[col] = !_zawDir[col];
    document.querySelectorAll('[id^="thz_"]').forEach(el => {
        el.textContent = el.textContent.replace(/ [▲▼]$/,'');
    });
    const thEl = document.getElementById('thz_' + col);
    if(thEl) thEl.textContent += _zawDir[col] ? ' ▼' : ' ▲';
    rows.sort((a, b) => {
        const av = a.cells[col]?.textContent.trim().replace('%','').replace(/[/].*$/,'') || '';
        const bv = b.cells[col]?.textContent.trim().replace('%','').replace(/[/].*$/,'') || '';
        const an = parseFloat(av); const bn = parseFloat(bv);
        if(!isNaN(an) && !isNaN(bn)) return _zawDir[col] ? bn - an : an - bn;
        return _zawDir[col] ? bv.localeCompare(av,'pl') : av.localeCompare(bv,'pl');
    });
    rows.forEach(r => tbody.appendChild(r));
}
window.addEventListener('DOMContentLoaded', () => { sortZaw(3); });
</script>"""

    return html_response(base(content, scripts, active="players"))

# ══════════════════════════════════════════════════════════════════════════════
# PROFIL ZAWODNIKA
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/zawodnik/<int:roster_id>")
@login_required
def profil_zawodnika(roster_id):
    sezon_filter = request.args.get("sezon", get_setting("current_season") or "")
    db = get_db(); cur = db.cursor()

    # Dane zawodnika — sprawdź obie tabele: roster (stary) i players (Struktura klubów)
    zawodnik = None
    use_player_id = False

    cur.execute("SELECT * FROM roster WHERE id=%s", (roster_id,))
    zawodnik_roster = cur.fetchone()
    if zawodnik_roster:
        zawodnik = zawodnik_roster
        use_player_id = False
    else:
        # Fallback: sprawdź czy to player_id z tabeli players
        cur.execute("SELECT * FROM players WHERE id=%s", (roster_id,))
        zawodnik_player = cur.fetchone()
        if zawodnik_player:
            zawodnik = zawodnik_player
            use_player_id = True

    if not zawodnik:
        flash("Zawodnik nie istnieje","error")
        return redirect(url_for("zawodnicy"))

    # Kolumna łączenia zależy od tabeli
    join_col = "ps.player_id" if use_player_id else "ps.roster_id"

    # Wszystkie sezony w których grał
    cur.execute(f"""SELECT DISTINCT m.sezon FROM player_stats ps
                   JOIN matches m ON ps.match_id=m.id
                   WHERE {join_col}=%s ORDER BY m.sezon DESC""", (roster_id,))
    sezony = [r["sezon"] for r in cur.fetchall()]

    # Statystyki per mecz w wybranym sezonie
    cur.execute(f"""
        SELECT ps.*, m.data_meczu, m.przeciwnik, m.wynik_gtk, m.wynik_opp,
               ms.poss as team_poss
        FROM player_stats ps
        JOIN matches m ON ps.match_id=m.id
        LEFT JOIN (SELECT match_id, SUM(poss) as poss FROM match_stats
                   WHERE druzyna='gtk' GROUP BY match_id) ms ON ms.match_id=ps.match_id
        WHERE {join_col}=%s AND m.sezon=%s AND ps.druzyna='gtk'
        AND (%s::int IS NULL OR m.team_id=%s::int)
        ORDER BY m.data_meczu ASC
    """, (roster_id, sezon_filter,
             int(request.args.get("team_id","") or 0) or None,
             int(request.args.get("team_id","") or 0) or None))
    mecze_stats = list(cur.fetchall())

    # ── Preload calc_play_time per mecz zawodnika ────────────────────────
    _pt_per_match = {}  # {match_id: secs×1.22 dla tego zawodnika}
    for _mr in mecze_stats:
        _mid = _mr["match_id"]
        _mnr = int(_mr.get("nr") or 0)
        _pt_per_match[_mid] = calc_play_time(_mid).get(_mnr, 0)
    _pt_total_zawodnik = sum(_pt_per_match.values())

    # Numer(y) w tym sezonie
    cur.execute(f"""SELECT DISTINCT ps.nr FROM player_stats ps
                   JOIN matches m ON ps.match_id=m.id
                   WHERE {join_col}=%s AND m.sezon=%s""", (roster_id, sezon_filter))
    numery = [str(r["nr"]) for r in cur.fetchall()]
    cur.close()

    if not mecze_stats:
        gtk_name = get_setting("gtk_name") or "GTK"
        content = f"""
<div class="d-flex align-items-center gap-2 mb-3">
  <a href="/zawodnicy" class="btn btn-outline-secondary btn-sm">← Statystyki</a>
  <div class="page-title mb-0">👤 {zawodnik['imie']} {zawodnik['nazwisko']}</div>
</div>
<div class="card p-4 text-center text-muted">Brak danych w sezonie {sezon_filter}.</div>"""
        return html_response(base(content, active="players"))

    # Agregaty sezonu
    def s(k): return sum(int(r.get(k,0) or 0) for r in mecze_stats)
    n = len(mecze_stats)
    pts_tot = s("pts"); fga_tot = s("p2a")+s("p3a"); fta_tot = s("fta")
    pm2_tot = s("p2m"); pm3_tot = s("p3m"); ftm_tot = s("ftm")
    br_tot = s("br"); fin_tot = s("finishes"); ast_tot = s("ast")
    stl_tot = s("stl"); blk_tot = s("blk"); fd_tot = s("fd")
    poss_tot = sum(int(r.get("team_poss",0) or 0) for r in mecze_stats)

    ppg   = f"{pts_tot/n:.1f}"
    efg   = f"{(pm2_tot+1.5*pm3_tot)/fga_tot:.1%}" if fga_tot else "—"
    ts    = f"{pts_tot/(2*(fga_tot+0.44*fta_tot)):.1%}" if (fga_tot+fta_tot) else "—"
    usg   = f"{(fga_tot+0.44*fta_tot+br_tot)/poss_tot:.1%}" if poss_tot else "—"
    p2pct = f"{pm2_tot/s('p2a'):.1%}" if s('p2a') else "—"
    p3pct = f"{pm3_tot/s('p3a'):.1%}" if s('p3a') else "—"
    ftpct = f"{ftm_tot/fta_tot:.1%}" if fta_tot else "—"
    tsum_tot = sum(float(r.get("time_sum") or 0) for r in mecze_stats)
    tcnt_tot = sum(int(r.get("time_cnt") or 0) for r in mecze_stats)
    avg_t_tot = f"{tsum_tot/tcnt_tot:.1f}s" if tcnt_tot else "—"

    # KPI cards
    def kpi(val, lbl, color="#1a2b4a", subtitle="", tooltip=""):
        sub_html = f'<div style="font-size:.65rem;color:#bbb;margin-top:1px">{subtitle}</div>' if subtitle else ""
        tip_attr = f' title="{tooltip}"' if tooltip else (f' title="{subtitle}"' if subtitle else "")
        cur = 'help' if (tooltip or subtitle) else 'default'
        # Bootstrap tooltip via data-bs-toggle jeśli jest dłuższy opis
        bs_tip = ""
        if tooltip:
            safe = tooltip.replace('"', '&quot;')
            bs_tip = f' data-bs-toggle="tooltip" data-bs-placement="top" data-bs-title="{safe}"'
        return (f'<div class="col"><div class="stat-card"{tip_attr}{bs_tip} style="cursor:{cur}">'
                f'<div class="stat-val sm" style="color:{color}">{val}</div>'
                f'<div class="stat-lbl">{lbl}</div>'
                f'{sub_html}'
                f'</div></div>')

    kpi_html = (
        kpi(p2pct,"2PT%",    "#1a2b4a", "2-point FG%",
            "Skuteczność rzutów za 2 punkty. Celne za 2 / wszystkie próby za 2.") +
        kpi(p3pct,"3PT%",    "#1a2b4a", "3-point FG%",
            "Skuteczność rzutów za 3 punkty. Celne za 3 / wszystkie próby za 3.") +
        kpi(ftpct,"FT%",     "#1a2b4a", "free throw %",
            "Skuteczność rzutów wolnych. Celne wolne / wszystkie próby wolne.") +
        kpi(efg,  "eFG%",    "#1a2b4a", "effective FG%",
            "Skuteczność rzutów z wagą dla trójek. Wzór: (2PM + 1.5×3PM) / FGA. Trójka warta więcej niż dwójka.") +
        kpi(ts,   "TS%",     "#1a2b4a", "true shooting %",
            "Prawdziwa skuteczność uwzględniająca rzuty wolne. Wzór: PTS / (2 × (FGA + 0.44×FTA)).") +
        kpi(usg,  "USG%",    "#D85A30", "usage rate",
            "Procent akcji drużyny zakończonych przez zawodnika. Wzór: (FGA + 0.44×FTA + TO) / akcje drużyny.") +
        kpi(avg_t_tot, "avg. play duration", "#555", "",
            "Średni czas trwania akcji zakończonych przez zawodnika (w sekundach). Im niższy — tym szybsze decyzje.")
    )

    # Tabela meczów
    match_rows = ""
    for r in mecze_stats:
        dt = r["data_meczu"].strftime("%d.%m") if r["data_meczu"] else ""
        pts = int(r.get("pts",0) or 0)
        p2a = int(r.get("p2a",0) or 0); p2m = int(r.get("p2m",0) or 0)
        p3a = int(r.get("p3a",0) or 0); p3m = int(r.get("p3m",0) or 0)
        fta = int(r.get("fta",0) or 0); ftm = int(r.get("ftm",0) or 0)
        br  = int(r.get("br",0) or 0)
        fin = int(r.get("finishes",0) or 0)
        ast = int(r.get("ast",0) or 0)
        fga = p2a + p3a
        efg_m = f"{(p2m+1.5*p3m)/fga:.0%}" if fga else "—"
        ts_m  = f"{pts/(2*(fga+0.44*fta)):.0%}" if (fga+fta) else "—"
        tposs = int(r.get("team_poss",0) or 0)
        usg_m = f"{(fga+0.44*fta+br)/tposs:.0%}" if tposs else "—"
        tsum_m = float(r.get("time_sum") or 0)
        tcnt_m = int(r.get("time_cnt") or 0)
        avg_t_m = f"{tsum_m/tcnt_m:.1f}s" if tcnt_m else "—"
        wg = int(r["wynik_gtk"] or 0); wo = int(r["wynik_opp"] or 0)
        wynik_badge = f'<span class="badge" style="background:{"#e8f5e9;color:#1a5c2a" if wg>wo else "#ffebee;color:#8b1a1a"}">{wg}:{wo}</span>'
        match_rows += f"""<tr>
            <td style="font-size:.78rem;font-weight:500">{r['przeciwnik']}</td>
            <td style="font-size:.75rem;color:#888">{dt}</td>
            <td class="text-center">{wynik_badge}</td>
            <td class="text-center fw-bold">{pts}</td>
            <td class="text-center">{p2m}/{p2a}</td><td class="text-center">{p3m}/{p3a}</td>
            <td class="text-center">{ftm}/{fta}</td>
            <td class="text-center">{ast}</td>
            <td class="text-center">{int(r.get("oreb",0) or 0) + int(r.get("dreb",0) or 0)}</td>
            <td class="text-center">{int(r.get("stl",0) or 0)}</td>
            <td class="text-center">{int(r.get("blk",0) or 0)}</td>
            <td class="text-center">{br}</td>
            <td class="text-center">{fin}</td>
            <td class="text-center" style="color:#888">{avg_t_m}</td>
            <td class="text-center"><b>{efg_m}</b></td>
            <td class="text-center">{ts_m}</td>
            <td class="text-center">{usg_m}</td>
        </tr>"""

    # Dane do wykresów JS
    import json
    labels_js  = json.dumps([r["przeciwnik"][:8] for r in mecze_stats])
    pts_js     = json.dumps([int(r.get("pts",0) or 0) for r in mecze_stats])
    efg_js     = json.dumps([round((int(r.get("p2m",0) or 0)+1.5*int(r.get("p3m",0) or 0))/(int(r.get("p2a",0) or 0)+int(r.get("p3a",0) or 0))*100,1) if (int(r.get("p2a",0) or 0)+int(r.get("p3a",0) or 0)) else None for r in mecze_stats])
    p2pct_js   = round(pm2_tot/s('p2a')*100,1) if s('p2a') else 0
    p3pct_js   = round(pm3_tot/s('p3a')*100,1) if s('p3a') else 0
    ftpct_js   = round(ftm_tot/fta_tot*100,1) if fta_tot else 0
    avg_pts    = round(pts_tot/n, 1)

    initials = (zawodnik['imie'][0] + zawodnik['nazwisko'][0]).upper()
    # Zmiana 1: nr w awatarze, nie w podpisie
    nr_display = numery[0] if len(numery) == 1 else ""
    nr_str = " / ".join(f"#{nr}" for nr in numery) if numery else ""
    gtk_name = get_setting("gtk_name") or "GTK"
    season_opts = "".join([f'<option value="{s}" {"selected" if s==sezon_filter else ""}>{s}</option>' for s in sezony])

    # Tabela: najmłodszy → najstarszy (DESC)
    match_rows_rev = ""
    for r in reversed(mecze_stats):
        dt = r["data_meczu"].strftime("%d.%m.%y") if r["data_meczu"] else ""
        dt_sort = r["data_meczu"].strftime("%Y%m%d") if r["data_meczu"] else "0"
        pts = int(r.get("pts",0) or 0)
        p2a = int(r.get("p2a",0) or 0); p2m = int(r.get("p2m",0) or 0)
        p3a = int(r.get("p3a",0) or 0); p3m = int(r.get("p3m",0) or 0)
        fta = int(r.get("fta",0) or 0); ftm = int(r.get("ftm",0) or 0)
        br  = int(r.get("br",0) or 0)
        fin = int(r.get("finishes",0) or 0)
        ast = int(r.get("ast",0) or 0)
        fd_m = int(r.get("fd",0) or 0)
        oreb_m = int(r.get("oreb",0) or 0)
        dreb_m = int(r.get("dreb",0) or 0)
        stl_m  = int(r.get("stl",0) or 0)
        blk_m  = int(r.get("blk",0) or 0)
        fga = p2a + p3a
        efg_m = f"{(p2m+1.5*p3m)/fga*100:.1f}%" if fga else "—"
        ts_m  = f"{pts/(2*(fga+0.44*fta))*100:.1f}%" if (fga+fta) else "—"
        tposs = int(r.get("team_poss",0) or 0)
        usg_m = f"{(fga+0.44*fta+br)/tposs*100:.1f}%" if tposs else "—"
        _pt_secs_m = _pt_per_match.get(r["match_id"], 0)
        if _pt_secs_m:
            min_m = f"{int(_pt_secs_m)//60}:{int(_pt_secs_m)%60:02d}"
        else:
            tsum_m = float(r.get("time_sum") or 0)
            min_m = f"{int(tsum_m/60*1.22)}:{int((tsum_m/60*1.22%1)*60):02d}" if tsum_m else "—"
        p2pct_m = f"{p2m/p2a*100:.1f}%" if p2a else "—"
        p3pct_m = f"{p3m/p3a*100:.1f}%" if p3a else "—"
        ftpct_m = f"{ftm/fta*100:.1f}%" if fta else "—"
        match_rows_rev += f"""<tr data-date="{dt_sort}">
            <td style="font-size:.78rem;font-weight:500;white-space:nowrap">{r['przeciwnik']}</td>
            <td class="text-center" style="font-size:.75rem;color:#888;white-space:nowrap">{dt}</td>
            <td class="text-center" style="color:#633806;white-space:nowrap">{min_m}</td>
            <td class="text-center fw-bold" onclick="switchMetric('PTS')" style="cursor:pointer">{pts}</td>
            <td class="text-center" style="border-left:1px solid #f0f0f0;white-space:nowrap">{p2m}/{p2a}</td>
            <td class="text-center;white-space:nowrap">{p2pct_m}</td>
            <td class="text-center" style="border-left:1px solid #f0f0f0;white-space:nowrap">{p3m}/{p3a}</td>
            <td class="text-center;white-space:nowrap">{p3pct_m}</td>
            <td class="text-center" style="border-left:1px solid #f0f0f0;white-space:nowrap">{ftm}/{fta}</td>
            <td class="text-center;white-space:nowrap">{ftpct_m}</td>
            <td class="text-center" onclick="switchMetric('OREB')" style="cursor:pointer;border-left:1px solid #f0f0f0">{oreb_m}</td>
            <td class="text-center" onclick="switchMetric('DREB')" style="cursor:pointer">{dreb_m}</td>
            <td class="text-center" onclick="switchMetric('REB')"  style="cursor:pointer;font-weight:500">{oreb_m + dreb_m}</td>
            <td class="text-center" onclick="switchMetric('AST')" style="cursor:pointer">{ast}</td>
            <td class="text-center" onclick="switchMetric('TO')" style="cursor:pointer">{br}</td>
            <td class="text-center" onclick="switchMetric('STL')" style="cursor:pointer">{stl_m}</td>
            <td class="text-center" onclick="switchMetric('BLK')" style="cursor:pointer">{blk_m}</td>
            <td class="text-center">{fd_m}</td>
            <td class="text-center" onclick="switchMetric('EFG')" style="cursor:pointer"><b>{efg_m}</b></td>
            <td class="text-center" onclick="switchMetric('TS')" style="cursor:pointer">{ts_m}</td>
            <td class="text-center" onclick="switchMetric('USG')" style="cursor:pointer">{usg_m}</td>
        </tr>"""

    # Dane JS — kolejność odwrócona (najstarszy→najnowszy = lewa→prawa)
    import json
    mecze_chron = mecze_stats  # chronologicznie (stare→nowe)
    labels_js   = json.dumps([r["przeciwnik"][:8] for r in mecze_chron])
    pts_js      = json.dumps([int(r.get("pts",0) or 0) for r in mecze_chron])
    ast_js      = json.dumps([int(r.get("ast",0) or 0) for r in mecze_chron])
    br_js       = json.dumps([int(r.get("br",0) or 0) for r in mecze_chron])
    fin_js      = json.dumps([int(r.get("finishes",0) or 0) for r in mecze_chron])
    reb_js      = json.dumps([int(r.get("oreb",0) or 0)+int(r.get("dreb",0) or 0) for r in mecze_chron])
    oreb_js     = json.dumps([int(r.get("oreb",0) or 0) for r in mecze_chron])
    dreb_js     = json.dumps([int(r.get("dreb",0) or 0) for r in mecze_chron])
    stl_js      = json.dumps([int(r.get("stl",0) or 0) for r in mecze_chron])
    blk_js      = json.dumps([int(r.get("blk",0) or 0) for r in mecze_chron])
    efg_js      = json.dumps([round((int(r.get("p2m",0) or 0)+1.5*int(r.get("p3m",0) or 0))/(int(r.get("p2a",0) or 0)+int(r.get("p3a",0) or 0))*100,1) if (int(r.get("p2a",0) or 0)+int(r.get("p3a",0) or 0)) else None for r in mecze_chron])
    def _ts(r):
        p = int(r.get("pts",0) or 0)
        fga_r = int(r.get("p2a",0) or 0)+int(r.get("p3a",0) or 0)
        fta_r = int(r.get("fta",0) or 0)
        return round(p/(2*(fga_r+0.44*fta_r))*100,1) if (fga_r+fta_r) else None
    def _usg(r):
        fga_r = int(r.get("p2a",0) or 0)+int(r.get("p3a",0) or 0)
        fta_r = int(r.get("fta",0) or 0)
        br_r  = int(r.get("br",0) or 0)
        tp    = int(r.get("team_poss",0) or 0)
        return round((fga_r+0.44*fta_r+br_r)/tp*100,1) if tp else None
    ts_js       = json.dumps([_ts(r) for r in mecze_chron])
    usg_js      = json.dumps([_usg(r) for r in mecze_chron])
    p2pct_js    = round(pm2_tot/s('p2a')*100,1) if s('p2a') else 0
    p3pct_js    = round(pm3_tot/s('p3a')*100,1) if s('p3a') else 0
    ftpct_js    = round(ftm_tot/fta_tot*100,1) if fta_tot else 0
    efg_season  = round((pm2_tot+1.5*pm3_tot)/(fga_tot)*100,1) if fga_tot else 0
    ts_season   = round(pts_tot/(2*(fga_tot+0.44*fta_tot))*100,1) if (fga_tot+fta_tot) else 0
    usg_season  = round((fga_tot+0.44*fta_tot+br_tot)/poss_tot*100,1) if poss_tot else 0
    avg_pts     = round(pts_tot/n, 1)
    # Sumy sezonowe dla prawego wykresu per metryka
    ast_tot_s   = round(ast_tot/n, 1)
    br_tot_s    = round(br_tot/n, 1)
    fin_tot_s   = round(fin_tot/n, 1)
    reb_tot_s   = round((sum(int(r.get("oreb",0) or 0)+int(r.get("dreb",0) or 0) for r in mecze_stats))/n, 1)
    # Per mecz dla prawego wykresu (każda metryka: [min, avg, max])
    def season_bar(vals):
        nz = [v for v in vals if v is not None]
        if not nz: return [0, 0, 0]
        return [min(nz), round(sum(nz)/len(nz),1), max(nz)]
    pts_season  = season_bar([int(r.get("pts",0) or 0) for r in mecze_stats])
    ast_season  = season_bar([int(r.get("ast",0) or 0) for r in mecze_stats])
    br_season   = season_bar([int(r.get("br",0) or 0) for r in mecze_stats])
    fin_season  = season_bar([int(r.get("finishes",0) or 0) for r in mecze_stats])
    reb_season  = season_bar([int(r.get("oreb",0) or 0)+int(r.get("dreb",0) or 0) for r in mecze_stats])
    oreb_season = season_bar([int(r.get("oreb",0) or 0) for r in mecze_stats])
    dreb_season = season_bar([int(r.get("dreb",0) or 0) for r in mecze_stats])
    stl_season  = season_bar([int(r.get("stl",0) or 0) for r in mecze_stats])
    blk_season  = season_bar([int(r.get("blk",0) or 0) for r in mecze_stats])
    efg_season_bar = season_bar([_ts(r) for r in mecze_stats if _ts(r) is not None] or [0])
    ts_season_bar  = season_bar([_ts(r) for r in mecze_stats if _ts(r) is not None] or [0])
    usg_season_bar = season_bar([_usg(r) for r in mecze_stats if _usg(r) is not None] or [0])
    import json as _json
    pts_season_js  = _json.dumps(pts_season)
    ast_season_js  = _json.dumps(ast_season)
    br_season_js   = _json.dumps(br_season)
    fin_season_js  = _json.dumps(fin_season)
    reb_season_js  = _json.dumps(reb_season)
    oreb_season_js = _json.dumps(oreb_season)
    dreb_season_js = _json.dumps(dreb_season)
    stl_season_js  = _json.dumps(stl_season)
    blk_season_js  = _json.dumps(blk_season)
    efg_season_bar_js = _json.dumps(efg_season_bar)
    ts_season_bar_js  = _json.dumps(ts_season_bar)
    usg_season_bar_js = _json.dumps(usg_season_bar)
    # Zmiana 5: szerokość słupka zależna od liczby meczów
    bar_w = max(16, min(50, 300 // max(n, 1)))

    content = f"""
<div class="d-flex align-items-center gap-2 mb-3 flex-wrap">
  <a href="/zawodnicy?sezon={sezon_filter}" class="btn btn-outline-secondary btn-sm">← Statystyki</a>
  <div class="page-title mb-0">Profil zawodnika</div>
</div>

<!-- HEADER -->
<div class="card mb-3">
  <div class="card-body p-3">
    <div class="d-flex align-items-center gap-3 flex-wrap justify-content-between">
      <!-- Zmiana 1: nr w kółku zamiast initials gdy jeden numer -->
      <div style="display:flex;align-items:center;gap:16px;flex-shrink:0">
        <div style="width:56px;height:56px;border-radius:50%;background:#1a2b4a;display:flex;flex-direction:column;align-items:center;justify-content:center;flex-shrink:0">
          {"<span style='font-size:9px;color:rgba(255,255,255,.5);line-height:1'>#</span><span style='font-size:20px;font-weight:700;color:#fff;line-height:1.1'>" + nr_display + "</span>" if nr_display else "<span style='font-size:20px;font-weight:700;color:#fff'>" + initials + "</span>"}
        </div>
        <div>
          <div style="font-size:20px;font-weight:700;color:#1a2b4a">{zawodnik['imie']} {zawodnik['nazwisko']}</div>
          <div style="font-size:.82rem;color:#888;margin-top:2px">{gtk_name}</div>
          <div style="margin-top:6px">
            <span class="badge" style="background:{"#e8f5e9;color:#1a5c2a" if zawodnik.get('aktywny', True) else "#ffebee;color:#8b1a1a"}">{"Aktywny" if zawodnik.get('aktywny', True) else "Nieaktywny"}</span>
          </div>
        </div>
      </div>

      <!-- Quick stats bar -->
      <div style="display:flex;gap:0;background:#1a2b4a;border-radius:10px;padding:6px 10px;overflow-x:auto;-webkit-overflow-scrolling:touch;max-width:100%">
        {"".join([
          f'<div style="display:flex;flex-direction:column;align-items:center;padding:6px 14px;border-radius:8px;transition:background .15s" title="{tip}"><span style="font-size:17px;font-weight:500;color:#fff;line-height:1.2">{val}</span><span style="font-size:10px;color:rgba(255,255,255,.45);margin-top:3px;text-transform:uppercase;letter-spacing:.5px">{lbl}</span></div>'
          + ('' if i == 8 else '<div style="width:0.5px;background:rgba(255,255,255,.15);align-self:stretch;margin:4px 0"></div>')
          for i, (val, lbl, tip) in enumerate([
            (ppg, "PPG", "Średnia punktów na mecz"),
            (f"{ast_tot/n:.1f}", "AST", "Średnia asyst na mecz"),
            (f"{(sum(int(r.get('oreb',0) or 0)+int(r.get('dreb',0) or 0) for r in mecze_stats))/n:.1f}", "REB", "Średnia zbiórek na mecz (OFF + DEF)"),
            (f"{stl_tot/n:.1f}", "STL", "Średnia przechwytów na mecz"),
            (f"{blk_tot/n:.1f}", "BLK", "Średnia bloków na mecz"),
            (f"{br_tot/n:.1f}", "TO", "Średnia strat na mecz (TO = Turnover = BR = Brak Rzutu)"),
            (f"{fd_tot/n:.1f}", "FD", "Średnia faulów wymuszonych na mecz"),
            (f"{fin_tot/n:.1f}", "FIN", "Średnia wykończeń na mecz"),
            (str(n), "Mecze", "Liczba meczów w sezonie"),
          ])
        ])}
      </div>

      <form method="GET" class="d-flex gap-2 align-items-center ms-auto">
        <label style="font-size:.82rem;font-weight:600">Sezon:</label>
        <select name="sezon" class="form-select form-select-sm" style="width:120px" onchange="this.form.submit()">
          {season_opts}
        </select>
      </form>
    </div>
  </div>
</div>

<!-- KPI -->
<div class="row g-2 mb-3">{kpi_html}</div>

<!-- WYKRESY — Zmiana 4: przełącznik metryki -->
<div class="row g-3 mb-3">
  <div class="col-lg-7">
    <div class="card h-100"><div class="card-body p-2" style="position:relative">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:4px">
        <div class="section-hdr mb-0" id="chartMainTitle">Punkty</div>
        <div style="font-size:10px;color:#aaa;text-transform:uppercase;letter-spacing:.6px">Per mecz</div>
      </div>
      <div style="position:relative;height:180px">
        <canvas id="chartMain"></canvas>
        <div id="chartMainCenter" style="display:none;position:absolute;top:50%;left:50%;transform:translate(-50%,-65%);text-align:center;pointer-events:none">
          <div id="chartMainCenterVal" style="font-size:22px;font-weight:500;color:#1a2b4a;line-height:1.1"></div>
          <div style="font-size:11px;color:#888;margin-top:2px">śr. per mecz</div>
        </div>
      </div>
    </div></div>
  </div>
  <div class="col-lg-5">
    <div class="card h-100"><div class="card-body p-2" style="position:relative">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:4px">
        <div class="section-hdr mb-0" id="chartShootTitle">Skuteczność rzutów</div>
        <div style="font-size:10px;color:#aaa;text-transform:uppercase;letter-spacing:.6px">Sezon</div>
      </div>
      <div style="position:relative;height:180px">
        <canvas id="chartShoot"></canvas>
        <div id="chartShootCenter" style="display:none;position:absolute;top:50%;left:50%;transform:translate(-50%,-65%);text-align:center;pointer-events:none">
          <div id="chartShootCenterVal" style="font-size:22px;font-weight:500;color:#1a2b4a;line-height:1.1"></div>
          <div style="font-size:11px;color:#888;margin-top:2px">sezon</div>
        </div>
      </div>
    </div></div>
  </div>
</div>

<!-- TABELA MECZÓW -->
<div class="card mb-3"><div class="card-body p-2">
  <div class="section-hdr">Przebieg sezonu — mecz po meczu</div>
  <div class="table-responsive">
  <table class="table table-hover mb-0" id="tbl-mecze-profil" style="min-width:900px">
    <thead>
      <tr>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 8px;text-align:left;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2)" rowspan="3">Zawodnik</th>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2)" rowspan="3">Data</th>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2)" rowspan="3">MIN<br>(szac.)</th>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2);cursor:pointer" rowspan="3" id="thPTS" onclick="switchMetric('PTS')">PTS ↓</th>
        <th style="background:#1a2b4a;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center" colspan="2">2PT</th>
        <th style="background:#1a2b4a;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center" colspan="2">3PT</th>
        <th style="background:#1a2b4a;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center" colspan="2">FT</th>
        <th style="background:#152236;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.1);text-align:center" colspan="3">ZB</th>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2);cursor:pointer" rowspan="3" id="thAST" onclick="switchMetric('AST')">AST</th>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2);cursor:pointer" rowspan="3" id="thTO" onclick="switchMetric('TO')">TO</th>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2);cursor:pointer" rowspan="3" id="thSTL" onclick="switchMetric('STL')">STL</th>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2);cursor:pointer" rowspan="3" id="thBLK" onclick="switchMetric('BLK')">BLK</th>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2)" rowspan="3">FD</th>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2);cursor:pointer" rowspan="3" id="thEFG" onclick="switchMetric('EFG')">eFG%</th>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2);cursor:pointer" rowspan="3" id="thTS" onclick="switchMetric('TS')">TS%</th>
        <th style="background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 4px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2);cursor:pointer" rowspan="3" id="thUSG" onclick="switchMetric('USG')">USG%</th>
      </tr>
      <tr>
        <th style="background:#1a2b4a;color:rgba(255,255,255,.8);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center;white-space:nowrap">M/A</th>
        <th style="background:#1a2b4a;color:rgba(255,255,255,.8);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center;white-space:nowrap">%</th>
        <th style="background:#1a2b4a;color:rgba(255,255,255,.8);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center;white-space:nowrap">M/A</th>
        <th style="background:#1a2b4a;color:rgba(255,255,255,.8);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center;white-space:nowrap">%</th>
        <th style="background:#1a2b4a;color:rgba(255,255,255,.8);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center;white-space:nowrap">M/A</th>
        <th style="background:#1a2b4a;color:rgba(255,255,255,.8);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center;white-space:nowrap">%</th>
        <th id="thOREB" onclick="switchMetric('OREB')" style="background:#152236;color:rgba(255,255,255,.75);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center;white-space:nowrap;cursor:pointer">A</th>
        <th id="thDREB" onclick="switchMetric('DREB')" style="background:#152236;color:rgba(255,255,255,.75);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center;white-space:nowrap;cursor:pointer">O</th>
        <th id="thREB"  onclick="switchMetric('REB')"  style="background:#152236;color:rgba(255,255,255,.75);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center;white-space:nowrap;cursor:pointer">S</th>
      </tr>
    </thead>
    <tbody id="tbody-mecze-profil">{match_rows_rev}</tbody>
  </table>
  </div>
</div></div>
"""

    scripts = f"""<script>
  var labels = {labels_js};
  var nMecze = labels.length;
  var barW = Math.max(16, Math.min(50, Math.floor(280 / Math.max(nMecze, 1))));

  var dataSets = {{
    'PTS':  {pts_js},
    'AST':  {ast_js},
    'REB':  {reb_js},
    'OREB': {oreb_js},
    'DREB': {dreb_js},
    'STL':  {stl_js},
    'BLK':  {blk_js},
    'TO':   {br_js},
    'FIN':  {fin_js},
    'EFG':  {efg_js},
    'TS':   {ts_js},
    'USG':  {usg_js}
  }};
  var titles = {{
    'PTS':  'Punkty',
    'AST':  'Asysty',
    'REB':  'Rebounds',
    'OREB': 'Off. Reb',
    'DREB': 'Def. Reb',
    'STL':  'Przechwyty',
    'BLK':  'Bloki',
    'TO':   'Straty',
    'FIN':  'Wykończenia',
    'EFG':  'eFG%',
    'TS':   'TS%',
    'USG':  'USG%'
  }};
  var colors = {{
    'PTS':  '#1a2b4a',
    'AST':  '#378ADD',
    'REB':  '#1D9E75',
    'OREB': '#27ae60',
    'DREB': '#16875a',
    'STL':  '#185FA5',
    'BLK':  '#0C447C',
    'TO':   '#D85A30',
    'FIN':  '#555555',
    'EFG':  '#534AB7',
    'TS':   '#0F6E56',
    'USG':  '#D85A30'
  }};
  var ptsAvg = {avg_pts};
  var pieMetrics = new Set(['EFG','TS','USG']);

  // dane dla wykresu kołowego per mecz
  var piePerMatch = {{
    'EFG': {{ data: {efg_js}, color: '#534AB7' }},
    'TS':  {{ data: {ts_js},  color: '#0F6E56' }},
    'USG': {{ data: {usg_js}, color: '#D85A30' }}
  }};

  // dane sezonowe do wykresu kołowego (prawy panel)
  var pieSeasonData = {{
    'EFG': [{efg_season}, {round(100-efg_season,1)}],
    'TS':  [{ts_season},  {round(100-ts_season,1)}],
    'USG': [{usg_season}, {round(100-usg_season,1)}]
  }};
  var pieSeasonColors = {{
    'EFG': ['#534AB7','#EEEDFE'],
    'TS':  ['#0F6E56','#E1F5EE'],
    'USG': ['#D85A30','#FAECE7']
  }};
  var pieSeasonLabels = {{
    'EFG': ['eFG%','pozostałe'],
    'TS':  ['TS%','pozostałe'],
    'USG': ['USG%','pozostałe']
  }};
  var pieSeasonTitles = {{
    'EFG': 'eFG% (sezon: {efg_season}%)',
    'TS':  'TS% (sezon: {ts_season}%)',
    'USG': 'USG% (sezon: {usg_season}%)'
  }};

  var mainChart = null;

  function getBarColors(metric, data) {{
    var c = colors[metric] || '#1a2b4a';
    return data.map(function() {{ return c; }});
  }}

  function switchMetric(metric) {{
    ['PTS','AST','REB','OREB','DREB','STL','BLK','TO','FIN','EFG','TS','USG'].forEach(function(m) {{
      var th = document.getElementById('th' + m);
      if (th) {{
        th.style.background = (m === metric) ? '#1a2b4a' : '';
        th.style.color = (m === metric) ? '#fff' : '';
      }}
    }});
    var titleEl = document.getElementById('chartMainTitle');
    if (titleEl) titleEl.textContent = titles[metric] || metric;
    if (mainChart) {{ mainChart.destroy(); mainChart = null; }}
    var canvas = document.getElementById('chartMain');
    if (!canvas) return;

    if (pieMetrics.has(metric)) {{
      // WYKRES KOŁOWY per mecz — pokaż center label
      var pd = piePerMatch[metric];
      var validData = pd.data.filter(function(v){{ return v !== null; }});
      var avgVal = validData.length ? (validData.reduce(function(a,b){{return a+b;}},0)/validData.length).toFixed(1) : '0.0';
      var cl = document.getElementById('chartMainCenter');
      var cv = document.getElementById('chartMainCenterVal');
      if (cl) cl.style.display = 'block';
      if (cv) cv.textContent = avgVal + '%';
      mainChart = new Chart(canvas.getContext('2d'), {{
        type: 'doughnut',
        data: {{
          labels: [metric + '% (śr.)', 'pozostałe'],
          datasets: [{{ data: [parseFloat(avgVal), Math.max(0, 100-parseFloat(avgVal))],
            backgroundColor: [pd.color, pd.color+'33'],
            borderWidth: 0 }}]
        }},
        options: {{
          responsive: true, maintainAspectRatio: false, cutout: '68%',
          plugins: {{
            legend: {{ display: true, position: 'bottom', labels: {{ font: {{ size: 11 }}, boxWidth: 12 }} }},
            tooltip: {{ callbacks: {{ label: function(c) {{ return c.parsed.toFixed(1)+'%'; }} }} }}
          }}
        }}
      }});
    }} else {{
      // WYKRES SŁUPKOWY — ukryj center label
      var cl = document.getElementById('chartMainCenter');
      if (cl) cl.style.display = 'none';
      var data = dataSets[metric];
      var avg = data.reduce(function(a,b){{return a+(b||0);}}, 0) / Math.max(data.length, 1);
      if (metric === 'PTS') avg = ptsAvg;
      var datasets = [{{
        type: 'bar', label: metric, data: data,
        backgroundColor: getBarColors(metric, data),
        borderRadius: 3, barThickness: barW, order: 2
      }}];
      if (data.length > 1) {{
        datasets.push({{
          type: 'line', label: 'Średnia',
          data: labels.map(function() {{ return Math.round(avg * 10) / 10; }}),
          borderColor: '#D85A30', borderWidth: 2, borderDash: [4, 3],
          pointRadius: 0, fill: false, order: 1
        }});
      }}
      mainChart = new Chart(canvas.getContext('2d'), {{
        type: 'bar',
        data: {{ labels: labels, datasets: datasets }},
        options: {{
          responsive: true, maintainAspectRatio: false,
          plugins: {{
            legend: {{ display: false }},
            tooltip: {{ callbacks: {{ label: function(ctx) {{ return ctx.dataset.label + ': ' + ctx.parsed.y; }} }} }}
          }},
          scales: {{
            x: {{ ticks: {{ font: {{ size: 9 }}, maxRotation: 40, autoSkip: false }} }},
            y: {{ min: 0, ticks: {{ font: {{ size: 10 }} }}, grid: {{ color: 'rgba(0,0,0,0.05)' }} }}
          }}
        }}
      }});
    }}
    if (typeof updateShootChart === 'function') updateShootChart(metric);
  }}

  // Dane dla prawego wykresu
  var shootData = {{
    'PTS': {{
      labels: ['2PT%', '3PT%', 'FT%'],
      data:   [{p2pct_js}, {p3pct_js}, {ftpct_js}],
      colors: ['#1a2b4a','#378ADD','#1D9E75'],
      title:  'Skuteczność rzutów',
      unit: '%', type: 'bar'
    }},
    'AST': {{ labels: ['Min', 'Śr.', 'Max'], data: {ast_season_js}, colors: ['#aad4f5','#378ADD','#1a5fa0'], title: 'Asysty', unit: '', type: 'bar' }},
    'REB':  {{ labels: ['Min', 'Śr.', 'Max'], data: {reb_season_js},  colors: ['#a8e6cf','#1D9E75','#0d6b4f'], title: 'Rebounds', unit: '', type: 'bar' }},
    'OREB': {{ labels: ['Min', 'Śr.', 'Max'], data: {oreb_season_js}, colors: ['#b6e8cc','#27ae60','#1a6e3f'], title: 'Off. Reb',  unit: '', type: 'bar' }},
    'DREB': {{ labels: ['Min', 'Śr.', 'Max'], data: {dreb_season_js}, colors: ['#a8e6cf','#16875a','#0a4f35'], title: 'Def. Reb',  unit: '', type: 'bar' }},
    'TO':   {{ labels: ['Min', 'Śr.', 'Max'], data: {br_season_js},   colors: ['#f5c6a0','#D85A30','#9e3a18'], title: 'Straty', unit: '', type: 'bar' }},
    'FIN': {{ labels: ['Min', 'Śr.', 'Max'], data: {fin_season_js}, colors: ['#c5ccd6','#555','#1a2b4a'],    title: 'Wykończenia', unit: '', type: 'bar' }},
    'STL': {{ labels: ['Min', 'Śr.', 'Max'], data: {stl_season_js}, colors: ['#b5d4f4','#185FA5','#0C447C'], title: 'Przechwyty', unit: '', type: 'bar' }},
    'BLK': {{ labels: ['Min', 'Śr.', 'Max'], data: {blk_season_js}, colors: ['#b5d4f4','#185FA5','#042C53'], title: 'Bloki', unit: '', type: 'bar' }},
    'EFG': {{ labels: pieSeasonLabels['EFG'], data: pieSeasonData['EFG'], colors: pieSeasonColors['EFG'], title: 'eFG%', unit: '%', type: 'doughnut' }},
    'TS':  {{ labels: pieSeasonLabels['TS'],  data: pieSeasonData['TS'],  colors: pieSeasonColors['TS'],  title: 'TS%',  unit: '%', type: 'doughnut' }},
    'USG': {{ labels: pieSeasonLabels['USG'], data: pieSeasonData['USG'], colors: pieSeasonColors['USG'], title: 'USG%', unit: '%', type: 'doughnut' }}
  }};

  var shootChart = null;

  function updateShootChart(metric) {{
    var cfg = shootData[metric] || shootData['PTS'];
    var titleEl = document.getElementById('chartShootTitle');
    if (titleEl) titleEl.textContent = cfg.title;
    if (shootChart) {{ shootChart.destroy(); shootChart = null; }}
    var canvas = document.getElementById('chartShoot');
    if (!canvas) return;

    if (cfg.type === 'doughnut') {{
      var scl = document.getElementById('chartShootCenter');
      var scv = document.getElementById('chartShootCenterVal');
      if (scl) scl.style.display = 'block';
      if (scv) scv.textContent = cfg.data[0].toFixed(1) + '%';
      shootChart = new Chart(canvas.getContext('2d'), {{
        type: 'doughnut',
        data: {{ labels: cfg.labels, datasets: [{{ data: cfg.data, backgroundColor: cfg.colors, borderWidth: 0 }}] }},
        options: {{
          responsive: true, maintainAspectRatio: false, cutout: '68%',
          plugins: {{
            legend: {{ display: true, position: 'bottom', labels: {{ font: {{ size: 11 }}, boxWidth: 12 }} }},
            tooltip: {{ callbacks: {{ label: function(c) {{ return c.parsed.toFixed(1)+'%'; }} }} }}
          }}
        }}
      }});
    }} else {{
      var scl = document.getElementById('chartShootCenter');
      if (scl) scl.style.display = 'none';
      var isPct = cfg.unit === '%';
      shootChart = new Chart(canvas.getContext('2d'), {{
        type: 'bar',
        data: {{ labels: cfg.labels, datasets: [{{ data: cfg.data, backgroundColor: cfg.colors, borderRadius: 4, barThickness: 40 }}] }},
        options: {{
          responsive: true, maintainAspectRatio: false,
          plugins: {{
            legend: {{ display: false }},
            tooltip: {{ callbacks: {{ label: function(c) {{ return isPct ? c.parsed.y.toFixed(1)+'%' : c.parsed.y; }} }} }}
          }},
          scales: {{
            x: {{ ticks: {{ font: {{ size: 12 }} }} }},
            y: {{ min: 0, max: isPct ? 100 : undefined,
              ticks: {{ font: {{ size: 10 }}, callback: function(v) {{ return isPct ? v+'%' : v; }} }},
              grid: {{ color: 'rgba(0,0,0,0.05)' }}
            }}
          }}
        }}
      }});
    }}
  }}

  // Inicjalizacja Bootstrap tooltipów na kafelkach KPI
  document.querySelectorAll('[data-bs-toggle="tooltip"]').forEach(function(el) {{
    new bootstrap.Tooltip(el, {{ trigger: 'hover', html: false }});
  }});

  // Inicjalizacja domyślna
  switchMetric('PTS');
</script>"""

    return html_response(base(content, scripts, active="players"))

# ══════════════════════════════════════════════════════════════════════════════
# USTAWIENIA
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/set_context", methods=["POST"])
@login_required
def set_context():
    import json as _json
    klub    = request.form.get("klub", "")
    sezon   = request.form.get("sezon", "")
    druzyna = request.form.get("druzyna", "")
    redirect_to = request.form.get("redirect", "/")

    if klub:    set_setting("current_klub", klub)
    if sezon:   set_setting("current_season", sezon)
    if druzyna: set_setting("current_druzyna", druzyna)

    if not redirect_to or redirect_to == "/set_context":
        redirect_to = "/"
    return redirect(redirect_to)



@app.route("/set_context_ajax", methods=["POST"])
@login_required
def set_context_ajax():
    import json as _j
    try:
        data  = request.get_json()
        field = data.get("field","")
        val   = data.get("value","")
        if field == "klub":
            set_setting("current_klub",    val)
            set_setting("current_season",  "")
            set_setting("current_druzyna", "")
        elif field == "sezon":
            set_setting("current_season",  val)
            set_setting("current_druzyna", "")
        elif field == "druzyna":
            set_setting("current_druzyna", val)
        return _j.dumps({"ok": True})
    except Exception as e:
        return _j.dumps({"ok": False, "error": str(e)}), 400



@app.route("/clear_context", methods=["POST", "GET"])
@login_required
def clear_context():
    set_setting("current_klub",    "")
    set_setting("current_season",  "")
    set_setting("current_druzyna", "")
    return redirect(request.referrer or "/")


@app.route("/admin/upload-logo", methods=["POST"])
@login_required
def upload_logo():
    import base64 as _b64
    if "logo" not in request.files:
        flash("Brak pliku", "danger")
        return redirect(url_for("ustawienia"))
    f = request.files["logo"]
    if not f or f.filename == "":
        flash("Brak pliku", "danger")
        return redirect(url_for("ustawienia"))
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in {"png", "jpg", "jpeg", "webp"}:
        flash("Dozwolone formaty: PNG, JPG, WEBP", "danger")
        return redirect(url_for("ustawienia"))
    klub = request.form.get("klub") or get_setting("current_klub") or ""
    if not klub:
        flash("Brak aktywnego klubu — ustaw kontekst przed wgraniem logo.", "danger")
        return redirect(url_for("ustawienia"))
    mime = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg", "webp": "image/webp"}.get(ext, "image/png")
    img_bytes = f.read()
    data_uri = f"data:{mime};base64," + _b64.b64encode(img_bytes).decode()
    safe_klub = klub.replace(" ", "_").replace("/", "-")
    set_setting(f"logo_{safe_klub}", data_uri)
    flash(f"Logo dla «{klub}» zapisane!", "success")
    return redirect(url_for("ustawienia"))


@app.route("/admin/upload-app-logo", methods=["POST"])
@login_required
def upload_app_logo():
    import base64 as _b64
    if "logo" not in request.files:
        flash("Brak pliku", "danger")
        return redirect(url_for("ustawienia"))
    f = request.files["logo"]
    if not f or f.filename == "":
        flash("Brak pliku", "danger")
        return redirect(url_for("ustawienia"))
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in {"png", "jpg", "jpeg", "webp"}:
        flash("Dozwolone formaty: PNG, JPG, WEBP", "danger")
        return redirect(url_for("ustawienia"))
    mime = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg", "webp": "image/webp"}.get(ext, "image/png")
    data_uri = f"data:{mime};base64," + _b64.b64encode(f.read()).decode()
    set_setting("app_logo_b64", data_uri)
    flash("Logo platformy zapisane!", "success")
    return redirect(url_for("ustawienia"))


@app.route("/admin/publish-portal", methods=["POST"])
@login_required
def publish_portal():
    set_setting("portal_klub",    get_setting("current_klub")    or "")
    set_setting("portal_sezon",   get_setting("current_season")  or "")
    set_setting("portal_druzyna", get_setting("current_druzyna") or "")
    flash("Portal zaktualizowany!", "success")
    return redirect(url_for("ustawienia"))


@app.route("/druzyny/js")
def druzyny_js():
    """Serve the druzyny JavaScript file"""
    from flask import Response
    js = open('/app/druzyny.js').read() if __import__('os').path.exists('/app/druzyny.js') else DRUZYNY_JS
    return Response(js, mimetype='application/javascript')


@app.route("/druzyny")
@login_required
def druzyny():
    try: init_db()
    except: pass
    import json as _j
    kj = get_setting("kluby_json") or "[]"
    try:    kluby = _j.loads(kj)
    except: kluby = []
    db_json = _j.dumps(kluby, ensure_ascii=True)

    page = """
<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;min-height:520px">
  <div style="background:#fff;border-radius:12px;box-shadow:0 1px 6px rgba(0,0,0,.08);overflow:hidden;display:flex;flex-direction:column">
    <div style="background:#1a2b4a;padding:10px 14px;display:flex;justify-content:space-between;align-items:center;flex-shrink:0">
      <span style="color:#fff;font-size:11px;font-weight:600;letter-spacing:.5px;text-transform:uppercase">Struktura klub&#243;w</span>
      <button id="btn-add" style="background:#EF9F27;color:#fff;border:none;padding:4px 14px;border-radius:6px;cursor:pointer;font-size:12px;font-weight:600">+ Dodaj</button>
    </div>
    <div style="padding:12px;flex:1;overflow-y:auto" id="tree-root"></div>
  </div>
  <div style="background:#fff;border-radius:12px;box-shadow:0 1px 6px rgba(0,0,0,.08);overflow:hidden;display:flex;flex-direction:column">
    <div id="sbar" style="display:none;align-items:center;padding:10px 14px 9px;border-bottom:1px solid #f0f0f0;flex-shrink:0"></div>
    <div id="rbody" style="flex:1;padding:40px 20px;text-align:center;color:#bbb;font-size:12px;line-height:2">
      Kliknij &#8943; przy klubie aby edytowa&#263;. Kliknij + Dodaj aby doda&#263; nowy klub.
    </div>
    <div id="rfoot" style="display:none;padding:10px 14px;gap:8px;flex-shrink:0"></div>
  </div>
</div>
<div id="ov-del" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:9000;align-items:center;justify-content:center">
  <div style="background:#fff;border-radius:12px;padding:24px;max-width:300px;width:90%">
    <div id="ov-title" style="font-size:15px;font-weight:600;color:#1a2b4a;margin-bottom:6px;text-align:center"></div>
    <div id="ov-msg" style="font-size:12px;color:#888;margin-bottom:18px;line-height:1.6;text-align:center"></div>
    <div style="display:flex;gap:8px">
      <button id="ov-cancel" style="flex:1;background:none;border:1px solid #ddd;color:#888;padding:8px;border-radius:8px;cursor:pointer;font-size:12px">Anuluj</button>
      <button id="ov-ok" style="flex:1;background:#E24B4A;color:#fff;border:none;padding:8px;border-radius:8px;cursor:pointer;font-size:12px;font-weight:600">Usu&#324;</button>
    </div>
  </div>
</div>
<div id="ov-player" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:9000;align-items:center;justify-content:center">
  <div style="background:#fff;border-radius:12px;padding:22px;max-width:360px;width:90%">
    <div id="pm-title" style="font-size:15px;font-weight:600;color:#1a2b4a;margin-bottom:3px">Nowy zawodnik</div>
    <div id="pm-sub" style="font-size:12px;color:#888;margin-bottom:14px"></div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px">
      <div><div style="font-size:11px;color:#666;margin-bottom:3px">Imi&#281;</div>
        <input id="pm-imie" placeholder="np. Jan" style="width:100%;padding:7px 10px;border:1px solid #e0e0e0;border-radius:7px;font-size:12px;box-sizing:border-box"></div>
      <div><div style="font-size:11px;color:#666;margin-bottom:3px">Nazwisko</div>
        <input id="pm-nazw" placeholder="np. Kowalski" style="width:100%;padding:7px 10px;border:1px solid #e0e0e0;border-radius:7px;font-size:12px;box-sizing:border-box"></div>
      <div><div style="font-size:11px;color:#666;margin-bottom:3px">Numer</div>
        <input id="pm-num" type="number" placeholder="5" style="width:100%;padding:7px 10px;border:1px solid #e0e0e0;border-radius:7px;font-size:12px;box-sizing:border-box"></div>
      <div><div style="font-size:11px;color:#666;margin-bottom:3px">Pozycja</div>
        <select id="pm-poz" style="width:100%;padding:7px 10px;border:1px solid #e0e0e0;border-radius:7px;font-size:12px;background:#fff;box-sizing:border-box">
          <option value="" disabled selected>— wybierz pozycję —</option>
          <option>Rozgrywaj&#261;cy</option><option>Rzucaj&#261;cy</option>
          <option>Ma&#322;y skrzyd&#322;owy</option><option>Skrzyd&#322;owy</option>
          <option>Silny skrzyd&#322;owy</option><option>&#346;rodkowy</option>
        </select></div>
    </div>
    <div style="margin-bottom:12px">
      <div style="font-size:11px;color:#666;margin-bottom:3px">Status</div>
      <select id="pm-aktywny" style="width:100%;padding:7px 10px;border:1px solid #e0e0e0;border-radius:7px;font-size:12px;background:#fff;box-sizing:border-box">
        <option value="1">Aktywny</option>
        <option value="0">Nieaktywny</option>
      </select>
    </div>
    <div style="display:flex;gap:8px">
      <button id="pm-cancel" style="flex:1;background:none;border:1px solid #ddd;color:#888;padding:8px;border-radius:8px;cursor:pointer;font-size:12px">Anuluj</button>
      <button id="pm-save" style="flex:2;background:#EF9F27;color:#fff;border:none;padding:8px;border-radius:8px;cursor:pointer;font-size:12px;font-weight:600">Zapisz zawodnika</button>
    </div>
  </div>
</div>
"""

    import json as _j2
    # Build PLAYERS_INIT from proper DB tables
    # Structure: {club_index: {season_name: {team_name: [{imie,nazwisko,num,poz}]}}}
    players_data = {}
    try:
        db2 = get_db()
        cur2 = db2.cursor()
        for ki, klub in enumerate(kluby):
            kname = klub.get("name","")
            cur2.execute("SELECT id FROM clubs WHERE name=%s", (kname,))
            cr = cur2.fetchone()
            if not cr: continue
            club_id = cr["id"]
            players_data[str(ki)] = {}
            sezony = klub.get("sezony", {})
            for sname in sezony:
                cur2.execute("SELECT id FROM seasons WHERE club_id=%s AND name=%s", (club_id, sname))
                sr = cur2.fetchone()
                if not sr: continue
                season_id = sr["id"]
                players_data[str(ki)][sname] = {}
                for tname in sezony[sname]:
                    cur2.execute("SELECT id FROM teams WHERE season_id=%s AND name=%s", (season_id, tname))
                    tr = cur2.fetchone()
                    if not tr:
                        players_data[str(ki)][sname][tname] = []
                        continue
                    team_id = tr["id"]
                    cur2.execute("SELECT imie, nazwisko, numer, pozycja FROM players WHERE team_id=%s ORDER BY id", (team_id,))
                    rows = cur2.fetchall()
                    players_data[str(ki)][sname][tname] = [
                        {"imie": r["imie"], "nazwisko": r["nazwisko"],
                         "num": r["numer"], "poz": r["pozycja"]}
                        for r in rows
                    ]
        cur2.close()
    except Exception:
        players_data = {}
    players_json = _j2.dumps(players_data, ensure_ascii=True)

    init_script = ("<script>var DB_INIT=" + db_json + ";"
                   "var PLAYERS_INIT=" + players_json + ";</script>")
    js_include = '<script src="/druzyny/js"></script>'

    return html_response(base(page + init_script + js_include, active="druzyny"))


@app.route("/druzyny/save", methods=["POST"])
@login_required
def druzyny_save():
    import json as _j
    try:
        data = request.get_json()
        db = get_db()
        cur = db.cursor()

        if "kluby" in data:
            # Save to settings for fast sidebar access
            set_setting("kluby_json", _j.dumps(data["kluby"], ensure_ascii=False))

            # Sync to proper tables
            kluby = data["kluby"]
            for ki, klub in enumerate(kluby):
                kname = klub.get("name","")
                kext  = bool(klub.get("ext", False))
                cur.execute("""INSERT INTO clubs (name, ext) VALUES (%s,%s)
                               ON CONFLICT (name) DO UPDATE SET ext=%s
                               RETURNING id""", (kname, kext, kext))
                club_row = cur.fetchone()
                club_id = club_row["id"]

                sezony = klub.get("sezony", {})
                for sname, teams in sezony.items():
                    cur.execute("""INSERT INTO seasons (club_id, name) VALUES (%s,%s)
                                   ON CONFLICT (club_id, name) DO NOTHING
                                   RETURNING id""", (club_id, sname))
                    sr = cur.fetchone()
                    if not sr:
                        cur.execute("SELECT id FROM seasons WHERE club_id=%s AND name=%s", (club_id, sname))
                        sr = cur.fetchone()
                    season_id = sr["id"]

                    for tname in teams:
                        cur.execute("""INSERT INTO teams (season_id, name) VALUES (%s,%s)
                                       ON CONFLICT (season_id, name) DO NOTHING""", (season_id, tname))

        if "players" in data:
            players = data["players"]
            # players structure: {club_index: {season: {team: [{imie,nazwisko,num,poz}]}}}
            # Convert to use DB ids
            for ki_str, seasons in players.items():
                ki = int(ki_str)
                # Get club from kluby_json by index
                kluby_raw = get_setting("kluby_json") or "[]"
                kluby_list = _j.loads(kluby_raw)
                if ki >= len(kluby_list): continue
                kname = kluby_list[ki]["name"]

                cur.execute("SELECT id FROM clubs WHERE name=%s", (kname,))
                cr = cur.fetchone()
                if not cr: continue
                club_id = cr["id"]

                for sname, teams in seasons.items():
                    cur.execute("SELECT id FROM seasons WHERE club_id=%s AND name=%s", (club_id, sname))
                    sr = cur.fetchone()
                    if not sr: continue
                    season_id = sr["id"]

                    for tname, pl_list in teams.items():
                        cur.execute("SELECT id FROM teams WHERE season_id=%s AND name=%s", (season_id, tname))
                        tr = cur.fetchone()
                        if not tr: continue
                        team_id = tr["id"]

                        # Delete existing players for this team and re-insert
                        cur.execute("DELETE FROM players WHERE team_id=%s", (team_id,))
                        for pl in pl_list:
                            cur.execute("""INSERT INTO players (team_id, imie, nazwisko, numer, pozycja)
                                           VALUES (%s,%s,%s,%s,%s)""",
                                        (team_id,
                                         pl.get("imie",""),
                                         pl.get("nazwisko",""),
                                         int(pl.get("num",0)),
                                         pl.get("poz","")))

        # Keep players_json in sync as backup
        if "players" in data:
            set_setting("players_json", _j.dumps(data["players"], ensure_ascii=False))

        db.commit()
        cur.close()
        return _j.dumps({"ok":True})
    except Exception as e:
        try: get_db().rollback()
        except: pass
        return _j.dumps({"ok":False,"error":str(e)}), 400


@app.route("/ustawienia", methods=["GET","POST"])
@login_required
def ustawienia():
    import json as _json
    if request.method == "POST":
        set_setting("gtk_name", request.form.get("gtk_name","GTK"))
        # current_season managed by context selectors
        flash("Ustawienia zapisane!","success")
        return redirect(url_for("ustawienia"))

    gtk_name   = get_setting("gtk_name") or "GTK"
    season     = get_setting("current_season") or ""
    klub       = get_setting("current_klub") or ""
    druzyna    = get_setting("current_druzyna") or ""
    kluby_json = get_setting("kluby_json") or "[]"


    content = f"""
<div class="page-title">⚙️ Ustawienia</div>
<div class="row g-3">
<div class="col-lg-6">

  <div class="card p-3 mb-3" style="border-color:#c9a340">
    <div class="section-hdr" style="color:#c9a340">&#127760; Portal zawodnika</div>
    <div class="mt-2 mb-2" style="font-size:.78rem;color:#6b7280">Portal aktualnie pokazuje:</div>
    <div class="d-flex gap-2 flex-wrap mb-3">
      {'<span class="badge" style="background:#1a2b4a;color:#c9a340;font-size:.8rem;padding:5px 12px">' + (get_setting("portal_klub") or "<i>nie ustawiono</i>") + '</span>'}
      {'<span class="badge" style="background:#1a2b4a;color:#c9a340;font-size:.8rem;padding:5px 12px">' + (get_setting("portal_sezon") or "<i>nie ustawiono</i>") + '</span>'}
      {'<span class="badge" style="background:#1a2b4a;color:#c9a340;font-size:.8rem;padding:5px 12px">' + (get_setting("portal_druzyna") or "<i>nie ustawiono</i>") + '</span>'}
    </div>
    <div class="mb-2" style="font-size:.78rem;color:#6b7280">Twój aktywny kontekst:</div>
    <div class="d-flex gap-2 flex-wrap mb-3">
      {'<span class="badge" style="background:#e6f1fb;color:#0c447c;font-size:.8rem;padding:5px 12px">' + klub + '</span>' if klub else '<span style="font-size:.8rem;color:#aaa;font-style:italic">brak</span>'}
      {'<span class="badge" style="background:#e8f5e9;color:#1a5c2a;font-size:.8rem;padding:5px 12px">' + season + '</span>' if season else ''}
      {'<span class="badge" style="background:#fff4e5;color:#854F0B;font-size:.8rem;padding:5px 12px">' + druzyna + '</span>' if druzyna else ''}
    </div>
    <form method="POST" action="/admin/publish-portal">
      <button type="submit" class="btn btn-warning w-100 fw-bold"
              {'disabled' if not (klub and season and druzyna) else ''}
              onclick="return confirm('Opublikować {druzyna} / {season} do portalu?')">
        &#128640; Opublikuj {druzyna or '—'} / {season or '—'} do portalu
      </button>
    </form>
    {'<div class="form-text mt-1 text-center">Ustaw kontekst w sidebarze aby opublikować.</div>' if not (klub and season and druzyna) else ''}
  </div>

  <div class="card p-3 mb-3">
    <div class="section-hdr">Aktywny kontekst</div>
    <div class="d-flex gap-2 flex-wrap mt-2">
      {'<span class="badge" style="background:#e6f1fb;color:#0c447c;font-size:.8rem;padding:5px 12px">' + klub + '</span>' if klub else ''}
      {'<span class="badge" style="background:#e8f5e9;color:#1a5c2a;font-size:.8rem;padding:5px 12px">' + season + '</span>' if season else ''}
      {'<span class="badge" style="background:#fff4e5;color:#854F0B;font-size:.8rem;padding:5px 12px">' + druzyna + '</span>' if druzyna else ''}
      {'<span style="font-size:.8rem;color:#aaa;font-style:italic">Brak aktywnego kontekstu</span>' if not klub and not season and not druzyna else ''}
    </div>
    <div class="form-text mt-2">Zmień kontekst przez selektory w sidebarze. Zarządzaj strukturą w <a href="/druzyny">Struktura klubów</a>.</div>
    <form method="POST" action="/clear_context" class="mt-2">
      <button type="submit" class="btn btn-sm btn-outline-danger">Wyczyść kontekst</button>
    </form>
  </div>

  <div class="card p-3 mb-3" style="border-color:#f5c6cb">
    <div class="section-hdr" style="color:#721c24">Strefa niebezpieczna</div>
    <div class="form-text mt-1 mb-2">Usuwa wszystkie mecze i statystyki z bazy. Operacji nie można cofnąć.</div>
    <form method="POST" action="/admin/reset-clubs" class="mb-2" onsubmit="return confirm('Wyczyścić i odświeżyć strukturę klubów z aktualnych ustawień?')">
      <button type="submit" class="btn btn-sm btn-outline-warning">&#8635; Odśwież strukturę klubów w bazie</button>
    </form>
    <form method="POST" action="/admin/reset-matches" onsubmit="return confirm('Usunąć WSZYSTKIE mecze i statystyki? Tej operacji nie można cofnąć.')">
      <button type="submit" class="btn btn-sm btn-danger">&#128465; Usuń wszystkie mecze i statystyki</button>
    </form>
  </div>

  <div class="card p-3 mb-3">
    <div class="section-hdr">Logo klubu</div>
    {'<div class="alert alert-warning py-2 mt-2" style="font-size:.8rem">Ustaw aktywny klub w sidebarze, aby przypisać logo.</div>' if not klub else f'''
    <div class="mt-2 mb-2" style="font-size:.75rem;color:#6b7280">Aktywny klub: <strong>{klub}</strong></div>
    <div class="mb-3 d-flex align-items-center gap-3">
      <img src="{get_setting(f'logo_{klub.replace(" ","_").replace("/","-")}') or ''}" id="logo-preview"
           style="height:64px;width:64px;object-fit:contain;border-radius:8px;background:#1a2b4a;padding:4px{';display:none' if not get_setting(f'logo_{klub.replace(" ","_").replace("/","-")}') else ''}">
      <div style="font-size:.8rem;color:#6b7280">Logo wyświetlane w portalu zawodnika</div>
    </div>
    <form method="POST" action="/admin/upload-logo" enctype="multipart/form-data">
      <input type="hidden" name="klub" value="{klub}">
      <input type="file" name="logo" accept="image/png,image/jpeg,image/webp"
             class="form-control form-control-sm mb-2"
             onchange="var r=new FileReader();r.onload=function(e){{var el=document.getElementById('logo-preview');el.src=e.target.result;el.style.display='block';}}; r.readAsDataURL(this.files[0])">
      <button type="submit" class="btn btn-sm btn-primary w-100">Wgraj logo dla {klub}</button>
    </form>'''}
  </div>

  <div class="card p-3 mb-3">
    <div class="section-hdr">Logo platformy (sidebar portalu)</div>
    <div class="mt-2 mb-3 d-flex align-items-center gap-3">
      {'<img src="' + (get_setting("app_logo_b64") or "") + '" id="app-logo-preview" style="height:56px;width:56px;object-fit:contain;border-radius:10px;background:#1a2b4a;padding:4px">' if get_setting("app_logo_b64") else '<div id="app-logo-preview" style="height:56px;width:56px;background:#1a2b4a;border-radius:10px;display:flex;align-items:center;justify-content:center;font-weight:900;color:#c9a340;font-size:.9rem">BK</div>'}
      <div style="font-size:.8rem;color:#6b7280">Wyświetlane w lewym sidebarze portalu</div>
    </div>
    <form method="POST" action="/admin/upload-app-logo" enctype="multipart/form-data">
      <input type="file" name="logo" accept="image/png,image/jpeg,image/webp"
             class="form-control form-control-sm mb-2"
             onchange="var r=new FileReader();r.onload=function(e){{var el=document.getElementById('app-logo-preview');el.style.cssText='height:56px;width:56px;object-fit:contain;border-radius:10px;background:#1a2b4a;padding:4px';el.outerHTML='<img id=\\'app-logo-preview\\' src=\\''+e.target.result+'\\' style=\\'height:56px;width:56px;object-fit:contain;border-radius:10px;background:#1a2b4a;padding:4px\\'>';}}; r.readAsDataURL(this.files[0])">
      <button type="submit" class="btn btn-sm btn-primary w-100">Wgraj logo platformy</button>
    </form>
  </div>

  <div class="card p-3 mb-3">
    <div class="section-hdr">Podstawowe ustawienia</div>
    <form method="POST">
      <div class="mb-3 mt-2">
        <label class="form-label fw-bold">Nazwa wyświetlana drużyny głównej</label>
        <input type="text" name="gtk_name" class="form-control" value="{gtk_name}" placeholder="np. GTK Gliwice U13">
        <div class="form-text">Nazwa w nagłówkach raportów i tabel.</div>
      </div>
      <button type="submit" class="btn btn-primary w-100">Zapisz</button>
    </form>
  </div>

  <div class="card p-3">
    <div class="section-hdr">Pobierz szablony</div>
    <div class="row g-2 mt-1">
      <div class="col-6">
        <a href="/template/zapis" class="btn btn-outline-primary w-100 btn-sm">📝 Zapis meczu</a>
      </div>
      <div class="col-6">
        <a href="/template/szablon" class="btn btn-outline-success w-100 btn-sm">📋 Szablon raportu</a>
      </div>
    </div>
  </div>

</div>
</div>"""

    return html_response(base(content, active="settings"))

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT (z match_id)
# ══════════════════════════════════════════════════════════════════════════════

def _get_szablon_b64():
    return (
    "UEsDBBQABgAIAAAAIQDFxOklggEAALgHAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADMlc9OAjEQxu8mvsOmV8MWUIkxLBxQj0oiPkBtZ9mG"
    "btu0BeHtnS1/YswKIWziXrbZdub7fp2kM8PxulTJCpyXRmekl3ZJApobIfU8Ix+zl84DSXxgWjBlNGRk"
    "A56MR9dXw9nGgk8wW/uMFCHYR0o9L6BkPjUWNJ7kxpUs4K+bU8v4gs2B9rvdAeVGB9ChEyoNMho+Qc6W"
    "KiTPa9zekjhQniSTbWDllRFmrZKcBSSlKy1+uXR2DilmxhhfSOtvEIPQWofq5G+DXd4blsZJAcmUufDK"
    "SsSga0W/jFt8GrNIj4vUUJo8lxyE4csSK5B664AJXwCEUqVxTUsm9Z77iH8M9jQuvYZBqvtF4TM5+i3h"
    "uG0Jx11LOO5bwjH4J46AfQlo/F7+VKLMiYfhw0aBb7o9RNFTzgVzIN6Dww7eOMBP7RMcnCk+KbCVNVyE"
    "g+4xf+yvU2esx0nj4HyA/SipsjsWhcAFCYdhUteUD444pS6+MVRzUICo8aZx7o6+AQAA//8DAFBLAwQU"
    "AAYACAAAACEAtVUwI/QAAABMAgAACwAIAl9yZWxzLy5yZWxzIKIEAiigAAIAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAKySTU/DMAyG70j8h8j31d2QEEJLd0FIuyFUfoBJ3A+1jaMkG92/JxwQVBqD"
    "A0d/vX78ytvdPI3qyCH24jSsixIUOyO2d62Gl/pxdQcqJnKWRnGs4cQRdtX11faZR0p5KHa9jyqruKih"
    "S8nfI0bT8USxEM8uVxoJE6UchhY9mYFaxk1Z3mL4rgHVQlPtrYawtzeg6pPPm3/XlqbpDT+IOUzs0pkV"
    "yHNiZ9mufMhsIfX5GlVTaDlpsGKecjoieV9kbMDzRJu/E/18LU6cyFIiNBL4Ms9HxyWg9X9atDTxy515"
    "xDcJw6vI8MmCix+o3gEAAP//AwBQSwMEFAAGAAgAAAAhAD3UBfWMAwAA/wgAAA8AAAB4bC93b3JrYm9v"
    "ay54bWykVutu4jgU/r/SvoOV/2niQNIQlY4SknTQAEVA2+0IaeQmpnibxIxjCtVonmofYV9sj8OlAVYr"
    "toPAju3j79y+c8LVp3WeoVcqSsaLtoYvTA3RIuEpK57b2t0k1l0NlZIUKcl4QdvaGy21T9e//3a14uLl"
    "ifMXBABF2dbmUi48wyiTOc1JecEXtICTGRc5kbAUz0a5EJSk5ZxSmWeGZZqOkRNWaBsET5yDwWczltCQ"
    "J8ucFnIDImhGJJhfztmi3KHlyTlwOREvy4We8HwBEE8sY/KtAtVQnnjd54IL8pSB22tso7WArwM/bMJg"
    "7TTB0YmqnCWCl3wmLwDa2Bh94j82DYwPQrA+jcF5SE1D0Femcri3SjgftMrZYznvYNj8ZTQM1Kq44kHw"
    "Pohm722ztOurGcvo/Ya6iCwWA5KrTGUaykgpo5RJmra1S1jyFT3YEMtFsGQZnFqtS8vWjOs9nYcCFpB7"
    "P5NUFETSDi8kUG1r+q/SqsLuzDmQGI3o9yUTFGoHKATuwEgSjzyVQyLnaCmytnbjTft///UnSt/Kl2ng"
    "j79Ek6n/xQ+jftefdnq3d2GE/IHfe/waTWt8JKfk/x+MJIkKiAFB2Bi6eT4OCNgrvB3rhlIgeO6GPYj8"
    "mLxCHiDb6bZMuxBo99uPVux2HMfCOg7dQG9GLUcPYtfWg1YYRX7sW66Lf4IXwvESTpZyvs2twmxrTUjk"
    "yVGfrHcn2PSWLH3X/8PcfnQ1Hw27s5/KU9XF7hldle8sUEu0fmBFyldtTccmdMG3w+WqOnxgqZwDjRqW"
    "DdWy2ftM2fMcLMZWU20C25Vlbe3AonBjUQwfXQ0HFhk1k6p+CaZVMyoqjk8iv49uokE08nvQnlVHVTHG"
    "GhKe0iS6KVae1e8Me/5jNBrXxKGD7cWtY/FedxDdDevijZp441h8/Pl2gibdfndwU9PQrF1pHl/p3N6N"
    "Jujr7SCqa7FrV6rCrPvQi8Dp0K9pgPjufXAq2u7ilZAsGQqkpio2LWxaLSVB17JXymqGImOQl8B2A7PR"
    "svRmjGO9iVumHgROU7fDuGFf4rAT2bEipnrReWuFOPtg/3KN6jYlcgmFr2q+WntqjLe7+83ZZmOb84MC"
    "9kahcmV7+78Ex/Aiz+iZwvH9mYKdQX/SP1O2F02+PcTnCvv9IPTPl/dHI/9xEv2xU2H8a0ANyDl0sV3m"
    "jd1/l+t/AAAA//8DAFBLAwQUAAYACAAAACEAZwxbpSkBAAABBgAAGgAIAXhsL19yZWxzL3dvcmtib29r"
    "LnhtbC5yZWxzIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAvJTLboMwEEX3lfoPlvfFQBL6UEwWrSpl29IP"
    "GJkBo4CNbPfB39eiakikyNkgNpZmRp57PLbvdvfTteQLjW204jSJYkpQCV02qub0o3i9e6DEOlAltFoh"
    "pwNaustvb7Zv2ILzm6xsekt8F2U5lc71T4xZIbEDG+kela9U2nTgfGhq1oM4QI0sjeOMmdMeND/rSfYl"
    "p2Zfev1i6L3y9d66qhqBL1p8dqjcBQlm3dD6A5ACTI2O07848oyUXZZfzSn/rc3BSkQ3ERxTlo2VVQjm"
    "fk4Y5+8IJ5AxZOOahBjShQeShmCShWGCk8kWhslCk9ksDLMJXpN3lfk+sYBWPEto1PR4j6kQxXrhkaxD"
    "MI+z2poEg+W7M961T93tNP0Pw86MO/8FAAD//wMAUEsDBBQABgAIAAAAIQBYz9TTERQAALJ2AAAYAAAA"
    "eGwvd29ya3NoZWV0cy9zaGVldDEueG1srF1rUyI7t/5+qs5/sPg0e/odEYFWKPWtAAICykW56DdGcYba"
    "Kr7AXPY5df77WelOoDtPVoyMuzaO85B1y+1JJyvTJ//+/fy093O2XM0XL6eZ3P5BZm/2cr94mL98O80M"
    "b+pfjjN7q/X05WH6tHiZnWb+ma0y/z777/86+bVY/r36Pput90jDy+o08329fi1ns6v777Pn6Wp/8Tp7"
    "oW8eF8vn6Zr+uvyWXb0uZ9OHSOj5KXt4cBBmn6fzl0ysobz00bF4fJzfz2qL+x/Ps5d1rGQ5e5quyf/V"
    "9/nrSmt7vvdR9zxd/v3j9cv94vmVVHydP83X/0RKM3vP9+WLby+L5fTrE8X9O1eY3u/9XtL/h/TJazMR"
    "Dpae5/fLxWrxuN4nzdnYZwy/lC1lp/cbTRi/l5pcIbuc/ZzLBtyqOtzNpVxxo+twqyy/o7Jwo0xW17L8"
    "Y/5wmvnfA/XfF/ozJ38cbH/o7/4vc3byMKcWllHtLWePpxmRK4turpTJnp1EPWg0n/1aJX7fW31f/Gos"
    "5w+d+cuMuiN15PX06/XsaXa/npHdXGZPdtmvi8XfUvSCoAOy8jp9me39vn6lho/K/KN+LZD44rUze1xX"
    "Z09Pp5lKMbM3vV/Pf856JHGa+bpYrxfPg/m37+togKwJe1wu/mf2EnkYmZW+S/2nGVIVF5XeG9/FmqQl"
    "m2Tazn82VXGVk6Wzm6pI/q6rpR6Nvd5y72H2OP3xtB4sfjVn0mOKtLBPEUadt/zwT222uqfRRDWyn5dq"
    "7xdPpIN+7j3P5axAg2H6O67C+cP6O/1Gs8L9jxXVwFgBSiwWoL4TCeRDqvS4wJFTIE9fRxIF2UyxhNVE"
    "NnYtCrs2XU/PTpaLX3s0ckhuRXVNDZ8rSyUyyMPifomaLXZ0EzkTNsUr9Qip6DQTkkOkYUU95OfZwUn2"
    "J9XvvSpSiYsUC1HMUqgKSA2Qc0DqgDQAaQJyAUgLkDYgHUAuAaEuJUNPxNUFpAdIH5ABINeA3AAyBGQE"
    "yBiQCSC3gNwBIlQrJ0IV2KoCm1VguwpsWIEtK7BpBbatwMYV2LoCm1dg+wpsYKFbuBjNHTRyNsOHRiwM"
    "n1x+h+EjFVEfSg6fnDF8VJHE8AGkBsg5IHVAGoA0AbkApAVIG5AOIJeAXAHSBaQHSB+QASDXgNwAMgRk"
    "BMgYkAkgt4DcASIEQhWEqghhuwpsWIEtK7BpBbatwMYV2LoCm1dg+wpsYKFbGIcPrW9w+IT7B+9mH6ko"
    "IrMN+RwaoycuERKzbork00WqcZHE5FYD5FwhUSgR8dUVQrS7UVxIK26AmiYgF6C4pTyWCzLNqMW04jao"
    "6QByCYqvFJKsijCtuAtCPeUNccLGm6O0UB9sDwC5BsU3SrFcOOowj9OKh6BmBGrGMUI/N1pKaS0TZShZ"
    "JGcsUW7B0h0gQiBUQUj3pm1PEao7hdS7N07mjFle6B62neaF7mIJVbpLJTUZPV6oTpY2Z/R5oftdwpzq"
    "eMWEOd3RUnVn9HKBfU/ozpfQrntfAkr1tmySXqlTwPxQ2CxwfZemUot7cohLpMjXGGpVSxFj0NQsRYwh"
    "cm4pYnT2elwkOZ3kjJ7cwCKHRkduWooYPe3CUsToQq24SCnRyw6NDtS2FDE6RsdSxKjdS0sRo3avVCsm"
    "p8NDo3q7tjJG/fZUmeREdmhUcN9SJm/U8MBWxqjia1sZo45vVBki/S0xGZU8tJUxanlkK2NU89hWxqjn"
    "iSqTHOp5o55vbWWMer6zlTHqWQhLoYJR0aJiK2ROm2p85pJuF8wZUY3QZNMXzAlRDdFUGXOmU2M0Vcao"
    "aqEGaaqMUdVCjdJUGaOqhRqmqTJGVQs1TlNlzKpWAzVHlbldTphVrYZqupBZ1WqwpguZVa2Ha8qcWdd6"
    "vKYKbSs7RQVEmrhULO3wpCUV0UZFcrwVjcaTm1bpJ/oqIDVAzgGpA9IApAnIBSAtQNqAdAC5BOQKkC4g"
    "PUD6gAwAuQbkBpAhICNAxoBMALkF5A4QIRDCVhXYrALbVWDDCmxZgU0rsG0FNq7A1hXYvALbV2ADC2xh"
    "oZsYn77oAeBjnr6kIlpgUZVvpxZjrquoMuF28w+QmkKOaPfwMVL1fbqcPWTijexaWK7l5JbUPNqBvqif"
    "DwbdwadK+KUa/uvgr5Pso2XP8VypVMup9dKq+jwsn2vVucyZVl0Ns5XwX5kvGa38S/qJog4RNABpOmNq"
    "huWmNny4NVwPgwYf04VXTBdh+UKrzidVZ5uumFoQQRuQjjOmTljuaMOFreFWGLT5mC69YroMy5dadTGp"
    "OttxxXQFEXQB6QHSB2TgjHsQlgfauXDrXC8M+nzc115xX4fla62aBofun70wO3DFfQMRDAEZOWMaheWR"
    "Nny8NXwTBkM+pjEYmQByC8gdIEIghNOIqGIpPZFsJxuhJ4IEpIevfboR9bBMG9Jqwiltg2+GAfU2bsIR"
    "egpgtDZIa2Mz19BJ2nayCerh58OgFX7O89r1dKIeyZj5TNC8Qtvkynf6c2OFaisraq5OI/T08oYJml9o"
    "212bSExeVANvmdCTzBsmWmSitTGRmMQiE+fOKPSs5Z74RZtMtDcmEtPVJ5qDc/vFz63wr+ynqMn/cjCB"
    "0FPiG+ZobqQjB11piSlMRvTp8PMn6nXBwX6h8LkS/uU06DdhCpox6UBDG0xMS5M3G0nPmm800hVZuNpY"
    "SMxOFbJQd7aRmoUP1SM915m7ZKG7sZCYhqjSPucODvjelnqAoN2dj1ntSEXmasd4aquoMonVDiA1hcBE"
    "ES9w6DgZTlDPlQjby6RodMD988xcq4D9BiBNl0eHcuEGHl34eBSdjYNHLbDfBqTj8ijaQQSPLn08itbC"
    "4NEV2O8C0gOkD8jA5XXUJcDrax+vj+ImMFr2BuwPARm5PDq2tuwYlEwAuQXkDhAhEMLhIapYSg+QJItj"
    "qbortCjtBSpb6K5vHXq5A2t9CD062PkwGnz20Sf0OHELqzFmNLDQI8UtbB9lQg8q95yhxpJpWY8/t7B9"
    "NAmvkZhT48G0rMeiO2b7iBBq2PLEEjWV6vdbyym6kPl7ZmJMbpejSanIpAtjA6+iyiToApCaQt5DF0pk"
    "F7oA+w1Ami6PGLrw8YihC7DfBqTj8oihCx+PGLoA+11AeoD0ARm4vGbowsdrhi7A/hCQkcsjhi5AyQSQ"
    "W0DuABECIRweooql9ABJ0gWWqrtC4+jCJcPShRLajS68hDm68BLm6MKnZ9HWf7QYBLrwEubowkuYowuv"
    "mDm6iIX/iC6omT+GLqQiky6Ms5yKKpOgC0BqCnkPXSgRqg25jYuPY46nC7DfAKTp8oihCx+PGLoA+21A"
    "Oi6PGLrw8YihC7DfBaQHSB+Qgctrhi58vGboAuwPARm5PGLoApRMALkF5A4QIRDC4SGqWEoPkCRdYKm6"
    "KzSOLlwyLF0oIfrDMfi4pwsvYY4uvIQ5ulDC7jmDowsvYY4uvIQ5uvCKmaOLWPiP6IK6wQfxRaTJIIwQ"
    "Mu+lOSqUYAwtt4VqGnoPZ2iZXUgDXWgg1HR6xfCGl1cMcaALbYQ6Tq8Y7vDyiiEPdKGLUA+hPkIDp+8M"
    "g3j5zlAIujBEaOT0imERVDNB6BahO4TobgMMErrcgFjVgm2GTpJNLOXqziA5PnEK0Ze2jVk6/lGu78Yo"
    "ftIcpfhJc5yipXcjFT9pjlX8pDla8Yub4xUl/WfEIs96PmTfKjpxoctgyayy0LyTogvR+d/J4+aUsRIG"
    "laOgchxUSgENIfY4s2oXr4ZB9SioHgfVUkCjjRWP00HIxch6MmHEcYSiTMqbeO9+yrG7S0d29aOgfhzU"
    "SwENcNbd+NhXu7uprAYlexwFjeOgUQqI/1jx+DyXi5YjQtkd4nuHfLQcEcayce1u3KU8jtZR0DoOWqWg"
    "5XA3PlWFaNuUBnIUtI+DdikgXmWjjU9JuWg5gvWJliNYa7RXYXB1FFwdB1el4Mrhbnw+CdF2w6B7FHSP"
    "g24pIL5mo+3ZexbljvSOgt5x0CsFxO2seN8u3qfUk6Ogfxz0SwGtA1jxOG2Fq2tuQeBT19yCwFrXlFVy"
    "cxTcHAc3peDG4e7QHu2QklKOguFxMCwFtL5go40TWrhouYVG0mNITRvn6JL1NoMskSwyDoPxUTA+Dsal"
    "YOxwapKKyZzK4uvbcI51u4vQ3S5CtDjiKiDaFGJWLpWdpNKs4FkXdMN0Fw/jRuM6A1PvKi2Ik2KXcU4H"
    "2WWcGmc7LuO8pNllnJc0u4zzmSHYrWQ/nmaXcV622WWcV9zsMk61M5vW4nH6KO8xfEy6e6SJbhar/hNd"
    "zA/Ny5G6UPJqPlxpq2Gpc4TqCDUQaiJ0gVALoTZCHYQuEbpCqItQD6E+QgOErhG6QWiI0AihMUIThG4R"
    "ukOIZm7LdUULVrVglhYnmkN9ljanJEcsZ2l1yiHEcpZ2pyw9LGdpecp2w3KWtqeMMlUO8+LlzZYPeoaK"
    "ry4dJm8lFc3M+MicsTmnrjwlN+cUxCTH5wqUHS83JikDjB4gt49iucKXaq7AJ8hr6+zzUPRPp5yT/vON"
    "/mSOPP0zOhXS78iSx/gaCDU1ZI+PvqZMeR1fMlU+V6CnJj6+C7/4qBily2v9yXx5io+MO+JrYTBthDru"
    "+OhryprX9pNp8xQfqWPb79IvPipGqfNafzJ3nuIj4474rjCYLkI9hPoIDdy1QF9TDr32MplET7VA6tha"
    "uParBSpGifRafzKTnmqBjDtq4QaDGSI0csdHX1M+vbafTKin+EgdG98YTU0QukXoDiFiBJhdaKMTMZo2"
    "zFMDWmUjRnMDlKu7K4JW0AWZXa+rIpleT1UhOySfYP+GapoNZIq9Up1PptiTajIsk+zppyvLXptgF93x"
    "vyclZyXaaNWmUnn21KNkbbmSxjdz0xtm5OREPKnNpHLtPcxsJqg3zFA5mW+vzaTy7aUZamdXNJtJz80k"
    "RORkpr0xk8q5pwaKku5zBZl1H/cEd96938CnpHsyuplg86nMe4otTr0ne3HuPZl3J997WpWzLi1IdI0m"
    "0+892m0z8b7RblROpuBrK8kUfGmFeryr2fRc/mYWvrQi/yG5aImRT6Xh02iK8/A5U6nMSnlP8IOWV/El"
    "z/TyykzFj8wZy6tYLnEcWtOlrGefVDfWdHwt5NxLpg10W7JUHf1qINR0+8Xt+qr43H5x275QOW30q+P2"
    "i9uf9fKL26AFv7roVw+hPkIDt/fcjqeX99yWJ3g/RL9Gbr+4vUlQPUHVtwjdIURrA1BGawPEqhZsM4ZS"
    "h6Aou+n59rHGbp8pTVYponnmFFRJsXOofGDKc3k1Okq3NLt95mWb3T7z6W55LrVGe+6cAogHrXmcl37S"
    "7PaZV9zs9lks7T4FJf4xPE9TjLwT+jGnoPHt0jTFmOn7RCSQXgNQTZd6H8W8dUE4euDnKAacaKCrTbdf"
    "HMV4+cVRDPjVRr86br84ivHyi6MY8KuLfvUQ6iM0cHvPUYyX9xzFgPdD9Gvk9oujGFA9QdW3CN0hRBRj"
    "udZtwaoWbDOGUhSD+uruKFmKUZreSTFKakeK8ZJmKcZLmqUYn+7GU4yXNEsxXtIsxXjFzVJMLP1nFCPv"
    "HX4MxcQ3GNMUY6b85+DmYxWhmobeRzE+d0/Zpxjwq4F+Nd1+cRTj5RdHMeBXG/3quP3iKMbLL45iwK8u"
    "+tVDqI/QwO09RzFe3nMUA94P0a+R2y+OYkD1BFXfInSHEFGM5c6xBbMMIdqzQ9lzC1Z3R8lSjOu6Mv8U"
    "o6R2pBgvaZZivKRZivHpbjzFeEmzFOMlzVKMV9wsxcTSf0Yx8l7ax1CMuuGWPIfESwLqDl3ykgBANZk4"
    "KTMK30cxSmi3jTJwoqGd2LradPvFUYyXXxzFgF9t9Kvj9oujGC+/OIoBv7roVw+hPkIDt/ccxXh5z1EM"
    "eD9Ev0ZuvziKAdUTVH2L0B1CRDGgjDbKEKtasM0YSj3FoGzdHSVLMa4RylOMktqRYrykWYrxkmYpxqe7"
    "8RTjJc1SjJc0SzFecbMUE0v/GcXI85yPoRipiXJPkv+WK14XUIXo/C15XYAO3miTmT50byBHFwdyx44r"
    "A1YVdFZNVwWK9KG7A6SChp3j2kBShZlfyp71xEJySPKZ9OxTktVpOpmj8+gifegGATlN491xecCqgg67"
    "Kf2lSB+6RUAqiBkdFwiccXMU6RU3R5FWp+mgmU7gi/ShuwTkdMvhdHxQLN9JlOoydIhMaTFF+tB9AlJB"
    "zOu4SuCMm6Ngr7g5CrbGTQe1dImgSB+6VUBOXzmcjk9aIW46raWLBEX60M0CUkHM7rhUYPWCMnXoMkGR"
    "PnS7gFTQSsBxscCqglJxKB2nSB+6YUAqaOXguFzgrH1uCeFV+9wSwuo0ZfDQtYIifeieATl943B6GJ1z"
    "Q+1Tig6l6RTpQ3cNSAWtTBzXDJxxc0sUTihKOBnnSvKegSwjT62S9wzIszF5NibPxuTZ2OHZJBWcOf+R"
    "VttJ2u0uQne7CNHSylVzjH+0+NpFqrqTVJxpqbuHZw2qhnuvVJwkxUkdsotAZ22wp6Vq2NEfjuti7Gmp"
    "lzS7CPSSZheBPhMGvwj0kmYXgV7S7CLQK252Eaja2XnZgDstjV+DGL8P8Hm2/DaT72xc7d0vfsiXGlKS"
    "09nJBt6+TTJ+haL5RTtP6UFR6xjfXOXLVlwcUtZS1BkMgXq+TG+loUkIbEsTlLaNX7XyZXrVDOK9fJne"
    "OGRRVSRNUXMaNiZkgt7cgxI3+TK9dMiiqUYS9IIe/KaSL9O7fCwSTZKgd+zQN9ltpcs3an6bXU6X3+Yv"
    "q70nequlfKvkEZ2FL+M3T8Z/oRdiRm+yiV9vGf36nd4NO6OXKR7sU+HHxWKt/yINbN42e/b/AAAA//8D"
    "AFBLAwQUAAYACAAAACEAHtNDiO0PAAD5aAAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbKyda1Pi"
    "2BaGv5+q8x8oPvW003LRaJtSp0YwkJArCvKVRmypUbGAvp1T57+ftZO9Icmbldlk6Jpp7ce93n19k5C1"
    "Yy7/+Pn6Uvs+X60Xy7ereuu4Wa/N32bLx8Xb16v66N769LleW2+mb4/Tl+Xb/Kr+a76u/3H9739d/liu"
    "/lo/z+ebGim8ra/qz5vNu9lorGfP89fp+nj5Pn+jnzwtV6/TDf1z9bWxfl/Np49x0OtLo91snjVep4u3"
    "eqJgrnQ0lk9Pi9m8u5x9e52/bRKR1fxluqH2r58X72ul9jrTkXudrv769v5ptnx9J4kvi5fF5lcsWq+9"
    "zkz769tyNf3yQv3+2Tqdzmo/V/Rfm/4/UdXEHGp6XcxWy/XyaXNMyo2kzdj9i8ZFYzrbKmH/tWRap43V"
    "/PtCTOBOql2tSS1jq9XeiZ1UFDvbionhWpnfFo9X9f825Z9P9LUl/mp+arbEX6k//6tfXz4uaIZFr2qr"
    "+dNV/c+WOTlp1xvXl/ECGi/mP9ap72vr5+WP3mrx6C7e5rQaaR2LFfplufxLFLWp5iaJvk/f5rWfd+80"
    "z1d16uIv+S01cLN8d+dPm8785eWq3jmt16azzeL7PKSIq/qX5WazfB0uvj5vYj9siD2tlv+Zv8Utmr/M"
    "qTA1Vehf1UkqKSpam/tZoiRqKopM10PBjW1f09+rfluxt8JV7XH+NP32shkuf/Tnoolk5NNj6kG8OM3H"
    "X935ekZuoSE4PhG1zpYvpEF/114XwvW02Kc/468/Fo+bZxFer82+ranLDxLIsCSABi4OoK8y4Ky0PI1u"
    "XJ5KyfLnpeXpp3H5C039Fs120oNdi8praKkmtXZtKu9DSzWqTcOl1Yv2dpioMhlBR9OycaVhTwaWvimr"
    "o5HMYLw6utPN9PpytfxRowMItW1Na5AM0DLbJCLWgviaVLpdHczSoDUhRP4UKlf1M+oxha/JNt+vz04u"
    "G99pEc5kmZukjHEaLwwR1QHSBXILxALSA9IHYgNxgAyAuEA8ID6QAEgIJAIyBHIH5B7ICMgYyAOQiSSG"
    "mIsGrYTtcqAVCMuhRXC/5SBUaDnQ2t0th9PcckjKGHET4kXUSci5OLJuF5GRjerKqN0iugViAekB6QOx"
    "oT2ObA8dKrbt+ZxtzgCC3ISk+3CRjfGkLH3ZdfMsW8aH5gVAQqg7kiQtfJ4VHmLzznJ9upPto2PArn25"
    "PtxDa0bQmnFCMovgvJltzgPoTNI6mZVJQwor83R7dNQ9UAmV+KC37dt5K7cwC4q0s0U6SZH00e48d7Tr"
    "FhTJOeC2oEhuuVsFRXJLpVdQJDfp/YIiuUm3C4rk5tyR45Jet59z8zkoKpMbX7egTN4iskjaRZ9zc+AX"
    "lclNQlBUJjcLYVGZ3DREuCROc9MwLOp5rsydLJN21ufcXN0XlclN1qioTG4Ix9jki7z5Cork5mpSUGQ3"
    "Dxl7UqfwxLG3PYUKnTjoMm7rTyM3GTdJmfR1BJAukFsgFpAekD4QG4gDZADEBeIB8YEEQEIgEZAhkDsg"
    "90BGQMZAHoBMJMHrCKNoOZwdN4nvdykhhOiK9Gx71XiTELoMF59L4utIIF1JzulK9CleSs/T1fyxnnww"
    "7BpmV1zFrxfxRzzbuh0Og+GHjnHUN35v/nbZeBLXrzm33EIlVmkllmFaqpJW/VpVcltaSQ8q6QOxgThA"
    "BkBcIB4QH0gAJAQSlQ5DZJiRGob2bhh6xlHH+Ng+ujU+nrBDPkyU2+fbmb6TdckLy82qcH7vDPNO1Xmy"
    "q1NMcOvY+Hhr/Nb40DWOLOO33+uf6mq+P2XP8fdaVd0b5r2q6nRXVWQ0PrQ/JpUcNY9PTz/2jd/Kahtp"
    "1TYyzJGqzdjV5lFtnhjQI+oVLbDyjo21qhob5lhVdVa/pnGStsgN0wPM0USSz/DZQlgu/1GzVeWYIISy"
    "x4SEpI8JQLqSwDEhOQzQwQQ9DyJWmUhL9BhEeiDSB2IDcYAMgLhAPCA+kABICCQq62Z8Sw26OUxC0maV"
    "IqxZxcDHd5e+X+fNpxMafwaE0JFOaHy+gtCxTmh82oHQB+j9RBK0AX1sOIwNhFDWBglJ2wBIV5J9bAAi"
    "VpkIYwMQ6QOxgThABkBcIB4QH0gAJAQSlXWTsUESkraBFKliA51QxgY6oYwNdEIZG0DvJ5KgDUSa6CBn"
    "AyGUtUFC0jYA0pVkHxuAiFUmwtgARPpAbCAOkAEQF4gHxAcSAAmBRGXdZGyQhKRtIEXoi/hoh5duJWcD"
    "nVDGBjqhjA10QhkbQO8nkqAN6KPuYWwghLI2SEjaBkC6kuxjAxCxykQYG4BIH4gNxAEyAOIC8YD4QAIg"
    "IZCorJuMDZKQtA2kSBUb6IQyNtAJZWygE8rYAHo/kQRtIBJyh/FBrJQ1gkRpJyDqKrSPF1DGKpVh3IAy"
    "fUQ2IgfRAJGLyEPkIwoQhYii0h4zxpAxaWcomSrW0IplvKEVy5hDK5ZxBw7BRKECf4iM3UEul1pJ7i99"
    "R02ijD+SUinUVaX28gfIWKUynD9Apo9tthE5iAaIXEQeIh9RgChEFJX2mPOHnKXUvS8lU8kfcgBLYzl/"
    "6MRy/tCJ5fwBQzBRa7fAHyJveBh/JBnIjD8SlPEHoK5Ik4vc4l7+ABmrVIbzB8j0lczuNrmNyEE0QOQi"
    "8hD5iAJEIaKotMecP+Qspf0hB6GSP3RiOX/oxHL+0Inl/AFDMJEj2S7wh8jkHcYfSU4w4w+ZJkylZMR2"
    "KWGGHeoqtJc/QMYqleH8ATJ9bKCNyEE0QOQi8hD5iAJEIaKotMecP+Qspf0hB6GSP3RiOX/oxHL+0Inl"
    "/AFDMJEjWeQPkck8jD+SnGjGHwnKnD8AdcUWyr3PHyBjlcpw/gCZvpJJnz+glIOlBohcRB4iH1GAKEQU"
    "lfaY84ecpbQ/ZPcq+UMnlvOHTiznD51Yzh8wBBM5kkX+EFm9w/hD5gdTGf0WJJU7iLoK7XX+AGWrVIbz"
    "B8j0sYE2IgfRAJGLyEPkIwoQhYii0h5z/sDcu5Kp5I+/SzqLu7+cP3RiOX/oxHL+wNS2HIIif4h032H8"
    "IROHaX9AgrYT5+Zz11dlSdtmYW76FmWSTSvcxxjOH9DAPirbiBxEA0QuIg+RjyhAFCJK9qdwPeb8gelu"
    "pVzJH/8g4a1VL+ePf5DylvWmbuFNFCr4/CHygIfxh8wo0q3k3f7k/B5e8ewFXUudiCd4nq7vRt6HG8O8"
    "aZ0xm7s6WL5jmGQypnxXlY/9EG9Cu0WJW8MkgzESFkr0MhKwbY1+bIZxETpW0klQ9KpnmD22ij6rJ462"
    "8ZUPbOOwK8Q4FWIGFWLcCjFehRi/QkxQISaZTLVK0/PNz0+Ey2aIK29omEN2WdyhxD2iEaIxoges+8Ew"
    "H9i6JxmJzAbilsgNwRGCDqZ7PomUpJjO6ebh7vCQ2zt+E9d1Vc88i5TEpVAXS90ishD1EPUR2YgcRANE"
    "LiIPkY8oQBQiihANEd0hukc0QjRG9IBoolDBE0oHyxS2hVI2UyhR+pM6oq5CzPbidtPsnoib13SU/bzb"
    "ptlpN4/67Sa/wxhrssproh+blqrpIrXJ+G9q6mFN1LD8DTsbkYNogMhF5CHyEQWIQkRR+aDQj81IDQol"
    "1LZbr6nXRzQHYtcx/V2y7Vjqp3OTqkr22i5+4viOqr7bVp3a9f1BzH2897jdFJuP6V80d6Xbj/VqvKca"
    "77c1pjZY0ygkW5CTquQmZKqzdBeyXqUjqnS0rTS1w5pmmXYiJ8N8JDpJA/13HR3r1UnFzPG2Ttpqze9I"
    "xumbKITXpeLZ3MNcl8ZKuaMJpC07slTqANNVqPC+Rtxi3JiMMlapzEXx1mSU6SOyETmIBohcRB4iH1GA"
    "KEQUlfaYnF+0G3uopil140/plH5wo9RjrJffqawXLD9F5oJHesHFe6THesHFN1UecBgmChWY5GDJVfGE"
    "e/6Ui8lVWSpjkrLkKmcSTK4q5UKvcSbB5Co20EbkIBogchF5iHxEAaIQUVTaY9YkmF1VOtVMopPmpORl"
    "kcNGejVzJtGqmTMJZljVAi4wycEyrG3MsEqUuS7FDKsqtdeZBDOspTKcSTDDim22ETmIBohcRB4iH1GA"
    "KEQUlfaYNQmmWJVONZPo5DpZk2gFcybRCuZMgmlWtYALTHKwNKv4HSX5MwmmWWWpzJmkLM3KnUkwzaqU"
    "9zqTYJoVG2gjchANELmIPEQ+ogBRiCgq7TFrEsyzKp1qJtFJeLIm0QrmTKIVzJkEc61qAReY5GC51jbm"
    "WiXKnEkgu9lVpfY6k2CutVSGO5NgrhXbbCNyEA0QuYg8RD6iAFGIKCrtMWsSTLYqnWom0cl6sibRCuZM"
    "ohXMmQQTrmoBF5jkYAlX2qkDZxJMuMpSmTNJWcKVO5OAsqWU9zqTYMIVG2gjchANELmIPEQ+ogBRiCgq"
    "7TFrEsy4Kp1qJtFJfbIm0QrmTKIVzJkEhmGiFnCBSQ6WdaWboWASeF62I0tlTFL2DC1nElC2lPJeJgGZ"
    "PjbQRuQgGiByEXmIfEQBohBRVNpj1iT4+LHSqWYSnWeBWZNoBXMm0QrmTILPIasFXGASGpkD3QIWSrlb"
    "wAnKXG4B6tKGIn7rJ2cSkLFKZbjLLZDpK5nU1k9EDqIBIheRh8hHFCAKEUWlPWZNIqcpfQtYjkI1k2gF"
    "c3e3tII5k2gFcyaBYZjI4SzY39YuzM5X+eUtsVLOJPCsbUeWypxJyp6/5UwCypZS3utMAjJ9bKCNyEE0"
    "QOQi8hD5iAJEIaKotMesSfDRZaVTzSQ6zxGzZxKtYM4kWsGcSfAZZrWA8UxC280OdCaJlbImkSh9JkHU"
    "VWifD+4oY5XKMGcSlOkjshE5iAaIXEQeIh9RgChEFJX2mDOJDErvFFA6lUyiF8ycSfSCGZPoBTMmwWGY"
    "KFRgkoNl3OPtADmTYMZdlkqfSRTayyT4JHOpDGcSfJIZG2gjchANELmIPEQ+ogBRiCjZVsPuDWcy7mqa"
    "UpdbSrqaSXSeKebOJHo1cybRqpkzCT7OrEamwCQHy7jTGxDindD0hd8urQrRXh61XZr229yctLj90hhA"
    "O5zMDhvQVQHx4CQbplGDtguZt6yGhRq9jAbumD5pm2FchHbi0bMbyZZpqqTHVtJnFcVuPtIo2lRiVwly"
    "qgQNqgS5VYK8KkF+laCgSlAyrbR3Ol6y+b3TzDxFuIKGuAqHtECG7AK5Q417RCNEY0QPWPkDVf7AVi7e"
    "ZxK7ObZRsn06eeNH8k6H1/nq61y8j2Rdmy2/ifd30K7Qyy3dvRcl3syU43dtk36LO63vHB+3Tfqt7Mj/"
    "bJs3hfzUpN8LjOWdtkm/L79Ap3Vhiv2++BOP7FsU0Wmb9Ov6xasLdh0Wb2r5Ovemq6+Lt3Xthd6WIl5e"
    "ck4phVXygpPkH/SilfjXvyevTYm/faZXDM3pdRTNYyr8tFxu1D9EBduXFl3/HwAA//8DAFBLAwQUAAYA"
    "CAAAACEAPL8da30kAABkMQEAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQzLnhtbKydW1PcyLKF30/E+Q8E"
    "T7PHGzcNNLY7bHaAaJBAwrpwcfPG4LZNjAEHMPbMOXH++0mpqhpJS5lkl3rHbF8WlVWrqqSvJXUq/f4/"
    "f99+X/k5e3i8ub/7sDp8vb66Mru7vv98c/f1w+rZ6cHa29WVx6eru89X3+/vZh9W/5k9rv5n57//6/2v"
    "+4c/H7/NZk8r1MPd44fVb09PP8aDweP1t9nt1ePr+x+zO/rJl/uH26sn+uvD18Hjj4fZ1ecq6Pb7YGN9"
    "fXtwe3Vzt2p6GD9o+rj/8uXmerZ/f/3X7ezuyXTyMPt+9UT+H7/d/Hh0vd1ea7q7vXr4868fa9f3tz+o"
    "iz9uvt88/VN1urpyez2Ovt7dP1z98Z3m/fdw6+p65e8H+m+D/r/phql0GOn25vrh/vH+y9Nr6nlgPOP0"
    "3w3eDa6u5z3h/FXdDLcGD7OfN+UGPne14WdpOJr3tfHc2aZnZ9vzzsrlehj/dfP5w+r/rtv/rdHvw/KX"
    "9bV1OhaqP7mf/d/qzvvPN7TD5axWHmZfPqzuDse7e2/WVwc776sj6Pxm9uux9ueVx2/3vw4fbj7HN3cz"
    "OhzpQC4P0T/u7/8sm0Y09Dr1+uPqbrbyd/GDNpoO99WVf+wfyeHT/Y949uUpmH3//mF1b2t15er66ebn"
    "LKWID6t/3D893d/mN1+/PVUnxBNpXx7u/2d2VzmafZ9RY/Ja9v9hlboyTUu3rZ+ZnsqRuiLr41DwYD7X"
    "+p/dvA+qkyt9WPk8+3L11/en/P5XOCst0tS2XtMMqqNz/Pmf/dnjNZ0utASvN8tRr++/Ux/068rtTXna"
    "09F+9Xf1+6+bz0/fPqxu0PZf//VIU74wwtCGmQD6aRVAv9sAwoTQnlbXDLA9D6A/CQHDNy7i3TyCJMkS"
    "bbjxRLOxpuQxyjmaiJF2DOrQRNB0pYkPzAJXm7d/9XS18/7h/tcKneDk7ZEOETo+h+MN6qTcqg3aKDOx"
    "+eYxO0dbVnayW/byYXWbFoTCH+mo/rnzbvP94CcdI9e2zZ5pM9qq9q2MCkDZB2UCygEoh6CEoESgHIFy"
    "DEoMSgLKCSgfQUlByUDJQSlAOQXlDJRzUC5A+QTKFJRLUHbtNte2cNft6qjc1QEdU/MDi45lOLCG85NY"
    "e2CVvdCBRQfm84G11TqwTJtRZaE6HAOjvCkR6g7H1tG4b4Oej8YJKAfQ8aHtmE7PZzujpp0Q+olAOYKe"
    "j23PdD4/97zd7DmGfhJQTqDnj0ahU/q54zfNjtOOJm+bTbKOjRgOm21ycFOAcgr+zqxS8zfcaHZ83jV4"
    "a0MvYKhPMNTULnJjLd41x7qEfnZ3Udqr99047OmYg8N+a/5ZpD3sy14qNs/3bLi+3jrsu9q0diQwbepY"
    "ftPa+/2OJq29n2CT4Xpriw662rS26NC0qZ/Lb1pNwo4mrdM9wibD9Vabo642rfP02LShT/P5Gr9pNYk7"
    "mrTOyQSbDNdbbU662rS24aPdTfqNP0072rxt7UPWdVS09jPHNlutBSy6ummdKacdbYatg/QM22y3Zn7e"
    "4aZl+KJrpNah/qmrTWtxprZN/fx/27Jz2dFmOGwdpru7na1ai7i719nq+SBrgINM4eflwuAoe6HPy/pR"
    "PRy2jsfyToIa1a/EQNkHZQLKASiHoISgRKAcgXIMSgxKAsoJKB9BSUHJQMlBKUA5BeUMlHNQLkD5BMoU"
    "lEtQdndRcruKV2J0tYJHFomLXeKXvdCtwfb88n0PlMAow+c2+6BMrGI/BZ4e6GbhS8W8b1cPs8+r5t56"
    "MhpPhnQQP95UN8nRwSTPP+a/BaPB/ujfq2ur/3o/+FLeZKw1PxMPbOdvOjs9GI0PXKfD1R3X6f5oLRj9"
    "e9112SLYIcwgBCVSzSkajSM3/Mbz8IejQSjN6Uic09FofOQ63XzuNBytHfJzOoYZxKAkqjklo3Hiht96"
    "Hv54NIilOZ2IczoZjU9cp6PnTuPR2jE/p48wgxSUDJQclMIqlqDM8VmMxoWzuP1sMR8NMmnep+K8T0fj"
    "U9cpHcHu+AxHr2gxuePzTOzybDQ+c12+fe4yGL06HP2+8ep49Psm2/O5aiXOR+NzN8K75xHOXliJC1Xn"
    "F6PxxfyMpYdlbkmo91xa50+wrVPV4Twdjafz8WqI+O1w9Gr4evT78ehfg9+q/fiXQKFLM9YGPf8pr+WY"
    "Q+hyNL6cj1Xjwdnod7rWl4+j3V07xJs5jXctjjfpJ8Ko1IyeWs7HrSHjcrRG3XZPq3GhQreoS/g4KXtp"
    "fpyAEhil/nECysQq7MeJ+QShhxP4YWFD4cOiDKmeLv7caX8YwPghKJHG0Ua5b+DoSHJUPSYFR8cwfgxK"
    "onFUPQgBRyeSo+pKAxx9hPFTUDJQclAKq7AgLreputIA16eS6+qUAddnUshbs1utg+FcY/Bd50ZfaEKH"
    "1VcLMLlPsFJTzf4O7RHdumK6NLE8q6qToftw3d21wXUKGYmnUNWfPZafvTQAQ48rlgCYspcmYEAJjFIH"
    "DCgTq/gAxoYuAhgYPwQl0jhiACM5YgAD48egJBpHDGAkRwxgYPwUlAyUHJTCKj6AkVwzgJFCGMBoDDKA"
    "0YRygIGVmmr2lwOMifUEjA2uA8ZIvQBTZhC0v/Ma0nXTYjfEZS9NwIASGKUOGFAmVvEBjA1dBDAwfghK"
    "pHHEAEZyxAAGxo9BSTSOGMBIjhjAwPgpKBkoOSiFVXwAI7lmACOFMIDRGGQAownlAAMrNdXsLwcYE+sJ"
    "GBtcB4yRegGmfHLVHzBlL03AgBIYpQ4YUCZWoXl13xgKt0g2dBHAwPghKJHGEQMYyREDGBg/BiXROGIA"
    "IzliAAPjp6BkoOSgFFah3/idZW6RJNcMYKQQBjAagwxgNKEcYGClppr95QBjYj0BY4PrgDFSL8DQxJdB"
    "mKqbJmJQCqxUhwxKEyf5YMbFLsIZtBCiFKlcMagRXTGsQQsxSonKFYMb0RXDG7SQopShlKNUOMkHOqJ3"
    "hjpiDIMdlUeGO6pYatT1AO8TrtdUtdccemywJ3tcdB0+VutHnzI1rP/1DU26fYGDUmClBn1MYE2auFZe"
    "9LHdLUQfsBCi0UjliqOP5IqjD7iK0VWicsXRR3LF0QdcpegqQylHqXCSF30k7xx9pBiOPjZG9MjRRxPL"
    "0gcWeqraa5Y+9gRlv8gSHw67c7lBH9NjP/qUuXdLoI9J4asnNJTpqc07rsBKDfqYVg36WMmLPjZ2IfqA"
    "hRCNRk4SXXH0kVxx9AFXMbpKVK44+kiuOPqAqxRdZSjlKBVO8qKP5J2jjxTD0cfGeNFHE8vSBxZ6qtpr"
    "lj72bPSkj41u0Mdo/ehT5vQtgT4mNbBBH5CCoU0grCVUoTRxkhd97AgL0QdchegqUrni6CO54ugDrmJ0"
    "lahccfSRXHH0AVcpuspQylEqnORFH8k7Rx8phqOPjfGijyaWpQ8s9FS11yx97NnoSR8b3aCP0frRp8z7"
    "XAJ9TPpogz4gBUMjNa59QJq4Vl70sd0tRB+wEKLRSOWKo4/kiqMPuIrRVaJyxdFHcsXRB1yl6CpDKUep"
    "cJIXfSTvHH2kGI4+NsaLPppYlj6w0FPVXrP0sWejJ31sdIM+RutHnzLrcAn0scmLtVTy8tv39p0X5HDu"
    "21aNO6+X8jqFL7dcdwvRB1yF6CpyktedF5fQW86Eow+4itFVonLF0UdyxdEHXKXoKkMpR6lwkhd9JO8c"
    "faQYjj42xos+mliWPrDQU9Ves/SxZ6MnfWx0gz5G60efMrVwCfSxSYt1+oAUEGOqNzDrd14gTVwrr2sf"
    "KeOUST9GVyFKkcoVd+0jueLoAwsTo6tE5Yqjj+SKow+4StFVhlKOUuEkL/pI3jn6SDEcfWyMF300sSx9"
    "YKGnqr1m6WPPRk/62OgGfYzWjz5l3uES6GMzGuv0ASkoi42U73/X6QPSxLXyoo+UjsrRByyEaDRSueLo"
    "I7ni6AOuYnSVqFxx9JFccfQBVym6ylDKUSqc5EUfyTtHHymGo4+N8aKPJpalDyz0VLXXLH3s2ehJHxvd"
    "oI/R+tGnTEpcAn1sumOdPiAFQyM16APSxLXyoo/tbqE7L7AQotFI5Yqjj+SKow+4itFVonLF0UdyxdEH"
    "XKXoKkMpR6lwkhd9JO8cfaQYjj42xos+mliWPrDQU9Ves/SxZ6MnfWx0gz5G60cfWtdl0KfsppVtCFJQ"
    "vZvauvYxrRrPfazkRR8buxB9wEKIRs3r/eRddMXRR3LF0QdcxejKvKD/kiuOPpIrjj7gKkVXGUo5SuYl"
    "e/LuRR/JO0cfKYajj43xoo8mlqUPLLR5e/2lvWbpY89GT/rY6AZ9jNaLPhud2YZ0ii32ulbVDZTGaVX/"
    "2bOt6rVxUNpHaYLSAUqHKIUoRSgdoXSMUoxSgtIJSh9RSlHKUMpRKlA6RekMpXOULlD6hNIUpUuUqIyb"
    "ST5rlC90GlbNKYtxLuFDr+qm+aGHUmCl+iU3ShMnsR8vVV1aajaebJaPPMo6BrWqLDTKYH9jQ6qf40Zg"
    "CuhQzwfznmulWajXNeqdL6KDkwlRinTzo2bjaO6iVn3lkOZH3QrzO5LnRz8eH817rpdgoflR7+z8jnEy"
    "MUqJbn7UbJzMXdSqttAoA+pWmN+JPD/68fhk3nOtWgv1uka9s/P7iJNJUcpQylEqnMR+XFZHMTUbF84r"
    "5b/Pa79QlwMaSViFU3kV6Mfj03nPtSovdOy8KteXLbUj93tG/Z7N+61VdKEz4xUdPWW9HfpVKLijW5lz"
    "Gud8Pk6tggsZeGFlLnQjULPxxXyEGkHKEWj9hbX/hBs+1R321Gw8nQ9ag8tvtHam+s7GRll+x+ySWIDH"
    "jvhSBR4a8XI+Yg0kNE1Tg0ecKn2gdOQYWu3FMjw09u7efPAaay7pXCy7VlTi2Sgzivo/Faq6aX1E2fyl"
    "5wdFgW3V+IiCnKuJayXeAdGHUmdBHhfcfWPWfbdxiL5ClCKdr+53JY9kX913EsdoIkYp0fmyNx6t2i0n"
    "si/7hkGras1HNJGilKGUo1Q4SbzzobOwa7dPRfd0XVYFteuJykH21rpdqEdnE4rSVLWkL3TB3Uf0J1yz"
    "qWrHN+zB3q7WY4Pl98FMWh8UDiKodGQlWk2+Q9uwxzhTsKesF78MDtm0pdrT6arnBpoCKzU4BNlXE9fK"
    "j0NS6hllVnQdmYfoK0Qp0vniOCT64jiEmYnoK9H54jgk+uI4hLmJ6CtDKUepcJIfhyT3LIfEII5DmhRD"
    "OiG7cHmhmyPHIcxPVO04yyFNfiLPoY78RGunH4fKLKMlXA/ZBKY6h0AKaGna39GjNHGSH4ekJDSWQ+Ar"
    "RF+RzhfHIdEXxyHwFaOvROeL45Doi+MQ+ErRV4ZSjlLhJD8OSe5ZDolBHIdskGyT45AqmOMQLPZUteMs"
    "h+x5KT6x5jlko+tPrK2dfhwqH74tgUM2lanOIZACmh1wCKSJa+XHISkdjeUQmAjRaqTzxXFI9MVxCHzF"
    "6CvR+eI4JPriOAS+UvSVoZSjVDjJj0OSe5ZDYhDHIRvkxyFVMMchWOypasdZDtnz0pNDNrrBIaP141CZ"
    "ebQEDtmkpjqHQAroHhE4BNLEtfLjkJSYxnIITIRoNdL54jgk+uI4BL5i9JXofHEcEn1xHAJfKfrKUMpR"
    "KpzkxyHJPcshMYjjkA3y45AqmOMQLPZUteMsh+x56ckhG93gkNH6cajMQVoCh2x6U51DIAXlP2LYyp1G"
    "aeIkPw7ZIRZ7Tg2+QvQV6XxxHBJ9cRwCXzH6SnS+OA6JvjgOga8UfWUo5SgVTvLjkOSe5ZAYxHHIBvlx"
    "SBXMcQgWe6racZZD9rz05JCNbnDIaP04VGYjLYFDNtGpziGQgg0jNZ5TgzRxrfw4ZPtbjENgIkSrkc4X"
    "xyHRF8ch8BWjr0Tni+OQ6IvjEPhK0VeGUo5S4SQ/DknuWQ6JQRyHbJAfh1TBHIdgsaeqHWc5ZM9LTw7Z"
    "6AaHjNaLQ5vLqd5addP83h6lwEp1DqE0cZIXh1zwQhxCEyFKkc4XwyHZF8MhNBGjlOh8MRySfTEcQhMp"
    "ShlKOUqFk7w4JLrnOCQHMRzS2WSeU+uCGQ7hmk1VO85xyAZ7fm/vouscslo/DnVmVtND8cUyqzdtmcja"
    "9RBKgZUaHIJ6lRPXyo9DUrFO7vkQ+gpRinS+OA6JvjgOweLE6CvR+eI4JPriOAS+UvSVoZSjVDjJj0OS"
    "e5ZDYhDHIU1RVu57e90cOQ5hRVfVjrMc0lR0Zb8vc2d1g0Oaiq4v5A9tLifVvuqmdT1kMzSf0RTYVg0O"
    "QeXKiWvlxyGpbCfLITARotVI54vjkOiL4xD4itFXovPFcUj0xXEIfKXoK0MpR6lwkh+HJPcsh8QgjkM2"
    "yOu+TDdHjkOw2FPVjrMcsuel332ZO9EbHDI99rseWk4+9SYkT++hFFipwSHMp3at/DgkFfBkOQQmQrQa"
    "6XxxHBJ9cRwCXzH6SnS+OA6JvjgOga8UfWUo5SgVTvLjkOSe5ZAYxHHIBvlxSBXMcQgWe6racZZD9lT1"
    "5FBHPrW1049Dy8mn3sR8apQCKzU4hPnUrpUfh3zyqdFXiFKk88VxyCefGk3EKCU6XxyHfPKp0USKUoZS"
    "jlLhJD8O+eRTuxE7HyHSO1VdKffnOpvc8yFVMjbHIcynVu04y6Fe+dR2aHOXVb21Qq+WmR77cajMj+z/"
    "fVnpoVX3A6XASg0OQbboxLXy45CYH8y814G+QpQinS+OQz751GgiRinR+eI45JNPjSZSlDKUcpQKJ/lx"
    "yCef2o24GIf65FPr5shxCPOpVTvOcqhXPrUduskh02P5I/7fqHzp+VCZC7kEDtk0y/pzapCC6p3cZv0h"
    "lEx9hRfrrDDvubr+Fvu+DPOp0Zepi/CiL45DPvnUaCJGydQzeNEXxyGffGo0kaKUoZSjZOoQvFSFiHvP"
    "Vdxt9r7MJ5/ajeR3X9YnnxrXzLzK/2ItLOY9V9uf7/dlHfnUtsd+HCpzkIBDJC74fZlJZXpDT85LJK7u"
    "vP+5Mxy+fT/4ufN+cL3zvrqC29s0reqViFDaR2mC0gFKhyiFKEUoHaF0jFKMUoLSCUofUUpRylDKUSpQ"
    "OkXpDKVzlC5Q+oTSFKVLlHZ3O7T5fmMlos0y5WQJH382m6X+8QdSUA3W+viDpJyJa8VehptKRJvvxpNR"
    "+ZCEKhFRPuG8hguNMtjffCdVInIjMJWIqOeDec+1SjbU6xr1zlciwvmFKEW6+VGzceRc0Bfy8/kd0vyo"
    "W6kSkTy/I+r5aN5zvUYNzY965ysR4WRilBLd/KjZOJm7qFW0Oab5UbdSJSJ5fifU88m851oNG+p1jXrn"
    "KxHhZFKUMpRylAonsZ+XphIReS3mXmvVcKjLAY0kVSKSV+GUej6d91wreUPHzqtyfdlKRHK/Z9Tv2bzf"
    "WmEbOjNe0dFTViKiX4VKRLqVOadxzufj1GrYkIEXVuZCNwI1G1/MR6gRpByB1l+qRIQbPtUd9tRsPJ0P"
    "WoPLb7R2phLR5ruyEpHZJbESkR3xpUpENOKlG3GrBhKapqlEJE6VPlA6MhWtxl/xVEc3PSl6R5WI5oPX"
    "WHNJ52LZtaISEXlexkdU1U3zG3yUAivVnxShNHGS+KSIPpQ6KxG54O7HAN3fBB2iiRClSOWLPky6fB2J"
    "vig9o+sx6TGaiFFKdL66i9eeyL66q9d+RBMpShlKOUqFk8RbH3om2lmJSHbf/VzwTA6yd9jtSkQ6m/Yb"
    "0FbxnwtdcPcR/QnXbKrb8e5/qPrSBst3aNSoa8F3d110/Rt8q8l3aFv2GGcqEdGPl8IhzGisem5WIrJS"
    "g0OY0eha+XFITFfjOAQmQrQaqXyxHJJ8sRzCjEb0leh8cRwSfXEcwoxG9JWhlKNUOMmPQ6J7jkNiEMch"
    "TUbjJschVTDHIVjsqW7HOQ7ZU1X8Bp/nkI1ucMho/Ti0nIzGLUhf3EMpsFKDQ5jR6Fr5cUhMV+M4hBmN"
    "aDVS+WI5JPliOYQZjegr0fniOCT64jiEGY3oK0MpR6lwkh+HRPcch8QgjkOajEaWQ6pgjkOY0ajbcY5D"
    "moxGnkM2usEho/XjUPnwrf+jwy2b6FR7dIhSYKUGhyBfa+Ja+XFITFfjOAQmQrQaqXyxHJJ8sRwCXzH6"
    "SnS+OA6JvjgOga8UfWUo5SgVTvLjkOie45AYxHFIk5TIckgVzHEIFnuq23GOQ/ZU9bwestENDhmtH4eW"
    "k9G4hRmNKAVWanAIMxpdKz8OiTluHIfARIhWI5UvlkOSL5ZD4CtGX4nOF8ch0RfHIfCVoq8MpRylwkl+"
    "HBLdcxwSgzgOaZISWQ6pgjkOYUajbsc5DmkyGvnrIRvd4JDR+nFoORmNW5jRiFJgpQaHMKPRtfLjkJjj"
    "xnEITIRoNVL5Yjkk+WI5BL5i9JXofHEcEn1xHAJfKfrKUMpRKpzkxyHRPcchMYjjkCajkeWQKpjjECz2"
    "VLfjHIfsqep5PWSjGxwyWj8OLSejcQszGlEKrNTgECQTTlwrPw6J6WochzCjEa1GKl8shyRfLIfAV4y+"
    "Ep0vjkOiL45D4CtFXxlKOUqFk/w4JLrnOCQGcRzSJCWyHFIFcxyCxZ7qdpzjkD1VPTnUkdFo7fTjUFnd"
    "cQnPh2zhyPrzIZCCLSh2uY/SxEl+HBLLf3IcAl8h+opUvlgOSb5YDoGvGH0lOl8ch0RfHIfAV4q+MpRy"
    "lAon+XFIdM9xSAziOGSDZJvc92WqYI5DsNhT3Y5zHLLnpSeHbHTjesho/TjUmVlNF1qLZVZvmZxpU8DE"
    "pFGjFFipcT1kAmvSxLXy45Dtb7H8ITARotVI5YvlkOSL5RD4itFXovPFcUj0xXEIfKXoK0MpR6lwkh+H"
    "RPcch8QgjkM2yI9DqmCOQ7DYU92Ocxyyp6onh2x0g0NG68ehMo9yCddDkFe/twVSYKUGh0yrBoes5Mch"
    "G7wYh8BEiFYjJ4m+WA5JvlgOga8YfSU6XxyHRF8ch8BXir4ylHKUCif5cUh0z3FIDOI4ZIP8OKQK5jgE"
    "iz3V7TjHIXteenLIRjc4ZLReHBotJ5+66qaZT41SYKU6h1CaOMmLQy54IQ6hiRClSOWL45Doi+MQmohR"
    "SnS+GA7JvhgOoYkUpQylHKXCSV4ckt0zHJKDGA7pbDL3ZbpghkO4ZlPdjjMcssGe+dQuus4hq/XjUJkL"
    "2f96aGTTLGvPh1AKrNTgEGSLTlwrPw5JybJUabzzX3ZFXyFKkcoXyyGffGo0EaOU6HxxHPLJp0YTKUoZ"
    "SjlKhZP8OOSTT+1G7PyoorfIOysR6WxyHOqTT41rNtXtOMehXvnUduhGBRCr9eNQmQu5BA7ZNMs6h0AK"
    "RpAauo/SxEl+HJKSZVkOga8QfUUqXyyHxCRe5v0yNBGjlOh8cRwSfXHXQ7BeKfrKUMpRKpzkxyHRPXc9"
    "JAZxHFKlRHMcUgVz10Ow2FPdjnMcsuel332ZHbrJoSXkU1evyS6BQ5hPXfXcfL/MSo3rIcyndq38OOST"
    "T42+QpRMXYSX6r+wHPLJp0YTMUqmnsGLvjgO+eRTo4kUpQylHCVTh+ClSkTce66uv+5LG45DPvnUbiSv"
    "50O6YI5DmE+tOkPobfmuV1XNW/l0XnpyqCOf2trpdz1U5iD1r0Q0MqlMrUpE71qViGyreiUilPZRmqB0"
    "gNIhSiFKEUpHKB2jFKOUoHSC0keUUpQylHKUCpROUTpD6RylC5Q+oTRF6RKl3d0Obc9pWIlotJy0taqb"
    "1mNJyGQLbKvGxx+mrblW7MefqUQ02h5P3pQPVakSEb3n9VyJaLQ92B9tS5WI3AhMJSLq+WDec62SDfW6"
    "RnPgKxHh/EKUIt38qNk4mruo1ag5pPlRt1IlInl+R9Tz0bzneo0amh/1zlciwsnEKCW6+VGzcTJ3Uato"
    "c0zzo26lSkTy/E6o55N5z7UaNtTrGvXOVyLCyaQoZSjlKBVOYj8vTSUi8lrMvdaq4VCXAxpJqkQkr8Ip"
    "9Xw677lW8oaOnVfl+rKViOR+z6jfM9cvPZCsn3ev6OgpKxHRr0IlIt3KnNM45/NxajVsyMALK3OhG4Ga"
    "jS/mI9QIUo5A6y9VIsINn+oOe2o2ns4HrcHlN1o7U4lotF1WIjK7JFYisiO+VImIRrycj1gDCU3TVCIS"
    "p0ofKB2ZilZ7sRIRjb27Nx+8xppLOhfLrhWViEZlNtIS7tBsolP9SRFIQTVYs1geShMniXdo9KHUWYnI"
    "BXdes9Obgt1PrDGjEX1FOl/dNXOOZF/d9xLHaCJGKdH5sk9AWkVyTmRf9slHqyzPRzSRopShlKNUOEm8"
    "9aHUmc5KRLL77u8nzsQggm73E2tNUiI9uu2yeaGaIz2D6Qr+hGs2Ve043Td136FpMhrNzcvPndbhQlDp"
    "yGi0mnyHRlfJLTuDh/tfO+/pl5WHD6ujMhtpCRyyiU51DoEUVIO1OAT5WhPXyo9Dtr/FOAQmQrQa6Xxx"
    "HBJ9cRwCXzH6SnS+OA6JvjgOga8UfWUo5SgVTvLjkOie45AUxHLIBok2WQ6pgjkOwWJPVTvOcsiel+KT"
    "Ip5DNrrxDb7R+nGozEZaAodsolOdQyAFI0jO2kdp4iQ/DtkhFuMQ+ArRV6TzxXFI9MVxCHzF6CvR+eI4"
    "JPriOAS+UvSVoZSjVDjJj0Oie45DUhDLIRvkxyFVMMchWOypasdZDtnz0pNDNrrBIaP14tD2cjIaq26a"
    "jw5RCqxUf3SI0sRJXhxywQtxCE2EKEU6XwyHZF8Mh9BEjFKi88VwSPbFcAhNpChlKOUoFU7y4pDsnuGQ"
    "GMRxSGWTux7SBTMcwjWbqnac45ANlr85Y6+HXHSdQ1brx6HlZDRuY0YjSoGVGhzCjEbXyo9DUo4b93wI"
    "fYUoRTpfHIdEXxyHYHFi9JXofHEcEn1xHAJfKfrKUMpRKpzkxyHRPcchKYjlkCYpkeWQKpjjECz2VLXj"
    "LIc0GY08h2x0g0NG68eh5WQ0bmNGI0qBlRocgnytiWvlxyEpXY3lEGY0otVI54vjkOiL4xD4itFXovPF"
    "cUj0xXEIMxrRV4ZSjlLhJD8Oie45DklBLIc0SYksh1TBHIcwo1G14yyHNBmNPIdsdINDRuvHoTJDqf/z"
    "oW3MaEQpsFKDQ5jR6Fr5cUhKV2M5BCZCtBrpfHEcEn1xHAJfMfpKdL44Dom+OA6BrxR9ZSjlKBVO8uOQ"
    "6J7jkBTEcsgGeT0fUs2R+74M12yq2nGWQ/ZU9Xs+5M7qBodMj/04VFZ3XAKHbOHI2nPqbZACKzU4BPUv"
    "J66VH4ek8p8sh8BEiFYjnS+OQ6IvjkPgK0Zfic4XxyHRF8ch8JWirwylHKXCSX4cEt1zHJKCWA7ZID8O"
    "qYK56yFY7Klqx1kO2fPSk0M2usEho/XjUGdm9cKViLZt4cg6h0AKbKsGh6D+5cS18uOQVP6T5RCYCNFq"
    "pPPFcUj0xXEIfMXoK9H54jgk+uI4BL5S9JWhlKNUOMmPQ6J7jkNSEMshG+THIVUwxyFY7Klqx1kO2fPS"
    "k0M2usEho/Xj0HJS7bchr34PpcBKDQ5hqr1r5cchqfwnyyEwEaLVSOeL45Doi+MQ+IrRV6LzxXFI9MVx"
    "CHyl6CtDKUepcJIfh0T3HIekIJZDNsiPQ6pgjkOw2FPVjrMcsqeqJ4c68q6tnX4cWk4+9TbmU6MUWKnB"
    "IUhlnrhWfhyy/S32vT3mU6PVSOeL45Doi+MQ+IrRV6LzxXFI9MVxCHyl6CtDKUepcJIfh0T3HIekIJZD"
    "ffKpVXNknw/BYk9VO85yqFc+tTurG9dDpsd+HCpzIZfwfMimWdbvy0AKtiE1dB+liZP8OGSHWIxD4CtE"
    "X5HOF8ch0RfHIfAVo69E54vjkOiL4xD4StFXhlKOUuEkPw6J7jkOSUEsh2yQ3/WQKpi7HoLFnqp2nOWQ"
    "PS89r4dsdINDRuvHoTIXcgkcMimV9UrV2yAFVmpcD5lW9QqxrpUfh2x/i3EITIRoNdL54jgk+uI4BL5i"
    "9JXofHEcEn1xHAJfKfrKUMpRKpzkxyHRPcchKYjlkA3y45AqmOMQLPZUteMsh+x56ckhG93gkNF6cah6"
    "TbY/h6pumvnUKAVWqnMIJVNf4aXKOtx7rq6/hTiEJkKUTF2EF30xHJJ9MRxCEzFKpp7Bi74YDsm+GA6h"
    "iRSlDKUcJVOH4KVKRNx7rrJ7hkNiEMchF+TFIV0wwyFcM/Mq/0s7znHI9ueZT+2i6xyymh+HBo/fZrOn"
    "/aunq533t7OHr7Ng9v3748r1/V93TzRFqpDwLK9U5Suo2Ay932/em51HmB9lG+PTahlb+u4WRVRnZusH"
    "5xvjT50BG0OK2KjeW273tfmWfrRZvXnd/tHGeK+zt7Kzrr4ON8ZHXQHHG+OTLn26UboqXxhujRxsjA8q"
    "ffC8hjvvf1x9nSVXD19v7h5Xvs++0Hquv35DXyo83Hz9Nv/L0/0PWufVlT/un57ub6s/fptdfZ7Ry8fr"
    "r6nxl/v7J/cXGnjw6/7hz2rPdv4fAAD//wMAUEsDBBQABgAIAAAAIQA28DM9JBIAAJJ5AAAYAAAAeGwv"
    "d29ya3NoZWV0cy9zaGVldDQueG1srJ1bU+NIEoXfN2L/A+GnmfF2Y124RsPEDLR1v5phn91gGscAJmz3"
    "bTf2v2+WqgpUOlVl5HbHTGO+zkqp5JNSouNCH37//viw93W2XM0XT2cD5/1osDd7ulnczp8+nw3+uhq/"
    "Ox7srdbTp9vpw+Jpdjb4MVsNfj//5z8+fFss/17dz2brPcrwtDob3K/Xz6f7+6ub+9njdPV+8Tx7on+5"
    "Wywfp2v6dvl5f/W8nE1vm0GPD/vuaHS4/zidPw14htPlW3Is7u7mN7PLxc2Xx9nTmidZzh6ma9r/1f38"
    "eSWzPd68Jd3jdPn3l+d3N4vHZ0rxaf4wX/9okg72Hm9Oo89Pi+X00wPN+7vjT2/2vi/pP5f+9+RmGg5b"
    "epzfLBerxd36PWXe5/uM0z/ZP9mf3rxkwvm/KY3j7y9nX+fsDXxN5W63S87BSy73NZm3ZbLDl2TscC1P"
    "v8xvzwb/HYk/7+irw/4avRt57K/Wn/8Nzj/czukdZrPaW87uzgZ/OKfX/miwf/6hEdD1fPZt1Xq9t7pf"
    "fAuW89t0/jQjNZKOmUI/LRZ/s9CItjyipM/Tp9ne98kzvc9nA9qrH68v14vndHa3vpg9PJwNLv3B3vRm"
    "Pf86K2nE2eDTYr1ePNbzz/frph7WxO6Wi//Mnpo9mj3MKJh2leU/G1AqHsr2tvNvPBPbkm5kezs0eP9l"
    "ru3Xct7jprbK5d7t7G765WFdL76FM7aLVMj+e5pBI87T2x+Xs9UNVQsdgvce2+rN4oFy0N97j3NW9ST2"
    "6ffm67f57fqeXh0O9m6+rGjK/xZADOMDSBvNAPoqBtBZwhJPB5rHvw440g3Y5zvWTPpyup6ef1guvu1R"
    "XdAerujQ0vvqnLqUhE3RpQnybb5M2jBjmipL8gfLcjY4pE3T8BWp4eu5444+7H+lg3sjgv7kQQd+M2E2"
    "7ALIJZCPQMZAAiAhkAhIDCQBkgLJgORACiAlkApIDWQC5ArIX0CuBTlgx3mf3uaX95reXnivHYL93muW"
    "5WxwoLzVTuet1sW4asyFLsZTYy55zBGpvKUrXw36KBK96moMJAASAomAxEASICmQDEgOpABSAqkEad7H"
    "psxqTg7plPByOA47h34Co67EKOUYHqjH8C9dzKEac62LOXqJUZRGbxgozR+9H/knrT8vJ6q3nmVY1uaM"
    "1RJDV3m6mK7ydDFd5fGY9pF23OOO8HjMiXJgT9SYsSaP1zkxBpo8XmdeoS5PZ16RLk9nXrEuT6egEl2e"
    "jmBSXZ6OYDJdnlfBNJLOdXk6x7nQ5ekc51KTx+8c50qTx+8c51qjDb9znCe6mM5xvhIxdCVtibVbddqg"
    "btlpgwx1R5vDMzydqvud4VkW6uGoU3ndd78jkj9FUPuSf9DZ9QtNjON3DgLrB1lVH760BR+BjIEEQEIg"
    "EZAYSAIkBZIByYEUQEogFZBakCPqmO6ag30/Xc5uB7wvr/3T2mUX83nTYV/6w7E/DP1h7A9Tf5j7w5Le"
    "kzvWa3WEPrGmnfinE5nWGZx/9IeBP4z8YeIPM39Y+MPKkPZKpBUnxfVSu9dX/umVTO8OzqPxx7ou6l9q"
    "f5/+H078X/81eDf4Vez4u85lSLeFfimuRQpxVuY72Urxm2vfD+UyRpWzg3JiWTaWkwiylpMmxvE7NXfJ"
    "g9rlBGQMJAASAomAxEASICmQDEgOpABSAqmA1IJAOfEKonYYS8U2xGGnJBhyJYYYy4BtzeVDu/rWDX0V"
    "5wFJ82A4ObCVyLVIYdD3wW+ePYmibzrD70DfLMtGfYsgq741MY7faRgueVBb30DGQAIgIZAISAwkAZIC"
    "yYDkQAogJZAKSC1IH33bhhj0LYZso2/d0Bd9/1IfDp33B7/Vh7+SRg+Hk0Or0EUug9APmxO5JYkidJLd"
    "DoTOsmwUughShN7R8IUmBvsiHtQWOpAxkABICCQCEgNJgKRAMiA5kAJICaQCUgvSR+i2IQahiyHbCF03"
    "9PVEfkT6PhpOjqz6FikM+j5q9G1JouibWQndu3gOXSn69f0sy0Z9iyDa+5cfDg66+tbEYKPCg9r6BjIG"
    "EgAJgURAYiAJkBRIBiQHUgApgVRAakH66Ns2xKBvMYS+sLcK+3VLo6Ib+qrvY9L38XBybNW3SGHQ93HT"
    "qFiSKPqmJDvQN8uyUd8iyKpvTQw2KjyorW8gYyABkBBIBCQGkgBJgWRAciAFkBJIBaQWpI++bUMM+hZD"
    "ttG3bmirUTnhjcoJa1ROhpMTq9BFLvryWmavtXLSnMgtSRShO+Su7UDpTZqNUpdRitY7N98udEHYrIio"
    "ttoRjREFiEJEEaIYUYIoRZQhyhEViEpEFaJaoj7at44xiF+O2Ub92rGvknVGJHtnNJw4I6vwZRqD8p1R"
    "I31rIlX87GbXz7cxDjfANty/lFF28YtU7SDsZEQqRfx8YAuNMSpAFCKKEMWIEkQpogxRjqhAVCKqENUS"
    "9RK/OFTaMSbxizFbiV83tiV+h4nfIfE7dvGLNCbxO02DY02kip/ZazsQP3fpNolfRNnFrwnCNocZyJ37"
    "94jGiAJEIaIIUYwoQZQiyhDliApEJaIKUS1RL/GLo9dL/GLMVuLXjW01Po7LOx/HZa0PfTehV5Y79tdy"
    "0qYycPk1wJZILQNmd+2gDLhrtqkMRJRSBh1n8cLRBGkaIGHTtWwsMVC5BkBUgFEhoghRjChBlCLKEOWI"
    "CkQlogpRLVGvMhDHpVcZiDFblYFubOsawG6iOx6J37OLX6Qxid/j4rclUsXPzKsdiJ97YJvEL6Ls4tcE"
    "aRogcDY/OoDGiAJEIaIIUYwoQZQiyhDliApEJaIKUS1RL/Gb7Fp2f8bUAG3yYi33duQ+KoXTEj9zah2y"
    "ah2rV3st05jE7/MGyJZIFT9ztnYgfm6QbRK/iLKLXxOkaYDAh/zIbr2qPdEYUYAoRBQhihEliFJEGaIc"
    "UYGoRFQhqiXqJf4tDFi5na3O/FYLloR/IBqgg6YBIjvWsfuxcmdMZXDArwG2RGoZMN9rB2XA7bNNZSCi"
    "2mVw2Pk0yQX7yDK7b6r+ENz9GI+IUn4IBgdzjFEBohBRhChGlCBKEWWIckQFohJRhaiWqFcZbOHTyu1s"
    "VQZWp5amwK4B5K46do9W7oJJ/NyltSZSxc9MsR2In3trm8Qvouzi1wRpGiAepYgf0NgBFCAKEUWIYkQJ"
    "ohRRhihHVCAqEVWIaol6iV8cl17dvxizlfh1Y1sNELNvHbJeHbuBK6dqEv8Rb4BsiVTxM8dsB+Lnxtsm"
    "8Ysou/g1QZoGSDh97R99AY0dQAGiEFGEKEaUIEoRZYhyRAWiElGFqJaol/jFceklfjFmK/HrxrbvAB2L"
    "Bui4aYDIoXXsPq+ctKkMjnkDZEuklgHz03ZQBtyW21QGIkopg85nwi8cTZDmDpDwAdtlAGgscrWuFAGi"
    "EFGEKEaUIEoRZYhyRAWiElGFqJaoVxmIQ9WrDMSYrcpAN7Z1DThh1wBybR279yunahI/d3+tiRTxu7vx"
    "f5s0G/1fGWUVvy4IGyAR1W6AEI0RBYhCRBGiGFGCKEWUIcoRFYhKRBWiWqI+4reOMdwBkmO2Eb927Kv4"
    "Xeb/uuT/unb/V6YxiN8dNQ2QNZEq/t34v83qg83i11i73cV0FzKV+qNv9yPJIkoRP/q/GBUgChFFiGJE"
    "CaIUUYYoR1QgKhFViPhyFbpp2Uv8W/i/ctNbid/q/5LwHd4AuQ5rgOg7Wi1jdQHkzpjKwGkaIGsitQx2"
    "Y4G5wrcSx4gvyvY7H+/5U0aJnW+iYB2XLggbIBGllAH4XWOMChCFiCJEMaIEUYooQ5QjKhCViCpEtURQ"
    "Bs3vW6B/Pq3pFy40C7u8wfml6w3psAxp0kOa0pB2eEi7M6SNGdZ32fNPKP9E5vdphRflomM8pCM4pOMz"
    "pNkPaW5D2nN9/iuZ31hezTwo7PRKbuegtdTLZfYZbYB2xLraS7uZ1mXobWmuZRpT/XEXzro/av3txoVj"
    "v0WB3TbdUH8iyl5/miBND4YunNiHVkmOEQWIQkQRohhRgihFlCHKERWISkQVoloi7WWo+T0duBDMOqb5"
    "3QK4EkyOsV6GmvX7X8+7S8G0Y1viZy6cS+aZa3fhZBqT+LkLZ02kin83LpwrrLMN4hdRdvFrgvAmlNig"
    "cvFBFw6jAkQhoghRjChBlCLKEOWICkQlogpRLVEv8dtcOJP4dU5aexEzs6BN4t/gwrnChXMbF46+m9Ar"
    "28eQ5KRNZcBdOGsitQx248K5wjrbUAYiSimD7poamaodpOnBwHL7KAYq1wCICjAqRBQhihEliFJEGaIc"
    "UYGoRFQhqiXqVQY2F85UBm9ZLmkqA7sL5zIXziUXzrW7cHKqJvFzF86aSBX/blw4V1hnG8Qvouzi1wRp"
    "GiB04cQ+KOJHFw6jQkQRohhRgihFlCHKERWISkQVolqiXuK3uXAm8b/FhTOJ3+7CucyFc8k8c+0unJyq"
    "SfzchbMmUsXP/JGftx9cbrNs6v5FlF38miBNAyR8nZb9IPZBET9EBRgVIooQxYgSRCmiDFGOqEBUIqoQ"
    "1RL1Er84LtoxJvGLMVt1/7qxLRfOFS6c27hw9N2EXlkbIJHQVAbchbMmUsuAEu2iDFiazT8EiyilDLoL"
    "0VxNkKYB4lHKzwGAxiJX24VDFCKKEMWIEkQpogxRjqhAVCKqENUS9SoDcah6lYEYs1UZ6Ma2fghmLpxL"
    "Lpxrd+HkVE3i5y6cNZEifm83LlyTZqP4ZZRV/LogbIBEVFv8iMaIAkQhoghRjChBlCLKEOWICkQlogpR"
    "LVEf8VvHGK4Bcsw24teOfRW/x1w4j1w4z+7CyTQG8XvchbMmUsW/GxfOE/6avfuXUXbxi1TtIGyARCpF"
    "/OjCYVSAKEQUIYoRJYhSRBmiHFGBqERUIaol6iV+mwtnEv9bVmEaun+5j0rhtBogT7hwXuPC0XcTemVr"
    "gGRCUxlwF86aSC2D3azH9MQiyg1lIKKUMuguRJOp1DLofg5bRCllAEs0xxgVIAoRRYhiRAmiFFGGKEdU"
    "ICoRVYhqiXqVgW09pqkM3rIe01QG1vWYNAV2DaDFk559Faacqkn8fBWmNZEq/t1Y0N6bLGgZZRe/SKWK"
    "v/vLD0UqRfxoQWNUgChEFCGKESWIUkQZohxRgahEVCGqJeolftsqTJP437IK0yR++ypMj/m/HtnInn0V"
    "ppyqSfxe8zEkayJV/Lvxf703+b8yyi5+rf/b/RiSSKWIH1dhYlSAKEQUIYoRJYhSRBmiHFGBqERUIaol"
    "6iV+2ypMk/jfsgrTJH7d2HYD5POPIXk++xiSR04wvbI2QCKhqQz85mNI1kRqGezGCfbe5ATLqHYZwEI0"
    "XRDeARJRShmgE4xRAaIQUYQoRpQgShFliHJEBaISUYWolqhXGWzhBMvtbPVDsNUJpikw8ZP/69n9X7kL"
    "JvFz/9eaSBX/bvxftvpg8+1PGWUXv8Yk1twBQv9XZG+7AIgCRCGiCFGMKEGUIsoQ5YgKRCWiClEtUS/x"
    "b+H/yu1sJX67/+sx/9cj/9ez+79yF0ziP+QNkC2RKv7d+L9swfAbxK+xdvHMr/V/oQFC/1fsgyJ+9H8x"
    "KkQUIYoRJYhSRBmiHFGBqERUIaol6iX+LfxfuZ2txG/1f0n4R6IBOmoaIHKCPbsTLHfGVAb8V+pSWnMi"
    "tQyYpfbzTjCzgN9QBiJKuQZ0F6LJVBvuAAkrsOUEi4FKGUBUgFEhoghRjChBlCLKEOWICkQlogpRLVGv"
    "MhDHpY8FJrezVRmI7elvhNIU2DWA/F/P7v/KXTCJn/u/1kSq+CnRLsTP0my2wESUXfyaIE0DxKOU7h/Q"
    "2AMUIAoRRYhiRAmiFFGGKEdUICoRVYhqiXqJXxyXXuIXY7YSv25sywJj/q9H/q9n93/lVE3iP+ENkC2R"
    "Iv5mUcPPn/mbNBvFL6Os4tcFoQUmotriRzRGFCAKEUWIYkQJohRRhihHVCAqEVWI+PIa00I0wwoAmaaP"
    "+OWYbcSvHdu6A+SPeAPkj1gDRN/Roh7rHSCZ0FAGPv99vNZEvAz4w1H5c0IfZ8vPM/bo1tXezeILe9Qp"
    "LVn68EL5o6ou3dOqeeRNh7NHyzbrVzu8dukZVM3zKV/Ts0fIfp5l0+Xn+dNq74Ee48qeqnpE9weW/Mmr"
    "/Bt6AmzzDED+PNfm5T09+3hGzxMdvafgu8ViLb9hD8B8eZry+f8BAAD//wMAUEsDBBQABgAIAAAAIQAr"
    "ixNtTAoAAD4+AAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDUueG1srJtbU+M6EoDft2r/Q8pPu+ccSGxM"
    "gBTh1JBgiB3YU+zt2ZM44JokTtnmMru1/31bUssXdSQcO9QMCe3ullr6JMtq6/r3z8269x6lWZxsx5Z9"
    "OrB60XaRLOPty9j65z+8k0url+Xhdhmuk200tn5GmfX7zZ//dP2RpD+y1yjKe+Bhm42t1zzfjfr9bPEa"
    "bcLsNNlFW7iyStJNmMOf6Us/26VRuORGm3XfGQyG/U0Yby3hYZQ28ZGsVvEimiaLt020zYWTNFqHOdQ/"
    "e413mfS2WTRxtwnTH2+7k0Wy2YGL7/E6zn9yp1ZvsxjNXrZJGn5fQ9yfthsuep8p/HPg/5kshstJSZt4"
    "kSZZsspPwXNf1JmGf9W/6oeLwhONv5Eb2+2n0XvMOrB05bSrkn1e+HJKZ2ctnQ0LZ6y50tFbvBxb/x3g"
    "zwl82uzX4GTgsl+Vn/9ZN9fLGHqYRdVLo9XY+maPnhzH6t9cc4D+FUcfWeV7L3tNPu7TeDmPtxHQCBwz"
    "Qr8nyQ+mOoOSB+B0F26j3uffd9DPYwtq9bP8mie7ebTKJ9F6PbamrtULF3n8Hv0BFmPre5LnyeY5fnnN"
    "+XjIQbZKk/9EW16jaB2BMlSV+R9b4Eqostoq14QnVtI+y2o5YNwvYq1+l3F7fGz9kfaW0Sp8W+fPycdD"
    "xKoIA9k9hQg4nKPlz2mULWC0QBOcnrFSF8kafMDv3iZmox5gDz/550e8zF/H1rnVW7xlEPG/xd82Wgl9"
    "QIPrwyfqQ8eYDKChuQF8ooFtNoC6c4NhYXBhLACucn34xALA0hABTGpcHz5R3+z/CvXhE/XB0uDfBvxE"
    "kzIORRuCrcmi6ASI3VSnvug8DsY0zMOb6zT56MHcAQ4ywA/Yt0c2OGEYOPApCi3A0FABODAn35iXsTWE"
    "5gDzDEbM+43tXl333wHABSrdCqVzl1PBzCZEMiWSOyLxiOSeSB6IZEYkPpEERDInkkcieULJOYurD81a"
    "tC3QStq2RLhp2zIv0LaATtm25wOlbYVStW1RwmvFO2kqJBcwSiud5NYd3RFHHpHcE8kDKWyGhcGAqhR2"
    "WS/MJ44CIpkTySORPFWLr/UATB2kB9zh6aCYqZp2AnPEB0ulD2ylD/bpOHWdyT6ds7rOVOjU+1vtpn06"
    "53U/3j6dYV3nfo+Oo/h52Ofnou5nJnRgsqq0j9rb+3SUOSLYozNUWJ/v01H64nGfjtIXT9gXMNmVlR6W"
    "nVEDCZToUG4DEnM0tly2qoA5UiEIL8KEUqmR0vWTwkNVSekzthBhvA6LufaOSDyUXEBVVtzXa5hGS0ss"
    "nDx35LH7cxbzJdDMu3t+/tvzX6bur3fub4O/XvdXLACla+7RJbKbp3td37uje+natm5K133P/c06saTz"
    "k3rrPKBzhEzj/MEdPdhwH2L1dqrOf3HM7mekgXwiCYxNFrijQMZ1VhY9c3/19U02b9Rkc3c0l67dqut+"
    "YGqyx0ZN9uiOHmWTnVedQ5MZ3T8hioAJ41XTI08uX4LzHhmW7u/dk7mm7rWRB1P1cUYec1SMPGUuuMWL"
    "9ZGnTJaTwoNp5Aml6sgjEg8lZOSJwQbLCDqy0EQ7spgpX2+/36jjBk2140YMFV6qYjojVfeJJDAFw58b"
    "SDDzJsHwZSIJ5rFJMHzNQ0yfsPu0tLJ24BNm1bTGIqxmjsMicwQswiwBXa3c/2/xIgyuyl1AueFOCg9V"
    "JeWOOxVKVRaJxEPJISyiSRsW0bQNi6TqPpEEpmA0LDYJRsNik2A0LGL3dWCR3ePUBze7zYqEOSpYVBYb"
    "t3ixzqKyaJsUHkwsCqUqi0TioeQQFtGkDYto2oZFUnWfSAJTMBoWmwSjYbFJMBoWsfs6sMh2e4/CInNU"
    "sKisaW/xYo3FC2UFOik8mFgUSlUWicRDySEsokkbFtG0DYuk6j6RBKZgNCw2CUbDYpNgNCxi93VgEZrw"
    "OCwyR8AirIjhHq0sBW/xYm1740J5nJsUHiosXigrz6lQqrJIJB5KDmERTdqwiKZtWCRV94kkMAWjYbFJ"
    "MBoWmwSjYRG7rwOLbEP3ODByTwWNymLwVl6t46isKyelDxOPqFUFkoo8KToESWnThklp2wZKWn2figJj"
    "RBouG0WkAbNRRBoyZU92QZPtWh/lnm2L/W+cKJVHkFt5tY6muqdV+jCiKUqqoUlEHvqCHQ1lX8vwdC1t"
    "WqGJVWiFJqm+L6tSbt0Fxoh0aKJnY0Q6NJtEpEMTaeiCJtt2PA6aIjGAaJJ8FF6to6msOicsYSPWAUY0"
    "hVYNTSLy0NdhaKKbVmiibSs0SfV9Wf0qmqi1d7Dp0GwSkQ7NJhHp0MSe7ILmvoRSq6duW2QfEE1bzefJ"
    "y3U21Q3J0omRTUx0VJIBaFjB1ZOig6ZN9NyKTbRtxSaJyKcRBcaIdGw2iUjHZpOIdGwiDl3YZLvxx5k2"
    "cV9fPPrYapaKvSbAZsQ6m+oGpdSCXddyG5M8/aBWbd4kaRdPah3E5leJFcOOuSyvFZs0j0SDDIwR6dhs"
    "EpGOza8yQaw1dGx+leb5euPcPloWh3sqHoRsNY8jL9fZVBalk9KJkU2aykHD2rzZIpkj3bSaNzukc2j1"
    "fSoKpOige3qHlI4szzjadGx2T+pAZx5r3sRdfZw31bwOL4jMm+pmutT6Yt4kuY87NKyxacqHDNi7ATTR"
    "KN20YrNJRoS/jUmyczNafZ+KAik6iM0OKR5ZXis2uyd5eFr+OPd03OVHNtU8Dy9IZfNS3VyXWl+wSVM9"
    "aFhjs0WyR7ppxWaHdA+tvk9FgRQdxGaHlI8srxWb3ZM+NturPw6buOvPGg5eilLzPrwgYBN7XShdqpvt"
    "UgsCK9ebl+puO2rV1puYrigfjzypddB6s0nCRPOGhiyv1pPli1H25S9nfaiU4eWoGQ3Mp6LAGJhu2dkk"
    "MN2yE233BwaVhsCgUobAnmS/dnkyYjv5xyEVcwJIqpoVsvFynVSyD184MZKKmYvqUzsReVjiYTtK6KbV"
    "LIq2OlKvOKlXRlJJFL6MorqxhFoHTaZNAtORagxsZrPAAtsU2JPs/Q6kOkdLGHFPMGkiqWrGSF6uk6pu"
    "y5dOTKSiVnVOpSJPig6ZU6VNG1KlrYZUZ8BIdQYmUmkUPhUFxsA0c2qjwDSkmgOb8cCgUqY5VfZrF1KP"
    "lj9yMGOApKoJJHm5Tqq6S186MZJKE0hoWF2ZStFBpDZJt2ju/rI8Hak2J9U2kkrzSDSwwBiYjtQOeSRz"
    "YDOHBQaVMpLaPZ3E35s+yt2feyrnVDWfJC/XSVX37EsnRlJpPgkNa6Sasi+a53vpptWcui/7Uq5THYeT"
    "6hhJpWklGph4jR5eaD7k7t8oMN2cagxsxgODShlJ7ZBdEgc2xbm8TZS+ROw4adZbJG/s+CW88H9dSMXp"
    "jJkD79Sz/RtFPnVGcBaLyr85o8leORyP5bNSvyyWHXd9iR7D9CXeZr01HDllJ0DZsbFUnBIVf8BpVX4g"
    "Spw95V9f4Zx2BOf6BqegvEqSXP7BDsYVJ79v/g8AAP//AwBQSwMEFAAGAAgAAAAhAIPmjfHhBQAAnyEA"
    "ABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0Ni54bWysWtuO4kYQfY+Uf7D8PtjdGBgQsFqwN1kpkaIol2eP"
    "MWAtxsj2XFZR/j3VDYNx1fFoMvJol0vP6dPt6lPVx+6Zf3rJD85TWlZZcVy4auC7TnpMik123C3cP//4"
    "cnfvOlUdHzfxoTimC/d7Wrmflj/+MH8uym/VPk1rhxiO1cLd1/Vp5nlVsk/zuBoUp/RIv9kWZR7X9LXc"
    "edWpTOON7ZQfPO37Yy+Ps6N7ZpiV7+EottssScMieczTY30mKdNDXNP8q312ql7Z8uQ9dHlcfns83SVF"
    "fiKKh+yQ1d8tqevkyezr7liU8cOBrvtFBXHivJT0T9P/4eswtl2MlGdJWVTFth4Qs3ees7z8qTf14uTK"
    "JK//XTQq8Mr0KTML2FDpj01Jja5cuiEbfpBsfCUz4Spnj9lm4f7jX37u6F2ZF//OH5mXm59/3eV8k9EK"
    "m6tyynS7cD+rWaSnrrecWwH9laXP1c1np9oXzz+V2eaX7JiSGknHRqEPRfHNQL/SyL7p7IneX6xCfyud"
    "TbqNHw/178Xzz2m229eUDsEgoJmbpZ9tvodplZDmiGgwNFRJcaAZ0KuTZyZ3SDLxi31/zjb13nR3neSx"
    "qov870vDpdu5A0XYdqD31w6UbG90oCjaDvT+2mH8Zgca33ag90uHgAZ7Y4TRpQO9XzoMYQfvfO02lmFc"
    "x8t5WTw7JGAKQnWKTTlQMyIxQdSjwZQ+nke9RrYjrBRPQ/PZ8Czc8cR1iKEiNTwt1f1k7j3RCiYX0OoM"
    "GgU2qqbbWrSEoiW6tIysGmja17nTld7OHS776/wMli6tNb17Nj2EmbYxa4CZ+m1MiDCqjYkQRl8x3u1l"
    "knr4EqnxwP/fS2R4Fu6oFYPpkMXgDAps3tqVXYuWULREty2tuZOQe5m74aG5U5Fo5DUN2NwhaMQWEILG"
    "bAUhiKk5gqBGU6040GL1EgfDQ3GglLiJA9PoCoBo72ZxgCCm0hCCGplagUQQ1OiqFQcqf73EwfAwPWif"
    "6wGCuB4giOsBgrgeIKhDD1Qne4mD4WF60D7XAwIprgcI4nqAIK6HK8js3reFzDhTttd8qJAZHlbItOKF"
    "7Ay6LWSiJRQt0W1La+7TnuZuePiCKS5cCOLChSAuXAjiwoWgDuEqKsG9LKIl4imsuHQhSnPtYhQXL0Zx"
    "9WJURzlTwD59SNKWiGtac11cULeilk2hbIpaTS1dK2airIv+iLuwRHw9NRPtCqGUZqpdQy6fcYV4RMYV"
    "YVSTAu1o9OW11MVs3W7UWnPDiVHcckKUjAYaccgyJcIjdkWjL/dm7qvEdj1k+bnCKJafa4iS0YAjsj0i"
    "wiM2WdfWRl8eTiHDNBSZAlEiU6D5EpkCufg2gOfVFY2+nJxCtmkoMgWiRKa8x/GFcMRAZAriChrNtrXR"
    "l59TyGEFLAdWGMXUvYYomSlwRLbzRHjErmj0ZfTUxend3vbpQGQKRIlMQSgZDcglMgWiuhxTX9ZRIZ8W"
    "CMcEUEqzK1hDLhkNNOJIZApEdWhD9+UfLRH3GyOeKRjFMwWiRDQwF88UjGo026obui//aIm4fxyxHFhd"
    "ULf+UTaFsilqNbWvoC//qM9PxFqPePSIP0HEKLZbrDGKZUoIUWOubozqUndf/lEjNzcW6oYooW6EGolo"
    "QC6hbojqUndf/lEjNzcW6oYoXvkgl4wG5GI6i/C8msi2M6Uv/6iRm5sw3a4winnuNUTJaMAR+Z0zHrHj"
    "zln35R8tEd8HJvzOGaP4IxWIktFAznDC7zvxiB13Wrov/2iJ+POliTjoQJ5vIo46EEpGA6HuRRWFqK4q"
    "2pd/1Min3YsqClGiiiKUjAbkElUUorqqaF/+kQ5j5bPHe1FFIUpUUej5xJ4CuUQVhSheRc/nwOezyzwt"
    "d+k6PRwqJykezQmvohPHa2tz/qzsmWEDX85P8S79NS532bFyDunWHg5PqLyV5wNkf2C+1MXJnow+FDWd"
    "BduPe/pDiJROLP0B/X5bFPXrF/OQ+/qnFcv/AAAA//8DAFBLAwQUAAYACAAAACEA6aYluGYGAABTGwAA"
    "EwAAAHhsL3RoZW1lL3RoZW1lMS54bWzsWc1uGzcQvhfoOxB7TyzZkmIZkQNLluI2cWLYSoocqV1qlxF3"
    "uSApO7oVybFAgaJp0UuB3noo2gZIgF7Sp3Gbok2BvEKH5EpaWlRsJwb6Fx1sLffj/M9whrp67UHK0CER"
    "kvKsFVQvVwJEspBHNItbwZ1+79J6gKTCWYQZz0grmBAZXNt8/72reEMlJCUI9mdyA7eCRKl8Y2VFhrCM"
    "5WWekwzeDblIsYJHEa9EAh8B3ZStrFYqjZUU0yxAGU6B7O3hkIYE9TXJYHNKvMvgMVNSL4RMHGjSxNlh"
    "sNGoqhFyIjtMoEPMWgHwifhRnzxQAWJYKnjRCirmE6xsXl3BG8UmppbsLe3rmU+xr9gQjVYNTxEPZkyr"
    "vVrzyvaMvgEwtYjrdrudbnVGzwBwGIKmVpYyzVpvvdqe0iyB7NdF2p1KvVJz8SX6awsyN9vtdr1ZyGKJ"
    "GpD9WlvAr1cata1VB29AFl9fwNfaW51Ow8EbkMU3FvC9K81GzcUbUMJoNlpAa4f2egX1GWTI2Y4Xvg7w"
    "9UoBn6MgGmbRpVkMeaaWxVqK73PRA4AGMqxohtQkJ0McQhR3cDoQFGsGeIPg0hu7FMqFJc0LyVDQXLWC"
    "D3MMGTGn9+r596+eP0Wvnj85fvjs+OFPx48eHT/80dJyNu7gLC5vfPntZ39+/TH64+k3Lx9/4cfLMv7X"
    "Hz755efP/UDIoLlEL7588tuzJy+++vT37x574FsCD8rwPk2JRLfIEdrnKehmDONKTgbifDv6CabODpwA"
    "bQ/prkoc4K0JZj5cm7jGuyugePiA18f3HVkPEjFW1MP5RpI6wF3OWZsLrwFuaF4lC/fHWexnLsZl3D7G"
    "hz7eHZw5ru2Oc6ia06B0bN9JiCPmHsOZwjHJiEL6HR8R4tHuHqWOXXdpKLjkQ4XuUdTG1GuSPh04gTTf"
    "tENT8MvEpzO42rHN7l3U5syn9TY5dJGQEJh5hO8T5pjxOh4rnPpI9nHKyga/iVXiE/JgIsIyrisVeDom"
    "jKNuRKT07bktQN+S029gqFdet++ySeoihaIjH82bmPMycpuPOglOc6/MNEvK2A/kCEIUoz2ufPBd7maI"
    "fgY/4Gypu+9S4rj79EJwh8aOSPMA0W/GoqjaTv1Nafa6YswoVON3xXh6Om3B0eRLiZ0TJXgZ7l9YeLfx"
    "ONsjEOuLB8+7uvuu7gb/+bq7LJfPWm3nBRaa5HlfbLrkdGmTPKSMHagJIzel6ZMlHBZRDxZNA2+muNnQ"
    "lCfwtSjuDi4W2OxBgquPqEoOEpxDj101I18sC9KxRDmXMNuZZTN8khO0zThJoc02k2Fdzwy2Hkisdnlk"
    "l9fKs+GMjJkUYzN/ThmtaQJnZbZ25e2YVa1US83mqlY1oplS56g2Uxl8uKgaLM6sCV0Igt4FrNyAEV3L"
    "DrMJZiTSdrdz89QtmvWFukgmOCKFj7Teiz6qGidNY2UaRh4f6TnvFB+VuDU12bfgdhYnldnVlrCbeu9t"
    "vDQdbude0nl7Ih1ZVk5OlqGjVtCsr9YDFOK8FQxhrIWvaQ5el7rxwyyGu6FQCRv2pyazCde5N5v+sKzC"
    "TYW1+4LCTh3IhVTbWCY2NMyrIgRYZoZwI/9qHcx6UQrYSH8DKdbWIRj+NinAjq5ryXBIQlV2dmnF3FEY"
    "QFFK+VgRcZBER2jAxmIfg/t1qII+EZVwO2Eqgn6AqzRtbfPKLc5F0pUvsAzOrmOWJ7gotzpFp5ls4SaP"
    "ZzKYJyutEQ9088pulDu/KiblL0iVchj/z1TR5wlcF6xF2gMh3OQKjHS+tgIuVMKhCuUJDXsCLrlM7YBo"
    "getYeA1BBffJ5r8gh/q/zTlLw6Q1TH1qn8ZIUDiPVCII2YOyZKLvFGLV4uyyJFlByERUSVyZW7EH5JCw"
    "vq6BDX22ByiBUDfVpCgDBncy/tznIoMGsW5y/qmdj03m87YHujuwLZbdf8ZepFYq+qWjoOk9+0xPNSsH"
    "rznYz3nU2oq1oPFq/cxHbQ6XPkj/gfOPipDZHyf0gdrn+1BbEfzWYNsrBFF9yTYeSBdIWx4H0DjZRRtM"
    "mpRtWIru9sLbKLiRLjrdGV/I0jfpdM9p7Flz5rJzcvH13ef5jF1Y2LF1udP1mBqS9mSK6vZoOsgYx5hf"
    "tco/PPHBfXD0Nlzxj5mS9mr/AVzxwZRhfySA5LfONVs3/wIAAP//AwBQSwMEFAAGAAgAAAAhAKSOwAbz"
    "BgAAqkkAAA0AAAB4bC9zdHlsZXMueG1s7Fxfj5s4EH8/6b4DQrqnUxZMgE22Sar8AalSrzqpe9I99IUQ"
    "krVqIAJnm7S6735jA8FRliTscgmuLittwDHjn2fG4xl7zOD9NiTKc5CkOI6GKrrTVSWI/HiBo9VQ/evR"
    "7fRUJaVetPBIHAVDdRek6vvRr78MUrojweenIKAKkIjSofpE6fpB01L/KQi99C5eBxH8soyT0KNwm6y0"
    "dJ0E3iJlD4VEM3Td1kIPR2pG4SH0LyESesnXzbrjx+Hao3iOCaY7TktVQv/hwyqKE29OAOoWmZ6vbJGd"
    "GMo2KRrhpUfthNhP4jRe0jugq8XLJfaDY7h9ra95fkkJKL+OErI03Tjo+zZ5JSVTS4JnzMSnjgbRJnRD"
    "mip+vInoULX2RUr2y4cFyNg2VSWTyjReAJ/0O/03VSsePqhpHdWsqGgfVvzyu/7uS0d/p1dUvz+qfscf"
    "gP+ApuIh0EQRNrQBwFkz7KvohJazYDRYxlHJCQQAueQfvkbxt8hlvwErgD+s2miQfleePQIliDXuxyRO"
    "FAp6DOzhJZEXBlmNqUfwPMGs2tILMdllxQYr4Kqf1wsxKCIr1LIWsv9zVqtoq1u2lazmQ9XNP7z/+wbH"
    "CfbISUK9Izom/7ydTgaoLp1+Q/0ShVGbQQWXj8GgsT3pTi/pVDUNnX/ezpj7mTXWM4U/I/BqLL0JGqNx"
    "XSyIt5pp+quZe6x5l2tM0aFjGvUFdEzjcqYcMLYc5y8Pu2opGPxzWgrcFKRgODAheyNtcCsEJaMBTGg0"
    "SCIXbpT8+nG3BhsUwdybmQBejz1/ovYq8XbIsC5/II0JXjBbuJpyy5frAxobE5MrliYgYxbtEhQVRB3H"
    "NV2XY2uOqOFYep8rdINISzVskGiplw0StWam3r1vmKfmGJk9biYbRAouX99qmujE7dp600Qd20KZZW5S"
    "+Xuu5fAZqUGiYHJnTtPSB6Ld6axhlXJdZ+I4DROd9hx7+h/wtOdwz69JQU0d02laT12L/TUuKDaPVxLl"
    "EwDMY/M4WUDoWIQbJkwhWdFoQIIlBV83wasn9k3jNfyfx5RCeDUaLLC3iiOPMPe4eEJ8EkJOiC6HKn2C"
    "6LDwx70NjXN3XGPkc+pn63IMHMLZqgCzQHm2btaZ83055MJNIbSRVbn8QZv8gJDPTO5/L/cqxXyj7VKI"
    "SWF9gkVSLJBll+CI5JeZ+mQ3TK1Eahltgez9/avoKtvlvoEqVN0SlaEqJSoEAWj+tOKt12THYlCmzvkd"
    "9KS8m/CBVd6PCV5FYZA9MBpACJrdKk9xgr8DIRa7+vB7AOscsJpDsS+WfEu89WOw5c0x3myX1VxliwM5"
    "V+FSavx9yfFDiCa3AGQfALAAIaEABAsk5QgW8MP6o3wWSMAPq45S44d1Wqnxw5Kx1PilnMEE/ZdzBhM7"
    "IOUMwHZjch8OtWUOftHvBO4W3jzbEiowgwBODNtPm3AeJC7fVxQc6Ku50zV6ImNHsu3BLMRiW26FUE77"
    "Eq0UCuh+oV6S9wTZfCv2p5CKEMILwWZbXL06duq0bW3lkBCYD87FfnC3mvsvK0xbfOs6CnOaza1UGJOp"
    "Sb7iJvscLRpR6fvCU26yCUHyvrRpQjg3mi+esdpu/YXV8zPGv2JlGjhVtQYuMOk2VrqGFE8DlEiKZzh9"
    "oRSRzXPzfjZHk63kFs7ObXydc/suLFC/cdB+FqLAxBaNa2H6QOLSTUvlLBhHlpXaivXJs5K/IWi21c8z"
    "EY52LZHNU4J/OlslMrvVy3fiyLsh6GoNqQLYln2NWuPuyqDrc7Utuy21uHpl0BdxVRhLVx7/HN4bM0gQ"
    "W6osVo2vvIPVNH7jyikYjeAXQrt2qE+NnCM5Uy6qUtGA/YdJX2Vi14nwuYoaOLevoCYYEym384W46PR2"
    "WvsT6qRMZxGs+ZXDlQpjWMOaSJl/IwxYKfNXBIW5ssK/XWGkT7mUPef1/5zpIiP8ijnfwrqelPxno7aI"
    "OKTvwC1s5tmIWWCwlBZG1JBbeDF1GCznHCRyWMq0c7EDckahYg+kDDVuOErrHZ7ip8vgPJlwaO3gyNr+"
    "8JnCXigwVD+xZFUS7YQAfr7BhOLohQNrQHWxLY/A8WPtlL3Shh+O27cDJnkRLL0NoY/7H4dqef1HsMCb"
    "EKx1XutP/BxTTmKoltcf2QFN2IcCAwUHxD6mcKoSvpVNgofqD2dy3585rtHp6ZNex+wGVqdvTWYdy5xO"
    "ZjO3rxv69B/hxTpveK0Ofw8QZBQg8yEl8PKdJO9sDv5zWTZUhZsMPt8eAtgi9r5h62M4wt1xuzrqmLbX"
    "6/TsrtVxLWTMbHPiwKlZAbv1ytfv6BpC2Yt8GHjrgeIwIDgqZFVISCwFIcHtiU5ohSS08iVLo38BAAD/"
    "/wMAUEsDBBQABgAIAAAAIQCPmK7gpQsAAOMkAAAUAAAAeGwvc2hhcmVkU3RyaW5ncy54bWykWktvI8cR"
    "vhPgf2jIWEALi6RI7fqFlYyhOFTGpDgTchhBuhhNsiWNSM4w81gueQhsxWvn7DXi2A6wCXLNzUgOsW8S"
    "r/kR+wvyE/L1DKmVuppSHoAhS1PV3dXVVfXVY599/GI8Ys9FGHmBv7tRLm5vMOH3g4Hnn+1udN164YMN"
    "FsXcH/BR4IvdjZmINj7ey+eeRVHMsNaPdjfO43jyUakU9c/FmEfFYCJ8UE6DcMxj/BmelaJJKPggOhci"
    "Ho9Kle3t90pj7vkbrB8kfry78eR9nJv43q8TsZ99qXy4vbH3LPL2nsV7HddwjzvuccNitXZ38dNxyz4y"
    "2ZvPvmGfmDWzxQ7N/ZNnpXjvWUnyZ2v+9fr1n9nip6tXi8/dY7bLJjyMRMg2oyQMpoINuC8eM8b+9fqb"
    "v7GWZVYts4Ptd5mUOllccsaTOEgZXv+JVbvtzokrj5VbjUUczoacNRwLDG/+8BLf/NWm7HnEOt3DY1We"
    "Wpgsfp75PJ8rscaUhzFXOdonXUh6ZDdbpko6MViFOd1WwyX7grSzhuQ0u52CTXer2VXLbVjkkKplt6ED"
    "XEglGR1X97njtvEwrMTqRreJB6ELHbtjGTWjZRnqlk7DVT+ZdRMXPGrZi++vv1SJrapRMpt4arct7QBn"
    "aY6ru0Y+FwwGRTbnrKxuUXcP87m+GPliDZmNvSjK53xPpFx0/aN8LhomsejP/WDxw/VXKkfFuX1ChZIf"
    "OKHiZPLDOFmUjImJVJyHJNi5I8GOKsGO84AEOw9IsPOgBJV3y/kcH/YvBDl8LcX1JmzMB2L5POrClHzv"
    "00gOKCyfW1xev8TreORwcDykO7ttVqVh5XPzXpEFp8KPisSN7/AMhJbphmetNEuHgqKiWRTPyIVtGFoc"
    "8njGqm3iOfncJJyLcMp9j1hIvZbPnRbZdDZOornwiRqcenZD8EwQpxeXPuI+4ao2GxlbbxQMPVWAjtvM"
    "qFKM/vl0Ri8Av+9AzCDyOCItFdPYd0HXWkn9AD5wWjwL+IjxOCYv4MiVk8Qf0lONUnZuum8Jp9PFjoPF"
    "w3gNESdLolx//Z16bVE/eLSMO/mcOIUG8YG8XOctTxwmcOPzIKCXcO1bbEnoB0BgslfdDW/Oq7sM5kAe"
    "ym7HZzc8gZQJXIBuRSwJ1B9FE94HgAOJAYTPxcYesOv3f2fMPbI/MVbASuK09Uv1ipbuk+bbr8hKm4R8"
    "IKV64r2yto+PjCa1x5v0wGrVjo+sWtcAiLIT48iutazG1asjcoea+iVFX7qzBLjSenhzO+qSQ6ulA8IV"
    "bi1hTA9f3dqJZSw+V7c84dNg4HtD9fs7+dwwiObJaKgBihsg0qDQCiQJ6RZ6aKBjtYyQbgGrBnRXywhp"
    "GWXUa63C8LrQuz5Ykki5CnY8GRHXUaIYCRVZJEUgJZTlE+dzY89PaBS6FSfWRwd6CaBTvIwEKrFutUCd"
    "zobB4ov+XCy+UBkW37eL+ycGAuNmJIaPVfK7pYI2q2sZTQv5Y6ex+NyowUvY04LdsavIcvd/kabXdt1s"
    "deByzGJ2tW23SHTIlhKDRbqq+UbtRqaNVBPqF5ln3+s3jnf9MkZCvumHLPMIjRLser1UM+tEt64m7qZJ"
    "ksqJ5It+0nIiSVI5l4kXsdBbOEnsLIMzIvABzaVl0pYBlwZqxHgSrDEMAh9a9Nir3UYZ5DxFPVtLxG8Z"
    "FbH/CwS6Y3n/7zY6q30YYf5nEVJ8ylZr/cW1EDwOmAQb6W+yhHXsLOrjFyZ92DLx26pqYpvbbz57VXkS"
    "EZ9e1ZPEWbS1pTubqIz7psTIkix+s18nKI9lRjeYe3xxyfpzHiVss3L17fusl/SHIr76cUrEME9Pia3L"
    "AEgM3XGIAGJUzOe2I/V7C+WXlrCfLigXnqxboiNli54WPli3SEfKFn1YKFfWrdLSlgLuFMrvrRVRS1wu"
    "fL9QWauOspaYLayUCzCRNXrUE//5FyZr3Fn/XF0GCsonWeaSgHR/plhxSHq3Qz8ZTZK/aZM6YwlO5Q8Z"
    "2gxmnVVtq9OgACRp5Os7RBd8PuXoY4XilBRbdi+a81BjnbQukqYp+wcqsxHHsrM05sQ/nFXBoStl8Hoo"
    "9MPZlI80j5c+gZ4qz9Muk+6op0AQPaFm1euoZzQ+2xTTGZMVINucBIMUWcWY3tCusYbdOTEPic+H/D/Z"
    "ID1mcYlCU1x/xTZPedijeqwb7Sp55sX3YTAQEE+/xknPf2jn7JY8EiMPj7E59gbkiodWrdA2WgckV8Fa"
    "VHjDcD4b4Bj94oOrH0POhqMEAVXPIeV8cJvsMr175UyvIoZRXFx/0Izdw1DjIzH02Eg+vFYTS4ZJqlkt"
    "RyqDz8Ng8TNKFraJEED0uVNIe5mqvaja1C3NtLm4TIbAJh0DUeY6phl7QMqmeWC2akaG5o7VYUedE3Sh"
    "GxaS44bd7B62kBZndZ16E7NpHpotWu022levyFf3mECjjePUPbtHxoEFvEYzHFK1SQH75pvP8J+UzG4f"
    "s+wvdY87TXGVaLWcLhFO9snZVFwsfuh7aHGra44mXoTwBR6ZOcxXXXb8MvLwQnMOevHFKHqhiW9MDGZx"
    "MOVw+vD6O23T7m53Xt2jbrcPuyhd1O9Bb+TB2XxSFRto6WMkMUvPSiVeMi6lN18gut8rqUahd6YDKh2l"
    "i/ppOT0gnDJABPC7aTTki39I30FzEU2uqWw190kTzvD5yEuvQq755q9/ZJ0To01mBa0S0dUkiWhXCTDH"
    "BoHcfMZicYHIP+55Pu9fECmWRpfNYzAI0ZudHJG4ptGEx50tLpFGCtLLkDuoKsFMBYEzGCVj+pQ1aXPI"
    "U4Ke6EsLZUE4O4OMSGnG6fwldVtU7/yMX1y/TOcz8iJyM+j4tn2qx570vCAcelsMTdothnx9i6FTsQXH"
    "a24xhxSQWSsHOfuh0TYWv90/QfmsMYNlBbvGGuApHG4U4cb0rofpvMkDhqBbnO2DIdZI8DP0FwdB6qKD"
    "gCGLAVj7iXqEhPYt5nbwQ9Z7W2j1yZ+o1lhb1yzcWz7qzaglDSzqrsg0RD/2ngtdF3QTgxD2LisXn159"
    "i57SY4QtfDLwCfMFAgeo7cFgvr0TBkoWRXtXtlM7sp2KBqeu8aqKiE5xejC7+pZtoq2M47eLT57IPzEo"
    "ekzksHElP5JX0utFak93BvYrb2/jKNl9Tu0ua0/D3FPCsg1O2zby4vz5GftNubxdKJefqpvX0vHCeoHk"
    "Q6prbMf5VF78Rqj0w0ow31v8HMGrd4HxExgMC3phQC1uvW1kNT45NK37f/d1alqpBnx+kZ5wO55dcDSw"
    "EMqQMWHAOiCGqtqBeoq7bI8x+vi25tPtB1kOU6QE7BEmH+lcYvEF25RnLn4Ii3iDnUfEIuqhEMw9x8RY"
    "mgRBPzToVRlhWHIeCmOTtT3OQx19/V2EUfJgOQBm80HQA2Yy5JsenTJkStDZvxN4fhyVnCCKRCTH9Xp7"
    "vzHCtLlTkhZIhIR33fE4Xchgsh+QVkwEYFBLlOaasdEhRnjS8CZx9OlzPoK7pqr4VMeLOCfxC127KYwh"
    "ojZ4G1owQcZ4WRuHltE6G9cBLMkjyXayer801dG1RBt4qLnohbIBMgkkwKRV8piF86QPVMxknY3VDd98"
    "+TWCq1lFjI1EP/AHrH/O/T7SopjU5SuJs+GhTuSaRuSHJEvLQZoKQC65G+TSBQtDO30E7hEAg15QA6Zq"
    "kS2irHmQqiUWWm1gk5Jrp+OoQN3tUNs3B9KqjPvoQbGzcAZ4zEYgPHWrYNK/QNTSFOLyHRwRFnbeA1Tj"
    "/0+2txi63+q+Vd1YEyCvu3aKqkMAM0wCVrAsvLWalgMN6TVnfEzM0Fk7K0VeoVX3ckHfk7aIH0MaLOR1"
    "5Rxj3aF1Ofi4b9BL0xmyJN18pX+d3deDZMTiMEh6dMjijJKoJJ+beIHmVdpXP8qisc9Zip/IE9nZ4O3b"
    "s94MmgCm9gIPvRFVlKM015P52ltKCf9Iae/fAAAA//8DAFBLAwQUAAYACAAAACEARMUWsSUTAAD9dgAA"
    "EAAAAHhsL2NhbGNDaGFpbi54bWx0nVuP3DiSRt8XmP9g1Pu0u66zM2j3IMspZUpZUrVTt5HeDHdtdwO+"
    "NGxjsPvvVxgxyNB3lI91KoLiJUgGySDzp3/+76ePr/798vXbH18+v7m6/uHHq1cvnz98+fWPz7+9uera"
    "/K//ffXq2/f3n399//HL55c3V//38u3qnz//5b9++vD+44e3v7//4/OrOYXP395c/f79+5//eP3624ff"
    "Xz69//bDlz9fPs//+Z8vXz+9/z7/+fW319/+/Pry/tdvv7+8fP/08fXNjz8+vP40J3D1808fXn19c3W6"
    "ubl69cebq/urVx/nrFy9DrxKfP5M4sfITfIAkoOcbq7DV1L6Rl6t0s8heYxkLXm6mavtPzlPaRpZSx4h"
    "eQDJQU7Xf9f0I5E8Q/J4QfJ0PbfsKs9HkBzkEMn6u9X13yS1J5ATyBHkAFJHoiXVL1bXD5KHE8hTJJqa"
    "6h4vSFbXs32u6u0JpI5E7ASSR5Ac5HAxtTtYhREtnUoer7clT9e3SNOIWDIkc5BDJGot1tOtvzxdKzmB"
    "HEEOIPlGOujv19v9/RC55SoHOV7Qra51BDiBPEWiNam6OXQPF3TrC/yEAcPA+tsHlcsVHLcVKx09Tgqe"
    "DIgtqlyFkUPBSUEOCe29TwbWH6+38VHVDwpyBZWOAicMCwaktVUuV3DYVnzSPnxC99/u00eVOyjIFTR3"
    "i0nexUn4DNJFsi5gc7tYjNMF6SJR3cU6km53a2QteY7cOmtzu1iF/66S9nZpxyTTgJxB+kg0t0tL+S8q"
    "6W+NrHX726XOk24L0oCcQbpINP1l2PYlVdLdGtG6peQyPPvUlJxvlfS3y0DrW1NJA5k2Es0VU1MrbW5h"
    "tyB9JNKaN7BbkP4G9gnSgjQg5410YKs3ShqQM0gfiZZOLf98g75wo9bbRiJtAcnmRu35DNLcqF2dQfrg"
    "Yrt+AdKAnEG6SLQeYCHB3XZfDC6zs3aQM0gfiXwxuMxulADpI5FeHJxfnzdYBWTOIF0kmj5sILi9rs9G"
    "IjYAySY4qr6kGA+jjOZE7acLLqpvBYyZUUbzxtQwHgbn1OdWZfooo22K0S84nim1PriKviZ1BGsh04Cc"
    "N9JRG26DG+jrCnYOmT4SKR0GQgWdDoONgv7CtA1FTNoKOgOSSYxnMGXMzgq67bm5wVCmYPzbUru30T/q"
    "QFqQBqQGqUBKkAIkB8lApoelFVOed48Rret2hOgA0oN0IC1IA1KDVCAlSAGSg2Qg48NifqkKBpAOZIpk"
    "XUtz9Vly63+0SKIBqUEqkBKkAMlBMpDxYelCzlpBWpAGpAapQEqQAiQHyUCmh6U7r6zVkFqrig5Q7kE6"
    "kBakAalBKpASpADJQTKQ6WEZnlZVYEirQEUHKPcgHUgL0oDUIBVICVKA5CAZyPSwjLyrKjCkVaCiA5R7"
    "kA6kBWlAapAKpAQpQHKQDGR8WJwS33mVTFEGI5SJ6gilSTT4TA1SgZQgBUgOkoGMD4s/5QuppIVMA1KD"
    "VCAlSAGSg2Qg08Piza1s05DapooOUO5BOpAWpAGpQSqQEqQAyUEykPFB3aABpAOZIoG1WnJqrfqZBonW"
    "IBVICVKA5CAZyHivnlQH0oI0IDVIBVKCFCA5SAaye7xX32cCGUEGkB6kA2lBGpAapAIpQQqQHCQDmQuv"
    "HtAEMoIMID1IB9KCNCA1SAVSghQgOUi2UVJ6UvcXPKnIbVN1AOlA+guptZBsQGqQCqQEKUBykAxkvMdM"
    "CjJFomPThX+0SKIBqUEqkBKkAMlBMpDxHjMpSAvSgNQgFUgJUoDkIBnIdM+ZNCKZSSE6gPQgHUgL0oDU"
    "IBVICVKA5CDZVoF1jhvvMbmCdCBTJDKHQrIBqUEqkBKkAMlBMpDxDnMoSAvSgNQgFUgJUoDkIBnI7vEO"
    "cyjICDKA9CAdSAvSgNQgFUgJUoDkINlGSXUKnavDkHTSyOM0AtKDdCAtSANSg1QgJUgBkoNkIOOdzqQd"
    "yBSJTiAX/tEiiQakBqlASpACJAfJQMY7bCWA9CAdyBQJKmJ7D6NFEg1IDVKBlCAFSA6SgYx3uqHQgbQg"
    "DUgNUoGUIAVIDpKBTHfq4syddHsLYIToANKDdCAtSANSg1QgJUgBkoNkIOOdOkADSAcyRQJrteRkPkUS"
    "DUgNUoGUIAVIDpKBjHfqQ3UgLUgDUoNUICVIAZKDZCBTiCDxOygR6ZQCZwjKPYiFqKQPtJCxUJckU0Om"
    "AilBCpAcJAOZQiCMr4KIpAogOoD0IBZp46oAMhax46oAMhVICVKA5CAZyBiia9zeH8gUiXbPC/+wQJ+U"
    "qAX6+ELqdF7hwyVIAZKDZCBjCANKXx9AOpApEhT7whyKJBqQGqQCKUEKkBwkAxlDdJJvW51VLe7JN5XK"
    "1EinAilBCpAcJAOZQlTUqktemEMhOoD0IBZ25bsktuOhVYNUICVIAZKDZCBjCOXy1optBchMkcBaL8yh"
    "SKIBqUEqkBKkAMlBMpAxRJ55a9VZ1aLTvLWqTI10KpASpADJQTKQKUSzrax1ext9hOgAYsFxvgp05m2h"
    "ZWF3fmxVrQpaJUgBkoNkIFMI0/NVEJHMoRAdQHqQDqQFaUBqkAqkBClAcpAMZAzxhK7ZQCx2cR5qVtdc"
    "LF7RWTR0a5AKpAQpQHKQDGQMEYxu8AHpQKZIdPC58I8WSVh4pbNiyFQgJUgBkoNkIGMIx/RNqP5JC5kG"
    "pAapQEqQAiQHyUCmEAa66nkXIhQgOoD0IB2IRaR6a9UNixpaFUgJUoDkIBnIGKJcvbVi3wAyUySw1u3o"
    "hhZJWMCtt1b9cAWtEqQAyUEykDEE7nprVR+mhYwFAPs8q1YFrRKkAMlBMpDdY7gMmD4/gYwgA0gP0oG0"
    "IA1IDWJXHVMOS8gUIHbJMWllkNk9hvBkV3iQEWQAsTBn1/KQaUEakBqkAilBCpAcJNsol+7bdyHg2tfG"
    "djDfXHEXovyQRANSg1QgJUgBYhcxXQtDZgwx3r5hcOodZXTcufCPFok2IHZDM33YboM6M4ZWAZKDZCBj"
    "CDr3hcQsCZkGpAax26M+z5pyAa0cJAOZQsh7Snm2qQuzJEQHkB6kA2lBLDbfDb2Qsautvgp0bi2gZRdW"
    "vW2q1hji9JPMAGLx/r5Lbk+Gc/VdmCWRaANSg1QgJUgBkoNkIGO4PuCtFbMkZBqQGqQCKUEKELuh65sK"
    "sYThIsPKWrfX7yNEB5AepANpQRqQGqQCKUEKELsg7KtAtzamcONiVQUXIvsgOoDYRRBvBdhWgJZdBPEd"
    "VrUqaJUgBYjdavZVoClP4WrIqgoubCtAdACxmya+CrCtAK0GxK47p3Ts4rUfszTlAunYBWtfBYi/QICC"
    "gsmATqbbvFX9RkGtoFJQKigUhAvcrmQqMcIZUtAqCNd7nC2qRLgP7ppBJQoFuYIM0RUMObgQcaCCg4Ie"
    "VyAUtAoaBbWCcHXdFVolCgXh6rprHpUYeZvhP08yuL6jEtN2ON98QSVoyo6P6jcKwj15181UooRvpCBc"
    "mXclVYlR3YRwHyuptCrRYK2vIFzMd20CF0ZBuIHvcooQiLBXfJPuw4P8AvIMYrvZKR3bzU7kCVonENvx"
    "Tlq2453IEVoHENsDT1pvIfMIYve6k9a7sC+diO1dJ7KHTBfJ2j5tEzrpdmHT139xGa8TsW3peTZfP+MD"
    "3T2IbWCrru0Fu5yEvVf3XRDbC3a5hYzt6rr6gYzt2PqvL0NkIna33H9LZWzX1X9LZWwXmDWwjE8+fSW2"
    "velzZTLrttiHbc8k2YdNRa+7jBP+i0psc9LXjMnoF1XXdhl9+kt/T8R2/nyNqUwfNsV8OsuyxKejxDby"
    "tJ73SK0PG1g+/cVZ9+kr2UPLdqt8OovHm4jtVqHvhH0kXw+qa/tatJzFqfRfUfIu3NL3JVIZu8nPGlPJ"
    "Ptzk9yXVUcIe1/IlUpkuprO2pSFsRyRdu2mfyC+QeQaxDZOkZRsmidjzWYnY81mJlEjZNlWSjD2xlYg9"
    "sZWIbbMk8hYpP4L0YaPD1TaIbY8kmS7KSD+Fru1z+PTRT8NuhGtNkD4sy3066JVRRuYO6NouBawRkrb1"
    "gH4RluU+P+jXkNmD2Mra1632zXdh/ZtkbK2Nng7JPYitylki9PSw6PRlVBlbhvr8m4zaBnp6WNH59OEP"
    "RBltU5XcI7Xugm6/rKVclhWERV6SCIs8l1NVCSs2Z8MqER5TcGlgilewVxCebEDbYXrH7K5gryC80ICk"
    "MY9jGlewx2StPf4dhgAFewVhzaDZ29nLiOlFyd0hHB04ZKcJDu0p9ZbIDlyc4g5SE4idwSS1fzFtOyqZ"
    "+8XK17XjmqTcQdkOcFTXDm2S7jvo/gLyDGIHOykde74yETvYcY95Ih076tF8so3sCCilxubYaI3ncITi"
    "GshOZxyydzIdsvMZh+ztzJUtLV3QITvq0SLt7HzHyT7F3MniCLk+g9hBlH4npRnfKb3wFZYnFXqdnwxf"
    "37NezQdaVfUyjKyqGshcnFXtQ8qcnFXtq1QfPJgkZN5bIna4xYrT1MxjU0nzwLxlq675W0lmj7ztnoND"
    "tKqyZXhcVQaQPVy6qgxIlTF52bw0l82p22GZltU8QF9/+qEzimFPsCYtO3rT9FkSO1zzNYei2TGWK4E9"
    "s7qqk2WecMhOslZFNynthaptB2zevFSGObO3Y725qFYWXFpfaOTdDrFc3k/BJV3Vw+L5ruoB6DkqaqFV"
    "1HxcX2iVYc6ekDF7jnbVGJaQGKgd4Km12PGbryYUzU5MVnWyeJerOgGyYxQM3VlwdZ36nugt0SPRDsiO"
    "olLqI2T+BWJnUEmrg4w9PZZk7JxKi/gOuvYkmUr+AslnkBrkCeQEYudaKbf2Wm8itqDwfcnacW1DbI2N"
    "xngOawDXsHbY5FBJqSORPd+7su9lGeKQnW5xnFVROxnz/U5lmFl7gNhXEDJh7xaze6moLZt8BS0rF9+7"
    "QMKTw74qIBMOwlARKnhWEFZbqhjO0dI3n/DJ8GIy5nMVDO8hp6Qyldgj7XA45oocHkd2JLyf7E0L/mM4"
    "QvPVC5kDyLMRnekhWW9LhuWnMzbVDOeDcJtULlbEOiso11EVw9FgykJY33rboz8JEo7jfAVC5kAn1IhW"
    "ICSftiXD4aOrQNUMZ42oQLid2x84qlwsqDjrKhfW8q5CnulPgZzoTtIzpWMKEp7ZTl/v4T4qCEeVqCiV"
    "qwysKyC89O0Hn23BTBMMWxNJ02L7Zlds9XMVO85wu/AcuftsOGH1GcGkl3wOW67tODNOuMC0e4xIdgcg"
    "aq9QuxPXsHjz7+8uRpN2n+y12qTV82kVpGNnae491HDi5b6FV71tRX4fT15tEZvI7hkjLsd3blTsuAux"
    "S5sO8Vlx7r6EmIvUdhar41rTAsZcn0dCNou77cQQX+62D0HajZ3ypYVSe0xb7z9ZHIgYBUR7KtNj57YY"
    "d8XCrOCLpxvU0zLQp5zvHo1INlUwTEquzLhJMgd8bydmD7o7A8bz9nZTOBlnGFK8JS779olYoN48Jqz2"
    "5QodTSwkKZnHxjqCzrMFoDpDs984WY0kOsyHwApnVjiMt4fDU50MvIaLZzrnEBOzPWkviPYki6rr/gDx"
    "9fz4QyH4cZw6Elmt4v1xO3tNH9xx/T2PCrq4Drv+rmYQfD0tOt6OjWh3k2CeXjV7jIFBxA+ci1IihzAo"
    "eFNcbE5NcReimPyUjxIPvE+2pOaet1bQAsSnvq3x7OjdpYLH1+3w1U9Ay8ojES6wJzzPOFvmMtjofbcB"
    "oj0JXvO2qFXXCLDNetluSG0Q3LIEgm/pWmmxeG2k4A6m/jobqg4hITbKGbPFZqem7XBq2IJYuLH7mLk2"
    "KSG7dpWEBjziuXvWYSfEESalCW9/zU/zXWgmiPYk+AmDDsTuzDl70gOiDjETdnXJz8N6RtstDZJEQgdw"
    "Oiox4W7AHMq+CMFKIWq7rX7SMt3YwZY53hmpgnA+iTHXT9J6Or4LUaduyOB+zITHaOanZ7bD9AeI9nxZ"
    "B75hT4PjztGEmPe5hrcv7w8Qtdf+/UCvnpXdtUgV2CDqxO6IO7vfuox2IVsQtUCglC129HlhoLNWi3zZ"
    "wiQNP9Y5nG/PhyrwRFTo1s6G+OMHIWjQjfSLLTjTVNAqmLZeYLvgxUK0BxmwSBkssbi+spsf3rfXXvzE"
    "FUE64VzP9hPu68+Ls2VPEh0eohZm6VxcPOq/e4xI/AyI9vytAaw4zbrdGIZ4Grt46kcMXSFOuCo8++IX"
    "BjqIWkhgKrcN0N5vWHqP9xJAaIP4iagOpAUJm1FuFl9WFW4WN7BuAxuenODSSd3SlXuy8xpUd1rtOpcb"
    "UPCY0jzaXjAsiPZ8Kwq+st23caudpbO4whhYl/oQTnfcKBN/7Sv9UOeS1yTD9bKtc92IETqea/T4C1aW"
    "crfx6106B9pq1rVoSNm7z9uls6W+9wyWsrgxFNMyboZZ8KibzPEcSY9Hvno8o9njceoeP/9gLp+vNvWo"
    "n/BbpzbKOQPAb6naBWasMSx6ca4Z3RtTH7HGl+2Az/USRnikCJg4dqfAE/noasB+HX+49uf/BwAA//8D"
    "AFBLAwQUAAYACAAAACEAJcRnYVMBAABlAgAAEQAIAWRvY1Byb3BzL2NvcmUueG1sIKIEASigAAEAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAjJLNTsMwEITvSLxD5HviuIECVppKgHoplZAIAnGz7G0bEf/INqTlyLvxXjhpG1LB"
    "gaN3xp9nVs6nG1lH72BdpdUEkSRFESiuRaVWE/RYzuJLFDnPlGC1VjBBW3BoWpye5NxQri3cW23A+gpc"
    "FEjKUW4maO29oRg7vgbJXBIcKohLbSXz4WhX2DD+ylaAR2k6xhI8E8wz3AJj0xPRHil4jzRvtu4AgmOo"
    "QYLyDpOE4B+vByvdnxc6ZeCUld+a0Gkfd8gWfCf27o2remPTNEmTdTFCfoKfF3cPXdW4Uu2uOKAiF5xy"
    "C8xrW7T9zXZT53gwbBdYM+cXYdfLCsT1tphrJ1k011+f/CPHv/XA7CrswCCiEIruKhyUp+zmtpyhYpSO"
    "xnGaxeSsJCklF5SQl/b5o/ttyN1A7kP8h3heplc0zWg2GhAPgKLLffwxim8AAAD//wMAUEsDBBQABgAI"
    "AAAAIQA5rZb7ugEAAJUDAAAQAAgBZG9jUHJvcHMvYXBwLnhtbCCiBAEooAABAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AJyT32/aMBDH3yftf7D8XpJ2E5qQ4yqCrCDxIyKh0vbmORewGuzINgj61++SqDS07GVvd/c9f/Xxnc0e"
    "T/uKHME6ZXRE7wchJaClKZTeRnST/7z7QYnzQheiMhoiegZHH/nXLyy1pgbrFTiCFtpFdOd9PQoCJ3ew"
    "F26AskalNHYvPKZ2G5iyVBImRh72oH3wEIbDAE4edAHFXX0xpJ3j6Oj/17QwsuFzz/m5RmDO4rqulBQe"
    "b8kXSlrjTOlJcpJQsaAvMqTLQB6s8mcesqCfskyKCsZozEtROWDBe4FNQTRDS4WyjrOjHx1BemOJU684"
    "tgdK/ggHDU5Ej8IqoT1iNW1d0sZV7bzlsX05uFd0R7WrtGG/sR+r73zYNmBw3dgYdBQoXPPlylfgVmUq"
    "rL+BO+zjtgwdbIeTJ/GCPCXLZB3P+5QX3nQe/0rW2U1tPlsmm/S2lk1XOclni9ny6ebZ8Wqzzsnv1TL5"
    "h3eCVJP40+TaTeAMPtx6rvSL29S5mQgPbyu9LrJsJywU+AouK78U2BS3aavGZLwTegvFW89noXmAz90v"
    "4/fDQfgtxLfVq7Hg/T/xvwAAAP//AwBQSwECLQAUAAYACAAAACEAxcTpJYIBAAC4BwAAEwAAAAAAAAAA"
    "AAAAAAAAAAAAW0NvbnRlbnRfVHlwZXNdLnhtbFBLAQItABQABgAIAAAAIQC1VTAj9AAAAEwCAAALAAAA"
    "AAAAAAAAAAAAALsDAABfcmVscy8ucmVsc1BLAQItABQABgAIAAAAIQA91AX1jAMAAP8IAAAPAAAAAAAA"
    "AAAAAAAAAOAGAAB4bC93b3JrYm9vay54bWxQSwECLQAUAAYACAAAACEAZwxbpSkBAAABBgAAGgAAAAAA"
    "AAAAAAAAAACZCgAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHNQSwECLQAUAAYACAAAACEAWM/U0xEU"
    "AACydgAAGAAAAAAAAAAAAAAAAAACDQAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAi0AFAAGAAgA"
    "AAAhAB7TQ4jtDwAA+WgAABgAAAAAAAAAAAAAAAAASSEAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbFBL"
    "AQItABQABgAIAAAAIQA8vx1rfSQAAGQxAQAYAAAAAAAAAAAAAAAAAGwxAAB4bC93b3Jrc2hlZXRzL3No"
    "ZWV0My54bWxQSwECLQAUAAYACAAAACEANvAzPSQSAACSeQAAGAAAAAAAAAAAAAAAAAAfVgAAeGwvd29y"
    "a3NoZWV0cy9zaGVldDQueG1sUEsBAi0AFAAGAAgAAAAhACuLE21MCgAAPj4AABgAAAAAAAAAAAAAAAAA"
    "eWgAAHhsL3dvcmtzaGVldHMvc2hlZXQ1LnhtbFBLAQItABQABgAIAAAAIQCD5o3x4QUAAJ8hAAAYAAAA"
    "AAAAAAAAAAAAAPtyAAB4bC93b3Jrc2hlZXRzL3NoZWV0Ni54bWxQSwECLQAUAAYACAAAACEA6aYluGYG"
    "AABTGwAAEwAAAAAAAAAAAAAAAAASeQAAeGwvdGhlbWUvdGhlbWUxLnhtbFBLAQItABQABgAIAAAAIQCk"
    "jsAG8wYAAKpJAAANAAAAAAAAAAAAAAAAAKl/AAB4bC9zdHlsZXMueG1sUEsBAi0AFAAGAAgAAAAhAI+Y"
    "ruClCwAA4yQAABQAAAAAAAAAAAAAAAAAx4YAAHhsL3NoYXJlZFN0cmluZ3MueG1sUEsBAi0AFAAGAAgA"
    "AAAhAETFFrElEwAA/XYAABAAAAAAAAAAAAAAAAAAnpIAAHhsL2NhbGNDaGFpbi54bWxQSwECLQAUAAYA"
    "CAAAACEAJcRnYVMBAABlAgAAEQAAAAAAAAAAAAAAAADxpQAAZG9jUHJvcHMvY29yZS54bWxQSwECLQAU"
    "AAYACAAAACEAOa2W+7oBAACVAwAAEAAAAAAAAAAAAAAAAAB7qAAAZG9jUHJvcHMvYXBwLnhtbFBLBQYA"
    "AAAAEAAQABwEAABrqwAAAAA="
    )


@app.route("/mecz/<int:match_id>/export/xlsx")
@login_required
def export_match_xlsx(match_id):
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
    m = cur.fetchone()
    if not m: return redirect(url_for("historia"))

    gtk_name = (m.get("team_name_a","") or m.get("nazwa_gtk","") or "").strip() or get_setting("gtk_name") or "GTK"
    name_opp = (m.get("team_name_b","") or m["przeciwnik"])
    dt = m['data_meczu'].strftime('%d.%m.%Y') if m['data_meczu'] else ""

    cur.execute("SELECT * FROM match_stats WHERE match_id=%s ORDER BY kwarta", (match_id,))
    all_stats = list(cur.fetchall())
    cur.execute("SELECT * FROM player_stats WHERE match_id=%s", (match_id,))
    all_players = list(cur.fetchall())
    cur.execute("SELECT * FROM timing_stats WHERE match_id=%s", (match_id,))
    all_timing = list(cur.fetchall())
    cur.close()

    # Wczytaj szablon (wbudowany w kod)
    import base64 as _b64
    tmpl = openpyxl.load_workbook(io.BytesIO(_b64.b64decode(_get_szablon_b64())))

    def qd(druzyna, kwarta):
        r = next((dict(x) for x in all_stats if x["druzyna"]==druzyna and x["kwarta"]==kwarta), {})
        return r

    def td(druzyna, bucket):
        r = next((dict(x) for x in all_timing if x["druzyna"]==druzyna and x["bucket"]==bucket), {})
        return r

    CTR = Alignment(horizontal="center", vertical="center")
    YEL = PatternFill("solid", fgColor="FFF9C4")

    def fill_cell(ws, row, col, val):
        c = ws.cell(row, col, val)
        c.alignment = CTR

    # ── TEAM GENERAL ─────────────────────────────────────────────────────────
    # Kolumny wg szablonu (wiersz 4):
    # A=drużyna/kwarta, B=FTA, C=FTM, D=FT miss, E=FT%,
    # F=2PM, G=2P miss, H=2PA, I=2P%, J=3PM, K=3P miss, L=3PA, M=3P%,
    # N=2+1, O=3+1, P=Tip made, Q=Tip miss, R=Tip sum, S=Tip%,
    # T=OREB, U=DREB, V=REB, W=AST, X=TO, Y=P, Z=FD,
    # AD=POSS, AG=PTS, AK=PPP, AM=eFG%, AN=TS%, AO=TO%, AP=FTr, AQ=ORtg
    ws1 = tmpl['TEAM GENERAL']
    ws1['A1'] = f"STATYSTYKI DRUŻYNOWE — {gtk_name} vs {name_opp} | {dt}"

    Q_NAMES = {1:"IQ", 2:"IIQ", 3:"IIIQ", 4:"IVQ"}

    for team_label, druzyna, start_row in [(gtk_name,"gtk",6),(name_opp,"opp",14)]:
        ws1.cell(start_row-1, 1, f"  ▶  {team_label}")
        for qi, qn in enumerate([1,2,3,4]):
            r = start_row + qi
            q = qd(druzyna, qn)
            ws1.cell(r,1, Q_NAMES[qn])
            fill_cell(ws1, r, 2, q.get("fta",0))    # FTA
            fill_cell(ws1, r, 3, q.get("ftm",0))    # FTM
            fill_cell(ws1, r, 6, q.get("p2m",0))    # 2PM
            fill_cell(ws1, r, 8, q.get("p2a",0))    # 2PA
            fill_cell(ws1, r, 7, q.get("p2a",0)-q.get("p2m",0))  # 2P miss
            fill_cell(ws1, r, 10, q.get("p3m",0))   # 3PM
            fill_cell(ws1, r, 12, q.get("p3a",0))   # 3PA
            fill_cell(ws1, r, 11, q.get("p3a",0)-q.get("p3m",0)) # 3P miss
            fill_cell(ws1, r, 24, q.get("br",0))    # TO/BR
            fill_cell(ws1, r, 26, q.get("fd",0))    # FD
            fill_cell(ws1, r, 30, q.get("poss",0))  # POSS
            fill_cell(ws1, r, 33, q.get("pts",0))   # PTS
        # OT row
        ws1.cell(start_row+4, 1, "OT")

    # ── PLAYERS ──────────────────────────────────────────────────────────────
    # Kolumny: A=Zawodnik, B=#, C=2PM, D=2PA, E=3PM, F=3PA, G=FTM, H=FTA,
    #          I=BLK, J=OREB, K=DREB, L=AST, M=TO, N=PF, O=STL, P=FD,
    #          Q=PTS, R=MIN, ...
    ws2 = tmpl['PLAYERS']
    start = 4
    for druzyna, label in [("gtk", gtk_name), ("opp", name_opp)]:
        ws2.cell(start, 1, f"  ▶  {label}")
        players = sorted([p for p in all_players if p["druzyna"]==druzyna],
                         key=lambda x: x["pts"], reverse=True)
        for i, p in enumerate(players):
            r = start + 1 + i
            fill_cell(ws2, r, 1, "")           # Zawodnik (brak nazwy w DB)
            fill_cell(ws2, r, 2, p["nr"])      # #
            fill_cell(ws2, r, 3, p["p2m"])     # 2PM
            fill_cell(ws2, r, 4, p["p2a"])     # 2PA
            fill_cell(ws2, r, 5, p["p3m"])     # 3PM
            fill_cell(ws2, r, 6, p["p3a"])     # 3PA
            fill_cell(ws2, r, 7, p["ftm"])     # FTM
            fill_cell(ws2, r, 8, p["fta"])     # FTA
            fill_cell(ws2, r, 10, p["oreb"])   # OREB
            fill_cell(ws2, r, 11, p["dreb"])   # DREB
            fill_cell(ws2, r, 12, p["ast"])    # AST
            fill_cell(ws2, r, 13, p["br"])     # TO
            fill_cell(ws2, r, 16, p["fd"])     # FD
            fill_cell(ws2, r, 17, p["pts"])    # PTS
        start = start + len(players) + 2
        if start < 20: start = 20  # separator między drużynami
        ws2.cell(start, 1, f"  ▶  {name_opp if druzyna=='gtk' else ''}")

    # ── SHOT TIMING ────────────────────────────────────────────────────────
    # Kolumny D-Q = celne/niecelne per bucket (0s,1-4s,5-8s,9-12s,13-16s,17-20s,21-24s)
    # Col D=Cel.0s, E=Niec.0s, F=Cel.1-4s, G=Niec.1-4s, ..., P=Cel.21-24s, Q=Niec.21-24s
    ws4 = tmpl['SHOT TIMING']
    bucket_cols = {
        "0s":    (4,5),
        "1-4s":  (6,7),
        "5-8s":  (8,9),
        "9-12s": (10,11),
        "13-16s":(12,13),
        "17-20s":(14,15),
        "21-24s":(16,17),
    }
    Q_LABEL = {1:"IQ", 2:"IIQ", 3:"IIIQ", 4:"IVQ"}

    # Mapowanie wierszy: TWOJA DRUŻYNA IQ 2PT=row4, IQ 3PT=row5, IQ ALL=row6, ...
    def timing_row(druzyna_label, kwarta_label, typ):
        base = 4 if druzyna_label == gtk_name else 22
        q_idx = {"IQ":0,"IIQ":1,"IIIQ":2,"IVQ":3,"OT":4,"SUMA":5}[kwarta_label]
        typ_idx = {"2PT":0,"3PT":1,"ALL":2}[typ]
        return base + q_idx*3 + typ_idx

    for druzyna, label in [("gtk", gtk_name), ("opp", name_opp)]:
        # Wpisz nazwy drużyn i kwart
        for qi, qn in enumerate([1,2,3,4]):
            for typ in ["2PT","3PT","ALL"]:
                r = timing_row(label, Q_LABEL[qn], typ)
                ws4.cell(r, 1, label)
                ws4.cell(r, 2, Q_LABEL[qn])
                ws4.cell(r, 3, typ)

            # Dane per bucket
            t2 = td(druzyna, "0s")  # placeholder
            for b, (col_made, col_miss) in bucket_cols.items():
                t = td(druzyna, b)
                q_timing_all = [x for x in all_timing
                                if x["druzyna"]==druzyna and x["bucket"]==b]
                # 2PT row
                r2 = timing_row(label, Q_LABEL[qn], "2PT")
                r3 = timing_row(label, Q_LABEL[qn], "3PT")

            # Suma per bucket (wszystkie kwarty)
            for b, (col_made, col_miss) in bucket_cols.items():
                t = td(druzyna, b)
                r_sum2 = timing_row(label, "SUMA", "2PT")
                r_sum3 = timing_row(label, "SUMA", "3PT")
                fill_cell(ws4, r_sum2, col_made, t.get("made2",0))
                fill_cell(ws4, r_sum2, col_miss, t.get("att2",0)-t.get("made2",0))
                fill_cell(ws4, r_sum3, col_made, t.get("made3",0))
                fill_cell(ws4, r_sum3, col_miss, t.get("att3",0)-t.get("made3",0))

    buf = io.BytesIO()
    tmpl.save(buf)
    buf.seek(0)
    filename = f"raport_{gtk_name}_vs_{name_opp}_{dt}.xlsx".replace(" ","_").replace("/","")
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route("/mecz/<int:match_id>/export/pdf")
@login_required
def export_match_pdf(match_id):
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
    m = cur.fetchone()
    if not m: return redirect(url_for("historia"))

    gtk_name = (m.get("team_name_a","") or m.get("nazwa_gtk","") or "").strip() or get_setting("gtk_name") or "GTK"
    name_opp = (m.get("team_name_b","") or m["przeciwnik"])
    dt = m['data_meczu'].strftime('%d.%m.%Y') if m['data_meczu'] else ""

    cur.execute("SELECT * FROM match_stats WHERE match_id=%s ORDER BY kwarta", (match_id,))
    all_stats = list(cur.fetchall())
    cur.execute("SELECT * FROM player_stats WHERE match_id=%s", (match_id,))
    all_players = list(cur.fetchall())
    cur.execute("SELECT * FROM timing_stats WHERE match_id=%s", (match_id,))
    all_timing = list(cur.fetchall())
    nr_name_map = build_nr_name_map(cur, match_id)
    lineup_html = build_lineup_section_html(cur, match_id, nr_name_map)
    play_time_secs = calc_play_time(match_id)
    cur.close()

    def build_suma(druzyna):
        s = {"pts":0,"poss":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"br":0,"fd":0,
             "ast":0,"oreb":0,"dreb":0,"stl":0,"blk":0,"d2m":0,"d2a":0,"przerw":0}
        for row in all_stats:
            if row["druzyna"]==druzyna:
                for k in s: s[k] += row.get(k,0) or 0
        return s

    suma_gtk = build_suma("gtk"); suma_opp = build_suma("opp")
    kpi_gtk  = calc_kpi(suma_gtk); kpi_opp = calc_kpi(suma_opp)

    winner = gtk_name if m['wynik_gtk']>m['wynik_opp'] else (name_opp if m['wynik_opp']>m['wynik_gtk'] else "Remis")

    def q_row(druzyna, qn, label, druzyna_rywal):
        q     = next((dict(r) for r in all_stats if r["druzyna"]==druzyna       and r["kwarta"]==qn), {})
        q_opp = next((dict(r) for r in all_stats if r["druzyna"]==druzyna_rywal and r["kwarta"]==qn), {})
        k       = calc_kpi(q)
        k_opp   = calc_kpi(q_opp)
        bg = "#f5f8ff" if qn%2==0 else "#fff"
        # NetRtg per kwarta = OFFrtg(A,Q) - DRtg(A,Q)
        # OFFrtg(A,Q) = pts_A * 100 / poss_A  — efektywność ataku A w kwarcie Q
        # DRtg(A,Q)   = ORtg rywala B w kwarcie Q = pts_B * 100 / poss_B — efektywność obrony A przez pryzmat ataku B
        net_val = None
        try:
            off_rtg = float(str(k['ortg']).replace('%',''))
            def_rtg = float(str(k_opp['ortg']).replace('%',''))
            net_val = off_rtg - def_rtg
            net_str = f"{net_val:+.1f}"
            net_col = "#1a6b3c" if net_val > 0 else ("#8b1a1a" if net_val < 0 else "#555")
        except:
            net_str = "—"; net_col = "#888"
        net_class = "net-pos" if (net_val is not None and net_val > 0) else ("net-neg" if (net_val is not None and net_val < 0) else "")
        return f"""<tr style="background:{bg}">
            <td class="left">{label}</td>
            <td>{q.get('pts',0)}</td>
            <td>{q.get('p2m',0)}/{q.get('p2a',0)}</td>
            <td>{q.get('p3m',0)}/{q.get('p3a',0)}</td>
            <td>{q.get('ftm',0)}/{q.get('fta',0)}</td>
            <td>{q.get('br',0)}</td>
            <td>{k['efg']}</td>
            <td class="{net_class}" style="color:{net_col};font-weight:700">{net_str}</td>
        </tr>"""

    def player_rows(druzyna):
        players = sorted([p for p in all_players if p["druzyna"]==druzyna],
                         key=lambda x: x.get("pts",0) or 0, reverse=True)
        rows = ""
        for i, p in enumerate(players):
            fga  = (p.get("p2a",0) or 0) + (p.get("p3a",0) or 0)
            fta  = p.get("fta",0) or 0
            pm2  = p.get("p2m",0) or 0; pa2 = p.get("p2a",0) or 0
            pm3  = p.get("p3m",0) or 0; pa3 = p.get("p3a",0) or 0
            ftm  = p.get("ftm",0) or 0
            pts  = p.get("pts",0) or 0
            ast  = p.get("ast",0) or 0
            oreb = p.get("oreb",0) or 0
            dreb = p.get("dreb",0) or 0
            stl  = p.get("stl",0) or 0
            blk  = p.get("blk",0) or 0
            br   = p.get("br",0) or 0
            fin  = p.get("finishes",0) or 0
            efg  = f"{(pm2+1.5*pm3)/fga:.0%}" if fga else "—"
            ts   = f"{pts/(2*(fga+0.44*fta)):.0%}" if (fga+fta) else "—"
            p2pct = f"{pm2/pa2:.0%}" if pa2 else "—"
            p3pct = f"{pm3/pa3:.0%}" if pa3 else "—"
            ftpct = f"{ftm/fta:.0%}" if fta else "—"
            team_poss = max(sum(r.get("poss",0) or 0 for r in all_stats if r["druzyna"]==druzyna), 1)
            usg  = f"{(fga+0.44*fta+br)/team_poss:.0%}" if team_poss else "—"
            name = nr_name_map.get(str(p['nr']), f"#{p['nr']}")
            nr_html = f'<span style="color:#888;font-weight:400">#{p["nr"]}</span>'
            bg = "#f5f8ff" if i%2==0 else "#fff"
            _pt_secs_pdf = play_time_secs.get(int(p.get("nr") or 0), 0)
            czas_pdf = f"{int(_pt_secs_pdf)//60}:{int(_pt_secs_pdf)%60:02d}" if _pt_secs_pdf>0 else "—"
            rows += f"""<tr style="background:{bg}">
                <td class="left" style="min-width:90px">{name} {nr_html}</td>
                <td style="text-align:center">{czas_pdf}</td>
                <td style="text-align:center;font-weight:700;color:#1a2b4a">{pts}</td>
                <td style="text-align:center">{pm2}/{pa2}</td>
                <td style="text-align:center">{p2pct}</td>
                <td style="text-align:center">{pm3}/{pa3}</td>
                <td style="text-align:center">{p3pct}</td>
                <td style="text-align:center">{ftm}/{fta}</td>
                <td style="text-align:center">{ftpct}</td>
                <td style="text-align:center">{oreb}</td>
                <td style="text-align:center">{dreb}</td>
                <td style="text-align:center">{ast}</td>
                <td style="text-align:center">{br}</td>
                <td style="text-align:center">{stl}</td>
                <td style="text-align:center">{blk}</td>
                <td style="text-align:center;font-weight:600">{efg}</td>
                <td style="text-align:center">{ts}</td>
                <td style="text-align:center">{usg}</td>
                <td style="text-align:center">{fin}</td>
            </tr>"""
        return rows

    def timing_rows():
        rows = ""
        for i, b in enumerate(BUCKETS):
            # Sumuj po wszystkich kwartach (nie tylko pierwsza pasująca)
            grs = [r for r in all_timing if r["druzyna"]=="gtk" and r["bucket"]==b]
            ors = [r for r in all_timing if r["druzyna"]=="opp" and r["bucket"]==b]
            gm2 = sum(r.get("made2",0) for r in grs); ga2 = sum(r.get("att2",0) for r in grs)
            gm3 = sum(r.get("made3",0) for r in grs); ga3 = sum(r.get("att3",0) for r in grs)
            om2 = sum(r.get("made2",0) for r in ors); oa2 = sum(r.get("att2",0) for r in ors)
            om3 = sum(r.get("made3",0) for r in ors); oa3 = sum(r.get("att3",0) for r in ors)
            g_br = sum(r.get("br",0) for r in grs); o_br = sum(r.get("br",0) for r in ors)
            g_ft = sum(r.get("poss_ft",0) or (1 if r.get("ftm",0)>0 else 0) for r in grs)
            o_ft = sum(r.get("poss_ft",0) or (1 if r.get("ftm",0)>0 else 0) for r in ors)
            gm = gm2+gm3; ga = ga2+ga3
            om = om2+om3; oa = oa2+oa3
            if ga+ga2+ga3+g_br+g_ft+oa+oa2+oa3+o_br+o_ft == 0: continue
            # RAZEM = trafione + posiadania FT + chybione + BR
            g_total = gm + g_ft + (ga-gm) + g_br
            o_total = om + o_ft + (oa-om) + o_br
            ge = f"{gm/ga:.0%}" if ga else "—"
            oe = f"{om/oa:.0%}" if oa else "—"
            bg = "#f5f8ff" if i%2==0 else "#fff"
            rows += f"""<tr style="background:{bg}">
                <td style="font-weight:700;padding:3px 5px">{b}</td>
                <td style="text-align:center">{gm2 or '—'}</td>
                <td style="text-align:center">{gm3 or '—'}</td>
                <td style="text-align:center;color:#1a6b3c">{g_ft or '—'}</td>
                <td style="text-align:center;color:#8b1a1a">{(ga2-gm2) or '—'}</td>
                <td style="text-align:center;color:#8b1a1a">{(ga3-gm3) or '—'}</td>
                <td style="text-align:center;color:#8b1a1a">{g_br or '—'}</td>
                <td style="text-align:center;font-weight:700">{g_total or '—'}</td>
                <td style="text-align:center;font-weight:700;color:#1a6b3c">{ge}</td>
                <td style="width:8px"></td>
                <td style="text-align:center">{om2 or '—'}</td>
                <td style="text-align:center">{om3 or '—'}</td>
                <td style="text-align:center;color:#1a6b3c">{o_ft or '—'}</td>
                <td style="text-align:center;color:#8b1a1a">{(oa2-om2) or '—'}</td>
                <td style="text-align:center;color:#8b1a1a">{(oa3-om3) or '—'}</td>
                <td style="text-align:center;color:#8b1a1a">{o_br or '—'}</td>
                <td style="text-align:center;font-weight:700">{o_total or '—'}</td>
                <td style="text-align:center;font-weight:700;color:#8b1a1a">{oe}</td>
            </tr>"""
        return rows

    TH  = "style='background:#1a2b4a;color:#fff;padding:5px 8px;font-size:10px;text-align:center'"
    THL = "style='background:#1a2b4a;color:#fff;padding:5px 8px;font-size:10px'"
    THG = "style='background:#1a6b3c;color:#fff;padding:5px 8px;font-size:10px;text-align:center'"
    THR = "style='background:#8b1a1a;color:#fff;padding:5px 8px;font-size:10px;text-align:center'"
    THNET = "style='background:#534AB7;color:#fff;padding:5px 8px;font-size:10px;text-align:center'"

    # ── Czy GTK wygrało? ─────────────────────────────────────────────────────
    gtk_won = m['wynik_gtk'] > m['wynik_opp']

    # ── Sekcja porównania paskami (jak v4) ───────────────────────────────────
    # ── Tabela celności ────────────────────────────────────────────────────────
    def _pct(m, a): return f"{m/a:.1%}" if a else "—"
    def _efg(d):
        fga = d.get("p2a",0)+d.get("p3a",0)
        return f"{(d.get('p2m',0)+1.5*d.get('p3m',0))/fga:.1%}" if fga else "—"
    def _ts(d):
        fga = d.get("p2a",0)+d.get("p3a",0); fta = d.get("fta",0)
        return f"{d.get('pts',0)/(2*(fga+0.44*fta)):.1%}" if (fga+fta) else "—"
    def _topct(d):
        poss = max(d.get("poss",1),1)
        return f"{d.get('br',0)/poss:.1%}"
    def _td_cmp(g_val, o_val, better="high"):
        try:
            gn = float(str(g_val).replace('%','').replace('—','0').replace('/',''))
            on = float(str(o_val).replace('%','').replace('—','0').replace('/',''))
            g_wins = gn > on if better=="high" else gn < on
        except: g_wins = False
        gs = "font-weight:700;color:#0F6E56" if g_wins else ""
        os_ = "font-weight:700;color:#A32D2D" if (not g_wins and g_val!=o_val) else ""
        return f'<td style="text-align:center;{gs}">{g_val}</td><td style="text-align:center;{os_}">{o_val}</td>'

    cmp_section = f"""
<div style="margin-bottom:10px">
  <table style="width:100%;border-collapse:collapse;font-size:8.5px">
    <thead>
      <tr style="background:#1a2b4a;color:#fff">
        <th style="padding:4px 6px;text-align:left">Drużyna</th>
        <th style="padding:4px 5px;text-align:center" colspan="2">2PT M/A</th>
        <th style="padding:4px 5px;text-align:center">2PT%</th>
        <th style="padding:4px 5px;text-align:center" colspan="2">3PT M/A</th>
        <th style="padding:4px 5px;text-align:center">3PT%</th>
        <th style="padding:4px 5px;text-align:center" colspan="2">FT M/A</th>
        <th style="padding:4px 5px;text-align:center">FT%</th>
        <th style="padding:4px 5px;text-align:center">eFG%</th>
        <th style="padding:4px 5px;text-align:center">TS%</th>
        <th style="padding:4px 5px;text-align:center">TO%</th>
      </tr>
    </thead>
    <tbody>
      <tr style="background:#f0fff4">
        <td style="padding:4px 6px;font-weight:700;color:#0F6E56">{gtk_name}</td>
        <td style="text-align:center">{suma_gtk.get('p2m',0)}</td>
        <td style="text-align:center">{suma_gtk.get('p2a',0)}</td>
        <td style="text-align:center;font-weight:700">{_pct(suma_gtk.get('p2m',0),suma_gtk.get('p2a',0))}</td>
        <td style="text-align:center">{suma_gtk.get('p3m',0)}</td>
        <td style="text-align:center">{suma_gtk.get('p3a',0)}</td>
        <td style="text-align:center;font-weight:700">{_pct(suma_gtk.get('p3m',0),suma_gtk.get('p3a',0))}</td>
        <td style="text-align:center">{suma_gtk.get('ftm',0)}</td>
        <td style="text-align:center">{suma_gtk.get('fta',0)}</td>
        <td style="text-align:center;font-weight:700">{_pct(suma_gtk.get('ftm',0),suma_gtk.get('fta',0))}</td>
        <td style="text-align:center;font-weight:700">{_efg(suma_gtk)}</td>
        <td style="text-align:center;font-weight:700">{_ts(suma_gtk)}</td>
        <td style="text-align:center">{_topct(suma_gtk)}</td>
      </tr>
      <tr style="background:#fff5f5">
        <td style="padding:4px 6px;font-weight:700;color:#A32D2D">{name_opp}</td>
        <td style="text-align:center">{suma_opp.get('p2m',0)}</td>
        <td style="text-align:center">{suma_opp.get('p2a',0)}</td>
        <td style="text-align:center;font-weight:700">{_pct(suma_opp.get('p2m',0),suma_opp.get('p2a',0))}</td>
        <td style="text-align:center">{suma_opp.get('p3m',0)}</td>
        <td style="text-align:center">{suma_opp.get('p3a',0)}</td>
        <td style="text-align:center;font-weight:700">{_pct(suma_opp.get('p3m',0),suma_opp.get('p3a',0))}</td>
        <td style="text-align:center">{suma_opp.get('ftm',0)}</td>
        <td style="text-align:center">{suma_opp.get('fta',0)}</td>
        <td style="text-align:center;font-weight:700">{_pct(suma_opp.get('ftm',0),suma_opp.get('fta',0))}</td>
        <td style="text-align:center;font-weight:700">{_efg(suma_opp)}</td>
        <td style="text-align:center;font-weight:700">{_ts(suma_opp)}</td>
        <td style="text-align:center">{_topct(suma_opp)}</td>
      </tr>
    </tbody>
  </table>
</div>"""

    # ── Clutch — ostatnie 1/3 posiadań Q4/OT (ceil), taka sama logika jak web ──
    import math as _math
    def _clutch_val(druzyna):
        """Zbiera Q4+ statystyki i aplikuje ratio ceil(poss/3)/poss."""
        qtrs = sorted({r["kwarta"] for r in all_stats
                       if r.get("kwarta",0) >= 4 and r.get("druzyna")==druzyna})
        out = {k:0 for k in ["pts","poss","p2m","p2a","p3m","p3a","ftm","fta","br"]}
        for qn in qtrs:
            qd = next((dict(r) for r in all_stats if r["druzyna"]==druzyna and r["kwarta"]==qn), {})
            poss = qd.get("poss",0) or 0
            cp   = _math.ceil(poss/3) if poss else 0
            ratio = cp/poss if poss else 0
            for k in ["pts","p2m","p2a","p3m","p3a","ftm","fta","br"]:
                out[k] += round((qd.get(k,0) or 0) * ratio)
            out["poss"] += cp
        return out

    cg = _clutch_val("gtk"); co = _clutch_val("opp")
    c_pts_g = cg['pts']; c_pts_o = co['pts']
    c_poss_g = max(cg['poss'],1); c_poss_o = max(co['poss'],1)
    c_p2m_g = cg['p2m']; c_p2a_g = cg['p2a']
    c_p3m_g = cg['p3m']; c_p3a_g = cg['p3a']
    c_ftm_g = cg['ftm']; c_fta_g = cg['fta']
    c_p2m_o = co['p2m']; c_p2a_o = co['p2a']
    c_p3m_o = co['p3m']; c_p3a_o = co['p3a']
    c_ftm_o = co['ftm']; c_fta_o = co['fta']
    c_br_g = cg['br']; c_br_o = co['br']
    # Wynik kwarty Q4 (pełny) do nagłówka
    clutch_gtk_q4 = next((dict(r) for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==4), {})
    clutch_opp_q4 = next((dict(r) for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==4), {})

    def efg_str(pm2, pm3, fga):
        return f"{(pm2+1.5*pm3)/fga:.0%}" if fga else "—"
    def bar_clutch(lbl, g_val, g_denom, o_val, o_denom):
        pg = round(g_val/g_denom*100) if g_denom else 50
        po = 100-pg
        gs = f"{g_val}/{g_denom}" if g_denom else "—"
        os = f"{o_val}/{o_denom}" if o_denom else "—"
        gp = f"{g_val/g_denom:.0%}" if g_denom else "—"
        op = f"{o_val/o_denom:.0%}" if o_denom else "—"
        return f"""<div style="display:flex;align-items:center;gap:5px;margin-bottom:3px">
          <div style="width:22px;font-size:7px;color:#666;text-align:right">{lbl}</div>
          <div style="flex:1;height:12px;background:#f0f0f0;border-radius:3px;overflow:hidden;display:flex">
            <div style="background:#1a6b3c;width:{pg}%;height:100%;display:flex;align-items:center;justify-content:flex-end;padding-right:2px"><span style="font-size:7px;color:#fff;white-space:nowrap">{gs}</span></div>
            <div style="background:#8b1a1a;width:{po}%;height:100%;display:flex;align-items:center;padding-left:2px"><span style="font-size:7px;color:#fff;white-space:nowrap">{os}</span></div>
          </div>
          <div style="width:58px;font-size:7px;color:#888;text-align:right">{gp} vs {op}</div>
        </div>"""

    c_fga_g = c_p2a_g + c_p3a_g; c_fga_o = c_p2a_o + c_p3a_o
    clutch_section = f"""
<div style="background:#fff8e1;border:1px solid #ffe082;border-radius:5px;padding:8px 10px;margin-bottom:10px">
  <div style="font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;color:#5c3d00;margin-bottom:5px">
    Clutch — IV kwarta &nbsp;<span style="background:#EF9F27;color:#fff;font-size:7px;font-weight:700;padding:2px 6px;border-radius:10px">Q4</span>
    <span style="font-size:7px;color:#7a5000;font-weight:400;margin-left:6px">Wynik kwarty: {gtk_name} <strong>{clutch_gtk_q4.get('pts',0)}</strong> – {name_opp} <strong>{clutch_opp_q4.get('pts',0)}</strong> &nbsp;·&nbsp; ostatnie 1/3 posiadań</span>
  </div>
  <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:5px;margin-bottom:7px">
    <div style="background:#fff;border-radius:4px;padding:4px 3px;text-align:center;border:1px solid #ffe082">
      <div style="display:flex;justify-content:center;align-items:baseline;gap:5px"><span style="font-size:12px;font-weight:700;color:#1a6b3c">{c_pts_g}</span><span style="font-size:7px;color:#aaa">vs</span><span style="font-size:10px;font-weight:700;color:#8b1a1a">{c_pts_o}</span></div>
      <div style="font-size:7px;color:#888;text-transform:uppercase">Punkty</div>
    </div>
    <div style="background:#fff;border-radius:4px;padding:4px 3px;text-align:center;border:1px solid #ffe082">
      <div style="display:flex;justify-content:center;align-items:baseline;gap:5px"><span style="font-size:12px;font-weight:700;color:#1a6b3c">{efg_str(c_p2m_g,c_p3m_g,c_fga_g)}</span><span style="font-size:7px;color:#aaa">vs</span><span style="font-size:10px;font-weight:700;color:#8b1a1a">{efg_str(c_p2m_o,c_p3m_o,c_fga_o)}</span></div>
      <div style="font-size:7px;color:#888;text-transform:uppercase">eFG%</div>
    </div>
    <div style="background:#fff;border-radius:4px;padding:4px 3px;text-align:center;border:1px solid #ffe082">
      <div style="display:flex;justify-content:center;align-items:baseline;gap:5px"><span style="font-size:12px;font-weight:700;color:#1a6b3c">{cg['poss']}</span><span style="font-size:7px;color:#aaa">vs</span><span style="font-size:10px;font-weight:700;color:#8b1a1a">{co['poss']}</span></div>
      <div style="font-size:7px;color:#888;text-transform:uppercase">Posiadania</div>
    </div>
    <div style="background:#fff;border-radius:4px;padding:4px 3px;text-align:center;border:1px solid #ffe082">
      <div style="display:flex;justify-content:center;align-items:baseline;gap:5px"><span style="font-size:12px;font-weight:700;color:#1a6b3c">{c_br_g}</span><span style="font-size:7px;color:#aaa">vs</span><span style="font-size:10px;font-weight:700;color:#8b1a1a">{c_br_o}</span></div>
      <div style="font-size:7px;color:#888;text-transform:uppercase">Straty</div>
    </div>
  </div>
  {bar_clutch('2PT', c_p2m_g, c_p2a_g, c_p2m_o, c_p2a_o)}
  {bar_clutch('3PT', c_p3m_g, c_p3a_g, c_p3m_o, c_p3a_o)}
  {bar_clutch('FT',  c_ftm_g, c_fta_g, c_ftm_o, c_fta_o)}
</div>"""

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  @page{{size:A4 landscape;margin:1cm 1.2cm}}
  body{{font-family:Arial,sans-serif;font-size:11px;color:#222;background:#fff;width:100%}}
  .page{{width:100%;padding:0}}
  h3{{font-size:10px;text-transform:uppercase;letter-spacing:.5px;margin:0 0 5px;font-weight:700}}
  .hero{{background:#1a2b4a;padding:11px 18px;border-radius:6px;margin-bottom:10px;display:flex;justify-content:space-between;align-items:center}}
  .hero h2{{font-size:16px;font-weight:700;color:#fff;margin-bottom:2px}}
  .hero .meta{{font-size:9px;color:#a8c4e0}}
  .score{{font-size:30px;font-weight:700;letter-spacing:3px;color:#EF9F27}}
  .winner{{font-size:11px;font-weight:700;color:#5DCAA5}}
  .two-col{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:10px}}
  .section{{margin-bottom:10px}}
  table{{width:100%;border-collapse:collapse;font-size:9.5px;margin-bottom:0}}
  th{{padding:5px 6px;color:#fff;font-size:8.5px;text-align:center}}
  th.left{{text-align:left}}
  td{{padding:4px 6px;border-bottom:1px solid #f0f0f0;text-align:center}}
  td.left{{text-align:left;font-weight:700}}
  tr.sum td{{font-weight:700}}
  .sep{{height:1px;background:#eee;margin:8px 0}}
  .kpi-header{{text-align:center;margin-bottom:5px}}
  .kpi-header-title{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.5px}}
  .kpi-grid{{display:grid;grid-template-columns:repeat(7,1fr);gap:4px;margin-bottom:7px}}
  .kpi-box{{background:#f5f8ff;border-radius:4px;padding:5px 4px;text-align:center}}
  .kpi-val{{font-size:13px;font-weight:700;color:#1a2b4a}}
  .kpi-lbl{{font-size:7px;color:#888;text-transform:uppercase;line-height:1.2}}
  .kpi-box.kpi-net{{background:#e8f0fb;border:1px solid #c5d5f5}}
  .kpi-box.kpi-net .kpi-val{{color:#534AB7;font-size:12px;line-height:1.4}}
  .kpi-box.kpi-net .kpi-lbl{{color:#534AB7;font-weight:700}}
  .kpi-sub{{font-size:7px;color:#888;font-weight:400;display:block;margin-top:2px}}
  .net-pos{{color:#1a6b3c;font-weight:700}}
  .net-neg{{color:#8b1a1a;font-weight:700}}
  @media print{{body{{-webkit-print-color-adjust:exact;print-color-adjust:exact}}}}
</style>
</head><body>
<div class="page">

<!-- HERO -->
<div class="hero">
  <div>
    <h2>{gtk_name} vs {name_opp}</h2>
    <div class="meta">Sezon {m['sezon']} · {dt} · Basket Kołcz Analytics</div>
  </div>
  <div class="score">{m['wynik_gtk']} : {m['wynik_opp']}</div>
  <div class="winner">{'✓ ' if gtk_won else ''}{winner}</div>
</div>

<!-- KPI — dwie kolumny -->
<div class="two-col" style="margin-bottom:6px">
  <div>
    <div class="kpi-header"><div class="kpi-header-title" style="color:#1a6b3c">{gtk_name} — metryki kluczowe</div></div>
    <div class="kpi-grid" style="grid-template-columns:repeat(5,1fr);margin-bottom:4px">
      <div class="kpi-box"><div class="kpi-val">{suma_gtk.get('pts',0)}</div><div class="kpi-lbl">Punkty</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_gtk.get('oreb',0)}&thinsp;/&thinsp;{suma_gtk.get('dreb',0)}</div><div class="kpi-lbl">Zb Off/Def</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_gtk.get('ast',0)}</div><div class="kpi-lbl">Asysty</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_gtk.get('stl',0)}</div><div class="kpi-lbl">Przechwyty</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_gtk.get('br',0)}</div><div class="kpi-lbl">Straty</div></div>
    </div>
    <div class="kpi-grid" style="grid-template-columns:repeat(5,1fr)">
      <div class="kpi-box"><div class="kpi-val">{suma_gtk.get('d2m',0)}/{suma_gtk.get('d2a',0)}</div><div class="kpi-lbl">Dobitki M/A</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_gtk.get('przerw',0)}</div><div class="kpi-lbl">Przerwania</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_gtk.get('fd',0)}</div><div class="kpi-lbl">Faule wym.</div></div>
      <div class="kpi-box"><div class="kpi-val">{kpi_gtk['ppp']}</div><div class="kpi-lbl">PPP</div></div>
      <div class="kpi-box kpi-net"><div class="kpi-val">{float(kpi_gtk['ortg'])-float(kpi_opp['ortg']):+.1f}</div><div class="kpi-lbl">NETrtg</div></div>
    </div>
  </div>
  <div>
    <div class="kpi-header"><div class="kpi-header-title" style="color:#8b1a1a">{name_opp} — metryki kluczowe</div></div>
    <div class="kpi-grid" style="grid-template-columns:repeat(5,1fr);margin-bottom:4px">
      <div class="kpi-box"><div class="kpi-val">{suma_opp.get('pts',0)}</div><div class="kpi-lbl">Punkty</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_opp.get('oreb',0)}&thinsp;/&thinsp;{suma_opp.get('dreb',0)}</div><div class="kpi-lbl">Zb Off/Def</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_opp.get('ast',0)}</div><div class="kpi-lbl">Asysty</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_opp.get('stl',0)}</div><div class="kpi-lbl">Przechwyty</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_opp.get('br',0)}</div><div class="kpi-lbl">Straty</div></div>
    </div>
    <div class="kpi-grid" style="grid-template-columns:repeat(5,1fr)">
      <div class="kpi-box"><div class="kpi-val">{suma_opp.get('d2m',0)}/{suma_opp.get('d2a',0)}</div><div class="kpi-lbl">Dobitki M/A</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_opp.get('przerw',0)}</div><div class="kpi-lbl">Przerwania</div></div>
      <div class="kpi-box"><div class="kpi-val">{suma_opp.get('fd',0)}</div><div class="kpi-lbl">Faule wym.</div></div>
      <div class="kpi-box"><div class="kpi-val">{kpi_opp['ppp']}</div><div class="kpi-lbl">PPP</div></div>
      <div class="kpi-box kpi-net"><div class="kpi-val">{float(kpi_opp['ortg'])-float(kpi_gtk['ortg']):+.1f}</div><div class="kpi-lbl">NETrtg</div></div>
    </div>
  </div>
</div>

<!-- PORÓWNANIE PASKAMI -->
{cmp_section}

<!-- PER KWARTA — dwie kolumny -->
<div class="two-col">
  <div class="section">
    <h3 style="color:#1a6b3c">{gtk_name} — per kwarta</h3>
    <table>
      <thead><tr>
        <th style="background:#1a2b4a" class="left">Q</th>
        <th style="background:#1a2b4a">PKT</th>
        <th style="background:#1a2b4a">2PM/A</th>
        <th style="background:#1a2b4a">3PM/A</th>
        <th style="background:#1a2b4a">FTM/A</th>
        <th style="background:#1a2b4a">TO</th>
        <th style="background:#1a2b4a">eFG%</th>
        <th style="background:#534AB7">NetRtg</th>
      </tr></thead>
      <tbody>
        {q_row('gtk',1,'1Q','opp')}{q_row('gtk',2,'2Q','opp')}{q_row('gtk',3,'3Q','opp')}{q_row('gtk',4,'4Q','opp')}
        <tr class="sum" style="background:#e8f0fb">
          <td class="left">SUMA</td>
          <td>{suma_gtk.get('pts',0)}</td>
          <td>{suma_gtk.get('p2m',0)}/{suma_gtk.get('p2a',0)}</td>
          <td>{suma_gtk.get('p3m',0)}/{suma_gtk.get('p3a',0)}</td>
          <td>{suma_gtk.get('ftm',0)}/{suma_gtk.get('fta',0)}</td>
          <td>{suma_gtk.get('br',0)}</td>
          <td>{kpi_gtk['efg']}</td>
          <td class="{'net-pos' if float(kpi_gtk['ortg'])>float(kpi_opp['ortg']) else 'net-neg'}">{float(kpi_gtk['ortg'])-float(kpi_opp['ortg']):+.1f}</td>
        </tr>
      </tbody>
    </table>
  </div>
  <div class="section">
    <h3 style="color:#8b1a1a">{name_opp} — per kwarta</h3>
    <table>
      <thead><tr>
        <th style="background:#1a2b4a" class="left">Q</th>
        <th style="background:#1a2b4a">PKT</th>
        <th style="background:#1a2b4a">2PM/A</th>
        <th style="background:#1a2b4a">3PM/A</th>
        <th style="background:#1a2b4a">FTM/A</th>
        <th style="background:#1a2b4a">TO</th>
        <th style="background:#1a2b4a">eFG%</th>
        <th style="background:#534AB7">NetRtg</th>
      </tr></thead>
      <tbody>
        {q_row('opp',1,'1Q','gtk')}{q_row('opp',2,'2Q','gtk')}{q_row('opp',3,'3Q','gtk')}{q_row('opp',4,'4Q','gtk')}
        <tr class="sum" style="background:#fce8e8">
          <td class="left">SUMA</td>
          <td>{suma_opp.get('pts',0)}</td>
          <td>{suma_opp.get('p2m',0)}/{suma_opp.get('p2a',0)}</td>
          <td>{suma_opp.get('p3m',0)}/{suma_opp.get('p3a',0)}</td>
          <td>{suma_opp.get('ftm',0)}/{suma_opp.get('fta',0)}</td>
          <td>{suma_opp.get('br',0)}</td>
          <td>{kpi_opp['efg']}</td>
          <td class="{'net-pos' if float(kpi_opp['ortg'])>float(kpi_gtk['ortg']) else 'net-neg'}">{float(kpi_opp['ortg'])-float(kpi_gtk['ortg']):+.1f}</td>
        </tr>
      </tbody>
    </table>
  </div>
</div>

<div class="sep"></div>

<!-- ZAWODNICY -->
<div class="section">
  <h3 style="color:#1a6b3c">{gtk_name} — Zawodnicy</h3>
  <table style="table-layout:fixed;width:100%">
    <thead>
      <tr style="background:#1a2b4a;color:#fff;font-size:7.5px">
        <th style="text-align:left;padding:3px 4px" rowspan="2">Zawodnik</th>
        <th style="text-align:center;padding:3px 3px" rowspan="2">MIN<br>(szac.)</th>
        <th style="text-align:center;padding:3px 3px" rowspan="2">PTS</th>
        <th style="text-align:center;padding:3px 3px;background:#152236;color:rgba(255,255,255,.55)" colspan="2">2PT</th>
        <th style="text-align:center;padding:3px 3px;background:#152236;color:rgba(255,255,255,.55)" colspan="2">3PT</th>
        <th style="text-align:center;padding:3px 3px;background:#152236;color:rgba(255,255,255,.55)" colspan="2">FT</th>
        <th style="text-align:center;padding:3px 3px;background:#0d1b2e;color:rgba(255,255,255,.45)" colspan="2">ZB</th>
        <th style="text-align:center;padding:3px 3px" rowspan="2">AST</th>
        <th style="text-align:center;padding:3px 3px" rowspan="2">TO</th>
        <th style="text-align:center;padding:3px 3px" rowspan="2">STL</th>
        <th style="text-align:center;padding:3px 3px" rowspan="2">BLK</th>
        <th style="text-align:center;padding:3px 3px" rowspan="2">eFG%</th>
        <th style="text-align:center;padding:3px 3px" rowspan="2">TS%</th>
        <th style="text-align:center;padding:3px 3px" rowspan="2">USG%</th>
        <th style="text-align:center;padding:3px 3px" rowspan="2">FIN</th>
      </tr>
      <tr style="background:#1a2b4a;color:rgba(255,255,255,.7);font-size:7px">
        <th style="text-align:center;padding:2px 3px;background:#152236">M/A</th>
        <th style="text-align:center;padding:2px 3px;background:#152236">%</th>
        <th style="text-align:center;padding:2px 3px;background:#152236">M/A</th>
        <th style="text-align:center;padding:2px 3px;background:#152236">%</th>
        <th style="text-align:center;padding:2px 3px;background:#152236">M/A</th>
        <th style="text-align:center;padding:2px 3px;background:#152236">%</th>
        <th style="text-align:center;padding:2px 3px;background:#0d1b2e">A</th>
        <th style="text-align:center;padding:2px 3px;background:#0d1b2e">O</th>
      </tr>
    </thead>
    <tbody>{player_rows('gtk') or '<tr><td colspan="19" style="text-align:center;color:#aaa">Brak danych</td></tr>'}</tbody>
  </table>
</div>

<div class="sep"></div>

<!-- CLUTCH -->
{clutch_section}

<!-- TIMING AKCJI -->
<div class="section">
  <h3>Timing Akcji — skuteczność rzutów w czasie posiadania</h3>
  <table style="width:100%">
    <thead>
      <tr>
        <th rowspan="2" style="background:#1a2b4a;text-align:left;width:50px">Czas</th>
        <th colspan="8" style="background:#1a6b3c">{gtk_name}</th>
        <th style="background:#fff;width:8px"></th>
        <th colspan="8" style="background:#8b1a1a">{name_opp}</th>
      </tr>
      <tr>
        <th style="background:#085041">2PT✓</th><th style="background:#085041">3PT✓</th><th style="background:#085041">FT</th>
        <th style="background:#5a1515">2PT✗</th><th style="background:#5a1515">3PT✗</th><th style="background:#5a1515">BR</th>
        <th style="background:#1a2b4a">RAZEM</th><th style="background:#1a2b4a">EFG%</th>
        <th style="background:#fff;width:8px"></th>
        <th style="background:#085041">2PT✓</th><th style="background:#085041">3PT✓</th><th style="background:#085041">FT</th>
        <th style="background:#5a1515">2PT✗</th><th style="background:#5a1515">3PT✗</th><th style="background:#5a1515">BR</th>
        <th style="background:#1a2b4a">RAZEM</th><th style="background:#1a2b4a">EFG%</th>
      </tr>
    </thead>
    <tbody>{timing_rows()}</tbody>
  </table>
</div>

<div style="margin-top:10px;text-align:center;font-size:7.5px;color:#aaa;border-top:1px solid #eee;padding-top:5px">
  Basket Kołcz Analytics · {gtk_name} vs {name_opp} · {dt}
</div>

</div>
</body></html>"""

    # Wstaw piątki przed footer
    if lineup_html:
        html = html.replace(
            '  Basket Kołcz Analytics · ' + gtk_name + ' vs ' + name_opp + ' · ' + dt,
            lineup_html + '\n<div style="text-align:center;font-size:7.5px;color:#aaa;border-top:1px solid #eee;padding-top:5px">\n  Basket Kołcz Analytics · ' + gtk_name + ' vs ' + name_opp + ' · ' + dt
        )

    html_print = html.replace(
        "</body></html>",
        "<script>window.onload=function(){window.print();}</script></body></html>"
    )
    return Response(html_print, mimetype="text/html")


# ══════════════════════════════════════════════════════════════════════════════
# SZABLONY (identyczne jak w app_v2)
# ══════════════════════════════════════════════════════════════════════════════


@app.route("/template/sklad")
@login_required
def template_sklad():
    import base64 as _b64
    from io import BytesIO
    data = (
        "UEsDBBQAAAAIANymdlxGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZ"
        "ih6kDkQ9ip68zy51hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGM"
        "sRoPpB8eA8OibdeAhTEMOMzit7Dp1C5GZ3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+L"
        "U8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIANymdlwuZU2A7gAAACsCAAARAAAAZG9jUHJvcHMvY29yZS54"
        "bWzNksFOwzAMhl8F5d66TVkPUZcLiBNISEwCcYsSb4tomigxavf2tGHrhOABOMb+8/mz5E4HoX3E5+gD"
        "RrKYbibXD0nosGVHoiAAkj6iU6mcE8Pc3PvoFM3PeICg9Ic6IPCqasEhKaNIwQIswkpksjNa6IiKfDzj"
        "jV7x4TP2GWY0YI8OB0pQlzUwuUwMp6nv4ApYYITRpe8CmpWYq39icwfYOTklu6bGcSzHJufmHWp4e3p8"
        "yesWdkikBo3zr2QFnQJu2WXya3N3v3tgkle8Laqm4HzHK7G5FZv2fXH94XcVdt7Yvf3HxhdB2cGvu5Bf"
        "UEsDBBQAAAAIANymdlyZXJwjEAYAAJwnAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbO1aW3PaOBR+76/Q"
        "eGf2bQvGNoG2tBNzaXbbtJmE7U4fhRFYjWx5ZJGEf79HNhDLlg3tkk26mzwELOn7zkVH5+g4efPuLmLo"
        "hoiU8nhg2S/b1ru3L97gVzIkEUEwGaev8MAKpUxetVppAMM4fckTEsPcgosIS3gUy9Zc4FsaLyPW6rTb"
        "3VaEaWyhGEdkYH1eLGhA0FRRWm9fILTlHzP4FctUjWWjARNXQSa5iLTy+WzF/NrePmXP6TodMoFuMBtY"
        "IH/Ob6fkTlqI4VTCxMBqZz9Wa8fR0kiAgsl9lAW6Sfaj0xUIMg07Op1YznZ89sTtn4zK2nQ0bRrg4/F4"
        "OLbL0otwHATgUbuewp30bL+kQQm0o2nQZNj22q6RpqqNU0/T933f65tonAqNW0/Ta3fd046Jxq3QeA2+"
        "8U+Hw66JxqvQdOtpJif9rmuk6RZoQkbj63oSFbXlQNMgAFhwdtbM0gOWXin6dZQa2R273UFc8FjuOYkR"
        "/sbFBNZp0hmWNEZynZAFDgA3xNFMUHyvQbaK4MKS0lyQ1s8ptVAaCJrIgfVHgiHF3K/99Ze7yaQzep19"
        "Os5rlH9pqwGn7bubz5P8c+jkn6eT101CznC8LAnx+yNbYYcnbjsTcjocZ0J8z/b2kaUlMs/v+QrrTjxn"
        "H1aWsF3Pz+SejHIju932WH32T0duI9epwLMi15RGJEWfyC265BE4tUkNMhM/CJ2GmGpQHAKkCTGWoYb4"
        "tMasEeATfbe+CMjfjYj3q2+aPVehWEnahPgQRhrinHPmc9Fs+welRtH2Vbzco5dYFQGXGN80qjUsxdZ4"
        "lcDxrZw8HRMSzZQLBkGGlyQmEqk5fk1IE/4rpdr+nNNA8JQvJPpKkY9psyOndCbN6DMawUavG3WHaNI8"
        "ev4F+Zw1ChyRGx0CZxuzRiGEabvwHq8kjpqtwhErQj5iGTYacrUWgbZxqYRgWhLG0XhO0rQR/FmsNZM+"
        "YMjszZF1ztaRDhGSXjdCPmLOi5ARvx6GOEqa7aJxWAT9nl7DScHogstm/bh+htUzbCyO90fUF0rkDyan"
        "P+kyNAejmlkJvYRWap+qhzQ+qB4yCgXxuR4+5Xp4CjeWxrxQroJ7Af/R2jfCq/iCwDl/Ln3Ppe+59D2h"
        "0rc3I31nwdOLW95GblvE+64x2tc0LihjV3LNyMdUr5Mp2DmfwOz9aD6e8e362SSEr5pZLSMWkEuBs0Ek"
        "uPyLyvAqxAnoZFslCctU02U3ihKeQhtu6VP1SpXX5a+5KLg8W+Tpr6F0PizP+Txf57TNCzNDt3JL6raU"
        "vrUmOEr0scxwTh7LDDtnPJIdtnegHTX79l125COlMFOXQ7gaQr4Dbbqd3Do4npiRuQrTUpBvw/npxXga"
        "4jnZBLl9mFdt59jR0fvnwVGwo+88lh3HiPKiIe6hhpjPw0OHeXtfmGeVxlA0FG1srCQsRrdguNfxLBTg"
        "ZGAtoAeDr1EC8lJVYDFbxgMrkKJ8TIxF6HDnl1xf49GS49umZbVuryl3GW0iUjnCaZgTZ6vK3mWxwVUd"
        "z1Vb8rC+aj20FU7P/lmtyJ8MEU4WCxJIY5QXpkqi8xlTvucrScRVOL9FM7YSlxi84+bHcU5TuBJ2tg8C"
        "Mrm7Oal6ZTFnpvLfLQwJLFuIWRLiTV3t1eebnK56Inb6l3fBYPL9cMlHD+U751/0XUOufvbd4/pukztI"
        "TJx5xREBdEUCI5UcBhYXMuRQ7pKQBhMBzZTJRPACgmSmHICY+gu98gy5KRXOrT45f0Usg4ZOXtIlEhSK"
        "sAwFIRdy4+/vk2p3jNf6LIFthFQyZNUXykOJwT0zckPYVCXzrtomC4Xb4lTNuxq+JmBLw3punS0n/9te"
        "1D20Fz1G86OZ4B6zh3OberjCRaz/WNYe+TLfOXDbOt4DXuYTLEOkfsF9ioqAEativrqvT/klnDu0e/GB"
        "IJv81tuk9t3gDHzUq1qlZCsRP0sHfB+SBmOMW/Q0X48UYq2msa3G2jEMeYBY8wyhZjjfh0WaGjPVi6w5"
        "jQpvQdVA5T/b1A1o9g00HJEFXjGZtjaj5E4KPNz+7w2wwsSO4e2LvwFQSwMEFAAAAAgA3KZ2XNMdJvaB"
        "BAAAOxUAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWy9mFtvozoUhf8KYqR5qkpMLr0lkabF1qk0"
        "HVWtzpxnN3ESJoA5xilDfv2YSwFfds/beUmwv7XtxSbtEl6WXByLA2PS+50mWbHyD1Lmt0FQbA4spcUl"
        "z1mmyI6LlEo1FPugyAWj26YoTYJwMlkEKY0zf71s5p7FeslPMokz9iy84pSmVFT3LOHlykf+x8RLvD/I"
        "eiJYL3O6Z69M/p0/CzUK+lW2ccqyIuaZJ9hu5X9DtyRsChrFz5iVxejaq2/ljfNjPXjcrvyJXy+dMa96"
        "zZO42cyTPP/OdvKBJYlaMPQ9upHxO3tWspX/xqXkac2VTUmlmtoJfmZZsydLmNIqM7klbhfpFq3v8d/O"
        "sN/fT21qfP3hnDSNVY16owV74Mk/8VYeVv61723Zjp4S+cLLv1jXrHm93oYnRfPpla124XubU6HMdLXK"
        "QBpn7Tf93fV4pEfXQEHYFYRGQTgBCqZdwdQogDaYdfqZucEMKJh3BXPzFqCCRVewaDrftqrpc0QlXS8F"
        "Lz3RqOt+huHHKn2H1U9mUyuap9j+YlZ+nNU/5lcpFI3VgnL9Pb9cBlLtUA+DTVd0/3nRYxozR9XD51U/"
        "6LmMiyN3VEb/UXlKmXCU4c/Lnvm52vyijkLyeeGr+pM5FXpdoDretz3suxs2C4XNQvU/jvc1Wgbv41aa"
        "iqFdIIlAgkFCXEQzPe1NTxvpdGQ6NEybisE0SCKQYJAQF9FMz3rTM6vTU8O0qRhMgyQCCQYJcRHN9Lw3"
        "Pbc6PTNMm4rBNEgikGCQEBfRTC960wur03PDtKkYTIMkAgkGCXERzfRVb/rK6vTCMG0qBtMgiUCCQUJc"
        "RDN93Zu+tjp9ZZg2FYNpkEQgwSAhLqKZvulN31idvjZMm4rBNEgikGCQEBfRTKPJEIMTq9c3hm1LMko2"
        "EEUwwjAiTqR7H0U4slqOJqZ5UzMyD6IIRhhGxIl080NCIkdEmhlpaUbm4ZSEEYYRcSLd/JCUyI5KZGal"
        "pRmZh9MSRhhGxIl080NiIjsykZmZlmZkHk5NGGEYESfSzQ/JiezoRGZ2WpqReTg9YYRhRJxINz8kKLIj"
        "FJkZamlG5uEUhRGGEXEi3fyQpMiOUmRmqaUZmYfTFEYYRsSJdPNDoiI7UpGZqZZmZB5OVRhhGBEn0s0P"
        "yYrsaEVmtlqakXk4XWGEYUScSH8RGgI2tAMWmQlraUYvQ3DCwgjDiDiRbn5I2NBO2NBMWEszMg8nLIww"
        "jIgTteaD0SnAVn3+pEmsvmOeFd6GnzLZvp7q6OMgB4e37bYHXkaC5xEvs/p8qZl4zPKTfGJFQfesn8RC"
        "cDGepEnCy/uEZsf2BbrK1XwSF1JtWh+unRKK1v4LP+9FVdJfX7+EC3S3qS5ezqfNaPhEv36ZhuFd5RVH"
        "ca627YiX1cWrOY6TzFapq9niTvDtUY38ZdBvvQz0W4daQcJb8r+04ttRVmVWXfyIGW0vP7drTBTtseIT"
        "FftYPeKE7dQTnlxeqTwS7VlPO5A8b1y0x3nt0RCjWyZqgeI7zuXHoD5S6s9L138AUEsDBBQAAAAIANym"
        "dlzgX4QB9AIAAAEOAAANAAAAeGwvc3R5bGVzLnhtbN1XUW+bMBD+K4gfMEJoUZhCpJYp0qRtqtQ+7NUJ"
        "BiwZzIypkv76+WwCpPFV6baHbURN7Pvu++58PmN13akjp48Vpco71LzpUr9Sqv0YBN2+ojXpPoiWNhop"
        "hKyJ0lNZBl0rKck7INU8WC4WcVAT1vibddPX21p13l70jUr9hR9s1oVoJsvStwbtSmrqPROe+hnhbCeZ"
        "8SU140drXoJhL7iQntKp0NQPwdK9WDi0M8hy0KlZIyQYAxvBfu8G90lNljud2mJrnkvJgW1+Oq3COB/T"
        "v/WtYbNuiVJUNls9MRxjvIC8Yfx0bHX+pSTHcHnrX03oBGc5hCyzeebh3fL+5s7IzKi/KbpdbZPtHxcd"
        "a+wUNT+6xjshcyrPmsSaNmtOC6XpkpUV/CrRBgAqJWo9yBkpRUPMFpwYc6Zn2jv1VWXa82z/P5nH5Aau"
        "Q4wrGcbXpHMlQXue8r6SYZ1nCxsGul57yvkjiHwvxqKFWupQePYEfs7h8HnQwqehrvQwtDJ2AoHmalZ7"
        "JnvzS7Jey56Fuu/1Choz/9ELRR8kLdjBzA/FGB9TDyf15Vxd20nb8uMdZ2VTU7v2qwNu1uTE8yoh2YuO"
        "Bmd/rw1U+t4zlYrtZxao0KG4qgjRv5HmzV+QZjC02ayXzzp5tHpwT6T+N7h++BTY2/WMK9YMs4rlOW0u"
        "GlrLK7LT99uZvvbPaUF6rp5GMPWn8Veas75ORq8HKMbgNY2/wBsgjMdrRcdiTU4PNM+GqT7SZy9D+wDh"
        "NTK9Ji8RjGMxNwIYFgfLAONYFhbnf1rPCl2PxbDcVk5khXJWKMeyXEhmPlgcNyfRj3ulSRJFcYxVNMuc"
        "GWRY3eIY/txqWG7AwOJApPfVGt9tvEPe7gNsT9/qEGyleCdiK8VrDYi7bsBIEvduY3GAge0C1jsQ3x0H"
        "esrNiSLYVSw37ATjSJJgCPSiu0fjGKlODB/3/mCnJIqSxI0A5s4gijAETiOOYBlADhgSReYefHUfBad7"
        "Kpj+6dv8BFBLAwQUAAAACADcpnZcl4q7HMAAAAATAgAACwAAAF9yZWxzLy5yZWxznZK5bsMwDEB/xdCe"
        "MAfQIYgzZfEWBPkBVqIP2BIFikWdv6/apXGQCxl5PTwS3B5pQO04pLaLqRj9EFJpWtW4AUi2JY9pzpFC"
        "rtQsHjWH0kBE22NDsFosPkAuGWa3vWQWp3OkV4hc152lPdsvT0FvgK86THFCaUhLMw7wzdJ/MvfzDDVF"
        "5UojlVsaeNPl/nbgSdGhIlgWmkXJ06IdpX8dx/aQ0+mvYyK0elvo+XFoVAqO3GMljHFitP41gskP7H4A"
        "UEsDBBQAAAAIANymdlzWt1fHQQEAACwCAAAPAAAAeGwvd29ya2Jvb2sueG1sjVFBbsIwEPxK5Ac0AbVI"
        "RYRLUVukqkWl4m7iDVlhe6P1BgrHvqtX/lUnUVSkXnqyZ3Y1nhnPjsT7LdE++XTWh1xVIvU0TUNRgdPh"
        "hmrwcVISOy0R8i4NNYM2oQIQZ9Nxlk1Sp9Gr+WzQWnF6DUigECQfyZbYIBzD77yFyQEDbtGinHLV3S2o"
        "xKFHh2cwucpUEio6PhPjmbxouy6YrM3VqB9sgAWLP/S6Nfmht6FjRG/fdTSSq0kWBUvkIN1Gp6+jxwPE"
        "5R41Qo9oBXihBZ6Ymhr9rpWJKdKrGF0Pw9mXOOX/1EhliQUsqGgceOl7ZLCtQR8qrINKvHaQq/X+8qVN"
        "Yri5fJ/8qU0Wn1qaPqVEe1ed8RTjgJemNzq4M1CiB/MaBUPkY1PFipP26HTGt3ej+9hIY+1D5N78C2kz"
        "hB0+av4DUEsDBBQAAAAIANymdlwkHpuirQAAAPgBAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJl"
        "bHO1kT0OgzAMha8S5QA1UKlDBUxdWCsuEAXzIxISxa4Kty+FAZA6dGGyni1/78lOn2gUd26gtvMkRmsG"
        "ymTL7O8ApFu0ii7O4zBPahes4lmGBrzSvWoQkii6QdgzZJ7umaKcPP5DdHXdaXw4/bI48A8wvF3oqUVk"
        "KUoVGuRMwmi2NsFS4stMlqKoMhmKKpZwWiDiySBtaVZ9sE9OtOd5Fzf3Ra7N4wmu3wxweHT+AVBLAwQU"
        "AAAACADcpnZcZZB5khkBAADPAwAAEwAAAFtDb250ZW50X1R5cGVzXS54bWytk01OwzAQha8SZVslLixY"
        "oKYbYAtdcAFjTxqr/pNnWtLbM07aSqASFYVNrHjevM+el6zejxGw6J312JQdUXwUAlUHTmIdIniutCE5"
        "SfyatiJKtZNbEPfL5YNQwRN4qih7lOvVM7Ryb6l46XkbTfBNmcBiWTyNwsxqShmjNUoS18XB6x+U6kSo"
        "uXPQYGciLlhQiquEXPkdcOp7O0BKRkOxkYlepWOV6K1AOlrAetriyhlD2xoFOqi945YaYwKpsQMgZ+vR"
        "dDFNJp4wjM+72fzBZgrIyk0KETmxBH/HnSPJ3VVkI0hkpq94IbL17PtBTluDvpHN4/0MaTfkgWJY5s/4"
        "e8YX/xvO8RHC7r8/sbzWThp/5ovhP15/AVBLAQIUAxQAAAAIANymdlxGx01IlQAAAM0AAAAQAAAAAAAA"
        "AAAAAACAAQAAAABkb2NQcm9wcy9hcHAueG1sUEsBAhQDFAAAAAgA3KZ2XC5lTYDuAAAAKwIAABEAAAAA"
        "AAAAAAAAAIABwwAAAGRvY1Byb3BzL2NvcmUueG1sUEsBAhQDFAAAAAgA3KZ2XJlcnCMQBgAAnCcAABMA"
        "AAAAAAAAAAAAAIAB4AEAAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECFAMUAAAACADcpnZc0x0m9oEEAAA7"
        "FQAAGAAAAAAAAAAAAAAAgIEhCAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAhQDFAAAAAgA3KZ2"
        "XOBfhAH0AgAAAQ4AAA0AAAAAAAAAAAAAAIAB2AwAAHhsL3N0eWxlcy54bWxQSwECFAMUAAAACADcpnZc"
        "l4q7HMAAAAATAgAACwAAAAAAAAAAAAAAgAH3DwAAX3JlbHMvLnJlbHNQSwECFAMUAAAACADcpnZc1rdX"
        "x0EBAAAsAgAADwAAAAAAAAAAAAAAgAHgEAAAeGwvd29ya2Jvb2sueG1sUEsBAhQDFAAAAAgA3KZ2XCQe"
        "m6KtAAAA+AEAABoAAAAAAAAAAAAAAIABThIAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAhQD"
        "FAAAAAgA3KZ2XGWQeZIZAQAAzwMAABMAAAAAAAAAAAAAAIABMxMAAFtDb250ZW50X1R5cGVzXS54bWxQ"
        "SwUGAAAAAAkACQA+AgAAfRQAAAAA"
    )
    raw = _b64.b64decode("".join(data.split()))
    from flask import send_file
    if request.args.get("fmt") == "csv":
        csv_rows = ["Lp.,Imie,Nazwisko,Numer,Pozycja,Status"]
        for i in range(1, 21):
            csv_rows.append(f"{i},,,,Rozgrywający,Aktywny")
        return send_file(
            BytesIO("\n".join(csv_rows).encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name="sklad_druzyny.csv"
        )
    return send_file(BytesIO(raw),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name="sklad_druzyny.xlsx")


@app.route("/template/zapis")
@login_required
def template_zapis():
    import base64 as _b64
    import io
    _data = (
    "UEsDBBQAAAAIAEyjeVxGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZ"
    "ih6kDkQ9ip68zy51hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGM"
    "sRoPpB8eA8OibdeAhTEMOMzit7Dp1C5GZ3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+L"
    "U8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIAEyjeVyRVLGgFgEAAF4CAAARAAAAZG9jUHJvcHMvY29yZS54"
    "bWzNkkFOwzAQRa9SeZ9M4pQAVuoFIDalEhKRQN1ZzrSNiGPLNkrCkrtxL5KQplRwAJb+//vNH2kyaZjU"
    "Fh+tNmh9iW7Rqqp2TJoVOXhvGICTB1TChX2i7s2dtkr4/mn3YIR8FXsEGkUpKPSiEF7AAAzMTCQ8KyST"
    "FoXXdsIXcsabN1uNsEICVqiw9g7iMAbCh4mma6sMToAB5tEq9y1gMRNH9U/s6ACZkq0r51TTNGGTjLl+"
    "hxheNg9P47pBWTsvaon9L1cy3xlckePk5+T2Lr8nnEY0DaIkoMs8vmbxJYuutkPXs36nwkoX5a78H40v"
    "choxmjK6/NH4WJBn/VlUwvnNJNx0fK2dEou1/vyQ7xn89kft/JT4F1BLAwQUAAAACABMo3lc6aYluLIF"
    "AABTGwAAEwAAAHhsL3RoZW1lL3RoZW1lMS54bWztWU2P20QYviPxH0a+t44TO82umq022aSFdtvVblrU"
    "48SZ2NOMPdbMZLe5ofaIhIQoiAsSNw4IqNRKXMqvWSiCIvUv8PojyXgz3mbbRRS1OSSe8fN+f/gd5/KV"
    "+xFDh0RIyuO25VysWYjEPh/ROGhbtwf9Cy0LSYXjEWY8Jm1rRqR1ZevDDy7jTRWSiCCgj+UmbluhUsmm"
    "bUsftrG8yBMSw70xFxFWsBSBPRL4CPhGzK7Xak07wjS2UIwjYHtrPKY+QYOUpbU1Z95j8BUrmW74TBz4"
    "mUSdIsOOJk76I2eyywQ6xKxtgZwRPxqQ+8pCDEsFN9pWLftY9tZle0HEVAWtRtfPPgVdQTCa1DM6EQwX"
    "hE7f3bi0s+Bfz/mv4nq9XrfnLPhlAOz7YKmzgnX7Lacz56mB8stV3t2aV3PLeI1/YwW/0el0vI0SvrHE"
    "uyv4Vq3pbtdLeHeJ91b172x3u80S3lvimyv4/qWNplvGZ6CQ0Xiygk7juYjMAjLm7JoR3gJ4a54AS5St"
    "ZVdOH6uqXIvwPS76AMiCixWNkZolZIx9wHVxNBQUpwLwJsHanXzLlytbqSwkfUET1bY+TjBUxBLy8tmP"
    "L589QS+fPT5+8PT4wS/HDx8eP/jZQHgNx4FO+OL7L/7+9lP015PvXjz6yoyXOv73nz777dcvzUClA59/"
    "/fiPp4+ff/P5nz88MsC3BR7q8AGNiEQ3yRHa5xHYZhBAhuJsFIMQ0xIFDgFpAPZUWALenGFmwnVI2Xl3"
    "BDQAE/Dq9F5J14NQTBU1AK+HUQm4yznrcGE053oqSzdnGgdm4WKq4/YxPjTJ7p4IbW+aQCZTE8tuSEpq"
    "7jGINg5ITBRK7/EJIQayu5SW/LpLfcElHyt0l6IOpkaXDOhQmYmu0QjiMjMpCKEu+Wb3DupwZmK/Qw7L"
    "SCgIzEwsCSu58SqeKhwZNcYR05E3sApNSh7MhF9yuFQQ6YAwjnojIqWJ5paYldS9jqETGcO+y2ZRGSkU"
    "nZiQNzDnOnKHT7ohjhKjzjQOdexHcgIpitEeV0YleLlC0jXEAceV4b5DiTpbWd+mQWhOkPTOVBRdu9R/"
    "Ixqf1owZhW78vhnP4dvwaDKVxMkWXIX7HzbeHTyN9wjk+vu++77vvot9t6qW1+22ywZr63Nxxi+qHJLH"
    "lLEDNWPkhsxaswSlR33YzBYZ0WImT0K4LMSVcIHA2TUSXH1CVXgQ4gTEOJmEQBasA4kSLuEkYFXyzo6T"
    "FIzP9rz5GRDQWO3yUb7d0M+GCzbZKpC6oEbKYF1hjUtvJszJgWtKczyzNO9UabbmTagGhNOTv9Os56Ih"
    "YzAjo9TvOYN5WM49RDLEI1LEyDEa4jTWdFvr1V7TpG003kzaOkHSxbkV4rxziFJtJUr2ajmyuLxCR6CV"
    "V/cs5OOkbY1hkoLLKAF+Mm1AmAVx2/JVYcori/mkwea0dGqVBpdEJEKqHSzDnCq7NX91Ei/1r3tu6ofz"
    "McDQjdbTotFy/kMt7JOhJeMx8VXFznJZ3ONTRcRBODpCQzYV+xj0dvPsGlEJz4z6fCGgQt0i8cqVX1TB"
    "yVc0RXVgloS46EktLfY5PLte6JCtNPXsCt1f05TGOZrivbumpJkLY2tjlB2oYAwQGKU52ra4UCGHLpSE"
    "1O8LGBwyWaAXgrJIVUIsfeGc6koOl30r55E3uSBU+zRAgkKnU6EgZE8Vdr6CmVPXn69zRkWfWagrk/x3"
    "SA4JG6TV20ztt1A47yaFIzLcyaDZpuoaBv23ePJxKyaf08eDpSD3LLOIqzV97VGw8WYqnPFRWzdbXPfW"
    "ftQmcPhA6Rc0bip8tpxvB3wfoo8WEyWCRLzQKspvsTkEnVuacSmrf3eMWoagVRHv8xw+NWc3Kpx9urjX"
    "d7Zn8LV3uqvt1RK1tYNMtlr544kP74HsHTgoTZmS+duk+3DU7M7/MgA+9pJ06x9QSwMEFAAAAAgATKN5"
    "XJ4CYaUpAwAAgwkAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWyVlmtv2jAUhv+KlUpVJ1UEAuVW"
    "QCKl26qNqmrVVtqXyhADBsfObKdp+PU7diALaoDyARK/PpfHl/i4lwi5UgtCNPoIGVd9Z6F11HVdNV2Q"
    "EKuKiAiHnpmQIdbQlHNXRZLgwDqFzPWq1aYbYsqdQc9qD3LQE7FmlJMHiVQchlimPmEi6Ts1Zys80vlC"
    "G8Ed9CI8J09EP0cPElpuHiWgIeGKCo4kmfWdYa3r1zzjYC1eKElU4R2phUh+SBr8hswwkKqDzOAmQqxM"
    "911gJLAnjEy1CYrh8U5uCGN9x/eA7K9N43tdv+PkIMa3+L5N+d3OCIxwghW5EeyVBnrRd9oOCsgMx0w/"
    "iuQn2YyyUWmYiFPBlP1HSWbtgfk0VlqEG3czQTplpO/UOw4KKbdSiD82U1XwrXvHfb2Nr501N0tv6UdY"
    "40FPigRJ62govXalerWNmaNDUBsQchvb4UYAcOil3Czzk5bQTSG2HoyG97dofHvz5xmdn7W9mneNkjQi"
    "52d1z7vmdIkiuSYBWolAJJhTEvZcDWjG2Z3CD5ByLm/D5X2Fy7NCbR+WjAGhXb1OOUZDdKETscTfdnPb"
    "OH4Wx7NxzJ7+RFU/har+dSofXcg0wayUqn6MqnEKVSMT9lHB3kAhma7jMpLMt1Hdj3J1CsrVQZRHsZ7D"
    "pKxoGUnmerUfpHkKSPMwSMwDjFz0SzCyXOEynOYxnNYpOK2DOGNKlmpK0EUgQqBK0iVeB6Ubp3WMqn0K"
    "VdsKzT1UrymnKzQsw8gcW/sxOqdgdL6A4ZdhdI5h1KqncBhrZaewFOQ5wfPSnbvxO7AqUOfyJNmJ1NmT"
    "5B6vE4wiRldxFwWFA+UNv72rtx1lUnbYuoWCEBI5txVRoamIuR1zQS3UYVtQ/ptDsYYAL5hReEJx3fHf"
    "7crrrNmaULJHUkQjkXBbsY1wx6NYj4lScC3IxVsphSyKmMGNwmeYr7LzFYpM32FUaUhobisxw7WBA9/H"
    "ZfZ1XHISa4kZT52emxv03F24T4LKridjLOcURsXIDAZVrbRg7WS2K7KGFpHlmAgNOyarqXBLItIYQP9M"
    "CL1tmKnL712Df1BLAwQUAAAACABMo3lccHVMf+IHAAB4MAAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQy"
    "LnhtbKWbbXPaOBDHv4qGzGR6c21syTw2DzOBJC3XtGEIuUz7TgEFXMDibFMXPv1JtnkwSFo5fZEEG++u"
    "9GMl/VcoFwkPp9GEsRj9ns+C6LIyiePFR8eJhhM2p9EZX7BAvPPKwzmNxWU4dqJFyOgoNZrPHOK6dWdO"
    "/aBydZHe64VXF3wZz/yA9UIULedzGq7abMaTywqubG70/fEkljecq4sFHbNHFj8teqG4crZeRv6cBZHP"
    "AxSy18vKNf546zWlQfrEvz5Lor3XKJrw5FPoj+5FZNERt4Jk5144n8q3uyN5SzzPZmwYS6dU/PnFOmw2"
    "k75Fy/7bhLlJm+VsXe+/3oS8S4mIHr7QiHX47NkfxZPLSrOCRuyVLmdxnyefWd7L6llVehzyWZT+Rkn2"
    "NK5W0HAZxXyem8tmxKsZu6x4rQqa+0F6a05/56j2bV3YluS25MDWq8O2Xm7rHdoS2Laa26a9drJup9Ru"
    "aEyvLkKeoDA1lHRI7axV2/jcIhNOU4citnz2Or8h+Ip3/UCm12Mcird94Tu++vJw8x1df+n800WnJ02C"
    "ybn461Ub50P/9ITU8fmYJlN64cSiPdLCGYof0Y5tY0jWGNwEW0LSGxK/uiU0ZmMe+gfBUts2ZMtHCqsO"
    "YPWw8COF2Q1g1gvXq6mARMg5HSEHLRM69g2EvJxQ3ebj8tIb1bom9mPnod/99kmFKLdsaCyJClBuo0sO"
    "McaDFVpTRNBiGqtQ5Q5aqQM5lf2SgX7tPXILxAjXyxit0TjMAhk4VstwrKY3asef4Q5Y1QzMUwHTuj0A"
    "5umAVY+BeQfAgBhFYJ4BWK0MsJruY9oBqwEZ9jdWIdM6zq0kKfQ3Gqbo+s8qaDUYGhBFriw8ocHqPYpD"
    "+upThB18FK1Ar16GXh1OtzpEz1XR0zou0gt8ZgJYh4cpEGgP4HCyehEAXQhgowzABpx+DWC8qtNP6zi3"
    "ski/xjG96gE9IMqGHhXZJ5bVqnf+c0o3qWhA2CyDsAnnYBNCqMxBreMiQiAHm/AgBgJpKGb5aKDYKkOx"
    "BSdiCxjJNyqIWr+Z0Q1/8eMpTfNQpYBuWvAQBkLE/uKDH6AFz4ixwF8aoEn5Y09NPg0lX/6MgZt6BOt9"
    "K9CJPJRpgt5l8/tfKpSbhphS0dyhIilcihSGE2zzjIGUmH1Vo6yjd2+C5RpgYTjxzH0qwiKlYBGLtCJw"
    "Wulgad2/NbOIRWYZ+1SEVap0wEDt0Pn8vd29/da9VVUPGCgfXEdZQGBA3X/brAmmIiJ3UsN71NxDatpA"
    "x9RKFQrYolLAQKngOspiQe/6GI62YNg4McIx9qEIp1RRgC2qAgyUBSJzlAui3neBDkWjbBQq4dQs4Bj7"
    "UIRTSvNjC9GPAdXvypmcqCcnQJEf8tmbzIl2fqpb8DJ2q8irlMTHFhofAyLflZO5hhegwbW8sIGXQvrj"
    "Q17GbhV5ldLz2ELQY0DRC15ExwtQ21pexMBLIfKPxIKxW0VepZQ7bpnXP2VZ0saAmtemG6C2ZVbJrZqs"
    "UElQwsVMP5woqSm0/VGWGeuS4o5sKeVOLJQ7gZS7hpHe99bOlhFRaPbDzDJ3pciolGYnFpqdAJpdN83r"
    "fW/trBlheHY3d6XIqJRUJxZSnQBSHTuemhGg04WdNSOFQj8ca+auFBmVUujEKGJzRtA+vo4RoMRJCUYW"
    "u/rmrhQZldLjxEKPE2jrXscIEOReCUYWG/nmrhQZlZLlxEKWE1CWaxgBstwtwchCnZu7UmRUSp0TC3VO"
    "AHWu2zfQ+97apYwyOOjdgqfSaYkStAjXLEx8ppRPRLFPfzQ52ctzUkqeEwt5TiB5rgMGaHP3rcAaFhlm"
    "r89Jc0ejCYlJ9APdd792B0/K76sBXS5V5b2SFKDKs2IFrdHMn/vxMsXkK9EopPhRLgHBTk+aTbd2Xk2j"
    "MDQKl6cnXtM9X4nPKEHThIZDn5nOBuxll8W+OoE21rXQAC1OrKFZ7LJDwf4UmuduoXkW2twDtLmrg6Z3"
    "vhOeVtA2jkyDEAr2x9DwDhoGR+7gtvP5THmGAhDxckEYKGECKj7bM0YxG04Cf7gOVkqQiq32wyELBdrO"
    "nLtQcxM3suNmIeA9QMC7Wj6Agnet+BCLRAMCleXj7fhAR3MG/evBtTKrAEnf7iuRAYJevKYxRQs/O4o0"
    "VY/NfGudaJx8UBjdQpFfQjpNl+vle7TMW8Ejn45oYPwe1qvuYFbNMO+un1STVtsDlP+dkiQg++/ktlZA"
    "UbhKqMiMd8k64In8fpSJ2Y+Kbiq1x8ZrObRAU17YGlbXRai1HdSaGWr7/uGLEipQKrTvVWYdD6gU2jM+"
    "3R4qXNMXcZmeIEk7qCRaewtRoBVBiGTk5c/sOOOQjblcUfjsDPVMWOs7rHUz1l7/h1hPnr+rJr62B5QY"
    "jwP14gxUGD2hioeTZBVD47/+FqRAcIF0sYmv4vpg4trYcW3AXPvP6hkVqEN6SqZAEdJLKw0qhz2dDn+q"
    "eTbewhMIvKbjMA0b8HFG8708TLMcycGyJW2CuqtbPKBuGXS/3j48qVMVqFnUCztQQgz8OfvANeO9+RaY"
    "QMCsXqRyPs8kZEZ0lAKOQxawULlGOXvnrOcsHKcH3CM05MsgrZb37qKD8+67x7Mz+V9pOPaDCM3YqzB1"
    "zxpiigqzqju7iPkiLcBfeCwq8uxAN6MjFsoHxPuvnMebCxlg+88GV/8DUEsDBBQAAAAIAEyjeVxMKdtC"
    "iSAAAAaZAQAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDMueG1svd1tb1tHevDx9/kUghdYbFs3Np/JrGNg"
    "Pc8zOTODTdoF+k6x6ViNbLqSUjf76W9K9vGdc3auvzN9USCxbP54HVGc4Tj4Q4yefTjd/Hz75ni8u/if"
    "t9fvbr999Obu7v03T57cvnxzfHt5+/Xp/fHdWV6fbt5e3p3/ePPTk9v3N8fLVw9Db6+fLJ8+3T55e3n1"
    "7tHzZw+31Zvnz06/3F1fvTvWm4vbX96+vbz59cXx+vTh20eLR+MNf7366c3d/Q1Pnj97f/nT8fvj3b+9"
    "rzfnPz35fJVXV2+P726vTu8ubo6vv330l8U3dfn0YeLhLv9+dfxw+5vfX9y+OX1wN1evvjt/6vNX8vTR"
    "xf1X9+Pp9PM9h1f3N91/tnfHi1+/f3999fD5L+5O7787vr5Tx+vr8+dYP7q4fHl39d/Her7bt49+PN3d"
    "nd7e+/mR313enW96fXP6+/Hdw6M4Xh/P9z0/vvf/cOePF/l4Ubs5D//Xw9dw/u3nL/H+Qf329+PXYh+e"
    "6/Nz9+Pl7VGdrv929eruzbeP9o8uXh1fX/5yfffX0wd//PT8rb9e31/x5en69uHXiw8f7314dPHyl9vz"
    "w/k0ff/M3/16fX6UqzO9vXr3cNPby//5tAa/Gd1/eXT5aXQ5G13svt7vt7vFfnf+kn883t7ZT0/yly64"
    "+nTB1fyCT788u/40u+7/OjbjU/B0Nis/fev1+PR9fv7+F0/gYnwGF//wFP6Or3gxPl2Lf/iaf/P8S1/A"
    "ZrzK56/+46b8uIcetqC+vLt8/uzm9OHi5mH0fqutPj+wz5vvfM2HR3V+CPd3/MvHG5bnh3nWq3f3R8D3"
    "dzdnvjpf+O55+nB5c3f51Z/++duLu6v3/3p6/fqfnj25O3/me37y8tNVXnzhKurvl7df/en2+PPXrWn1"
    "pcdwenVx+fPL/7xqzOpPX9DT9uz9CfnN7fvLl+cn8XwE3h5v/vv46Pn5DudX5VeNy5lPl1sID+U/Lj+c"
    "Xr27+vmrRWPY/t7hZWPY/d7hVWPY/97hdWM4/N7hTWM4fhpeCsM/nP8+OP/N8tWffvj27vLn1uqnj1dY"
    "S58+nf74h9Vy/eeXf//1q/td8Mc/LPeLP1/86d1N62rffXo8K+Fqf7n99fx3wsUf/3DYrVZ//kq4yvCF"
    "q/zHj1fnv6UuL4q1X7pU/p2X0uaLlyqfLrUVLlVv/n58+ebDr3cX3//w3ZcuVr9wsRfXp58vXnyX4DpP"
    "zofN5xNn+fHEWWy/frr54qGz/HjD+uFz3/9XyP8/SWRSMmmZjExWJieTlynIFGVKH2m9/Ef6Tp4aZMoy"
    "FZlqkyZrvOpZ49XHGzaNNZZJyaRlMjJZmZxMXqYgU5QpfaT1qrHG8tQgU5apyFSbNFnjdc8ar+U1Xsuv"
    "Y5m0TEYmK5OTycsUZIoypbX8OpanBpmyTEWm2qTJGm961ngjr7FMSiYtk5HJyuRk8jIFmaJMaSO/juWp"
    "QaYsU5GpNmmyxtueNd7Ka7yVX8cyaZmMTFYmJ5OXKcgUZUpb+XUsTw0yZZmKTLVJkzXe9azxTl5jmZRM"
    "WiYjk5XJyeRlCjJFmdJOfh3LU4NMWaYiU23SZI33PWu8l9d4L7+OZdIyGZmsTE4mL1OQKcqU9vLrWJ4a"
    "ZMoyFZlqkyZrfOhZ44O8xjIpmbRMRiYrk5PJyxRkijKlg/w6lqcGmbJMRabapMka3zfC37/I9/eWVnm0"
    "1ksZTIMZMAvmwDxYAItg6ZM1X9MwN4BlsAJW2zZd9EXXoi9g0WVTYBrMgFkwB+bBAlgES5+s+SKHuQEs"
    "gxWw2rbponelsMUSFh1iGJgGM2AWzIF5sAAWwdICqhjMDWAZrIDVtk0XvauNLSCOgSkwDWbALJgD82AB"
    "LIKlBWQymBvAMlgBq22bLnpXLFtALRut+UqHXgZmwCyYA/NgASyCpQV0M5gbwDJYAattmy56Vz1bQD4D"
    "U2AazIBZMAfmwQJYBEsLCGkwN4BlsAJW2zZd9K6ctoCetoCgBqbBDJgFc2AeLIBFsLSAsgZzA1gGK2C1"
    "bdNF7+prCwhsYApMgxkwC+bAPFgAi2BpAakN5gawDFbAatumi94V3BZQ3BaQ3MA0mAGzYA7MgwWwCJYW"
    "0N5gbgDLYAWstm266F0FbgEJDkyBaTADZsEcmAcLYBEsLSDGwdwAlsEKWG3b9DtWuorcEorcEoocmAYz"
    "YBbMgXmwABbB0hKKHMwNYBmsgNW2TRe9q8gtociBKTANZsAsmAPzYAEsgqUlFDmYG8AyWAGrbZsuet83"
    "p0GRW0KRA9NgBsyCOTAPFsAiWFrS96nRN6rRd6rRt6rR96p9ucgtu4rcEoocmALTYAbMgjkwDxbAIlha"
    "QpGDuQEsgxWw2rbponcVuSUUOTA1WvOVDkUOzII5MA8WwCJYWkKRg7kBLIMVsNq26aJ3FbklVDcwDWZG"
    "ay6sbA7MgwWwCI8zLaG6wdwAlsEKWG3bdGG7qttSrj0vwBSYBjNgFsyBebAAFsHSEqobzA1gGayA1bZN"
    "F72rui2huoEpMA1mwCyYA/NgASyCpSVUN5gbwDJYAattmy56V3VbyrXnBZgC02AGzII5MA8WwCJYWkJ1"
    "g7kBLIMVsNq26aJ3VbclVDcwBabBDJgFc2AeLIBFsLSE6gZzA1gGK2C1bdP3EHVVt5Vce16AKTANZsAs"
    "mAPzYAEsgqUVVDeYG8AyWAGrbZsueld1W0F1A1NgGsyAWTAH5sECWARLK6huMDeAZbACVts2XfSu6raC"
    "94SCKTANZsAsmAPzYAEsgqUVVDeYG8AyWAGrbZsuet97ROlNovQuUXqbKL1PlN4oSu8UpbeK0ntF6c2i"
    "9G5RersovV+U3jBK7xilt4x+ubqtuqrbSq49L8AUmAYzYBbMgXmwABbB0gqqG8wNYBmsgNW2TRe9q7qt"
    "5BL0AkyBaTADZsEcmAcLYBEsraDIwdwAlsEKWG3bdNG7itwKihyYAtNgBsyCOTAPFsAiWFpBkYO5ASyD"
    "FbDatumidxW5FRQ5MAWmwQyYBXNgHiyARbC0giIHcwNYBitgtW3TRe8qcisocmAKTIMZMAvmwDxYAItg"
    "aQVFDuYGsAxWwGrbpoveVeRWUOTAFJgGM2AWzIF5sAAWwdIKihzMDWAZrIDVtk3/jy9dRW4NRQ5MgWkw"
    "A2bBHJgHC2ARLK2hyMHcAJbBClht23TRu4rcGoocmALTYAbMgjkwDxbAIlhaQ5GDuQEsgxWw2rbponcV"
    "uTUUOTAFpsEMmAVzYB4sgEWwtIYiB3MDWAYrYLVt00XvKnJrKHJgCkyDGTAL5sA8WACLYGkNRQ7mBrAM"
    "VsBq26aL3lXk1lDkwBSYBjNgFsyBebAAFsHSGooczA1gGayA1bZNF72ryK2hyIEpMA1mwCyYA/NgASyC"
    "pTUUOZgbwDJYAattmy56V5FbQ5EDU2AazIBZMAfmwQJYBEtrKHIwN4BlsAJW2zZd9K4it4YiB6bANJgB"
    "s2AOzIMFsAiW1lDkYG4Ay2AFrLZtuuhdRW4NRQ5MgWkwA2bBHJgHC2ARLK2hyMHcAJbBClht23TRu4rc"
    "GoocmALTYAbMgjkwDxbAIlhaQ5GDuQEsgxWw2rbp/5+3q8htoMiBKTANZsAsmAPzYAEsgqUNFDmYG8Ay"
    "WAGrbZsueleR20CRA1NgGsyAWTAH5sECWARLGyhyMDeAZbACVts2XfSuIreBIgemwDSYAbNgDsyDBbAI"
    "ljZQ5GBuAMtgBay2bbroXUVuA0UOTIFpMANmwRyYBwtgESxtoMjB3ACWwQpYbdt00buK3AaKHJgC02AG"
    "zII5MA8WwCJY2kCRg7kBLIMVsNq26aJ3FbkNFDkwBabBDJgFc2AeLIBFsLSBIgdzA1gGK2C1bdNF7ypy"
    "GyhyYApMgxkwC+bAPFgAi2BpA0UO5gawDFbAatumi95V5DZQ5MAUmAYzYBbMgXmwABbB0gaKHMwNYBms"
    "gNW2TRe9q8htoMiBKTANZsAsmAPzYAEsgqUNFDmYG8AyWAGrbZsueleR20CRA1NgGsyAWTAH5sECWARL"
    "GyhyMDeAZbACVts2/WlKXUVuC0UOTIFpMANmwRyYBwtgESxtocjB3ACWwQpYbdt00buK3BaKHJgC02AG"
    "zII5MA8WwCJY2kKRg7kBLIMVsNq26aJ3FbktFDkwBabBDJgFc2AeLIBFsLSFIgdzA1gGK2C1bdNF7ypy"
    "WyhyYApMgxkwC+bAPFgAi2BpC0UO5gawDFbAatumi95V5LZQ5MAUmAYzYBbMgXmwABbB0haKHMwNYBms"
    "gNW2TRe9q8htociBKTANZsAsmAPzYAEsgqUtFDmYG8AyWAGrbZsuet8PQ4UiB6bANJgBs2AOzIMFsAiW"
    "tvRzUekHo9JPRqUfjUo/G/XLRW7bVeS2UOTAFJgGM2AWzIF5sAAWwdIWihzMDWAZrIDVtk0XvavIbaHI"
    "gSkwDWbALJgD82ABLIKlLRQ5mBvAMlgBq22bLnpXkdtCkQNTYBrMgFkwB+bBAlgES1socjA3gGWwAlbb"
    "Nv3Z111FbgdFDkyBaTADZsEcmAcLYBEs7aDIwdwAlsEKWG3bdNG7itwOihyYAtNgBsyCOTAPFsAiWNpB"
    "kYO5ASyDFbDatumidxW5HRQ5MAWmwQyYBXNgHiyARbC0gyIHcwNYBitgtW3TRe8qcjsocmAKTIMZMAvm"
    "wDxYAItgaQdFDuYGsAxWwGrbpoveVeR2UOTAFJgGM2AWzIF5sAAWwdIOihzMDWAZrIDVtk0XvavI7aDI"
    "gSkwDWbALJgD82ABLIKlHRQ5mBvAMlgBq22bLnpXkdtBkQNTYBrMgFkwB+bBAlgESzsocjA3gGWwAlbb"
    "Nl30riK3gyIHpsA0mAGzYA7MgwWwCJZ2UORgbgDLYAWstm266F1FbgdFDkyBaTADZsEcmAcLYBEs7aDI"
    "wdwAlsEKWG3bdNG7itwOihyYAtNgBsyCOTAPFsAiWNpBkYO5ASyDFbDatsmi77uK3B6KHJgC02AGzII5"
    "MA8WwCJY2kORg7kBLIMVsNq26aJ3Fbk9FDkwBabBDJgFc2AeLIBFsLSHIgdzA1gGK2C1bdNF7ypyeyhy"
    "YApMgxkwC+bAPFgAi2BpD0UO5gawDFbAatumi95V5PZQ5MAUmAYzYBbMgXmwABbB0h6KHMwNYBmsgNW2"
    "TRe9q8jtociBKTANZsAsmAPzYAEsgqU9FDmYG8AyWAGrbZsueleR20ORA1NgGsyAWTAH5sECWARLeyhy"
    "MDeAZbACVts2XfSuIreHIgemwDSYAbNgDsyDBbAIlvZQ5GBuAMtgBay2bbroXUVuD0UOTIFpMANmwRyY"
    "BwtgESztocjB3ACWwQpYbdt00buK3B6KHJgC02AGzII5MA8WwCJY2kORg7kBLIMVsNq26aJ3Fbk9FDkw"
    "BabBDJgFc2AeLIBFsLSHIgdzA1gGK2C1bZNFP3QVuQMUOTAFpsEMmAVzYB4sgEWwdIAiB3MDWAYrYLVt"
    "00XvKnIHKHJgCkyDGTAL5sA8WACLYOkARQ7mBrAMVsBq26aL3lXkDlDkwBSYBjNgFsyBebAAFsHSAYoc"
    "zA1gGayA1bZNF72ryB2gyIEpMA1mwCyYA/NgASyCpQMUOZgbwDJYAattmy56V5E7QJEDU2AazIBZMAfm"
    "wQJYBEsHKHIwN4BlsAJW2zZd9K4id4AiB6bANJgBs2AOzIMFsAiWDlDkYG4Ay2AFrLZtuuhdRe4ARQ5M"
    "gWkwA2bBHJgHC2ARLB2gyMHcAJbBClht23TRu4rcAYocmALTYAbMgjkwDxbAIlg6QJGDuQEsgxWw2rbp"
    "oncVuQMUOTAFpsEMmAVzYB4sgEWwdIAiB3MDWAYrYLVt00XvKnIHKHJgCkyDGTAL5sA8WACLYOkARQ7m"
    "BrAMVsBq2yaLvnjaleQe7i691gkVoSY0hJbQEXrCQBgJ04jNVz1NDoSZsBBWAWe7oKvRPdxdevETKkJN"
    "aAgtoSP0hIEwEqYRm8cATQ6EmbAQVgFnu6Ar2j3cXT4LINsRakJDaAkdoScMhJEwjSicBRDwCDNhIawC"
    "znZBV8V7uLt8FkDHI9SEhtASOkJPGAgjYRpROAug6BFmwkJYBZztgq6s93B3+SyAsEeoCQ2hJXSEnjAQ"
    "RsI0onAWQOIjzISFsAo42wVdne/h7vJZAKWPUBMaQkvoCD1hIIyEaUThLIDmR5gJC2EVcLYLusLfw93l"
    "swDSH6EmNISW0BF6wkAYCdOIwlkAEZAwExbCKuBsF3SVwIe7y2cBtEBCTWgILaEj9ISBMBKmEYWzAKog"
    "YSYshFXA2S7oSoMPd5fPAoiDhJrQEFpCR+gJA2EkTCMKZwFkQsJMWAirgLNd0NUKH+4unwVQCwk1oSG0"
    "hI7QEwbCSJhGFM4C6IaEmbAQVgGnu2DR1w4X1A4BFaEmNISW0BF6wkAYCdOI7bMAJgfCTFgIq4CzXdDX"
    "DhfUDgEVoSY0hJbQEXrCQBgJ04jtswAmB8JMWAirgLNd0NcOF9QOARWhJjSEltAResJAGAnTiMJZQO0Q"
    "MBMWwirgbBf0tcMFtUNARagJDaEldISeMBBGwjSicBZQOwTMhIWwCjjbBX3tcEHtEFARakJDaAkdoScM"
    "hJEwjSicBdQOATNhIawCznZBXztcUDsEVISa0BBaQkfoCQNhJEwjCmcBtUPATFgIq4CzXdDXDhfUDgEV"
    "oSY0hJbQEXrCQBgJ04jCWUDtEDATFsIq4GwX9LXDBbVDQEWoCQ2hJXSEnjAQRsI0onAWUDsEzISFsAo4"
    "2wV97XBB7RBQEWpCQ2gJHaEnDISRMI0onAXUDgEzYSGsAs52QV87XFA7BFSEmtAQWkJH6AkDYSRMIwpn"
    "AbVDwExYCKuA012w7GuHS2qHgIpQExpCS+gIPWEgjIRpxPZZAJMDYSYshFXA2S7oa4dLaoeAilATGkJL"
    "6Ag9YSCMhGnE9lkAkwNhJiyEVcDZLuhrh0tqh4CKUBMaQkvoCD1hIIyEaUThLKB2CJgJC2EVcLYL+trh"
    "ktohoCLUhIbQEjpCTxgII2EaUTgLqB0CZsJCWAWc7YK+drikdgioCDWhIbSEjtATBsJImEYUzgJqh4CZ"
    "sBBWAWe7oK8dLqkdAipCTWgILaEj9ISBMBKmEYWzgNohYCYshFXA2S7oa4dLaoeAilATGkJL6Ag9YSCM"
    "hGlE4SygdgiYCQthFXC2C/ra4ZLaIaAi1ISG0BI6Qk8YCCNhGlE4C6gdAmbCQlgFnO2Cvna4pHYIqAg1"
    "oSG0hI7QEwbCSJhGFM4CaoeAmbAQVgFnu6CvHS6pHQIqQk1oCC2hI/SEgTASphGFs4DaIWAmLIRVwOku"
    "WPW1wxW1Q0BFqAkNoSV0hJ4wEEbCNGL7LIDJgTATFsIq4GwX9LXDFbVDQEWoCQ2hJXSEnjAQRsI0Yvss"
    "gMmBMBMWwirgbBf0tcMVtUNARagJDaEldISeMBBGwjSicBZQOwTMhIWwCjjbBX3tcEXtEFARakJDaAkd"
    "oScMhJEwjSicBdQOATNhIawCznZBXztcUTsEVISa0BBaQkfoCQNhJEwjCmcBtUPATFgIq4CzXdDXDlfU"
    "DgEVoSY0hJbQEXrCQBgJ04jCWUDtEDATFsIq4GwX9LXDFbVDQEWoCQ2hJXSEnjAQRsI0onAWUDsEzISF"
    "sAo42wV97XBF7RBQEWpCQ2gJHaEnDISRMI0onAXUDgEzYSGsAs52QV87XFE7BFSEmtAQWkJH6AkDYSRM"
    "IwpnAbVDwExYCKuAs13Q1w5X1A4BFaEmNISW0BF6wkAYCdOIwllA7RAwExbCKuB0F6z72uGa2iGgItSE"
    "htASOkJPGAgjYRqxfRbA5ECYCQthFXC2C/ra4ZraIaAi1ISG0BI6Qk8YCCNhGrF9FsDkQJgJC2EVcLYL"
    "+trhmtohoCLUhIbQEjpCTxgII2EaUTgLqB0CZsJCWAWc7YK+drimdgioCDWhIbSEjtATBsJImEYUzgJq"
    "h4CZsBBWAWe7oK8drqkdAipCTWgILaEj9ISBMBKmEYWzgNohYCYshFXA2S7oa4draoeAilATGkJL6Ag9"
    "YSCMhGlE4SygdgiYCQthFXC2C/ra4ZraIaAi1ISG0BI6Qk8YCCNhGlE4C6gdAmbCQlgFnO2Cvna4pnYI"
    "qAg1oSG0hI7QEwbCSJhGFM4CaoeAmbAQVgFnu6CvHa6pHQIqQk1oCC2hI/SEgTASphGFs4DaIWAmLIRV"
    "wNku6GuHa2qHgIpQExpCS+gIPWEgjIRpROEsoHYImAkLYRVwugs2fe1wQ+0QUBFqQkNoCR2hJwyEkTCN"
    "2D4LYHIgzISFsAo42wV97XBD7RBQEWpCQ2gJHaEnDISRMI3YPgtgciDMhIWwCjjbBX3tcEPtEFARakJD"
    "aAkdoScMhJEwjSicBdQOATNhIawCznZBXzvcUDsEVISa0BBaQkfoCQNhJEwjCmcBtUPATFgIq4CzXdDX"
    "DjfUDgEVoSY0hJbQEXrCQBgJ04jCWUDtEDATFsIq4GwX9LXDDbVDQEWoCQ2hJXSEnjAQRsI0onAWUDsE"
    "zISFsAo42wV97XBD7RBQEWpCQ2gJHaEnDISRMI0onAXUDgEzYSGsAs52QV873FA7BFSEmtAQWkJH6AkD"
    "YSRMIwpnAbVDwExYCKuAs13Q1w431A4BFaEmNISW0BF6wkAYCdOIwllA7RAwExbCKuBsF/S1ww21Q0BF"
    "qAkNoSV0hJ4wEEbCNKJwFlA7BMyEhbAKON0F2752uKV2CKgINaEhtISO0BMGwkiYRmyfBTA5EGbCQlgF"
    "nO2Cvna4pXYIqAg1oSG0hI7QEwbCSJhGbJ8FMDkQZsJCWAWc7YK+drildgioCDWhIbSEjtATBsJImEYU"
    "zgJqh4CZsBBWAWe7oK8dbqkdAipCTWgILaEj9ISBMBKmEYWzgNohYCYshFXA2S7oa4dbaoeAilATGkJL"
    "6Ag9YSCMhGlE4SygdgiYCQthFXC2C/ra4ZbaIaAi1ISG0BI6Qk8YCCNhGlE4C6gdAmbCQlgFnO2Cvna4"
    "pXYIqAg1oSG0hI7QEwbCSJhGFM4CaoeAmbAQVgFnu6CvHW6pHQIqQk1oCC2hI/SEgTASphGFs4DaIWAm"
    "LIRVwNku6GuHW2qHgIpQExpCS+gIPWEgjIRpROEsoHYImAkLYRVwtgv62uGW2iGgItSEhtASOkJPGAgj"
    "YRpROAuoHQJmwkJYBZzugl1fO9xROwRUhJrQEFpCR+gJA2EkTCO2zwKYHAgzYSGsAs52QV873FE7BFSE"
    "mtAQWkJH6AkDYSRMI7bPApgcCDNhIawCznZBXzvcUTsEVISa0BBaQkfoCQNhJEwjCmcBtUPATFgIq4Cz"
    "XdDXDnfUDgEVoSY0hJbQEXrCQBgJ04jCWUDtEDATFsIq4GwX9LXDHbVDQEWoCQ2hJXSEnjAQRsI0onAW"
    "UDsEzISFsAo42wV97XBH7RBQEWpCQ2gJHaEnDISRMI0onAXUDgEzYSGsAs52QV873FE7BFSEmtAQWkJH"
    "6AkDYSRMIwpnAbVDwExYCKuAs13Q1w531A4BFaEmNISW0BF6wkAYCdOIwllA7RAwExbCKuBsF/S1wx21"
    "Q0BFqAkNoSV0hJ4wEEbCNKJwFlA7BMyEhbAKONsFfe1wR+0QUBFqQkNoCR2hJwyEkTCNKJwF1A4BM2Eh"
    "rAJOd8G+rx3uqR0CKkJNaAgtoSP0hIEwEqYR22cBTA6EmbAQVgFnu6CvHe6pHQIqQk1oCC2hI/SEgTAS"
    "phHbZwFMDoSZsBBWAWe7oK8d7qkdAipCTWgILaEj9ISBMBKmEYWzgNohYCYshFXA2S7oa4d7aoeAilAT"
    "GkJL6Ag9YSCMhGlE4SygdgiYCQthFXC2C/ra4Z7aIaAi1ISG0BI6Qk8YCCNhGlE4C6gdAmbCQlgFnO2C"
    "vna4p3YIqAg1oSG0hI7QEwbCSJhGFM4CaoeAmbAQVgFnu6CvHe6pHQIqQk1oCC2hI/SEgTASphGFs4Da"
    "IWAmLIRVwNku6GuHe2qHgIpQExpCS+gIPWEgjIRpROEsoHYImAkLYRVwtgv62uGe2iGgItSEhtASOkJP"
    "GAgjYRpROAuoHQJmwkJYBZztgr52uKd2CKgINaEhtISO0BMGwkiYRhTOAmqHgJmwEFYBp7vg0NcOD9QO"
    "ARWhJjSEltAResJAGAnTiO2zACYHwkxYCKuAs13Q1w4P1A4BFaEmNISW0BF6wkAYCdOI7bMAJgfCTFgI"
    "q4CzXdDXDg/UDgEVoSY0hJbQEXrCQBgJ04jCWUDtEDATFsIq4GwX9LXDA7VDQEWoCQ2hJXSEnjAQRsI0"
    "onAWUDsEzISFsAo42wV97fBA7RBQEWpCQ2gJHaEnDISRMI0onAXUDgEzYSGsAs52QV87PFA7BFSEmtAQ"
    "WkJH6AkDYSRMIwpnAbVDwExYCKuAs13Q1w4P1A4BFaEmNISW0BF6wkAYCdOIwllA7RAwExbCKuBsF/S1"
    "wwO1Q0BFqAkNoSV0hJ4wEEbCNKJwFlA7BMyEhbAKONsFfe3wQO0QUBFqQkNoCR2hJwyEkTCNKJwF1A4B"
    "M2EhrALOdkFfOzxQOwRUhJrQEFpCR+gJA2EkTCMKZwG1Q8BMWAirgJNdsHza1Q4f7i6dBYSKUBMaQkvo"
    "CD1hIIyEacTmWUCTA2EmLIRVwNku6GqHD3eXzgJCRagJDaEldISeMBBGwjRi8yygyYEwE5bP2N4FTfy4"
    "C57cvjke7/Tl3eXzZ6/Ov/775fXV+ePV6d3txcvTL+/OA6tHc7q4/a+b4+v73fDNpw3x5vRB35ze69OH"
    "d98+evrxhvDu/S93w/H29vKn4+cbzc3N6ea3N15eX58+vLi+fPfzw267+/X9+fbrq9u782d9fbp5+8v1"
    "5eL5o8XjxT8/Xj5ePV4/evbk883Pnkwfl/Q41fIb9X/zOO8f4vJfFud/nz5enT+uzh+X+vzP/W36X54+"
    "Wfzt4Q/3H58+WZ7/Xd1/1I+f3uPyb/cfF/cfH35ZPtzy8Y8PA+dfXvz1sX384rv0+Psfvntc/zfPRlx+"
    "E/9vno0f+OHNbrh9/uz9+frD5c1PV+fdd318fd58T7++f2fdzcfT6OMf7k7vHz7pj6e780n18Ns3x8tX"
    "x5v7O5z99el0N/7hyfmzfDjd/Pywy5//P1BLAwQUAAAACABMo3lc/jUQcsgdAACujAEAGAAAAHhsL3dv"
    "cmtzaGVldHMvc2hlZXQ0LnhtbL3da28b552G8ff5FIICFNnWG4mkjqltoH7O50GTbYG+U2w61loWXYmp"
    "m3z6pejDZqbzv+ynLwoklsUfZ0jxFuXgAh0+fre5e33/ar3eHvzzzc3t/ZPDV9vt2++Oju6fv1q/ubr/"
    "dvN2fbuTl5u7N1fb3ad3Px3dv71bX73YH/Tm5mh5fHx29Obq+vbw6eP9ZcPd08ebn7c317fr4e7g/uc3"
    "b67ufnm2vtm8e3K4OPx4wZ+vf3q1fbjg6Onjt1c/rb9fb//n7XC3++zo01leXL9Z395fb24P7tYvnxz+"
    "afHdsDzeH7G/yl+u1+/uf/P7g/tXm3fu7vpF3t307is5PjzYXv34/fpm/Xy7frG/8Yev9sfN5vXD1cOL"
    "h6s83Prt+uCX79/eXO/vz8F28zavX27V+uZmd5vLw4Or59vrf6yH3dWeHP642W43bx5895Vsr7a7i17e"
    "bX5d3+7v1f6mHu7v23+58vuTvD+pXTw8Dn/ff1EPv//0RT/crd/+/uNXZ/eP/u7R/PHqfq02N3+9frF9"
    "9eTw4vDgxfrl1c832z9v3vn1h0f05NuThzM+39zc7389ePf+2peHB89/vt/doQ9HP9yH7S83u/u52tGb"
    "69v9RW+u/vlhld8cevH5Q5cfDl1ODl0sP3/s6sOxq+mxx58/9uTDsSf9d/n041d7PDn2Sx6pTw/Vv/FY"
    "LT4+WIt/ebQWX3D0x4dr8e88XouPD9hi+oh9yVf96SE73X/Tvv8e23+L6qvt1dPHd5t3B3f7Qx++FVef"
    "7s+nb87dOffn293ywxX/9P6C5e4Gdnp9+/BD4/vt3Y6vdyfePk3vru62V1998/snB9vrt/+9efnyvx4f"
    "bXe3/MBHzz+c5dlnzqJ+vbr/6pv79etv545Wn7sPmxcHV6+f/+/1zLH6wxd0PH/sw8/U7+7fXj3fPYi7"
    "H5r367t/rA+f7q6we9Z+NXM68+F0C+Gu/O3q3ebF7fXrrxYzB9svPXg5c7D70oNXMwf7Lz34ZObg8KUH"
    "n84cHD8cvBQO/mH3J8juz6Kvvvnhyfbq9dz66cMZVuL6v/t6tTz54/Nff/nq4bvgd18vLxZ/PPjm9m7u"
    "bPkzZ/vT/S+7PzUOfvf15flq9cevhLOUz5zlbz9e7/4cuzpo1n7uVPULT6XNZ0/VPpzqTDjVcPfr+vmr"
    "d79sD77/IX/uZMNnTvbsZvP64FlOcJ6j3Q+bTz9xlu9/4izOvj0+/ewPneX7C072t/3w3y3//5NEJiWT"
    "lsnIZGVyMnmZgkxRpiRTlqnIVGVqMg2zNNp41bPxSt74A53ObCyTlsnIZGVyMnmZgkxRpiRTlqnIVGVq"
    "Mg2zNNr4pGfjE3ljmZRMWiYjk5XJyeRlCjJFmZJMWaYiU5WpyTTM0mjj056NT+WNT+XnsUxaJiOTlcnJ"
    "5GUKMkWZkkxZpiJTlanJNMzSaOOzno3P5I1lUjJpmYxMViYnk5cpyBRlSjJlmYpMVaYm0zBLo43PezY+"
    "lzc+l5/HMmmZjExWJieTlynIFGVKMmWZikxVpibTMEujjS96Nr6QN5ZJyaRlMjJZmZxMXqYgU5QpyZRl"
    "KjJVmZpMwyyNNr7s2fhS3vhSfh7LpGUyMlmZnExepiBTlCnJlGUqMlWZmkzDLI02fkiDXz7yw7WllcEU"
    "mAYzYBbMgXmwABbBElgGK2AVrIEN8zYefdE1+gJGX8jPbTANZsAsmAPzYAEsgiWwDFbAKlgDG+ZtPHpX"
    "CltACwNTYBrMfLTZ0WVzYB4sgEW4nwksgxWwCtbAhnkbj97VxhYQxxZQx8A0mAGzYA7MgwWwCJbAMlgB"
    "q2ANbJi38ehdsWwBtQxMgWkwA2bBHJgHC2ARLIFlsAJWwRrYMG/j0bvq2QLy2UebfaZDQAMzYBbMgXmw"
    "ABbBElgGK2AVrIEN8zYevSunLaCngSkwDWbALJgD82ABLIIlsAxWwCpYAxvmbTx6V19bQGBbQGED02AG"
    "zII5MA8WwCJYAstgBayCNbBh3sajdwW3BRQ3MAWmwQyYBXNgHiyARbAElsEKWAVrYMO8jUfvKnALSHAL"
    "aHBgGsyAWTAH5sECWARLYBmsgFWwBjbM2/gVK11FbglFDkyBaTADZsEcmAcLYBEsgWWwAlbBGtgwb+PR"
    "u4rcEorcEoocmAYzYBbMgXmwABbBElgGK2AVrIEN8zYeve/FafTqNHp5Gr0+jV6gRq9Qo5eo0WvU6EVq"
    "9Co1epkavU6NXqhGr1Sjl6rRa9U+X+SWXUVuCUVuCUUOTIMZMAvmwDxYAItgCSyDFbAK1sCGeRuP3lXk"
    "llDkwBSYBjNgFsyBebAAFsESWAYrYBWsgQ3zNh69q8gt5RL0DEyBaTADZsEcmAcLYBEsgWWwAlbBGtgw"
    "b+PRu4rcEoocmALTYAbMgjkwDxbAIlgCy2AFrII1sGHexqN3FbmlXIKegSkwDWbALJgD82ABLIIlsAxW"
    "wCpYAxvmbTx6V5FbQpEDU2AazIBZMAfmwQJYBEtgGayAVbAGNszbePSuIreUS9AzMAWmwQyYBXNgHiyA"
    "RbAElsEKWAVrYMO8jf9+UVeRW0GRA1NgGsyAWTAH5sECWARLYBmsgFWwBjbM23j0riK3kkvQMzAFpsEM"
    "mAVzYB4sgEWwBJbBClgFa2DDvI1H7ypyKyhyYApMgxkwC+bAPFgAi2AJLIMVsArWwIZ5G4/e9/dH5RL0"
    "DEyBaTADZsEcmAcLYBEsgWWwAlbBGtgwb+PRu4rcCoocmALTYAbMgjkwDxbAIlgCy2AFrII1sGHexqN3"
    "FbkVFDkwBabBDJgFc2AeLIBFsASWwQpYBWtgw7yNR+8qcisocmAKTIMZMAvmwDxYAItgCSyDFbAK1sCG"
    "eRuP3lXkVlDkwBSYBjNgFsyBebAAFsESWAYrYBWsgQ3zNh69q8itoMiBKTANZsAsmAPzYAEsgiWwDFbA"
    "KlgDG+ZtPHpXkVtBkQNTYBrMgFkwB+bBAlgES2AZrIBVsAY2zNv4/wbTVeROoMiBKTANZsAsmAPzYAEs"
    "giWwDFbAKlgDG+ZtPHpXkTuBIgemwDSYAbNgDsyDBbAIlsAyWAGrYA1smLfx6F1F7gSKHJgC02AGzII5"
    "MA8WwCJYAstgBayCNbBh3sajdxW5EyhyYApMgxkwC+bAPFgAi2AJLIMVsArWwIZ5G4/eVeROoMiBKTAN"
    "ZsAsmAPzYAEsgiWwDFbAKlgDG+ZtPHpXkTuBIgemwDSYAbNgDsyDBbAIlsAyWAGrYA1smLfx6F1F7gSK"
    "HJgC02AGzII5MA8WwCJYAstgBayCNbBh3sajdxW5EyhyYApMgxkwC+bAPFgAi2AJLIMVsArWwIZ5G4/e"
    "VeROoMiBKTANZsAsmAPzYAEsgiWwDFbAKlgDG+ZtPHpXkTuBIgemwDSYAbNgDsyDBbAIlsAyWAGrYA1s"
    "mLfx/7u3q8idQpEDU2AazIBZMAfmwQJYBEtgGayAVbAGNszbePSuIncKRQ5MgWkwA2bBHJgHC2ARLIFl"
    "sAJWwRrYMG/j0buK3CkUOTAFpsEMmAVzYB4sgEWwBJbBClgFa2DDvI1H7ypyp1DkwBSYBjNgFsyBebAA"
    "FsESWAYrYBWsgQ3zNh69q8idQpEDU2AazIBZMAfmwQJYBEtgGayAVbAGNszbePSuIncKRQ5MgWkwA2bB"
    "HJgHC2ARLIFlsAJWwRrYMG/j0buK3CkUOTAFpsEMmAVzYB4sgEWwBJbBClgFa2DDvI1H7ypyp1DkwBSY"
    "BjNgFsyBebAAFsESWAYrYBWsgQ3zNh69q8idQpEDU2AazIBZMAfmwQJYBEtgGayAVbAGNszbePSuIncK"
    "RQ5MgWkwA2bBHJgHC2ARLIFlsAJWwRrYMG/jd1rqKnJnUOTAFJgGM2AWzIF5sAAWwRJYBitgFayBDfM2"
    "Hr2ryJ1BkQNTYBrMgFkwB+bBAlgES2AZrIBVsAY2zNt49K4idwZFDkyBaTADZsEcmAcLYBEsgWWwAlbB"
    "Gtgwb+PRu4rcGRQ5MAWmwQyYBXNgHiyARbAElsEKWAVrYMO8jUfvKnJnUOTAFJgGM2AWzIF5sAAWwRJY"
    "BitgFayBDfM2Hr2ryJ1BkQNTYBrMgFkwB+bBAlgES2AZrIBVsAY2zNt49L43SqV3SqW3SqX3SqU3S6V3"
    "S6W3S6X3S6U3TKV3TKW3TKX3TKU3TaV3TaW3TaX3Tf18kTvrKnJnUOTAFJgGM2AWzIF5sAAWwRJYBitg"
    "FayBDfM2Hr2ryJ1BkQNTYBrMgFkwB+bBAlgES2AZrIBVsAY2zNt49K4idwZFDkyBaTADZsEcmAcLYBEs"
    "gWWwAlbBGtgwb+P3xe4qcudQ5MAUmAYzYBbMgXmwABbBElgGK2AVrIEN8zYevavInUORA1NgGsyAWTAH"
    "5sECWARLYBmsgFWwBjbM23j0riJ3DkUOTIFpMANmwRyYBwtgESyBZbACVsEa2DBv49G7itw5FDkwBabB"
    "DJgFc2AeLIBFsASWwQpYBWtgw7yNR+8qcudQ5MAUmAYzYBbMgXmwABbBElgGK2AVrIEN8zYevavInUOR"
    "A1NgGsyAWTAH5sECWARLYBmsgFWwBjbM23j0riJ3DkUOTIFpMANmwRyYBwtgESyBZbACVsEa2DBv49G7"
    "itw5FDkwBabBDJgFc2AeLIBFsASWwQpYBWtgw7yNR+8qcudQ5MAUmAYzYBbMgXmwABbBElgGK2AVrIEN"
    "8zYevavInUORA1NgGsyAWTAH5sECWARLYBmsgFWwBjbM22j0i64idwFFDkyBaTADZsEcmAcLYBEsgWWw"
    "AlbBGtgwb+PRu4rcBRQ5MAWmwQyYBXNgHiyARbAElsEKWAVrYMO8jUfvKnIXUOTAFJgGM2AWzIF5sAAW"
    "wRJYBitgFayBDfM2Hr2ryF1AkQNTYBrMgFkwB+bBAlgES2AZrIBVsAY2zNt49K4idwFFDkyBaTADZsEc"
    "mAcLYBEsgWWwAlbBGtgwb+PRu4rcBRQ5MAWmwQyYBXNgHiyARbAElsEKWAVrYMO8jUfvKnIXUOTAFJgG"
    "M2AWzIF5sAAWwRJYBitgFayBDfM2Hr2ryF1AkQNTYBrMgFkwB+bBAlgES2AZrIBVsAY2zNt49K4idwFF"
    "DkyBaTADZsEcmAcLYBEsgWWwAlbBGtgwb+PRu4rcBRQ5MAWmwQyYBXNgHiyARbAElsEKWAVrYMO8jUa/"
    "7Cpyl1DkwBSYBjNgFsyBebAAFsESWAYrYBWszdt42K7qdgnVDUyBaTADZsEcmAcLYBEsgWWwAlbB2ryN"
    "h+0qa5dQ1sAUmAYzYBbMgXmwABbBElgGK2AVrM3beNiuenYJ9QxMgWkwA2bBHJgHC2ARLIFlsAJWwdq8"
    "jYftKmSXUMjAFJgGM2AWzIF5sAAWwRJYBitgFazN23jYrgp2CRUMTIFpMANmwRyYBwtgESyBZbACVsHa"
    "vI2H7Spdl1C6wBSYBjNgFsyBebAAFsESWAYrYBWszdt42K6adQk1C0yBaTADZsEcmAcLYBEsgWWwAlbB"
    "2ryNh+0qVpdQrMAUmAYzYBbMgXmwABbBElgGK2AVrM3beNiuKnUJVQpMgWkwA2bBHJgHC2ARLIFlsAJW"
    "wdq8jYZdHHelp/3VpecsoSLUhIbQEjpCTxgII2EizISFsBI2ASdDd6Wo/dWl5zChItSEhtASOkJPGAgj"
    "YSLMhIWwEjYBJ0N3pan91eVnNMQpQk1oCC2hI/SEgTASJsJMWAgrYRNwMnRXqtpfXX5GQ6wi1ISG0BI6"
    "Qk8YCCNhIsyEhbASNgEnQ3elq/3V5Wc0xCtCTWgILaEj9ISBMBImwkxYCCthE3AydFfK2l9dfkZDzCLU"
    "hIbQEjpCTxgII2EizISFsBI2ASdDd6Wt/dXlZzTELUJNaAgtoSP0hIEwEibCTFgIK2ETcDJ0V+raX11+"
    "RkPsItSEhtASOkJPGAgjYSLMhIWwEjYBJ0N3pa/91eVnNMQvQk1oCC2hI/SEgTASJsJMWAgrYRNwMnRX"
    "CttfXX5GQwwj1ISG0BI6Qk8YCCNhIsyEhbASNgHHQy/6ytiCyhigItSEhtASOkJPGAgjYSLMhIWwEjYB"
    "J0P3lbEFlTFARagJDaEldISeMBBGwkSYCQthJWwCTobuK2MLKmOAilATGkJL6Ag9YSCMhIkwExbCStgE"
    "nAzdV8YWVMYAFaEmNISW0BF6wkAYCRNhJiyElbAJOBm6r4wtqIwBKkJNaAgtoSP0hIEwEibCTFgIK2ET"
    "cDJ0XxlbUBkDVISa0BBaQkfoCQNhJEyEmbAQVsIm4GTovjK2oDIGqAg1oSG0hI7QEwbCSJgIM2EhrIRN"
    "wMnQfWVsQWUMUBFqQkNoCR2hJwyEkTARZsJCWAmbgJOh+8rYgsoYoCLUhIbQEjpCTxgII2EizISFsBI2"
    "ASdD95WxBZUxQEWoCQ2hJXSEnjAQRsJEmAkLYSVsAo6HXvaVsSWVMUBFqAkNoSV0hJ4wEEbCRJgJC2El"
    "bAJOhu4rY0sqY4CKUBMaQkvoCD1hIIyEiTATFsJK2AScDN1XxpZUxgAVoSY0hJbQEXrCQBgJE2EmLISV"
    "sAk4GbqvjC2pjAEqQk1oCC2hI/SEgTASJsJMWAgrYRNwMnRfGVtSGQNUhJrQEFpCR+gJA2EkTISZsBBW"
    "wibgZOi+MrakMgaoCDWhIbSEjtATBsJImAgzYSGshE3AydB9ZWxJZQxQEWpCQ2gJHaEnDISRMBFmwkJY"
    "CZuAk6H7ytiSyhigItSEhtASOkJPGAgjYSLMhIWwEjYBJ0P3lbEllTFARagJDaEldISeMBBGwkSYCQth"
    "JWwCTobuK2NLKmOAilATGkJL6Ag9YSCMhIkwExbCStgEHA+96itjKypjgIpQExpCS+gIPWEgjISJMBMW"
    "wkrYBJwM3VfGVlTGABWhJjSEltAResJAGAkTYSYshJWwCTgZuq+MraiMASpCTWgILaEj9ISBMBImwkxY"
    "CCthE3AydF8ZW1EZA1SEmtAQWkJH6AkDYSRMhJmwEFbCJuBk6L4ytqIyBqgINaEhtISO0BMGwkiYCDNh"
    "IayETcDJ0H1lbEVlDFARakJDaAkdoScMhJEwEWbCQlgJm4CTofvK2IrKGKAi1ISG0BI6Qk8YCCNhIsyE"
    "hbASNgEnQ/eVsRWVMUBFqAkNoSV0hJ4wEEbCRJgJC2ElbAJOhu4rYysqY4CKUBMaQkvoCD1hIIyEiTAT"
    "FsJK2AScDN1XxlZUxgAVoSY0hJbQEXrCQBgJE2EmLISVsAk4Hvqkr4ydUBkDVISa0BBaQkfoCQNhJEyE"
    "mbAQVsIm4GTovjJ2QmUMUBFqQkNoCR2hJwyEkTARZsJCWAmbgJOh+8rYCZUxQEWoCQ2hJXSEnjAQRsJE"
    "mAkLYSVsAk6G7itjJ1TGABWhJjSEltAResJAGAkTYSYshJWwCTgZuq+MnVAZA1SEmtAQWkJH6AkDYSRM"
    "hJmwEFbCJuBk6L4ydkJlDFARakJDaAkdoScMhJEwEWbCQlgJm4CTofvK2AmVMUBFqAkNoSV0hJ4wEEbC"
    "RJgJC2ElbAJOhu4rYydUxgAVoSY0hJbQEXrCQBgJE2EmLISVsAk4GbqvjJ1QGQNUhJrQEFpCR+gJA2Ek"
    "TISZsBBWwibgZOi+MnZCZQxQEWpCQ2gJHaEnDISRMBFmwkJYCZuA46FP+8rYKZUxQEWoCQ2hJXSEnjAQ"
    "RsJEmAkLYSVsAk6G7itjp1TGABWhJjSEltAResJAGAkTYSYshJWwCTgZuq+MnVIZA1SEmtAQWkJH6AkD"
    "YSRMhJmwEFbCJuBk6L4ydkplDFARakJDaAkdoScMhJEwEWbCQlgJm4CTofvK2CmVMUBFqAkNoSV0hJ4w"
    "EEbCRJgJC2ElbAJOhu4rY6dUxgAVoSY0hJbQEXrCQBgJE2EmLISVsAk4GbqvjJ1SGQNUhJrQEFpCR+gJ"
    "A2EkTISZsBBWwibgZOi+MnZKZQxQEWpCQ2gJHaEnDISRMBFmwkJYCZuAk6H7ytgplTFARagJDaEldISe"
    "MBBGwkSYCQthJWwCTobuK2OnVMYAFaEmNISW0BF6wkAYCRNhJiyElbAJOB76rK+MnVEZA1SEmtAQWkJH"
    "6AkDYSRMhJmwEFbCJuBk6L4ydkZlDFARakJDaAkdoScMhJEwEWbCQlgJm4CTofvK2BmVMUBFqAkNoSV0"
    "hJ4wEEbCRJgJC2ElbAJOhu4rY2dUxgAVoSY0hJbQEXrCQBgJE2EmLISVsAk4GbqvjJ1RGQNUhJrQEFpC"
    "R+gJA2EkTISZsBBWwibgZOi+MnZGZQxQEWpCQ2gJHaEnDISRMBFmwkJYCZuAk6H7ytgZlTFARagJDaEl"
    "dISeMBBGwkSYCQthJWwCTobuK2NnVMYAFaEmNISW0BF6wkAYCRNhJiyElbAJOBm6r4ydURkDVISa0BBa"
    "QkfoCQNhJEyEmbAQVsIm4GTovjJ2RmUMUBFqQkNoCR2hJwyEkTARZsJCWAmbgOOhz/vK2DmVMUBFqAkN"
    "oSV0hJ4wEEbCRJgJC2ElbAJOhu4rY+dUxgAVoSY0hJbQEXrCQBgJE2EmLISVsAk4GbqvjJ1TGQNUhJrQ"
    "EFpCR+gJA2EkTISZsBBWwibgZOi+MnZOZQxQEWpCQ2gJHaEnDISRMBFmwkJYCZuAk6H7ytg5lTFARagJ"
    "DaEldISeMBBGwkSYCQthJWwCTobuK2PnVMYAFaEmNISW0BF6wkAYCRNhJiyElbAJOBm6r4ydUxkDVISa"
    "0BBaQkfoCQNhJEyEmbAQVsIm4GTovjJ2TmUMUBFqQkNoCR2hJwyEkTARZsJCWAmbgJOh+8rYOZUxQEWo"
    "CQ2hJXSEnjAQRsJEmAkLYSVsAk6G7itj51TGABWhJjSEltAResJAGAkTYSYshJWwCTge+qKvjF1QGQNU"
    "hJrQEFpCR+gJA2EkTISZsBBWwibgZOi+MnZBZQxQEWpCQ2gJHaEnDISRMBFmwkJYCZuAk6H7ytgFlTFA"
    "RagJDaEldISeMBBGwkSYCQthJWwCTobuK2MXVMYAFaEmNISW0BF6wkAYCRNhJiyElbAJOBm6r4xdUBkD"
    "VISa0BBaQkfoCQNhJEyEmbAQVsIm4GTovjJ2QWUMUBFqQkNoCR2hJwyEkTARZsJCWAmbgJOh+8rYBZUx"
    "QEWoCQ2hJXSEnjAQRsJEmAkLYSVsAk6G7itjF1TGABWhJjSEltAResJAGAkTYSYshJWwCTgZuq+MXVAZ"
    "A1SEmtAQWkJH6AkDYSRMhJmwEFbCJuBk6L4ydkFlDFARakJDaAkdoScMhJEwEWbCQlgJm4DjoS/7ytgl"
    "lTFARagJDaEldISeMBBGwkSYCQthJWwCTobuK2OXVMYAFaEmNISW0BF6wkAYCRNhJiyElbAJOBm6r4xd"
    "UhkDVISa0BBaQkfoCQNhJEyEmbAQVsIm4GTovjJ2SWUMUBFqQkNoCR2hJwyEkTARZsJCWAmbgJOh+8rY"
    "JZUxQEWoCQ2hJXSEnjAQRsJEmAkLYSVsAk6G7itjl1TGABWhJjSEltAResJAGAkTYSYshJWwCTgZuq+M"
    "XVIZA1SEmtAQWkJH6AkDYSRMhJmwEFbCJuBk6L4ydkllDFARakJDaAkdoScMhJEwEWbCQlgJm4CTofvK"
    "2CWVMUBFqAkNoSV0hJ4wEEbCRJgJC2ElbAJOhu4rY5dUxgAVoSY0hJbQEXrCQBgJE2EmLISVsAk4Gnp5"
    "3FXG9leXntGEilATGkJL6Ag9YSCMhIkwExbCStgEnAzdVcb2V5ee0YSKUBMaQkvoCD1hIIyEiTATFsJK"
    "2AR8P/TR/av1equvtldPH7/Y/fqXq5vr3cfrze39wfPNz7e7A1aHUzq4//vd+uXD4N992PzV5p2+27zV"
    "m3e3Tw6P318Qbt/+vC3r+/urn9afLjR3d5u73154dXOzeffs5ur29f4bavvL293lN9f3292tvtzcvfn5"
    "5mrx9HDxaPH7R8tHq0cnh4+PPl38+Gh8v6T7qZbfqf/M/Xy4i8s/LHb/Hj9a7T6udh+XevfPw2X6D8dH"
    "i7/uP3n4eHy03P27evioHx0/4PKvDx8XDx/3vyz3l7z/dH/A7pdnf35kHz3L6dH3P+RHw7/zaMTld/E/"
    "82j8wHdvcsH908dvd+cvV3c/Xe+++27WL3fffMffPvw9o7v3P3Def7LdvN3f6I+b7e6H0f63r9ZXL9Z3"
    "D1fY+cvNZvvxk6Pdrbzb3L3ef5c//T9QSwMEFAAAAAgATKN5XMSqCT/TBAAALBAAABgAAAB4bC93b3Jr"
    "c2hlZXRzL3NoZWV0NS54bWydmFlz2zYQgN/7K3bkTEbupKJISZZs2p6RfKSp5WNyNNO+dCASkmAeYEGw"
    "DPnruwApWp5SaJgH28Sxu/h2scDC5zkXQbqlVMK3KIzTi95WyuTMslJvSyOSDnhCYxxZcxERiU2xsdJE"
    "UOJroSi0nOHwxIoIi3uX57rvSVye80yGLKZPAtIsiogoFjTk+UXP7u06PrLNVqoO6/I8IRv6icovyZPA"
    "ltVo8VlE45TxGARdX/Tm9tnCnigBPeN3RvN07xvSLc/fC+Yv0TKCDHug4FacB2r4g6+6cD4NqSeVUoJ/"
    "/qFXNAwveosZruxvbQY/m1Uowf3vnb1b7Q7EW5GUXvHwK/Pl9qKHSny6JlkoP/L8V1ojjgdjpdHjYap/"
    "Q17NdnCFXpZKHtXiyjuyCOlFb3Tag4jFuisi32o/7clOvkPWqWUdzVOZ16u/JpJcngueg9CCapXOZHA6"
    "2elslo5KtUK0rebO6w7kxFEWqxh/kgKHGeqWl8ub9zcP13O4e1x+uX84tyTaVSOWhz9orzHq1EadwfD/"
    "jTpVxwGbc3h7NHNsx4W7nAhJXlvVGhaVhskBDfY7cN7B6B2MIcxWYP8MfcmSX/h6DQmjIk9L+gyBUl4c"
    "G5hGXZhGRqZFw3RVkhT6KQ0Gx21gIyOYFiaB98wAs4MGWewTbwvch4R75dsj58R2ZZDVbNAfaqu2ezIc"
    "mkDHXUDHRtCrl+DhqvRS2zArJeMDSr4WKwxTCSWEDHNABzFPWFqCQMaZ7XplzCj0EyJx1t3j9R8mukkX"
    "uomR7rqhw148F9rQJsYIVnIgykxmZ2DX8ZloxKX19mg0PnGtJ+iHNCe6OXUF92lgJYLkxIR50gXzxIh5"
    "Uy/rQ4P7J8m5HzOvaCOulDkHlD1kERUFTGAjiFcWEBNYcZYG2QCeBLbLiBGM5hmedniLuDHPcePGyQCm"
    "ru2YgKddgKdG4N8a0M94Q+Fd14Y5Ne9ZvT8/wzOtohYySERJfZBVWupcqD9XBc5x0Jpss/YKctYFcmaE"
    "vNtLTWV/7Kpw1MvCpGpjnplDK6CsNkZAIFCg45GLscbd7ZEa0YLgxVrd92IT3hjYT7uwnxrZlw37PC3S"
    "9jvl9IdYS0KURp7vgPt4aQdcBb8Aj4ZxEWl/MHpshLWHXWjVbAPu/UvirhiWTAQeb28xFJ8+L9vIa22H"
    "0H2RIdts6BYx+Wt+BjG6Qh3QpN7PHt1w4Gus7IocM/mnV/MXer5KBW+bFzLblzH6w+7kD9voj4f/+OP6"
    "5rbVE/YPbYK8ZBVWvQfQPfVYUO9zrCIr/9RHgBG9UzFlm6upxwZ9EfIA+q+D01qD7DROO6fCCk28ZIK6"
    "5MygnSos21xivWlANzkjpR+0Jnmt5BDaHQ+zKC6AlzFeVjym8AbSOmQ88Z55TELs3FnCk0Ri8hcBAxb7"
    "BZbymRo3IXeqtWxzsbVY3uHp6rdyjs0xrK4rpSBHFeEArmDjF1XImuupCWhcuCqLVetVAteyjybgTuWX"
    "ba6/8Pw6CDz5HmCl4BVww9qcUO7h86qWvIf+4lhXbI/Qn7eWKNbeswzLn41+lKbg8SzW2Hu9e09h/ax7"
    "mV69o++J2LA4hZCuUXQ4mCKoqNxXNSRPtCdXXKJrq/cfPuepUBNwfM253DWUgeYfBJf/AlBLAwQUAAAA"
    "CABMo3lcRpW6++gEAABVOgAADQAAAHhsL3N0eWxlcy54bWzdW22vqjgQ/iuEH7C8CcJGTZSDySa7m5vc"
    "82G/olZtUl4W6lm9v37bggJH5izeU8npQoy003nmmem0FC2zkl4I+n5EiGrnhKTlXD9Smv9qGOX2iJK4"
    "/CXLUcok+6xIYsqKxcEo8wLFu5IrJcSwTdMzkhin+mKWnpJ1Qkttm51SOtdN3VjM9lna1FiOXtWwtnGC"
    "tLeYzPUwJnhTYNE4TjC5VNU2r9hmJCs0yrggps1ryh+V2KpKnGaNk+A0K3ilUVl4b2dZ4Jhw+aZGaAwU"
    "h81cX9fHvZWfBrRegmjqtgGDIXgdDFMcD2PAnJb2arKUhxetg7U9lYfnrt3IfZwfHo7nP9wHjjjk+Xif"
    "aIM4wcZhSZtDXyvxVbLWmJDbOLXFOGU1i1keU4qKdM0KQklU3om0+vr1krOBeijii2W7+mCFMiN4x00e"
    "wv5UNVqqnwSN/GgZvcgGtViWRZJB2RGEE9mgPhuvsmO6XkbRi3TQMFpFK9kdtWITgmzQZoqRCNrcO2TG"
    "1J2uralkUHPtRa4nP/nrKVIm03AymYZSQU0z8ln3B5JBnzD2GajPT8mgoR954RPcl937HJRNKHInaQbq"
    "8hMEFV/s9rrJih0qbjfYiX6tWswI2lOmXuDDkX/TLOdLh4zSLGEXOxwfsjQWN9+rRltTE0v4uU6PYgne"
    "WWaE4hDceNPaxkAN0VbQGajAWl55D9SoGn/OxxdxPOJjS2OYjy2FgT62ND7vIxtiD/ZjR2OIjx2FQT52"
    "NK4+3hyrL1jebxEh3znIX/vmKZBBnfda9bT4244/KGp8FXq9ZCOmvqxgqgI31EarsFuwrvNTuFqO3zK6"
    "OjEXUlH++5RR9K1Ae3wW5fP+RgBCtxt0p41u6Vqc5+SyJPiQJqhyfrDBxSy+6mnHrMA/mDW+fOc5oGtv"
    "qKB4y8tb1gAVuvZPEeev6Ezr1b5x3sOMnYbxRA3GE5Vj7CrH2FODsatyVkzVYOwpF+PpU2f7cWeiq/+f"
    "i4g/6nwvh/NEQc7jxlnGWGkx9lWJ8kRBzuPGWXJmjLB+kZ4ZynAeN86SM2OENYz0zFCG87hxlpwZgSpR"
    "nijIedw4S86MEZ4tpWeGMpzHjbOMzAgaxrYqUW5xtkwVSVsqklYyPZT5jcBq/UhudZ5ebUVYe6qQbueH"
    "Ms8pjnJrDgu4s5hfOMpP+KeKD2Z1/5QYeS6Vkncjz6SyOY8wj8qg/MFQeRbjr0UJnHH+s7NbM6PV6W3n"
    "y/a21fqX3Zo+n7Ok5Uk70r6CkQ5UjLRtqhdp21Im0m3W9ldjbdT7glqbjzpbj261Gt9iP9f/5O+2kPTS"
    "8NA2J0woTm+rqWbjEUOl8YagLixT2aF9fCL09Sac6831H2iHT0lwa/WNB6Ru1Vz/zndqWd7tbQFmC6c7"
    "dEa7sC4Wh01n82Hzbsh7SbOB8l4C6VSyfgmXQXYgBpBOpQXZ+T/544P+VDKIm98r8UEdH9SptPokoTgh"
    "O/06ATv6PQ0Cx/E8KKLVPsE7BiEUN8/jn340iBvXgOxwS4/FGu5tOEM+zgOoTz/KEMhTOBMhT+FYc0l/"
    "3MQLKEF/b0N2uAbUC1DucPv9dnhO9es4znX3aR83aATDkiCAJDwX+3PU84DoePzs7x9olDhOEPRLuKyf"
    "AX8nrV/CRyMsgRhwDpCkevvNeHc/Mq73KaN5kXTxL1BLAwQUAAAACABMo3lcl4q7HMAAAAATAgAACwAA"
    "AF9yZWxzLy5yZWxznZK5bsMwDEB/xdCeMAfQIYgzZfEWBPkBVqIP2BIFikWdv6/apXGQCxl5PTwS3B5p"
    "QO04pLaLqRj9EFJpWtW4AUi2JY9pzpFCrtQsHjWH0kBE22NDsFosPkAuGWa3vWQWp3OkV4hc152lPdsv"
    "T0FvgK86THFCaUhLMw7wzdJ/MvfzDDVF5UojlVsaeNPl/nbgSdGhIlgWmkXJ06IdpX8dx/aQ0+mvYyK0"
    "elvo+XFoVAqO3GMljHFitP41gskP7H4AUEsDBBQAAAAIAEyjeVy/K66hlwEAAIUEAAAPAAAAeGwvd29y"
    "a2Jvb2sueG1svZTdatwwEIVfxei+seP9oSyrhdDdJEvTJDQhIVdBtsbrIfox0jjO5tl6m/eqbGPqElh6"
    "s72SzhkYfRxmtGyse8msfYnetDKes5KoWsSxz0vQwp/YCkyoFNZpQUG6XewrB0L6EoC0itMkmcdaoGGr"
    "5dDr1sWrZXt5QGj8H7+V0St6zFAh7Tnr7gpYpNGgxneQnCUs8qVtLq3Dd2tIqLvcWaU4O+0LD+AI80/2"
    "XctzLzLfOW+PaKRtOPtymnxl0f5v2XTqESWVnKWTdDYfvEvAXUmhRTptTRLZT0FoOZsngatA56l7qMMU"
    "OeErhDc5mwRVkz1HReDWguDC2bpCs2tpQhjxKI0uueHsY1+4fwneFgXmsLZ5rcFQn7wD1QIaX2LlWWSE"
    "Bs5+bO7P2lRC/63sE6LANMrbLTAU3FZ2dMcj+X6zfhqRpAdI0uOSSFd//Nob8SxGPJMDPJP/xJONeKYH"
    "eKbH5bnaXGyu1+OxmR2AmXVDPUyyhAINyOvQyAc/LGd+66L22PYLXdRKfQvyxlxZIYedGL6F1W9QSwME"
    "FAAAAAgATKN5XIU5SJ3HAAAAPAQAABoAAAB4bC9fcmVscy93b3JrYm9vay54bWwucmVsc8WUTQ6CMBBG"
    "r0J6AEYBMTHAyg1b4wWaOlDCT5vOGPH2oiygiQs3hlXzTdP3vc00u2AnuTED6cZSMPbdQLnQzPYEQEpj"
    "Lyk0FofppjKulzxFV4OVqpU1QrTbpeDWDFFka2ZwfVr8hWiqqlF4Nure48BfwPAwriWNyCK4Slcj5wLG"
    "bhkTfI59OJFFUN5y4crbXsDWQpEnFG0vFHtC8fZCiSeUbC908IQOfxQifnZIi82cvfr0j/U8vcWl/RPn"
    "ob9Gx7cDeJ9F8QJQSwMEFAAAAAgATKN5XFDd/8srAQAA7wUAABMAAABbQ29udGVudF9UeXBlc10ueG1s"
    "zZRNT8MwDIb/StXr1GaMjwNadwGusAN/ILTuGjVfir3R/XvcdpsEGhVTJ9FLosT2+7yxpSzf9x4waoy2"
    "mMUVkX8UAvMKjMTUebAcKV0wkvgYNsLLvJYbEIv5/EHkzhJYSqjViFfLZyjlVlP00vA1KmezOIDGOHrq"
    "E1tWFkvvtcolcVzsbPGDkhwIKVd2OVgpjzNOiMVZQhv5HXCoe9tBCKqAaC0DvUrDWaLRAmmvAdNhiTMe"
    "XVmqHAqXbw2XpOgDyAIrADI67UVnw2TiDkO/3ozmdzJDQM5cB+eRJxbgctxxJG114lkIAqnhJ56ILD36"
    "fdBOu4Dij2xu76cLdTcPFN02vsffZ3zSv9DHYiI+bifi424iPu7/0ceHc/W1v6B2T41U9sgX3T+/+gJQ"
    "SwECFAMUAAAACABMo3lcRsdNSJUAAADNAAAAEAAAAAAAAAAAAAAAgAEAAAAAZG9jUHJvcHMvYXBwLnht"
    "bFBLAQIUAxQAAAAIAEyjeVyRVLGgFgEAAF4CAAARAAAAAAAAAAAAAACAAcMAAABkb2NQcm9wcy9jb3Jl"
    "LnhtbFBLAQIUAxQAAAAIAEyjeVzppiW4sgUAAFMbAAATAAAAAAAAAAAAAACAAQgCAAB4bC90aGVtZS90"
    "aGVtZTEueG1sUEsBAhQDFAAAAAgATKN5XJ4CYaUpAwAAgwkAABgAAAAAAAAAAAAAAICB6wcAAHhsL3dv"
    "cmtzaGVldHMvc2hlZXQxLnhtbFBLAQIUAxQAAAAIAEyjeVxwdUx/4gcAAHgwAAAYAAAAAAAAAAAAAACA"
    "gUoLAAB4bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWxQSwECFAMUAAAACABMo3lcTCnbQokgAAAGmQEAGAAA"
    "AAAAAAAAAAAAgIFiEwAAeGwvd29ya3NoZWV0cy9zaGVldDMueG1sUEsBAhQDFAAAAAgATKN5XP41EHLI"
    "HQAArowBABgAAAAAAAAAAAAAAICBITQAAHhsL3dvcmtzaGVldHMvc2hlZXQ0LnhtbFBLAQIUAxQAAAAI"
    "AEyjeVzEqgk/0wQAACwQAAAYAAAAAAAAAAAAAACAgR9SAAB4bC93b3Jrc2hlZXRzL3NoZWV0NS54bWxQ"
    "SwECFAMUAAAACABMo3lcRpW6++gEAABVOgAADQAAAAAAAAAAAAAAgAEoVwAAeGwvc3R5bGVzLnhtbFBL"
    "AQIUAxQAAAAIAEyjeVyXirscwAAAABMCAAALAAAAAAAAAAAAAACAATtcAABfcmVscy8ucmVsc1BLAQIU"
    "AxQAAAAIAEyjeVy/K66hlwEAAIUEAAAPAAAAAAAAAAAAAACAASRdAAB4bC93b3JrYm9vay54bWxQSwEC"
    "FAMUAAAACABMo3lchTlInccAAAA8BAAAGgAAAAAAAAAAAAAAgAHoXgAAeGwvX3JlbHMvd29ya2Jvb2su"
    "eG1sLnJlbHNQSwECFAMUAAAACABMo3lcUN3/yysBAADvBQAAEwAAAAAAAAAAAAAAgAHnXwAAW0NvbnRl"
    "bnRfVHlwZXNdLnhtbFBLBQYAAAAADQANAFYDAABDYQAAAAA="
    )
    buf = io.BytesIO(_b64.b64decode("".join(_data.split())))

    # Dodaj Runda i Kolejka do arkusza META (po wierszu Rozgrywki)
    try:
        wb_t = openpyxl.load_workbook(buf)
        for sn in wb_t.sheetnames:
            if sn.upper() == "META":
                ws_m = wb_t[sn]
                rozg_row = None
                runda_exists = False
                kolejka_exists = False
                for row in ws_m.iter_rows(min_row=1):
                    v = str(row[0].value or "").strip().lower()
                    if "rozgrywki" in v:
                        rozg_row = row[0].row
                    if "runda" in v:
                        runda_exists = True
                    if "kolejka" in v:
                        kolejka_exists = True
                if rozg_row:
                    ins_row = rozg_row + 1
                    rows_needed = (0 if runda_exists else 1) + (0 if kolejka_exists else 1)
                    if rows_needed:
                        ws_m.insert_rows(ins_row, rows_needed)
                        r = ins_row
                        if not runda_exists:
                            ws_m.cell(r, 1).value = "Runda"
                            r += 1
                        if not kolejka_exists:
                            ws_m.cell(r, 1).value = "Kolejka"
                break
        buf2 = io.BytesIO()
        wb_t.save(buf2)
        buf2.seek(0)
        return send_file(buf2, as_attachment=True,
                         download_name="SZABLON_MECZU_v3.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception:
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name="SZABLON_MECZU_v3.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/template/szablon")
@login_required
def template_szablon():
    wb = openpyxl.Workbook()
    HDR=PatternFill("solid",fgColor="1A2B4A"); HDR_F=Font(color="FFFFFF",bold=True,size=10)
    YEL=PatternFill("solid",fgColor="FFF8E1"); CTR=Alignment(horizontal="center",vertical="center")
    BORDER=Border(bottom=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),
                  left=Side(style="thin",color="CCCCCC"),top=Side(style="thin",color="CCCCCC"))

    def hdr(ws,row,col,val,w=10):
        c=ws.cell(row,col,val); c.fill=HDR; c.font=HDR_F; c.alignment=CTR; c.border=BORDER
        ws.column_dimensions[get_column_letter(col)].width=w

    sheets=[
        ("TEAM GENERAL",["Q","PKT","POSS","2PM","2PA","2P%","3PM","3PA","3P%","FTM","FTA","FT%","eFG%","TS%","ORtg","DRtg","NetRtg","PPP","TO%","FT Rate"]),
        ("PLAYERS",["#","MIN","PTS","2PM","2PA","2P%","3PM","3PA","3P%","FTM","FTA","FT%","eFG%","TS%","AST","OREB","DREB","TO","FD","FIN"]),
        ("LINEUPS",["Skład","POSS","PKT","PPP","eFG%","ORtg","DRtg","NetRtg","TO","FD"]),
        ("SHOT TIMING",["Czas","2PT Made","2PT Att","2PT%","3PT Made","3PT Att","3PT%","Eff% łącznie"]),
    ]
    ws_first=None
    for sheet_name, cols in sheets:
        if ws_first is None:
            ws=wb.active; ws.title=sheet_name; ws_first=ws
        else:
            ws=wb.create_sheet(sheet_name)
        ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
        t=ws["A1"]; t.value=sheet_name; t.fill=HDR; t.font=Font(color="FFFFFF",bold=True,size=12); t.alignment=CTR
        for i,c in enumerate(cols): hdr(ws,2,i+1,c)
        for r in range(3,18):
            for col in range(1,len(cols)+1):
                c=ws.cell(r,col); c.fill=YEL; c.alignment=CTR; c.border=BORDER
        ws.freeze_panes="B3"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="SZABLON_MECZ.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════════════════════════════
# ROSTER — zarządzanie zawodnikami
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/roster")
@login_required
def roster():
    try: init_db()
    except: pass
    sort = request.args.get("sort","nazwisko")
    order = request.args.get("order","asc")
    if sort not in ("nazwisko","imie","aktywny"): sort = "nazwisko"
    sql_order = f"r.{sort} {'DESC' if order=='desc' else 'ASC'}, r.imie ASC"

    db = get_db(); cur = db.cursor()
    cur.execute(f"""
        SELECT r.id, r.imie, r.nazwisko, r.pseudonim, r.aktywny,
               COALESCE(string_agg(DISTINCT pa.nr::text, ', '), '—') as numery
        FROM roster r
        LEFT JOIN player_aliases pa ON pa.roster_id = r.id
        GROUP BY r.id, r.imie, r.nazwisko, r.pseudonim, r.aktywny
        ORDER BY {sql_order}
    """)
    players = cur.fetchall()
    cur.close()

    def sort_arrow(col):
        if sort == col:
            return " ↓" if order=="asc" else " ↑"
        return " ↕"

    def sort_url(col):
        new_order = "desc" if (sort==col and order=="asc") else "asc"
        return f"/roster?sort={col}&order={new_order}"

    rows = ""
    for i, p in enumerate(players):
        bg = "background:#f8f9ff" if i%2==0 else ""
        is_active = p['aktywny']
        # Toggle button
        toggle_label = "aktywny" if is_active else "nieaktywny"
        toggle_style = "background:#c8e6c9;color:#1a6b3c" if is_active else "background:#e0e0e0;color:#555"
        toggle_title = "Kliknij aby dezaktywować" if is_active else "Kliknij aby aktywować"
        rows += f"""<tr style="{bg}" id="row_{p['id']}">
            <td class="fw-bold">{p['nazwisko']} {p['imie']}</td>
            <td style="color:#888;font-size:.82rem">{p['pseudonim'] or '—'}</td>
            <td style="font-size:.8rem">{p['numery']}</td>
            <td>
              <button type="button"
                onclick="toggleStatus({p['id']}, this)"
                data-active="{1 if is_active else 0}"
                title="{toggle_title}"
                style="border:none;border-radius:20px;padding:3px 10px;font-size:.7rem;font-weight:700;cursor:pointer;transition:.2s;{toggle_style}">
                {toggle_label}
              </button>
            </td>
            <td class="text-center">
              <a href="/roster/{p['id']}/edit" class="btn btn-outline-primary btn-sm" style="font-size:.72rem">Edytuj</a>
              <a href="/roster/{p['id']}/delete" class="btn btn-outline-danger btn-sm ms-1" style="font-size:.72rem"
                 onclick="return confirm('Usunąć {p['imie']} {p['nazwisko']}?')">✕</a>
            </td>
        </tr>"""

    content = f"""
<div class="d-flex justify-content-between align-items-center mb-3 flex-wrap gap-2">
  <div class="page-title mb-0">👥 Skład drużyny <span style="font-size:.8rem;color:#aaa;font-weight:400">({len(players)} zawodników)</span></div>
  <div class="d-flex gap-2 flex-wrap">
    <a href="/roster/szablon" class="btn btn-outline-secondary btn-sm">📥 Szablon Excel</a>
    <button class="btn btn-outline-success btn-sm" onclick="document.getElementById('importFile').click()">📤 Importuj</button>
    <form method="POST" action="/roster/import" enctype="multipart/form-data" style="display:none">
      <input type="file" id="importFile" name="file" accept=".xlsx" onchange="this.form.submit()">
    </form>
    <a href="/roster/nowy" class="btn btn-primary btn-sm">+ Dodaj</a>
  </div>
</div>
<div class="card">
  <div class="card-body p-2">
    <div class="table-responsive">
      <table class="table table-hover mb-0">
        <thead><tr>
          <th>
            <a href="{sort_url('nazwisko')}" class="text-white text-decoration-none">
              Nazwisko i imię{sort_arrow('nazwisko')}
            </a>
          </th>
          <th>Pseudonim</th>
          <th>Numery koszulek</th>
          <th>
            <a href="{sort_url('aktywny')}" class="text-white text-decoration-none">
              Status{sort_arrow('aktywny')}
            </a>
          </th>
          <th class="text-center">Akcje</th>
        </tr></thead>
        <tbody id="rosterBody">
          {rows if rows else '<tr><td colspan="5" class="text-center text-muted py-4">Brak zawodników</td></tr>'}
        </tbody>
      </table>
    </div>
  </div>
</div>"""

    scripts = """<script>
function toggleStatus(id, btn) {
    const isActive = btn.dataset.active === '1';
    fetch('/roster/' + id + '/toggle', {method:'POST'})
    .then(r => r.json())
    .then(data => {
        if(data.ok) {
            btn.dataset.active = data.aktywny ? '1' : '0';
            btn.textContent = data.aktywny ? 'aktywny' : 'nieaktywny';
            btn.style.background = data.aktywny ? '#c8e6c9' : '#e0e0e0';
            btn.style.color = data.aktywny ? '#1a6b3c' : '#555';
            btn.title = data.aktywny ? 'Kliknij aby dezaktywować' : 'Kliknij aby aktywować';
        }
    });
}
</script>"""

    return html_response(base(content, scripts, active="roster"))


@app.route("/roster/szablon")
@login_required
def roster_szablon():
    """Pobierz szablon Excel do importu zawodników"""
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "ZAWODNICY"

    HDR_FILL = PatternFill("solid", fgColor="1A2B4A")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    YEL      = PatternFill("solid", fgColor="FFF9C4")
    EMPTY    = PatternFill("solid", fgColor="FAFAFA")
    CTR      = Alignment(horizontal="center", vertical="center")
    BORDER   = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right =Side(style="thin", color="CCCCCC"),
        left  =Side(style="thin", color="CCCCCC"),
        top   =Side(style="thin", color="CCCCCC"),
    )

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "SZABLON IMPORTU ZAWODNIKÓW — Basket Kołcz Analytics"
    t.fill = HDR_FILL; t.font = Font(color="FFFFFF", bold=True, size=12)
    t.alignment = CTR
    ws.row_dimensions[1].height = 26

    headers = [("A","Imię *",18),("B","Nazwisko",20),("C","Pseudonim",16),("D","Status",14)]
    for col, label, width in headers:
        ws.column_dimensions[col].width = width
        c = ws[f"{col}2"]
        c.value = label; c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = CTR; c.border = BORDER
    ws.row_dimensions[2].height = 22

    examples = [
        ("Jan","Kowalski","KOW","aktywny"),
        ("Piotr","Nowak","NOW","aktywny"),
        ("Marek","Wiśniewski","WIS","aktywny"),
        ("Tomasz","Kowalczyk","","aktywny"),
    ]
    for i, row_data in enumerate(examples):
        r = 3+i
        for j, v in enumerate(row_data):
            c = ws.cell(r, j+1, v)
            c.fill = YEL if v else EMPTY
            c.alignment = CTR; c.border = BORDER
        ws.row_dimensions[r].height = 20

    tbl = Table(displayName="Zawodnicy", ref="A2:D6")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)

    dv = DataValidation(type="list", formula1='"aktywny,nieaktywny"', allow_blank=True)
    ws.add_data_validation(dv); dv.add("D3:D200")

    ws.column_dimensions["H"].width = 52
    ws["H1"] = "INSTRUKCJA:"
    ws["H1"].font = Font(bold=True, color="1A2B4A", size=10)
    instructions = [
        "* Imię jest wymagane — pozostałe pola opcjonalne",
        "* Pseudonim: skrót używany w raportach (np. KOW)",
        "* Status: wybierz z listy — aktywny / nieaktywny",
        "* Numery koszulek przypisuj po imporcie w zakładce Roster",
        "* Nie usuwaj wiersza 2 z nagłówkami",
    ]
    for i, text in enumerate(instructions):
        cell = ws.cell(2+i, 8, text)
        cell.font = Font(italic=True, color="555555", size=9)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[2+i].height = 20

    ws.freeze_panes = "A3"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="SZABLON_ZAWODNICY.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route("/roster/import", methods=["POST"])
@login_required
def roster_import():
    """Importuj zawodników z pliku Excel"""
    if "file" not in request.files:
        flash("Nie wybrano pliku","error"); return redirect(url_for("roster"))
    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        flash("Plik musi być w formacie .xlsx","error"); return redirect(url_for("roster"))

    try:
        init_db()
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        # Znajdź arkusz z danymi
        ws = None
        for sn in wb.sheetnames:
            if "zawodnicy" in sn.lower() or "roster" in sn.lower() or sn == wb.sheetnames[0]:
                ws = wb[sn]; break
        if not ws:
            flash("Nie znaleziono arkusza z danymi","error"); return redirect(url_for("roster"))

        db = get_db(); cur = db.cursor()
        dodani = 0; zaktualizowani = 0; bledy = []

        for i, row in enumerate(ws.iter_rows(min_row=3, max_row=500, values_only=True), 3):
            if not row[0] and not (len(row) > 1 and row[1]): continue
            imie = str(row[0] or "").strip()
            if not imie: continue
            if imie.lower() in ("jan","piotr","marek","tomasz","imię"): continue

            nazwisko  = str(row[1] or "").strip() if len(row) > 1 else ""
            pseudonim = str(row[2] or "").strip() if len(row) > 2 else ""
            col3 = str(row[3] or "").strip() if len(row) > 3 else ""
            col5 = str(row[5] or "").strip() if len(row) > 5 else ""
            if col3.lower() in ("aktywny","nieaktywny",""):
                status = col3.lower() if col3 else "aktywny"
                numery_raw = ""; sezon_raw = ""
            else:
                numery_raw = col3
                sezon_raw  = str(row[4] or "").strip() if len(row) > 4 else ""
                status     = col5.lower() if col5 else "aktywny"
            aktywny = status != "nieaktywny"

            try:
                # Sprawdź czy istnieje po nazwisku (główny klucz)
                cur.execute("SELECT id FROM roster WHERE LOWER(nazwisko)=LOWER(%s)", (nazwisko,))
                existing = cur.fetchone()
                if not existing and imie:
                    # Próbuj też po imieniu + nazwisku
                    cur.execute("SELECT id FROM roster WHERE LOWER(imie)=LOWER(%s) AND LOWER(nazwisko)=LOWER(%s)", (imie, nazwisko))
                    existing = cur.fetchone()

                if existing:
                    cur.execute("UPDATE roster SET pseudonim=%s, aktywny=%s WHERE id=%s",
                                (pseudonim, aktywny, existing["id"]))
                    roster_id = existing["id"]
                    zaktualizowani += 1
                else:
                    cur.execute("INSERT INTO roster (imie,nazwisko,pseudonim,aktywny) VALUES (%s,%s,%s,%s) RETURNING id",
                                (imie, nazwisko, pseudonim, aktywny))
                    roster_id = cur.fetchone()["id"]
                    dodani += 1

                # Dodaj numery koszulek
                if numery_raw:
                    import re as _re
                    for nr_part in numery_raw.split(","):
                        nr_part = nr_part.strip()
                        if not nr_part: continue
                        try:
                            nr = int(_re.sub(r'[^\d]','', nr_part))
                            sezon = sezon_raw
                            cur.execute("""INSERT INTO player_aliases (roster_id,nr,sezon)
                                          VALUES (%s,%s,%s) ON CONFLICT DO NOTHING""",
                                        (roster_id, nr, sezon))
                        except: pass

            except Exception as e:
                bledy.append(f"Wiersz {i}: {str(e)[:60]}")
                try: get_db().rollback()
                except: pass

        db.commit(); cur.close()

        msg = f"✓ Import zakończony: {dodani} nowych, {zaktualizowani} zaktualizowanych"
        if bledy: msg += f" | Błędy: {', '.join(bledy[:3])}"
        flash(msg, "success" if not bledy else "error")

    except Exception as e:
        try: get_db().rollback()
        except: pass
        flash(f"Błąd importu: {str(e)}", "error")

    return redirect(url_for("roster"))


@app.route("/roster/nowy", methods=["GET","POST"])
@app.route("/roster/<int:player_id>/edit", methods=["GET","POST"])
@login_required
def roster_edit(player_id=None):
    try: init_db()
    except: pass
    db = get_db(); cur = db.cursor()

    if request.method == "POST":
        imie     = request.form.get("imie","").strip()
        nazwisko = request.form.get("nazwisko","").strip()
        pseudonim= request.form.get("pseudonim","").strip()
        aktywny  = request.form.get("aktywny","1") == "1"
        # Numery — lista par (nr, sezon)
        numery_raw = request.form.get("numery","").strip()

        if not imie:
            flash("Imię jest wymagane","error")
            return redirect(request.url)

        if player_id:
            cur.execute("UPDATE roster SET imie=%s,nazwisko=%s,pseudonim=%s,aktywny=%s WHERE id=%s",
                        (imie,nazwisko,pseudonim,aktywny,player_id))
            cur.execute("DELETE FROM player_aliases WHERE roster_id=%s", (player_id,))
        else:
            cur.execute("INSERT INTO roster (imie,nazwisko,pseudonim,aktywny) VALUES (%s,%s,%s,%s) RETURNING id",
                        (imie,nazwisko,pseudonim,aktywny))
            player_id = cur.fetchone()["id"]

        # Parsuj numery: "5, 12 (2024/25), 7 (2025/26)"
        for part in numery_raw.split(","):
            part = part.strip()
            if not part: continue
            import re as _re
            m = _re.match(r'(\d+)\s*(?:\(([^)]+)\))?', part)
            if m:
                nr = int(m.group(1))
                sezon = (m.group(2) or "").strip()
                try:
                    cur.execute("INSERT INTO player_aliases (roster_id,nr,sezon) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                                (player_id, nr, sezon))
                except: pass

        db.commit(); cur.close()
        flash(f"✓ Zawodnik {imie} {nazwisko} zapisany","success")
        return redirect(url_for("roster"))

    # GET
    player = None; aliases = []
    if player_id:
        cur.execute("SELECT * FROM roster WHERE id=%s", (player_id,))
        player = cur.fetchone()
        cur.execute("SELECT nr,sezon FROM player_aliases WHERE roster_id=%s ORDER BY nr", (player_id,))
        aliases = cur.fetchall()
    cur.close()

    numery_str = ", ".join(
        f"{a['nr']}" + (f" ({a['sezon']})" if a['sezon'] else "")
        for a in aliases
    )

    title = "Edytuj zawodnika" if player_id else "Nowy zawodnik"
    content = f"""
<div class="page-title">{title}</div>
<div class="row justify-content-center">
<div class="col-lg-6">
  <div class="card p-3">
    <form method="POST">
      <div class="row g-3">
        <div class="col-6">
          <label class="form-label fw-bold">Imię *</label>
          <input type="text" name="imie" class="form-control" value="{''+player['imie'] if player else ''}" required>
        </div>
        <div class="col-6">
          <label class="form-label fw-bold">Nazwisko</label>
          <input type="text" name="nazwisko" class="form-control" value="{''+player['nazwisko'] if player else ''}">
        </div>
        <div class="col-6">
          <label class="form-label fw-bold">Pseudonim / inicjały</label>
          <input type="text" name="pseudonim" class="form-control" value="{''+player['pseudonim'] if player else ''}" placeholder="np. KOW, Kowal">
        </div>
        <div class="col-6">
          <label class="form-label fw-bold">Status</label>
          <select name="aktywny" class="form-select">
            <option value="1" {'selected' if not player or player['aktywny'] else ''}>Aktywny</option>
            <option value="0" {'selected' if player and not player['aktywny'] else ''}>Nieaktywny</option>
          </select>
        </div>
        <div class="col-12">
          <label class="form-label fw-bold">Numery koszulek</label>
          <input type="text" name="numery" class="form-control" value="{numery_str}"
                 placeholder="np. 5, 12 (2024/25), 7 (2025/26)">
          <div class="form-text">Wpisz numery oddzielone przecinkiem. Opcjonalnie dodaj sezon w nawiasie.</div>
        </div>
      </div>
      <div class="d-flex gap-2 mt-3">
        <button type="submit" class="btn btn-primary">Zapisz</button>
        <a href="/roster" class="btn btn-outline-secondary">Anuluj</a>
      </div>
    </form>
  </div>
</div></div>"""
    return html_response(base(content, active="players"))


@app.route("/roster/<int:player_id>/toggle", methods=["POST"])
@login_required
def roster_toggle(player_id):
    from flask import jsonify
    try:
        db = get_db(); cur = db.cursor()
        cur.execute("SELECT aktywny FROM roster WHERE id=%s", (player_id,))
        row = cur.fetchone()
        if not row: return jsonify({"ok": False})
        new_status = not row["aktywny"]
        cur.execute("UPDATE roster SET aktywny=%s WHERE id=%s", (new_status, player_id))
        db.commit(); cur.close()
        return jsonify({"ok": True, "aktywny": new_status})
    except Exception as e:
        try: get_db().rollback()
        except: pass
        return jsonify({"ok": False, "error": str(e)})


@app.route("/roster/<int:player_id>/delete")
@login_required
def roster_delete(player_id):
    db = get_db(); cur = db.cursor()
    cur.execute("DELETE FROM roster WHERE id=%s", (player_id,))
    db.commit(); cur.close()
    flash("Zawodnik usunięty","success")
    return redirect(url_for("roster"))


# ══════════════════════════════════════════════════════════════════════════════
# EDYTOR MECZU — przypisanie numerów do zawodników z rostera
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/mecz/<int:match_id>/edytuj", methods=["GET","POST"])
@login_required
def mecz_edytuj(match_id):
    try: init_db()
    except: pass
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
    m = cur.fetchone()
    if not m: return redirect(url_for("historia"))

    gtk_name = (m.get("team_name_a","") or m.get("nazwa_gtk","") or "").strip() or get_setting("gtk_name") or "GTK"

    cur.execute("""SELECT id,nr,pts,p2m,p2a,p3m,p3a,ftm,fta,roster_id,player_id
                   FROM player_stats WHERE match_id=%s AND druzyna='gtk'
                   ORDER BY nr""", (match_id,))
    players = list(cur.fetchall())

    # Pobierz zawodników z tabeli players (Struktura klubów) jeśli mecz ma team_id
    # Fallback na roster jeśli brak team_id
    match_team_id = m.get("team_id")
    if match_team_id:
        try:
            cur.execute("""SELECT p.id, p.imie, p.nazwisko,
                                  NULL as pseudonim
                           FROM players p
                           WHERE p.team_id=%s
                           ORDER BY p.nazwisko ASC, p.imie ASC""",
                        (match_team_id,))
            roster_list = list(cur.fetchall())
        except Exception:
            roster_list = []
        if not roster_list:
            cur.execute("""SELECT r.id, r.imie, r.nazwisko, r.pseudonim
                           FROM roster r WHERE r.aktywny=TRUE
                           ORDER BY r.nazwisko ASC, r.imie ASC""")
            roster_list = list(cur.fetchall())
    else:
        # Brak team_id — spróbuj znaleźć drużynę przez kontekst sesji
        # i pobierz z players; fallback na roster
        roster_list = []
        try:
            # Pobierz wszystkich zawodników ze wszystkich drużyn (players)
            cur.execute("""SELECT p.id, p.imie, p.nazwisko, NULL as pseudonim,
                                  t.name as team_name
                           FROM players p
                           JOIN teams t ON p.team_id = t.id
                           ORDER BY t.name, p.nazwisko ASC, p.imie ASC""")
            roster_list = list(cur.fetchall())
        except Exception:
            roster_list = []
        if not roster_list:
            cur.execute("""SELECT r.id, r.imie, r.nazwisko, r.pseudonim
                           FROM roster r WHERE r.aktywny=TRUE
                           ORDER BY r.nazwisko ASC, r.imie ASC""")
            roster_list = list(cur.fetchall())

    if request.method == "POST":
        sezon = m.get("sezon","")
        # Sprawdź czy roster_list pochodzi z players (nowy) czy roster (stary)
        roster_ids_from_players = {r["id"] for r in roster_list if not r.get("pseudonim") or r.get("team_name")}
        for p in players:
            rid = request.form.get(f"roster_{p['id']}","")
            rid_val = int(rid) if rid and rid.isdigit() else None
            if rid_val:
                # Jeśli mecz ma team_id LUB wybrany id pochodzi z tabeli players
                if match_team_id or rid_val in roster_ids_from_players:
                    cur.execute("UPDATE player_stats SET player_id=%s WHERE id=%s", (rid_val, p['id']))
                else:
                    cur.execute("UPDATE player_stats SET roster_id=%s WHERE id=%s", (rid_val, p['id']))
                    if p['nr'] is not None:
                        try:
                            cur.execute("""INSERT INTO player_aliases (roster_id,nr,sezon)
                                           VALUES (%s,%s,%s) ON CONFLICT DO NOTHING""",
                                        (rid_val, p['nr'], sezon))
                        except: pass
        db.commit(); cur.close()
        flash("✓ Przypisania zawodników zapisane","success")
        return redirect(url_for("mecz", match_id=match_id))

    cur.close()

    # JSON roster dla JS
    import json as _json
    roster_json = _json.dumps([
        {"id": r["id"],
         "name": f"{r['nazwisko']} {r['imie']}" +
                 (f" [{r['team_name']}]" if r.get('team_name') else "") +
                 (f" ({r['pseudonim']})" if r.get('pseudonim') else "")}
        for r in roster_list
    ])

    # Aktualne przypisania {ps_id: roster_id}
    # Użyj player_id jeśli mecz ma team_id, inaczej roster_id
    if match_team_id:
        current = {p['id']: (p.get('player_id') or p.get('roster_id') or "") for p in players}
    else:
        current = {p['id']: (p.get('roster_id') or "") for p in players}
    current_json = _json.dumps({str(k): str(v) if v else "" for k,v in current.items()})

    # Wiersze tabeli — same numery i statystyki, select wypełniany przez JS
    rows = ""
    for i, p in enumerate(players):
        fga = p.get('p2a',0)+p.get('p3a',0)
        efg = f"{(p.get('p2m',0)+1.5*p.get('p3m',0))/fga:.0%}" if fga else "—"
        bg = "background:#f8f9ff" if i%2==0 else ""
        rows += f"""<tr style="{bg}">
            <td class="fw-bold" style="width:55px;font-size:1rem">#{p['nr']}</td>
            <td class="fw-bold" style="color:#1a2b4a;width:50px">{p['pts']}</td>
            <td style="font-size:.78rem;color:#888;width:90px">{p.get('p2m',0)}/{p.get('p2a',0)} | {p.get('p3m',0)}/{p.get('p3a',0)}</td>
            <td style="font-size:.78rem;color:#888;width:55px">{efg}</td>
            <td>
              <select name="roster_{p['id']}" id="sel_{p['id']}"
                      class="form-select form-select-sm roster-sel"
                      data-ps-id="{p['id']}"
                      onchange="updateSelects()">
              </select>
            </td>
        </tr>"""

    dt = m['data_meczu'].strftime('%d.%m.%Y') if m['data_meczu'] else ""

    content = f"""
<div class="d-flex justify-content-between align-items-center mb-3">
  <div>
    <div class="page-title mb-0">✏️ Przypisz zawodników — {gtk_name}</div>
    <div style="font-size:.8rem;color:#888">{m['przeciwnik']} · {dt}</div>
  </div>
  <a href="/mecz/{match_id}" class="btn btn-outline-secondary btn-sm">← Wróć</a>
</div>

<div class="card p-2 mb-3" style="background:#e8f5e9;border:1px solid #a5d6a7">
  <div style="font-size:.82rem;color:#1a6b3c">
    <b>Wybierz zawodnika</b> z listy dla każdego numeru koszulki.
    Wybrany zawodnik znika z pozostałych list.
    Numer zostanie automatycznie przypisany do zawodnika w rosterze.
    <a href="/roster/nowy" style="color:#1a6b3c">+ Dodaj zawodnika do rostera</a>
  </div>
</div>

<div class="card">
  <div class="card-body p-3">
    <form method="POST" id="assignForm">
      <div class="table-responsive">
        <table class="table table-hover mb-0" style="table-layout:fixed">
          <thead><tr>
            <th style="width:55px">#</th>
            <th style="width:50px">PTS</th>
            <th style="width:90px">2PT|3PT</th>
            <th style="width:55px">eFG%</th>
            <th>Zawodnik</th>
          </tr></thead>
          <tbody>{rows}</tbody>
        </table>
      </div>
      <div class="d-flex gap-2 mt-3">
        <button type="submit" class="btn btn-primary fw-bold">✓ Zapisz</button>
        <a href="/mecz/{match_id}" class="btn btn-outline-secondary">Anuluj</a>
        <button type="button" class="btn btn-outline-warning btn-sm ms-auto"
                onclick="clearAll()">Wyczyść wszystkie</button>
      </div>
    </form>
  </div>
</div>"""

    scripts = f"""<script>
const ROSTER = {roster_json};
const CURRENT = {current_json};

function getSelected() {{
  const sel = {{}};
  document.querySelectorAll('.roster-sel').forEach(s => {{
    if(s.value) sel[s.value] = s.dataset.psId;
  }});
  return sel;
}}

function updateSelects() {{
  const selected = getSelected();
  document.querySelectorAll('.roster-sel').forEach(sel => {{
    const curVal = sel.value;
    sel.innerHTML = '<option value="">— nie przypisany —</option>';
    ROSTER.forEach(r => {{
      const usedBy = selected[r.id.toString()];
      // Pokaż jeśli: nie wybrany przez nikogo LUB wybrany przez ten sam select
      if(!usedBy || usedBy === sel.dataset.psId) {{
        const opt = document.createElement('option');
        opt.value = r.id;
        opt.textContent = r.name;
        if(r.id.toString() === curVal) opt.selected = true;
        sel.appendChild(opt);
      }}
    }});
  }});
}}

function clearAll() {{
  document.querySelectorAll('.roster-sel').forEach(s => s.value = '');
  updateSelects();
}}

// Inicjalizacja — wypełnij z aktualnych przypisań
document.addEventListener('DOMContentLoaded', () => {{
  // Najpierw wypełnij wszystkie puste
  document.querySelectorAll('.roster-sel').forEach(sel => {{
    sel.innerHTML = '<option value="">— nie przypisany —</option>';
    ROSTER.forEach(r => {{
      const opt = document.createElement('option');
      opt.value = r.id; opt.textContent = r.name;
      sel.appendChild(opt);
    }});
  }});
  // Ustaw zapisane wartości
  const psIds = Object.keys(CURRENT);
  psIds.forEach(psId => {{
    const rid = CURRENT[psId];
    if(rid) {{
      const sel = document.querySelector(`[data-ps-id="${{psId}}"]`);
      if(sel) sel.value = rid;
    }}
  }});
  // Teraz odśwież listy
  updateSelects();
}});
</script>"""

    return html_response(base(content, scripts, active="history"))


@app.route("/reparse/<int:match_id>")
@login_required
def reparse_match_dev(match_id):
    """Re-parsuje mecz z pliku /tmp/reparse_match.xlsx i nadpisuje dane w bazie."""
    import openpyxl, io as _io
    REPARSE_FILE = "/mnt/user-data/outputs/reparse_match.xlsx"
    try:
        if not os.path.exists(REPARSE_FILE):
            return f"<p>Brak pliku {REPARSE_FILE}</p>", 404
        wb = openpyxl.load_workbook(REPARSE_FILE, data_only=True)
        report = validate_workbook(wb)
        meta = report["meta"]
        name_a, name_b = report["names"]
        # Usun stare dane
        db = get_db(); cur = db.cursor()
        cur.execute("DELETE FROM match_stats WHERE match_id=%s", (match_id,))
        cur.execute("DELETE FROM player_stats WHERE match_id=%s", (match_id,))
        cur.execute("DELETE FROM lineup_stats WHERE match_id=%s", (match_id,))
        cur.execute("DELETE FROM timing_stats WHERE match_id=%s", (match_id,))
        cur.execute("DELETE FROM score_flow WHERE match_id=%s", (match_id,))
        db.commit()
        # Parsuj
        stats_gtk = parse_team_sheet(wb[name_a], sheet_type="A")
        stats_opp = parse_team_sheet(wb[name_b], sheet_type="B")
        try: def_lineups = build_gtk_def_lineups(wb[name_a], wb[name_b])
        except: def_lineups = None
        # Zapisz match_stats — dreb/stl/blk GTK z arkusza B (stats_opp)
        for qn in [1,2,3,4]:
            qg = dict(stats_gtk["quarter"].get(qn, {}))
            qo = dict(stats_opp["quarter"].get(qn, {}))
            qg["dreb"] = qo.get("dreb", 0)
            qg["stl"]  = qo.get("stl",  0)
            qg["blk"]  = qo.get("blk",  0)
            for druzyna, qd in [("gtk", qg), ("opp", qo)]:
                cur.execute("""
                    INSERT INTO match_stats
                    (match_id,druzyna,kwarta,pts,poss,p2m,p2a,p3m,p3a,ftm,fta,br,fd,ast,oreb,dreb,stl,blk,d2m,d2a,przerw)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (match_id, druzyna, qn,
                      qd.get("pts",0), qd.get("poss",0),
                      qd.get("p2m",0), qd.get("p2a",0),
                      qd.get("p3m",0), qd.get("p3a",0),
                      qd.get("ftm",0), qd.get("fta",0),
                      qd.get("br",0),  qd.get("fd",0),
                      qd.get("ast",0), qd.get("oreb",0), qd.get("dreb",0),
                      qd.get("stl",0), qd.get("blk",0),
                      qd.get("d2m",0), qd.get("d2a",0), qd.get("przerw",0)))
        # Zapisz player_stats — dreb/stl/blk GTK z arkusza B (stats_opp)
        opp_players_r = dict(stats_opp["players"])
        for druzyna, stats in [("gtk", stats_gtk), ("opp", stats_opp)]:
            for nr, pd in stats["players"].items():
                pts = pd.get("p2m",0)*2 + pd.get("p3m",0)*3 + pd.get("ftm",0)
                if druzyna == "gtk" and int(nr) in opp_players_r:
                    op = opp_players_r[int(nr)]
                    dreb = op.get("dreb",0); stl = op.get("stl",0); blk = op.get("blk",0)
                else:
                    dreb = pd.get("dreb",0); stl = pd.get("stl",0); blk = pd.get("blk",0)
                cur.execute("""
                    INSERT INTO player_stats
                    (match_id,druzyna,nr,pts,p2m,p2a,p3m,p3a,ftm,fta,ast,oreb,dreb,br,fd,finishes,time_sum,time_cnt,stl,blk)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (match_id, druzyna, int(nr), pts,
                      pd.get("p2m",0), pd.get("p2a",0),
                      pd.get("p3m",0), pd.get("p3a",0),
                      pd.get("ftm",0), pd.get("fta",0),
                      pd.get("ast",0), pd.get("oreb",0), dreb,
                      pd.get("br",0),  pd.get("fd",0), pd.get("finishes",0),
                      pd.get("time_sum",0), pd.get("time_cnt",0),
                      stl, blk))
        # Zapisz lineup_stats
        for druzyna, stats in [("gtk", stats_gtk), ("opp", stats_opp)]:
            for lineup_key, lu in stats["lineups"].items():
                # Dla GTK: STL i BLK są w def_lineups (arkusz B), nie w stats_gtk
                _def_lu = (def_lineups or {}).get(lineup_key, {}) if druzyna == "gtk" else {}
                cur.execute("""
                    INSERT INTO lineup_stats (match_id,druzyna,lineup,pts,poss,p2m,p2a,p3m,p3a,ftm,fta,br,fd,ast,oreb,dreb,stl,blk)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (match_id, druzyna, lineup_key,
                      lu.get("pts",0), lu.get("poss",0),
                      lu.get("p2m",0), lu.get("p2a",0),
                      lu.get("p3m",0), lu.get("p3a",0),
                      lu.get("ftm",0), lu.get("fta",0),
                      lu.get("br",0),  lu.get("fd",0),
                      lu.get("ast",0), lu.get("oreb",0),
                      _def_lu.get("dreb",0) or lu.get("dreb",0),
                      _def_lu.get("stl",0) or lu.get("stl",0),
                      _def_lu.get("blk",0) or lu.get("blk",0)))
        if def_lineups:
            for lineup_key, lu in def_lineups.items():
                cur.execute("""
                    INSERT INTO lineup_stats (match_id,druzyna,lineup,pts,poss,p2m,p2a,p3m,p3a,ftm,fta,br,fd,ast,oreb,dreb,stl,blk)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (match_id, "gtk_def", lineup_key,
                      lu.get("pts",0), lu.get("poss",0),
                      lu.get("p2m",0), lu.get("p2a",0),
                      lu.get("p3m",0), lu.get("p3a",0),
                      lu.get("ftm",0), lu.get("fta",0),
                      lu.get("br",0),  lu.get("fd",0),
                      lu.get("ast",0), lu.get("oreb",0),
                      lu.get("dreb",0),lu.get("stl",0), lu.get("blk",0)))
        db.commit(); cur.close()
        flash(f"✅ Mecz #{match_id} przeliczony z nową logiką parsera.", "success")
        return redirect(url_for("mecz", match_id=match_id))
    except Exception as e:
        import traceback
        return f"<pre>BŁĄD: {traceback.format_exc()}</pre>", 500


@app.route("/debug-version")
def debug_version():
    import hashlib, os
    path = os.path.abspath(__file__)
    with open(path, 'rb') as f:
        h = hashlib.md5(f.read()).hexdigest()
    with open(path, 'r', encoding='utf-8') as f:
        src = f.read()
    has_piatki = "gtktab-gtk_l" in src
    has_accordion = "cmp-accordion" in src

    # DB diagnostics for latest match
    db_info = ""
    try:
        db2 = get_db()
        cur2 = db2.cursor()
        cur2.execute("SELECT id FROM matches ORDER BY id DESC LIMIT 1")
        row = cur2.fetchone()
        if row:
            mid = row[0]
            cur2.execute("""SELECT ps.id, ps.nr, ps.player_id, ps.roster_id,
                                   p.imie, p.nazwisko, r.imie as ri, r.nazwisko as rn
                            FROM player_stats ps
                            LEFT JOIN players p ON ps.player_id = p.id
                            LEFT JOIN roster r ON ps.roster_id = r.id
                            WHERE ps.match_id=%s AND ps.druzyna='gtk'
                            ORDER BY ps.nr""", (mid,))
            rows2 = cur2.fetchall()
            db_info = f"Match {mid} player_stats (gtk):\n"
            for r2 in rows2:
                pname = f"{r2['imie']} {r2['nazwisko']}" if r2.get('imie') else "—"
                rname = f"{r2['ri']} {r2['rn']}" if r2.get('ri') else "—"
                db_info += f"  nr={r2['nr']} player_id={r2['player_id']} roster_id={r2['roster_id']} -> players:{pname} roster:{rname}\n"
        cur2.close()
    except Exception as e:
        db_info = f"DB error: {e}"

    return Response(f"""<pre>
File: {path}
MD5: {h}
Size: {os.path.getsize(path)} bytes
Has Piatki tab (OLD): {has_piatki}
Has CMP accordion (NEW): {has_accordion}
gtktab-gtk_l count: {src.count("gtktab-gtk_l")}

{db_info}
</pre>""", mimetype='text/html')


@app.route("/porownaj/zbiorcze")
@login_required
def porownaj_zbiorcze():
    """Raport zbiorczy z wielu meczów — HTML/PDF."""
    ids_raw = request.args.getlist("ids")
    if not ids_raw:
        ids_raw = (request.args.get("ids","") or "").split(",")
    match_ids = [int(x) for x in ids_raw if str(x).isdigit()]
    if not match_ids:
        flash("Nie wybrano żadnych meczów.")
        return redirect(url_for("porownaj"))

    db = get_db(); cur = db.cursor()

    # Pobierz dane każdego meczu
    mecze_data = []
    all_match_stats = {}
    all_player_stats = {}
    all_lineups_data = {}
    for mid in match_ids:
        cur.execute("SELECT * FROM matches WHERE id=%s", (mid,))
        m = cur.fetchone()
        if not m: continue
        cur.execute("SELECT * FROM match_stats WHERE match_id=%s ORDER BY kwarta", (mid,))
        mstats = list(cur.fetchall())
        cur.execute("SELECT * FROM player_stats WHERE match_id=%s", (mid,))
        pstats = list(cur.fetchall())
        mecze_data.append(m)
        all_match_stats[mid] = mstats
        all_player_stats[mid] = pstats
        all_lineups_data[mid] = {"off": [], "def": []}
        try:
            cur.execute("""SELECT * FROM lineup_stats WHERE match_id=%s AND druzyna='gtk' ORDER BY poss DESC""", (mid,))
            all_lineups_data[mid]["off"] = list(cur.fetchall())
            cur.execute("""SELECT * FROM lineup_stats WHERE match_id=%s AND druzyna='opp' ORDER BY poss DESC""", (mid,))
            all_lineups_data[mid]["def"] = list(cur.fetchall())
        except Exception:
            pass

    if not mecze_data:
        flash("Brak danych dla wybranych meczów.")
        return redirect(url_for("porownaj"))

    n = len(mecze_data)
    gtk_name = (mecze_data[0].get("team_name_a","") or mecze_data[0].get("nazwa_gtk","") or "GTK").strip()
    opp_names = list({m.get("przeciwnik","OPP") for m in mecze_data})
    opp_label = opp_names[0] if len(opp_names)==1 else "OPP"

    # ── Agregaty per mecz ────────────────────────────────────────────────────
    def qv(stats, druzyna, field, kwarta=None):
        rows = [r for r in stats if r["druzyna"]==druzyna and (kwarta is None or r["kwarta"]==kwarta)]
        return sum(r.get(field,0) or 0 for r in rows)

    def pct(made, att): return round(made/att*100,1) if att else 0

    def efg(p2m,p2a,p3m,p3a): return round((p2m+1.5*p3m)/(p2a+p3a)*100,1) if (p2a+p3a) else 0

    # Tabela meczów
    rows_mecze = []
    wins = 0
    sum_gtk = sum_opp = 0
    sum_q = {q:{"g":0,"o":0} for q in [1,2,3,4]}
    for m in mecze_data:
        mid = m["id"]
        ms  = all_match_stats.get(mid,[])
        wg  = m.get("wynik_gtk",0) or 0
        wo  = m.get("wynik_opp",0) or 0
        win = wg > wo
        if win: wins += 1
        sum_gtk += wg; sum_opp += wo
        dt = m["data_meczu"].strftime("%d.%m.%Y") if m["data_meczu"] else "—"
        qs = []
        for q in [1,2,3,4]:
            qg = qv(ms,"gtk","pts",q); qo = qv(ms,"opp","pts",q)
            sum_q[q]["g"] += qg; sum_q[q]["o"] += qo
            qs.append(f"{qg}–{qo}")
        opp_n = m.get("przeciwnik","OPP")
        miejsce = m.get("miejsce","—") or "—"
        badge = '<span style="background:#d4edda;color:#155724;border-radius:50%;width:18px;height:18px;display:inline-flex;align-items:center;justify-content:center;font-size:9px;font-weight:500">W</span>' if win else '<span style="background:#f8d7da;color:#721c24;border-radius:50%;width:18px;height:18px;display:inline-flex;align-items:center;justify-content:center;font-size:9px;font-weight:500">P</span>'
        rows_mecze.append(f'''<tr style="{'background:#fafbfd' if len(rows_mecze)%2 else ''}">
          <td style="text-align:left">{dt}</td><td style="text-align:left">{opp_n}</td><td>{miejsce}</td>
          {''.join(f'<td>{q}</td>' for q in qs)}
          <td style="color:{'#0F6E56' if win else '#A32D2D'};font-weight:500">{wg}</td>
          <td style="color:{'#A32D2D' if win else '#0F6E56'};font-weight:500">{wo}</td>
          <td>{badge}</td>
        </tr>''')

    # Wiersz SUMA i ŚREDNIA
    avg_gtk = round(sum_gtk/n,1); avg_opp = round(sum_opp/n,1)
    row_suma = f'''<tr style="background:#E6F1FB;font-weight:500">
      <td style="text-align:left" colspan="3">SUMA</td>
      {''.join(f'<td>{sum_q[q]["g"]}–{sum_q[q]["o"]}</td>' for q in [1,2,3,4])}
      <td style="color:#0C447C">{sum_gtk}</td><td style="color:#A32D2D">{sum_opp}</td><td></td>
    </tr>'''
    row_avg = f'''<tr style="background:#EEEDFE;font-weight:500">
      <td style="text-align:left" colspan="3">ŚREDNIA <span style="background:#534AB7;color:#fff;border-radius:8px;padding:1px 6px;font-size:8px;margin-left:4px">per mecz</span></td>
      {''.join(f'<td>{round(sum_q[q]["g"]/n,1)}–{round(sum_q[q]["o"]/n,1)}</td>' for q in [1,2,3,4])}
      <td style="color:#534AB7">{avg_gtk}</td><td style="color:#A32D2D">{avg_opp}</td><td></td>
    </tr>'''

    # ── KPI uśrednione ───────────────────────────────────────────────────────
    def avg_field(druzyna, field):
        total = 0
        for mid, ms in all_match_stats.items(): total += qv(ms,druzyna,field)
        return round(total/n,1)

    def avg_pct_fields(druzyna, made_f, att_f):
        tm = ta = 0
        for mid, ms in all_match_stats.items():
            tm += qv(ms,druzyna,made_f); ta += qv(ms,druzyna,att_f)
        return pct(tm,ta)

    def avg_efg(druzyna):
        tm2=ta2=tm3=ta3=0
        for mid, ms in all_match_stats.items():
            tm2+=qv(ms,druzyna,"p2m"); ta2+=qv(ms,druzyna,"p2a")
            tm3+=qv(ms,druzyna,"p3m"); ta3+=qv(ms,druzyna,"p3a")
        return efg(tm2,ta2,tm3,ta3)

    def avg_ortg(druzyna):
        tp=tposs=0
        for mid,ms in all_match_stats.items():
            tp+=qv(ms,druzyna,"pts"); tposs+=qv(ms,druzyna,"poss")
        return round(tp/tposs*100,1) if tposs else 0

    def avg_ts(druzyna):
        tp=tf=0
        for mid,ms in all_match_stats.items():
            pts=qv(ms,druzyna,"pts"); p2a=qv(ms,druzyna,"p2a"); p3a=qv(ms,druzyna,"p3a"); fta=qv(ms,druzyna,"fta")
            tp+=pts; tf+=p2a+p3a+0.44*fta
        return round(tp/(2*tf)*100,1) if tf else 0

    def kpi_box(val, lbl, cls=""):
        return f'<div class="kpi-box {cls}"><div class="kpi-val">{val}</div><div class="kpi-lbl">{lbl}</div></div>'

    for_druzyna = {}
    for d in ["gtk","opp"]:
        ortg   = avg_ortg(d)
        drtg   = avg_ortg("opp" if d=="gtk" else "gtk")
        net    = round(ortg-drtg,1)
        net_s  = ("+"+str(net)) if net>0 else str(net)
        net_c  = "net-pos" if net>0 else "net-neg"
        pts    = avg_field(d,"pts"); reb_o = avg_field(d,"oreb"); reb_d = avg_field(d,"dreb")
        ast    = avg_field(d,"ast"); br    = avg_field(d,"br")
        stl    = avg_field(d,"stl"); blk   = avg_field(d,"blk")
        d2m    = avg_field(d,"d2m"); d2a   = avg_field(d,"d2a")
        przerw = avg_field(d,"przerw"); fd = avg_field(d,"fd")
        p2     = avg_pct_fields(d,"p2m","p2a"); p3 = avg_pct_fields(d,"p3m","p3a")
        ft     = avg_pct_fields(d,"ftm","fta"); eg = avg_efg(d); ts = avg_ts(d)
        poss   = avg_field(d,"poss"); ppp = round(pts/poss,2) if poss else 0
        for_druzyna[d] = {
            "pts":pts,"reb_o":reb_o,"reb_d":reb_d,"ast":ast,"br":br,"stl":stl,
            "d2m":d2m,"d2a":d2a,"przerw":przerw,"fd":fd,
            "p2":p2,"p3":p3,"ft":ft,"efg":eg,"ts":ts,"ortg":ortg,"ppp":ppp,"net":net_s,"net_c":net_c
        }

    def kpi_section(d, label, color):
        kd = for_druzyna[d]
        tag_bg = "#0F6E56" if d=="gtk" else "#A32D2D"
        return f'''
        <div>
          <div style="font-size:9px;text-transform:uppercase;letter-spacing:.5px;font-weight:500;color:{color};margin:0 0 4px">
            {label} <span style="background:{tag_bg};color:#fff;border-radius:8px;padding:1px 6px;font-size:8px;margin-left:4px">avg {n} mecze</span>
          </div>
          <div class="kpi-grid" style="grid-template-columns:repeat(5,1fr);margin-bottom:3px">
            {kpi_box(kd["pts"],"Punkty")}
            {kpi_box(f'{kd["reb_o"]}/{kd["reb_d"]}','Zb Off/Def')}
            {kpi_box(kd["ast"],"Asysty")}
            {kpi_box(kd["stl"],"Przechwyty")}
            {kpi_box(kd["br"],"Straty")}
          </div>
          <div class="kpi-grid" style="grid-template-columns:repeat(5,1fr)">
            {kpi_box(f'{kd["d2m"]}/{kd["d2a"]}','Dobitki M/A')}
            {kpi_box(kd["przerw"],"Przerwania")}
            {kpi_box(kd["fd"],"Faule wym.")}
            {kpi_box(kd["ppp"],"PPP")}
            {kpi_box(f'<span class="{kd["net_c"]}">{kd["net"]}</span>',"NETrtg","kpi-net")}
          </div>
        </div>'''

    # ── Porównanie paskami ────────────────────────────────────────────────────
    def _zbr_pct(tm, ta): return f"{tm/ta:.1%}" if ta else "—"
    def _zbr_efg(d):
        fga = d.get("p2a",0)+d.get("p3a",0); return f"{(d.get('p2m',0)+1.5*d.get('p3m',0))/fga:.1%}" if fga else "—"
    def _zbr_ts(d):
        fga=d.get("p2a",0)+d.get("p3a",0); fta=d.get("fta",0)
        return f"{d.get('pts',0)/(2*(fga+0.44*fta)):.1%}" if (fga+fta) else "—"
    def _zbr_topct(d):
        poss=max(d.get("poss",1),1); return f"{d.get('br',0)/poss:.1%}"

    # Sumy GTK i OPP z wszystkich meczów
    _sum_gtk = {k: sum(qv(ms,"gtk",k) for ms in all_match_stats.values())
                for k in ["p2m","p2a","p3m","p3a","ftm","fta","pts","br","poss"]}
    _sum_opp = {k: sum(qv(ms,"opp",k) for ms in all_match_stats.values())
                for k in ["p2m","p2a","p3m","p3a","ftm","fta","pts","br","poss"]}

    cmp_html = f"""
<table style="width:100%;border-collapse:collapse;font-size:8.5px;margin-bottom:8px">
  <thead>
    <tr style="background:#1a2b4a;color:#fff">
      <th style="padding:4px 6px;text-align:left">Drużyna</th>
      <th style="padding:4px 5px;text-align:center" colspan="2">2PT M/A</th>
      <th style="padding:4px 5px;text-align:center">2PT%</th>
      <th style="padding:4px 5px;text-align:center" colspan="2">3PT M/A</th>
      <th style="padding:4px 5px;text-align:center">3PT%</th>
      <th style="padding:4px 5px;text-align:center" colspan="2">FT M/A</th>
      <th style="padding:4px 5px;text-align:center">FT%</th>
      <th style="padding:4px 5px;text-align:center">eFG%</th>
      <th style="padding:4px 5px;text-align:center">TS%</th>
      <th style="padding:4px 5px;text-align:center">TO%</th>
    </tr>
  </thead>
  <tbody>
    <tr style="background:#f0fff4">
      <td style="padding:4px 6px;font-weight:700;color:#0F6E56">{gtk_name}</td>
      <td style="text-align:center">{_sum_gtk['p2m']}</td>
      <td style="text-align:center">{_sum_gtk['p2a']}</td>
      <td style="text-align:center;font-weight:700">{_zbr_pct(_sum_gtk['p2m'],_sum_gtk['p2a'])}</td>
      <td style="text-align:center">{_sum_gtk['p3m']}</td>
      <td style="text-align:center">{_sum_gtk['p3a']}</td>
      <td style="text-align:center;font-weight:700">{_zbr_pct(_sum_gtk['p3m'],_sum_gtk['p3a'])}</td>
      <td style="text-align:center">{_sum_gtk['ftm']}</td>
      <td style="text-align:center">{_sum_gtk['fta']}</td>
      <td style="text-align:center;font-weight:700">{_zbr_pct(_sum_gtk['ftm'],_sum_gtk['fta'])}</td>
      <td style="text-align:center;font-weight:700">{_zbr_efg(_sum_gtk)}</td>
      <td style="text-align:center;font-weight:700">{_zbr_ts(_sum_gtk)}</td>
      <td style="text-align:center">{_zbr_topct(_sum_gtk)}</td>
    </tr>
    <tr style="background:#fff5f5">
      <td style="padding:4px 6px;font-weight:700;color:#A32D2D">{opp_label}</td>
      <td style="text-align:center">{_sum_opp['p2m']}</td>
      <td style="text-align:center">{_sum_opp['p2a']}</td>
      <td style="text-align:center;font-weight:700">{_zbr_pct(_sum_opp['p2m'],_sum_opp['p2a'])}</td>
      <td style="text-align:center">{_sum_opp['p3m']}</td>
      <td style="text-align:center">{_sum_opp['p3a']}</td>
      <td style="text-align:center;font-weight:700">{_zbr_pct(_sum_opp['p3m'],_sum_opp['p3a'])}</td>
      <td style="text-align:center">{_sum_opp['ftm']}</td>
      <td style="text-align:center">{_sum_opp['fta']}</td>
      <td style="text-align:center;font-weight:700">{_zbr_pct(_sum_opp['ftm'],_sum_opp['fta'])}</td>
      <td style="text-align:center;font-weight:700">{_zbr_efg(_sum_opp)}</td>
      <td style="text-align:center;font-weight:700">{_zbr_ts(_sum_opp)}</td>
      <td style="text-align:center">{_zbr_topct(_sum_opp)}</td>
    </tr>
  </tbody>
</table>
"""

    # ── Zawodnicy GTK — średnie ───────────────────────────────────────────────
    from collections import defaultdict
    # Grupuj zawodników po player_id (gdy przypisany) lub nr koszulki
    # Klucz: player_id jeśli dostępny, else "nr_<nr>"
    player_totals = defaultdict(lambda: {"pts":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"ast":0,"oreb":0,"dreb":0,"br":0,"stl":0,"blk":0,"mecze":0})
    player_key_to_nr = {}   # klucz → nr (dla fallback nazwy)
    player_key_to_pid = {}  # klucz → player_id
    _seen_player_match = set()  # (mid, klucz) — raz per mecz
    for mid, pstats in all_player_stats.items():
        for ps in pstats:
            if ps["druzyna"] != "gtk": continue
            nr = ps["nr"]
            pid = ps.get("player_id") or ps.get("roster_id")
            # Klucz: preferuj player_id, fallback nr
            pkey = ("pid", pid) if pid else ("nr", nr)
            seen_key = (mid, pkey)
            if seen_key in _seen_player_match:
                continue
            _seen_player_match.add(seen_key)
            player_totals[pkey]["mecze"] += 1
            player_key_to_nr[pkey] = nr
            if pid: player_key_to_pid[pkey] = pid
            for f in ["pts","p2m","p2a","p3m","p3a","ftm","fta","ast","oreb","dreb","br","stl","blk"]:
                player_totals[pkey][f] += ps.get(f,0) or 0

    # Pobierz nazwy — po player_id lub nr
    pkey_names = {}
    all_pids = [v for k,v in player_key_to_pid.items()]
    all_nrs  = [v for k,v in player_key_to_nr.items() if k[0]=="nr"]
    if all_pids or all_nrs:
        # Najpierw po player_id
        if all_pids:
            cur.execute("""
                SELECT p.id, COALESCE(p.imie,'') as imie, COALESCE(p.nazwisko,'') as nazwisko
                FROM players p WHERE p.id = ANY(%s)
            """, (all_pids,))
            pid_to_name = {}
            for row in cur.fetchall():
                pi=row["imie"] or ""; pn=row["nazwisko"] or ""
                pid_to_name[row["id"]] = (pi[:1]+". "+pn) if pi else pn
            for pkey, pid in player_key_to_pid.items():
                if pid in pid_to_name:
                    pkey_names[pkey] = pid_to_name[pid]
        # Fallback: po nr z player_stats
        match_ids_list = list(all_match_stats.keys())
        if match_ids_list:
            cur.execute("""
                SELECT DISTINCT ON (ps.nr) ps.nr,
                       COALESCE(p.imie,r.imie,'') as imie,
                       COALESCE(p.nazwisko,r.nazwisko,'') as nazwisko
                FROM player_stats ps
                LEFT JOIN players p ON ps.player_id=p.id
                LEFT JOIN roster  r ON ps.roster_id=r.id
                WHERE ps.match_id = ANY(%s) AND ps.druzyna='gtk'
                  AND (p.id IS NOT NULL OR r.id IS NOT NULL)
                ORDER BY ps.nr
            """, (match_ids_list,))
            nr_to_name = {}
            for row in cur.fetchall():
                pi=row["imie"] or ""; pn=row["nazwisko"] or ""
                nr_to_name[row["nr"]] = (pi[:1]+". "+pn) if pi else pn
            for pkey in player_totals:
                if pkey not in pkey_names:
                    nr = player_key_to_nr.get(pkey)
                    if nr in nr_to_name:
                        pkey_names[pkey] = nr_to_name[nr]

    rows_players = []
    sorted_players = sorted(player_totals.items(), key=lambda x: x[1]["pts"], reverse=True)
    for pkey, tot in sorted_players:
        nr     = player_key_to_nr.get(pkey, "?")
        m_cnt  = tot["mecze"]   # liczba meczów TEGO zawodnika w wybranych spotkaniach
        name   = pkey_names.get(pkey, f"#{nr}")
        # Sumy (raw) — dzielone przez m_cnt zawodnika
        p2m=tot["p2m"]; p2a=tot["p2a"]; p3m=tot["p3m"]; p3a=tot["p3a"]
        ftm=tot["ftm"]; fta=tot["fta"]; pts_t=tot["pts"]
        ast_t=tot["ast"]; oreb_t=tot.get("oreb",0); dreb_t=tot.get("dreb",0)
        br_t=tot["br"]; stl_t=tot.get("stl",0); blk_t=tot.get("blk",0)
        fin_t=tot.get("finishes",0)
        fga = p2a + p3a
        # Średnie per mecz zawodnika
        def _avg(v): return round(v/m_cnt,1)
        pts_avg  = _avg(pts_t)
        p2m_avg  = round(p2m/m_cnt,1); p2a_avg = round(p2a/m_cnt,1)
        p3m_avg  = round(p3m/m_cnt,1); p3a_avg = round(p3a/m_cnt,1)
        ftm_avg  = round(ftm/m_cnt,1); fta_avg = round(fta/m_cnt,1)
        oreb_avg = _avg(oreb_t); dreb_avg = _avg(dreb_t)
        ast_avg  = _avg(ast_t); br_avg = _avg(br_t)
        stl_avg  = _avg(stl_t); blk_avg = _avg(blk_t); fin_avg = _avg(fin_t)
        # Procenty z sum (nie ze średnich)
        p2pct = f"{p2m/p2a:.0%}" if p2a else "—"
        p3pct = f"{p3m/p3a:.0%}" if p3a else "—"
        ftpct = f"{ftm/fta:.0%}" if fta else "—"
        efg_v = round((p2m+1.5*p3m)/fga*100,1) if fga else None
        efg_s = f"{efg_v:.0f}%" if efg_v is not None else "—"
        tsd   = 2*(fga+0.44*fta)
        ts_s  = f"{pts_t/tsd:.0%}" if tsd else "—"
        # USG% — team_poss z meczów zawodnika
        _p_match_ids = {m_id for m_id, mk in _seen_player_match if mk == pkey}
        team_poss_t = sum(qv(ms,"gtk","poss") for mid, ms in all_match_stats.items() if mid in _p_match_ids)
        usg_s = f"{(fga+0.44*fta+br_t)/max(team_poss_t,1):.0%}" if team_poss_t else "—"
        # MIN (szac.) — brak w zbiorczym (brak pliku Excel per mecz), pokazujemy —
        eg_c  = "color:#0F6E56" if efg_v and efg_v>=50 else ("color:#A32D2D" if efg_v and efg_v<35 else "")
        bg    = "background:#fafbfd" if len(rows_players)%2 else ""
        rows_players.append(f'''<tr style="{bg};font-size:7.5px">
          <td style="text-align:left;padding:2px 5px">{name}</td>
          <td style="text-align:center;color:#534AB7;padding:2px 3px">{m_cnt}/{n}</td>
          <td style="text-align:center;padding:2px 3px">—</td>
          <td style="text-align:center;font-weight:700;padding:2px 3px">{pts_avg}</td>
          <td style="text-align:center;padding:2px 3px">{p2m_avg}/{p2a_avg}</td>
          <td style="text-align:center;padding:2px 3px">{p2pct}</td>
          <td style="text-align:center;padding:2px 3px">{p3m_avg}/{p3a_avg}</td>
          <td style="text-align:center;padding:2px 3px">{p3pct}</td>
          <td style="text-align:center;padding:2px 3px">{ftm_avg}/{fta_avg}</td>
          <td style="text-align:center;padding:2px 3px">{ftpct}</td>
          <td style="text-align:center;padding:2px 3px">{oreb_avg}</td>
          <td style="text-align:center;padding:2px 3px">{dreb_avg}</td>
          <td style="text-align:center;padding:2px 3px">{ast_avg}</td>
          <td style="text-align:center;padding:2px 3px">{br_avg}</td>
          <td style="text-align:center;padding:2px 3px">{stl_avg}</td>
          <td style="text-align:center;padding:2px 3px">{blk_avg}</td>
          <td style="text-align:center;padding:2px 3px;{eg_c}">{efg_s}</td>
          <td style="text-align:center;padding:2px 3px">{ts_s}</td>
          <td style="text-align:center;padding:2px 3px">{usg_s}</td>
          <td style="text-align:center;padding:2px 3px">{fin_avg}</td>
        </tr>''')


    # ── Clutch IV kwarta ─────────────────────────────────────────────────────
    q4_pts_g = q4_pts_o = q4_efg_g = q4_efg_o = q4_poss_g = q4_to_g = 0
    q4_p2m=q4_p2a=q4_p3m=q4_p3a=q4_ftm=q4_fta=0
    q4_op2m=q4_op2a=q4_op3m=q4_op3a=q4_oftm=q4_ofta=0
    for mid, ms in all_match_stats.items():
        q4_pts_g += qv(ms,"gtk","pts",4); q4_pts_o += qv(ms,"opp","pts",4)
        q4_poss_g += qv(ms,"gtk","poss",4); q4_to_g += qv(ms,"gtk","br",4)
        q4_p2m+=qv(ms,"gtk","p2m",4); q4_p2a+=qv(ms,"gtk","p2a",4)
        q4_p3m+=qv(ms,"gtk","p3m",4); q4_p3a+=qv(ms,"gtk","p3a",4)
        q4_ftm+=qv(ms,"gtk","ftm",4); q4_fta+=qv(ms,"gtk","fta",4)
        q4_op2m+=qv(ms,"opp","p2m",4); q4_op2a+=qv(ms,"opp","p2a",4)
        q4_op3m+=qv(ms,"opp","p3m",4); q4_op3a+=qv(ms,"opp","p3a",4)
        q4_oftm+=qv(ms,"opp","ftm",4); q4_ofta+=qv(ms,"opp","fta",4)

    q4_efg_g = efg(q4_p2m,q4_p2a,q4_p3m,q4_p3a)
    q4_efg_o = efg(q4_op2m,q4_op2a,q4_op3m,q4_op3a)
    q4_efg_c  = "color:#0F6E56" if q4_efg_g>=50 else "color:#A32D2D"
    q4_avg_pts  = round(q4_pts_g/n,1); q4_avg_ptso = round(q4_pts_o/n,1)
    q4_avg_poss = round(q4_poss_g/n,1); q4_avg_to  = round(q4_to_g/n,1)

    # ── Bilans ───────────────────────────────────────────────────────────────
    losses = n - wins
    bilans = f"{wins}W–{losses}P"
    bilans_c = "#EF9F27" if wins==losses else ("#5DCAA5" if wins>losses else "#F09595")

    # ── Zbiorcze piątki ─────────────────────────────────────────────────────
    from collections import defaultdict as _dd
    zbr_off = _dd(lambda: {"pts":0,"poss":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"br":0,"oreb":0,"dreb":0,"ast":0,"stl":0,"blk":0,"mecze":0})
    zbr_def = _dd(lambda: {"pts":0,"poss":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"br":0,"oreb":0,"dreb":0,"ast":0,"stl":0,"blk":0,"mecze":0})
    zbr_nr_map = {}
    for mid_z, lu_data in all_lineups_data.items():
        _nr_map = build_nr_name_map(cur, mid_z)
        zbr_nr_map.update(_nr_map)
        for lu in lu_data["off"]:
            k = lu["lineup"]
            for _f in ["pts","poss","p2m","p2a","p3m","p3a","ftm","fta","br","oreb","dreb","ast","stl","blk"]:
                zbr_off[k][_f] += int(lu.get(_f,0) or 0)
            zbr_off[k]["mecze"] += 1
        for lu in lu_data["def"]:
            k = lu["lineup"]
            for _f in ["pts","poss","p2m","p2a","p3m","p3a","ftm","fta","br","oreb","dreb","ast","stl","blk"]:
                zbr_def[k][_f] += int(lu.get(_f,0) or 0)
            zbr_def[k]["mecze"] += 1

    zbr_off_rtg = {k: d["pts"]*100/d["poss"] for k,d in zbr_off.items() if d["poss"]>0}
    zbr_def_rtg = {k: d["pts"]*100/d["poss"] for k,d in zbr_def.items() if d["poss"]>0}
    zbr_net_list = []
    for k in set(zbr_off_rtg)|set(zbr_def_rtg):
        d = dict(zbr_off.get(k, zbr_def.get(k,{})))
        d["lineup"] = k
        d["ortg"] = zbr_off_rtg.get(k)
        d["drtg"] = zbr_def_rtg.get(k)
        net = (zbr_off_rtg[k]-zbr_def_rtg[k]) if (k in zbr_off_rtg and k in zbr_def_rtg) else None
        d["net_rtg"] = round(net,1) if net is not None else None
        d["poss"] = zbr_off.get(k,{}).get("poss", zbr_def.get(k,{}).get("poss",0))
        zbr_net_list.append(d)
    zbr_net_list.sort(key=lambda x: int(x.get("poss",0) or 0), reverse=True)

    def zbr_lu_name(lineup_str):
        return " · ".join(zbr_nr_map.get(str(nr), "#"+str(nr)) for nr in lineup_str.split("-"))

    def zbr_make_rows(data_dict, mode):
        rows = ""
        sorted_items = sorted(data_dict.items(), key=lambda x: int(x[1].get("poss",0) or 0), reverse=True)
        for i,(k,d) in enumerate(sorted_items):
            p2m=d["p2m"]; p2a=d["p2a"]; p3m=d["p3m"]; p3a=d["p3a"]
            ftm=d["ftm"]; fta=d["fta"]; pts=d["pts"]; poss=d["poss"]
            br=d["br"]; m_cnt=d["mecze"]
            oreb=d.get("oreb",0) or 0; dreb=d.get("dreb",0) or 0
            ast=d.get("ast",0) or 0; stl=d.get("stl",0) or 0; blk=d.get("blk",0) or 0
            fga=p2a+p3a
            efg_v=round((p2m+1.5*p3m)/fga*100) if fga else None
            ppp_v=pts/poss if poss else None
            ts_denom=2*(fga+0.44*fta)
            ts_v=round(pts/ts_denom*100,1) if ts_denom else None
            efg_s=(str(efg_v)+"%") if efg_v is not None else "—"
            ts_s=(str(ts_v)+"%") if ts_v is not None else "—"
            ppp_s=("%.2f"%ppp_v) if ppp_v is not None else "—"
            p2pct=("%.0f%%"%(p2m/p2a*100)) if p2a else "—"
            p3pct=("%.0f%%"%(p3m/p3a*100)) if p3a else "—"
            ftpct=("%.0f%%"%(ftm/fta*100)) if fta else "—"
            if mode=="off":
                ppp_c="#0F6E56" if ppp_v and ppp_v>=0.9 else ("#A32D2D" if ppp_v and ppp_v<0.7 else "#444")
                efg_c="#0F6E56" if efg_v and efg_v>=50 else ("#A32D2D" if efg_v and efg_v<35 else "#444")
                ts_c="#0F6E56" if ts_v and ts_v>=50 else ("#A32D2D" if ts_v and ts_v<35 else "#444")
            else:
                ppp_c="#0F6E56" if ppp_v and ppp_v<0.7 else ("#A32D2D" if ppp_v and ppp_v>=0.9 else "#444")
                efg_c="#0F6E56" if efg_v and efg_v<35 else ("#A32D2D" if efg_v and efg_v>=50 else "#444")
                ts_c="#0F6E56" if ts_v and ts_v<35 else ("#A32D2D" if ts_v and ts_v>=50 else "#444")
            bg="#f5f8ff" if i%2==0 else "#fff"
            br_c="#A32D2D" if br>=4 else "#444"
            name_html=zbr_lu_name(k)+(' <span style="font-size:7px;color:#888">('+str(m_cnt)+'M)</span>' if m_cnt<n else "")
            rows += (
                '<tr style="background:'+bg+'">'
                '<td style="font-size:7.5px;text-align:left;padding:2px 5px">'+name_html+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+str(poss)+'</td>'
                '<td style="text-align:center;font-weight:700;color:#1a2b4a;padding:2px 4px">'+str(pts)+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+str(p2m)+"/"+str(p2a)+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+p2pct+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+str(p3m)+"/"+str(p3a)+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+p3pct+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+str(ftm)+"/"+str(fta)+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+ftpct+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+str(oreb)+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+str(dreb)+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+str(ast)+'</td>'
                '<td style="text-align:center;padding:2px 4px;color:'+br_c+'">'+str(br)+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+str(stl)+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+str(blk)+'</td>'
                '<td style="text-align:center;padding:2px 4px;color:'+efg_c+'">'+efg_s+'</td>'
                '<td style="text-align:center;padding:2px 4px;color:'+ts_c+'">'+ts_s+'</td>'
                '<td style="text-align:center;font-weight:700;padding:2px 4px;color:'+ppp_c+'">'+ppp_s+'</td>'
                '</tr>'
            )
        return rows or '<tr><td colspan="18" style="text-align:center;color:#888;padding:6px">Brak danych</td></tr>'


    def zbr_net_rows_fn():
        rows=""
        for i,lu in enumerate(zbr_net_list):
            poss=int(lu.get("poss",0) or 0); pts=int(lu.get("pts",0) or 0)
            ortg=lu.get("ortg"); drtg=lu.get("drtg"); net=lu.get("net_rtg")
            ortg_s=("%.1f"%ortg) if ortg is not None else "—"
            drtg_s=("%.1f"%drtg) if drtg is not None else "—"
            net_s=(("%+.1f"%net) if net is not None else "—")
            net_c="#0F6E56" if net and net>0 else ("#A32D2D" if net and net<0 else "#888")
            bg="#f5f8ff" if i%2==0 else "#fff"
            rows+=(
                '<tr style="background:'+bg+'">'
                '<td style="font-size:7.5px;text-align:left;padding:2px 5px">'+zbr_lu_name(lu["lineup"])+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+str(poss)+'</td>'
                '<td style="text-align:center;padding:2px 4px">'+str(pts)+'</td>'
                '<td style="text-align:center;padding:2px 4px;color:#0F6E56">'+ortg_s+'</td>'
                '<td style="text-align:center;padding:2px 4px;color:#A32D2D">'+drtg_s+'</td>'
                '<td style="text-align:center;font-weight:700;padding:2px 4px;color:'+net_c+'">'+net_s+'</td>'
                '</tr>'
            )
        return rows or '<tr><td colspan="6" style="text-align:center;color:#888;padding:6px">Brak danych</td></tr>'

    _th  = 'background:#1a2b4a;color:#fff;font-size:8px;font-weight:700;padding:3px 3px;text-align:center;white-space:nowrap'
    _thl = 'background:#1a2b4a;color:#fff;font-size:8px;font-weight:700;padding:3px 5px;text-align:left;white-space:nowrap'
    _thg = 'background:#1a2b4a;color:rgba(255,255,255,.55);font-size:7px;letter-spacing:.3px;padding:3px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center'
    _ths = 'background:#1a2b4a;color:rgba(255,255,255,.8);font-size:8px;font-weight:500;padding:2px 3px 4px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center'
    _thz = 'background:#152236;color:rgba(255,255,255,.55);font-size:7px;letter-spacing:.3px;padding:3px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.1);text-align:center'
    _thzs= 'background:#152236;color:rgba(255,255,255,.75);font-size:8px;font-weight:500;padding:2px 3px 4px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center'
    _than= 'background:#633806;color:#FAC775;font-size:8px;font-weight:700;padding:3px 3px;text-align:center'
    _vm  = 'vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.2)'
    _tbl = 'width:100%;border-collapse:collapse;font-size:8.5px;table-layout:fixed;margin-bottom:4px'
    _sec = 'font-size:9px;text-transform:uppercase;letter-spacing:.5px;font-weight:700;color:#1a2b4a;margin:10px 0 3px;padding-bottom:2px;border-bottom:1px solid #e0e0e0'
    _sub = 'font-size:7.5px;color:#888;margin-bottom:3px'

    _cg = ("<colgroup>"
        "<col style=\'width:auto\'>"
        "<col style=\'width:28px\'><col style=\'width:26px\'>"
        "<col style=\'width:32px\'><col style=\'width:24px\'>"
        "<col style=\'width:32px\'><col style=\'width:24px\'>"
        "<col style=\'width:32px\'><col style=\'width:24px\'>"
        "<col style=\'width:20px\'><col style=\'width:20px\'>"
        "<col style=\'width:22px\'><col style=\'width:22px\'>"
        "<col style=\'width:22px\'><col style=\'width:22px\'>"
        "<col style=\'width:28px\'><col style=\'width:28px\'><col style=\'width:28px\'>"
        "</colgroup>")
    _hdr_html = (
        "<thead><tr>"
        "<th style=\'"+_thl+";"+_vm+"\'  rowspan=\'3\'>Sk&#322;ad</th>"
        "<th style=\'"+_th+";"+_vm+"\'   rowspan=\'3\'>POSS</th>"
        "<th style=\'"+_th+";"+_vm+"\'   rowspan=\'3\'>PKT</th>"
        "<th style=\'"+_thg+"\'  colspan=\'2\'>2PT</th>"
        "<th style=\'"+_thg+"\'  colspan=\'2\'>3PT</th>"
        "<th style=\'"+_thg+"\'  colspan=\'2\'>FT</th>"
        "<th style=\'"+_thz+"\'  colspan=\'2\'>ZB</th>"
        "<th style=\'"+_th+";"+_vm+"\'   rowspan=\'3\'>AST</th>"
        "<th style=\'"+_th+";"+_vm+"\'   rowspan=\'3\'>TO</th>"
        "<th style=\'"+_th+";"+_vm+"\'   rowspan=\'3\'>STL</th>"
        "<th style=\'"+_th+";"+_vm+"\'   rowspan=\'3\'>BLK</th>"
        "<th style=\'"+_th+";"+_vm+"\'   rowspan=\'3\'>eFG%</th>"
        "<th style=\'"+_th+";"+_vm+"\'   rowspan=\'3\'>TS%</th>"
        "<th style=\'"+_th+";"+_vm+"\'   rowspan=\'3\'>PPP</th>"
        "</tr><tr>"
        "<th style=\'"+_ths+"\'>M/A</th><th style=\'"+_ths+"\'>%</th>"
        "<th style=\'"+_ths+"\'>M/A</th><th style=\'"+_ths+"\'>%</th>"
        "<th style=\'"+_ths+"\'>M/A</th><th style=\'"+_ths+"\'>%</th>"
        "<th style=\'"+_thzs+"\'>A</th><th style=\'"+_thzs+"\'>O</th>"
        "</tr></thead>"
    )
    _net_hdr_html = (
        "<thead><tr>"
        "<th style=\'"+_thl+";"+_vm+"\'              rowspan=\'2\'>Sk\u0142ad</th>"
        "<th style=\'"+_th+";"+_vm+"\'               rowspan=\'2\'>POSS</th>"
        "<th style=\'"+_th+";"+_vm+"\'               rowspan=\'2\'>PKT</th>"
        "<th style=\'"+_th+";color:#9FE1CB;"+_vm+"\' rowspan=\'2\'>ORtg</th>"
        "<th style=\'"+_th+";color:#F09595;"+_vm+"\' rowspan=\'2\'>DRtg</th>"
        "<th style=\'"+_than+";"+_vm+"\'             rowspan=\'2\'>Net RTG</th>"
        "</tr></thead>"
    )

    zbr_lineup_html = (
        "<div style=\'"+_sec+"\'>Pi\u0105tki \u2014 Atak (OFF) <span style=\'background:#0F6E56;color:#fff;border-radius:8px;padding:1px 6px;font-size:7px;margin-left:4px\'>sumy "+str(n)+" mecz\u00f3w</span></div>"
        "<div style=\'"+_sub+"\'>PPP: <span style=\'color:#0F6E56\'>&ge;0.90 dobry</span> / <span style=\'color:#A32D2D\'>&lt;0.70 s&#322;aby</span> &middot; sortowanie: POSS malej&#261;co</div>"
        "<table style=\'"+_tbl+"\'>"+_cg+_hdr_html+"<tbody>"+zbr_make_rows(zbr_off,"off")+"</tbody></table>"
        "<div style=\'"+_sec+";margin-top:12px\'>Pi\u0105tki \u2014 Obrona (DEF) <span style=\'background:#A32D2D;color:#fff;border-radius:8px;padding:1px 6px;font-size:7px;margin-left:4px\'>sumy "+str(n)+" mecz\u00f3w</span></div>"
        "<div style=\'"+_sub+"\'>PPP rywala: <span style=\'color:#0F6E56\'>&lt;0.70 dobry</span> / <span style=\'color:#A32D2D\'>&ge;0.90 s&#322;aby</span> &middot; sortowanie: POSS malej&#261;co</div>"
        "<table style=\'"+_tbl+"\'>"+_cg+_hdr_html+"<tbody>"+zbr_make_rows(zbr_def,"def")+"</tbody></table>"
        "<div style=\'"+_sec+";margin-top:12px\'>Pi\u0105tki \u2014 Net RTG <span style=\'font-size:7.5px;font-weight:400;color:#888\'>(ORtg OFF \u2212 DRtg DEF)</span></div>"
        "<div style=\'"+_sub+"\'>sortowanie: POSS malej&#261;co</div>"
        "<table style=\'"+_tbl+";table-layout:fixed\'>"
        "<colgroup><col style=\'width:auto\'><col style=\'width:36px\'><col style=\'width:36px\'>"
        "<col style=\'width:48px\'><col style=\'width:48px\'><col style=\'width:56px\'></colgroup>"
        +_net_hdr_html+"<tbody>"+zbr_net_rows_fn()+"</tbody></table>"
    )

    # ── HTML raportu ─────────────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="pl"><head><meta charset="UTF-8">
<style>
*{{box-sizing:border-box;margin:0;padding:0;font-family:Arial,sans-serif}}
body{{font-size:10px;color:#222;padding:16px;max-width:900px;margin:0 auto}}
.hero{{background:#1a2b4a;padding:10px 16px;border-radius:8px;margin-bottom:12px;display:flex;justify-content:space-between;align-items:center}}
.hero h2{{font-size:13px;font-weight:700;color:#fff;margin-bottom:2px}}
.hero .meta{{font-size:8px;color:#a8c4e0}}
.sec-title{{font-size:9px;text-transform:uppercase;letter-spacing:.5px;font-weight:700;color:#1a2b4a;margin:10px 0 4px;padding-bottom:3px;border-bottom:1px solid #e0e0e0}}
table{{width:100%;border-collapse:collapse;font-size:9px;margin-bottom:8px}}
th{{padding:4px 6px;color:#fff;font-size:8px;font-weight:700;text-align:center;background:#1a2b4a}}
th.left{{text-align:left}}
td{{padding:3px 6px;border-bottom:0.5px solid #eee;text-align:center}}
td.left{{text-align:left}}
.sep{{height:1px;background:#eee;margin:8px 0}}
.two-col{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:8px}}
.kpi-grid{{display:grid;grid-template-columns:repeat(7,minmax(0,1fr));gap:3px;margin-bottom:4px}}
.kpi-box{{background:#f5f8ff;border-radius:3px;padding:4px 2px;text-align:center}}
.kpi-val{{font-size:10px;font-weight:700;color:#1a2b4a}}
.kpi-lbl{{font-size:7px;color:#888;text-transform:uppercase;margin-top:1px}}
.kpi-net{{background:#EEEDFE;border:0.5px solid #AFA9EC}}
.kpi-net .kpi-val{{color:#3C3489}}
.net-pos{{color:#0F6E56}}.net-neg{{color:#A32D2D}}
.clutch-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:8px}}
.c-box{{background:#f5f8ff;border-radius:4px;padding:8px;text-align:center}}
.c-val{{font-size:14px;font-weight:700;color:#1a2b4a}}
.c-lbl{{font-size:8px;color:#888;text-transform:uppercase;margin-top:2px}}
@media print {{
  @page {{ margin: 12mm 10mm; size: A4 landscape; }}
  body {{ padding: 8px; max-width: 100%; }}
  .hero {{ -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; background: #1a2b4a !important; }}
  .hero * {{ -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }}
  th, .kpi-box, .kpi-net, .c-box {{ -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }}
  .no-print {{ display: none !important; }}
  canvas {{ display: block !important; }}
  .cmp-row {{ break-inside: avoid; }}
  .sec-title {{ break-after: avoid; }}
  tr {{ break-inside: avoid; }}
}}
</style></head><body>

<div class="hero" style="display:grid;grid-template-columns:1fr auto 1fr;align-items:center;gap:20px">

  <div>
    <div style="font-size:8px;color:rgba(255,255,255,.4);text-transform:uppercase;letter-spacing:.7px;margin-bottom:4px">{gtk_name}</div>
    <div style="font-size:26px;font-weight:500;color:#fff;line-height:1">{avg_gtk}</div>
    <div style="font-size:8px;color:rgba(255,255,255,.35);margin-top:3px">śr. punktów per mecz</div>
  </div>

  <div style="text-align:center;padding:0 20px;border-left:1px solid rgba(255,255,255,.1);border-right:1px solid rgba(255,255,255,.1)">
    <div style="font-size:8px;color:rgba(255,255,255,.35);text-transform:uppercase;letter-spacing:.7px;margin-bottom:5px">Analiza serii · Sezon {mecze_data[0].get("sezon","") if mecze_data else ""}</div>
    <div style="font-size:9px;color:rgba(255,255,255,.5);margin-bottom:5px">{gtk_name} vs {opp_label}</div>
    <div style="display:flex;gap:6px;justify-content:center;margin-bottom:5px">
      {"".join(
        '<div style="width:22px;height:22px;border-radius:50%;background:#0F6E56;display:flex;align-items:center;justify-content:center;font-size:9px;font-weight:500;color:#fff">W</div>'
        if m.get("wynik_gtk",0) > m.get("wynik_opp",0) else
        '<div style="width:22px;height:22px;border-radius:50%;background:#A32D2D;display:flex;align-items:center;justify-content:center;font-size:9px;font-weight:500;color:#fff">P</div>'
        for m in mecze_data
      )}
    </div>
    <div style="font-size:18px;font-weight:500;letter-spacing:2px;color:{bilans_c}">{bilans}</div>
    <div style="font-size:8px;color:rgba(255,255,255,.3);margin-top:3px">{n} {"mecz" if n==1 else "mecze" if 2<=n<=4 else "meczów"}</div>
  </div>

  <div style="text-align:right">
    <div style="font-size:8px;color:rgba(255,255,255,.4);text-transform:uppercase;letter-spacing:.7px;margin-bottom:4px">{opp_label}</div>
    <div style="font-size:26px;font-weight:500;color:#fff;line-height:1">{avg_opp}</div>
    <div style="font-size:8px;color:rgba(255,255,255,.35);margin-top:3px">śr. punktów per mecz</div>
  </div>

</div>

<div class="sec-title">Zestawienie meczów</div>
<table>
  <thead><tr>
    <th class="left">Data</th><th class="left">Rywal</th><th>Miejsce</th>
    <th style="background:#0F6E56">1Q</th><th style="background:#0F6E56">2Q</th>
    <th style="background:#0F6E56">3Q</th><th style="background:#0F6E56">4Q</th>
    <th>GTK</th><th>OPP</th><th>W/P</th>
  </tr></thead>
  <tbody>
    {"".join(rows_mecze)}
    {row_suma}
    {row_avg}
  </tbody>
</table>

<div class="sep"></div>

<div class="two-col">
  {kpi_section("gtk", gtk_name, "#0F6E56")}
  {kpi_section("opp", opp_label, "#A32D2D")}
</div>

<div class="sep"></div>

<div style="background:#fafbfd;border-radius:6px;padding:8px 12px;margin-bottom:10px">
  <div style="text-align:center;font-size:9px;font-weight:700;color:#1a2b4a;margin-bottom:6px">Porównanie bezpośrednie — średnie z {n} meczów</div>
  <div style="display:flex;justify-content:space-between;font-size:8px;font-weight:700;padding:0 42px;margin-bottom:4px">
    <span style="color:#0F6E56">{gtk_name}</span><span style="color:#A32D2D">{opp_label}</span>
  </div>
  {cmp_html}
</div>

<div class="sec-title">Zawodnicy GTK — średnie per mecz <span style="background:#534AB7;color:#fff;border-radius:8px;padding:1px 6px;font-size:7px;margin-left:4px">{n} mecze</span></div>
<table style="table-layout:fixed;width:100%">
  <thead>
    <tr style="background:#1a2b4a;color:#fff;font-size:7px">
      <th style="text-align:left;padding:3px 4px" rowspan="2">Zawodnik</th>
      <th style="text-align:center;padding:3px 2px;color:#534AB7;background:#26215C" rowspan="2">M</th>
      <th style="text-align:center;padding:3px 2px" rowspan="2">MIN<br>(szac.)</th>
      <th style="text-align:center;padding:3px 2px" rowspan="2">PTS</th>
      <th style="text-align:center;padding:3px 2px;background:#152236;color:rgba(255,255,255,.55)" colspan="2">2PT</th>
      <th style="text-align:center;padding:3px 2px;background:#152236;color:rgba(255,255,255,.55)" colspan="2">3PT</th>
      <th style="text-align:center;padding:3px 2px;background:#152236;color:rgba(255,255,255,.55)" colspan="2">FT</th>
      <th style="text-align:center;padding:3px 2px;background:#0d1b2e;color:rgba(255,255,255,.45)" colspan="2">ZB</th>
      <th style="text-align:center;padding:3px 2px" rowspan="2">AST</th>
      <th style="text-align:center;padding:3px 2px" rowspan="2">TO</th>
      <th style="text-align:center;padding:3px 2px" rowspan="2">STL</th>
      <th style="text-align:center;padding:3px 2px" rowspan="2">BLK</th>
      <th style="text-align:center;padding:3px 2px" rowspan="2">eFG%</th>
      <th style="text-align:center;padding:3px 2px" rowspan="2">TS%</th>
      <th style="text-align:center;padding:3px 2px" rowspan="2">USG%</th>
      <th style="text-align:center;padding:3px 2px" rowspan="2">FIN</th>
    </tr>
    <tr style="background:#1a2b4a;color:rgba(255,255,255,.65);font-size:6px">
      <th style="text-align:center;padding:2px;background:#152236">M/A</th>
      <th style="text-align:center;padding:2px;background:#152236">%</th>
      <th style="text-align:center;padding:2px;background:#152236">M/A</th>
      <th style="text-align:center;padding:2px;background:#152236">%</th>
      <th style="text-align:center;padding:2px;background:#152236">M/A</th>
      <th style="text-align:center;padding:2px;background:#152236">%</th>
      <th style="text-align:center;padding:2px;background:#0d1b2e">A</th>
      <th style="text-align:center;padding:2px;background:#0d1b2e">O</th>
    </tr>
  </thead>
  <tbody>{"".join(rows_players)}</tbody>
</table>

<div class="sep"></div>

<div class="sec-title">Clutch — IV kwarta <span style="background:#534AB7;color:#fff;border-radius:8px;padding:1px 6px;font-size:7px;margin-left:4px">{n} mecze</span></div>

<div style="display:grid;grid-template-columns:repeat(2,1fr);gap:10px;margin-bottom:10px">
  <div>
    <div style="font-size:8px;font-weight:700;color:#0F6E56;text-transform:uppercase;letter-spacing:.3px;margin-bottom:5px">{gtk_name}</div>
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:8px">
      <div class="c-box"><div class="c-val">{q4_avg_pts}</div><div class="c-lbl">PKT Q4</div></div>
      <div class="c-box"><div class="c-val" style="{q4_efg_c}">{q4_efg_g}%</div><div class="c-lbl">eFG% Q4</div></div>
      <div class="c-box"><div class="c-val">{q4_avg_poss}</div><div class="c-lbl">POSS Q4</div></div>
      <div class="c-box"><div class="c-val" style="{'color:#A32D2D' if q4_avg_to>=5 else ''}">{q4_avg_to}</div><div class="c-lbl">TO Q4</div></div>
    </div>
    <table>
      <thead><tr>
        <th class="left" style="background:#0F6E56">Rzuty Q4</th>
        <th style="background:#0F6E56">M/A</th>
        <th style="background:#0F6E56">%</th>
      </tr></thead>
      <tbody>
        <tr><td class="left">2PT</td>
          <td>{round(q4_p2m/n,1)}/{round(q4_p2a/n,1)}</td>
          <td>{"—" if not q4_p2a else str(round(q4_p2m/q4_p2a*100))+"%"}</td></tr>
        <tr style="background:#fafbfd"><td class="left">3PT</td>
          <td>{round(q4_p3m/n,1)}/{round(q4_p3a/n,1)}</td>
          <td>{"—" if not q4_p3a else str(round(q4_p3m/q4_p3a*100))+"%"}</td></tr>
        <tr><td class="left">FT</td>
          <td>{round(q4_ftm/n,1)}/{round(q4_fta/n,1)}</td>
          <td>{"—" if not q4_fta else str(round(q4_ftm/q4_fta*100))+"%"}</td></tr>
        <tr style="background:#fafbfd"><td class="left" style="color:#A32D2D">Straty</td>
          <td style="color:#A32D2D" colspan="2">{q4_avg_to}</td></tr>
      </tbody>
    </table>
  </div>
  <div>
    <div style="font-size:8px;font-weight:700;color:#A32D2D;text-transform:uppercase;letter-spacing:.3px;margin-bottom:5px">{opp_label}</div>
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:8px">
      <div class="c-box"><div class="c-val">{q4_avg_ptso}</div><div class="c-lbl">PKT Q4</div></div>
      <div class="c-box"><div class="c-val" style="{'color:#A32D2D' if q4_efg_o>=50 else 'color:#0F6E56'}">{q4_efg_o}%</div><div class="c-lbl">eFG% Q4</div></div>
      <div class="c-box"><div class="c-val">{round(qv(list(all_match_stats.values())[0],"opp","poss",4) if all_match_stats else 0,1)}</div><div class="c-lbl">POSS Q4</div></div>
      <div class="c-box"><div class="c-val">{round(sum(qv(ms,"opp","br",4) for ms in all_match_stats.values())/n,1)}</div><div class="c-lbl">TO Q4</div></div>
    </div>
    <table>
      <thead><tr>
        <th class="left" style="background:#A32D2D">Rzuty Q4</th>
        <th style="background:#A32D2D">M/A</th>
        <th style="background:#A32D2D">%</th>
      </tr></thead>
      <tbody>
        <tr><td class="left">2PT</td>
          <td>{round(q4_op2m/n,1)}/{round(q4_op2a/n,1)}</td>
          <td>{"—" if not q4_op2a else str(round(q4_op2m/q4_op2a*100))+"%"}</td></tr>
        <tr style="background:#fafbfd"><td class="left">3PT</td>
          <td>{round(q4_op3m/n,1)}/{round(q4_op3a/n,1)}</td>
          <td>{"—" if not q4_op3a else str(round(q4_op3m/q4_op3a*100))+"%"}</td></tr>
        <tr><td class="left">FT</td>
          <td>{round(q4_oftm/n,1)}/{round(q4_ofta/n,1)}</td>
          <td>{"—" if not q4_ofta else str(round(q4_oftm/q4_ofta*100))+"%"}</td></tr>
        <tr style="background:#fafbfd"><td class="left" style="color:#A32D2D">Straty</td>
          <td style="color:#A32D2D" colspan="2">{round(sum(qv(ms,"opp","br",4) for ms in all_match_stats.values())/n,1)}</td></tr>
      </tbody>
    </table>
  </div>
</div>

<div style="text-align:center;font-size:7px;color:#aaa;border-top:0.5px solid #eee;padding-top:6px;margin-top:8px">
{zbr_lineup_html}

<div style="text-align:center;font-size:7px;color:#aaa;border-top:0.5px solid #eee;padding-top:6px;margin-top:8px">
  Basket Kołcz Analytics · Raport serii · {gtk_name} vs {opp_label} · {n} {"mecz" if n==1 else "mecze" if 2<=n<=4 else "meczów"}
</div>

<script>
window.addEventListener('beforeprint', function() {{
  document.querySelectorAll('.hero, .hero *, th, .kpi-box, .kpi-net, .c-box, .cmp-g, .cmp-r').forEach(function(el) {{
    el.style.webkitPrintColorAdjust = 'exact';
    el.style.printColorAdjust = 'exact';
    el.style.colorAdjust = 'exact';
  }});
}});
</script>

</body></html>"""

    cur.close()
    return Response(html, mimetype="text/html",
                    headers={"Content-Disposition": f'inline; filename="raport_zbiorczy_{n}_meczow.html"'})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)


# ══════════════════════════════════════════════════════════════════════════════
# PORÓWNYWARKA MECZÓW I DRUŻYN
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/porownaj")
@login_required
def porownaj():
    """Strona konfiguracji porównania."""
    db = get_db(); cur = db.cursor()

    # Dane do dropdownów — mecze
    cur.execute("""
        SELECT m.id, m.data_meczu, m.przeciwnik, m.wynik_gtk, m.wynik_opp,
               COALESCE(m.team_name_a, m.nazwa_gtk, '') as nazwa_a,
               COALESCE(m.sezon,'') as sezon
        FROM matches m ORDER BY m.data_meczu DESC NULLS LAST LIMIT 200
    """)
    mecze = cur.fetchall()

    # Dane do dropdownów — kluby/sezony/drużyny
    cur.execute("""
        SELECT c.id as cid, c.name as cname, s.id as sid, s.name as sname,
               t.id as tid, t.name as tname
        FROM clubs c JOIN seasons s ON s.club_id=c.id JOIN teams t ON t.season_id=s.id
        ORDER BY c.name, s.name, t.name
    """)
    tree_rows = cur.fetchall()
    cur.close()

    # Mecze z URL (pre-selected)
    ids_param = request.args.get("ids","")
    presel = [x for x in ids_param.split(",") if x.isdigit()][:4]

    mecze_opts = "".join(
        f'<option value="{m["id"]}" {"selected" if str(m["id"]) in presel[:1] else ""}>'
        f'{m["data_meczu"].strftime("%d.%m.%Y") if m["data_meczu"] else "—"} · '
        f'{m["nazwa_a"]} vs {m["przeciwnik"]} ({m["wynik_gtk"]}:{m["wynik_opp"]})'
        f'</option>'
        for m in mecze
    )
    mecze_opts_b = "".join(
        f'<option value="{m["id"]}" {"selected" if str(m["id"]) in presel[1:2] else ""}>'
        f'{m["data_meczu"].strftime("%d.%m.%Y") if m["data_meczu"] else "—"} · '
        f'{m["nazwa_a"]} vs {m["przeciwnik"]} ({m["wynik_gtk"]}:{m["wynik_opp"]})'
        f'</option>'
        for m in mecze
    )

    team_opts = "".join(
        f'<option value="{r["tid"]}">{r["cname"]} · {r["sname"]} · {r["tname"]}</option>'
        for r in tree_rows
    )

    content = f"""
<div class="page-title">&#9654; Porównywarka meczów i drużyn</div>

<div class="card p-3 mb-3">
  <div style="font-size:.82rem;color:#666;margin-bottom:14px">
    Wybierz tryb porównania, skonfiguruj obie strony i wygeneruj raport PDF.
  </div>

  <!-- Tryb -->
  <div style="display:flex;gap:0;border:0.5px solid #dee2e6;border-radius:8px;overflow:hidden;margin-bottom:18px;width:fit-content">
    <button id="tab-mecze" onclick="switchTab('mecze')"
      style="padding:8px 20px;border:none;background:#1a2b4a;color:#fff;font-size:.82rem;font-weight:500;cursor:pointer">
      Dwa mecze
    </button>
    <button id="tab-druzyny" onclick="switchTab('druzyny')"
      style="padding:8px 20px;border:none;background:#f8f9fa;color:#666;font-size:.82rem;cursor:pointer">
      Dwie drużyny / sezony
    </button>
    <button id="tab-zbiorcze" onclick="switchTab('zbiorcze')"
      style="padding:8px 20px;border:none;background:#f8f9fa;color:#666;font-size:.82rem;cursor:pointer">
      Raport zbiorczy
    </button>
  </div>

  <!-- Tryb: 2 mecze -->
  <div id="pane-mecze">
    <form method="GET" action="/porownaj/pdf">
      <input type="hidden" name="tryb" value="mecze">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px">
        <div>
          <div style="font-size:.75rem;font-weight:600;color:#185FA5;text-transform:uppercase;letter-spacing:.4px;margin-bottom:6px">&#9632; Mecz A</div>
          <select name="mecz_a" class="form-select form-select-sm">
            <option value="">— wybierz mecz —</option>
            {mecze_opts}
          </select>
        </div>
        <div>
          <div style="font-size:.75rem;font-weight:600;color:#BA7517;text-transform:uppercase;letter-spacing:.4px;margin-bottom:6px">&#9632; Mecz B</div>
          <select name="mecz_b" class="form-select form-select-sm">
            <option value="">— wybierz mecz —</option>
            {mecze_opts_b}
          </select>
        </div>
      </div>
      <button type="submit" class="btn btn-sm fw-bold"
        style="background:#534AB7;color:#fff;border:none;padding:7px 20px">
        Generuj raport PDF &#8594;
      </button>
    </form>
  </div>

  <!-- Tryb: 2 drużyny -->
  <div id="pane-druzyny" style="display:none">
    <form method="GET" action="/porownaj/pdf">
      <input type="hidden" name="tryb" value="druzyny">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px">
        <div>
          <div style="font-size:.75rem;font-weight:600;color:#185FA5;text-transform:uppercase;letter-spacing:.4px;margin-bottom:6px">&#9632; Drużyna A</div>
          <select name="team_a" class="form-select form-select-sm">
            <option value="">— wybierz drużynę —</option>
            {team_opts}
          </select>
          <div style="font-size:.72rem;color:#888;margin-top:4px">Agregat ze wszystkich meczów tej drużyny</div>
        </div>
        <div>
          <div style="font-size:.75rem;font-weight:600;color:#BA7517;text-transform:uppercase;letter-spacing:.4px;margin-bottom:6px">&#9632; Drużyna B</div>
          <select name="team_b" class="form-select form-select-sm">
            <option value="">— wybierz drużynę —</option>
            {team_opts}
          </select>
          <div style="font-size:.72rem;color:#888;margin-top:4px">Agregat ze wszystkich meczów tej drużyny</div>
        </div>
      </div>
      <button type="submit" class="btn btn-sm fw-bold"
        style="background:#534AB7;color:#fff;border:none;padding:7px 20px">
        Generuj raport PDF &#8594;
      </button>
    </form>
  </div>
</div>

  <!-- Raport zbiorczy -->
  <div id="pane-zbiorcze" style="display:none">
    <p style="font-size:.82rem;color:#666;margin-bottom:10px">Zaznacz mecze do raportu zbiorczego. Raport zawiera: tabelę wyników, uśrednione KPI GTK i OPP, średnie zawodników oraz statystyki clutch IV kwarty.</p>
    <form method="GET" action="/porownaj/zbiorcze">
      <div style="border:0.5px solid #e0e0e0;border-radius:6px;overflow:hidden;margin-bottom:12px">
        <div style="background:#1a2b4a;padding:6px 10px;font-size:.78rem;color:#fff;font-weight:500;display:flex;justify-content:space-between;align-items:center">
          <span>Dostępne mecze</span>
          <label style="font-size:.75rem;font-weight:400;cursor:pointer"><input type="checkbox" id="zb-all" onchange="zbAllToggle(this)" style="margin-right:4px">Zaznacz wszystkie</label>
        </div>
        <div id="zbiorcze-list" style="padding:6px 8px;max-height:280px;overflow-y:auto"></div>
      </div>
      <button type="submit" style="font-size:.82rem;font-weight:500;cursor:pointer;padding:7px 20px;background:#534AB7;color:#fff;border:none;border-radius:6px">
        Generuj raport zbiorczy PDF &#8594;
      </button>
    </form>
  </div>

<div style="font-size:.78rem;color:#888;margin-top:4px">
  &#8592; <a href="/historia" style="color:#185FA5">Wróć do historii meczów</a>
</div>
"""

    _mecze_js_parts = []
    for _m in mecze:
        _dt   = _m["data_meczu"].strftime("%d.%m.%Y") if _m["data_meczu"] else "—"
        _gn   = (_m.get("nazwa_a","") or "GTK").replace('"','\\"')
        _on   = (_m.get("przeciwnik","") or "?").replace('"','\\"')
        _win  = "true" if ((_m.get("wynik_gtk") or 0) > (_m.get("wynik_opp") or 0)) else "false"
        _wnik = f'{_m.get("wynik_gtk",0)}:{_m.get("wynik_opp",0)}'
        _mecze_js_parts.append(f'{{"id":{_m["id"]},"data":"{_dt}","nazwa":"{_gn}","rywal":"{_on}","w":{_win},"wynik":"{_wnik}"}}')
    _mecze_js  = "[" + ",".join(_mecze_js_parts) + "]"
    _presel_js = "[" + ",".join(f'"{p}"' for p in presel) + "]"

    scripts = f"""<script>
var _allMecze = {_mecze_js};
var _zbiorczePresel = {_presel_js};
var _zbBuilt = false;

function zbAllToggle(cb) {{
    document.querySelectorAll('#zbiorcze-list input').forEach(function(c){{ c.checked=cb.checked; }});
}}

function buildZbiorczeList() {{
    if(_zbBuilt) return; _zbBuilt=true;
    var el=document.getElementById('zbiorcze-list'); if(!el) return;
    _allMecze.forEach(function(m) {{
        var lbl=document.createElement('label');
        lbl.style.cssText='display:flex;align-items:center;gap:8px;padding:4px 6px;border-radius:4px;cursor:pointer;font-size:.82rem';
        lbl.innerHTML='<input type="checkbox" name="ids" value="'+m.id+'" '+(_zbiorczePresel.indexOf(String(m.id))>=0?'checked':'')+' style="width:14px;height:14px"> '
            +'<span style="color:#888;min-width:72px">'+m.data+'</span>'
            +'<strong>'+m.nazwa+'</strong> <span style="color:#888;font-size:.78rem">vs</span> <strong>'+m.rywal+'</strong>'
            +' <span style="border-radius:10px;padding:1px 7px;font-size:10px;margin-left:4px;background:'+(m.w?'#d4edda':'#f8d7da')+';color:'+(m.w?'#155724':'#721c24')+'">'+( m.w?'W':'P')+'</span>'
            +' <span style="color:#888;font-size:10px">'+m.wynik+'</span>';
        el.appendChild(lbl);
    }});
}}

function switchTab(t) {{
    ['mecze','druzyny','zbiorcze'].forEach(function(p){{
        var pn=document.getElementById('pane-'+p), tb=document.getElementById('tab-'+p);
        if(pn) pn.style.display=t===p?'':'none';
        if(tb){{ tb.style.background=t===p?'#1a2b4a':'#f8f9fa'; tb.style.color=t===p?'#fff':'#666'; }}
    }});
    if(t==='zbiorcze') buildZbiorczeList();
}}
</script>"""

    # Auto-switch do zakładki zbiorczej gdy przekazano wiele meczów
    if len(presel) > 1:
        scripts = scripts.replace('</script>', """
<script>document.addEventListener('DOMContentLoaded',function(){switchTab('zbiorcze');});</script>
""")

    return html_response(base(content, scripts, active="history"))


@app.route("/porownaj/pdf")
@login_required
def porownaj_pdf():
    """Generuje raport porównawczy HTML (do wydruku/PDF)."""
    import json as _j

    tryb   = request.args.get("tryb","mecze")
    db     = get_db(); cur = db.cursor()

    def _agg_match(match_id):
        """Zwraca (meta, suma, kpi, player_rows) dla jednego meczu."""
        cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
        m = cur.fetchone()
        if not m:
            return None
        cur.execute("SELECT * FROM match_stats WHERE match_id=%s", (match_id,))
        all_stats = list(cur.fetchall())
        cur.execute("SELECT * FROM player_stats WHERE match_id=%s AND druzyna='gtk'", (match_id,))
        all_players = list(cur.fetchall())

        def build_suma(druzyna):
            s = {"pts":0,"poss":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,
                 "ftm":0,"fta":0,"br":0,"fd":0,"ast":0,"oreb":0,"dreb":0,"stl":0}
            for row in all_stats:
                if row["druzyna"] == druzyna:
                    for k in s: s[k] += row.get(k,0) or 0
            return s

        suma = build_suma("gtk")
        kpi  = calc_kpi(suma)

        # Wyniki per kwarta
        q_pts_g = [next((r["pts"] for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==q), 0) for q in [1,2,3,4]]
        q_pts_o = [next((r["pts"] for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==q), 0) for q in [1,2,3,4]]

        # Zawodnicy (top 8 wg PTS)
        nr_map = build_nr_name_map(cur, match_id)
        players_sorted = sorted(all_players, key=lambda x: x.get("pts",0) or 0, reverse=True)[:8]

        nazwa = (m.get("team_name_a","") or m.get("nazwa_gtk","") or "GTK").strip()
        dt    = m["data_meczu"].strftime("%d.%m.%Y") if m["data_meczu"] else ""

        meta = {
            "id": match_id,
            "nazwa": nazwa,
            "przeciwnik": m["przeciwnik"],
            "wynik_gtk": m["wynik_gtk"],
            "wynik_opp": m["wynik_opp"],
            "sezon": m["sezon"] or "",
            "dt": dt,
            "n": 1,
        }
        return meta, suma, kpi, players_sorted, nr_map, q_pts_g, q_pts_o

    def _agg_team(team_id):
        """Zwraca (meta, suma_avg, kpi_avg) jako średnie ze wszystkich meczów drużyny."""
        cur.execute("""
            SELECT m.id, m.data_meczu, m.sezon, m.wynik_gtk, m.wynik_opp,
                   COALESCE(m.team_name_a, m.nazwa_gtk,'GTK') as nazwa
            FROM matches m WHERE m.team_id=%s ORDER BY m.data_meczu
        """, (team_id,))
        mecze = cur.fetchall()
        if not mecze:
            return None

        cur.execute("SELECT name FROM teams WHERE id=%s", (team_id,))
        tr = cur.fetchone()
        team_name = tr["name"] if tr else "—"

        fields = ["pts","poss","p2m","p2a","p3m","p3a","ftm","fta","br","fd","ast","oreb","dreb","stl"]
        total  = {f:0 for f in fields}
        n      = len(mecze)
        wins   = sum(1 for m in mecze if m["wynik_gtk"] > m["wynik_opp"])

        for m in mecze:
            cur.execute("SELECT * FROM match_stats WHERE match_id=%s AND druzyna='gtk'", (m["id"],))
            for row in cur.fetchall():
                for f in fields: total[f] += row.get(f,0) or 0

        avg = {f: round(total[f]/n, 1) for f in fields}
        kpi = calc_kpi(avg)

        sezon = mecze[0]["sezon"] or ""
        meta = {
            "nazwa": team_name,
            "przeciwnik": f"{n} meczów",
            "wynik_gtk": round(sum(m["wynik_gtk"] for m in mecze)/n, 1),
            "wynik_opp": round(sum(m["wynik_opp"] for m in mecze)/n, 1),
            "sezon": sezon,
            "dt": f"Sezon {sezon}",
            "n": n,
            "wins": wins,
        }
        return meta, avg, kpi, [], {}, [], []

    # ── Pobierz dane A i B ────────────────────────────────────────────────────
    if tryb == "mecze":
        aid = request.args.get("mecz_a","")
        bid = request.args.get("mecz_b","")
        if not aid or not bid:
            return "Brak wybranych meczów.", 400
        res_a = _agg_match(int(aid))
        res_b = _agg_match(int(bid))
    else:
        aid = request.args.get("team_a","")
        bid = request.args.get("team_b","")
        if not aid or not bid:
            return "Brak wybranych drużyn.", 400
        res_a = _agg_team(int(aid))
        res_b = _agg_team(int(bid))

    # Piątki dla trybu mecze
    lineup_a_html = lineup_b_html = ""
    if tryb == "mecze":
        try:
            _aid = int(request.args.get("mecz_a","0"))
            _bid = int(request.args.get("mecz_b","0"))
            if _aid and res_a:
                _, _, _, _, nr_map_tmp_a, _, _ = res_a
                lineup_a_html = build_lineup_section_html(cur, _aid, nr_map_tmp_a)
            if _bid and res_b:
                _, _, _, _, nr_map_tmp_b, _, _ = res_b
                lineup_b_html = build_lineup_section_html(cur, _bid, nr_map_tmp_b)
        except Exception:
            pass

    cur.close()
    if not res_a or not res_b:
        return "Nie znaleziono danych.", 404

    meta_a, suma_a, kpi_a, players_a, nr_map_a, q_g_a, q_o_a = res_a
    meta_b, suma_b, kpi_b, players_b, nr_map_b, q_g_b, q_o_b = res_b

    gtk_name = get_setting("gtk_name") or "GTK"

    # ── Funkcje pomocnicze ────────────────────────────────────────────────────
    def kpi_box(val, lbl, bg="#f5f8ff"):
        return (f'<div style="background:{bg};border-radius:4px;padding:5px 4px;text-align:center">'
                f'<div style="font-size:12px;font-weight:700;color:#1a2b4a">{val}</div>'
                f'<div style="font-size:7.5px;color:#888;text-transform:uppercase">{lbl}</div></div>')

    def cmp_bar(lbl, ga, gb, low=False):
        try:
            gn = float(str(ga).replace("%","").replace("-","0") or 0)
            bn = float(str(gb).replace("%","").replace("-","0") or 0)
        except: gn = bn = 0
        tot = gn + bn or 1
        gp  = round(gn/tot*100)
        bp  = 100 - gp
        return (f'<div style="display:flex;align-items:center;gap:5px;margin-bottom:5px">'
                f'<div style="width:36px;font-size:8px;font-weight:700;color:#0C447C;text-align:right">{ga}</div>'
                f'<div style="width:60px;font-size:8.5px;color:#555;text-align:center">{lbl}</div>'
                f'<div style="flex:1;height:10px;background:#e8e8e8;border-radius:3px;overflow:hidden;display:flex">'
                f'<div style="width:{gp}%;background:#185FA5;height:100%"></div>'
                f'<div style="width:{bp}%;background:#BA7517;height:100%"></div>'
                f'</div>'
                f'<div style="width:36px;font-size:8px;font-weight:700;color:#633806">{gb}</div></div>')

    def player_table(players, nr_map, hdr_color):
        if not players:
            return '<div style="font-size:9px;color:#aaa;padding:6px">Brak danych zawodników</div>'
        rows = ""
        for i, p in enumerate(players):
            nr   = p.get("nr",0)
            name = nr_map.get(str(nr), f"#{nr}")
            pts  = p.get("pts",0) or 0
            fga  = (p.get("p2a",0) or 0) + (p.get("p3a",0) or 0)
            efg  = f"{(( p.get('p2m',0) or 0)+1.5*(p.get('p3m',0) or 0))/fga:.0%}" if fga else "—"
            bg   = "#f5f8ff" if i%2==0 else "#fff"
            rows += (f'<tr style="background:{bg}">'
                     f'<td style="padding:3px 6px;font-weight:700;font-size:9px">{name}</td>'
                     f'<td style="padding:3px 6px;text-align:center;font-weight:700;font-size:9px;color:#1a2b4a">{pts}</td>'
                     f'<td style="padding:3px 6px;text-align:center;font-size:9px">{p.get("p2m",0)}/{p.get("p2a",0)}</td>'
                     f'<td style="padding:3px 6px;text-align:center;font-size:9px">{p.get("p3m",0)}/{p.get("p3a",0)}</td>'
                     f'<td style="padding:3px 6px;text-align:center;font-size:9px">{p.get("ftm",0)}/{p.get("fta",0)}</td>'
                     f'<td style="padding:3px 6px;text-align:center;font-size:9px;font-weight:600">{efg}</td>'
                     f'<td style="padding:3px 6px;text-align:center;font-size:9px">{p.get("ast",0) or 0}</td>'
                     f'<td style="padding:3px 6px;text-align:center;font-size:9px">{p.get("br",0) or 0}</td>'
                     f'</tr>')
        th = f'background:{hdr_color};color:#fff;padding:4px 6px;font-size:8px;text-align:center'
        thl = f'background:{hdr_color};color:#fff;padding:4px 6px;font-size:8px'
        return (f'<table style="width:100%;border-collapse:collapse;font-size:9px">'
                f'<thead><tr>'
                f'<th style="{thl}">Zawodnik</th>'
                f'<th style="{th}">PTS</th><th style="{th}">2PM/A</th>'
                f'<th style="{th}">3PM/A</th><th style="{th}">FTM/A</th>'
                f'<th style="{th}">eFG%</th><th style="{th}">AST</th>'
                f'<th style="{th}">TO</th>'
                f'</tr></thead><tbody>{rows}</tbody></table>')

    # ── KPI gridy ────────────────────────────────────────────────────────────
    def kpi_grid(suma, kpi, bg):
        reb = (suma.get("oreb",0) or 0)+(suma.get("dreb",0) or 0)
        return (f'<div style="display:grid;grid-template-columns:repeat(6,1fr);gap:4px;margin-bottom:8px">'
                + kpi_box(suma.get("pts",0), "Punkty", bg)
                + kpi_box(kpi["efg"], "eFG%", bg)
                + kpi_box(kpi["ortg"], "ORtg", bg)
                + kpi_box(kpi["ppp"], "PPP", bg)
                + kpi_box(suma.get("br",0), "Straty", bg)
                + kpi_box(kpi["p2_pct"], "2PT%", bg)
                + kpi_box(kpi["p3_pct"], "3PT%", bg)
                + kpi_box(kpi["ft_pct"], "FT%", bg)
                + kpi_box(suma.get("poss",0), "Posiadania", bg)
                + kpi_box(suma.get("ast",0), "Asysty", bg)
                + kpi_box(reb, "Zbiórki", bg)
                + kpi_box(suma.get("stl",0), "Przechwyty", bg)
                + '</div>')

    tryb_lbl = "Porównanie meczów" if tryb == "mecze" else "Porównanie drużyn"
    n_lbl_a  = f" (avg {meta_a['n']} meczów)" if meta_a['n'] > 1 else ""
    n_lbl_b  = f" (avg {meta_b['n']} meczów)" if meta_b['n'] > 1 else ""

    # ── Wyniki kwart ─────────────────────────────────────────────────────────
    q_rows_a = ""
    q_rows_b = ""
    if tryb == "mecze":
        for qi, (pg, po) in enumerate(zip(q_g_a, q_o_a), 1):
            bg = "#fff" if qi%2==0 else "#f5f8ff"
            q_rows_a += f'<tr style="background:{bg}"><td style="padding:3px 6px;font-weight:700">{qi}Q</td><td style="text-align:center;padding:3px 6px;font-weight:700">{pg}</td><td style="text-align:center;padding:3px 6px">{po}</td></tr>'
        for qi, (pg, po) in enumerate(zip(q_g_b, q_o_b), 1):
            bg = "#fff" if qi%2==0 else "#fdf6ed"
            q_rows_b += f'<tr style="background:{bg}"><td style="padding:3px 6px;font-weight:700">{qi}Q</td><td style="text-align:center;padding:3px 6px;font-weight:700">{pg}</td><td style="text-align:center;padding:3px 6px">{po}</td></tr>'

    q_section_a = ""
    q_section_b = ""
    if tryb == "mecze":
        q_section_a = f"""
<h3 style="font-size:9px;text-transform:uppercase;letter-spacing:.5px;color:#185FA5;margin:8px 0 4px">Per kwarta</h3>
<table style="width:100%;border-collapse:collapse;font-size:9px;margin-bottom:8px">
<thead><tr>
  <th style="background:#1a2b4a;color:#fff;padding:4px 6px">Q</th>
  <th style="background:#1a6b3c;color:#fff;padding:4px 6px;text-align:center">{meta_a["nazwa"]}</th>
  <th style="background:#8b1a1a;color:#fff;padding:4px 6px;text-align:center">{meta_a["przeciwnik"]}</th>
</tr></thead>
<tbody>{q_rows_a}</tbody>
</table>"""
        q_section_b = f"""
<h3 style="font-size:9px;text-transform:uppercase;letter-spacing:.5px;color:#BA7517;margin:8px 0 4px">Per kwarta</h3>
<table style="width:100%;border-collapse:collapse;font-size:9px;margin-bottom:8px">
<thead><tr>
  <th style="background:#1a2b4a;color:#fff;padding:4px 6px">Q</th>
  <th style="background:#1a6b3c;color:#fff;padding:4px 6px;text-align:center">{meta_b["nazwa"]}</th>
  <th style="background:#8b1a1a;color:#fff;padding:4px 6px;text-align:center">{meta_b["przeciwnik"]}</th>
</tr></thead>
<tbody>{q_rows_b}</tbody>
</table>"""

    # ── HTML raportu ──────────────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
  @page{{size:A4 landscape;margin:1.2cm}}
  body{{font-family:Arial,sans-serif;font-size:10px;color:#222;margin:0;padding:0}}
  h2{{font-size:14px;margin:0 0 2px;color:#fff}} h3{{font-size:11px;color:#1a2b4a;margin:10px 0 4px;text-transform:uppercase;letter-spacing:.5px}}
  .hero{{background:#1a2b4a;color:#fff;padding:10px 16px;border-radius:6px;margin-bottom:12px;display:flex;justify-content:space-between;align-items:center}}
  .two-col{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:10px}}
  .section{{margin-bottom:10px;page-break-inside:avoid}}
  .sep{{height:1px;background:#eee;margin:10px 0}}
  @media print{{.no-print{{display:none}}}}
</style>
</head><body>

<div class="hero">
  <div>
    <h2>{meta_a["nazwa"]}{n_lbl_a} vs {meta_b["nazwa"]}{n_lbl_b}</h2>
    <div style="opacity:.7;font-size:9px">{tryb_lbl} · {meta_a["dt"]} / {meta_b["dt"]} · Basket Kołcz Analytics</div>
  </div>
  <div style="text-align:center">
    <div style="font-size:11px;opacity:.7;margin-bottom:2px">Śr. wynik</div>
    <div style="font-size:22px;font-weight:700;letter-spacing:2px;color:#EF9F27">
      {meta_a["wynik_gtk"]} : {meta_b["wynik_gtk"]}
    </div>
  </div>
  <div style="font-size:10px;opacity:.8">{meta_a["sezon"]} / {meta_b["sezon"]}</div>
</div>

<!-- KPI -->
<div class="two-col">
  <div class="section">
    <h3 style="color:#185FA5">{meta_a["nazwa"]}{n_lbl_a} — metryki</h3>
    {kpi_grid(suma_a, kpi_a, "#f0f4ff")}
    {q_section_a}
  </div>
  <div class="section">
    <h3 style="color:#BA7517">{meta_b["nazwa"]}{n_lbl_b} — metryki</h3>
    {kpi_grid(suma_b, kpi_b, "#fdf6ed")}
    {q_section_b}
  </div>
</div>

<div class="sep"></div>

<!-- Porównanie paskami -->
<div style="padding:8px 10px;background:#f9f9f9;border-radius:4px;margin-bottom:10px">
  <div style="text-align:center;font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:#1a2b4a;margin-bottom:8px">
    Porównanie bezpośrednie
  </div>
  <div style="display:flex;justify-content:space-between;margin-bottom:5px;padding:0 36px;font-size:8px;font-weight:700">
    <span style="color:#0C447C">{meta_a["nazwa"]}</span>
    <span style="color:#633806">{meta_b["nazwa"]}</span>
  </div>
  {cmp_bar("Punkty",       suma_a.get("pts",0),        suma_b.get("pts",0))}
  {cmp_bar("Zbiórki",      (suma_a.get("oreb",0) or 0)+(suma_a.get("dreb",0) or 0),
                            (suma_b.get("oreb",0) or 0)+(suma_b.get("dreb",0) or 0))}
  {cmp_bar("Asysty",       suma_a.get("ast",0),         suma_b.get("ast",0))}
  {cmp_bar("Przechwyty",   suma_a.get("stl",0),         suma_b.get("stl",0))}
  {cmp_bar("Straty (TO)",  suma_a.get("br",0),          suma_b.get("br",0),  low=True)}
  {cmp_bar("eFG%",         kpi_a["efg"],                kpi_b["efg"])}
  {cmp_bar("ORtg",         kpi_a["ortg"],               kpi_b["ortg"])}
  {cmp_bar("PPP",          kpi_a["ppp"],                kpi_b["ppp"])}
  {cmp_bar("TS%",          kpi_a["ts"],                 kpi_b["ts"])}
</div>

<div class="sep"></div>

<!-- Zawodnicy -->
<div class="two-col">
  <div class="section">
    <h3 style="color:#185FA5">{meta_a["nazwa"]} — top zawodnicy</h3>
    {player_table(players_a, nr_map_a, "#1a6b3c")}
  </div>
  <div class="section">
    <h3 style="color:#BA7517">{meta_b["nazwa"]} — top zawodnicy</h3>
    {player_table(players_b, nr_map_b, "#BA7517")}
  </div>
</div>

<div class="no-print" style="margin:14px 0;text-align:center">
  <button onclick="window.print()" style="background:#1a2b4a;color:#fff;border:none;border-radius:6px;padding:8px 24px;font-size:13px;cursor:pointer;font-weight:600">
    &#128438; Drukuj / Zapisz PDF
  </button>
  &nbsp;
  <a href="/porownaj" style="font-size:12px;color:#185FA5">&#8592; Nowe porównanie</a>
</div>

<div style="margin-top:10px;text-align:center;font-size:8px;color:#aaa;border-top:1px solid #eee;padding-top:5px">
  Basket Kołcz Analytics · {tryb_lbl} · {meta_a["nazwa"]} vs {meta_b["nazwa"]}
</div>

</body></html>"""

    # Wstaw piątki przed footer
    footer_marker = '  Basket Kołcz Analytics · ' + tryb_lbl + ' · ' + meta_a["nazwa"] + ' vs ' + meta_b["nazwa"]
    lineups_insert = ""
    if lineup_a_html:
        lineups_insert += f'<div style="margin-bottom:4px"><strong style="font-size:8px">{meta_a["nazwa"]}</strong></div>' + lineup_a_html
    if lineup_b_html:
        lineups_insert += f'<div style="margin-top:6px;margin-bottom:4px"><strong style="font-size:8px">{meta_b["nazwa"]}</strong></div>' + lineup_b_html
    if lineups_insert:
        html = html.replace(footer_marker, lineups_insert + '\n<div style="text-align:center;font-size:7.5px;color:#aaa;border-top:1px solid #eee;padding-top:5px">\n  ' + footer_marker[2:])

    html_print = html.replace(
        "</body></html>",
        "<script>window.onload=function(){window.print();}</script></body></html>"
    )
    return Response(html_print, mimetype="text/html")


@app.route("/dev/download-app")
@login_required
def dev_download_app():
    import os
    from flask import send_file as _sf
    path = os.path.abspath(__file__)
    return _sf(path, as_attachment=True, download_name="app.py", mimetype="text/x-python")


# ══════════════════════════════════════════════════════════════════════════════
# PORTAL ZAWODNIKA
# ══════════════════════════════════════════════════════════════════════════════

_PORTAL_LOGIN = os.environ.get("PORTAL_LOGIN", "")
_PORTAL_PASS  = os.environ.get("PORTAL_PASS", "")

_PORTAL_CSS = """
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Inter',sans-serif;background:#f0f2f7;color:#111827;min-height:100vh}
.topbar{background:#1a2b4a;border-bottom:1px solid rgba(0,0,0,.12);padding:0 24px;height:58px;
  display:flex;align-items:center;position:sticky;top:0;z-index:100;
  box-shadow:0 2px 8px rgba(0,0,0,.15)}
.tb-left{display:flex;align-items:center;gap:10px;flex-shrink:0}
.tb-logo{display:flex;align-items:center;gap:10px}
.tb-logo-icon{width:36px;height:36px;border-radius:8px;overflow:hidden;
  display:flex;align-items:center;justify-content:center;font-size:15px;flex-shrink:0}
.tb-logo-icon img{width:100%;height:100%;object-fit:cover;border-radius:8px}
.tb-logo-text{font-size:.95rem;font-weight:800;color:#fff;letter-spacing:-.02em}
.tb-logo-text span{color:#c9a340}
.tb-nav{display:flex;gap:4px;flex:1;justify-content:center}
.tb-btn{padding:6px 15px;border-radius:8px;font-size:.76rem;font-weight:600;cursor:pointer;
  border:none;background:transparent;color:rgba(255,255,255,.5);font-family:inherit;
  letter-spacing:.02em;transition:all .15s}
.tb-btn:hover{color:#fff;background:rgba(255,255,255,.1)}
.tb-btn.active{color:#c9a340;background:rgba(201,163,64,.15)}
.tb-right{display:flex;align-items:center;gap:8px}
.tb-name{font-size:.78rem;font-weight:600;color:rgba(255,255,255,.7)}
.tb-logout{font-size:.7rem;color:rgba(255,255,255,.45);cursor:pointer;background:none;
  border:1px solid rgba(255,255,255,.2);border-radius:6px;padding:4px 10px;font-family:inherit;
  text-decoration:none;transition:all .15s}
.tb-logout:hover{color:#fff;border-color:rgba(255,255,255,.5)}
.tb-selector{position:relative}
.tb-sel-btn{display:flex;align-items:center;gap:6px;padding:5px 10px;border-radius:8px;
  border:1px solid rgba(255,255,255,.18);background:rgba(255,255,255,.06);cursor:pointer;
  font-family:inherit;transition:all .18s;min-width:110px}
.tb-sel-btn:hover{background:rgba(255,255,255,.12);border-color:rgba(255,255,255,.32)}
.tb-sel-btn.open{background:rgba(201,163,64,.15);border-color:rgba(201,163,64,.5)}
.tb-sel-label{font-size:.56rem;font-weight:700;text-transform:uppercase;letter-spacing:.07em;
  color:rgba(255,255,255,.38);line-height:1;display:block}
.tb-sel-value{font-size:.73rem;font-weight:700;color:#fff;line-height:1.2;display:block;
  margin-top:1px;white-space:nowrap}
.tb-sel-btn.open .tb-sel-value{color:#c9a340}
.tb-chevron{margin-left:auto;color:rgba(255,255,255,.4);font-size:.58rem;transition:transform .18s;flex-shrink:0}
.tb-sel-btn.open .tb-chevron{transform:rotate(180deg);color:#c9a340}
.tb-dropdown{position:absolute;top:calc(100% + 6px);left:0;min-width:150px;
  background:#1e3055;border:1px solid rgba(255,255,255,.14);border-radius:10px;
  box-shadow:0 8px 28px rgba(0,0,0,.35);z-index:200;overflow:hidden;
  opacity:0;transform:translateY(-6px);pointer-events:none;
  transition:opacity .18s,transform .18s}
.tb-selector.open .tb-dropdown{opacity:1;transform:translateY(0);pointer-events:auto}
.tb-dd-item{display:block;padding:9px 14px;font-size:.78rem;font-weight:500;
  color:rgba(255,255,255,.75);cursor:pointer;transition:all .12s;
  border-bottom:1px solid rgba(255,255,255,.06)}
.tb-dd-item:last-child{border-bottom:none}
.tb-dd-item:hover{background:rgba(255,255,255,.08);color:#fff}
.tb-dd-item.active{color:#c9a340;font-weight:700;background:rgba(201,163,64,.1)}
.tb-dd-item.active::before{content:"✓ ";font-size:.68rem}
.tb-sep{width:1px;height:22px;background:rgba(255,255,255,.12);margin:0 2px}
@media(max-width:480px){
  .topbar{padding:0 12px;height:50px}
  .tb-logo-text{display:none}
  .tb-logo-icon{width:30px;height:30px}
  .tb-btn{padding:5px 10px;font-size:.7rem}
  .tb-logout{font-size:.65rem;padding:3px 8px}
  .tb-name{display:none}
  .tb-selector .tb-sel-btn{min-width:80px;padding:4px 8px}
  .tb-sel-value{font-size:.68rem}
}
.pane{display:none;padding:24px;max-width:1280px;margin:0 auto;width:100%}
.pane.active{display:block}
.eyebrow{font-size:.63rem;font-weight:700;text-transform:uppercase;letter-spacing:.1em;
  color:#c9a340;margin-bottom:6px}
.sec-title{font-size:1.4rem;font-weight:800;color:#1a2b4a;margin-bottom:18px;letter-spacing:-.02em}
.hero-strip{background:linear-gradient(135deg,#1a2b4a,#253d66);border-radius:14px;
  padding:22px 26px;display:flex;align-items:center;justify-content:space-between;
  margin-bottom:20px;gap:14px;flex-wrap:wrap;box-shadow:0 4px 18px rgba(26,43,74,.18)}
.hero-name{font-size:1.6rem;font-weight:900;color:#fff;letter-spacing:-.03em}
.hero-sub{font-size:.72rem;color:rgba(255,255,255,.4);margin-top:3px}
.hero-rec{display:flex;gap:18px}
.hr-stat-val{font-size:1.8rem;font-weight:900;line-height:1;text-align:center}
.hr-stat-lbl{font-size:.58rem;font-weight:600;text-transform:uppercase;letter-spacing:.08em;
  color:rgba(255,255,255,.4);margin-top:2px;text-align:center}
.hero-rtg{display:flex;gap:10px;background:rgba(0,0,0,.2);border-radius:10px;padding:14px 18px}
.rtg-item{text-align:center;min-width:54px}
.rtg-val{font-size:1.2rem;font-weight:800;color:#c9a340}
.rtg-lbl{font-size:.58rem;color:rgba(255,255,255,.38);text-transform:uppercase;letter-spacing:.07em;margin-top:2px}
.rtg-div{width:1px;background:rgba(255,255,255,.1)}
.kpi-row{display:grid;gap:10px;margin-bottom:10px}
.kpi-row.c7{grid-template-columns:repeat(7,1fr)}
.kpi-row.c5{grid-template-columns:repeat(5,1fr)}
.kc{perspective:600px;text-align:center;cursor:default}
.kc-inner{position:relative;height:72px;transform-style:preserve-3d;
  transition:transform .45s cubic-bezier(.4,0,.2,1)}
.kc:hover .kc-inner{transform:rotateY(180deg)}
.kc-front,.kc-back{position:absolute;inset:0;border-radius:10px;
  display:flex;flex-direction:column;align-items:center;justify-content:center;
  padding:10px 8px;backface-visibility:hidden;-webkit-backface-visibility:hidden}
.kc-front{background:#fff;border:1px solid #e8ecf3;box-shadow:0 1px 3px rgba(0,0,0,.05)}
.kc-back{background:linear-gradient(135deg,#1a2b4a,#253d66);
  box-shadow:0 4px 14px rgba(26,43,74,.2);transform:rotateY(180deg)}
.kc-val{font-size:1.35rem;font-weight:800;color:#1a2b4a}
.kc-lbl{font-size:.6rem;text-transform:uppercase;letter-spacing:.07em;color:#9ca3af;margin-top:3px}
.kc-back .kc-val{color:#fff}
.kc-back .kc-lbl{color:#c9a340}
.kc-static{background:#fff;border:1px solid #e8ecf3;border-radius:10px;padding:14px 10px;
  box-shadow:0 1px 3px rgba(0,0,0,.05);perspective:none}
.kc-static:hover .kc-inner{transform:none}
.matches-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
.mc{background:#fff;border:1px solid #e8ecf3;border-radius:12px;padding:14px;
  box-shadow:0 1px 3px rgba(0,0,0,.05);transition:all .2s}
.mc:hover{border-color:#c9a340;box-shadow:0 4px 14px rgba(201,163,64,.12)}
.mc-date{font-size:.62rem;text-transform:uppercase;letter-spacing:.07em;color:#9ca3af;margin-bottom:8px}
.mc-row{display:flex;justify-content:space-between;align-items:center;margin-bottom:6px}
.mc-team{font-size:.82rem;font-weight:600;color:#1a2b4a}
.mc-score{font-size:1.2rem;font-weight:900}
.mc-score.w{color:#1a6b3c}.mc-score.l{color:#8b1a1a}
.badge-w{font-size:.6rem;font-weight:700;padding:2px 7px;border-radius:4px;background:#e8f5e9;color:#1a6b3c}
.badge-l{font-size:.6rem;font-weight:700;padding:2px 7px;border-radius:4px;background:#ffebee;color:#8b1a1a}
.duel-grid{display:grid;grid-template-columns:5fr 3fr 4fr;gap:14px}
.sp{background:#fff;border:1px solid #e8ecf3;border-radius:12px;padding:18px;
  box-shadow:0 1px 3px rgba(0,0,0,.05)}
.sp-title{font-size:.62rem;font-weight:700;text-transform:uppercase;letter-spacing:.1em;
  color:#9ca3af;margin-bottom:12px}
.dh{display:grid;grid-template-columns:58px 1fr 58px;font-size:.63rem;font-weight:700;
  text-transform:uppercase;letter-spacing:.07em;margin-bottom:6px}
.dh-g{color:#1a6b3c;text-align:right}.dh-o{color:#8b1a1a}
.dr{display:grid;grid-template-columns:58px 1fr 58px;align-items:center;gap:7px;
  padding:5px 0;border-bottom:1px solid #f3f4f6}
.dr:last-child{border-bottom:none}
.dr-g{font-size:.9rem;font-weight:700;text-align:right;color:#1a6b3c}
.dr-g.lose{color:#9ca3af;font-weight:400}
.dr-o{font-size:.9rem;color:#9ca3af}
.dr-o.win{font-weight:700;color:#8b1a1a}
.dr-lbl{font-size:.58rem;font-weight:600;color:#9ca3af;text-align:center;margin-bottom:2px}
.dr-bar{display:flex;align-items:center;height:8px}
.db-g{height:8px;border-radius:3px 0 0 3px;background:#1D9E75}
.db-d{width:1px;height:12px;background:#d1d5db;flex-shrink:0}
.db-o{height:8px;border-radius:0 3px 3px 0;background:#E24B4A}
.net-card{background:linear-gradient(135deg,#e8f5e9,#f0faf2);border:1px solid #c8e6c9;
  border-radius:12px;padding:18px;text-align:center;margin-bottom:14px}
.net-val{font-size:2.6rem;font-weight:900;color:#1a6b3c;line-height:1}
.net-sub{font-size:.62rem;color:#9ca3af;text-transform:uppercase;letter-spacing:.08em;margin-top:3px}
.net-row{display:grid;grid-template-columns:1fr 1fr;gap:7px;margin-top:12px}
.net-box{background:rgba(255,255,255,.75);border-radius:7px;padding:8px;text-align:center}
.net-box-val{font-size:1rem;font-weight:800}
.net-box-lbl{font-size:.58rem;color:#9ca3af;text-transform:uppercase;letter-spacing:.07em;margin-top:1px}
.eff-row{display:grid;grid-template-columns:40px 1fr 40px;align-items:center;gap:6px;
  padding:5px 0;border-bottom:1px solid #f3f4f6}
.eff-row:last-child{border-bottom:none}
.eff-g{font-size:.85rem;font-weight:700;color:#1a6b3c;text-align:right}
.eff-g.lose{color:#9ca3af;font-weight:400}
.eff-o{font-size:.85rem;color:#9ca3af}
.eff-o.win{color:#8b1a1a;font-weight:700}
.eff-lbl{font-size:.58rem;font-weight:600;color:#9ca3af;text-align:center;margin-bottom:2px}
.eff-bar{display:flex;align-items:center;height:7px}
.eb-g{height:7px;border-radius:3px 0 0 3px;background:#1D9E75}
.eb-d{width:1px;height:10px;background:#d1d5db;flex-shrink:0}
.eb-o{height:7px;border-radius:0 3px 3px 0;background:#E24B4A}
.ind-wrap{background:#fff;border:1px solid #e8ecf3;border-radius:12px;overflow:hidden;
  box-shadow:0 1px 3px rgba(0,0,0,.05)}
.ind-tbl{width:100%;border-collapse:collapse;font-size:.78rem;min-width:760px}
.ind-wrap-scroll{overflow-x:auto;-webkit-overflow-scrolling:touch;border-radius:12px}
@media(max-width:768px){
  .ind-tbl{font-size:.68rem;min-width:700px}
  .ind-tbl thead tr:first-child th{padding:7px 4px;font-size:.6rem}
  .ind-tbl thead tr.sh th{padding:3px 4px;font-size:.57rem}
  .ind-tbl tbody td{padding:6px 4px}
  .ind-tbl tbody td.pn{padding-left:8px}
}
.ind-tbl thead tr:first-child th{background:#1a2b4a;color:#fff;font-size:.68rem;font-weight:600;
  padding:9px 7px;text-align:center;white-space:nowrap;border-bottom:1px solid rgba(255,255,255,.08)}
.ind-tbl thead tr.sh th{background:#253d66;color:rgba(255,255,255,.6);font-size:.63rem;
  padding:4px 5px;text-align:center;border-bottom:1px solid rgba(255,255,255,.08)}
.ind-tbl thead .zb{background:#1e3459}
.ind-tbl tbody tr{border-bottom:1px solid #f3f4f6;transition:background .15s}
.ind-tbl tbody tr:hover{background:#f8fafc}
.ind-tbl tbody td{padding:8px 7px;text-align:center;color:#374151;white-space:nowrap}
.ind-tbl tbody td.pn{text-align:left;font-weight:600;color:#111827;padding-left:14px}
.ind-tbl tbody td.pts{font-weight:800;color:#1a2b4a}
.ind-tbl tbody td.min{color:#9ca3af;font-size:.72rem}
@media(max-width:768px){
  .duel-grid{grid-template-columns:1fr}
  .matches-grid{grid-template-columns:1fr}
  .kpi-row.c7,.kpi-row.c5{grid-template-columns:repeat(3,1fr)}
  .tb-nav{display:none}
  .pane{padding:16px 12px}
  .hero-strip{flex-direction:column;padding:16px;gap:10px}
  .hero-name{font-size:1.3rem}
  .hero-rec{flex-wrap:wrap;gap:8px}
  .hero-rtg{flex-wrap:wrap;padding:10px 12px;gap:8px}
  .rtg-item{min-width:44px}
  .sec-title{font-size:1.1rem;margin-bottom:12px}
  .kc-inner{height:62px}
  .kc-val{font-size:1.15rem}
  .kc-front,.kc-back{padding:8px 6px}
  .sp{padding:14px}
}
@media(max-width:480px){
  .kpi-row.c7,.kpi-row.c5{grid-template-columns:repeat(2,1fr)}
  .pane{padding:10px 8px}
  .hero-strip{padding:12px}
  .hero-name{font-size:1.05rem}
  .hero-sub{font-size:.65rem}
  .kc-inner{height:56px}
  .kc-val{font-size:1rem}
  .kc-lbl{font-size:.55rem}
  .net-val{font-size:2rem}
  .net-sub{font-size:.58rem}
  .sec-title{font-size:1rem}
  .eyebrow{font-size:.58rem}
}
/* ── Sidebar layout ───────────────────────────── */
.portal-layout{display:flex;min-height:100vh}
.sidebar{width:220px;background:#fff;border-right:1px solid #e5e7eb;
  display:flex;flex-direction:column;position:sticky;top:0;height:100vh;
  overflow-y:auto;flex-shrink:0;z-index:50}
.portal-main{flex:1;min-width:0;background:#f0f2f7}
.pane{padding:clamp(12px,3vw,24px);max-width:1280px;margin:0 auto;width:100%}
.sb-brand{background:#1a2b4a;padding:18px 14px;display:flex;align-items:center;gap:10px}
.sb-logo-wrap{width:40px;height:40px;border-radius:10px;overflow:hidden;flex-shrink:0;
  background:#253d66;display:flex;align-items:center;justify-content:center}
.sb-logo-wrap img{width:100%;height:100%;object-fit:cover;display:block}
.sb-logo-fallback{font-size:.9rem;font-weight:900;color:#c9a340;letter-spacing:-.03em}
.sb-brand-text{font-size:.9rem;font-weight:800;color:#fff;letter-spacing:-.02em}
.sb-brand-text span{color:#c9a340}
.sb-divider{height:1px;background:#f0f0f0;margin:4px 10px}
.sb-selectors{padding:6px 0}
.sb-sel-group{position:relative}
.sb-sel-hdr{width:100%;display:flex;align-items:center;gap:9px;padding:9px 14px;
  background:none;border:none;cursor:pointer;font-family:inherit;text-align:left;
  transition:background .15s;min-width:0}
.sb-sel-hdr:hover,.sb-sel-hdr.open{background:#f8fafc}
.sb-sel-icon-box{width:30px;height:30px;background:#f3f4f6;border-radius:7px;
  display:flex;align-items:center;justify-content:center;font-size:15px;flex-shrink:0}
.sb-sel-texts{flex:1;min-width:0;text-align:left}
.sb-sel-lbl{font-size:.54rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;
  color:#9ca3af;display:block}
.sb-sel-val{font-size:.78rem;font-weight:600;color:#1a2b4a;white-space:nowrap;
  overflow:hidden;text-overflow:ellipsis;display:block;margin-top:1px}
.sb-sel-arr{font-size:.55rem;color:#c0c4cc;flex-shrink:0;transition:transform .18s}
.sb-sel-hdr.open .sb-sel-arr{transform:rotate(180deg);color:#1a2b4a}
.sb-dropdown{position:absolute;left:0;right:0;background:#fff;border:1px solid #e5e7eb;
  border-radius:10px;box-shadow:0 6px 24px rgba(0,0,0,.1);z-index:300;
  opacity:0;transform:translateY(-4px);pointer-events:none;
  transition:opacity .16s,transform .16s;overflow:hidden;min-width:170px}
.sb-sel-group.open .sb-dropdown{opacity:1;transform:translateY(0);pointer-events:auto}
.sb-dd-item{display:block;padding:8px 14px;font-size:.78rem;color:#374151;cursor:pointer;
  border-bottom:1px solid #f3f4f6;transition:background .1s}
.sb-dd-item:last-child{border-bottom:none}
.sb-dd-item:hover{background:#f8fafc}
.sb-dd-item.sel{font-weight:700;color:#1a2b4a;background:#eff6ff}
.sb-dd-item.sel::before{content:"✓ ";color:#1a6b3c;font-size:.65rem}
.sb-nav-section{padding:6px 8px}
.sb-nav-btn{width:100%;display:flex;align-items:center;gap:9px;padding:9px 10px;
  background:none;border:none;border-radius:8px;cursor:pointer;font-family:inherit;
  font-size:.8rem;font-weight:500;color:#6b7280;transition:all .15s;text-align:left}
.sb-nav-btn:hover{background:#f3f4f6;color:#1a2b4a}
.sb-nav-btn.active{background:#1a2b4a;color:#fff;font-weight:600}
.sb-nav-icon{font-size:14px;width:20px;text-align:center;flex-shrink:0}
.sb-footer{padding:12px 14px;border-top:1px solid #f0f0f0;margin-top:auto}
.sb-logout-btn{display:block;text-align:center;padding:7px;border-radius:7px;
  font-size:.72rem;color:#9ca3af;border:1px solid #e5e7eb;text-decoration:none;
  transition:all .15s}
.sb-logout-btn:hover{color:#8b1a1a;border-color:#f8a0a0}
@media(max-width:900px){
  .portal-layout{flex-direction:column}
  .sidebar{width:100%;height:auto;position:relative;flex-direction:row;
    flex-wrap:nowrap;overflow-x:auto;align-items:center;padding:0 8px;gap:6px;
    background:#1a2b4a;border-right:none;border-bottom:1px solid rgba(255,255,255,.1)}
  .sb-brand{padding:8px 4px;background:transparent}
  .sb-brand-text{display:none}
  .sb-selectors{display:flex;flex-direction:row;padding:0;gap:4px}
  .sb-sel-group{flex-shrink:0}
  .sb-sel-hdr{padding:5px 9px;background:rgba(255,255,255,.08);border-radius:8px;
    border:1px solid rgba(255,255,255,.15)}
  .sb-sel-hdr:hover,.sb-sel-hdr.open{background:rgba(255,255,255,.15)}
  .sb-sel-lbl,.sb-sel-arr{color:rgba(255,255,255,.5)}
  .sb-sel-val{color:#fff}
  .sb-sel-icon-box{background:rgba(255,255,255,.1);width:22px;height:22px;
    border-radius:5px;font-size:11px}
  .sb-nav-section{display:flex;padding:0;gap:4px}
  .sb-nav-btn{padding:6px 12px;white-space:nowrap;color:rgba(255,255,255,.6);
    background:transparent;font-size:.75rem}
  .sb-nav-btn:hover{background:rgba(255,255,255,.1);color:#fff}
  .sb-nav-btn.active{background:rgba(201,163,64,.2);color:#c9a340}
  .sb-nav-icon{display:none}
  .sb-divider{display:none}
  .sb-footer{margin:0 0 0 auto;padding:8px;border-top:none}
  .sb-logout-btn{background:rgba(255,255,255,.08);border-color:rgba(255,255,255,.2);
    color:rgba(255,255,255,.6);white-space:nowrap;padding:6px 12px}
  .sb-dropdown{top:calc(100% + 4px);background:#1e3055;border-color:rgba(255,255,255,.14)}
  .sb-dd-item{color:rgba(255,255,255,.75);border-color:rgba(255,255,255,.06)}
  .sb-dd-item:hover{background:rgba(255,255,255,.08)}
  .sb-dd-item.sel{background:rgba(201,163,64,.1);color:#c9a340}
  .sb-dd-item.sel::before{color:#c9a340}
}
</style>"""

def _portal_duel_row(lbl, vg, vo, higher_is_better=True, neutral=False):
    try:
        fg = float(str(vg).replace('%','').replace('—','0').replace('+','') or 0)
        fo = float(str(vo).replace('%','').replace('—','0').replace('+','') or 0)
        total = abs(fg)+abs(fo)
        pg = max(5,min(95,int(fg/total*100))) if total else 50
        po = 100-pg
        gw = (higher_is_better and fg>fo) or (not higher_is_better and fg<fo)
        ow = (higher_is_better and fo>fg) or (not higher_is_better and fo<fg)
        cg = "#1a6b3c" if gw else ("#9ca3af" if ow else "#555")
        co = "#8b1a1a" if ow else "#9ca3af"
        wg = "700" if gw else "400"
        wo = "700" if ow else "400"
        if neutral: bg,bo = "#aaa","#aaa"
        elif higher_is_better: bg,bo = "#1D9E75","#E24B4A"
        else: bg = "#1D9E75" if gw else "#E24B4A"; bo = "#1D9E75" if ow else "#E24B4A"
    except Exception:
        pg=po=50; cg="#1a6b3c"; co="#9ca3af"; wg=wo="400"; bg=bo="#aaa"
    bar = (f'<div class="dr-bar"><div class="db-g" style="width:{pg}%;background:{bg}"></div>'
           f'<div class="db-d"></div><div class="db-o" style="width:{po}%;background:{bo}"></div></div>')
    gl = "" if (gw or not (gw or ow)) else " lose"
    ol = " win" if ow else ""
    return (f'<div class="dr"><div class="dr-g{gl}" style="color:{cg};font-weight:{wg}">{vg}</div>'
            f'<div><div class="dr-lbl">{lbl}</div>{bar}</div>'
            f'<div class="dr-o{ol}" style="color:{co};font-weight:{wo}">{vo}</div></div>')

def _portal_eff_row(lbl, vg, vo):
    try:
        fg = float(str(vg).replace('%','') or 0)
        fo = float(str(vo).replace('%','') or 0)
        total = abs(fg)+abs(fo)
        pg = max(5,min(95,int(fg/total*100))) if total else 50
        po = 100-pg
        gw = fg>fo
        cg = "#1a6b3c" if gw else "#9ca3af"
        co = "#8b1a1a" if fo>fg else "#9ca3af"
        wg = "700" if gw else "400"
        wo = "700" if fo>fg else "400"
    except Exception:
        pg=po=50; cg="#1a6b3c"; co="#9ca3af"; wg=wo="400"
    bar = (f'<div class="eff-bar"><div class="eb-g" style="width:{pg}%"></div>'
           f'<div class="eb-d"></div><div class="eb-o" style="width:{po}%"></div></div>')
    gl = "" if gw else " lose"
    ol = " win" if fo>fg else ""
    return (f'<div class="eff-row"><div class="eff-g{gl}" style="color:{cg};font-weight:{wg}">{vg}</div>'
            f'<div><div class="eff-lbl">{lbl}</div>{bar}</div>'
            f'<div class="eff-o{ol}" style="color:{co};font-weight:{wo}">{vo}</div></div>')


@app.route("/portal")
def portal():
    if not session.get("portal_logged_in"):
        err = request.args.get("err","")
        err_html = '<div style="color:#8b1a1a;font-size:.78rem;text-align:center;margin-bottom:12px;background:#ffebee;padding:8px 12px;border-radius:8px">Błędny login lub hasło</div>' if err else ""
        login_page = f"""<!DOCTYPE html>
<html lang="pl"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Portal Zawodnika — BasketKołcz</title>
{_PORTAL_CSS}
<style>
body{{background:linear-gradient(135deg,#dde6f5,#e8eef8,#d8e4f2);display:flex;
  align-items:center;justify-content:center;min-height:100vh}}
.login-box{{background:#fff;border:1px solid #e2e8f0;border-radius:18px;padding:44px 40px;
  width:370px;max-width:94vw;box-shadow:0 18px 50px rgba(0,0,0,.10)}}
.ll{{font-size:.7rem;font-weight:600;color:#6b7280;text-transform:uppercase;
  letter-spacing:.06em;margin-bottom:5px}}
.li{{width:100%;background:#f8fafc;border:1px solid #e2e8f0;border-radius:9px;
  padding:11px 14px;color:#111827;font-size:.88rem;font-family:inherit;
  outline:none;margin-bottom:15px;transition:border-color .2s}}
.li:focus{{border-color:#c9a340;background:#fff}}
.li::placeholder{{color:#cbd5e1}}
.lb{{width:100%;padding:12px;border:none;border-radius:9px;cursor:pointer;
  background:linear-gradient(135deg,#1a2b4a,#253d66);color:#fff;font-weight:800;
  font-size:.9rem;font-family:inherit;transition:opacity .2s;margin-top:2px}}
.lb:hover{{opacity:.88}}
.lhint{{text-align:center;font-size:.7rem;color:#cbd5e1;margin-top:18px}}
</style>
</head><body>
<div class="login-box">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:28px;justify-content:center">
    <div style="width:40px;height:40px;background:linear-gradient(135deg,#c9a340,#e8c56a);
      border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:20px">🏀</div>
    <div style="font-size:1.1rem;font-weight:800;letter-spacing:-.02em;color:#1a2b4a">
      Basket<span style="color:#c9a340">Kołcz</span></div>
  </div>
  <div style="text-align:center;font-size:.75rem;color:#9ca3af;margin-bottom:26px;
    text-transform:uppercase;letter-spacing:.04em">Portal zawodnika</div>
  {err_html}
  <form method="POST" action="/portal/login">
    <div class="ll">Login</div>
    <input class="li" type="text" name="login" placeholder="Login" autocomplete="username">
    <div class="ll">Hasło</div>
    <input class="li" type="password" name="password" placeholder="••••••••" autocomplete="current-password">
    <button class="lb" type="submit">Zaloguj się</button>
  </form>
  <div class="lhint">Dane logowania otrzymujesz od trenera</div>
</div>
</body></html>"""
        return html_response(login_page)

    # ── Pobierz dane z bazy ─────────────────────────────────────────────────
    import json as _pj
    _p_klub, _p_sezon, _p_druz = get_portal_context()

    db = get_db(); cur = db.cursor()

    # Wczytaj strukturę klubów
    try:
        _all_kluby = _pj.loads(get_setting("kluby_json") or "[]")
    except Exception:
        _all_kluby = []

    # Lista wszystkich klubów
    portal_kluby = [k["name"] for k in _all_kluby if k.get("name")]

    # Wybrany klub (Klub → Drużyna → Sezon)
    _default_klub = _p_klub or (portal_kluby[0] if portal_kluby else "")
    ctx_klub = request.args.get("klub", _default_klub).strip() or _default_klub

    # Aktywny obiekt klubu
    _ak = next((k for k in _all_kluby if k["name"] == ctx_klub), None)

    # Sezony wybranego klubu
    if _ak:
        portal_sezony = sorted(_ak.get("sezony", {}).keys(), reverse=True)
    else:
        portal_sezony = []
    if not portal_sezony:
        try:
            cur.execute("SELECT DISTINCT sezon FROM matches ORDER BY sezon DESC")
            portal_sezony = [r["sezon"] for r in cur.fetchall() if r["sezon"]]
        except Exception:
            portal_sezony = []

    # Wybrany sezon
    _default_sezon = _p_sezon or (portal_sezony[0] if portal_sezony else "")
    sezon = request.args.get("sezon", _default_sezon).strip() or _default_sezon

    # Drużyny wybranego klubu w wybranym sezonie
    portal_druzyny = (_ak.get("sezony", {}).get(sezon, []) if _ak else [])

    # Wybrana drużyna
    _default_druz = _p_druz or (portal_druzyny[0] if portal_druzyny else "")
    ctx_druzyna = request.args.get("druzyna", _default_druz).strip() or _default_druz

    gtk_name = get_setting("gtk_name") or "GTK"

    # team_id filter
    team_id = None
    if ctx_druzyna:
        try:
            cur.execute("SELECT id FROM teams WHERE name=%s", (ctx_druzyna,))
            tr = cur.fetchone()
            if tr: team_id = tr["id"]
        except Exception: pass

    _cond        = " AND m.team_id=%s" if team_id else ""        # dla queries z aliasem JOIN matches m
    _cond_simple = " AND team_id=%s"   if team_id else ""        # dla prostych FROM matches (bez aliasu)
    _pg          = [sezon] + ([team_id] if team_id else [])
    _pg_opp      = [sezon] + ([team_id] if team_id else [])

    # Liczba meczów + bilans
    if team_id:
        cur.execute("SELECT wynik_gtk,wynik_opp FROM matches WHERE sezon=%s AND team_id=%s", (sezon,team_id))
    else:
        cur.execute("SELECT wynik_gtk,wynik_opp FROM matches WHERE sezon=%s", (sezon,))
    results = cur.fetchall()
    n_m   = max(len(results), 1)
    wins  = sum(1 for r in results if r["wynik_gtk"] > r["wynik_opp"])
    losses= sum(1 for r in results if r["wynik_gtk"] < r["wynik_opp"])

    # GTK agregaty
    try:
        cur.execute(f"""
            SELECT SUM(pts) as pts,SUM(poss) as poss,
                   SUM(p2m) as p2m,SUM(p2a) as p2a,SUM(p3m) as p3m,SUM(p3a) as p3a,
                   SUM(ftm) as ftm,SUM(fta) as fta,SUM(br) as br,SUM(fd) as fd,
                   COALESCE(SUM(ast),0) as ast,COALESCE(SUM(oreb),0) as oreb,
                   COALESCE(SUM(dreb),0) as dreb,COALESCE(SUM(stl),0) as stl,
                   COALESCE(SUM(blk),0) as blk
            FROM match_stats ms JOIN matches m ON ms.match_id=m.id
            WHERE m.sezon=%s AND ms.druzyna='gtk'{_cond}
        """, _pg)
        gtk_tot = dict(cur.fetchone() or {})
    except Exception:
        try: get_db().rollback()
        except: pass
        gtk_tot = {}

    # OPP agregaty
    try:
        cur.execute(f"""
            SELECT SUM(pts) as pts,SUM(poss) as poss,
                   SUM(p2m) as p2m,SUM(p2a) as p2a,SUM(p3m) as p3m,SUM(p3a) as p3a,
                   SUM(ftm) as ftm,SUM(fta) as fta,SUM(br) as br,SUM(fd) as fd,
                   COALESCE(SUM(ast),0) as ast,COALESCE(SUM(oreb),0) as oreb,
                   COALESCE(SUM(dreb),0) as dreb,COALESCE(SUM(stl),0) as stl,
                   COALESCE(SUM(blk),0) as blk
            FROM match_stats ms JOIN matches m ON ms.match_id=m.id
            WHERE m.sezon=%s AND ms.druzyna='opp'{_cond}
        """, _pg_opp)
        opp_tot = dict(cur.fetchone() or {})
    except Exception:
        try: get_db().rollback()
        except: pass
        opp_tot = {}

    # player_stats agregaty (dla GTK)
    try:
        cur.execute(f"""
            SELECT COALESCE(SUM(ps.stl),0) as stl,COALESCE(SUM(ps.blk),0) as blk,
                   COALESCE(SUM(ps.ast),0) as ast,COALESCE(SUM(ps.oreb),0) as oreb,
                   COALESCE(SUM(ps.dreb),0) as dreb
            FROM player_stats ps JOIN matches m ON ps.match_id=m.id
            WHERE m.sezon=%s AND ps.druzyna='gtk'{_cond}
        """, _pg)
        gtk_pl = dict(cur.fetchone() or {})
    except Exception:
        try: get_db().rollback()
        except: pass
        gtk_pl = {}

    # Ostatnie mecze (5) — dla dashboardu
    try:
        cur.execute(f"""
            SELECT id,data_meczu,przeciwnik,wynik_gtk,wynik_opp,miejsce
            FROM matches WHERE sezon=%s{_cond_simple}
            ORDER BY data_meczu DESC LIMIT 5
        """, _pg)
        recent = cur.fetchall()
    except Exception:
        try: get_db().rollback()
        except: pass
        recent = []

    # Wszystkie mecze — pełna historia (zakładka Mecze)
    try:
        cur.execute(f"""
            SELECT id,data_meczu,przeciwnik,wynik_gtk,wynik_opp,miejsce,
                   COALESCE(rozgrywki,'') as rozgrywki, COALESCE(runda,'') as runda,
                   COALESCE(kolejka,'') as kolejka
            FROM matches WHERE sezon=%s{_cond_simple}
            ORDER BY data_meczu DESC
        """, _pg)
        all_matches = cur.fetchall()
    except Exception:
        try: get_db().rollback()
        except: pass
        all_matches = []

    # Zawodnicy — statystyki indywidualne (identyczna struktura jak /zawodnicy)
    try:
        cur.execute("""
            SELECT
                grp_id, nazwa,
                SUM(pts) as pts, SUM(p2m) as p2m, SUM(p2a) as p2a,
                SUM(p3m) as p3m, SUM(p3a) as p3a,
                SUM(ftm) as ftm, SUM(fta) as fta,
                SUM(ast) as ast, SUM(oreb) as oreb, SUM(dreb) as dreb,
                SUM(br) as br, SUM(fd) as fd,
                SUM(stl) as stl, SUM(blk) as blk,
                SUM(finishes) as finishes,
                COUNT(DISTINCT match_id) as mecze,
                SUM(time_sum) as time_sum,
                SUM(poss) as poss
            FROM (
                SELECT
                    CASE WHEN p.id IS NOT NULL THEN 'p_'||p.id::text
                         WHEN r.id IS NOT NULL THEN r.id::text
                         ELSE 'nr_'||ps.nr::text END as grp_id,
                    CASE WHEN p.id IS NOT NULL THEN p.nazwisko||' '||p.imie
                         WHEN r.id IS NOT NULL THEN r.nazwisko||' '||r.imie
                         ELSE '— #'||ps.nr::text END as nazwa,
                    ps.match_id,
                    SUM(ps.pts) as pts, SUM(ps.p2m) as p2m, SUM(ps.p2a) as p2a,
                    SUM(ps.p3m) as p3m, SUM(ps.p3a) as p3a,
                    SUM(ps.ftm) as ftm, SUM(ps.fta) as fta,
                    SUM(ps.ast) as ast, SUM(ps.oreb) as oreb, SUM(ps.dreb) as dreb,
                    SUM(ps.br) as br, SUM(ps.fd) as fd,
                    SUM(COALESCE(ps.stl,0)) as stl, SUM(COALESCE(ps.blk,0)) as blk,
                    SUM(COALESCE(ps.finishes,0)) as finishes,
                    SUM(COALESCE(ps.time_sum,0)) as time_sum,
                    COALESCE((SELECT SUM(poss) FROM match_stats ms2
                               WHERE ms2.match_id=ps.match_id AND ms2.druzyna='gtk'),0) as poss
                FROM player_stats ps
                JOIN matches m ON ps.match_id=m.id
                LEFT JOIN players p ON ps.player_id=p.id
                LEFT JOIN roster r ON ps.roster_id=r.id
                WHERE m.sezon=%s AND ps.druzyna='gtk'
                GROUP BY p.id, p.imie, p.nazwisko, r.id, r.imie, r.nazwisko, ps.nr, ps.match_id
            ) sub
            GROUP BY grp_id, nazwa
            ORDER BY SUM(pts) DESC
        """, (sezon,))
        players = cur.fetchall()
    except Exception:
        try: get_db().rollback()
        except: pass
        cur = get_db().cursor()
        try:
            cur.execute("""
                SELECT ps.nr::text as grp_id,
                       '— #'||ps.nr::text as nazwa,
                       SUM(ps.pts) as pts, SUM(ps.p2m) as p2m, SUM(ps.p2a) as p2a,
                       SUM(ps.p3m) as p3m, SUM(ps.p3a) as p3a,
                       SUM(ps.ftm) as ftm, SUM(ps.fta) as fta,
                       SUM(ps.ast) as ast, SUM(ps.oreb) as oreb, SUM(ps.dreb) as dreb,
                       SUM(ps.br) as br, SUM(ps.fd) as fd,
                       SUM(COALESCE(ps.stl,0)) as stl, SUM(COALESCE(ps.blk,0)) as blk,
                       COUNT(DISTINCT ps.match_id) as mecze,
                       SUM(COALESCE(ps.time_sum,0)) as time_sum
                FROM player_stats ps JOIN matches m ON ps.match_id=m.id
                WHERE m.sezon=%s AND ps.druzyna='gtk'
                GROUP BY ps.nr ORDER BY SUM(ps.pts) DESC
            """, (sezon,))
            players = cur.fetchall()
        except Exception: players = []

    cur.close()

    # ── Helpers ─────────────────────────────────────────────────────────────
    def _av(d, k): return (d.get(k) or 0) / n_m
    def _af(d, k): return f"{_av(d,k):.1f}"
    def _pct(m, a): return f"{m/a*100:.0f}%" if a else "—"

    gtk_kpi = calc_kpi(gtk_tot)
    opp_kpi = calc_kpi(opp_tot)

    try:
        net_v = round(float(gtk_kpi["ortg"]) - float(opp_kpi["ortg"]), 1)
        net_s = f"+{net_v}" if net_v > 0 else str(net_v)
        net_c = "#1a6b3c" if net_v >= 0 else "#8b1a1a"
    except Exception: net_v=0; net_s="—"; net_c="#9ca3af"

    # ── Dashboard ────────────────────────────────────────────────────────────
    g_pts   = int(gtk_tot.get("pts") or 0)
    g_p2m   = int(gtk_tot.get("p2m") or 0); g_p2a = int(gtk_tot.get("p2a") or 0)
    g_p3m   = int(gtk_tot.get("p3m") or 0); g_p3a = int(gtk_tot.get("p3a") or 0)
    g_ftm   = int(gtk_tot.get("ftm") or 0); g_fta = int(gtk_tot.get("fta") or 0)
    g_ast   = int(gtk_pl.get("ast") or 0)
    g_oreb  = int(gtk_pl.get("oreb") or 0); g_dreb = int(gtk_pl.get("dreb") or 0)
    g_stl   = int(gtk_pl.get("stl") or 0)
    g_blk   = int(gtk_pl.get("blk") or 0)
    g_br    = int(gtk_tot.get("br") or 0)
    g_fd    = int(gtk_tot.get("fd") or 0)
    g_reb   = g_oreb + g_dreb
    g_fga   = g_p2a + g_p3a
    fg_pct  = _pct(g_p2m+g_p3m, g_fga)
    p2pct_s = _pct(g_p2m, g_p2a); p3pct_s = _pct(g_p3m, g_p3a)
    ftpct_s = _pct(g_ftm, g_fta)
    ts_s    = gtk_kpi["ts"]

    # Ostatnie mecze HTML
    mc_html = ""
    for r in recent:
        dm = r.get("data_meczu"); rywal = r.get("przeciwnik","?")
        wg = int(r.get("wynik_gtk") or 0); wo = int(r.get("wynik_opp") or 0)
        dat = dm.strftime("%d.%m.%Y") if hasattr(dm,"strftime") else str(dm or "")
        msc = r.get("miejsce","")
        msc_lbl = "Dom" if str(msc).lower() in ("dom","h","home","1") else "Wyjazd"
        w_cls = "w" if wg>wo else "l"
        mc_bg    = "background:linear-gradient(135deg,#f0faf4,#e8f5e9)" if wg>wo else "background:linear-gradient(135deg,#fff5f5,#ffebee)"
        mc_border= "border-color:#c8e6c9" if wg>wo else "border-color:#ffcdd2"
        mid = r.get("id","")
        mc_html += f"""<a href="/portal/mecz/{mid}" style="text-decoration:none;display:block">
<div class="mc" style="{mc_bg};{mc_border};cursor:pointer;transition:transform .15s,box-shadow .15s"
  onmouseover="this.style.transform='translateY(-2px)';this.style.boxShadow='0 6px 18px rgba(0,0,0,.1)'"
  onmouseout="this.style.transform='';this.style.boxShadow=''">
  <div class="mc-date">{dat} &middot; {msc_lbl}</div>
  <div class="mc-row"><div class="mc-team">{gtk_name}</div><div class="mc-score {'w' if wg>wo else 'l'}">{wg}</div></div>
  <div class="mc-row"><div class="mc-team" style="color:#9ca3af">{rywal}</div><div class="mc-score {'l' if wg>wo else 'w'}">{wo}</div></div>
  <div style="text-align:center;margin-top:10px;font-size:.68rem;font-weight:700;color:#6b7280">
    ORtg {gtk_kpi['ortg']} &nbsp;&middot;&nbsp; DRtg {opp_kpi['ortg']}
  </div>
</div></a>"""

    if not mc_html:
        mc_html = '<div style="color:#9ca3af;font-size:.82rem;padding:12px">Brak danych o meczach</div>'

    _safe_klub = ctx_klub.replace(" ", "_").replace("/", "-")
    _klub_logo = get_setting(f"logo_{_safe_klub}") or ""
    _logo_html = (f'<img src="{_klub_logo}" style="height:80px;width:80px;object-fit:contain;flex-shrink:0">'
                  if _klub_logo else "")

    dash_html = f"""
<div class="eyebrow">Sezon {sezon} &middot; {gtk_name}</div>
<div class="hero-strip">
  <div style="display:flex;align-items:center;gap:16px">
    {_logo_html}
    <div>
      <div class="hero-name">{ctx_klub or gtk_name}</div>
      <div class="hero-sub">{ctx_druzyna}</div>
    </div>
  </div>
  <div class="hero-rec">
    <div><div class="hr-stat-val" style="color:#22c55e">{wins}</div><div class="hr-stat-lbl">Wygrane</div></div>
    <div><div class="hr-stat-val" style="color:#ef4444">{losses}</div><div class="hr-stat-lbl">Przegrane</div></div>
    <div><div class="hr-stat-val" style="color:#c9a340">{wins/n_m:.0%}</div><div class="hr-stat-lbl">Win %</div></div>
  </div>
  <div class="hero-rtg">
    <div class="rtg-item"><div class="rtg-val">{gtk_kpi['ortg']}</div><div class="rtg-lbl">ORtg</div></div>
    <div class="rtg-div"></div>
    <div class="rtg-item"><div class="rtg-val">{opp_kpi['ortg']}</div><div class="rtg-lbl">DRtg</div></div>
    <div class="rtg-div"></div>
    <div class="rtg-item"><div class="rtg-val" style="color:#22c55e">{net_s}</div><div class="rtg-lbl">Net</div></div>
  </div>
</div>
<div class="kpi-row c7" style="margin-bottom:10px">
  <div class="kc"><div class="kc-inner">
    <div class="kc-front"><div class="kc-val">{_af(gtk_tot,'pts')}</div><div class="kc-lbl">PPG</div></div>
    <div class="kc-back"><div class="kc-val">{g_pts}</div><div class="kc-lbl">Punkty</div></div>
  </div></div>
  <div class="kc"><div class="kc-inner">
    <div class="kc-front"><div class="kc-val">{g_ast/n_m:.1f}</div><div class="kc-lbl">AST</div></div>
    <div class="kc-back"><div class="kc-val">{g_ast}</div><div class="kc-lbl">Asysty</div></div>
  </div></div>
  <div class="kc"><div class="kc-inner">
    <div class="kc-front"><div class="kc-val">{g_reb/n_m:.1f}</div><div class="kc-lbl">REB</div></div>
    <div class="kc-back"><div class="kc-val">{g_reb}</div><div class="kc-lbl">Zbiórki</div></div>
  </div></div>
  <div class="kc"><div class="kc-inner">
    <div class="kc-front"><div class="kc-val">{g_stl/n_m:.1f}</div><div class="kc-lbl">STL</div></div>
    <div class="kc-back"><div class="kc-val">{g_stl}</div><div class="kc-lbl">Przechwyty</div></div>
  </div></div>
  <div class="kc"><div class="kc-inner">
    <div class="kc-front"><div class="kc-val">{g_blk/n_m:.1f}</div><div class="kc-lbl">BLK</div></div>
    <div class="kc-back"><div class="kc-val">{g_blk}</div><div class="kc-lbl">Bloki</div></div>
  </div></div>
  <div class="kc"><div class="kc-inner">
    <div class="kc-front"><div class="kc-val">{_af(gtk_tot,'br')}</div><div class="kc-lbl">TO</div></div>
    <div class="kc-back"><div class="kc-val">{g_br}</div><div class="kc-lbl">Straty</div></div>
  </div></div>
  <div class="kc"><div class="kc-inner">
    <div class="kc-front"><div class="kc-val">{_af(gtk_tot,'fd')}</div><div class="kc-lbl">FD</div></div>
    <div class="kc-back"><div class="kc-val">{g_fd}</div><div class="kc-lbl">Faulowane</div></div>
  </div></div>
</div>
<div class="kpi-row c5" style="margin-bottom:22px">
  <div class="kc kc-static"><div class="kc-val">{p2pct_s}</div><div class="kc-lbl">2PT%</div></div>
  <div class="kc kc-static"><div class="kc-val">{p3pct_s}</div><div class="kc-lbl">3PT%</div></div>
  <div class="kc kc-static"><div class="kc-val">{ftpct_s}</div><div class="kc-lbl">FT%</div></div>
  <div class="kc kc-static"><div class="kc-val">{fg_pct}</div><div class="kc-lbl">FG%</div></div>
  <div class="kc kc-static"><div class="kc-val">{ts_s}</div><div class="kc-lbl">TS%</div></div>
</div>
<div class="eyebrow" style="margin-bottom:10px">Ostatnie mecze</div>
<div class="matches-grid">{mc_html}</div>"""

    # ── Statystyki drużynowe (duel panels) ──────────────────────────────────
    def _drow(lbl, vg, vo, hib=True, neu=False):
        return _portal_duel_row(lbl, vg, vo, hib, neu)
    def _erow(lbl, vg, vo):
        return _portal_eff_row(lbl, vg, vo)

    dh = ('<div class="dh"><div class="dh-g">'+ gtk_name +'</div><div></div>'
          '<div class="dh-o">Przeciwnicy</div></div>')
    dhs = ('<div class="dh"><div class="dh-g">'+ gtk_name +'</div><div></div>'
           '<div class="dh-o">Prz.</div></div>')

    scoring_h = (dh +
        _drow("PKT",   _af(gtk_tot,"pts"),  _af(opp_tot,"pts")) +
        _drow("AST",   f"{(gtk_pl.get('ast') or 0)/n_m:.1f}", f"{(gtk_pl.get('ast') or 0)/n_m:.1f}".replace(f"{(gtk_pl.get('ast') or 0)/n_m:.1f}", _af(opp_tot,"ast"))) +
        _drow("OREB",  f"{g_oreb/n_m:.1f}", f"{(gtk_pl.get('oreb') or 0)/n_m:.1f}".replace(f"{(gtk_pl.get('oreb') or 0)/n_m:.1f}", _af(opp_tot,"oreb"))) +
        _drow("TO ↓",  _af(gtk_tot,"br"),   _af(opp_tot,"br"),   False) +
        _drow("FD",    _af(gtk_tot,"fd"),   _af(opp_tot,"fd")) +
        _drow("PPP",   gtk_kpi["ppp"],      opp_kpi["ppp"]) +
        _drow("POSS",  _af(gtk_tot,"poss"), _af(opp_tot,"poss"), True, True))

    # fix AST/OREB to use correct gtk_pl values
    def _aplf(k): return f"{(gtk_pl.get(k) or 0)/n_m:.1f}"
    def _aopf(k): return _af(opp_tot, k)

    scoring_h = (dh +
        _drow("PKT",   _af(gtk_tot,"pts"),  _af(opp_tot,"pts")) +
        _drow("AST",   _aplf("ast"),        _aopf("ast")) +
        _drow("OREB",  _aplf("oreb"),       _aopf("oreb")) +
        _drow("TO ↓",  _af(gtk_tot,"br"),   _af(opp_tot,"br"),   False) +
        _drow("FD",    _af(gtk_tot,"fd"),   _af(opp_tot,"fd")) +
        _drow("PPP",   gtk_kpi["ppp"],      opp_kpi["ppp"]) +
        _drow("POSS",  _af(gtk_tot,"poss"), _af(opp_tot,"poss"), True, True))

    defense_h = (dhs +
        _drow("DREB",     _aplf("dreb"),     _aopf("dreb")) +
        _drow("STL",      _aplf("stl"),      _aopf("stl")) +
        _drow("BLK",      _aplf("blk"),      _aopf("blk")) +
        _drow("Faule ↓",  _af(opp_tot,"fd"), _af(gtk_tot,"fd"),  False) +
        _drow("DRtg ↓",   opp_kpi["ortg"],   gtk_kpi["ortg"],    False))

    shooting_h = (dhs +
        _erow("2PT%",    gtk_kpi["p2_pct"],  opp_kpi["p2_pct"]) +
        _erow("3PT%",    gtk_kpi["p3_pct"],  opp_kpi["p3_pct"]) +
        _erow("FT%",     gtk_kpi["ft_pct"],  opp_kpi["ft_pct"]) +
        _erow("eFG%",    gtk_kpi["efg"],     opp_kpi["efg"]) +
        _erow("TS%",     gtk_kpi["ts"],      opp_kpi["ts"]) +
        _erow("FT Rate", gtk_kpi["ftr"],     opp_kpi["ftr"]))

    net_bg = "linear-gradient(135deg,#e8f5e9,#f0faf2)" if net_v >= 0 else "linear-gradient(135deg,#ffebee,#fff5f5)"
    net_border = "#c8e6c9" if net_v >= 0 else "#ffcdd2"

    # ── Mecze — pełna historia ───────────────────────────────────────────────
    mh_rows = ""
    prev_month = None
    for i, am in enumerate(all_matches):
        wg = int(am.get("wynik_gtk") or 0); wo = int(am.get("wynik_opp") or 0)
        is_win = wg > wo
        dm = am.get("data_meczu")
        dat_s = dm.strftime("%d.%m.%Y") if hasattr(dm, "strftime") else str(dm or "—")
        month_s = dm.strftime("%B %Y").capitalize() if hasattr(dm, "strftime") else ""
        msc = str(am.get("miejsce","") or "").lower()
        msc_lbl = "Dom" if msc in ("dom","h","home","1") else ("Wyjazd" if msc else "—")
        rozg = am.get("rozgrywki","") or ""
        runda = am.get("runda","") or ""
        kolejka = am.get("kolejka","") or ""
        # Unikaj duplikacji gdy runda i kolejka mają tę samą wartość
        seen_rk = set()
        rk_parts = []
        for p in [runda, kolejka]:
            if p and p not in seen_rk:
                rk_parts.append(p)
                seen_rk.add(p)
        rk_s = " / ".join(rk_parts)
        mid = am.get("id","")
        wynik_s = f"{wg}:{wo}"
        badge_bg  = "#e8ecf3"
        badge_tc  = "#1a6b3c" if is_win else "#8b1a1a"
        badge_txt = "W" if is_win else "P"
        row_bg    = "#fff"
        row_brd   = "#e8ecf3"
        place_c   = "#185FA5" if msc_lbl == "Dom" else "#6b7280"

        # Separator miesiąca
        if month_s and month_s != prev_month:
            prev_month = month_s
            mh_rows += (f'<div style="font-size:.7rem;font-weight:700;color:#9ca3af;'
                        f'text-transform:uppercase;letter-spacing:.06em;padding:14px 4px 6px">'
                        f'{month_s}</div>')

        mh_rows += f"""<a href="/portal/mecz/{mid}" style="text-decoration:none;display:block;margin-bottom:6px">
<div style="display:flex;align-items:center;gap:10px;padding:10px 14px;border-radius:10px;
  background:{row_bg};border:1px solid {row_brd};transition:box-shadow .15s,transform .15s"
  onmouseover="this.style.boxShadow='0 4px 14px rgba(0,0,0,.09)';this.style.transform='translateY(-1px)'"
  onmouseout="this.style.boxShadow='';this.style.transform=''">
  <div style="width:24px;height:24px;border-radius:6px;background:{badge_bg};display:flex;
    align-items:center;justify-content:center;font-size:.7rem;font-weight:800;color:{badge_tc};flex-shrink:0">{badge_txt}</div>
  <div style="flex:0 0 82px;font-size:.75rem;color:#6b7280">{dat_s}</div>
  {'<div style="flex:0 0 90px;font-size:.68rem;color:#9ca3af;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">' + rozg + '</div>' if rozg else '<div style="flex:0 0 90px"></div>'}
  {'<div style="flex:0 0 70px;font-size:.68rem;color:#6b7280;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">' + rk_s + '</div>' if rk_s else '<div style="flex:0 0 70px"></div>'}
  <div style="flex:0 0 54px;font-size:.7rem;font-weight:600;color:{place_c}">{msc_lbl}</div>
  <div style="flex:1;font-size:.85rem;font-weight:700;color:#1a2b4a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">
    vs. {am.get('przeciwnik','?')}</div>
  <div style="flex:0 0 52px;text-align:right;font-size:.92rem;font-weight:800;color:#1a2b4a">{wynik_s}</div>
  <div style="flex:0 0 20px;text-align:right;color:#9ca3af;font-size:.85rem">›</div>
</div></a>"""

    if not mh_rows:
        mh_rows = '<div style="color:#9ca3af;font-size:.82rem;padding:24px;text-align:center">Brak meczów w tym sezonie</div>'

    n_all = len(all_matches)
    w_all = sum(1 for am in all_matches if (am.get("wynik_gtk") or 0) > (am.get("wynik_opp") or 0))
    l_all = n_all - w_all
    win_pct = f"{w_all/n_all:.0%}" if n_all else "—"

    team_html = f"""

<div style="display:flex;align-items:center;gap:16px;margin-bottom:16px;flex-wrap:wrap">
  <div style="display:flex;gap:10px">
    <div style="background:#e8f5e9;border-radius:8px;padding:8px 18px;text-align:center">
      <div style="font-size:1.4rem;font-weight:800;color:#1a6b3c">{w_all}</div>
      <div style="font-size:.68rem;font-weight:600;color:#4b5563;text-transform:uppercase;letter-spacing:.05em">Wygrane</div>
    </div>
    <div style="background:#ffebee;border-radius:8px;padding:8px 18px;text-align:center">
      <div style="font-size:1.4rem;font-weight:800;color:#8b1a1a">{l_all}</div>
      <div style="font-size:.68rem;font-weight:600;color:#4b5563;text-transform:uppercase;letter-spacing:.05em">Przegrane</div>
    </div>
    <div style="background:#f0f4ff;border-radius:8px;padding:8px 18px;text-align:center">
      <div style="font-size:1.4rem;font-weight:800;color:#1a2b4a">{win_pct}</div>
      <div style="font-size:.68rem;font-weight:600;color:#4b5563;text-transform:uppercase;letter-spacing:.05em">Win%</div>
    </div>
  </div>
  <div style="margin-left:auto;font-size:.75rem;color:#9ca3af">{n_all} meczów</div>
</div>
<div>{mh_rows}</div>"""

    # ── Zawodnicy — karty + tabela ───────────────────────────────────────────
    th_s = 'background:#1a2b4a;color:#fff;font-size:.65rem;font-weight:600;padding:8px 6px;text-align:center;white-space:nowrap;cursor:pointer;user-select:none'
    th_sc= 'background:#1a2b4a;color:#fff;font-size:.65rem;font-weight:600;padding:8px 6px;text-align:center;white-space:nowrap'
    th_l = 'background:#1a2b4a;color:#fff;font-size:.65rem;font-weight:600;padding:8px 10px;text-align:left;white-space:nowrap'
    th_g = 'background:#1a2b4a;color:rgba(255,255,255,.55);font-size:.62rem;padding:3px 4px 2px;border-bottom:.5px solid rgba(255,255,255,.15);text-align:center'
    th_gs= 'background:#1a2b4a;color:rgba(255,255,255,.75);font-size:.63rem;padding:2px 4px 5px;text-align:center;cursor:pointer;user-select:none'
    th_z = 'background:#1e3459;color:rgba(255,255,255,.55);font-size:.62rem;padding:3px 4px 2px;border-bottom:.5px solid rgba(255,255,255,.1);text-align:center'
    th_zs= 'background:#1e3459;color:rgba(255,255,255,.75);font-size:.63rem;padding:2px 4px 5px;text-align:center;cursor:pointer;user-select:none'
    vm   = 'vertical-align:middle;border-bottom:.5px solid rgba(255,255,255,.15)'
    thead_p = f"""<thead>
      <tr>
        <th style="{th_l};{vm}" rowspan="3">Zawodnik</th>
        <th style="{th_sc};{vm}" rowspan="3" onclick="sortP(1)"><span id="ph1">MIN<br>(szac.)</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortP(2)"><span id="ph2">G</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortP(3)"><span id="ph3">PTS ↓</span></th>
        <th style="{th_g}" colspan="2">2PT</th>
        <th style="{th_g}" colspan="2">3PT</th>
        <th style="{th_g}" colspan="2">FT</th>
        <th style="{th_z}" colspan="3">ZB</th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortP(13)"><span id="ph13">AST</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortP(14)"><span id="ph14">TO</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortP(15)"><span id="ph15">STL</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortP(16)"><span id="ph16">BLK</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortP(17)"><span id="ph17">FD</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortP(18)"><span id="ph18">eFG%</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortP(19)"><span id="ph19">TS%</span></th>
        <th style="{th_s};{vm}" rowspan="3" onclick="sortP(20)"><span id="ph20">USG%</span></th>
      </tr>
      <tr>
        <th style="{th_gs}" onclick="sortP(4)"><span id="ph4">M/A</span></th>
        <th style="{th_gs}" onclick="sortP(5)"><span id="ph5">%</span></th>
        <th style="{th_gs}" onclick="sortP(6)"><span id="ph6">M/A</span></th>
        <th style="{th_gs}" onclick="sortP(7)"><span id="ph7">%</span></th>
        <th style="{th_gs}" onclick="sortP(8)"><span id="ph8">M/A</span></th>
        <th style="{th_gs}" onclick="sortP(9)"><span id="ph9">%</span></th>
        <th style="{th_zs}" onclick="sortP(10)"><span id="ph10">A</span></th>
        <th style="{th_zs}" onclick="sortP(11)"><span id="ph11">O</span></th>
        <th style="{th_zs}" onclick="sortP(12)"><span id="ph12">S</span></th>
      </tr>
    </thead>"""

    # ── Preload calc_play_time for portal zawodnicy (niezależnie od DB time_sum) ──
    _pt_by_grp_portal = {}
    try:
        _cur_pt = get_db().cursor()
        if team_id:
            _cur_pt.execute("SELECT id FROM matches WHERE sezon=%s AND team_id=%s", (sezon, team_id))
        else:
            _cur_pt.execute("SELECT id FROM matches WHERE sezon=%s", (sezon,))
        _match_ids_pt = [_r["id"] for _r in _cur_pt.fetchall()]
        _pt_all = {}
        for _mid in _match_ids_pt:
            for _nr, _secs in calc_play_time(_mid).items():
                _pt_all[(_mid, int(_nr))] = _secs
        _args_pt = [sezon] + ([team_id] if team_id else [])
        _cur_pt.execute(f"""
            SELECT DISTINCT
                CASE WHEN p.id IS NOT NULL THEN 'p_'||p.id::text
                     WHEN r.id IS NOT NULL THEN r.id::text
                     ELSE 'nr_'||ps.nr::text END as grp_id,
                ps.match_id, ps.nr
            FROM player_stats ps JOIN matches m ON ps.match_id=m.id
            LEFT JOIN players p ON ps.player_id=p.id
            LEFT JOIN roster r ON ps.roster_id=r.id
            WHERE m.sezon=%s AND ps.druzyna='gtk'
            {"AND m.team_id=%s" if team_id else ""}
        """, _args_pt)
        for _gr in _cur_pt.fetchall():
            _g = _gr["grp_id"]
            _s = _pt_all.get((_gr["match_id"], int(_gr["nr"] or 0)), 0)
            _pt_by_grp_portal[_g] = _pt_by_grp_portal.get(_g, 0) + _s
        _cur_pt.close()
    except Exception:
        pass

    p_rows = ""
    for pl in players:
        def _pi(k): return int(pl.get(k,0) or 0)
        n = max(_pi("mecze"),1)
        pm2=_pi("p2m"); pa2=_pi("p2a"); pm3=_pi("p3m"); pa3=_pi("p3a")
        ftm=_pi("ftm"); fta=_pi("fta"); pts=_pi("pts")
        ast=_pi("ast"); oreb=_pi("oreb"); dreb=_pi("dreb")
        br=_pi("br"); fd=_pi("fd"); stl=_pi("stl"); blk=_pi("blk")
        fga=pa2+pa3
        efg_v = f"{(pm2+1.5*pm3)/fga*100:.1f}%" if fga else "—"
        ts_v  = f"{pts/(2*(fga+0.44*fta))*100:.1f}%" if (fga+fta) else "—"
        poss_pl = float(pl.get("poss") or 0)
        usg_v = f"{(fga+0.44*fta+br)/poss_pl*100:.1f}%" if poss_pl else "—"
        fin_v = _pi("finishes")
        p2p   = f"{pm2/pa2*100:.0f}%" if pa2 else "—"
        p3p   = f"{pm3/pa3*100:.0f}%" if pa3 else "—"
        ftp   = f"{ftm/fta*100:.0f}%" if fta else "—"
        _grp_id_pl = pl.get("grp_id", "")
        _pt_total_pl = _pt_by_grp_portal.get(_grp_id_pl, 0)
        if _pt_total_pl and n:
            am = _pt_total_pl / n / 60
            min_s = f"{int(am)}:{int((am%1)*60):02d}"
        else: min_s = "—"
        def _a(v): return f"{v/n:.1f}"
        grp = pl.get("grp_id",""); nazwa = pl.get("nazwa","?")
        if grp.startswith("p_"):
            pid = grp[2:]
            nc = f'<a href="/portal/zawodnik/{pid}?sezon={sezon}" style="color:#1a2b4a;text-decoration:none;font-weight:600">{nazwa}</a>'
        elif grp.isdigit():
            nc = f'<a href="/portal/zawodnik/{grp}?sezon={sezon}" style="color:#1a2b4a;text-decoration:none;font-weight:600">{nazwa}</a>'
        else:
            nc = f'<span style="font-weight:500">{nazwa}</span>'
        efg_c = "#0F6E56" if fga and (pm2+1.5*pm3)/fga>=0.5 else ("#A32D2D" if fga and (pm2+1.5*pm3)/fga<0.35 else "inherit")
        _total_min_pl = _pt_total_pl / 60 if _pt_total_pl else 0
        p_rows += f"""<tr data-n="{n}" data-min="{_total_min_pl:.3f}"
            data-pts="{pts}" data-p2m="{pm2}" data-p2a="{pa2}"
            data-p3m="{pm3}" data-p3a="{pa3}" data-ftm="{ftm}" data-fta="{fta}"
            data-oreb="{oreb}" data-dreb="{dreb}" data-ast="{ast}" data-br="{br}"
            data-stl="{stl}" data-blk="{blk}" data-fd="{fd}" data-fin="{fin_v}">
            <td class="pn">{nc}</td>
            <td class="min" data-cell="min">{min_s}</td>
            <td style="color:#9ca3af;font-size:.72rem">{n}</td>
            <td class="pts" data-cell="pts">{_a(pts)}</td>
            <td data-cell="p2ma">{pm2/n:.1f}/{pa2/n:.1f}</td><td>{p2p}</td>
            <td data-cell="p3ma">{pm3/n:.1f}/{pa3/n:.1f}</td><td>{p3p}</td>
            <td data-cell="ftma">{ftm/n:.1f}/{fta/n:.1f}</td><td>{ftp}</td>
            <td data-cell="oreb">{_a(oreb)}</td><td data-cell="dreb">{_a(dreb)}</td>
            <td data-cell="zbs">{_a(oreb+dreb)}</td>
            <td data-cell="ast">{_a(ast)}</td><td data-cell="br">{_a(br)}</td>
            <td data-cell="stl">{_a(stl)}</td><td data-cell="blk">{_a(blk)}</td><td data-cell="fd">{_a(fd)}</td>
            <td style="color:{efg_c};font-weight:700">{efg_v}</td>
            <td>{ts_v}</td>
            <td>{usg_v}</td>
        </tr>"""

    if not p_rows:
        p_rows = '<tr><td colspan="21" style="text-align:center;color:#9ca3af;padding:24px">Brak danych</td></tr>'

    players_html = f"""
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px;flex-wrap:wrap;gap:10px">
  <div style="border-left:3px solid #1a2b4a;padding-left:11px">
    <div style="font-size:.95rem;font-weight:700;color:#1a2b4a;line-height:1.2">Statystyki indywidualne</div>
    <div id="p-per-desc" style="font-size:.67rem;color:#bbb;margin-top:2px">Wartości uśrednione &middot; G = liczba meczów</div>
  </div>
  <div style="display:flex;align-items:center;gap:10px">
    <div style="display:flex;background:#1a2b4a;border-radius:20px;padding:2px;gap:0;box-shadow:0 2px 8px rgba(0,0,0,.18)">
      <button id="pStatAvg" onclick="pSetStatMode('avg')"
        style="border:none;border-radius:18px;padding:5px 13px;font-size:.74rem;font-weight:600;
               cursor:pointer;background:#fff;color:#1a2b4a;transition:.18s;white-space:nowrap">Średnie</button>
      <button id="pStatSum" onclick="pSetStatMode('sum')"
        style="border:none;border-radius:18px;padding:5px 13px;font-size:.74rem;font-weight:600;
               cursor:pointer;background:transparent;color:rgba(255,255,255,.7);transition:.18s;white-space:nowrap">Sumaryczne</button>
    </div>
    <div style="position:relative;display:inline-block" id="pPerWrap">
      <select id="pPerSelect" onchange="pSetPer(this.value)"
        style="background:#fff;color:#1a2b4a;border:none;border-radius:20px;
               padding:6px 32px 6px 16px;font-size:.78rem;font-weight:600;
               cursor:pointer;appearance:none;-webkit-appearance:none;
               outline:none;box-shadow:0 2px 8px rgba(0,0,0,.18);
               letter-spacing:.2px;min-width:115px">
        <option value="game" selected>Per Mecz</option>
        <option value="36">Per 36 min</option>
        <option value="40">Per 40 min</option>
        <option value="100">Per 100 pos</option>
      </select>
      <span style="position:absolute;right:12px;top:50%;transform:translateY(-50%);
                   pointer-events:none;color:#1a2b4a;font-size:.6rem">▼</span>
    </div>
  </div>
</div>
<div style="overflow-x:auto;-webkit-overflow-scrolling:touch" class="ind-wrap ind-wrap-scroll">
  <table class="ind-tbl" id="ptbl">
    {thead_p}
    <tbody id="ptbody">{p_rows}</tbody>
  </table>
</div>"""

    # ── Złóż stronę ──────────────────────────────────────────────────────────
    # Sidebar logo: use app logo from DB (app_logo_b64), fallback to "BK"
    _app_logo_uri = get_setting("app_logo_b64") or ""
    _logo_html = (f'<img src="{_app_logo_uri}" alt="BasketKołcz" style="width:100%;height:100%;object-fit:cover;border-radius:10px">'
                  if _app_logo_uri else '<span class="sb-logo-fallback">BK</span>')

    # Build sidebar dropdown items
    def _sb_dd(items, active_val, nav_k, nav_d, nav_s):
        if not items:
            return f'<div class="sb-dd-item sel">{active_val or "—"}</div>'
        rows = []
        for v in items:
            k = nav_k(v); d = nav_d(v); s = nav_s(v)
            cls = ' sel' if v == active_val else ''
            rows.append(f'<div class="sb-dd-item{cls}" onclick="navPortal(event,\\"{k}\\",\\"{d}\\",\\"{s}\\")">{v}</div>')
        return "".join(rows)

    _klub_dd  = _sb_dd(portal_kluby,   ctx_klub,     lambda v: v,       lambda v: "",          lambda v: "")
    _sezon_dd = _sb_dd(portal_sezony,  sezon,        lambda v: ctx_klub, lambda v: "",          lambda v: v)
    _druz_dd  = _sb_dd(portal_druzyny, ctx_druzyna,  lambda v: ctx_klub, lambda v: v,           lambda v: sezon)

    portal_js = """<script>
function showPane(id,btn){
  document.querySelectorAll('.pane').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.sb-nav-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById('pane-'+id).classList.add('active');
  if(btn) btn.classList.add('active');
}
var _pd={};
var _pPerMode='game', _pStatMode='avg';
function _pf1(v){return v.toFixed(1);}
function pSetStatMode(mode){
  _pStatMode=mode;
  var ab=document.getElementById('pStatAvg'), sb=document.getElementById('pStatSum');
  if(ab&&sb){
    if(mode==='avg'){ab.style.background='#fff';ab.style.color='#1a2b4a';sb.style.background='transparent';sb.style.color='rgba(255,255,255,.7)';}
    else{sb.style.background='#fff';sb.style.color='#1a2b4a';ab.style.background='transparent';ab.style.color='rgba(255,255,255,.7)';}
  }
  var pw=document.getElementById('pPerWrap');
  if(pw){pw.style.opacity=mode==='sum'?'0.38':'1';pw.style.pointerEvents=mode==='sum'?'none':'';}
  if(mode==='avg'){pSetPer(_pPerMode);}
  else{
    var desc=document.getElementById('p-per-desc');
    if(desc) desc.textContent='Wartości sumaryczne za cały sezon  ·  G = liczba meczów';
    var tb=document.getElementById('ptbody'); if(!tb) return;
    tb.querySelectorAll('tr').forEach(function(row){
      row.querySelectorAll('[data-cell]').forEach(function(td){
        var c=td.dataset.cell, iv=function(k){return parseInt(row.dataset[k])||0;};
        if(c==='pts')      td.textContent=iv('pts');
        else if(c==='p2ma') td.textContent=iv('p2m')+'/'+iv('p2a');
        else if(c==='p3ma') td.textContent=iv('p3m')+'/'+iv('p3a');
        else if(c==='ftma') td.textContent=iv('ftm')+'/'+iv('fta');
        else if(c==='oreb') td.textContent=iv('oreb');
        else if(c==='dreb') td.textContent=iv('dreb');
        else if(c==='zbs')  td.textContent=iv('oreb')+iv('dreb');
        else if(c==='ast')  td.textContent=iv('ast');
        else if(c==='br')   td.textContent=iv('br');
        else if(c==='stl')  td.textContent=iv('stl');
        else if(c==='blk')  td.textContent=iv('blk');
        else if(c==='fd')   td.textContent=iv('fd');
        else if(c==='min'){var m=parseFloat(row.dataset.min)||0;td.textContent=m?Math.floor(m)+':'+String(Math.round((m%1)*60)).padStart(2,'0'):'—';}
      });
    });
  }
}
function pSetPer(mode){
  _pPerMode=mode;
  if(_pStatMode==='sum') return;
  var desc=document.getElementById('p-per-desc');
  var labels={'game':'Wartości uśrednione per mecz w którym zawodnik zagrał  ·  G = liczba meczów','36':'Statystyki przeliczone na 36 minut gry  ·  G = liczba meczów','40':'Statystyki przeliczone na 40 minut gry  ·  G = liczba meczów','100':'Statystyki przeliczone na 100 posiadań  ·  G = liczba meczów'};
  if(desc) desc.textContent=labels[mode]||labels['game'];
  var tb=document.getElementById('ptbody'); if(!tb) return;
  var scale=mode==='36'?36:mode==='40'?40:mode==='100'?100:0;
  tb.querySelectorAll('tr').forEach(function(row){
    var n=parseFloat(row.dataset.n)||1, min=parseFloat(row.dataset.min)||0, has=min>0;
    function pg(k){var v=parseFloat(row.dataset[k])||0;if(scale===0)return _pf1(v/n);if(!has)return '—';return _pf1(v/min*scale);}
    function pgma(km,ka){var vm=parseFloat(row.dataset[km])||0,va=parseFloat(row.dataset[ka])||0;if(scale===0)return _pf1(vm/n)+'/'+_pf1(va/n);if(!has)return '—/—';return _pf1(vm/min*scale)+'/'+_pf1(va/min*scale);}
    row.querySelectorAll('[data-cell]').forEach(function(td){
      var c=td.dataset.cell;
      if(c==='pts') td.textContent=pg('pts');
      else if(c==='p2ma') td.textContent=pgma('p2m','p2a');
      else if(c==='p3ma') td.textContent=pgma('p3m','p3a');
      else if(c==='ftma') td.textContent=pgma('ftm','fta');
      else if(c==='oreb') td.textContent=pg('oreb');
      else if(c==='dreb') td.textContent=pg('dreb');
      else if(c==='zbs')  {var o=parseFloat(row.dataset.oreb)||0,d=parseFloat(row.dataset.dreb)||0;td.textContent=scale===0?_pf1((o+d)/n):(has?_pf1((o+d)/min*scale):'—');}
      else if(c==='ast')  td.textContent=pg('ast');
      else if(c==='br')   td.textContent=pg('br');
      else if(c==='stl')  td.textContent=pg('stl');
      else if(c==='blk')  td.textContent=pg('blk');
      else if(c==='fd')   td.textContent=pg('fd');
    });
  });
}
function sortP(col){
  var tb=document.getElementById('ptbody'); if(!tb) return;
  var rows=Array.from(tb.querySelectorAll('tr'));
  _pd[col]=!_pd[col];
  document.querySelectorAll('[id^="ph"]').forEach(function(el){
    el.textContent=el.textContent.replace(/ [▲▼]$/,'');
  });
  var ph=document.getElementById('ph'+col);
  if(ph) ph.textContent+=_pd[col]?' ▼':' ▲';
  rows.sort(function(a,b){
    var av=(a.cells[col]||{}).textContent||'';
    var bv=(b.cells[col]||{}).textContent||'';
    av=av.trim().replace('%','').split('/')[0];
    bv=bv.trim().replace('%','').split('/')[0];
    var an=parseFloat(av), bn=parseFloat(bv);
    if(!isNaN(an)&&!isNaN(bn)) return _pd[col]?bn-an:an-bn;
    return _pd[col]?bv.localeCompare(av,'pl'):av.localeCompare(bv,'pl');
  });
  rows.forEach(function(r){tb.appendChild(r);});
}
window.addEventListener('DOMContentLoaded',function(){
  sortP(3);
  var tab=new URLSearchParams(window.location.search).get('tab');
  if(tab==='players'){
    var btn=document.querySelector('.sb-nav-btn[data-tab="players"]');
    if(btn) showPane('players',btn);
  } else if(tab==='team'){
    var btn=document.querySelector('.sb-nav-btn[data-tab="team"]');
    if(btn) showPane('team',btn);
  }
});
function toggleSB(id){
  var all=document.querySelectorAll('.sb-sel-group');
  all.forEach(function(el){
    if(el.id!==id){el.classList.remove('open');var h=el.querySelector('.sb-sel-hdr');if(h)h.classList.remove('open');}
  });
  var el=document.getElementById(id);
  var h=el.querySelector('.sb-sel-hdr');
  el.classList.toggle('open');
  if(h) h.classList.toggle('open');
}
function navPortal(e,klub,druzyna,sezon){
  e.stopPropagation();
  var p=new URLSearchParams();
  if(klub)    p.set('klub',   klub);
  if(druzyna) p.set('druzyna',druzyna);
  if(sezon)   p.set('sezon',  sezon);
  window.location.href='/portal?'+p.toString();
}
document.addEventListener('click',function(e){
  if(!e.target.closest('.sb-sel-group')){
    document.querySelectorAll('.sb-sel-group').forEach(function(el){
      el.classList.remove('open');
      var h=el.querySelector('.sb-sel-hdr');if(h)h.classList.remove('open');
    });
  }
});
</script>"""

    full = f"""<!DOCTYPE html>
<html lang="pl"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Portal — BasketKołcz</title>
{_PORTAL_CSS}
</head><body>
<div class="portal-layout">

  <!-- ── SIDEBAR ─────────────────────────── -->
  <aside class="sidebar">

    <!-- Brand / logo -->
    <div class="sb-brand">
      <div class="sb-logo-wrap">{_logo_html}</div>
      <div class="sb-brand-text">Basket<span>Kołcz</span></div>
    </div>
    <div class="sb-divider" style="margin-top:0"></div>

    <!-- Selectors -->
    <div class="sb-selectors">

      <!-- Klub -->
      <div class="sb-sel-group" id="sb-klub">
        <button class="sb-sel-hdr" onclick="toggleSB('sb-klub')">
          <div class="sb-sel-icon-box">🏛</div>
          <div class="sb-sel-texts">
            <span class="sb-sel-lbl">Klub</span>
            <span class="sb-sel-val">{ctx_klub or '—'}</span>
          </div>
          <span class="sb-sel-arr">▼</span>
        </button>
        <div class="sb-dropdown">{_klub_dd}</div>
      </div>

      <!-- Sezon -->
      <div class="sb-sel-group" id="sb-sezon">
        <button class="sb-sel-hdr" onclick="toggleSB('sb-sezon')">
          <div class="sb-sel-icon-box">📅</div>
          <div class="sb-sel-texts">
            <span class="sb-sel-lbl">Sezon</span>
            <span class="sb-sel-val">{sezon or '—'}</span>
          </div>
          <span class="sb-sel-arr">▼</span>
        </button>
        <div class="sb-dropdown">{_sezon_dd}</div>
      </div>

      <!-- Drużyna -->
      <div class="sb-sel-group" id="sb-druzyna">
        <button class="sb-sel-hdr" onclick="toggleSB('sb-druzyna')">
          <div class="sb-sel-icon-box">🏀</div>
          <div class="sb-sel-texts">
            <span class="sb-sel-lbl">Drużyna</span>
            <span class="sb-sel-val">{ctx_druzyna or '—'}</span>
          </div>
          <span class="sb-sel-arr">▼</span>
        </button>
        <div class="sb-dropdown">{_druz_dd}</div>
      </div>

    </div>
    <div class="sb-divider"></div>

    <!-- Navigation -->
    <div class="sb-nav-section">
      <button class="sb-nav-btn active" data-tab="dash" onclick="showPane('dash',this)">
        <span class="sb-nav-icon">📊</span>Główny Pulpit
      </button>
      <button class="sb-nav-btn" data-tab="team" onclick="showPane('team',this)">
        <span class="sb-nav-icon">🏆</span>Mecze
      </button>
      <button class="sb-nav-btn" data-tab="players" onclick="showPane('players',this)">
        <span class="sb-nav-icon">👤</span>Statystyki
      </button>
    </div>

    <!-- Footer -->
    <div class="sb-footer">
      <a class="sb-logout-btn" href="/portal/logout">Wyloguj</a>
    </div>

  </aside>

  <!-- ── MAIN CONTENT ──────────────────── -->
  <main class="portal-main">
    <div id="pane-dash" class="pane active">{dash_html}</div>
    <div id="pane-team" class="pane">{team_html}</div>
    <div id="pane-players" class="pane">{players_html}</div>
  </main>

</div>
{portal_js}
</body></html>"""
    return html_response(full)


@app.route("/portal/mecz/<int:match_id>")
def portal_mecz(match_id):
    if not session.get("portal_logged_in"):
        return redirect(url_for("portal"))

    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
    m = cur.fetchone()
    if not m:
        cur.close()
        return redirect(url_for("portal"))

    gtk_name = (m.get("team_name_a","") or get_setting("gtk_name") or "GTK").strip()
    name_opp = (m.get("team_name_b","") or m.get("przeciwnik","Rywal")).strip()

    cur.execute("SELECT * FROM match_stats WHERE match_id=%s ORDER BY kwarta", (match_id,))
    all_stats = cur.fetchall()

    cur.execute("SELECT * FROM player_stats WHERE match_id=%s", (match_id,))
    all_players = cur.fetchall()

    cur.execute("SELECT * FROM timing_stats WHERE match_id=%s", (match_id,))
    all_timing = cur.fetchall()

    try:
        cur.execute("SELECT * FROM lineup_stats WHERE match_id=%s AND druzyna='gtk' ORDER BY poss DESC", (match_id,))
        all_lineups = list(cur.fetchall())
    except: all_lineups = []

    try:
        cur.execute("SELECT * FROM lineup_stats WHERE match_id=%s AND druzyna='gtk_def' ORDER BY poss DESC", (match_id,))
        all_lineups_def = list(cur.fetchall())
    except: all_lineups_def = []

    try:
        cur.execute("SELECT kwarta,czas_sek,pts_gtk,pts_opp FROM score_flow WHERE match_id=%s ORDER BY kwarta,czas_sek", (match_id,))
        flow_rows = list(cur.fetchall())
    except: flow_rows = []

    nr_name_map = build_nr_name_map(cur, match_id)
    cur.close()

    play_time_secs = calc_play_time(match_id)   # Excel-based timing (×1.22 already applied)

    # ── Helpers (analogiczne do mecz()) ─────────────────────────────────────
    def build_suma(druzyna):
        s = {"pts":0,"poss":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"br":0,"fd":0,
             "ast":0,"oreb":0,"dreb":0,"stl":0,"blk":0,"d2m":0,"d2a":0,"przerw":0}
        for row in all_stats:
            if row["druzyna"] == druzyna:
                for k in s: s[k] += row.get(k,0) or 0
        return s

    suma_gtk = build_suma("gtk")
    suma_opp = build_suma("opp")
    kpi_gtk  = calc_kpi(suma_gtk)
    kpi_opp  = calc_kpi(suma_opp)
    dt = m['data_meczu'].strftime('%d.%m.%Y') if m.get('data_meczu') else ""

    def kpi_cards(suma, kpi):
        reb = (suma.get("oreb",0) or 0) + (suma.get("dreb",0) or 0)
        ast = suma.get("ast",0) or 0
        ks  = 'background:#f4f6fb;border-radius:8px;padding:14px 12px;text-align:center'
        kv  = 'font-size:22px;font-weight:500;color:#1a2b4a'
        kl  = 'font-size:10px;color:#999;text-transform:uppercase;letter-spacing:.5px;margin-top:2px'
        ka  = 'background:#E6F1FB;border-radius:8px;padding:14px 12px;text-align:center'
        kav = 'font-size:22px;font-weight:500;color:#0C447C'
        kal = 'font-size:10px;color:#185FA5;text-transform:uppercase;letter-spacing:.5px;margin-top:2px'
        def pct_bar(val_str):
            try: v = float(str(val_str).replace("%",""))
            except: v = 0
            w = min(int(v), 100)
            return f'<div style="flex:1;height:5px;background:#e0e0e0;border-radius:3px;overflow:hidden;margin-left:8px"><div style="width:{w}%;height:100%;background:#185FA5;border-radius:3px"></div></div>'
        r1 = f'''<div style="display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:8px;margin-bottom:8px">
  <div style="{ks}"><div style="{kv}">{suma.get("pts",0)}</div><div style="{kl}">Punkty</div></div>
  <div style="{ks}"><div style="{kv}">{ast}</div><div style="{kl}">Asysty</div></div>
  <div style="{ks}"><div style="{kv}">{reb}</div><div style="{kl}">Zbiórki</div></div>
  <div style="{ks}"><div style="{kv}">{suma.get("br",0)}</div><div style="{kl}">Straty</div></div>
</div>'''
        r2 = f'''<div style="display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:8px;margin-bottom:8px">
  <div style="{ka}"><div style="{kav}">{kpi["ortg"]}</div><div style="{kal}">ORtg</div></div>
  <div style="{ka}"><div style="{kav}">{kpi["ppp"]}</div><div style="{kal}">PPP</div></div>
  <div style="{ka}"><div style="{kav}">{kpi["efg"]}</div><div style="{kal}">eFG%</div></div>
  <div style="{ka}"><div style="{kav}">{kpi["ts"]}</div><div style="{kal}">TS%</div></div>
</div>'''
        r3 = f'''<div style="display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:8px;margin-bottom:8px">
  <div style="{ks};display:flex;align-items:center;padding:12px">
    <div><div style="{kl}">2PT%</div><div style="{kv};font-size:18px">{kpi["p2_pct"]}</div></div>
    {pct_bar(kpi["p2_pct"])}</div>
  <div style="{ks};display:flex;align-items:center;padding:12px">
    <div><div style="{kl}">3PT%</div><div style="{kv};font-size:18px">{kpi["p3_pct"]}</div></div>
    {pct_bar(kpi["p3_pct"])}</div>
  <div style="{ks};display:flex;align-items:center;padding:12px">
    <div><div style="{kl}">FT%</div><div style="{kv};font-size:18px">{kpi["ft_pct"]}</div></div>
    {pct_bar(kpi["ft_pct"])}</div>
</div>'''
        return r1 + r2 + r3

    def q_table(druzyna):
        opp_druzyna = "opp" if druzyna == "gtk" else "gtk"
        rows = ""
        for qn in [1,2,3,4]:
            qd  = next((r for r in all_stats if r["druzyna"]==druzyna and r["kwarta"]==qn), {})
            oqd = next((r for r in all_stats if r["druzyna"]==opp_druzyna and r["kwarta"]==qn), {})
            p2m=qd.get("p2m",0) or 0; p2a=qd.get("p2a",0) or 0
            p3m=qd.get("p3m",0) or 0; p3a=qd.get("p3a",0) or 0
            ftm=qd.get("ftm",0) or 0; fta=qd.get("fta",0) or 0
            pts=qd.get("pts",0) or 0; poss=qd.get("poss",0) or 0
            ast=qd.get("ast",0) or 0; oreb=qd.get("oreb",0) or 0; dreb=qd.get("dreb",0) or 0
            to_=qd.get("br",0) or 0;  fd_=qd.get("fd",0) or 0; prz=qd.get("przerw",0) or 0
            opp_pts=oqd.get("pts",0) or 0; opp_poss=oqd.get("poss",0) or 0
            efg_v = round((p2m+1.5*p3m)/(p2a+p3a)*100) if (p2a+p3a) else None
            efg_s = ("%d%%" % efg_v) if efg_v is not None else "—"
            efg_cls = "mgood" if efg_v and efg_v>=50 else ("mbad" if efg_v and efg_v<35 else "")
            p2p_s = ("%d%%" % round(p2m/p2a*100)) if p2a else "—"
            p3p_s = ("%d%%" % round(p3m/p3a*100)) if p3a else "—"
            ftp_s = ("%d%%" % round(ftm/fta*100)) if fta else "—"
            ortg_s = ("%.1f" % (pts*100/poss)) if poss else "—"
            drtg_s = ("%.1f" % (opp_pts*100/opp_poss)) if opp_poss else "—"
            qd_cls = {1:"mq1",2:"mq2",3:"mq3",4:"mq4"}[qn]
            td = 'style="text-align:center;padding:7px 10px"'
            rows += (f'<tr>'
                f'<td style="text-align:left;padding:7px 10px"><span class="mqdot {qd_cls}">{qn}Q</span></td>'
                f'<td {td} style="text-align:center;padding:7px 10px;font-weight:500">{pts}</td>'
                f'<td {td}>{p2m}/{p2a}</td><td {td}>{p2p_s}</td>'
                f'<td {td}>{p3m}/{p3a}</td><td {td}>{p3p_s}</td>'
                f'<td {td}>{ftm}/{fta}</td><td {td}>{ftp_s}</td>'
                f'<td {td}>{oreb}</td><td {td}>{dreb}</td>'
                f'<td {td}>{ast}</td><td {td}>{prz}</td><td {td}>{to_}</td>'
                f'<td {td}>{poss}</td><td {td}>{fd_}</td>'
                f'<td {td} class="{efg_cls}">{efg_s}</td><td {td}>{ortg_s}</td><td {td}>{drtg_s}</td>'
                f'</tr>')
        def _s(f): return sum(r.get(f,0) or 0 for r in all_stats if r["druzyna"]==druzyna and r["kwarta"] in [1,2,3,4])
        def _os(f): return sum(r.get(f,0) or 0 for r in all_stats if r["druzyna"]==opp_druzyna and r["kwarta"] in [1,2,3,4])
        sp2m=_s("p2m");sp2a=_s("p2a");sp3m=_s("p3m");sp3a=_s("p3a")
        sftm=_s("ftm");sfta=_s("fta");spts=_s("pts");sposs=_s("poss")
        sast=_s("ast");soreb=_s("oreb");sdreb=_s("dreb");sprz=_s("przerw")
        sto=_s("br");sfd=_s("fd")
        sopp_pts=_os("pts");sopp_poss=_os("poss")
        sefg_v = round((sp2m+1.5*sp3m)/(sp2a+sp3a)*100) if (sp2a+sp3a) else None
        sefg_s = ("%d%%" % sefg_v) if sefg_v is not None else "—"
        sefg_c = "mgood" if sefg_v and sefg_v>=50 else ("mbad" if sefg_v and sefg_v<35 else "")
        sp2p_s = ("%d%%" % round(sp2m/sp2a*100)) if sp2a else "—"
        sp3p_s = ("%d%%" % round(sp3m/sp3a*100)) if sp3a else "—"
        sftp_s = ("%d%%" % round(sftm/sfta*100)) if sfta else "—"
        sortg_s= ("%.1f" % (spts*100/sposs)) if sposs else "—"
        sdrtg_s= ("%.1f" % (sopp_pts*100/sopp_poss)) if sopp_poss else "—"
        td = 'style="text-align:center;padding:7px 10px"'
        rows += (f'<tr class="msrow">'
            f'<td style="text-align:left;padding:7px 10px"><span class="mqdot mqs">&Sigma;</span></td>'
            f'<td {td}>{spts}</td>'
            f'<td {td}>{sp2m}/{sp2a}</td><td {td}>{sp2p_s}</td>'
            f'<td {td}>{sp3m}/{sp3a}</td><td {td}>{sp3p_s}</td>'
            f'<td {td}>{sftm}/{sfta}</td><td {td}>{sftp_s}</td>'
            f'<td {td}>{soreb}</td><td {td}>{sdreb}</td>'
            f'<td {td}>{sast}</td><td {td}>{sprz}</td><td {td}>{sto}</td>'
            f'<td {td}>{sposs}</td><td {td}>{sfd}</td>'
            f'<td {td} class="{sefg_c}">{sefg_s}</td><td {td}>{sortg_s}</td><td {td}>{sdrtg_s}</td>'
            f'</tr>')
        TH  = 'background:#1a2b4a;color:#fff;font-size:10px;font-weight:500;padding:8px 10px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.15)'
        THL = TH + ';text-align:left'
        GRP = 'background:#1a2b4a;color:rgba(255,255,255,.55);font-size:9px;font-weight:400;letter-spacing:.4px;padding:5px 10px 2px;text-align:center;white-space:nowrap;border-bottom:0.5px solid rgba(255,255,255,.12)'
        SUB = 'background:#1a2b4a;color:rgba(255,255,255,.85);font-size:10px;font-weight:500;padding:2px 10px 7px;text-align:center;white-space:nowrap;border-bottom:0.5px solid rgba(255,255,255,.2)'
        return ('<div class="mtw"><table style="width:100%;border-collapse:collapse;font-size:12px;table-layout:fixed">'
            '<colgroup><col style="width:32px"><col style="width:38px"><col style="width:54px"><col style="width:40px"><col style="width:54px"><col style="width:40px"><col style="width:54px"><col style="width:40px"><col style="width:38px"><col style="width:38px"><col style="width:38px"><col style="width:38px"><col style="width:36px"><col style="width:44px"><col style="width:36px"><col style="width:44px"><col style="width:50px"><col style="width:50px"></colgroup>'
            '<thead><tr>'
            f'<th style="{THL}" rowspan="3">Q</th>'
            f'<th style="{TH}" rowspan="3">PKT</th>'
            f'<th style="{GRP}" colspan="2">2PT</th>'
            f'<th style="{GRP}" colspan="2">3PT</th>'
            f'<th style="{GRP}" colspan="2">FT</th>'
            f'<th style="{GRP}" colspan="2">ZB</th>'
            f'<th style="{TH}" rowspan="3">AST</th>'
            f'<th style="{TH}" rowspan="3">PRZ</th>'
            f'<th style="{TH}" rowspan="3">TO</th>'
            f'<th style="{TH}" rowspan="3">POSS</th>'
            f'<th style="{TH}" rowspan="3">FD</th>'
            f'<th style="{TH}" rowspan="3">eFG%</th>'
            f'<th style="{TH}" rowspan="3">ORtg</th>'
            f'<th style="{TH}" rowspan="3">DRtg</th>'
            '</tr><tr>'
            f'<th style="{SUB}">M/A</th><th style="{SUB}">%</th>'
            f'<th style="{SUB}">M/A</th><th style="{SUB}">%</th>'
            f'<th style="{SUB}">M/A</th><th style="{SUB}">%</th>'
            f'<th style="{SUB}">A</th><th style="{SUB}">O</th>'
            '</tr></thead><tbody>' + rows + '</tbody></table></div>')

    def p_table(druzyna):
        nr_to_name_map = {}
        try:
            cur_p = db.cursor()
            cur_p.execute("""
                SELECT ps.nr,
                       COALESCE(p.imie, r.imie, '')        AS imie,
                       COALESCE(p.nazwisko, r.nazwisko, '') AS nazwisko,
                       CASE WHEN p.id IS NOT NULL THEN 1 WHEN r.id IS NOT NULL THEN 2 ELSE 3 END AS prio
                FROM player_stats ps
                LEFT JOIN players p ON ps.player_id = p.id
                LEFT JOIN roster  r ON ps.roster_id  = r.id
                WHERE ps.match_id = %s AND ps.druzyna = %s
                  AND (p.id IS NOT NULL OR r.id IS NOT NULL)
                ORDER BY ps.nr, prio
            """, (match_id, druzyna))
            for row in cur_p.fetchall():
                nr_int = int(row["nr"])
                if nr_int not in nr_to_name_map:
                    nr_to_name_map[nr_int] = (row["imie"] or "", row["nazwisko"] or "")
            cur_p.close()
        except Exception: pass

        team_poss = sum(r.get("poss",0) or 0 for r in all_stats if r["druzyna"]==druzyna)
        ps_by_nr = {}
        for ps in all_players:
            if ps.get("druzyna") != druzyna: continue
            nr = int(ps.get("nr") or 0)
            if nr not in ps_by_nr: ps_by_nr[nr] = []
            ps_by_nr[nr].append(ps)

        # Sortuj domyślnie po PKT malejąco
        sorted_nrs = sorted(ps_by_nr.keys(),
            key=lambda n: sum(int(ps.get("pts",0) or 0) for ps in ps_by_nr[n]),
            reverse=True)

        rows = ""
        for idx_p, nr in enumerate(sorted_nrs):
            agg = {}
            for ps in ps_by_nr[nr]:
                for k in ["pts","p2m","p2a","p3m","p3a","ftm","fta","ast","oreb","dreb","br","fd","stl","blk","finishes","time_sum","time_cnt"]:
                    agg[k] = agg.get(k,0) + (int(ps.get(k,0) or 0) if k != "time_sum" else float(ps.get(k,0) or 0))
            p2m=agg.get("p2m",0); p2a=agg.get("p2a",0)
            p3m=agg.get("p3m",0); p3a=agg.get("p3a",0)
            ftm=agg.get("ftm",0); fta=agg.get("fta",0)
            pts=agg.get("pts",0); ast=agg.get("ast",0)
            oreb=agg.get("oreb",0); dreb=agg.get("dreb",0)
            br=agg.get("br",0); fd=agg.get("fd",0)
            stl=agg.get("stl",0); blk=agg.get("blk",0)
            fin=agg.get("finishes",0)
            fga=p2a+p3a
            efg_v = round((p2m+1.5*p3m)/fga*100) if fga else None
            efg_s = ("%d%%" % efg_v) if efg_v is not None else "—"
            efg_cls = "mgood" if efg_v and efg_v>=50 else ("mbad" if efg_v and efg_v<35 else "")
            ts_v = round(pts/(2*(fga+0.44*fta))*100) if (fga+fta) else None
            ts_s = ("%d%%" % ts_v) if ts_v is not None else "—"
            usg_v = round((fga+0.44*fta+br)/team_poss*100) if team_poss else None
            usg_s = ("%d%%" % usg_v) if usg_v is not None else "—"
            p2p_s = ("%d%%" % round(p2m/p2a*100)) if p2a else "—"
            p3p_s = ("%d%%" % round(p3m/p3a*100)) if p3a else "—"
            ftp_s = ("%d%%" % round(ftm/fta*100)) if fta else "—"
            # MIN — priorytet: calc_play_time (Excel×1.22), fallback: time_sum×1.22
            _pt_secs = play_time_secs.get(nr, 0)
            if _pt_secs > 0:
                min_s = f"{int(_pt_secs)//60}:{int(_pt_secs)%60:02d}"
                min_dv = str(int(_pt_secs))
            else:
                ts_raw = float(agg.get("time_sum",0))
                if ts_raw:
                    _cm = ts_raw/60*1.22
                    min_s = f"{int(_cm)}:{int((_cm%1)*60):02d}"
                    min_dv = str(int(ts_raw))
                else:
                    min_s = "—"; min_dv = "0"
            imie, nazwisko = nr_to_name_map.get(nr, ("", ""))
            if nazwisko or imie:
                name_s = f"#{nr} {nazwisko} {imie[0]}." if imie else f"#{nr} {nazwisko}"
            else:
                name_s = nr_name_map.get(str(nr), f"#{nr}")
            bg = '#f8f9ff' if idx_p % 2 == 0 else '#fff'
            td = f'style="text-align:center;padding:7px 4px;border-bottom:0.5px solid #eceef2;background:{bg}"'
            tdl= f'style="text-align:left;padding:7px 8px;border-bottom:0.5px solid #eceef2;background:{bg};font-size:11px;font-weight:500;white-space:nowrap"'
            rows += (f'<tr>'
                f'<td {tdl}>{name_s}</td>'
                f'<td {td} data-v="{min_dv}">{min_s}</td>'
                f'<td {td} data-v="{pts}" style="text-align:center;padding:7px 4px;border-bottom:0.5px solid #eceef2;background:{bg};font-weight:600">{pts}</td>'
                f'<td {td}>{p2m}/{p2a}</td><td {td}>{p2p_s}</td>'
                f'<td {td}>{p3m}/{p3a}</td><td {td}>{p3p_s}</td>'
                f'<td {td}>{ftm}/{fta}</td><td {td}>{ftp_s}</td>'
                f'<td {td}>{oreb}</td><td {td}>{dreb}</td>'
                f'<td {td} data-v="{ast}">{ast}</td>'
                f'<td {td} data-v="{br}">{br}</td>'
                f'<td {td} data-v="{stl}">{stl}</td>'
                f'<td {td} data-v="{blk}">{blk}</td>'
                f'<td {td} data-v="{fd}">{fd}</td>'
                f'<td {td} class="{efg_cls}" data-v="{efg_v if efg_v is not None else -1}">{efg_s}</td>'
                f'<td {td} data-v="{ts_v if ts_v is not None else -1}">{ts_s}</td>'
                f'<td {td} data-v="{usg_v if usg_v is not None else -1}">{usg_s}</td>'
                f'<td {td} data-v="{fin}">{fin}</td>'
                f'</tr>')
        if not rows:
            rows = '<tr><td colspan="20" style="text-align:center;color:#9ca3af;padding:16px">Brak danych</td></tr>'
        TH  = 'background:#1a2b4a;color:#fff;font-size:10px;font-weight:500;padding:8px 6px;text-align:center;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.15)'
        THL = TH + ';text-align:left'
        THS = TH + ';cursor:pointer'  # sortable header
        GRP = 'background:#1a2b4a;color:rgba(255,255,255,.55);font-size:9px;padding:5px 6px 2px;text-align:center;border-bottom:0.5px solid rgba(255,255,255,.12)'
        SUB = 'background:#1a2b4a;color:rgba(255,255,255,.85);font-size:9px;padding:2px 6px 7px;text-align:center;border-bottom:0.5px solid rgba(255,255,255,.2)'
        tid = f'ptable-{druzyna}'
        # helper: sortable th — col is the 0-based td index in each row
        def _sth(label, col, default_active=False):
            arr = ' ▼' if default_active else ' ▼'
            op  = '1'  if default_active else '0.3'
            return (f'<th style="{THS}" rowspan="2" data-sc="{col}"'
                    f' onclick="sortPTable(\'{tid}\',{col})">'
                    f'{label}<span class="si" style="font-size:8px;opacity:{op}">{arr}</span></th>')
        return (f'<div class="mtw"><table id="{tid}" style="width:100%;border-collapse:collapse;font-size:12px">'
            '<thead><tr>'
            f'<th style="{THL}" rowspan="2">Zawodnik</th>'
            f'<th style="{TH}"  rowspan="2">MIN</th>'
            + _sth('PTS', 2, True) +
            f'<th style="{GRP}" colspan="2">2PT</th>'
            f'<th style="{GRP}" colspan="2">3PT</th>'
            f'<th style="{GRP}" colspan="2">FT</th>'
            f'<th style="{GRP}" colspan="2">ZB</th>'
            + _sth('AST', 11) + _sth('TO', 12) + _sth('STL', 13)
            + _sth('BLK', 14) + _sth('FD', 15) + _sth('eFG%', 16)
            + _sth('TS%', 17) + _sth('USG%', 18) + _sth('FIN', 19) +
            '</tr><tr>'
            f'<th style="{SUB}">M/A</th><th style="{SUB}">%</th>'
            f'<th style="{SUB}">M/A</th><th style="{SUB}">%</th>'
            f'<th style="{SUB}">M/A</th><th style="{SUB}">%</th>'
            f'<th style="{SUB}">A</th><th style="{SUB}">O</th>'
            '</tr></thead><tbody>' + rows + '</tbody></table></div>')

    # Piątki
    def lineup_table():
        if not all_lineups and not all_lineups_def:
            return '<p style="color:#9ca3af;padding:16px;font-size:.82rem">Brak danych piątek.</p>'

        def_map = {lu["lineup"]: lu for lu in all_lineups_def}
        off_rtg = {lu["lineup"]: lu["pts"]*100/lu["poss"] for lu in all_lineups if int(lu.get("poss",0) or 0)>0}
        def_rtg = {lu["lineup"]: lu["pts"]*100/lu["poss"] for lu in all_lineups_def if int(lu.get("poss",0) or 0)>0}

        rows = ""
        for i, lu in enumerate(sorted(all_lineups, key=lambda x: int(x.get("poss",0) or 0), reverse=True)):
            p2m=int(lu.get("p2m",0) or 0); p2a=int(lu.get("p2a",0) or 0)
            p3m=int(lu.get("p3m",0) or 0); p3a=int(lu.get("p3a",0) or 0)
            ftm=int(lu.get("ftm",0) or 0); fta=int(lu.get("fta",0) or 0)
            pts=int(lu.get("pts",0) or 0); poss=int(lu.get("poss",0) or 0)
            br =int(lu.get("br",0)  or 0)
            oreb=int(lu.get("oreb",0) or 0); dreb=int(lu.get("dreb",0) or 0)
            ast =int(lu.get("ast",0)  or 0)
            stl =int(lu.get("stl",0)  or 0)
            blk =int(lu.get("blk",0)  or 0)
            fd_ =int(lu.get("fd",0)   or 0)
            fga = p2a + p3a
            efg_v  = round((p2m+1.5*p3m)/fga*100) if fga else None
            ppp_v  = pts/poss if poss else None
            efg_s  = f"{efg_v}%" if efg_v is not None else "—"
            ppp_s  = f"{ppp_v:.2f}" if ppp_v is not None else "—"
            p2pct  = f"{p2m/p2a:.0%}" if p2a else "—"
            p3pct  = f"{p3m/p3a:.0%}" if p3a else "—"
            ftpct  = f"{ftm/fta:.0%}" if fta else "—"
            ppp_c  = "#0F6E56" if ppp_v and ppp_v>=0.9 else ("#A32D2D" if ppp_v and ppp_v<0.7 else "inherit")
            efg_c  = "#0F6E56" if efg_v and efg_v>=50 else ("#A32D2D" if efg_v and efg_v<35 else "inherit")

            dlu = def_map.get(lu["lineup"], {})
            dp2m=int(dlu.get("p2m",0) or 0); dp2a=int(dlu.get("p2a",0) or 0)
            dp3m=int(dlu.get("p3m",0) or 0); dp3a=int(dlu.get("p3a",0) or 0)
            dp_fga = dp2a+dp3a
            defg_v = round((dp2m+1.5*dp3m)/dp_fga*100) if dp_fga else None
            dposs  = int(dlu.get("poss",0) or 0)
            dpts   = int(dlu.get("pts",0) or 0)
            dppp_v = dpts/dposs if dposs else None
            defg_s = f"{defg_v}%" if defg_v is not None else "—"
            dppp_s = f"{dppp_v:.2f}" if dppp_v is not None else "—"
            defg_c = "#0F6E56" if defg_v and defg_v<35 else ("#A32D2D" if defg_v and defg_v>=50 else "inherit")
            dppp_c = "#0F6E56" if dppp_v and dppp_v<0.7 else ("#A32D2D" if dppp_v and dppp_v>=0.9 else "inherit")

            k = lu["lineup"]
            ortg_v = off_rtg.get(k); drtg_v = def_rtg.get(k)
            net_v  = round(ortg_v - drtg_v, 1) if (ortg_v is not None and drtg_v is not None) else None
            ortg_s = f"{ortg_v:.1f}" if ortg_v is not None else "—"
            drtg_s = f"{drtg_v:.1f}" if drtg_v is not None else "—"
            net_s  = f"{net_v:+.1f}" if net_v is not None else "—"
            net_c  = "#0F6E56" if net_v and net_v>0 else ("#A32D2D" if net_v and net_v<0 else "#888")
            ortg_c = "#0F6E56" if ortg_v and ortg_v>=90 else ("#A32D2D" if ortg_v and ortg_v<70 else "inherit")
            drtg_c = "#0F6E56" if drtg_v and drtg_v<70 else ("#A32D2D" if drtg_v and drtg_v>=90 else "inherit")

            # MIN szac.
            _lu_nrs = [int(n) for n in k.split("-") if n.strip().isdigit()]
            _lu_secs = [play_time_secs.get(n, 0) for n in _lu_nrs]
            _lu_secs_valid = [s for s in _lu_secs if s > 0]
            if _lu_secs_valid:
                _avg_secs = sum(_lu_secs_valid) / len(_lu_nrs)
                _min_str  = f"{int(_avg_secs)//60}:{int(_avg_secs)%60:02d}"
                _min_sv   = f"{_avg_secs:.1f}"
            else:
                _min_str = "—"; _min_sv = "-1"

            bg = "#f8f9ff" if i%2==0 else "#fff"
            br_c   = "#A32D2D" if br>=4 else "inherit"
            efg_sv = str(efg_v) if efg_v is not None else "-1"
            ppp_sv = f"{ppp_v*100:.1f}" if ppp_v is not None else "-1"
            defg_sv= str(defg_v) if defg_v is not None else "-1"
            dppp_sv= f"{dppp_v*100:.1f}" if dppp_v is not None else "-1"
            net_sv = f"{net_v:.1f}" if net_v is not None else "-999"
            skladniki = " · ".join(nr_name_map.get(n, f"#{n}") for n in k.split("-"))

            rows += f"""<tr style="background:{bg}">
                <td style="font-size:.78rem;text-align:left;padding:6px 8px;border-bottom:0.5px solid #eceef2;background:{bg}">{skladniki}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{poss}">{poss}</td>
                <td class="text-center" style="color:#633806;font-weight:500;padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{_min_sv}">{_min_str}</td>
                <td class="text-center fw-bold" style="color:#1a2b4a;padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{pts}">{pts}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{p2m*100+p2a}">{p2m}/{p2a}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{int(p2m/p2a*100) if p2a else -1}">{p2pct}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{p3m*100+p3a}">{p3m}/{p3a}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{int(p3m/p3a*100) if p3a else -1}">{p3pct}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{ftm*100+fta}">{ftm}/{fta}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{int(ftm/fta*100) if fta else -1}">{ftpct}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{oreb}">{oreb}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{dreb}">{dreb}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{ast}">{ast}</td>
                <td class="text-center" style="color:{br_c};padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{br}">{br}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{stl}">{stl}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{blk}">{blk}</td>
                <td class="text-center" style="padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{fd_}">{fd_}</td>
                <td class="text-center" style="color:{efg_c};padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{efg_sv}">{efg_s}</td>
                <td class="text-center fw-bold" style="color:{ppp_c};padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{ppp_sv}">{ppp_s}</td>
                <td class="text-center" style="color:{defg_c};padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{defg_sv}">{defg_s}</td>
                <td class="text-center fw-bold" style="color:{dppp_c};padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{dppp_sv}">{dppp_s}</td>
                <td class="text-center" style="color:{ortg_c};padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{ortg_v or -999}">{ortg_s}</td>
                <td class="text-center" style="color:{drtg_c};padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{drtg_v or 999}">{drtg_s}</td>
                <td class="text-center fw-bold" style="color:{net_c};padding:6px 4px;border-bottom:0.5px solid #eceef2;background:{bg}" data-v="{net_sv}">{net_s}</td>
            </tr>"""

        no_data = '<tr><td colspan="24" style="text-align:center;color:#9ca3af;padding:16px;font-size:.82rem">Brak danych</td></tr>'

        th  = 'background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 3px;text-align:center;white-space:nowrap;cursor:pointer;user-select:none;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.15)'
        thl = 'background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 8px;text-align:left;white-space:nowrap;vertical-align:middle;border-bottom:0.5px solid rgba(255,255,255,.15)'
        thg = 'background:#1a2b4a;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center'
        ths = 'background:#1a2b4a;color:rgba(255,255,255,.8);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.2);text-align:center;cursor:pointer;user-select:none'
        thz = 'background:#152236;color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.1);text-align:center'
        thzs= 'background:#152236;color:rgba(255,255,255,.75);font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center;cursor:pointer;user-select:none'
        thn = 'background:#412402;color:#FAC775;font-size:8px;letter-spacing:.3px;padding:4px 3px 2px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center'
        thns= 'background:#412402;color:#FAC775;font-size:9px;font-weight:500;padding:2px 3px 5px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center;cursor:pointer;user-select:none'

        cg = ('<colgroup>'
            '<col style="width:190px">'
            '<col style="width:36px"><col style="width:46px"><col style="width:34px">'
            '<col style="width:40px"><col style="width:30px">'
            '<col style="width:40px"><col style="width:30px">'
            '<col style="width:40px"><col style="width:30px">'
            '<col style="width:26px"><col style="width:26px">'
            '<col style="width:28px"><col style="width:28px"><col style="width:28px"><col style="width:28px"><col style="width:28px">'
            '<col style="width:34px"><col style="width:34px">'
            '<col style="width:34px"><col style="width:34px">'
            '<col style="width:38px"><col style="width:38px"><col style="width:42px">'
            '</colgroup>')

        hdr = (f'<thead><tr>'
            f'<th style="{thl}" rowspan="3">Skład</th>'
            f'<th style="{th}" rowspan="3" onclick="pmLuSort(this,1)">POSS ↕</th>'
            f'<th style="{th}" rowspan="3" onclick="pmLuSort(this,2)" title="Szac. czas gry składu">MIN<br>(szac.)</th>'
            f'<th style="{th}" rowspan="3" onclick="pmLuSort(this,3)">PKT ↕</th>'
            f'<th style="{thg}" colspan="2">2PT</th>'
            f'<th style="{thg}" colspan="2">3PT</th>'
            f'<th style="{thg}" colspan="2">FT</th>'
            f'<th style="{thz}" colspan="2">ZB</th>'
            f'<th style="{th}" rowspan="3" onclick="pmLuSort(this,12)">AST ↕</th>'
            f'<th style="{th}" rowspan="3" onclick="pmLuSort(this,13)">TO ↕</th>'
            f'<th style="{th}" rowspan="3" onclick="pmLuSort(this,14)">STL ↕</th>'
            f'<th style="{th}" rowspan="3" onclick="pmLuSort(this,15)">BLK ↕</th>'
            f'<th style="{th}" rowspan="3" onclick="pmLuSort(this,16)">FD ↕</th>'
            f'<th style="{thg}" colspan="2">OFF</th>'
            f'<th style="{thg}" colspan="2">DEF</th>'
            f'<th style="{thn}" colspan="3">NET RTG</th>'
            f'</tr><tr>'
            f'<th style="{ths}" onclick="pmLuSort(this,4)">M/A ↕</th><th style="{ths}" onclick="pmLuSort(this,5)">% ↕</th>'
            f'<th style="{ths}" onclick="pmLuSort(this,6)">M/A ↕</th><th style="{ths}" onclick="pmLuSort(this,7)">% ↕</th>'
            f'<th style="{ths}" onclick="pmLuSort(this,8)">M/A ↕</th><th style="{ths}" onclick="pmLuSort(this,9)">% ↕</th>'
            f'<th style="{thzs}" onclick="pmLuSort(this,10)">A ↕</th><th style="{thzs}" onclick="pmLuSort(this,11)">O ↕</th>'
            f'<th style="{ths}" onclick="pmLuSort(this,17)">eFG% ↕</th><th style="{ths}" onclick="pmLuSort(this,18)">PPP ↕</th>'
            f'<th style="{ths}" onclick="pmLuSort(this,19)">eFG% ↕</th><th style="{ths}" onclick="pmLuSort(this,20)">PPP ↕</th>'
            f'<th style="{thns}" onclick="pmLuSort(this,21)">ORtg ↕</th>'
            f'<th style="{thns}" onclick="pmLuSort(this,22)">DRtg ↕</th>'
            f'<th style="{thns}" onclick="pmLuSort(this,23)">Net ↕</th>'
            f'</tr></thead>')

        lu_js = """<script>
(function(){
  var _d={};
  window.pmLuSort=function(th,col){
    var tbl=document.getElementById('pm-lu-off'); if(!tbl) return;
    var tb=tbl.querySelector('tbody');
    var rows=Array.from(tb.querySelectorAll('tr'));
    var k='lu-'+col; _d[k]=!_d[k]; var asc=_d[k];
    rows.sort(function(a,b){
      var av=parseFloat(a.children[col]&&a.children[col].dataset.v);
      var bv=parseFloat(b.children[col]&&b.children[col].dataset.v);
      if(isNaN(av))av=asc?Infinity:-Infinity;
      if(isNaN(bv))bv=asc?Infinity:-Infinity;
      return asc?av-bv:bv-av;
    });
    rows.forEach(function(r,i){r.style.background=i%2===0?'#f8f9ff':'#fff';tb.appendChild(r);});
    tbl.querySelectorAll('th').forEach(function(h){h.innerHTML=h.innerHTML.replace(/ [▲▼]/,' ↕');});
    th.innerHTML=th.innerHTML.replace(' ↕',asc?' ▲':' ▼');
  };
})();
</script>"""

        sub = 'font-size:8px;color:#888;margin-bottom:6px'
        return (
            f'<div style="{sub}">PPP OFF: <span style="color:#0F6E56">≥0.90 dobry</span> / <span style="color:#A32D2D">&lt;0.70 słaby</span> &nbsp;·&nbsp; '
            f'PPP DEF: <span style="color:#0F6E56">&lt;0.70 dobry</span> / <span style="color:#A32D2D">≥0.90 słaby</span> &nbsp;·&nbsp; sortowanie: POSS malejąco</div>'
            f'<div class="mtw"><table id="pm-lu-off" style="width:100%;border-collapse:collapse;font-size:12px;min-width:900px">'
            f'{cg}{hdr}'
            f'<tbody>{rows or no_data}</tbody></table></div>'
            + lu_js
        )

    def momentum_table():
        if not all_stats:
            return '<p style="color:#9ca3af;padding:16px;font-size:.82rem">Brak danych.</p>'
        rows_html = ""
        q_results = []
        for qn in [1, 2, 3, 4]:
            qs_gtk = next((r for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==qn), {})
            qs_opp = next((r for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==qn), {})
            q_gtk = qs_gtk.get("pts", 0) or 0
            q_opp = qs_opp.get("pts", 0) or 0
            q_winner = "gtk" if q_gtk > q_opp else ("opp" if q_opp > q_gtk else "tie")
            q_results.append(q_winner)

            def _efg(qs):
                p2a=qs.get("p2a",0) or 0; p3a=qs.get("p3a",0) or 0
                p2m=qs.get("p2m",0) or 0; p3m=qs.get("p3m",0) or 0
                return str(round((p2m+1.5*p3m)/(p2a+p3a)*100)) + "%" if (p2a+p3a) else "—"

            qd_cls = {1:"mq1",2:"mq2",3:"mq3",4:"mq4"}[qn]
            if q_winner == "gtk":
                res_html = f'<span class="pmwin">W +{q_gtk-q_opp}</span>'
            elif q_winner == "opp":
                res_html = f'<span class="pmloss">L -{q_opp-q_gtk}</span>'
            else:
                res_html = '<span style="background:#eee;color:#666;border-radius:6px;padding:2px 8px;font-size:11px">=</span>'

            efg_g = _efg(qs_gtk); efg_o = _efg(qs_opp)
            to_g = qs_gtk.get("br",0) or 0; to_o = qs_opp.get("br",0) or 0
            ps_g = qs_gtk.get("poss",0) or 0; ps_o = qs_opp.get("poss",0) or 0

            def _efg_int(s):
                try: return int(s.rstrip("%"))
                except: return -1
            g_efg = "pmgood" if efg_g!="—" and efg_o!="—" and _efg_int(efg_g)>_efg_int(efg_o) else ("pmbad" if efg_g!="—" and efg_o!="—" and _efg_int(efg_g)<_efg_int(efg_o) else "")
            o_efg = "pmgood" if efg_g!="—" and efg_o!="—" and _efg_int(efg_o)>_efg_int(efg_g) else ("pmbad" if efg_g!="—" and efg_o!="—" and _efg_int(efg_o)<_efg_int(efg_g) else "")
            g_to  = "pmgood" if to_g < to_o else ("pmbad" if to_g > to_o else "")
            o_to  = "pmgood" if to_o < to_g else ("pmbad" if to_o > to_g else "")
            g_pkt = "pmgood" if q_gtk > q_opp else ("pmbad" if q_gtk < q_opp else "")
            o_pkt = "pmgood" if q_opp > q_gtk else ("pmbad" if q_opp < q_gtk else "")

            rows_html += (
                '<tr>'
                f'<td><span class="mqdot {qd_cls}">{qn}Q</span></td>'
                f'<td style="text-align:center">{res_html}</td>'
                f'<td style="text-align:center" class="{g_pkt}">{q_gtk}</td>'
                f'<td style="text-align:center" class="{g_efg}">{efg_g}</td>'
                f'<td style="text-align:center" class="{g_to}">{to_g}</td>'
                f'<td style="text-align:center">{ps_g}</td>'
                f'<td style="text-align:center;border-left:2px solid #e0e0e0" class="{o_pkt}">{q_opp}</td>'
                f'<td style="text-align:center" class="{o_efg}">{efg_o}</td>'
                f'<td style="text-align:center" class="{o_to}">{to_o}</td>'
                f'<td style="text-align:center">{ps_o}</td>'
                '</tr>'
            )

        streak_html = ""
        for r in q_results:
            col = "#0F6E56" if r=="gtk" else ("#A32D2D" if r=="opp" else "#888")
            lbl = "W" if r=="gtk" else ("L" if r=="opp" else "=")
            streak_html += (f'<span style="display:inline-flex;align-items:center;justify-content:center;'
                            f'width:28px;height:28px;border-radius:50%;background:{col};color:#fff;'
                            f'font-size:11px;font-weight:500;margin-right:4px">{lbl}</span>')
        best = cur_s = 0
        for r in q_results:
            cur_s = cur_s+1 if r=="gtk" else 0
            best = max(best, cur_s)
        best_html = (f'<span style="font-size:12px;color:#666">Najlepsza seria: '
                     f'<span style="color:#0F6E56;font-weight:500">{best}'
                     f'{"kwarty" if best>1 else "kwarta"}</span></span>') if best >= 1 else ""

        gtk_s = gtk_name.upper()[:16]; opp_s = name_opp.upper()[:16]
        th_qw  = "padding:8px 10px;font-size:10px;font-weight:500;text-transform:uppercase;letter-spacing:.5px;color:#888;border-bottom:0.5px solid #e0e0e0;background:#fff;white-space:nowrap"
        th_grp = "padding:8px 10px 4px;font-size:10px;font-weight:500;letter-spacing:.4px;text-align:center;background:#fff;border-bottom:none;white-space:nowrap"
        th_sub = "padding:4px 10px 8px;font-size:10px;font-weight:500;text-align:center;border-bottom:0.5px solid #e0e0e0;background:#fff;color:#888;white-space:nowrap"
        return (
            f'<div style="padding:10px 14px;display:flex;align-items:center;gap:8px;flex-wrap:wrap;border-bottom:0.5px solid #e0e0e0">'
            + streak_html + best_html +
            '</div>'
            '<div style="overflow-x:auto">'
            '<table style="width:100%;border-collapse:collapse;font-size:12px;min-width:500px">'
            '<thead><tr>'
            f'<th rowspan="2" style="{th_qw};text-align:left;vertical-align:middle">Q</th>'
            f'<th rowspan="2" style="{th_qw};text-align:center;vertical-align:middle">Wynik</th>'
            f'<th colspan="4" style="{th_grp};color:#185FA5">{gtk_s}</th>'
            f'<th colspan="4" style="{th_grp};color:#A32D2D;border-left:2px solid #e0e0e0">{opp_s}</th>'
            '</tr><tr>'
            f'<th style="{th_sub}">PKT</th><th style="{th_sub}">eFG%</th>'
            f'<th style="{th_sub}">TO</th><th style="{th_sub}">POSS</th>'
            f'<th style="{th_sub};border-left:2px solid #e0e0e0">PKT</th>'
            f'<th style="{th_sub}">eFG%</th><th style="{th_sub}">TO</th><th style="{th_sub}">POSS</th>'
            '</tr></thead>'
            f'<tbody>{rows_html}</tbody>'
            '</table></div>'
        )

    # Quarters per team
    pts_q_gtk = [next((r["pts"] for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==q),0) for q in [1,2,3,4]]
    pts_q_opp = [next((r["pts"] for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==q),0) for q in [1,2,3,4]]

    # ── Timing akcji ────────────────────────────────────────────────────────────
    def tim_table(druzyna):
        qcolors = {1:"mq1",2:"mq2",3:"mq3",4:"mq4"}
        def get_td(q, b):
            return next((r for r in all_timing if r["druzyna"]==druzyna and r["bucket"]==b and (r.get("kwarta") or 0)==q), {})
        def calc_efg(m2, a2, m3, a3, ftm_v):
            att = a2+a3
            if att==0 and ftm_v>0: return 100
            if att==0: return None
            base = round((m2+1.5*m3)/att*100)
            if ftm_v>0: return max(base, round((m2+1.5*m3+1)/(att+1)*100))
            return base
        def calc_udane(m2, m3, ftm_v): return m2+m3+(1 if ftm_v>0 else 0)
        def calc_nieudane(m2, a2, m3, a3, br_v): return (a2-m2)+(a3-m3)+br_v
        def pill(efg):
            if efg is None: return '<span style="color:#aaa">&#8212;</span>'
            c = "pill-good" if efg>=50 else ("pill-bad" if efg<35 else "pill-neu")
            return '<span class="epill %s">%d%%</span>' % (c, efg)
        sum_rows = []
        for b in BUCKETS:
            m2=a2=m3=a3=br_v=ftm_v=poss_ft_v=0
            for q in [1,2,3,4]:
                td=get_td(q,b); m2+=td.get("made2",0); a2+=td.get("att2",0)
                m3+=td.get("made3",0); a3+=td.get("att3",0)
                br_v+=td.get("br",0); ftm_v+=td.get("ftm",0)
                poss_ft_v+=td.get("poss_ft",0)
            td0=get_td(0,b)
            if td0 and not any(get_td(q,b) for q in [1,2,3,4]):
                m2=td0.get("made2",0); a2=td0.get("att2",0)
                m3=td0.get("made3",0); a3=td0.get("att3",0)
                br_v=td0.get("br",0); ftm_v=td0.get("ftm",0)
                poss_ft_v=td0.get("poss_ft",0)
            eff_poss_ft = poss_ft_v if poss_ft_v > 0 else (1 if ftm_v > 0 else 0)
            udane=calc_udane(m2,m3,eff_poss_ft); nieudane=calc_nieudane(m2,a2,m3,a3,br_v)
            total=udane+nieudane; efg=calc_efg(m2,a2,m3,a3,ftm_v)
            sum_rows.append((b,m2,a2,m3,a3,eff_poss_ft,br_v,udane,nieudane,total,efg))
        th  = 'style="padding:9px 14px;font-size:10px;font-weight:500;text-align:right;background:#1a2b4a;color:#fff;white-space:nowrap;text-transform:uppercase;letter-spacing:.4px"'
        tdr = 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:right;font-size:12px"'
        tdu = 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:right;font-size:12px;font-weight:500;color:#085041"'
        tdn = 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:right;font-size:12px;font-weight:500;color:#A32D2D"'
        tdc = 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:center;font-size:12px"'
        tdc_b='style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:center;font-size:12px;font-weight:700"'
        tdlc= 'style="padding:8px 14px;border-bottom:0.5px solid #e0e0e0;text-align:center;font-size:12px;font-weight:500"'
        s_rows=""
        for b,m2,a2,m3,a3,ftm_v,br_v,udane,nieudane,total,efg_v in sum_rows:
            chyb2=a2-m2; chyb3=a3-m3
            p2 = str(m2) if m2>0 else '<span style="color:#aaa">&#8212;</span>'
            p3 = str(m3) if m3>0 else '<span style="color:#aaa">&#8212;</span>'
            ft_poss_s = str(poss_ft_v) if poss_ft_v>0 else ('<span style="color:#aaa">1</span>' if ftm_v>0 else '<span style="color:#aaa">&#8212;</span>')
            c2_s = str(chyb2) if chyb2>0 else '<span style="color:#aaa">&#8212;</span>'
            c3_s = str(chyb3) if chyb3>0 else '<span style="color:#aaa">&#8212;</span>'
            br_s = str(br_v) if br_v>0 else '<span style="color:#aaa">&#8212;</span>'
            tot_s = ('<b>%d</b>' % total) if total else '<span style="color:#aaa">&#8212;</span>'
            efg_s = pill(efg_v) if (a2+a3) else '<span style="color:#aaa">&#8212;</span>'
            s_rows += ('<tr><td %s>%s</td><td %s>%s</td><td %s>%s</td><td %s>%s</td>'
                       '<td %s>%s</td><td %s>%s</td><td %s>%s</td>'
                       '<td %s>%s</td><td %s>%s</td></tr>'
                       % (tdlc,b, tdr,p2, tdr,p3, tdu,ft_poss_s,
                          tdn,c2_s, tdn,c3_s, tdn,br_s, tdc_b,tot_s, tdc,efg_s))
        q_to_vals = {q: next((r for r in all_stats if r["druzyna"]==druzyna and r["kwarta"]==q),{}).get("br",0) or 0 for q in [1,2,3,4]}
        q_rows=""
        for q in [1,2,3,4]:
            qcls=qcolors[q]; to_q=q_to_vals[q]
            q_rows += ('<tr style="background:#f0f2f7"><td colspan="9" style="padding:5px 14px;border-bottom:0.5px solid #e0e0e0">'
                       '<span class="mqdot %s">%dQ</span></td></tr>' % (qcls,q))
            has=False
            for b in BUCKETS:
                td=get_td(q,b); m2=td.get("made2",0); a2=td.get("att2",0)
                m3=td.get("made3",0); a3=td.get("att3",0)
                br_v=td.get("br",0); ftm_v=td.get("ftm",0)
                poss_ft_v=td.get("poss_ft",0)
                if a2+a3+br_v+ftm_v==0: continue
                has=True
                eff_poss_ft = poss_ft_v if poss_ft_v > 0 else (1 if ftm_v > 0 else 0)
                udane=calc_udane(m2,m3,eff_poss_ft); nieudane=calc_nieudane(m2,a2,m3,a3,br_v)
                total=udane+nieudane; efg_v=calc_efg(m2,a2,m3,a3,ftm_v)
                chyb2=a2-m2; chyb3=a3-m3
                p2=str(m2) if m2>0 else "&#8212;"; p3=str(m3) if m3>0 else "&#8212;"
                ft_s=str(eff_poss_ft) if eff_poss_ft>0 else "&#8212;"
                c2_s=str(chyb2) if chyb2>0 else "&#8212;"; c3_s=str(chyb3) if chyb3>0 else "&#8212;"
                br_s=str(br_v) if br_v>0 else "&#8212;"
                tot_s=('<b>%d</b>'%total) if total else "&#8212;"
                efg_s=pill(efg_v) if (a2+a3) else "&#8212;"
                q_rows += ('<tr><td style="padding:8px 14px 8px 24px;border-bottom:0.5px solid #e0e0e0;font-size:12px;color:#888;text-align:center">%s</td>'
                           '<td %s>%s</td><td %s>%s</td><td %s>%s</td>'
                           '<td %s>%s</td><td %s>%s</td><td %s>%s</td>'
                           '<td %s>%s</td><td %s>%s</td></tr>'
                           % (b, tdr,p2, tdr,p3, tdu,ft_s, tdn,c2_s, tdn,c3_s, tdn,br_s, tdc_b,tot_s, tdc,efg_s))
            if not has:
                q_rows += '<tr><td colspan="9" style="padding:6px 14px 6px 24px;border-bottom:0.5px solid #e0e0e0;font-size:11px;color:#aaa">brak danych</td></tr>'
        tbl_wrap='overflow-x:auto;border:0.5px solid #e0e0e0;border-radius:10px;overflow:hidden'
        tbl_s='width:100%;border-collapse:collapse;font-size:12px'
        th_c  = 'style="background:#1a2b4a;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:right;white-space:nowrap"'
        th_g1 = 'style="background:#085041;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:center;white-space:nowrap;border-bottom:1px solid #5DCAA5"'
        th_r1 = 'style="background:#A32D2D;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:center;white-space:nowrap;border-bottom:1px solid #F09595"'
        th_g2 = 'style="background:#0F6E56;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:right;white-space:nowrap"'
        th_r2 = 'style="background:#791F1F;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:right;white-space:nowrap"'
        th_mid= 'style="background:#1a2b4a;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;letter-spacing:.4px;text-transform:uppercase;text-align:center;white-space:nowrap;vertical-align:middle"'
        hdr = ('<tr><th rowspan="2" '+th_mid+' style="background:#1a2b4a;color:#fff;padding:7px 10px;font-size:10px;font-weight:500;text-align:center;white-space:nowrap;vertical-align:middle;width:80px">Czas</th>'
               '<th colspan="3" '+th_g1+'>Udane</th>'
               '<th colspan="3" '+th_r1+'>Nieudane</th>'
               '<th rowspan="2" '+th_mid+'>Razem</th>'
               '<th rowspan="2" '+th_mid+'>eFG%</th>'
               '</tr><tr>'
               '<th '+th_g2+'>2PT</th><th '+th_g2+'>3PT</th><th '+th_g2+'>FT&ge;1</th>'
               '<th '+th_r2+'>2PT</th><th '+th_r2+'>3PT</th><th '+th_r2+'>BR</th>'
               '</tr>')
        out  = '<div id="tim-wrap-' + druzyna + '">'
        out += '<div style="display:flex;border-bottom:0.5px solid #e0e0e0;margin-bottom:12px">'
        out += '<button onclick="pmTimSwitch(\'%s\',\'sum\')" id="pm-tim-btn-sum-%s" style="font-size:12px;padding:8px 14px;border:none;background:none;border-bottom:2px solid #1a2b4a;font-weight:500;color:#1a2b4a;cursor:pointer;margin-bottom:-1px">Suma</button>' % (druzyna,druzyna)
        out += '<button onclick="pmTimSwitch(\'%s\',\'q\')" id="pm-tim-btn-q-%s" style="font-size:12px;padding:8px 14px;border:none;background:none;border-bottom:2px solid transparent;color:#888;cursor:pointer;margin-bottom:-1px">Per kwarta</button>' % (druzyna,druzyna)
        out += '</div>'
        out += '<div id="pm-tim-sum-' + druzyna + '"><div style="' + tbl_wrap + '"><table style="' + tbl_s + '">'
        out += '<thead>' + hdr + '</thead><tbody>%s</tbody></table></div></div>' % s_rows
        out += '<div id="pm-tim-q-' + druzyna + '" style="display:none"><div style="' + tbl_wrap + '"><table style="' + tbl_s + '">'
        out += '<thead>' + hdr + '</thead><tbody>%s</tbody></table></div></div></div>' % q_rows
        return out

    # ── Clutch ──────────────────────────────────────────────────────────────────
    def clutch_stats():
        import math as _math
        # Clutch = ostatnie 1/3 posiadań Q4 (i dogrywek), zaokrąglone w górę.
        clutch_qtrs = sorted({r["kwarta"] for r in all_stats
                               if r.get("kwarta", 0) >= 4 and r.get("druzyna") in ("gtk","opp")})
        if not clutch_qtrs:
            return '<p style="color:#9ca3af;padding:16px;font-size:.82rem">Brak danych Q4/OT w tym meczu.</p>'
        keys_g = ["pts_g","poss_g","p2m_g","p2a_g","p3m_g","p3a_g","ftm_g","fta_g","br_g","fd_g","ast_g","oreb_g","dreb_g","stl_g","blk_g","d2m_g","d2a_g"]
        keys_o = [k.replace("_g","_o") for k in keys_g]
        c = {k: 0 for k in keys_g+keys_o}
        clutch_poss_g_total = 0
        clutch_poss_o_total = 0
        for qn in clutch_qtrs:
            qg = next((r for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==qn), {})
            qo = next((r for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==qn), {})
            poss_g = qg.get("poss", 0) or 0
            poss_o = qo.get("poss", 0) or 0
            cp_g = _math.ceil(poss_g / 3) if poss_g else 0
            cp_o = _math.ceil(poss_o / 3) if poss_o else 0
            ratio_g = cp_g / poss_g if poss_g else 0
            ratio_o = cp_o / poss_o if poss_o else 0
            clutch_poss_g_total += cp_g
            clutch_poss_o_total += cp_o
            def _sc(d, k, ratio): return round((d.get(k,0) or 0)*ratio)
            for k in ["pts","p2m","p2a","p3m","p3a","ftm","fta","br","fd","ast","oreb","dreb","stl","blk","d2m","d2a"]:
                c[f"{k}_g"] += _sc(qg, k, ratio_g)
                c[f"{k}_o"] += _sc(qo, k, ratio_o)
            c["poss_g"] += cp_g
            c["poss_o"] += cp_o
        def pct(m,a): return f"{round(m/a*100)}%" if a else "—"
        def ppp(p,po): return f"{p/po:.2f}" if po else "—"
        def topct(br,poss): return f"{round(br/poss*100)}%" if poss else "—"
        efg_g=pct(c["p2m_g"]+int(1.5*c["p3m_g"]),c["p2a_g"]+c["p3a_g"])
        efg_o=pct(c["p2m_o"]+int(1.5*c["p3m_o"]),c["p2a_o"]+c["p3a_o"])
        p2p_g=pct(c["p2m_g"],c["p2a_g"]); p2p_o=pct(c["p2m_o"],c["p2a_o"])
        p3p_g=pct(c["p3m_g"],c["p3a_g"]); p3p_o=pct(c["p3m_o"],c["p3a_o"])
        ftp_g=pct(c["ftm_g"],c["fta_g"]); ftp_o=pct(c["ftm_o"],c["fta_o"])
        ppp_g=ppp(c["pts_g"],c["poss_g"]); ppp_o=ppp(c["pts_o"],c["poss_o"])
        to_g=topct(c["br_g"],c["poss_g"]); to_o=topct(c["br_o"],c["poss_o"])
        ortg_g=round(c["pts_g"]/c["poss_g"]*100) if c["poss_g"] else 0
        ortg_o=round(c["pts_o"]/c["poss_o"]*100) if c["poss_o"] else 0
        net_g=ortg_g-ortg_o; net_o=ortg_o-ortg_g
        net_gs=("+"+str(net_g)) if net_g>0 else str(net_g)
        net_os=("+"+str(net_o)) if net_o>0 else str(net_o)
        win_gtk=c["pts_g"]>c["pts_o"]; win_opp=c["pts_o"]>c["pts_g"]
        score_c_g="#1a6b3c" if win_gtk else ("#8b1a1a" if win_opp else "#856404")
        score_c_o="#1a6b3c" if win_opp else ("#8b1a1a" if win_gtk else "#856404")
        ortg_gs=str(ortg_g) if c["poss_g"] else "—"
        ortg_os=str(ortg_o) if c["poss_o"] else "—"
        net_color="#1a6b3c" if net_g>0 else ("#8b1a1a" if net_g<0 else "#856404")
        wc="GTK lepsza" if win_gtk else ("Rywal lepszy" if win_opp else "Remis")
        wc_bg2="#e8f5e9" if win_gtk else ("#ffebee" if win_opp else "#fff8e1")
        wc_col2="#1a6b3c" if win_gtk else ("#8b1a1a" if win_opp else "#856404")
        qtrs_label="+".join(f"Q{q}" for q in sorted(clutch_qtrs))
        def col(vg, vo, higher=True):
            try:
                fg=float(str(vg).replace('%','').replace('—','0').replace('+',''))
                fo=float(str(vo).replace('%','').replace('—','0').replace('+',''))
                if higher: return ("#1a6b3c" if fg>fo else "#8b1a1a" if fg<fo else "#888"),("#1a6b3c" if fo>fg else "#8b1a1a" if fo<fg else "#888")
                else: return ("#1a6b3c" if fg<fo else "#8b1a1a" if fg>fo else "#888"),("#1a6b3c" if fo<fg else "#8b1a1a" if fo>fg else "#888")
            except: return "#1a2b4a","#1a2b4a"
        def eff_pill2(lbl, vg, vo, higher=True, neutral=False):
            if neutral: cg=co="#1a2b4a"
            else: cg,co=col(vg,vo,higher=higher)
            return (f'<div style="flex:1;background:#f4f6fb;border-radius:10px;padding:10px 8px;text-align:center;min-width:0;border:0.5px solid #e3e8f0">'
                    f'<div style="font-size:9px;color:#999;text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px">{lbl}</div>'
                    f'<div style="display:flex;align-items:center;justify-content:center;gap:6px">'
                    f'<span style="font-size:15px;font-weight:700;color:{cg}">{vg}</span>'
                    f'<span style="font-size:9px;color:#bbb">vs</span>'
                    f'<span style="font-size:15px;font-weight:700;color:{co}">{vo}</span>'
                    f'</div></div>')
        eff_strip=(eff_pill2("eFG%",efg_g,efg_o)+
                   eff_pill2("POSS",c["poss_g"],c["poss_o"],neutral=True)+
                   eff_pill2("TO%",to_g,to_o,higher=False)+
                   eff_pill2("PPP",ppp_g,ppp_o))
        def face_bar(lbl, mg, ag, mo, ao, pts_type=""):
            pg=pct(mg,ag); po=pct(mo,ao)
            max_w=46
            wg=round(float(str(pg).replace('%','') or 0)/100*max_w) if ag else 0
            wo=round(float(str(po).replace('%','') or 0)/100*max_w) if ao else 0
            badge_bg="#2e5090" if pts_type=="3PT" else ("#856404" if pts_type=="FT" else "#1a6b3c")
            return (f'<div style="margin-bottom:14px">'
                    f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:5px">'
                    f'<span style="font-size:10px;font-weight:700;color:#fff;background:{badge_bg};border-radius:4px;padding:1px 7px">{lbl}</span>'
                    f'<span style="font-size:10px;color:#888">{mg}/{ag}</span>'
                    f'<span style="flex:1;height:1px;background:#e8ecf0"></span>'
                    f'<span style="font-size:10px;color:#888">{mo}/{ao}</span></div>'
                    f'<div style="display:flex;align-items:center;gap:4px">'
                    f'<span style="font-size:13px;font-weight:700;color:{score_c_g};width:38px;text-align:right">{pg}</span>'
                    f'<div style="flex:1;display:flex;height:20px;border-radius:4px;overflow:hidden;background:#eef0f5">'
                    f'<div style="flex:1;display:flex;justify-content:flex-end;align-items:center">'
                    f'<div style="width:{wg}%;height:100%;background:#1a6b3c;border-radius:3px 0 0 3px"></div></div>'
                    f'<div style="width:2px;background:#fff;flex-shrink:0"></div>'
                    f'<div style="flex:1;display:flex;align-items:center">'
                    f'<div style="width:{wo}%;height:100%;background:#8b1a1a;border-radius:0 3px 3px 0"></div></div></div>'
                    f'<span style="font-size:13px;font-weight:700;color:{score_c_o};width:38px">{po}</span>'
                    f'</div></div>')
        face_bars=(face_bar("2PT",c["p2m_g"],c["p2a_g"],c["p2m_o"],c["p2a_o"],"2PT")+
                   face_bar("3PT",c["p3m_g"],c["p3a_g"],c["p3m_o"],c["p3a_o"],"3PT")+
                   face_bar("FT",c["ftm_g"],c["fta_g"],c["ftm_o"],c["fta_o"],"FT"))
        def srow(lbl, vg, vo, higher=True, neutral=False):
            if neutral: cg_s=co_s="#1a2b4a"
            else: rg,ro=col(vg,vo,higher=higher); cg_s=rg; co_s=ro
            return (f'<div style="display:flex;align-items:center;padding:5px 0;border-bottom:0.5px solid #eceef2">'
                    f'<span style="flex:1;font-size:12px;font-weight:600;color:{cg_s};text-align:right;padding-right:12px">{vg}</span>'
                    f'<span style="width:90px;font-size:10px;color:#999;text-align:center;text-transform:uppercase;letter-spacing:.4px;flex-shrink:0">{lbl}</span>'
                    f'<span style="flex:1;font-size:12px;font-weight:600;color:{co_s};text-align:left;padding-left:12px">{vo}</span></div>')
        stat_rows=(srow("AST",c["ast_g"],c["ast_o"])+srow("STL",c["stl_g"],c["stl_o"])+
                   srow("BLK",c["blk_g"],c["blk_o"])+srow("ZB Off",c["oreb_g"],c["oreb_o"])+
                   srow("ZB Def",c["dreb_g"],c["dreb_o"])+
                   srow("DOB M/A",f"{c['d2m_g']}/{c['d2a_g']}",f"{c['d2m_o']}/{c['d2a_o']}",neutral=True)+
                   srow("FD",c["fd_g"],c["fd_o"])+srow("TO",c["br_g"],c["br_o"],higher=False)+
                   srow("TO%",to_g,to_o,higher=False)+srow("NETrtg",net_gs,net_os))
        return f"""
<div style="border-radius:12px;overflow:hidden;border:0.5px solid #e3e8f0;background:#fff">
  <div style="padding:10px 16px;display:flex;align-items:center;gap:10px;border-bottom:0.5px solid #eceef2;background:#f8f9fc">
    <span style="background:#EF9F27;color:#fff;padding:2px 10px;border-radius:20px;font-size:11px;font-weight:700;letter-spacing:.4px">{qtrs_label} CLUTCH</span>
    <span style="font-size:11px;color:#aaa">Ostatnie 3 min Q4</span>
    <span style="margin-left:auto;background:{wc_bg2};color:{wc_col2};padding:2px 10px;border-radius:20px;font-size:11px;font-weight:600">{wc} w clutch</span>
  </div>
  <div style="display:grid;grid-template-columns:1fr auto 1fr;align-items:center;padding:18px 16px;gap:12px;border-bottom:0.5px solid #eceef2">
    <div style="text-align:right">
      <div style="font-size:11px;font-weight:600;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">{gtk_name}</div>
      <div style="font-size:48px;font-weight:800;color:{score_c_g};line-height:1">{c["pts_g"]}</div>
      <div style="font-size:11px;color:#aaa;margin-top:5px">ORtg <span style="color:#555;font-weight:600">{ortg_gs}</span></div>
    </div>
    <div style="text-align:center;padding:0 8px">
      <div style="font-size:11px;color:#bbb;font-weight:400;margin-bottom:8px">vs</div>
      <div style="background:#f4f6fb;border-radius:8px;padding:6px 12px;border:0.5px solid #e3e8f0">
        <div style="font-size:9px;color:#aaa;text-transform:uppercase;letter-spacing:.4px">NETrtg</div>
        <div style="font-size:18px;font-weight:800;color:{net_color}">{net_gs}</div>
      </div>
    </div>
    <div style="text-align:left">
      <div style="font-size:11px;font-weight:600;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">{name_opp}</div>
      <div style="font-size:48px;font-weight:800;color:{score_c_o};line-height:1">{c["pts_o"]}</div>
      <div style="font-size:11px;color:#aaa;margin-top:5px">ORtg <span style="color:#555;font-weight:600">{ortg_os}</span></div>
    </div>
  </div>
  <div style="display:flex;gap:6px;padding:10px 16px;border-bottom:0.5px solid #eceef2;background:#fafbfd">{eff_strip}</div>
  <div style="padding:14px 16px;border-bottom:0.5px solid #eceef2">
    <div style="font-size:9px;color:#aaa;text-transform:uppercase;letter-spacing:.6px;margin-bottom:10px;display:flex;align-items:center;gap:8px">
      <span style="color:#1a6b3c;font-weight:700">{gtk_name}</span>
      <span style="flex:1;height:1px;background:#eceef2"></span>
      <span>Rzuty z pola</span>
      <span style="flex:1;height:1px;background:#eceef2"></span>
      <span style="color:#8b1a1a;font-weight:700">{name_opp}</span>
    </div>
    {face_bars}
  </div>
  <div style="padding:10px 16px 14px">
    <div style="font-size:9px;color:#aaa;text-transform:uppercase;letter-spacing:.6px;margin-bottom:8px">Szczegółowe statystyki</div>
    <div style="display:flex;align-items:center;margin-bottom:4px">
      <span style="flex:1;font-size:10px;font-weight:700;color:#1a6b3c;text-align:right;padding-right:12px">{gtk_name}</span>
      <span style="width:90px;flex-shrink:0"></span>
      <span style="flex:1;font-size:10px;font-weight:700;color:#8b1a1a;text-align:left;padding-left:12px">{name_opp}</span>
    </div>
    {stat_rows}
  </div>
  <div style="padding:4px 16px 8px;font-size:10px;color:#bbb;border-top:0.5px solid #eceef2;background:#fafbfd">
    Ostatnie 1/3 posa\u0107 Q4 (zaokr. w g\u00f3r\u0119) \u2014 GTK: {clutch_poss_g_total} posa\u0107, Rywal: {clutch_poss_o_total} posa\u0107
  </div>
</div>"""

    import json as _json
    try:
        flow_gtk_js = _json.dumps([r["pts_gtk"] for r in flow_rows])
        flow_opp_js = _json.dumps([r["pts_opp"] for r in flow_rows])
    except: flow_gtk_js = "[]"; flow_opp_js = "[]"

    # Comparison data
    def _qval(druzyna, field, q):
        r = next((x for x in all_stats if x['druzyna']==druzyna and x['kwarta']==q), {})
        return r.get(field, 0) or 0
    def _qefg(druzyna, q):
        p2m=_qval(druzyna,'p2m',q); p3m=_qval(druzyna,'p3m',q)
        p2a=_qval(druzyna,'p2a',q); p3a=_qval(druzyna,'p3a',q)
        return round((p2m+1.5*p3m)/(p2a+p3a)*100,1) if (p2a+p3a) else 0
    def _qppp(druzyna, q):
        pts=_qval(druzyna,'pts',q); poss=_qval(druzyna,'poss',q)
        return round(pts/poss,2) if poss else 0
    def _qreb(druzyna, q):
        return (_qval(druzyna,'oreb',q) or 0)+(_qval(druzyna,'dreb',q) or 0)
    def _stl(druzyna):
        return sum(int(r.get("stl",0) or 0) for r in all_stats if r["druzyna"]==druzyna)
    def _fin(druzyna):
        m = sum(int(r.get("d2m",0) or 0) for r in all_stats if r["druzyna"]==druzyna)
        a = sum(int(r.get("d2a",0) or 0) for r in all_stats if r["druzyna"]==druzyna)
        return (m, a)
    def _netrtg_q(q):
        g_poss=_qval("gtk","poss",q); o_poss=_qval("opp","poss",q)
        return round(_qval("gtk","pts",q)/g_poss*100 - _qval("opp","pts",q)/o_poss*100, 1) if (g_poss and o_poss) else 0
    _g_ortg = float(kpi_gtk["ortg"]) if kpi_gtk["ortg"] not in ("-","—") else 0
    _o_ortg = float(kpi_opp["ortg"]) if kpi_opp["ortg"] not in ("-","—") else 0
    _netrtg_total = round(_g_ortg - _o_ortg, 1)
    cmp_list = [
        ("Punkty",      suma_gtk.get("pts",0), suma_opp.get("pts",0), False,
         [_qval("gtk","pts",q) for q in [1,2,3,4]], [_qval("opp","pts",q) for q in [1,2,3,4]]),
        ("Zbiórki",     (suma_gtk.get("oreb",0) or 0)+(suma_gtk.get("dreb",0) or 0),
                        (suma_opp.get("oreb",0) or 0)+(suma_opp.get("dreb",0) or 0), False,
         [_qreb("gtk",q) for q in [1,2,3,4]], [_qreb("opp",q) for q in [1,2,3,4]]),
        ("Asysty",      suma_gtk.get("ast",0), suma_opp.get("ast",0), False,
         [_qval("gtk","ast",q) for q in [1,2,3,4]], [_qval("opp","ast",q) for q in [1,2,3,4]]),
        ("Przechwyty",  _stl("gtk"),            _stl("opp"),           False,
         [_qval("gtk","stl",q) for q in [1,2,3,4]], [_qval("opp","stl",q) for q in [1,2,3,4]]),
        ("Straty (TO)", suma_gtk.get("br",0),  suma_opp.get("br",0),  True,
         [_qval("gtk","br",q) for q in [1,2,3,4]], [_qval("opp","br",q) for q in [1,2,3,4]]),
        ("Dobitki",     f"{_fin('gtk')[0]}/{_fin('gtk')[1]}", f"{_fin('opp')[0]}/{_fin('opp')[1]}", False,
         [_qval("gtk","d2m",q) for q in [1,2,3,4]], [_qval("opp","d2m",q) for q in [1,2,3,4]]),
        ("NetRtg",      _netrtg_total,          -_netrtg_total,        False,
         [_netrtg_q(q) for q in [1,2,3,4]], [-_netrtg_q(q) for q in [1,2,3,4]]),
        ("PPP",         kpi_gtk["ppp"],         kpi_opp["ppp"],        False,
         [_qppp("gtk",q) for q in [1,2,3,4]], [_qppp("opp",q) for q in [1,2,3,4]]),
        ("TS%",         kpi_gtk["ts"],          kpi_opp["ts"],         False,
         [round(_qval("gtk","pts",q)/(2*(_qval("gtk","p2a",q)+_qval("gtk","p3a",q)+0.44*_qval("gtk","fta",q)))*100,1) if (_qval("gtk","p2a",q)+_qval("gtk","p3a",q)+_qval("gtk","fta",q)) else 0 for q in [1,2,3,4]],
         [round(_qval("opp","pts",q)/(2*(_qval("opp","p2a",q)+_qval("opp","p3a",q)+0.44*_qval("opp","fta",q)))*100,1) if (_qval("opp","p2a",q)+_qval("opp","p3a",q)+_qval("opp","fta",q)) else 0 for q in [1,2,3,4]]),
        ("eFG%",        kpi_gtk["efg"],         kpi_opp["efg"],        False,
         [_qefg("gtk",q) for q in [1,2,3,4]], [_qefg("opp",q) for q in [1,2,3,4]]),
    ]
    cmp_js = _json.dumps([{"lbl":l,"g":str(g),"o":str(o),"low":low,"gq":gq,"oq":oq}
                          for l,g,o,low,gq,oq in cmp_list])

    badge = ('<span class="badge-win" style="font-size:.9rem;padding:6px 14px">WYGRANA</span>'
             if m['wynik_gtk']>m['wynik_opp'] else
             '<span class="badge-loss" style="font-size:.9rem;padding:6px 14px">PRZEGRANA</span>'
             if m['wynik_gtk']<m['wynik_opp'] else
             '<span class="badge-draw" style="font-size:.9rem;padding:6px 14px">REMIS</span>')

    q_scores = "".join(
        f'<div style="text-align:center"><div style="font-size:8px;opacity:.55;letter-spacing:.5px;margin-bottom:3px">{q}Q</div>'
        f'<div style="background:rgba(255,255,255,.13);border-radius:6px;padding:4px 12px;font-size:13px;font-weight:700;letter-spacing:1px">'
        f'{pts_q_gtk[q-1]} : {pts_q_opp[q-1]}</div></div>'
        for q in [1,2,3,4])

    content = f"""
<div class="hero mb-3">
  <div class="d-flex justify-content-between align-items-center flex-wrap gap-3">
    <div><div style="font-size:.8rem;opacity:.7">{gtk_name}</div><div style="font-size:2rem;font-weight:700">{m['wynik_gtk']}</div></div>
    <div class="text-center">{badge}</div>
    <div class="text-end"><div style="font-size:.8rem;opacity:.7">{name_opp}</div><div style="font-size:2rem;font-weight:700">{m['wynik_opp']}</div></div>
  </div>
  <div style="border-top:1px solid rgba(255,255,255,.15);margin-top:8px;padding-top:8px;display:flex;justify-content:center;gap:10px">
    {q_scores}
  </div>
</div>

<ul class="nav nav-tabs mb-2" id="mainTabs">
  <li class="nav-item"><button class="nav-link active" data-bs-toggle="tab" data-bs-target="#tabGTK">{gtk_name}</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tabOPP">{name_opp}</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tabCMP">Porównanie</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tabFLOW">Przebieg</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tabCLUTCH">Clutch</button></li>
</ul>

<div class="tab-content">

<div class="tab-pane fade show active" id="tabGTK">
  {kpi_cards(suma_gtk, kpi_gtk)}
  <div style="display:flex;gap:4px;margin:8px 0 4px;border-bottom:1px solid #dee2e6;padding-bottom:0">
    <button onclick="gtkTab('gtk_q')" id="gtktab-gtk_q" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid #1a2b4a;font-weight:500;color:#1a2b4a;cursor:pointer;margin-bottom:-1px">Per kwarta</button>
    <button onclick="gtkTab('gtk_p')" id="gtktab-gtk_p" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid transparent;color:#666;cursor:pointer;margin-bottom:-1px">Statystyki</button>
    <button onclick="gtkTab('gtk_l')" id="gtktab-gtk_l" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid transparent;color:#666;cursor:pointer;margin-bottom:-1px">Piątki</button>
    <button onclick="gtkTab('gtk_t')" id="gtktab-gtk_t" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid transparent;color:#666;cursor:pointer;margin-bottom:-1px">Timing akcji</button>
  </div>
  <div id="gpane-gtk_q" style="display:block"><div class="card mt-1"><div class="card-body" style="padding:0">{q_table('gtk')}</div></div></div>
  <div id="gpane-gtk_p" style="display:none"><div class="card mt-1"><div class="card-body" style="padding:0">{p_table('gtk')}</div></div></div>
  <div id="gpane-gtk_l" style="display:none"><div class="card mt-1"><div class="card-body" style="padding:0">{lineup_table()}</div></div></div>
  <div id="gpane-gtk_t" style="display:none"><div class="card mt-1"><div class="card-body p-2">{tim_table('gtk')}</div></div></div>
</div>

<div class="tab-pane fade" id="tabOPP">
  {kpi_cards(suma_opp, kpi_opp)}
  <div style="display:flex;gap:4px;margin:8px 0 4px;border-bottom:1px solid #dee2e6;padding-bottom:0">
    <button onclick="oppTab('opp_q')" id="opptab-opp_q" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid #1a2b4a;font-weight:500;color:#1a2b4a;cursor:pointer;margin-bottom:-1px">Per kwarta</button>
    <button onclick="oppTab('opp_t')" id="opptab-opp_t" style="font-size:.82rem;padding:6px 12px;border:none;background:none;border-bottom:2px solid transparent;color:#666;cursor:pointer;margin-bottom:-1px">Timing akcji</button>
  </div>
  <div id="opane-opp_q" style="display:block"><div class="card mt-1"><div class="card-body" style="padding:0">{q_table('opp')}</div></div></div>
  <div id="opane-opp_t" style="display:none"><div class="card mt-1"><div class="card-body p-2">{tim_table('opp')}</div></div></div>
</div>

<div class="tab-pane fade" id="tabCMP">
  <div style="margin:10px 0 8px;font-size:.78rem;color:#888">Kliknij metrykę aby rozwinąć wykres per kwarta</div>
  <div id="cmp-accordion"></div>
</div>

<div class="tab-pane fade" id="tabFLOW">
  <div class="card mt-1"><div class="card-body p-2">
    <div style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">Przebieg punktowy</div>
    <div style="background:#f4f6fb;border-radius:8px;padding:8px 12px;margin-bottom:8px;display:flex;gap:16px;align-items:center">
      <div style="display:flex;align-items:center;gap:6px"><div style="width:20px;height:3px;background:#1a6b3c;border-radius:2px"></div><span style="font-size:12px;color:#666">{gtk_name}</span></div>
      <div style="display:flex;align-items:center;gap:6px"><div style="width:20px;height:3px;background:#c0392b;border-radius:2px"></div><span style="font-size:12px;color:#666">{name_opp}</span></div>
    </div>
    <div style="border:0.5px solid #e0e0e0;border-radius:8px;padding:12px"><canvas id="mFlowChart" style="width:100%;height:200px"></canvas></div>
  </div></div>
  <div class="card mt-2" style="overflow:hidden">{momentum_table()}</div>
</div>

<div class="tab-pane fade" id="tabCLUTCH">
  <div class="card mt-1"><div class="card-body p-2">{clutch_stats()}</div></div>
</div>

</div>"""

    scripts = f"""<script>
const gtkName = {_json.dumps(gtk_name)};
const oppName = {_json.dumps(name_opp)};
const CMP_DATA = {cmp_js};
var _cmpCharts = {{}};
(function() {{
  var wrap = document.getElementById('cmp-accordion');
  if (!wrap) return;
  CMP_DATA.forEach(function(m, i) {{
    var gNum = parseFloat(String(m.g).replace('%','')) || 0;
    var oNum = parseFloat(String(m.o).replace('%','')) || 0;
    var total = gNum + oNum || 1;
    var gPct = Math.round(gNum / total * 100); var oPct = 100 - gPct;
    var diff = gNum - oNum;
    var ds = (diff > 0 ? '+' : '') + (diff % 1 === 0 ? Math.round(diff) : diff.toFixed(1));
    var gBetter = m.low ? gNum < oNum : gNum > oNum;
    var dc = diff === 0 ? 'color:#888' : (gBetter ? 'color:#1a6b3c' : 'color:#8b1a1a');
    var el = document.createElement('div');
    el.style.cssText = 'border:0.5px solid #e0e0e0;border-radius:8px;margin-bottom:8px;overflow:hidden';
    el.innerHTML =
      '<div style="display:flex;align-items:center;gap:0;padding:10px 14px;cursor:pointer;background:#fafafa" onclick="toggleCmp('+i+',this)">' +
        '<span style="font-weight:500;min-width:90px;font-size:.85rem">'+m.lbl+'</span>' +
        '<div style="display:flex;align-items:center;gap:10px;flex:1">' +
          '<div style="text-align:right;min-width:52px"><div style="font-size:9px;color:#185FA5;text-transform:uppercase;letter-spacing:.3px">'+gtkName+'</div><div style="font-size:15px;font-weight:600;color:#0C447C">'+m.g+'</div></div>' +
          '<div style="flex:1;height:10px;background:#e8e8e8;border-radius:5px;overflow:hidden;display:flex">' +
            '<div style="flex:'+gPct+';background:#185FA5;height:100%;border-radius:5px 0 0 5px"></div>' +
            '<div style="flex:'+oPct+';background:#c0392b;height:100%;border-radius:0 5px 5px 0"></div>' +
          '</div>' +
          '<div style="text-align:left;min-width:52px"><div style="font-size:9px;color:#c0392b;text-transform:uppercase;letter-spacing:.3px">'+oppName+'</div><div style="font-size:15px;font-weight:600;color:#a32d2d">'+m.o+'</div></div>' +
        '</div>' +
        '<span id="cmp-chev-'+i+'" style="font-size:11px;color:#aaa;margin-left:10px">&#9660;</span>' +
      '</div>' +
      '<div id="cmp-body-'+i+'" style="display:none;padding:0 14px 14px;border-top:0.5px solid #e8e8e8"><canvas id="cmp-chart-'+i+'" height="100"></canvas></div>';
    wrap.appendChild(el);
  }});
}})();
function toggleCmp(i, header) {{
  var body = document.getElementById('cmp-body-'+i);
  var chev = document.getElementById('cmp-chev-'+i);
  var isOpen = body.style.display !== 'none';
  body.style.display = isOpen ? 'none' : 'block';
  chev.style.transform = isOpen ? '' : 'rotate(180deg)';
  if (!isOpen && !_cmpCharts[i]) {{
    var m = CMP_DATA[i]; var ctx = document.getElementById('cmp-chart-'+i);
    if (!ctx) return;
    _cmpCharts[i] = new Chart(ctx, {{
      type: 'bar',
      data: {{ labels: ['1Q','2Q','3Q','4Q'],
        datasets: [
          {{label: gtkName, data: m.gq, backgroundColor: '#185FA5cc', borderColor: '#185FA5', borderWidth:1, borderRadius:4, barPercentage:.45}},
          {{label: oppName, data: m.oq, backgroundColor: '#c0392bcc', borderColor: '#c0392b', borderWidth:1, borderRadius:4, barPercentage:.45}}
        ]}},
      options: {{ responsive:true, maintainAspectRatio:false,
        plugins: {{legend:{{position:'top',labels:{{font:{{size:11}},boxWidth:12}}}},tooltip:{{mode:'index',intersect:false}}}},
        scales: {{x:{{grid:{{display:false}},ticks:{{font:{{size:11}}}}}},y:{{beginAtZero:true,grid:{{color:'rgba(0,0,0,.05)'}},ticks:{{font:{{size:10}}}}}}}}
      }}
    }});
  }}
}}
function gtkTab(id) {{
  ['gtk_q','gtk_p','gtk_l','gtk_t'].forEach(function(p) {{
    var pane=document.getElementById('gpane-'+p); var btn=document.getElementById('gtktab-'+p);
    if(!pane||!btn) return; var active=p===id;
    pane.style.display=active?'block':'none';
    btn.style.borderBottomColor=active?'#1a2b4a':'transparent';
    btn.style.color=active?'#1a2b4a':'#666'; btn.style.fontWeight=active?'500':'normal';
  }});
}}
(function() {{
  var flowTabBtn = document.querySelector('[data-bs-target="#tabFLOW"]');
  if(flowTabBtn) flowTabBtn.addEventListener('shown.bs.tab', function() {{ setTimeout(initFlow, 50); }});
}})();
function oppTab(id) {{
  ['opp_q','opp_t'].forEach(function(p) {{
    var pane=document.getElementById('opane-'+p); var btn=document.getElementById('opptab-'+p);
    if(!pane||!btn) return; var active=p===id;
    pane.style.display=active?'block':'none';
    btn.style.borderBottomColor=active?'#1a2b4a':'transparent';
    btn.style.color=active?'#1a2b4a':'#666'; btn.style.fontWeight=active?'500':'normal';
  }});
}}
(function() {{
  var gtkData = {flow_gtk_js}; var oppData = {flow_opp_js};
  window.initFlow = function() {{
    var canvas=document.getElementById('mFlowChart'); if(!canvas) return;
    if(window._flowChart) {{ window._flowChart.destroy(); }}
    window._flowChart = new Chart(canvas, {{
      type: 'line',
      data: {{ labels: gtkData.map(function(_,i){{return i;}}),
        datasets: [
          {{label: '{gtk_name}', data: gtkData, borderColor:'#1a6b3c', backgroundColor:'transparent', borderWidth:2, pointRadius:0, tension:.3}},
          {{label: '{name_opp}', data: oppData, borderColor:'#c0392b', backgroundColor:'transparent', borderWidth:2, pointRadius:0, tension:.3}}
        ]}},
      options: {{ responsive:true, maintainAspectRatio:false,
        plugins:{{legend:{{display:false}},tooltip:{{mode:'index',intersect:false}}}},
        scales:{{x:{{display:false}},y:{{beginAtZero:true,grid:{{color:'rgba(0,0,0,.05)'}},ticks:{{font:{{size:10}},stepSize:10}}}}}}
      }}
    }});
  }};
}})();
// ── Timing Akcji przełącznik ─────────────────────────────────────────────
function pmTimSwitch(druzyna, view) {{
  var sum = document.getElementById("pm-tim-sum-" + druzyna);
  var q   = document.getElementById("pm-tim-q-"   + druzyna);
  var bS  = document.getElementById("pm-tim-btn-sum-" + druzyna);
  var bQ  = document.getElementById("pm-tim-btn-q-"   + druzyna);
  if (!sum || !q) return;
  if (view === "sum") {{
    sum.style.display = ""; q.style.display = "none";
    bS.style.borderBottomColor="#1a2b4a"; bS.style.color="#1a2b4a"; bS.style.fontWeight="500";
    bQ.style.borderBottomColor="transparent"; bQ.style.color="#888"; bQ.style.fontWeight="normal";
  }} else {{
    sum.style.display = "none"; q.style.display = "";
    bQ.style.borderBottomColor="#1a2b4a"; bQ.style.color="#1a2b4a"; bQ.style.fontWeight="500";
    bS.style.borderBottomColor="transparent"; bS.style.color="#888"; bS.style.fontWeight="normal";
  }}
}}
// ── Sortowanie tabeli zawodników ─────────────────────────────────────────
var _ptSort = {{}};
function sortPTable(tid, col) {{
  var tbl = document.getElementById(tid); if (!tbl) return;
  var tbody = tbl.querySelector('tbody'); if (!tbody) return;
  var cur = _ptSort[tid] || {{col:-1, asc:false}};
  var asc = (cur.col === col) ? !cur.asc : false;
  _ptSort[tid] = {{col:col, asc:asc}};
  var rows = Array.from(tbody.querySelectorAll('tr'));
  rows.sort(function(a, b) {{
    var ac = a.children[col], bc = b.children[col];
    var av = parseFloat((ac && ac.dataset.v) || 0) || 0;
    var bv = parseFloat((bc && bc.dataset.v) || 0) || 0;
    return asc ? av - bv : bv - av;
  }});
  rows.forEach(function(r, i) {{
    var bg = i % 2 === 0 ? '#f8f9ff' : '#fff';
    r.style.background = bg;
    Array.from(r.children).forEach(function(td) {{ td.style.background = bg; }});
    tbody.appendChild(r);
  }});
  tbl.querySelectorAll('th[data-sc]').forEach(function(th) {{
    var ind = th.querySelector('.si'); if (!ind) return;
    if (parseInt(th.dataset.sc) === col) {{
      ind.textContent = asc ? ' ▲' : ' ▼'; ind.style.opacity = '1';
    }} else {{
      ind.textContent = ' ▼'; ind.style.opacity = '0.3';
    }}
  }});
}}
</script>"""

    # Wrap w portal layout z Bootstrap + Chart.js
    page = f"""<!DOCTYPE html>
<html lang="pl"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{gtk_name} vs {name_opp} — Portal BasketKołcz</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
{_PORTAL_CSS}
<style>
.hero{{background:linear-gradient(135deg,#1a2b4a,#253d66);border-radius:14px;padding:20px 24px;color:#fff;box-shadow:0 4px 18px rgba(26,43,74,.18)}}
.badge-win{{background:#c8f7c5;color:#1a5c2a;border-radius:20px;font-weight:700;display:inline-block}}
.badge-loss{{background:#ffd5d5;color:#8b1a1a;border-radius:20px;font-weight:700;display:inline-block}}
.badge-draw{{background:#fff3cd;color:#856404;border-radius:20px;font-weight:700;display:inline-block}}
.nav-tabs{{border-bottom:2px solid #dee2e6}}
.nav-link{{color:#666;border:none;background:none;padding:8px 16px;cursor:pointer;font-size:.88rem}}
.nav-link.active{{color:#1a2b4a;font-weight:600;border-bottom:2px solid #1a2b4a;margin-bottom:-2px}}
.tab-content{{padding-top:8px}}
.tab-pane{{display:none}}.tab-pane.show{{display:block}}
.card{{background:#fff;border:1px solid #e8ecf3;border-radius:10px}}
.card-body{{padding:16px}}
.mtw{{overflow-x:auto}}
.mgood{{color:#0F6E56;font-weight:600}}
.mbad{{color:#A32D2D;font-weight:600}}
.mqdot{{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}}
.mq1{{background:#c8e6c9;color:#1a5c2a}}.mq2{{background:#bbdefb;color:#0d47a1}}
.mq3{{background:#fff9c4;color:#f57f17}}.mq4{{background:#fce4ec;color:#880e4f}}
.mqs{{background:#e8ecf3;color:#1a2b4a}}
.msrow{{background:#f0f2f7!important;font-weight:600}}
.pmwin{{background:#e8f5e9;color:#1a5c2a;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600;display:inline-block}}
.pmloss{{background:#ffebee;color:#8b1a1a;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600;display:inline-block}}
.pmgood{{color:#0F6E56;font-weight:600}}
.pmbad{{color:#A32D2D;font-weight:600}}
.epill{{display:inline-block;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:600}}
.pill-good{{background:#e8f5e9;color:#0F6E56}}
.pill-bad{{background:#ffebee;color:#A32D2D}}
.pill-neu{{background:#f0f4ff;color:#1a2b4a}}
</style>
</head><body style="background:#f0f2f7;min-height:100vh">
<nav class="topbar">
  <div class="tb-logo"><div class="tb-logo-icon"><img src="/static/img/app_logo.png" onerror="this.parentElement.innerHTML='🏀'"></div><div class="tb-logo-text">Basket<span>Kołcz</span></div></div>
  <div class="tb-nav"><a href="/portal?tab=dash" style="text-decoration:none"><button class="tb-btn">← Powrót</button></a></div>
  <div class="tb-right">
    <a class="tb-logout" href="/portal/logout">Wyloguj</a>
  </div>
</nav>
<div style="padding:clamp(10px,3vw,20px);max-width:1200px;margin:0 auto">
  <div style="font-size:.8rem;color:#888;margin-bottom:12px">{gtk_name} vs {name_opp} &middot; {dt} &middot; Sezon {m.get('sezon','')}</div>
  {content}
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
{scripts}
</body></html>"""
    return html_response(page)


@app.route("/portal/login", methods=["POST"])
def portal_login():
    login = request.form.get("login","").strip()
    pwd   = request.form.get("password","").strip()
    if login == _PORTAL_LOGIN and pwd == _PORTAL_PASS:
        session["portal_logged_in"] = True
        session.permanent = True
        return redirect(url_for("portal"))
    return redirect(url_for("portal") + "?err=1")


@app.route("/portal/logout")
def portal_logout():
    session.pop("portal_logged_in", None)
    return redirect(url_for("portal"))


@app.route("/portal/preview-selectors")
def portal_preview_selectors():
    """Podgląd nowych dropdownów sezon/drużyna w topbarze portalu."""
    sezony = ["2025/2026", "2024/2025", "2023/2024"]
    druzyny = ["U18 Złota", "U16 Srebrna", "U14 Brązowa"]
    current_sezon = "2025/2026"
    current_druzyna = "U18 Złota"

    preview_html = f"""<!DOCTYPE html>
<html lang="pl"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Podgląd — Selektory Portalu</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Inter',sans-serif;background:#0f1e35;min-height:100vh;padding:40px 20px}}
.topbar{{background:#1a2b4a;border-bottom:1px solid rgba(0,0,0,.12);padding:0 24px;height:58px;
  display:flex;align-items:center;
  box-shadow:0 2px 8px rgba(0,0,0,.15);border-radius:12px}}
.tb-left{{display:flex;align-items:center;gap:10px;flex-shrink:0}}
.tb-logo{{display:flex;align-items:center;gap:10px}}
.tb-logo-icon{{width:32px;height:32px;background:linear-gradient(135deg,#c9a340,#e8c56a);
  border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:15px}}
.tb-logo-text{{font-size:.95rem;font-weight:800;color:#fff;letter-spacing:-.02em}}
.tb-logo-text span{{color:#c9a340}}
.tb-nav{{display:flex;gap:4px;flex:1;justify-content:center}}
.tb-btn{{padding:6px 15px;border-radius:8px;font-size:.76rem;font-weight:600;cursor:pointer;
  border:none;background:transparent;color:rgba(255,255,255,.5);font-family:inherit;
  letter-spacing:.02em;transition:all .15s}}
.tb-btn:hover{{color:#fff;background:rgba(255,255,255,.1)}}
.tb-btn.active{{color:#c9a340;background:rgba(201,163,64,.15)}}
.tb-right{{display:flex;align-items:center;gap:8px;flex-shrink:0}}
.tb-logout{{font-size:.7rem;color:rgba(255,255,255,.45);cursor:pointer;background:none;
  border:1px solid rgba(255,255,255,.2);border-radius:6px;padding:4px 10px;font-family:inherit;
  text-decoration:none;transition:all .15s}}
.tb-logout:hover{{color:#fff;border-color:rgba(255,255,255,.5)}}
.tb-selector{{position:relative}}
.tb-sel-btn{{display:flex;align-items:center;gap:6px;padding:5px 10px;border-radius:8px;
  border:1px solid rgba(255,255,255,.18);background:rgba(255,255,255,.06);cursor:pointer;
  font-family:inherit;transition:all .18s;min-width:120px}}
.tb-sel-btn:hover{{background:rgba(255,255,255,.12);border-color:rgba(255,255,255,.32)}}
.tb-sel-btn.open{{background:rgba(201,163,64,.15);border-color:rgba(201,163,64,.5)}}
.tb-sel-label{{font-size:.56rem;font-weight:700;text-transform:uppercase;letter-spacing:.07em;
  color:rgba(255,255,255,.38);line-height:1;display:block}}
.tb-sel-value{{font-size:.73rem;font-weight:700;color:#fff;line-height:1.2;display:block;
  margin-top:1px;white-space:nowrap}}
.tb-sel-btn.open .tb-sel-value{{color:#c9a340}}
.tb-chevron{{margin-left:auto;color:rgba(255,255,255,.4);font-size:.58rem;transition:transform .18s;flex-shrink:0}}
.tb-sel-btn.open .tb-chevron{{transform:rotate(180deg);color:#c9a340}}
.tb-dropdown{{position:absolute;top:calc(100% + 6px);left:0;min-width:150px;
  background:#1e3055;border:1px solid rgba(255,255,255,.14);border-radius:10px;
  box-shadow:0 8px 28px rgba(0,0,0,.35);z-index:200;overflow:hidden;
  opacity:0;transform:translateY(-6px);pointer-events:none;
  transition:opacity .18s,transform .18s}}
.tb-selector.open .tb-dropdown{{opacity:1;transform:translateY(0);pointer-events:auto}}
.tb-dd-item{{display:block;padding:9px 14px;font-size:.78rem;font-weight:500;
  color:rgba(255,255,255,.75);cursor:pointer;transition:all .12s;
  border-bottom:1px solid rgba(255,255,255,.06)}}
.tb-dd-item:last-child{{border-bottom:none}}
.tb-dd-item:hover{{background:rgba(255,255,255,.08);color:#fff}}
.tb-dd-item.active{{color:#c9a340;font-weight:700;background:rgba(201,163,64,.1)}}
.tb-dd-item.active::before{{content:"✓ ";font-size:.68rem}}
.tb-sep{{width:1px;height:22px;background:rgba(255,255,255,.12);margin:0 2px}}
.preview-note{{background:rgba(201,163,64,.12);border:1px solid rgba(201,163,64,.3);
  border-radius:10px;padding:14px 18px;margin-bottom:20px;color:rgba(255,255,255,.7);
  font-size:.78rem;max-width:960px;margin-left:auto;margin-right:auto}}
.preview-note strong{{color:#c9a340}}
</style>
</head><body>

<div class="preview-note" style="margin-bottom:16px">
  <strong>Podgląd UI</strong> — selektory sezon / drużyna między logo a nawigacją.
  Kliknij aby rozwinąć listę.
</div>

<div class="topbar" style="max-width:960px;margin:0 auto">
  <div class="tb-left">
    <div class="tb-logo">
      <div class="tb-logo-icon"><img src="/static/img/app_logo.png" onerror="this.parentElement.innerHTML='🏀'"></div>
      <div class="tb-logo-text">Basket<span>Kołcz</span></div>
    </div>

    <!-- Selektory po lewej, za logo -->
    <div style="display:flex;align-items:center;gap:6px;margin-left:12px">
    <div class="tb-selector" id="sel-sezon">
      <button class="tb-sel-btn" onclick="toggleSel('sel-sezon')">
        <div>
          <span class="tb-sel-label">Sezon</span>
          <span class="tb-sel-value" id="val-sezon">{current_sezon}</span>
        </div>
        <span class="tb-chevron">▼</span>
      </button>
      <div class="tb-dropdown">
        {''.join(f'<div class="tb-dd-item{" active" if s == current_sezon else ""}" onclick="selectSezon(this,\\"{s}\\")">{s}</div>' for s in sezony)}
      </div>
    </div>
    <div class="tb-sep"></div>
    <div class="tb-selector" id="sel-druzyna">
      <button class="tb-sel-btn" onclick="toggleSel('sel-druzyna')">
        <div>
          <span class="tb-sel-label">Drużyna</span>
          <span class="tb-sel-value" id="val-druzyna">{current_druzyna}</span>
        </div>
        <span class="tb-chevron">▼</span>
      </button>
      <div class="tb-dropdown">
        {''.join(f'<div class="tb-dd-item{" active" if d == current_druzyna else ""}" onclick="selectDruzyna(this,\\"{d}\\")">{d}</div>' for d in druzyny)}
      </div>
    </div>
    </div>
  </div>

  <div class="tb-nav">
    <button class="tb-btn active">Główny Pulpit</button>
    <button class="tb-btn">Mecze</button>
    <button class="tb-btn">Zawodnicy</button>
  </div>
  <div class="tb-right">
    <a class="tb-logout" href="#">Wyloguj</a>
  </div>
</div>

<script>
function toggleSel(id){{
  document.querySelectorAll('.tb-selector').forEach(function(el){{
    if(el.id!==id){{el.classList.remove('open');el.querySelector('.tb-sel-btn').classList.remove('open');}}
  }});
  var el=document.getElementById(id);
  el.classList.toggle('open');
  el.querySelector('.tb-sel-btn').classList.toggle('open');
}}
function selectSezon(item,val){{
  document.getElementById('val-sezon').textContent=val;
  item.closest('.tb-dropdown').querySelectorAll('.tb-dd-item').forEach(i=>i.classList.remove('active'));
  item.classList.add('active');
  toggleSel('__close__');
}}
function selectDruzyna(item,val){{
  document.getElementById('val-druzyna').textContent=val;
  item.closest('.tb-dropdown').querySelectorAll('.tb-dd-item').forEach(i=>i.classList.remove('active'));
  item.classList.add('active');
  toggleSel('__close__');
}}
document.addEventListener('click',function(e){{
  if(!e.target.closest('.tb-selector')){{
    document.querySelectorAll('.tb-selector').forEach(function(el){{
      el.classList.remove('open');
      var b=el.querySelector('.tb-sel-btn');if(b)b.classList.remove('open');
    }});
  }}
}});
</script>
</body></html>"""
    return html_response(preview_html)


@app.route("/portal/zawodnik/<int:pid>")
def portal_zawodnik(pid):
    if not session.get("portal_logged_in"):
        return redirect(url_for("portal"))

    sezon_filter = request.args.get("sezon", get_setting("current_season") or "")
    db = get_db(); cur = db.cursor()

    # Dane zawodnika — sprawdź obie tabele
    zawodnik = None
    use_player_id = False
    cur.execute("SELECT * FROM roster WHERE id=%s", (pid,))
    zawodnik_roster = cur.fetchone()
    if zawodnik_roster:
        zawodnik = zawodnik_roster
        use_player_id = False
    else:
        cur.execute("SELECT * FROM players WHERE id=%s", (pid,))
        zawodnik_player = cur.fetchone()
        if zawodnik_player:
            zawodnik = zawodnik_player
            use_player_id = True

    if not zawodnik:
        cur.close()
        return redirect(url_for("portal") + "?tab=players")

    join_col = "ps.player_id" if use_player_id else "ps.roster_id"

    # Wszystkie sezony zawodnika
    cur.execute(f"""SELECT DISTINCT m.sezon FROM player_stats ps
                   JOIN matches m ON ps.match_id=m.id
                   WHERE {join_col}=%s ORDER BY m.sezon DESC""", (pid,))
    sezony = [r["sezon"] for r in cur.fetchall()]

    # Statystyki per mecz w wybranym sezonie
    cur.execute(f"""
        SELECT ps.*, m.data_meczu, m.przeciwnik, m.wynik_gtk, m.wynik_opp,
               ms.poss as team_poss
        FROM player_stats ps
        JOIN matches m ON ps.match_id=m.id
        LEFT JOIN (SELECT match_id, SUM(poss) as poss FROM match_stats
                   WHERE druzyna='gtk' GROUP BY match_id) ms ON ms.match_id=ps.match_id
        WHERE {join_col}=%s AND m.sezon=%s AND ps.druzyna='gtk'
        ORDER BY m.data_meczu ASC
    """, (pid, sezon_filter))
    mecze_stats = list(cur.fetchall())

    # Numery w sezonie
    cur.execute(f"""SELECT DISTINCT ps.nr FROM player_stats ps
                   JOIN matches m ON ps.match_id=m.id
                   WHERE {join_col}=%s AND m.sezon=%s""", (pid, sezon_filter))
    numery = [str(r["nr"]) for r in cur.fetchall()]
    cur.close()

    # ── Preload calc_play_time per match (niezależnie od DB time_sum) ──────────
    _pt_per_match_portal = {}
    try:
        for _mr in mecze_stats:
            _mid = _mr["match_id"]
            _mnr = int(_mr.get("nr") or 0)
            _pt_per_match_portal[_mid] = calc_play_time(_mid).get(_mnr, 0)
    except Exception:
        pass

    gtk_name = get_setting("gtk_name") or "GTK"

    if not mecze_stats:
        page = f"""<!DOCTYPE html>
<html lang="pl"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{zawodnik['imie']} {zawodnik['nazwisko']} — Portal</title>
{_PORTAL_CSS}
</head><body>
<nav class="topbar">
  <div class="tb-logo"><div class="tb-logo-icon"><img src="/static/img/app_logo.png" onerror="this.parentElement.innerHTML='🏀'"></div><div class="tb-logo-text">Basket<span>Kołcz</span></div></div>
  <div class="tb-nav"><a href="/portal?tab=players" style="text-decoration:none"><button class="tb-btn">← Powrót</button></a></div>
  <div class="tb-right"><a class="tb-logout" href="/portal/logout">Wyloguj</a></div>
</nav>
<div style="padding:24px;max-width:1100px;margin:0 auto">
  <div style="background:#fff;border:1px solid #e8ecf3;border-radius:12px;padding:40px;text-align:center;color:#9ca3af">
    Brak danych w sezonie {sezon_filter}.
  </div>
</div></body></html>"""
        return html_response(page)

    # Agregaty sezonu
    def s(k): return sum(int(r.get(k,0) or 0) for r in mecze_stats)
    n = len(mecze_stats)
    pts_tot = s("pts"); fga_tot = s("p2a")+s("p3a"); fta_tot = s("fta")
    pm2_tot = s("p2m"); pm3_tot = s("p3m"); ftm_tot = s("ftm")
    br_tot = s("br"); fin_tot = s("finishes"); ast_tot = s("ast")
    stl_tot = s("stl"); blk_tot = s("blk"); fd_tot = s("fd")
    poss_tot = sum(int(r.get("team_poss",0) or 0) for r in mecze_stats)

    ppg   = f"{pts_tot/n:.1f}"
    efg   = f"{(pm2_tot+1.5*pm3_tot)/fga_tot:.1%}" if fga_tot else "—"
    ts    = f"{pts_tot/(2*(fga_tot+0.44*fta_tot)):.1%}" if (fga_tot+fta_tot) else "—"
    usg   = f"{(fga_tot+0.44*fta_tot+br_tot)/poss_tot:.1%}" if poss_tot else "—"
    p2pct = f"{pm2_tot/s('p2a'):.1%}" if s('p2a') else "—"
    p3pct = f"{pm3_tot/s('p3a'):.1%}" if s('p3a') else "—"
    ftpct = f"{ftm_tot/fta_tot:.1%}" if fta_tot else "—"
    tsum_tot = sum(float(r.get("time_sum") or 0) for r in mecze_stats)
    tcnt_tot = sum(int(r.get("time_cnt") or 0) for r in mecze_stats)
    avg_t_tot = f"{tsum_tot/tcnt_tot:.1f}s" if tcnt_tot else "—"

    # KPI cards
    def kpi_card(val, lbl, color="#1a2b4a", subtitle="", tooltip=""):
        sub_html = f'<div style="font-size:.65rem;color:#bbb;margin-top:1px">{subtitle}</div>' if subtitle else ""
        tip_attr = f' title="{tooltip}"' if tooltip else (f' title="{subtitle}"' if subtitle else "")
        cur2 = 'help' if (tooltip or subtitle) else 'default'
        bs_tip = ""
        if tooltip:
            safe = tooltip.replace('"', '&quot;')
            bs_tip = f' data-bs-toggle="tooltip" data-bs-placement="top" data-bs-title="{safe}"'
        return (f'<div class="col"><div class="stat-card"{tip_attr}{bs_tip} style="cursor:{cur2}">'
                f'<div class="stat-val sm" style="color:{color}">{val}</div>'
                f'<div class="stat-lbl">{lbl}</div>'
                f'{sub_html}'
                f'</div></div>')

    kpi_html = (
        kpi_card(p2pct,"2PT%",    "#1a2b4a", "2-point FG%",
            "Skuteczność rzutów za 2 punkty. Celne za 2 / wszystkie próby za 2.") +
        kpi_card(p3pct,"3PT%",    "#1a2b4a", "3-point FG%",
            "Skuteczność rzutów za 3 punkty. Celne za 3 / wszystkie próby za 3.") +
        kpi_card(ftpct,"FT%",     "#1a2b4a", "free throw %",
            "Skuteczność rzutów wolnych. Celne wolne / wszystkie próby wolne.") +
        kpi_card(efg,  "eFG%",    "#1a2b4a", "effective FG%",
            "Skuteczność rzutów z wagą dla trójek. Wzór: (2PM + 1.5×3PM) / FGA. Trójka warta więcej niż dwójka.") +
        kpi_card(ts,   "TS%",     "#1a2b4a", "true shooting %",
            "Prawdziwa skuteczność uwzględniająca rzuty wolne. Wzór: PTS / (2 × (FGA + 0.44×FTA)).") +
        kpi_card(usg,  "USG%",    "#D85A30", "usage rate",
            "Procent akcji drużyny zakończonych przez zawodnika. Wzór: (FGA + 0.44×FTA + TO) / akcje drużyny.") +
        kpi_card(avg_t_tot, "avg. play duration", "#555", "",
            "Średni czas trwania akcji zakończonych przez zawodnika (w sekundach). Im niższy — tym szybsze decyzje.")
    )

    # Numer/inicjały w awatarze
    initials = (zawodnik['imie'][0] + zawodnik['nazwisko'][0]).upper()
    nr_display = numery[0] if len(numery) == 1 else ""
    season_opts = "".join([f'<option value="{sv}" {"selected" if sv==sezon_filter else ""}>{sv}</option>' for sv in sezony])

    # Tabela meczów — odwrócona kolejność (najnowsze na górze)
    match_rows_rev = ""
    for r in reversed(mecze_stats):
        dt = r["data_meczu"].strftime("%d.%m.%y") if r["data_meczu"] else ""
        dt_sort = r["data_meczu"].strftime("%Y%m%d") if r["data_meczu"] else "0"
        pts = int(r.get("pts",0) or 0)
        p2a = int(r.get("p2a",0) or 0); p2m = int(r.get("p2m",0) or 0)
        p3a = int(r.get("p3a",0) or 0); p3m = int(r.get("p3m",0) or 0)
        fta = int(r.get("fta",0) or 0); ftm = int(r.get("ftm",0) or 0)
        br  = int(r.get("br",0) or 0)
        fin = int(r.get("finishes",0) or 0)
        ast = int(r.get("ast",0) or 0)
        fd_m = int(r.get("fd",0) or 0)
        oreb_m = int(r.get("oreb",0) or 0)
        dreb_m = int(r.get("dreb",0) or 0)
        stl_m  = int(r.get("stl",0) or 0)
        blk_m  = int(r.get("blk",0) or 0)
        fga = p2a + p3a
        efg_m = f"{(p2m+1.5*p3m)/fga*100:.1f}%" if fga else "—"
        ts_m  = f"{pts/(2*(fga+0.44*fta))*100:.1f}%" if (fga+fta) else "—"
        tposs = int(r.get("team_poss",0) or 0)
        usg_m = f"{(fga+0.44*fta+br)/tposs*100:.1f}%" if tposs else "—"
        _pt_secs_m = _pt_per_match_portal.get(r["match_id"], 0)
        if _pt_secs_m:
            min_m = f"{int(_pt_secs_m)//60}:{int(_pt_secs_m)%60:02d}"
        else:
            tsum_m = float(r.get("time_sum") or 0)
            min_m = f"{int(tsum_m/60*1.22)}:{int((tsum_m/60*1.22 % 1)*60):02d}" if tsum_m else "—"
        p2pct_m = f"{p2m/p2a*100:.1f}%" if p2a else "—"
        p3pct_m = f"{p3m/p3a*100:.1f}%" if p3a else "—"
        ftpct_m = f"{ftm/fta*100:.1f}%" if fta else "—"
        wg = int(r.get("wynik_gtk",0) or 0); wo = int(r.get("wynik_opp",0) or 0)
        match_rows_rev += f"""<tr data-date="{dt_sort}">
            <td style="font-size:.78rem;font-weight:500;white-space:nowrap">{r['przeciwnik']}</td>
            <td class="text-center" style="font-size:.75rem;color:#888;white-space:nowrap">{dt}</td>
            <td class="text-center" style="color:#633806;white-space:nowrap">{min_m}</td>
            <td class="text-center fw-bold" onclick="switchMetric('PTS')" style="cursor:pointer">{pts}</td>
            <td class="text-center" style="border-left:1px solid #f0f0f0;white-space:nowrap">{p2m}/{p2a}</td>
            <td class="text-center;white-space:nowrap">{p2pct_m}</td>
            <td class="text-center" style="border-left:1px solid #f0f0f0;white-space:nowrap">{p3m}/{p3a}</td>
            <td class="text-center;white-space:nowrap">{p3pct_m}</td>
            <td class="text-center" style="border-left:1px solid #f0f0f0;white-space:nowrap">{ftm}/{fta}</td>
            <td class="text-center;white-space:nowrap">{ftpct_m}</td>
            <td class="text-center" onclick="switchMetric('OREB')" style="cursor:pointer;border-left:1px solid #f0f0f0">{oreb_m}</td>
            <td class="text-center" onclick="switchMetric('DREB')" style="cursor:pointer">{dreb_m}</td>
            <td class="text-center" onclick="switchMetric('REB')"  style="cursor:pointer;font-weight:500">{oreb_m + dreb_m}</td>
            <td class="text-center" onclick="switchMetric('AST')" style="cursor:pointer">{ast}</td>
            <td class="text-center" onclick="switchMetric('TO')" style="cursor:pointer">{br}</td>
            <td class="text-center" onclick="switchMetric('STL')" style="cursor:pointer">{stl_m}</td>
            <td class="text-center" onclick="switchMetric('BLK')" style="cursor:pointer">{blk_m}</td>
            <td class="text-center">{fd_m}</td>
            <td class="text-center" onclick="switchMetric('EFG')" style="cursor:pointer"><b>{efg_m}</b></td>
            <td class="text-center" onclick="switchMetric('TS')" style="cursor:pointer">{ts_m}</td>
            <td class="text-center" onclick="switchMetric('USG')" style="cursor:pointer">{usg_m}</td>
        </tr>"""

    # Dane JS dla wykresów
    import json as _json
    mecze_chron = mecze_stats
    labels_js   = _json.dumps([r["przeciwnik"][:8] for r in mecze_chron])
    pts_js      = _json.dumps([int(r.get("pts",0) or 0) for r in mecze_chron])
    ast_js      = _json.dumps([int(r.get("ast",0) or 0) for r in mecze_chron])
    br_js       = _json.dumps([int(r.get("br",0) or 0) for r in mecze_chron])
    fin_js      = _json.dumps([int(r.get("finishes",0) or 0) for r in mecze_chron])
    reb_js      = _json.dumps([int(r.get("oreb",0) or 0)+int(r.get("dreb",0) or 0) for r in mecze_chron])
    oreb_js     = _json.dumps([int(r.get("oreb",0) or 0) for r in mecze_chron])
    dreb_js     = _json.dumps([int(r.get("dreb",0) or 0) for r in mecze_chron])
    stl_js      = _json.dumps([int(r.get("stl",0) or 0) for r in mecze_chron])
    blk_js      = _json.dumps([int(r.get("blk",0) or 0) for r in mecze_chron])
    efg_js      = _json.dumps([round((int(r.get("p2m",0) or 0)+1.5*int(r.get("p3m",0) or 0))/(int(r.get("p2a",0) or 0)+int(r.get("p3a",0) or 0))*100,1) if (int(r.get("p2a",0) or 0)+int(r.get("p3a",0) or 0)) else None for r in mecze_chron])

    def _ts_val(r):
        p = int(r.get("pts",0) or 0)
        fga_r = int(r.get("p2a",0) or 0)+int(r.get("p3a",0) or 0)
        fta_r = int(r.get("fta",0) or 0)
        return round(p/(2*(fga_r+0.44*fta_r))*100,1) if (fga_r+fta_r) else None
    def _usg_val(r):
        fga_r = int(r.get("p2a",0) or 0)+int(r.get("p3a",0) or 0)
        fta_r = int(r.get("fta",0) or 0)
        br_r  = int(r.get("br",0) or 0)
        tp    = int(r.get("team_poss",0) or 0)
        return round((fga_r+0.44*fta_r+br_r)/tp*100,1) if tp else None

    ts_js       = _json.dumps([_ts_val(r) for r in mecze_chron])
    usg_js      = _json.dumps([_usg_val(r) for r in mecze_chron])
    p2pct_js    = round(pm2_tot/s('p2a')*100,1) if s('p2a') else 0
    p3pct_js    = round(pm3_tot/s('p3a')*100,1) if s('p3a') else 0
    ftpct_js    = round(ftm_tot/fta_tot*100,1) if fta_tot else 0
    efg_season  = round((pm2_tot+1.5*pm3_tot)/fga_tot*100,1) if fga_tot else 0
    ts_season   = round(pts_tot/(2*(fga_tot+0.44*fta_tot))*100,1) if (fga_tot+fta_tot) else 0
    usg_season  = round((fga_tot+0.44*fta_tot+br_tot)/poss_tot*100,1) if poss_tot else 0
    avg_pts     = round(pts_tot/n, 1)

    def season_bar(vals):
        nz = [v for v in vals if v is not None]
        if not nz: return [0, 0, 0]
        return [min(nz), round(sum(nz)/len(nz),1), max(nz)]

    ast_season_js  = _json.dumps(season_bar([int(r.get("ast",0) or 0) for r in mecze_stats]))
    br_season_js   = _json.dumps(season_bar([int(r.get("br",0) or 0) for r in mecze_stats]))
    fin_season_js  = _json.dumps(season_bar([int(r.get("finishes",0) or 0) for r in mecze_stats]))
    reb_season_js  = _json.dumps(season_bar([int(r.get("oreb",0) or 0)+int(r.get("dreb",0) or 0) for r in mecze_stats]))
    oreb_season_js = _json.dumps(season_bar([int(r.get("oreb",0) or 0) for r in mecze_stats]))
    dreb_season_js = _json.dumps(season_bar([int(r.get("dreb",0) or 0) for r in mecze_stats]))
    stl_season_js  = _json.dumps(season_bar([int(r.get("stl",0) or 0) for r in mecze_stats]))
    blk_season_js  = _json.dumps(season_bar([int(r.get("blk",0) or 0) for r in mecze_stats]))

    bar_w = max(16, min(50, 300 // max(n, 1)))

    page = f"""<!DOCTYPE html>
<html lang="pl"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{zawodnik['imie']} {zawodnik['nazwisko']} — Portal BasketKołcz</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
{_PORTAL_CSS}
<style>
.stat-card{{background:#fff;border:1px solid #e8ecf3;border-radius:10px;padding:14px 10px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.05)}}
.stat-val{{font-size:1.35rem;font-weight:800;color:#1a2b4a;line-height:1.1}}
.stat-val.sm{{font-size:1.1rem}}
.stat-lbl{{font-size:.6rem;text-transform:uppercase;letter-spacing:.07em;color:#9ca3af;margin-top:3px}}
.section-hdr{{font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:#6b7280;margin-bottom:8px}}
.card{{background:#fff;border:1px solid #e8ecf3;border-radius:12px;box-shadow:0 1px 3px rgba(0,0,0,.05)}}
.card-body{{padding:16px}}
@media(max-width:768px){{
  .stat-card{{padding:10px 6px}}
  .stat-val{{font-size:1.1rem}}
  .stat-val.sm{{font-size:.95rem}}
  .stat-lbl{{font-size:.55rem}}
  .card-body{{padding:10px}}
  .page-title{{font-size:1rem}}
  .table thead th{{font-size:8px;padding:4px 5px}}
  .table tbody td{{font-size:.72rem;padding:5px 5px}}
}}
@media(max-width:480px){{
  .stat-val{{font-size:.95rem}}
  .stat-lbl{{font-size:.52rem}}
  .table thead th{{font-size:7px;padding:3px 4px}}
  .table tbody td{{font-size:.65rem;padding:4px 4px}}
}}
.table{{font-size:.78rem;margin-bottom:0}}
.table thead th{{background:#1a2b4a;color:#fff;font-size:9px;font-weight:500;padding:5px 8px;border-bottom:0.5px solid rgba(255,255,255,.2);white-space:nowrap;vertical-align:middle}}
.table tbody tr:hover{{background:#f8fafc}}
.fw-bold{{font-weight:700!important}}
.text-center{{text-align:center!important}}
.badge{{display:inline-block;padding:.25em .55em;font-size:.7em;font-weight:700;border-radius:.35rem}}
.page-title{{font-size:1.2rem;font-weight:700;color:#1a2b4a}}
</style>
</head><body style="background:#f0f2f7;min-height:100vh">
<nav class="topbar">
  <div class="tb-logo">
    <div class="tb-logo-icon"><img src="/static/img/app_logo.png" onerror="this.parentElement.innerHTML='🏀'"></div>
    <div class="tb-logo-text">Basket<span>Kołcz</span></div>
  </div>
  <div class="tb-nav">
    <a href="/portal?tab=players" style="text-decoration:none"><button class="tb-btn">← Powrót</button></a>
  </div>
  <div class="tb-right">
    <a class="tb-logout" href="/portal/logout">Wyloguj</a>
  </div>
</nav>
<div style="padding:clamp(10px,3vw,24px);max-width:1200px;margin:0 auto">

<!-- HEADER -->
<div class="card mb-3">
  <div class="card-body p-3">
    <div class="d-flex align-items-center gap-3 flex-wrap justify-content-between">
      <div style="display:flex;align-items:center;gap:16px;flex-shrink:0">
        <div style="width:56px;height:56px;border-radius:50%;background:#1a2b4a;display:flex;flex-direction:column;align-items:center;justify-content:center;flex-shrink:0">
          {"<span style='font-size:9px;color:rgba(255,255,255,.5);line-height:1'>#</span><span style='font-size:20px;font-weight:700;color:#fff;line-height:1.1'>" + nr_display + "</span>" if nr_display else "<span style='font-size:20px;font-weight:700;color:#fff'>" + initials + "</span>"}
        </div>
        <div>
          <div style="font-size:20px;font-weight:700;color:#1a2b4a">{zawodnik['imie']} {zawodnik['nazwisko']}</div>
          <div style="font-size:.82rem;color:#888;margin-top:2px">{gtk_name}</div>
          <div style="margin-top:6px">
            <span class="badge" style="background:{'#e8f5e9;color:#1a5c2a' if zawodnik.get('aktywny', True) else '#ffebee;color:#8b1a1a'}">{'Aktywny' if zawodnik.get('aktywny', True) else 'Nieaktywny'}</span>
          </div>
        </div>
      </div>
      <div style="display:flex;gap:0;margin:0 auto;background:#1a2b4a;border-radius:10px;padding:6px 10px;flex-shrink:0">
        {"".join([
          f'<div style="display:flex;flex-direction:column;align-items:center;padding:6px 14px;border-radius:8px" title="{tip}"><span style="font-size:17px;font-weight:500;color:#fff;line-height:1.2">{val}</span><span style="font-size:10px;color:rgba(255,255,255,.45);margin-top:3px;text-transform:uppercase;letter-spacing:.5px">{lbl}</span></div>'
          + ('' if i == 8 else '<div style="width:0.5px;background:rgba(255,255,255,.15);align-self:stretch;margin:4px 0"></div>')
          for i, (val, lbl, tip) in enumerate([
            (ppg, "PPG", "Średnia punktów na mecz"),
            (f"{ast_tot/n:.1f}", "AST", "Średnia asyst na mecz"),
            (f"{(sum(int(r.get('oreb',0) or 0)+int(r.get('dreb',0) or 0) for r in mecze_stats))/n:.1f}", "REB", "Średnia zbiórek na mecz"),
            (f"{stl_tot/n:.1f}", "STL", "Średnia przechwytów na mecz"),
            (f"{blk_tot/n:.1f}", "BLK", "Średnia bloków na mecz"),
            (f"{br_tot/n:.1f}", "TO", "Średnia strat na mecz"),
            (f"{fd_tot/n:.1f}", "FD", "Średnia faulów wymuszonych"),
            (f"{fin_tot/n:.1f}", "FIN", "Średnia wykończeń na mecz"),
            (str(n), "Mecze", "Liczba meczów w sezonie"),
          ])
        ])}
      </div>
      <form method="GET" class="d-flex gap-2 align-items-center ms-auto">
        <label style="font-size:.82rem;font-weight:600">Sezon:</label>
        <select name="sezon" class="form-select form-select-sm" style="width:120px" onchange="this.form.submit()">
          {season_opts}
        </select>
      </form>
    </div>
  </div>
</div>

<!-- KPI -->
<div class="row g-2 mb-3">{kpi_html}</div>

<!-- WYKRESY -->
<div class="row g-3 mb-3">
  <div class="col-lg-7">
    <div class="card h-100"><div class="card-body p-2" style="position:relative">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:4px">
        <div class="section-hdr mb-0" id="chartMainTitle">Punkty</div>
        <div style="font-size:10px;color:#aaa;text-transform:uppercase;letter-spacing:.6px">Per mecz</div>
      </div>
      <div style="position:relative;height:180px">
        <canvas id="chartMain"></canvas>
        <div id="chartMainCenter" style="display:none;position:absolute;top:50%;left:50%;transform:translate(-50%,-65%);text-align:center;pointer-events:none">
          <div id="chartMainCenterVal" style="font-size:22px;font-weight:500;color:#1a2b4a;line-height:1.1"></div>
          <div style="font-size:11px;color:#888;margin-top:2px">śr. per mecz</div>
        </div>
      </div>
    </div></div>
  </div>
  <div class="col-lg-5">
    <div class="card h-100"><div class="card-body p-2" style="position:relative">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:4px">
        <div class="section-hdr mb-0" id="chartShootTitle">Skuteczność rzutów</div>
        <div style="font-size:10px;color:#aaa;text-transform:uppercase;letter-spacing:.6px">Sezon</div>
      </div>
      <div style="position:relative;height:180px">
        <canvas id="chartShoot"></canvas>
        <div id="chartShootCenter" style="display:none;position:absolute;top:50%;left:50%;transform:translate(-50%,-65%);text-align:center;pointer-events:none">
          <div id="chartShootCenterVal" style="font-size:22px;font-weight:500;color:#1a2b4a;line-height:1.1"></div>
          <div style="font-size:11px;color:#888;margin-top:2px">sezon</div>
        </div>
      </div>
    </div></div>
  </div>
</div>

<!-- TABELA MECZÓW -->
<div class="card mb-3"><div class="card-body p-2">
  <div class="section-hdr">Przebieg sezonu — mecz po meczu</div>
  <div class="table-responsive">
  <table class="table table-hover mb-0" style="min-width:900px">
    <thead>
      <tr>
        <th rowspan="3" style="text-align:left">Zawodnik</th>
        <th rowspan="3" style="text-align:center">Data</th>
        <th rowspan="3" style="text-align:center">MIN<br>(szac.)</th>
        <th rowspan="3" id="thPTS" onclick="switchMetric('PTS')" style="text-align:center;cursor:pointer">PTS ↓</th>
        <th colspan="2" style="color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center">2PT</th>
        <th colspan="2" style="color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center">3PT</th>
        <th colspan="2" style="color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;border-bottom:0.5px solid rgba(255,255,255,.15);text-align:center">FT</th>
        <th colspan="3" style="color:rgba(255,255,255,.55);font-size:8px;letter-spacing:.3px;border-bottom:0.5px solid rgba(255,255,255,.15);background:#152236;text-align:center">ZB</th>
        <th rowspan="3" id="thAST" onclick="switchMetric('AST')" style="text-align:center;cursor:pointer">AST</th>
        <th rowspan="3" id="thTO"  onclick="switchMetric('TO')"  style="text-align:center;cursor:pointer">TO</th>
        <th rowspan="3" id="thSTL" onclick="switchMetric('STL')" style="text-align:center;cursor:pointer">STL</th>
        <th rowspan="3" id="thBLK" onclick="switchMetric('BLK')" style="text-align:center;cursor:pointer">BLK</th>
        <th rowspan="3" style="text-align:center">FD</th>
        <th rowspan="3" id="thEFG" onclick="switchMetric('EFG')" style="text-align:center;cursor:pointer">eFG%</th>
        <th rowspan="3" id="thTS"  onclick="switchMetric('TS')"  style="text-align:center;cursor:pointer">TS%</th>
        <th rowspan="3" id="thUSG" onclick="switchMetric('USG')" style="text-align:center;cursor:pointer">USG%</th>
      </tr>
      <tr>
        <th style="color:rgba(255,255,255,.8);font-size:9px;text-align:center">M/A</th>
        <th style="color:rgba(255,255,255,.8);font-size:9px;text-align:center">%</th>
        <th style="color:rgba(255,255,255,.8);font-size:9px;text-align:center">M/A</th>
        <th style="color:rgba(255,255,255,.8);font-size:9px;text-align:center">%</th>
        <th style="color:rgba(255,255,255,.8);font-size:9px;text-align:center">M/A</th>
        <th style="color:rgba(255,255,255,.8);font-size:9px;text-align:center">%</th>
        <th id="thOREB" onclick="switchMetric('OREB')" style="color:rgba(255,255,255,.75);font-size:9px;text-align:center;background:#152236;cursor:pointer">A</th>
        <th id="thDREB" onclick="switchMetric('DREB')" style="color:rgba(255,255,255,.75);font-size:9px;text-align:center;background:#152236;cursor:pointer">O</th>
        <th id="thREB"  onclick="switchMetric('REB')"  style="color:rgba(255,255,255,.75);font-size:9px;text-align:center;background:#152236;cursor:pointer">S</th>
      </tr>
    </thead>
    <tbody id="tbody-mecze-profil">{match_rows_rev}</tbody>
  </table>
  </div>
</div></div>

</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
<script>
  var labels = {labels_js};
  var nMecze = labels.length;
  var barW = Math.max(16, Math.min(50, Math.floor(280 / Math.max(nMecze, 1))));

  var dataSets = {{
    'PTS':  {pts_js},
    'AST':  {ast_js},
    'REB':  {reb_js},
    'OREB': {oreb_js},
    'DREB': {dreb_js},
    'STL':  {stl_js},
    'BLK':  {blk_js},
    'TO':   {br_js},
    'FIN':  {fin_js},
    'EFG':  {efg_js},
    'TS':   {ts_js},
    'USG':  {usg_js}
  }};
  var titles = {{
    'PTS': 'Punkty', 'AST': 'Asysty',
    'REB': 'Rebounds', 'OREB': 'Off. Reb', 'DREB': 'Def. Reb',
    'STL': 'Przechwyty', 'BLK': 'Bloki', 'TO': 'Straty',
    'FIN': 'Wykończenia', 'EFG': 'eFG%', 'TS': 'TS%', 'USG': 'USG%'
  }};
  var colors = {{
    'PTS': '#1a2b4a', 'AST': '#378ADD',
    'REB': '#1D9E75', 'OREB': '#27ae60', 'DREB': '#16875a',
    'STL': '#185FA5', 'BLK': '#0C447C', 'TO':  '#D85A30',
    'FIN': '#555555', 'EFG': '#534AB7', 'TS':  '#0F6E56', 'USG': '#D85A30'
  }};
  var ptsAvg = {avg_pts};
  var pieMetrics = new Set(['EFG','TS','USG']);
  var piePerMatch = {{
    'EFG': {{ data: {efg_js}, color: '#534AB7' }},
    'TS':  {{ data: {ts_js},  color: '#0F6E56' }},
    'USG': {{ data: {usg_js}, color: '#D85A30' }}
  }};
  var pieSeasonData = {{
    'EFG': [{efg_season}, {round(100-efg_season,1)}],
    'TS':  [{ts_season},  {round(100-ts_season,1)}],
    'USG': [{usg_season}, {round(100-usg_season,1)}]
  }};
  var pieSeasonColors = {{
    'EFG': ['#534AB7','#EEEDFE'], 'TS': ['#0F6E56','#E1F5EE'], 'USG': ['#D85A30','#FAECE7']
  }};
  var pieSeasonLabels = {{
    'EFG': ['eFG%','pozostałe'], 'TS': ['TS%','pozostałe'], 'USG': ['USG%','pozostałe']
  }};
  var shootData = {{
    'PTS': {{ labels: ['2PT%','3PT%','FT%'], data: [{p2pct_js},{p3pct_js},{ftpct_js}],
              colors: ['#1a2b4a','#378ADD','#1D9E75'], title: 'Skuteczność rzutów', unit: '%', type: 'bar' }},
    'AST': {{ labels: ['Min','Śr.','Max'], data: {ast_season_js}, colors: ['#aad4f5','#378ADD','#1a5fa0'], title: 'Asysty', unit: '', type: 'bar' }},
    'REB':  {{ labels: ['Min','Śr.','Max'], data: {reb_season_js},  colors: ['#a8e6cf','#1D9E75','#0d6b4f'], title: 'Rebounds', unit: '', type: 'bar' }},
    'OREB': {{ labels: ['Min','Śr.','Max'], data: {oreb_season_js}, colors: ['#b6e8cc','#27ae60','#1a6e3f'], title: 'Off. Reb',  unit: '', type: 'bar' }},
    'DREB': {{ labels: ['Min','Śr.','Max'], data: {dreb_season_js}, colors: ['#a8e6cf','#16875a','#0a4f35'], title: 'Def. Reb',  unit: '', type: 'bar' }},
    'TO':   {{ labels: ['Min','Śr.','Max'], data: {br_season_js},   colors: ['#f5c6a0','#D85A30','#9e3a18'], title: 'Straty', unit: '', type: 'bar' }},
    'FIN': {{ labels: ['Min','Śr.','Max'], data: {fin_season_js}, colors: ['#c5ccd6','#555','#1a2b4a'],    title: 'Wykończenia', unit: '', type: 'bar' }},
    'STL': {{ labels: ['Min','Śr.','Max'], data: {stl_season_js}, colors: ['#b5d4f4','#185FA5','#0C447C'], title: 'Przechwyty', unit: '', type: 'bar' }},
    'BLK': {{ labels: ['Min','Śr.','Max'], data: {blk_season_js}, colors: ['#b5d4f4','#185FA5','#042C53'], title: 'Bloki', unit: '', type: 'bar' }},
    'EFG': {{ labels: pieSeasonLabels['EFG'], data: pieSeasonData['EFG'], colors: pieSeasonColors['EFG'], title: 'eFG%', unit: '%', type: 'doughnut' }},
    'TS':  {{ labels: pieSeasonLabels['TS'],  data: pieSeasonData['TS'],  colors: pieSeasonColors['TS'],  title: 'TS%',  unit: '%', type: 'doughnut' }},
    'USG': {{ labels: pieSeasonLabels['USG'], data: pieSeasonData['USG'], colors: pieSeasonColors['USG'], title: 'USG%', unit: '%', type: 'doughnut' }}
  }};

  var mainChart = null, shootChart = null;

  function switchMetric(metric) {{
    ['PTS','AST','REB','OREB','DREB','STL','BLK','TO','FIN','EFG','TS','USG'].forEach(function(m) {{
      var th = document.getElementById('th'+m);
      if (th) {{ th.style.opacity = (m===metric) ? '1' : '0.6'; }}
    }});
    var titleEl = document.getElementById('chartMainTitle');
    if (titleEl) titleEl.textContent = titles[metric] || metric;
    if (mainChart) {{ mainChart.destroy(); mainChart = null; }}
    var canvas = document.getElementById('chartMain');
    if (!canvas) return;

    if (pieMetrics.has(metric)) {{
      var pd = piePerMatch[metric];
      var validData = pd.data.filter(function(v){{ return v !== null; }});
      var avgVal = validData.length ? (validData.reduce(function(a,b){{return a+b;}},0)/validData.length).toFixed(1) : '0.0';
      var cl = document.getElementById('chartMainCenter');
      var cv = document.getElementById('chartMainCenterVal');
      if (cl) cl.style.display = 'block';
      if (cv) cv.textContent = avgVal + '%';
      mainChart = new Chart(canvas.getContext('2d'), {{
        type: 'doughnut',
        data: {{ labels: [metric+'% (śr.)','pozostałe'],
          datasets: [{{ data: [parseFloat(avgVal), Math.max(0,100-parseFloat(avgVal))],
            backgroundColor: [pd.color, pd.color+'33'], borderWidth: 0 }}] }},
        options: {{ responsive: true, maintainAspectRatio: false, cutout: '68%',
          plugins: {{ legend: {{ display: true, position: 'bottom', labels: {{ font: {{ size: 11 }}, boxWidth: 12 }} }},
            tooltip: {{ callbacks: {{ label: function(c){{ return c.parsed.toFixed(1)+'%'; }} }} }} }} }}
      }});
    }} else {{
      var cl = document.getElementById('chartMainCenter');
      if (cl) cl.style.display = 'none';
      var data = dataSets[metric];
      var avg = data.reduce(function(a,b){{return a+(b||0);}},0)/Math.max(data.length,1);
      if (metric==='PTS') avg = ptsAvg;
      var datasets = [{{ type:'bar', label:metric, data:data,
        backgroundColor: data.map(function(){{ return colors[metric]||'#1a2b4a'; }}),
        borderRadius:3, barThickness:barW, order:2 }}];
      if (data.length>1) datasets.push({{ type:'line', label:'Średnia',
        data:labels.map(function(){{ return Math.round(avg*10)/10; }}),
        borderColor:'#D85A30', borderWidth:2, borderDash:[4,3],
        pointRadius:0, fill:false, order:1 }});
      mainChart = new Chart(canvas.getContext('2d'), {{
        type:'bar', data:{{ labels:labels, datasets:datasets }},
        options:{{ responsive:true, maintainAspectRatio:false,
          plugins:{{ legend:{{ display:false }},
            tooltip:{{ callbacks:{{ label:function(ctx){{ return ctx.dataset.label+': '+ctx.parsed.y; }} }} }} }},
          scales:{{ x:{{ ticks:{{ font:{{ size:9 }}, maxRotation:40, autoSkip:false }} }},
            y:{{ min:0, ticks:{{ font:{{ size:10 }} }}, grid:{{ color:'rgba(0,0,0,0.05)' }} }} }} }}
      }});
    }}
    updateShootChart(metric);
  }}

  function updateShootChart(metric) {{
    var cfg = shootData[metric] || shootData['PTS'];
    var titleEl = document.getElementById('chartShootTitle');
    if (titleEl) titleEl.textContent = cfg.title;
    if (shootChart) {{ shootChart.destroy(); shootChart = null; }}
    var canvas = document.getElementById('chartShoot');
    if (!canvas) return;
    if (cfg.type==='doughnut') {{
      var scl = document.getElementById('chartShootCenter');
      var scv = document.getElementById('chartShootCenterVal');
      if (scl) scl.style.display = 'block';
      if (scv) scv.textContent = cfg.data[0].toFixed(1)+'%';
      shootChart = new Chart(canvas.getContext('2d'), {{
        type:'doughnut',
        data:{{ labels:cfg.labels, datasets:[{{ data:cfg.data, backgroundColor:cfg.colors, borderWidth:0 }}] }},
        options:{{ responsive:true, maintainAspectRatio:false, cutout:'68%',
          plugins:{{ legend:{{ display:true, position:'bottom', labels:{{ font:{{ size:11 }}, boxWidth:12 }} }},
            tooltip:{{ callbacks:{{ label:function(c){{ return c.parsed.toFixed(1)+'%'; }} }} }} }} }}
      }});
    }} else {{
      var scl = document.getElementById('chartShootCenter');
      if (scl) scl.style.display = 'none';
      var isPct = cfg.unit==='%';
      shootChart = new Chart(canvas.getContext('2d'), {{
        type:'bar',
        data:{{ labels:cfg.labels, datasets:[{{ data:cfg.data, backgroundColor:cfg.colors, borderRadius:4, barThickness:40 }}] }},
        options:{{ responsive:true, maintainAspectRatio:false,
          plugins:{{ legend:{{ display:false }},
            tooltip:{{ callbacks:{{ label:function(c){{ return isPct ? c.parsed.y.toFixed(1)+'%' : c.parsed.y; }} }} }} }},
          scales:{{ x:{{ ticks:{{ font:{{ size:12 }} }} }},
            y:{{ min:0, max:isPct?100:undefined,
              ticks:{{ font:{{ size:10 }}, callback:function(v){{ return isPct?v+'%':v; }} }},
              grid:{{ color:'rgba(0,0,0,0.05)' }} }} }} }}
      }});
    }}
  }}

  document.querySelectorAll('[data-bs-toggle="tooltip"]').forEach(function(el) {{
    new bootstrap.Tooltip(el, {{ trigger:'hover', html:false }});
  }});

  switchMetric('PTS');
</script>
</body></html>"""
    return html_response(page)
